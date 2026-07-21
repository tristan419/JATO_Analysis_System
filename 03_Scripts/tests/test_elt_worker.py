from __future__ import annotations

import importlib.util
import json
from pathlib import Path
import sys

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook


MODULE_PATH = Path(__file__).resolve().parents[1] / "elt_worker.py"


def load_module():
    module_name = "elt_worker_test_module"
    if module_name in sys.modules:
        return sys.modules[module_name]

    script_root = str(MODULE_PATH.parent)
    if script_root not in sys.path:
        sys.path.insert(0, script_root)

    spec = importlib.util.spec_from_file_location(module_name, MODULE_PATH)
    assert spec is not None
    assert spec.loader is not None

    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


elt_worker_module = load_module()


def test_manifest_binds_output_parquet_sha_and_size(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "candidate.parquet"
    frame = pd.DataFrame(
        [{"Country": "Hungary", "2026 Jun": 1}]
    )
    frame.to_parquet(output_path, index=False)
    source_path = tmp_path / "source.xlsx"
    source_path.write_bytes(b"source")
    manifest_path = tmp_path / "manifest.json"

    elt_worker_module.write_manifest(
        manifest_path=manifest_path,
        source_files=[source_path],
        output_file=output_path,
        sheet_name="Data Export",
        df=frame,
        elapsed_seconds=1.0,
        validation_summary={},
        merge_summary={},
    )

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest["fileSizeBytes"] == output_path.stat().st_size
    assert manifest["sha256"] == elt_worker_module.sha256_file(
        output_path
    )


def _write_streaming_test_workbook(path: Path) -> None:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Data Export")
    worksheet.append(
        ["国家", "Version name", "MSRP规整", "2026 May", "2026 Jun"]
    )
    worksheet.append(["捷克", "T5", 31_000, 10, 12])
    worksheet.append(["丹麦", 5, 32_500.5, 8, 9])
    worksheet.append(["丹麦", "T5 EVO", "", 0, 4])
    for index in range(510):
        worksheet.append(["捷克", f"V{index}", 30_000 + index, 0, 1])
    workbook.save(path)


def test_streaming_xlsx_conversion_uses_active_schema_and_bounded_row_groups(
    tmp_path: Path,
    monkeypatch,
) -> None:
    source_path = tmp_path / "washed.xlsx"
    _write_streaming_test_workbook(source_path)
    schema_path = tmp_path / "active.parquet"
    pd.DataFrame(
        {
            "国家": pd.Series(["捷克"], dtype="string"),
            "Version name": pd.Series(["T5"], dtype="string"),
            "MSRP规整": pd.Series([30_000.0], dtype="float64"),
            "2026 May": pd.Series([10], dtype="int64"),
        }
    ).to_parquet(schema_path, index=False)
    output_path = tmp_path / "candidate.parquet"
    manifest_path = tmp_path / "manifest.json"

    monkeypatch.setattr(
        elt_worker_module,
        "read_excel_with_fallback",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("streaming path must not materialize the workbook")
        ),
    )

    elt_worker_module.convert_jato_xlsx_to_parquet_streaming(
        input_path=str(source_path),
        output_path=str(output_path),
        manifest_path=str(manifest_path),
        sheet_name="Data Export",
        schema_from_parquet=str(schema_path),
        batch_rows=256,
        job_id="streaming-test",
    )

    result = pd.read_parquet(output_path)
    assert len(result) == 513
    assert result["国家"].tolist()[:3] == ["捷克", "丹麦", "丹麦"]
    assert result["Version name"].tolist()[:3] == ["T5", "5", "T5 EVO"]
    assert result["MSRP规整"].tolist()[:2] == [31_000.0, 32_500.5]
    assert pd.isna(result["MSRP规整"].iloc[2])
    assert result["2026 Jun"].tolist()[:3] == [12.0, 9.0, 4.0]
    assert str(pq.read_schema(output_path).field("Version name").type) == "string"
    assert str(pq.read_schema(output_path).field("2026 May").type) == "int64"
    assert str(pq.read_schema(output_path).field("2026 Jun").type) == "double"
    unified_schema = pa.unify_schemas(
        [pq.read_schema(schema_path), pq.read_schema(output_path)],
        promote_options="permissive",
    )
    assert str(unified_schema.field("2026 May").type) == "int64"
    parquet_metadata = pq.ParquetFile(output_path).metadata
    assert parquet_metadata.num_rows == 513
    assert parquet_metadata.num_row_groups == 3
    assert all(
        parquet_metadata.row_group(index).num_rows <= 256
        for index in range(parquet_metadata.num_row_groups)
    )

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest["rows"] == 513
    assert manifest["columns"] == 5
    assert manifest["mergeSummary"]["streamingXlsx"] is True
    assert manifest["mergeSummary"]["streamingBatchRows"] == 256
    assert manifest["mergeSummary"]["schemaHintedColumnCount"] == 4


def test_streaming_xlsx_conversion_fails_closed_on_invalid_numeric_value(
    tmp_path: Path,
) -> None:
    source_path = tmp_path / "invalid.xlsx"
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Data Export")
    worksheet.append(["国家", "MSRP规整", "2026 Jun"])
    worksheet.append(["捷克", "not-a-price", 12])
    workbook.save(source_path)
    schema_path = tmp_path / "active.parquet"
    pd.DataFrame(
        {"国家": ["捷克"], "MSRP规整": [30_000.0], "2026 May": [10.0]}
    ).to_parquet(schema_path, index=False)
    output_path = tmp_path / "candidate.parquet"
    manifest_path = tmp_path / "manifest.json"

    try:
        elt_worker_module.convert_jato_xlsx_to_parquet_streaming(
            input_path=str(source_path),
            output_path=str(output_path),
            manifest_path=str(manifest_path),
            sheet_name="Data Export",
            schema_from_parquet=str(schema_path),
            batch_rows=256,
            job_id="streaming-invalid-test",
        )
    except ValueError as exc:
        assert "MSRP规整" in str(exc)
        assert "非数值" in str(exc)
    else:
        raise AssertionError("invalid numeric data must fail closed")

    assert not output_path.exists()
    assert not manifest_path.exists()


def test_supplement_missing_countries_from_parquet_adds_absent_country_rows(
    tmp_path: Path,
) -> None:
    source_path = tmp_path / "patch.xlsx"
    source_path.write_bytes(b"fake")
    current_df = elt_worker_module.add_source_tracking_columns(
        pd.DataFrame(
            [
                {"国家": "德国", "Model": "ID.4", "2026 Mar": 21},
            ]
        ),
        source_file=source_path,
        source_index=1,
    )
    current_df = elt_worker_module.normalize_dataframe(current_df)

    supplement_path = tmp_path / "active.parquet"
    pd.DataFrame(
        [
            {"国家": "德国", "Model": "ID.4", "2026 Mar": 21},
            {"国家": "瑞典", "Model": "EX30", "2026 Mar": 12},
        ]
    ).to_parquet(supplement_path, index=False)

    merged_df, summary = elt_worker_module.supplement_missing_countries_from_parquet(
        current_df,
        str(supplement_path),
        source_index=2,
    )

    assert summary["enabled"] is True
    assert summary["supplementedCountryCount"] == 1
    assert summary["supplementedCountries"] == ["瑞典"]
    assert summary["supplementedRowCount"] == 1
    assert summary["replacedCountryCount"] == 0
    assert summary["replacedCountries"] == []
    assert sorted(merged_df["国家"].astype("string").tolist()) == ["德国", "瑞典"]


def test_supplement_missing_countries_from_parquet_replaces_unuploaded_baseline_countries_with_active(
    tmp_path: Path,
) -> None:
    baseline_path = tmp_path / "baseline.xlsx"
    patch_path = tmp_path / "patch.xlsx"
    baseline_path.write_bytes(b"baseline")
    patch_path.write_bytes(b"patch")

    baseline_df = elt_worker_module.add_source_tracking_columns(
        pd.DataFrame(
            [
                {"国家": "德国", "Model": "ID.4", "2026 Jan": 11},
                {"国家": "西班牙", "Model": "Born", "2026 Jan": 5},
            ]
        ),
        source_file=baseline_path,
        source_index=1,
    )
    patch_df = elt_worker_module.add_source_tracking_columns(
        pd.DataFrame(
            [
                {"国家": "德国", "Model": "ID.4", "2026 Mar": 21},
            ]
        ),
        source_file=patch_path,
        source_index=2,
    )
    current_df = elt_worker_module.normalize_dataframe(
        pd.concat([baseline_df, patch_df], ignore_index=True, sort=False)
    )

    supplement_path = tmp_path / "active.parquet"
    pd.DataFrame(
        [
            {"国家": "德国", "Model": "ID.4", "2026 Mar": 21},
            {"国家": "西班牙", "Model": "Born", "2026 Mar": 18},
        ]
    ).to_parquet(supplement_path, index=False)

    merged_df, summary = elt_worker_module.supplement_missing_countries_from_parquet(
        current_df,
        str(supplement_path),
        source_index=3,
        patch_source_indices={2},
    )

    assert summary["supplementedCountryCount"] == 1
    assert summary["supplementedCountries"] == ["西班牙"]
    assert summary["replacedCountryCount"] == 1
    assert summary["replacedCountries"] == ["西班牙"]
    assert summary["supplementedRowCount"] == 1

    germany_rows = merged_df[merged_df["国家"] == "德国"].copy()
    spain_rows = merged_df[merged_df["国家"] == "西班牙"].copy()
    assert len(germany_rows) == 1
    assert len(spain_rows) == 1
    assert spain_rows["2026 Mar"].iloc[0] == 18
    assert pd.isna(spain_rows["2026 Jan"].iloc[0])
    assert germany_rows["2026 Mar"].iloc[0] == 21
    assert pd.isna(germany_rows["2026 Jan"].iloc[0])


def test_uploaded_country_replaces_baseline_rows_instead_of_double_counting(
    tmp_path: Path,
) -> None:
    baseline_path = tmp_path / "baseline.xlsx"
    patch_path = tmp_path / "patch.xlsx"
    baseline_path.write_bytes(b"baseline")
    patch_path.write_bytes(b"patch")

    baseline_df = elt_worker_module.add_source_tracking_columns(
        pd.DataFrame(
            [
                {"国家": "德国", "2025 Jan": 195969, "2026 Jan": 182526},
                {"国家": "瑞典", "2025 Jan": 19609, "2026 Jan": 16041},
            ]
        ),
        source_file=baseline_path,
        source_index=1,
    )
    patch_df = elt_worker_module.add_source_tracking_columns(
        pd.DataFrame(
            [
                {"国家": "德国", "2025 Jan": 195969, "2026 Jan": 182526, "2026 Feb": 195317, "2026 Mar": 275452},
            ]
        ),
        source_file=patch_path,
        source_index=2,
    )
    current_df = elt_worker_module.normalize_dataframe(
        pd.concat([baseline_df, patch_df], ignore_index=True, sort=False)
    )

    supplement_path = tmp_path / "active.parquet"
    pd.DataFrame(
        [
            {"国家": "德国", "2025 Jan": 413522, "2026 Jan": 387382},
            {"国家": "瑞典", "2025 Jan": 19632, "2026 Jan": 16041, "2026 Mar": 26578},
        ]
    ).to_parquet(supplement_path, index=False)

    merged_df, _summary = elt_worker_module.supplement_missing_countries_from_parquet(
        current_df,
        str(supplement_path),
        source_index=3,
        patch_source_indices={2},
    )

    germany_rows = merged_df[merged_df["国家"] == "德国"].copy()
    assert len(germany_rows) == 1
    assert germany_rows["2025 Jan"].iloc[0] == 195969
    assert germany_rows["2026 Jan"].iloc[0] == 182526
    assert germany_rows["2026 Feb"].iloc[0] == 195317
    assert germany_rows["2026 Mar"].iloc[0] == 275452


def test_supplement_matches_country_by_logical_key_without_double_counting(
    tmp_path: Path,
) -> None:
    baseline_path = tmp_path / "baseline.xlsx"
    patch_path = tmp_path / "patch.xlsx"
    baseline_path.write_bytes(b"baseline")
    patch_path.write_bytes(b"patch")
    baseline_df = elt_worker_module.add_source_tracking_columns(
        pd.DataFrame(
            [{"Country": "Hungary", "Model": "old", "2026 Apr": 10}]
        ),
        source_file=baseline_path,
        source_index=1,
    )
    patch_df = elt_worker_module.add_source_tracking_columns(
        pd.DataFrame(
            [
                {
                    "Country": "  hungary ",
                    "Model": "latest",
                    "2026 Apr": 10,
                    "2026 May": 12,
                }
            ]
        ),
        source_file=patch_path,
        source_index=2,
    )
    current_df = elt_worker_module.normalize_dataframe(
        pd.concat([baseline_df, patch_df], ignore_index=True, sort=False)
    )
    supplement_path = tmp_path / "active.parquet"
    pd.DataFrame(
        [
            {"Country": "Hungary", "Model": "old", "2026 Apr": 10},
            {"Country": "Germany", "Model": "ID.4", "2026 Apr": 20},
        ]
    ).to_parquet(supplement_path, index=False)

    merged_df, summary = (
        elt_worker_module.supplement_missing_countries_from_parquet(
            current_df,
            str(supplement_path),
            source_index=3,
            patch_source_indices={2},
        )
    )

    assert summary["supplementedCountries"] == ["Germany"]
    assert merged_df["Country"].tolist().count("Hungary") == 1
    assert "hungary" not in merged_df["Country"].tolist()
    hungary_rows = merged_df.loc[merged_df["Country"] == "Hungary"]
    assert len(hungary_rows) == 1
    assert hungary_rows["Model"].iloc[0] == "latest"
    assert hungary_rows["2026 Apr"].iloc[0] == 10
    assert hungary_rows["2026 May"].iloc[0] == 12


def test_supplement_rejects_ambiguous_country_display_variants(
    tmp_path: Path,
) -> None:
    current_df = elt_worker_module.normalize_dataframe(
        pd.DataFrame([{"Country": "Germany", "2026 May": 10}])
    )
    supplement_path = tmp_path / "active.parquet"
    pd.DataFrame(
        [
            {"Country": "Hungary", "2026 Apr": 10},
            {"Country": "hungary", "2026 Apr": 20},
        ]
    ).to_parquet(supplement_path, index=False)

    try:
        elt_worker_module.supplement_missing_countries_from_parquet(
            current_df,
            str(supplement_path),
            source_index=2,
        )
    except ValueError as exc:
        assert "同一逻辑国家" in str(exc)
        assert "重复累加" in str(exc)
    else:
        raise AssertionError("ambiguous logical countries must be rejected")
