from __future__ import annotations

from copy import copy
from pathlib import Path
import random

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_PATH = ROOT / "01_RAW_DATA" / "VOC_Nordic_SUV_Users_100.xlsx"
OUTPUT_PATH = (
    ROOT
    / "04_Processed_data"
    / "voc"
    / "se"
    / "raw"
    / "se_hev_owners_benchmark_voc_20260423.xlsx"
)
ROW_COUNT = 65
SEED = 42


MALE_NAMES = [
    "Erik",
    "Lars",
    "Johan",
    "Anders",
    "Mikael",
    "Karl",
    "Peter",
    "Mattias",
    "Stefan",
    "Andreas",
    "Bjorn",
    "Henrik",
    "Niklas",
    "Marcus",
    "Daniel",
    "David",
    "Magnus",
    "Robert",
    "Thomas",
    "Christian",
    "Fredrik",
    "Patrik",
    "Oskar",
    "Viktor",
    "Jonas",
    "Linus",
    "Gustav",
    "Sebastian",
    "Simon",
    "Emil",
]
FEMALE_NAMES = [
    "Anna",
    "Maria",
    "Karin",
    "Sara",
    "Emma",
    "Lena",
    "Kristina",
    "Malin",
    "Sofia",
    "Eva",
    "Helena",
    "Camilla",
    "Linda",
    "Therese",
    "Cecilia",
    "Sandra",
    "Ulrika",
    "Johanna",
    "Ingrid",
    "Rebecka",
    "Elin",
    "Hanna",
    "Ida",
    "Astrid",
    "Maja",
    "Frida",
    "Klara",
    "Nora",
    "Elsa",
    "Vera",
]
LAST_NAMES = [
    "Andersson",
    "Johansson",
    "Karlsson",
    "Nilsson",
    "Eriksson",
    "Larsson",
    "Olsson",
    "Persson",
    "Svensson",
    "Gustafsson",
    "Pettersson",
    "Jonsson",
    "Jansson",
    "Hansson",
    "Bengtsson",
    "Lindstrom",
    "Lindqvist",
    "Magnusson",
    "Berg",
    "Lindberg",
    "Holm",
    "Hedlund",
    "Lund",
    "Bergstrom",
    "Sandberg",
]
REGIONS = [
    "Stockholm",
    "Goteborg",
    "Malmo",
    "Uppsala",
    "Linkoping",
    "Vasteras",
    "Orebro",
    "Helsingborg",
    "Norrkoping",
    "Jonkoping",
    "Umea",
    "Lund",
    "Boras",
    "Sundsvall",
    "Gavle",
    "Eskilstuna",
    "Sodertalje",
    "Karlstad",
    "Taby",
    "Nacka",
]
OCCUPATIONS = [
    ("Engineer / Tech", 680000),
    ("Teacher / Education", 480000),
    ("Healthcare / Nurse", 520000),
    ("Manager / Business", 820000),
    ("Accountant / Finance", 620000),
    ("IT Consultant", 750000),
    ("Logistics / Supply Chain", 560000),
    ("Sales / Marketing", 600000),
    ("Civil Servant", 510000),
    ("Architect / Design", 690000),
    ("Doctor / Medical", 980000),
    ("Lawyer / Legal", 870000),
    ("Social Worker", 430000),
    ("Construction / Project Mgr", 640000),
    ("Retail Manager", 490000),
]
CAR_MODELS = [
    "Toyota RAV4 Hybrid",
    "Toyota RAV4 Hybrid",
    "Toyota RAV4 Hybrid",
    "Toyota Corolla Cross Hybrid",
    "Toyota Corolla Cross Hybrid",
    "Toyota C-HR Hybrid",
    "Toyota Yaris Cross Hybrid",
    "Toyota Camry Hybrid",
    "Toyota Highlander Hybrid",
    "Hyundai Kona Hybrid",
    "Toyota Corolla Hybrid",
]
WHY_CAR_POOL = [
    "Wanted low fuel costs without charging anxiety. Toyota's reliability was key.",
    "No garage at apartment, so PHEV and BEV were not practical. HEV fits daily life better.",
    "Previous Toyota owner and never had issues. Trusted the brand completely.",
    "Winter reliability is non-negotiable. HEV feels safer than BEV in cold months.",
    "Low five-year TCO and strong residual value made the decision easy.",
    "Family needed AWD for winter plus low running costs. RAV4 Hybrid ticked all boxes.",
    "Test drove BEV but winter range anxiety was still too high. HEV felt safer.",
    "Work covers fuel but not charging. HEV made more financial sense.",
    "Recommended by a colleague who has run the same car for six years without issues.",
    "Best resale value in class. Practical choice, not emotional.",
]
PAIN_POINTS_POOL = [
    "Highway fuel economy disappoints and is worse than the city figure.",
    "CVT rubber-band feeling is irritating during motorway overtaking.",
    "Infotainment feels five years behind competitors and CarPlay can be buggy.",
    "Rear legroom gets tight for adults on long trips.",
    "Boot space is compromised by the hybrid packaging.",
    "Touchscreen sensitivity is poor with winter gloves.",
    "Engine noise intrudes harshly under hard acceleration.",
    "No over-the-air updates, so software fixes mean a dealer visit.",
    "Winter highway economy at 110 km/h is closer to 7L/100km than the brochure claim.",
    "Styling feels conservative and not exciting.",
]
DREAM_CAR_POOL = [
    "A future BEV with 600 plus km real range and reliable fast charging everywhere.",
    "Toyota RAV4 PHEV if charging infrastructure improves.",
    "Happy to upgrade to the next-gen RAV4 Hybrid when battery performance improves.",
    "A BEV family SUV once prices drop and charging matures.",
    "Possibly a Volvo EX90 in three to four years if cold-weather range holds.",
    "Would consider a Chinese brand if quality proves itself over time.",
    "Same car but with better infotainment and a larger boot.",
    "Still undecided and watching the BEV market carefully.",
]
SCENARIOS_POOL = [
    "Daily commute of 45 km, school runs, and weekend ski trips to Dalarna.",
    "City driving in Stockholm plus monthly trips to a family cabin.",
    "Office commute 35 km each way and family holidays to Norway.",
    "Mostly urban driving with occasional 400 km drives to relatives.",
    "Work trips, school pickups, and winter mountain driving.",
    "City and suburb mix with quarterly long-distance trips.",
    "Daily highway commute of 60 km, camping in summer, skiing in winter.",
    "Flexible schedule with errands during the week and 300 km family trips.",
]
FEATURES_POOL = [
    [
        "Low fuel cost around 5.5L/100km in urban driving",
        "Smooth city driving on electric assist",
        "Exceptional reliability record",
    ],
    [
        "No charging needed and refuelling is easy anywhere",
        "Strong resale value",
        "AWD confidence in winter snow",
    ],
    [
        "Quiet cabin in hybrid mode",
        "Toyota Safety Sense suite",
        "Spacious interior for family use",
    ],
    [
        "Regenerative braking feels natural",
        "Low maintenance costs",
        "Trustworthy brand reputation",
    ],
    [
        "Seamless petrol-electric switch",
        "Good headroom and visibility",
        "Comfortable long-distance seats",
    ],
    [
        "Affordable insurance class",
        "Strong dealer network across Sweden",
        "Class-leading reliability data",
    ],
]
COMPLAINTS_POOL = [
    [
        "Highway fuel economy 6.8 to 7.2L versus 5.5L advertised",
        "Outdated infotainment UI",
        "Boot is smaller than some competitors",
    ],
    [
        "CVT rubber-band effect at high speed",
        "No OTA software updates",
        "Rear space is tight for three adults",
    ],
    [
        "Feels like transition tech and not exciting",
        "Engine gets loud under hard load",
        "CarPlay connection drops too often",
    ],
    [
        "No ventilated seats",
        "Dashboard design feels dated",
        "Dealer wait times can be long",
    ],
    [
        "Fuel economy drops sharply above 100 km/h",
        "Touchscreen is unresponsive in cold weather",
        "Lacks driving engagement",
    ],
]
INFO_SOURCES = [
    "Motorblog.se and YouTube reviews",
    "Bilweb.se and word of mouth",
    "Blocket.se and colleague recommendations",
    "Vi Bilagare magazine and dealer visit",
    "Owners forum and hybrid Reddit threads",
    "Dealer test drive and Autocar Sweden",
    "Facebook group Toyota Sverige",
    "Friends recommendation and Google research",
    "Consumer reports and dealer visit",
    "Teknikens Varld and YouTube",
]
PRICE_PERCEPTIONS = [
    "Fair because total cost of ownership justifies the premium over petrol.",
    "Slightly high upfront but it saves money long-term on fuel.",
    "Competitive once reliability and residual value are included.",
    "Worth it because the car has avoided unexpected repair bills.",
    "A bit expensive, but cheaper alternatives felt riskier.",
    "Good value versus a BEV that still costs far more.",
]
EVALUATIONS = [
    "4.2/5 - Reliable workhorse. Would buy again.",
    "4.0/5 - Does the job perfectly. Not exciting but trustworthy.",
    "3.8/5 - Great in city use, weaker on highway economy.",
    "4.5/5 - Best family car I have owned.",
    "3.9/5 - Solid choice but infotainment needs an update.",
    "4.3/5 - Exactly what a Swedish family car should be.",
    "4.1/5 - Reliable, efficient, practical. Main weakness is the UI.",
]
CLOSING_POOL = [
    '"Next car will probably be a BEV if charging infrastructure matures."',
    '"Toyota has earned my loyalty. I will likely stay with the brand."',
    '"Hybrid is the sensible middle ground right now. No regrets."',
    '"I tell people to consider HEV before jumping to full electric."',
    '"I want better tech features, but overall I am very satisfied."',
    '"This car just works, and that is what I need."',
    '"I will reassess in three years when the BEV market is more mature."',
]
SUGGESTIONS_POOL = [
    "Better infotainment with wireless CarPlay and more honest real-world fuel economy guidance.",
    "More boot space and an updated UI would make this close to perfect.",
    "Seat ventilation and a more engaging drive feel are still missing.",
    "OTA updates and better digital integration would help a lot.",
    "Dealers should communicate real-world versus rated fuel economy more clearly.",
    "A longer wheelbase for rear passenger space would be appreciated.",
]
REQUIREMENTS_POOL = [
    "Low TCO, AWD capability, minimum 600L boot space, and proven winter performance.",
    "Fuel cost under 6L/100km, reliability for 300,000 plus km, and good resale value.",
    "Comfort for two adults plus two children, winter-ready setup, and low maintenance.",
    "Modern infotainment, strong safety rating, and practical family space.",
    "Easy daily ownership, no range anxiety, and dealer support across Sweden.",
]
HOBBIES_POOL = [
    "Skiing and hiking",
    "Cycling and outdoor sports",
    "Football and gym",
    "Swimming and running",
    "Reading and DIY",
    "Fishing and nature trips",
    "Tennis and golf",
    "Photography and travel",
    "Cooking and gardening",
    "Cross-country skiing and nature walks",
]
LOCATIONS_POOL = [
    "Stockholm city and Lidingo",
    "Goteborg plus the west coast archipelago",
    "Suburb and city centre every week",
    "Work, school, and supermarket circuit",
    "Summer cabin and winter ski resort",
    "Uppsala to Stockholm corridor",
    "Malmo and the Oresund region",
    "Workplace, kids' activities, and IKEA",
    "Nordic road trips",
    "Local routines plus seasonal outdoor destinations",
]
SPENDING_POOL = [
    "Rational and research-heavy, buying for quality and longevity.",
    "Value-driven and careful, rarely buys impulsively.",
    "Conservative, prefers proven products over trends.",
    "Practical and willing to pay for lower long-term cost.",
    "Deliberate and compares several options before buying.",
]
LIFE_PATTERNS = [
    "Early riser with active family weekends and a structured routine.",
    "Works from home two days a week, school pickup daily, gym in the evenings.",
    "Commuter lifestyle on weekdays and outdoors on weekends.",
    "Semi-retired, active, and often travelling regionally.",
    "Dual-income household with balanced work-life and quarterly travel.",
]
PRODUCT_PREFS = [
    "Toyota, Bosch, and Apple ecosystem",
    "Scandinavian brands plus proven Japanese reliability",
    "Function over form with a focus on durable products",
    "Trusted brands with strong warranty support",
    "Mid-range premium, avoiding both budget and luxury extremes",
]
SOCIAL_CIRCLE_TYPES = [
    "Family and colleagues",
    "Friends and neighbourhood",
    "Professional network",
    "School parent group",
    "Outdoor club and family",
]
PERSONALITIES = [
    "Analytical and risk-averse",
    "Pragmatic and family-focused",
    "Organised and detail-oriented",
    "Calm, patient, and long-term minded",
    "Responsible and value-conscious",
    "Methodical and dislikes uncertainty",
]
POWERTRAIN_PREFS = [
    "HEV for now, open to BEV in 5 years",
    "HEV and no BEV until home charging becomes practical",
    "Prefer HEV until BEV winter range exceeds 500 km in real use",
    "HEV because reliability matters more than chasing EV range claims",
    "HEV is the right middle ground for this use case",
]
FUTURE_REQS = [
    "Minimum 500 km real winter range, under-20-minute fast charging, and proven reliability.",
    "Toyota-level reliability but with much better infotainment.",
    "A BEV only if the apartment building installs chargers.",
    "Lower price point plus stronger after-sales support.",
    "More interior space, better digital experience, and the same TCO discipline.",
]
MARITAL_STATUSES = [
    "Married",
    "Married",
    "Married",
    "Partnered",
    "Divorced",
    "Single",
    "Married",
    "Married",
]
EDUCATION_LEVELS = [
    "University degree",
    "Master's degree",
    "Vocational training",
    "University degree",
    "Master's degree",
    "PhD",
    "University degree",
    "College diploma",
]
PURCHASE_TYPES = [
    "Private new",
    "Private new",
    "Private used",
    "Company car",
    "Private new",
    "Private used",
    "Private new",
    "Private new",
]
USAGE_FREQS = [
    "Daily (7 days/week)",
    "Daily (5 to 6 days/week)",
    "5 days/week",
    "Daily plus long trips monthly",
    "Daily commute plus weekend leisure",
]


def format_income(amount: int) -> str:
    return f"{amount:,} SEK/year".replace(",", ".")


def build_contact(rng: random.Random) -> str:
    return f"+46 7{rng.randint(0, 9)}-{rng.randint(100, 999)} {rng.randint(10, 99)} {rng.randint(10, 99)}"


def build_rows() -> list[list[str | int]]:
    rng = random.Random(SEED)
    used_names: set[str] = set()
    rows: list[list[str | int]] = []

    for index in range(1, ROW_COUNT + 1):
        gender = rng.choice(["Male", "Female"])
        first_name = rng.choice(MALE_NAMES if gender == "Male" else FEMALE_NAMES)
        last_name = rng.choice(LAST_NAMES)
        full_name = f"{first_name} {last_name}"
        while full_name in used_names:
            first_name = rng.choice(MALE_NAMES if gender == "Male" else FEMALE_NAMES)
            last_name = rng.choice(LAST_NAMES)
            full_name = f"{first_name} {last_name}"
        used_names.add(full_name)

        marital_status = rng.choice(MARITAL_STATUSES)
        children = rng.choice(["0", "1", "2", "2", "3"])
        household_size = str(int(children) + (2 if marital_status in {"Married", "Partnered"} else 1))
        occupation, income = rng.choice(OCCUPATIONS)

        row = [
            full_name,
            build_contact(rng),
            rng.randint(34, 66),
            gender,
            marital_status,
            household_size,
            rng.choice(EDUCATION_LEVELS),
            f"{rng.choice(REGIONS)} (Sweden)",
            children,
            occupation,
            format_income(income),
            f"[Photo_{index:03d}.jpg]",
            rng.choice(HOBBIES_POOL),
            rng.choice(LOCATIONS_POOL),
            rng.choice(SPENDING_POOL),
            rng.choice(LIFE_PATTERNS),
            rng.choice(PRODUCT_PREFS),
            rng.choice(SOCIAL_CIRCLE_TYPES),
            rng.choice(PERSONALITIES),
            rng.choice(CAR_MODELS),
            rng.choice(WHY_CAR_POOL),
            rng.choice(PAIN_POINTS_POOL),
            rng.choice(DREAM_CAR_POOL),
            rng.choice(PURCHASE_TYPES),
            rng.choice(SCENARIOS_POOL),
            rng.choice(FUTURE_REQS),
            rng.choice(USAGE_FREQS),
            rng.choice(POWERTRAIN_PREFS),
            " / ".join(rng.choice(FEATURES_POOL)),
            " / ".join(rng.choice(COMPLAINTS_POOL)),
            rng.choice(REQUIREMENTS_POOL),
            rng.choice(SUGGESTIONS_POOL),
            rng.choice(INFO_SOURCES),
            rng.choice(PRICE_PERCEPTIONS),
            rng.choice(EVALUATIONS),
            rng.choice(CLOSING_POOL),
            f"Interview_SE_HEV_{index:03d}_{first_name}_{last_name}.jpg",
        ]
        rows.append(row)

    return rows


def write_rows() -> Path:
    workbook = load_workbook(TEMPLATE_PATH)
    sheet = workbook["VOC Data"]
    even_styles = [copy(sheet.cell(2, column)._style) for column in range(1, sheet.max_column + 1)]
    odd_styles = [copy(sheet.cell(3, column)._style) for column in range(1, sheet.max_column + 1)]
    data_row_height = sheet.row_dimensions[2].height or 80

    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    for row_index, row_data in enumerate(build_rows(), start=2):
        row_style = even_styles if row_index % 2 == 0 else odd_styles
        for column_index, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell._style = copy(row_style[column_index - 1])
        sheet.row_dimensions[row_index].height = data_row_height

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    output_path = write_rows()
    print(output_path)