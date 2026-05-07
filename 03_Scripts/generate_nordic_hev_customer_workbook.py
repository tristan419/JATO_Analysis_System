from __future__ import annotations

import argparse
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "Name",
    "Contact",
    "Age",
    "Gender",
    "Marital Status",
    "Household Size",
    "Education",
    "Region",
    "Children",
    "Occupation / Industry",
    "Annual Income",
    "Photo",
    "Sports / Hobbies",
    "Frequent Locations",
    "Spending Philosophy",
    "Daily Life Pattern",
    "Product Preference",
    "Social Circle",
    "Personality",
    "Car Ownership / Model",
    "Why This Car?",
    "Current Car Pain Points",
    "Dream Car",
    "Purchase Type",
    "Driving Scenarios",
    "Future Car Requirements",
    "Usage Frequency",
    "Powertrain Preference",
    "Top 3 Favourite Features",
    "Top 3 Complaints",
    "Customer Requirements",
    "Suggestions",
    "Information Source",
    "Price Perception",
    "Evaluation",
    "Closing Remarks",
    "Interview Screenshot",
]

MALE_NAMES = [
    "Erik", "Lars", "Johan", "Anders", "Mikael", "Karl", "Peter", "Mattias", "Stefan", "Andreas",
    "Bjorn", "Henrik", "Niklas", "Marcus", "Daniel", "David", "Magnus", "Robert", "Thomas", "Christian",
    "Fredrik", "Patrik", "Oskar", "Viktor", "Jonas", "Linus", "Gustav", "Sebastian", "Simon", "Emil",
]
FEMALE_NAMES = [
    "Anna", "Maria", "Karin", "Sara", "Emma", "Lena", "Kristina", "Malin", "Sofia", "Eva",
    "Helena", "Camilla", "Linda", "Therese", "Cecilia", "Sandra", "Ulrika", "Johanna", "Ingrid", "Rebecka",
    "Elin", "Hanna", "Ida", "Astrid", "Maja", "Frida", "Klara", "Nora", "Elsa", "Vera",
]
LAST_NAMES = [
    "Andersson", "Johansson", "Karlsson", "Nilsson", "Eriksson", "Larsson", "Olsson", "Persson",
    "Svensson", "Gustafsson", "Pettersson", "Jonsson", "Jansson", "Hansson", "Bengtsson", "Lindstrom",
    "Lindqvist", "Magnusson", "Berg", "Lindberg", "Holm", "Hedlund", "Lund", "Bergstrom", "Sandberg",
]
REGIONS = [
    "Stockholm", "Goteborg", "Malmo", "Uppsala", "Linkoping", "Vasteras", "Orebro", "Helsingborg",
    "Norrkoping", "Jonkoping", "Umea", "Lund", "Boras", "Sundsvall", "Gavle", "Eskilstuna", "Sodertalje",
    "Karlstad", "Taby", "Nacka", "Tampere", "Espoo", "Turku", "Oslo", "Bergen", "Trondheim", "Aarhus",
]
EDUCATION_LEVELS = [
    "University", "University", "University", "College", "Vocational", "College",
]
MARITAL_STATUS = ["Married", "Married", "Cohabiting", "Cohabiting", "Single"]
CHILDREN_LABELS = ["2 children", "1 child", "2 children", "No children", "3 children"]
OCCUPATIONS = [
    "Software Engineer / Tech",
    "Teacher / Education",
    "Healthcare / Nurse",
    "Manager / Business",
    "Accountant / Finance",
    "IT Consultant",
    "Logistics / Supply Chain",
    "Sales / Marketing",
    "Civil Servant",
    "Architect / Design",
    "Doctor / Medical",
    "Construction / Project Manager",
    "Retail Manager",
    "Public Service / Police",
]
SPORTS = [
    "Cross-country skiing", "Trail running", "Hiking / fell walking", "Road cycling", "Swimming",
    "Gardening", "Photography", "Sailing", "Ice hockey", "Snowmobiling",
]
LOCATIONS = [
    "Children's school & activities", "Summer cottage", "Supermarket / ICA / Prisma", "Ski resort / slopes",
    "Hardware store / Bauhaus", "Forest trails", "City centre", "National park", "Sports hall", "Ferry terminal",
]
SPENDING_PHILOSOPHY = [
    "Family-centred practical buyer",
    "Quality first, price secondary",
    "Value for money – functional over flashy",
    "Total cost of ownership focus",
    "Prefers to lease, keeps options open",
]
SOCIAL_CIRCLES = ["Family", "Family", "Colleagues", "Neighbours", "Parents at school"]
PERSONALITIES = ["Pragmatic", "Measured", "Practical", "Calm", "Methodical"]
PRODUCT_PREFERENCE = ["SUV", "SUV", "SUV", "Crossover"]
PURCHASE_TYPES = ["Finance", "Lease", "Finance", "Cash", "Finance"]
USAGE_FREQUENCY = ["Every day", "Every day", "5 days/week", "Varies seasonally"]
INFORMATION_SOURCES = [
    "Manufacturer website",
    "Recommendation from colleague",
    "Dealership test drive event",
    "YouTube comments",
    "Reddit r/cars / r/electricvehicles",
]

PRIMARY_MODELS = [
    ("Toyota RAV4 Hybrid", 34),
    ("Toyota Corolla Cross Hybrid", 22),
    ("Toyota Yaris Cross Hybrid", 14),
    ("Toyota C-HR Hybrid", 10),
    ("Toyota Highlander Hybrid", 8),
    ("Lexus NX 350h", 8),
    ("Honda CR-V e:HEV", 6),
    ("Kia Niro Hybrid", 6),
    ("Hyundai Tucson Hybrid", 6),
    ("Nissan Qashqai e-Power", 6),
]


def weighted_choice(randomizer: random.Random, weighted_items: list[tuple[str, int]]) -> str:
    total = sum(weight for _, weight in weighted_items)
    threshold = randomizer.uniform(0, total)
    running = 0.0
    for value, weight in weighted_items:
        running += weight
        if running >= threshold:
            return value
    return weighted_items[-1][0]


def build_name(randomizer: random.Random, gender: str) -> str:
    first_name = randomizer.choice(MALE_NAMES if gender == "Male" else FEMALE_NAMES)
    last_name = randomizer.choice(LAST_NAMES)
    return f"{first_name} {last_name}"


def build_income(randomizer: random.Random) -> str:
    value = randomizer.randrange(460_000, 980_001, 10_000)
    return str(value)


def build_hobbies(randomizer: random.Random) -> str:
    return "; ".join(randomizer.sample(SPORTS, k=2))


def build_locations(randomizer: random.Random) -> str:
    return "; ".join(randomizer.sample(LOCATIONS, k=2))


def build_daily_life_pattern(randomizer: random.Random, household_size: int) -> str:
    commute_km = randomizer.choice([28, 34, 42, 48, 56, 63])
    family_clause = (
        "evenings revolve around school runs and shopping"
        if household_size >= 4
        else "evenings mix gym, errands and visiting relatives"
    )
    return f"Commutes {commute_km} km/day by car; {family_clause}"


def build_driving_scenarios(randomizer: random.Random) -> str:
    scenarios = [
        "Daily urban commute + school run",
        "Mixed city + rural daily use",
        "Weekend ski trips to the mountains",
        "Long-distance family road trips (Germany/Norway)",
        "Summer cottage",
        "Business travel between cities",
    ]
    return "; ".join(randomizer.sample(scenarios, k=2))


def build_powertrain_preference(randomizer: random.Random) -> str:
    choices = [
        ("Mild hybrid (HEV) – no plug needed", 72),
        ("Open to any if TCO makes sense", 16),
        ("PHEV (daily EV, petrol on long runs)", 8),
        ("Full BEV (500 km+ real-world range)", 4),
    ]
    return weighted_choice(randomizer, choices)


def build_feature_triplet(model: str) -> str:
    if "Toyota" in model:
        return "1. Fuel economy  2. Toyota reliability  3. Heated seats"
    if "Lexus" in model:
        return "1. Quiet cabin  2. Low running cost  3. Heated seats"
    return "1. Fuel economy  2. Winter traction  3. Simple controls"


def build_complaint_triplet(randomizer: random.Random) -> str:
    options = [
        "1. Tyre noise  2. Base audio  3. Rear-seat width",
        "1. Slightly firm ride  2. Boot lip height  3. Small infotainment icons",
        "1. Modest acceleration  2. Road noise  3. Rear bench space",
        "1. Camera quality  2. Seat base length  3. Wireless charging heat",
    ]
    return randomizer.choice(options)


def build_why_this_car(model: str) -> str:
    if "Toyota" in model:
        return "Toyota reliability, low fuel consumption and family-friendly packaging made the choice easy"
    if "Lexus" in model:
        return "Wanted hybrid smoothness with premium comfort and predictable ownership cost"
    return "Needed a hybrid SUV that feels easy to own, efficient in winter and strong on resale"


def build_future_requirements(model: str) -> str:
    if "Toyota" in model:
        return "No plug needed, lower fuel cost, heated seats and reliable winter traction"
    return "Self-charging hybrid, quiet cabin, low running cost and strong resale value"


def build_customer_requirements(model: str) -> str:
    if "Toyota" in model:
        return "Needs a self-charging SUV without a new charging routine and wants Toyota dependability"
    return "Prefers no home charger dependency, simple maintenance cost and a calm family cabin"


def build_suggestions(model: str) -> str:
    if "Toyota" in model:
        return "Explain Toyota hybrid system benefits in plain Nordic family use cases and bundle winter package clearly"
    return "Show running-cost math against diesel SUVs and explain winter efficiency with simple charts"


def build_price_perception(model: str) -> str:
    if "Toyota" in model:
        return "Higher upfront, but justified by reliability, fuel savings and resale value"
    return "Feels reasonable if running cost stays low and ownership remains predictable"


def build_evaluation(model: str) -> str:
    if "Toyota" in model:
        return "A dependable family hybrid that avoids charging hassle and feels easy to recommend"
    return "Feels like a measured, low-stress choice for families that want hybrid efficiency without complexity"


def build_row(randomizer: random.Random, index: int) -> list[str | int]:
    gender = randomizer.choice(["Male", "Female"])
    name = build_name(randomizer, gender)
    household_size = randomizer.choice([2, 3, 4, 4, 5])
    age = randomizer.randint(31, 58)
    model = weighted_choice(randomizer, PRIMARY_MODELS)
    region = randomizer.choice(REGIONS)
    email = f"{name.lower().replace(' ', '.')}@example.com"
    powertrain_preference = build_powertrain_preference(randomizer)
    purchase_type = randomizer.choice(PURCHASE_TYPES)
    children = randomizer.choice(CHILDREN_LABELS if household_size >= 3 else ["No children", "1 child"])
    dream_car = randomizer.choice([
        "Toyota RAV4 Hybrid Lounge",
        "Toyota Highlander Hybrid",
        "Toyota Corolla Cross GR Sport Hybrid",
        "Lexus NX 350h Takumi",
        "Honda CR-V e:HEV Advance",
    ])

    return [
        name,
        email,
        age,
        gender,
        randomizer.choice(MARITAL_STATUS),
        household_size,
        randomizer.choice(EDUCATION_LEVELS),
        region,
        children,
        randomizer.choice(OCCUPATIONS),
        build_income(randomizer),
        "",
        build_hobbies(randomizer),
        build_locations(randomizer),
        randomizer.choice(SPENDING_PHILOSOPHY),
        build_daily_life_pattern(randomizer, household_size),
        randomizer.choice(PRODUCT_PREFERENCE),
        randomizer.choice(SOCIAL_CIRCLES),
        randomizer.choice(PERSONALITIES),
        f"Yes – {model}",
        build_why_this_car(model),
        randomizer.choice([
            "Cabin can feel tight with winter gear",
            "Road noise is noticeable on rough asphalt",
            "Rear-seat space is only just enough with child seats",
            "Would like clearer digital HVAC shortcuts",
        ]),
        dream_car,
        purchase_type,
        build_driving_scenarios(randomizer),
        build_future_requirements(model),
        randomizer.choice(USAGE_FREQUENCY),
        powertrain_preference,
        build_feature_triplet(model),
        build_complaint_triplet(randomizer),
        build_customer_requirements(model),
        build_suggestions(model),
        randomizer.choice(INFORMATION_SOURCES),
        build_price_perception(model),
        build_evaluation(model),
        randomizer.choice([
            "Reliable and calm",
            "Hybrid without friction",
            "Toyota just makes sense for this stage of life",
            "We want the family car to stay simple",
        ]),
        "",
    ]


def autosize_columns(worksheet) -> None:
    for index, header in enumerate(HEADERS, start=1):
        max_length = len(header)
        for cell in worksheet[get_column_letter(index)]:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 42)


def write_workbook(output_path: Path, rows: int, seed: int) -> None:
    randomizer = random.Random(seed)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "VOC Data"
    worksheet.append(HEADERS)

    header_fill = PatternFill(fill_type="solid", fgColor="103A5E")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index in range(rows):
        worksheet.append(build_row(randomizer, row_index))

    worksheet.freeze_panes = "A2"
    autosize_columns(worksheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a Nordic HEV / Toyota customer benchmark workbook.")
    parser.add_argument("--rows", type=int, default=120, help="Number of synthetic respondents to generate.")
    parser.add_argument("--seed", type=int, default=42, help="Random seed for reproducible output.")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "01_RAW_DATA" / "VOC_Nordic_HEV_Toyota_Users.xlsx",
        help="Output workbook path.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    write_workbook(output_path=args.output, rows=args.rows, seed=args.seed)
    print(f"Generated {args.output} with {args.rows} rows")


if __name__ == "__main__":
    main()