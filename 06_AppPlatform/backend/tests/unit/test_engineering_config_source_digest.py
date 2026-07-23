from pathlib import Path
import sys

import pandas as pd
import pytest
from openpyxl import Workbook

from app.services import engineering_config_source_digest as source_digest
from app.services.engineering_config_source_digest import build_source_digest, build_workbook_digest_from_frames


def _pdf_text(value: str) -> str:
    return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _write_text_pdf(path: Path, lines: list[str]) -> None:
    commands = ["BT /F1 12 Tf 14 TL 72 720 Td"]
    for index, line in enumerate(lines):
        if index:
            commands.append("T*")
        commands.append(f"({_pdf_text(line)}) Tj")
    commands.append("ET")
    stream = " ".join(commands).encode("utf-8")
    objects = [
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n",
        b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n",
        b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj\n",
        b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n",
        b"5 0 obj << /Length " + str(len(stream)).encode("ascii") + b" >> stream\n" + stream + b"\nendstream endobj\n",
    ]
    content = b"%PDF-1.4\n"
    offsets = []
    for item in objects:
        offsets.append(len(content))
        content += item
    xref_offset = len(content)
    content += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode("ascii")
    for offset in offsets:
        content += f"{offset:010d} 00000 n \n".encode("ascii")
    content += (
        f"trailer << /Root 1 0 R /Size {len(objects) + 1} >>\n"
        f"startxref\n{xref_offset}\n%%EOF"
    ).encode("ascii")
    path.write_bytes(content)


def test_split_pdf_table_line_preserves_internal_empty_cells() -> None:
    row = source_digest._split_pdf_table_line("Feature | Basic | | Premium")

    assert row == ["Feature", "Basic", "", "Premium"]


def test_paddleocr_geometry_rebuilds_table_rows() -> None:
    result = {
        "rec_texts": [
            "Category", "Feature", "Basic", "Premium",
            "Comfort", "Seat heat", "-", "S",
            "ADAS", "Camera", "Rear", "360",
        ],
        "rec_boxes": [
            [0, 0, 50, 10], [80, 0, 140, 10], [180, 0, 230, 10], [260, 0, 330, 10],
            [0, 30, 60, 40], [80, 30, 160, 40], [180, 30, 190, 40], [260, 30, 270, 40],
            [0, 60, 50, 70], [80, 60, 130, 70], [180, 60, 230, 70], [260, 60, 300, 70],
        ],
    }

    assert source_digest._collect_paddleocr_text_lines(result) == [
        "Category | Feature | Basic | Premium",
        "Comfort | Seat heat | - | S",
        "ADAS | Camera | Rear | 360",
    ]


def test_paddleocr_geometry_rebuilds_table_rows_from_mapping_like_result() -> None:
    class MappingLikeOcrResult:
        def __init__(self) -> None:
            self.payload = {
                "rec_texts": [
                    "Drive form", "FWD", "FWD", "FWD",
                    "Number of", "seats", "5", "5",
                    "Exterior", "package", "-", "S",
                ],
                "rec_boxes": [
                    [0, 0, 90, 16], [150, 0, 210, 16], [250, 0, 310, 16], [350, 0, 410, 16],
                    [0, 40, 110, 56], [150, 40, 220, 56], [250, 40, 270, 56], [350, 40, 370, 56],
                    [0, 80, 80, 96], [150, 80, 240, 96], [250, 80, 265, 96], [350, 80, 365, 96],
                ],
            }

        def get(self, key: str, default: object = None) -> object:
            return self.payload.get(key, default)

    assert source_digest._collect_paddleocr_text_lines(MappingLikeOcrResult()) == [
        "Drive form | FWD | FWD | FWD",
        "Number of | seats | 5 | 5",
        "Exterior | package | - | S",
    ]


def test_workbook_digest_detects_same_model_row_variant_compare_group() -> None:
    frame = pd.DataFrame(
        [
            ["T19CFL EV", "", "", "", "", "", ""],
            ["", "", "", "", "", "", ""],
            ["", "整备质量 (Kg)", "", "", "行驶质量 (Kg)", "", ""],
            ["Type", "Alex 1", "Alex 2", "Total", "Alex 1", "Alex 2", "Total"],
            ["FL EV 左舵-option1", 900, 810, 1710, 935, 850, 1785],
            ["FL EV 左舵-option2", 911, 831, 1742, 946, 871, 1817],
        ]
    )

    digest = build_workbook_digest_from_frames(
        [("备注", frame)],
        file_name="T19CFL EV Basic Parameters(04.30).xlsx",
    )

    assert digest["status"] == "ready"
    assert digest["modelName"] == "T19CFL EV"
    assert digest["summary"]["comparableGroupCount"] == 1
    assert digest["summary"]["candidateTrimCount"] == 2

    group = digest["compareGroups"][0]
    assert group["modelName"] == "T19CFL EV"
    assert group["trimCount"] == 2
    assert [trim["trimName"] for trim in group["trims"]] == [
        "FL EV 左舵-option1",
        "FL EV 左舵-option2",
    ]

    rows_by_feature = {row["featureName"]: row for row in group["rows"]}
    assert rows_by_feature["整备质量 (Kg) / Total"]["comparisonType"] == "DIFFERENT_VALUE"
    assert [
        value["rawValue"] if value else None
        for value in rows_by_feature["整备质量 (Kg) / Total"]["values"]
    ] == ["1710", "1742"]


def test_workbook_digest_uses_shared_value_normalization_for_unit_suffixes() -> None:
    frame = pd.DataFrame(
        [
            ["Category", "Feature", "Basic", "Premium"],
            ["Charging", "DC最大充电功率（千瓦）", "70kW", "70"],
        ]
    )

    digest = build_workbook_digest_from_frames(
        [("Specs", frame)],
        file_name="rival-config.xlsx",
    )

    group = digest["compareGroups"][0]
    row = group["rows"][0]
    assert row["featureName"] == "DC最大充电功率（千瓦）"
    assert row["comparisonType"] == "COMMON_SAME"
    assert [value["displayValue"] for value in row["values"]] == ["70", "70"]


def test_workbook_digest_detects_eu_config_resource_table_without_remark_columns() -> None:
    index_frame = pd.DataFrame(
        [
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["Serial number", "车型 Code name", "品牌 Brand", "配置 Configuration", "物料号 BOM"],
            [1, "T19C MY ICE", "OMODA", "Comfort-FWD", "6150001"],
        ]
    )
    model_frame = pd.DataFrame(
        [
            ["The Specification of Omoda 5 T19C MY2025", "", "", "", "", "", ""],
            ["", "", "", "", "", "", ""],
            ["基本参数", "国家", "Country", "欧盟(23+3)\n舒适型-升级版", "", "欧盟(23+3)\n尊贵型-升级版", "删除或减少"],
            ["", "配置版型", "Configuration version", "两驱基本型\nBasic-FWD", "两驱舒适型\nComfort-FWD", "两驱尊贵型\nPremium-FWD", "备注"],
            ["", "物料号", "Material NO.", "6150001", "6150002", "6150003", ""],
            ["基本参数", "长宽高", "L/W/H", "4400/1830/1588", "4400/1830/1588", "4400/1830/1588", "no"],
            ["安全", "侧气帘", "Side curtain airbag", "—", "S", "S", "减少"],
            ["舒适", "座椅加热", "Seat heating", "", "O", "S", "新增"],
        ]
    )

    digest = build_workbook_digest_from_frames(
        [("总表", index_frame), ("T19C MY ICE ", model_frame), ("配置修订记录", pd.DataFrame([["x"]]))],
        file_name="欧盟在售车型可控资源表20260226.xlsx",
    )

    assert digest["status"] == "ready"
    assert digest["workbookFormat"] == "eu_config_resource_table"
    assert digest["summary"]["sheetCount"] == 3
    assert digest["summary"]["comparableGroupCount"] == 1
    assert [group["sourceSheet"] for group in digest["compareGroups"]] == ["T19C MY ICE "]

    group = digest["compareGroups"][0]
    assert group["trimCount"] == 3
    assert [trim["trimName"] for trim in group["trims"]] == [
        "两驱基本型 Basic-FWD",
        "两驱舒适型 Comfort-FWD",
        "两驱尊贵型 Premium-FWD",
    ]
    assert [trim["materialNo"] for trim in group["trims"]] == ["6150001", "6150002", "6150003"]
    assert all(trim["trimName"] != "备注" for trim in group["trims"])
    assert "总表" not in {group["modelName"] for group in digest["compareGroups"]}
    assert group["summary"]["confirmedDifferenceCount"] == 2
    assert group["summary"]["rawConfirmedDifferenceCount"] == 1
    assert group["summary"]["inferredDifferenceCount"] == 1

    rows_by_feature = {row["featureName"]: row for row in group["rows"]}
    side_curtain = rows_by_feature["Side curtain airbag / 侧气帘"]
    assert side_curtain["comparisonType"] == "UNIQUE_OR_PARTIAL"
    assert [value["valueState"] for value in side_curtain["values"]] == ["marker_value", "marker_value", "marker_value"]
    assert [value["availability"] for value in side_curtain["values"]] == ["NOT_AVAILABLE", "STANDARD", "STANDARD"]
    seat_heating = rows_by_feature["Seat heating / 座椅加热"]
    assert seat_heating["comparisonType"] == "OPTIONAL_DIFFERENT"
    assert [value["valueState"] for value in seat_heating["values"]] == ["blank", "marker_value", "marker_value"]
    assert [value["availability"] for value in seat_heating["values"]] == ["NOT_AVAILABLE", "OPTIONAL", "STANDARD"]
    assert seat_heating["values"][0]["inferred"] is True
    assert seat_heating["values"][0]["displayValue"] == "不配备*"


def test_eu_config_digest_reads_full_sheet_and_keeps_profile_out_of_features() -> None:
    index_frame = pd.DataFrame(
        [
            ["Serial number", "车型 Code name", "品牌 Brand", "配置 Configuration", "物料号 BOM"],
            [1, "T19C-BEV", "OMODA", "Comfort-FWD", "T7000WU**MY0001"],
            [2, "T19C-BEV", "OMODA", "Premium-FWD", "T7000WU**MY0002"],
        ]
    )
    rows = [
        ["The Specification of Omoda 5 BEV", "", "", "", "", "更新时间：20260226"],
        ["The configuration is for reference only.", "", "", "", "", ""],
        ["基本参数", "国家", "Country", "欧盟(23+3)", "", "删除或减少"],
        ["", "配置版型", "Configuration version", "两驱长续航舒适型\nComfort-FWD", "两驱长续航尊贵型\nPremium-FWD", "备注"],
        ["", "物料号", "Material No.", "T7000WU**MY0001", "T7000WU**MY0002", ""],
        ["", "家族识别码", "Interpolation family identifier", "IP-BEV-COMFORT", "IP-BEV-PREMIUM", "WVTA"],
        ["", "版本号", "Variant/Version", "L3E3308", "L3E3309", "WVTA"],
    ]
    late_features = {
        150: ("舒适便利", "外后视镜电动折叠", "Power folding outside rearview mirror", "●", "●"),
        162: ("舒适便利", "热泵空调", "Heat pump A/C", "", "●"),
        173: ("舒适便利", "电动尾门", "Power tailgate", "", "●"),
        186: ("科技配置", "8扬声器（品牌音响）", "8 speakers (branded audio)", "", "●"),
        198: ("科技配置", "品牌音响（sony）", "Branded audio (sony)", "", "●"),
    }
    for index in range(205):
        category, zh, en, comfort_value, premium_value = late_features.get(
            index,
            (
                "驾驶辅助" if index < 110 else "舒适便利" if index < 170 else "科技配置",
                f"配置项{index + 1}",
                f"Feature {index + 1}",
                "●",
                "●" if index % 9 else "",
            ),
        )
        rows.append([category if index % 12 == 0 else "", zh, en, comfort_value, premium_value, ""])

    digest = build_workbook_digest_from_frames(
        [("总表", index_frame), ("T19C-BEV（2025款）", pd.DataFrame(rows))],
        file_name="欧盟在售车型可控资源表20260226.xlsx",
    )

    group = digest["compareGroups"][0]

    assert group["trimCount"] == 2
    assert group["featureCount"] == 205
    assert group["summary"]["totalFeatures"] == 205
    assert group["summary"]["categoryCounts"]["舒适便利"] > 0
    assert group["summary"]["categoryCounts"]["科技配置"] > 0
    assert [trim["materialNo"] for trim in group["trims"]] == ["T7000WU**MY0001", "T7000WU**MY0002"]
    assert [trim["profile"]["variantVersion"] for trim in group["trims"]] == ["L3E3308", "L3E3309"]

    rows_by_feature = {row["featureName"]: row for row in group["rows"]}
    assert "Country / 国家" not in rows_by_feature
    assert "Configuration version / 配置版型" not in rows_by_feature
    assert "Material No. / 物料号" not in rows_by_feature
    assert "Heat pump A/C / 热泵空调" in rows_by_feature
    assert "Power tailgate / 电动尾门" in rows_by_feature
    assert "8 speakers (branded audio) / 8扬声器（品牌音响）" in rows_by_feature
    assert "Branded audio (sony) / 品牌音响（sony）" in rows_by_feature
    heat_pump = rows_by_feature["Heat pump A/C / 热泵空调"]
    assert heat_pump["comparisonType"] == "UNIQUE_OR_PARTIAL"
    assert [value["availability"] for value in heat_pump["values"]] == ["NOT_AVAILABLE", "STANDARD"]
    assert heat_pump["values"][0]["inferenceReason"] == "blank_as_not_equipped_by_eu_matrix_policy"
    assert rows_by_feature["Heat pump A/C / 热泵空调"]["featureKey"].startswith(
        "T19C-BEV（2025款）|舒适便利|热泵空调|Heat pump A/C"
    )


def test_source_digest_expands_merged_cells_for_common_trim_values(tmp_path) -> None:
    workbook = Workbook()
    index_sheet = workbook.active
    index_sheet.title = "总表"
    index_sheet.append(["Serial number", "车型 Code name", "品牌 Brand", "配置 Configuration", "物料号 BOM"])
    index_sheet.append([1, "T19C MY ICE", "OMODA", "Basic-FWD", "T71607V**MM0001"])

    sheet = workbook.create_sheet("T19C MY ICE ")
    sheet.append(["The Specification of Omoda 5", "", "", "", "", "", ""])
    sheet.append(["The configuration is for reference only.", "", "", "", "", "", ""])
    sheet.append(["基本参数", "国家", "Country", "欧盟(23+3)", "欧盟(23+3)", "欧盟(23+3)", "来源"])
    sheet.append(["", "配置版型", "Configuration version", "两驱基本型 Basic-FWD", "两驱舒适型 Comfort-FWD", "两驱尊贵型 Premium-FWD", "备注"])
    sheet.append(["", "物料号", "Material No.", "T71607V**MM0001", "T71607V**MM0002", "T71607V**MM0003", ""])
    sheet.append(["", "家族识别码", "Interpolation family identifier", "ICE-BASIC", "ICE-COMFORT", "ICE-PREMIUM", ""])
    sheet.append(["", "版本号", "Variant/Version", "L3E3001", "L3E3002", "L3E3003", ""])
    sheet.append(["基本参数", "座椅", "Number of seats", "", "", "", "WVTA"])
    sheet.merge_cells("D8:F8")
    sheet["D8"] = "5"
    sheet.append(["尺寸和重量", "长×宽×高 mm", "Length*Width*Height (mm)", "", "", "", "WVTA"])
    sheet.merge_cells("D9:F9")
    sheet["D9"] = "4447*1824*1588"
    sheet.append(["", "油箱 L", "Fuel tank L", "", "55", "", "WVTA"])
    sheet.append(["舒适便利", "座椅加热", "Seat heating", "", "●", "●", "配置表"])

    source_path = tmp_path / "merged-common-values.xlsx"
    workbook.save(source_path)
    workbook.close()

    digest = build_source_digest(source_path, source_path.name)
    assert digest is not None
    assert digest["workbookFormat"] == "eu_config_resource_table"
    group = digest["compareGroups"][0]
    rows_by_feature = {row["featureName"]: row for row in group["rows"]}

    seats = rows_by_feature["Number of seats / 座椅"]
    assert seats["comparisonType"] == "COMMON_SAME"
    assert [value["rawValue"] if value else None for value in seats["values"]] == ["5", "5", "5"]
    assert [value["valueState"] for value in seats["values"]] == ["numeric_value", "numeric_value", "numeric_value"]
    assert seats["values"][1]["source"]["cell"] == "E8"
    assert seats["values"][1]["source"]["sourceCell"] == "D8"
    assert seats["values"][1]["source"]["mergedRange"] == "D8:F8"

    dimensions = rows_by_feature["Length*Width*Height (mm) / 长×宽×高 mm"]
    assert dimensions["comparisonType"] == "COMMON_SAME"
    assert [value["rawValue"] if value else None for value in dimensions["values"]] == [
        "4447*1824*1588",
        "4447*1824*1588",
        "4447*1824*1588",
    ]
    fuel_tank = rows_by_feature["Fuel tank L / 油箱 L"]
    assert fuel_tank["comparisonType"] == "MISSING_UNKNOWN"
    assert [value["availability"] for value in fuel_tank["values"]] == ["UNKNOWN", "VALUE", "UNKNOWN"]
    assert all(not value["inferred"] for value in fuel_tank["values"])

    seat_heating = rows_by_feature["Seat heating / 座椅加热"]
    assert seat_heating["comparisonType"] == "UNIQUE_OR_PARTIAL"
    assert seat_heating["values"][0]["valueState"] == "blank"
    assert seat_heating["values"][0]["availability"] == "NOT_AVAILABLE"
    assert seat_heating["values"][0]["displayValue"] == "不配备*"
    assert seat_heating["values"][0]["source"]["inferenceReason"] == "blank_as_not_equipped_by_eu_matrix_policy"
    assert group["summary"]["confirmedDifferenceCount"] == 1
    assert group["summary"]["rawConfirmedDifferenceCount"] == 0
    assert group["summary"]["inferredDifferenceCount"] == 1


def test_tabular_csv_digest_detects_simple_config_columns(tmp_path) -> None:
    source_path = tmp_path / "rival-config.csv"
    source_path.write_text(
        "\n".join(
            [
                "Category,Feature,Basic,Premium",
                "Comfort,Seat heating,-,S",
                "Safety,Blind spot,,S",
                "Infotainment,Speaker count,6,8",
            ]
        ),
        encoding="utf-8",
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "tabular"
    assert digest["summary"]["comparableGroupCount"] == 1
    group = digest["compareGroups"][0]
    assert group["trimCount"] == 2
    assert [trim["trimName"] for trim in group["trims"]] == ["Basic", "Premium"]
    rows_by_feature = {row["featureName"]: row for row in group["rows"]}
    assert rows_by_feature["Seat heating"]["comparisonType"] == "UNIQUE_OR_PARTIAL"
    assert [value["availability"] for value in rows_by_feature["Seat heating"]["values"]] == ["NOT_AVAILABLE", "STANDARD"]
    assert rows_by_feature["Blind spot"]["comparisonType"] == "MISSING_UNKNOWN"
    assert rows_by_feature["Speaker count"]["comparisonType"] == "DIFFERENT_VALUE"


def test_tabular_tsv_digest_detects_simple_config_columns(tmp_path) -> None:
    source_path = tmp_path / "rival-config.tsv"
    source_path.write_text(
        "\n".join(
            [
                "Category\tFeature\tBasic\tPremium",
                "Comfort\tSeat heating\t-\tS",
                "Safety\tBlind spot\t\tS",
                "Infotainment\tSpeaker count\t6\t8",
            ]
        ),
        encoding="utf-8",
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "tabular"
    assert digest["summary"]["comparableGroupCount"] == 1
    group = digest["compareGroups"][0]
    assert group["trimCount"] == 2
    assert [trim["trimName"] for trim in group["trims"]] == ["Basic", "Premium"]
    rows_by_feature = {row["featureName"]: row for row in group["rows"]}
    assert rows_by_feature["Seat heating"]["comparisonType"] == "UNIQUE_OR_PARTIAL"
    assert [value["availability"] for value in rows_by_feature["Seat heating"]["values"]] == ["NOT_AVAILABLE", "STANDARD"]
    assert rows_by_feature["Blind spot"]["comparisonType"] == "MISSING_UNKNOWN"
    assert rows_by_feature["Speaker count"]["comparisonType"] == "DIFFERENT_VALUE"


def test_tabular_csv_digest_detects_price_list_rows_as_trim_columns(tmp_path) -> None:
    source_path = tmp_path / "competitor-price-list.csv"
    source_path.write_text(
        "\n".join(
            [
                "Brand,Model,Trim,Market,Model Year,Powertrain,MSRP,Currency",
                "OMODA,T19C,Basic,Germany,2026,ICE,23000,EUR",
                "OMODA,T19C,Comfort,Germany,2026,ICE,25000,EUR",
                "OMODA,T19C,Premium,Germany,2026,ICE,28000,EUR",
                "OMODA,T19C-BEV,Comfort,Germany,2026,BEV,31000,EUR",
                "OMODA,T19C-BEV,Premium,Germany,2026,BEV,35000,EUR",
            ]
        ),
        encoding="utf-8",
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "tabular"
    assert digest["summary"]["comparableGroupCount"] == 2
    assert digest["summary"]["candidateTrimCount"] == 5

    ice_group = digest["compareGroups"][0]
    assert ice_group["sourceKind"] == "price_list"
    assert ice_group["modelName"] == "T19C"
    assert ice_group["title"] == "OMODA / T19C / Germany / 2026 / ICE / 价格单"
    assert [trim["trimName"] for trim in ice_group["trims"]] == ["Basic", "Comfort", "Premium"]
    assert [trim["market"] for trim in ice_group["trims"]] == ["Germany", "Germany", "Germany"]
    assert [trim["modelYear"] for trim in ice_group["trims"]] == ["2026", "2026", "2026"]
    assert [trim["energyType"] for trim in ice_group["trims"]] == ["ICE", "ICE", "ICE"]
    assert all(trim["dataOrigin"] == "external_or_scraped" for trim in ice_group["trims"])

    rows_by_feature = {row["featureName"]: row for row in ice_group["rows"]}
    assert rows_by_feature["MSRP"]["category"] == "价格 Pricing"
    assert rows_by_feature["MSRP"]["comparisonType"] == "DIFFERENT_VALUE"
    assert [value["rawValue"] for value in rows_by_feature["MSRP"]["values"]] == ["23000", "25000", "28000"]
    assert [value["availability"] for value in rows_by_feature["MSRP"]["values"]] == ["VALUE", "VALUE", "VALUE"]
    assert rows_by_feature["Currency"]["comparisonType"] == "COMMON_SAME"
    assert rows_by_feature["MSRP"]["values"][0]["source"]["cell"] == "G2"


def test_workbook_price_list_keeps_powertrain_groups_separate(tmp_path) -> None:
    workbook = Workbook()
    intro_sheet = workbook.active
    intro_sheet.title = "Readme"
    intro_sheet.append(["Downloaded competitor price list"])
    price_sheet = workbook.create_sheet("Germany Price List")
    price_sheet.append(["OMODA Germany public price list"])
    price_sheet.append([])
    price_sheet.append(["Brand", "Model", "Trim", "Market", "Model Year", "Powertrain", "MSRP", "Currency"])
    price_sheet.append(["OMODA", "T19C", "Basic ICE", "Germany", "2026", "ICE", 23000, "EUR"])
    price_sheet.append(["OMODA", "T19C", "Premium ICE", "Germany", "2026", "ICE", 28000, "EUR"])
    price_sheet.append(["OMODA", "T19C", "Basic BEV", "Germany", "2026", "BEV", 31000, "EUR"])
    price_sheet.append(["OMODA", "T19C", "Premium BEV", "Germany", "2026", "BEV", 35000, "EUR"])

    source_path = tmp_path / "competitor-price-list.xlsx"
    workbook.save(source_path)
    workbook.close()

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "workbook"
    assert digest["summary"]["comparableGroupCount"] == 2
    assert digest["summary"]["candidateTrimCount"] == 4
    groups_by_powertrain = {
        group["trims"][0]["energyType"]: group
        for group in digest["compareGroups"]
    }
    assert set(groups_by_powertrain) == {"ICE", "BEV"}

    ice_group = groups_by_powertrain["ICE"]
    bev_group = groups_by_powertrain["BEV"]
    assert ice_group["sourceKind"] == "price_list"
    assert bev_group["sourceKind"] == "price_list"
    assert ice_group["title"] == "OMODA / T19C / Germany / 2026 / ICE / 价格单"
    assert bev_group["title"] == "OMODA / T19C / Germany / 2026 / BEV / 价格单"
    assert [trim["trimName"] for trim in ice_group["trims"]] == ["Basic ICE", "Premium ICE"]
    assert [trim["trimName"] for trim in bev_group["trims"]] == ["Basic BEV", "Premium BEV"]

    ice_msrp = {row["featureName"]: row for row in ice_group["rows"]}["MSRP"]
    bev_msrp = {row["featureName"]: row for row in bev_group["rows"]}["MSRP"]
    assert [value["rawValue"] for value in ice_msrp["values"]] == ["23000", "28000"]
    assert [value["rawValue"] for value in bev_msrp["values"]] == ["31000", "35000"]
    assert ice_msrp["featureKey"] != bev_msrp["featureKey"]
    assert ice_msrp["values"][0]["source"]["cell"] == "G4"
    assert bev_msrp["values"][0]["source"]["cell"] == "G6"


def test_tabular_html_digest_detects_simple_config_columns(tmp_path) -> None:
    source_path = tmp_path / "rival-config.html"
    source_path.write_text(
        """
        <table>
          <tr><td>Feature</td><td>Comfort</td><td>Premium</td></tr>
          <tr><td>360 camera</td><td>-</td><td>S</td></tr>
          <tr><td>Wireless charging</td><td></td><td>O</td></tr>
        </table>
        """,
        encoding="utf-8",
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "tabular"
    group = digest["compareGroups"][0]
    assert [trim["trimName"] for trim in group["trims"]] == ["Comfort", "Premium"]
    rows_by_feature = {row["featureName"]: row for row in group["rows"]}
    assert rows_by_feature["360 camera"]["comparisonType"] == "UNIQUE_OR_PARTIAL"
    assert [value["displayValue"] for value in rows_by_feature["Wireless charging"]["values"]] == ["待确认", "选装"]


def test_pdf_text_digest_detects_simple_config_columns(tmp_path) -> None:
    source_path = tmp_path / "rival-config.pdf"
    _write_text_pdf(
        source_path,
        [
            "Rival C-SUV configuration sheet",
            "Category | Feature | Basic | Premium",
            "Comfort | Seat heating | - | S",
            "Safety | 360 camera | - | O",
            "Infotainment | Speaker count | 6 | 8",
        ],
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "pdf_text"
    assert digest["sourceFormat"] == "pdf_text"
    assert digest["status"] == "ready"
    assert digest["summary"]["comparableGroupCount"] == 1
    group = digest["compareGroups"][0]
    assert [trim["trimName"] for trim in group["trims"]] == ["Basic", "Premium"]
    rows_by_feature = {row["featureName"]: row for row in group["rows"]}
    assert rows_by_feature["Seat heating"]["comparisonType"] == "UNIQUE_OR_PARTIAL"
    assert [value["availability"] for value in rows_by_feature["Seat heating"]["values"]] == ["NOT_AVAILABLE", "STANDARD"]
    assert rows_by_feature["360 camera"]["values"][1]["displayValue"] == "选装"
    assert rows_by_feature["Speaker count"]["comparisonType"] == "DIFFERENT_VALUE"
    source = rows_by_feature["Seat heating"]["values"][0]["source"]
    assert source["sourceType"] == "pdf_text"
    assert source["pageNumber"] == 1
    assert source["cell"] == "P1R3C3"


def test_pdf_table_parser_preserves_grade_prices_and_page_location() -> None:
    table = [
        [
            "Motor",
            "kW/PS",
            "Antrieb",
            "Getriebe",
            "Sportage",
            "Power Ed.",
            "Style",
            "Black Ed.",
            "GT-Line",
        ],
        [
            "1.6 T-GDi PHEV (Benzin)",
            "212/288",
            "4×4",
            "Automat",
            "47’750.-",
            "51’950.-",
            "56’050.-",
            "–",
            "59’950.-",
        ],
    ]

    parsed = source_digest._pdf_table_standard_rows(
        table,
        page_number=1,
        table_number=1,
    )

    assert parsed is not None
    variants, rows = parsed
    assert variants == ("Sportage", "Power Ed.", "Style", "Black Ed.", "GT-Line")
    assert rows[0][0][1] == "MSRP / 1.6 T-GDi PHEV (Benzin) / 212/288 / 4×4 / Automat"
    assert rows[0][0][2:] == ["47’750.-", "51’950.-", "56’050.-", "–", "59’950.-"]
    assert rows[0][1][0]["pageNumber"] == 1
    assert rows[0][1][0]["sourceCell"] == "P1T1R2C5"


def test_pdf_table_parser_repairs_missing_grade_prices_from_page_text() -> None:
    table = [
        [
            "Motor",
            "kW/PS",
            "Antrieb",
            "Getriebe",
            "Sportage",
            "Power Ed.",
            "Style",
            "Black Ed.",
            "GT-Line",
        ],
        [
            None,
            "212/288",
            "4×4",
            "Automat",
            "47’750.-",
            "51’950.-",
            "56’050.-",
            "–",
            None,
        ],
    ]
    page_text = (
        "1.6 T-GDi PHEV (Benzin) 212/288 4×4 Automat "
        "47’750.- 51’950.- 56’050.- – 59’950.-"
    )

    parsed = source_digest._pdf_table_standard_rows(
        table,
        page_number=1,
        table_number=1,
        page_text=page_text,
    )

    assert parsed is not None
    _variants, rows = parsed
    assert rows[0][0][1] == "MSRP / 1.6 T-GDi PHEV (Benzin) 212/288 4×4 Automat"
    assert rows[0][0][2:] == ["47’750.-", "51’950.-", "56’050.-", "–", "59’950.-"]
    assert rows[0][1][-1]["pageNumber"] == 1


def test_pdf_table_parser_reuses_grade_header_for_equipment() -> None:
    table = [
        ["Komfort", "Sportage", "Power Ed.", "Style", "Black Ed.", "GT-Line"],
        ["Heckklappe elektrisch", "-", "-", "●", "●", "●"],
        ["Lenkradheizung", "●", "●", "●", "●", "●"],
    ]

    parsed = source_digest._pdf_table_standard_rows(
        table,
        page_number=3,
        table_number=3,
    )

    assert parsed is not None
    variants, rows = parsed
    assert variants[-1] == "GT-Line"
    assert rows[0][0] == ["Komfort", "Heckklappe elektrisch", "-", "-", "●", "●", "●"]
    assert rows[0][1][-1]["pageNumber"] == 3


def test_pdf_text_digest_without_table_stays_pending(tmp_path) -> None:
    source_path = tmp_path / "scanned-or-narrative.pdf"
    _write_text_pdf(
        source_path,
        [
            "This PDF has narrative text only.",
            "It does not contain enough repeated table columns.",
        ],
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "pdf_text"
    assert digest["status"] == "pending"
    assert digest["summary"]["comparableGroupCount"] == 0
    assert digest["compareGroups"] == []


def test_pdf_ocr_digest_detects_simple_config_columns(tmp_path, monkeypatch) -> None:
    source_path = tmp_path / "scanned-config.pdf"
    _write_text_pdf(source_path, [])
    page_image = tmp_path / "page-1.png"
    page_image.write_bytes(b"\x89PNG\r\n\x1a\n")
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._render_pdf_pages_for_ocr",
        lambda _path, _output_dir: ([page_image], None),
    )
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._extract_ocr_text_candidates",
        lambda _path, compare_engines=False: (
            [
                (
                    "\n".join(
                        [
                            "Category | Feature | Basic | Premium",
                            "Comfort | Seat heating | - | S",
                            "Safety | 360 camera | - | O",
                        ]
                    ),
                    "unit_pdf_ocr",
                )
            ],
            [],
        ),
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "pdf_ocr"
    assert digest["sourceFormat"] == "pdf_ocr"
    assert digest["ocrEngine"] == "unit_pdf_ocr"
    assert digest["ocrEngineCandidates"][0]["engine"] == "unit_pdf_ocr"
    assert digest["ocrEngineCandidates"][0]["selected"] is True
    assert digest["ocrEngineCandidates"][0]["score"]["rowCount"] == 3
    assert digest["ocrEngineCandidates"][0]["lineCount"] == 3
    assert "Category | Feature | Basic | Premium" in digest["ocrEngineCandidates"][0]["textPreview"]
    assert digest["status"] == "ready"
    group = digest["compareGroups"][0]
    assert [trim["trimName"] for trim in group["trims"]] == ["Basic", "Premium"]
    rows_by_feature = {row["featureName"]: row for row in group["rows"]}
    assert rows_by_feature["Seat heating"]["comparisonType"] == "UNIQUE_OR_PARTIAL"
    source = rows_by_feature["Seat heating"]["values"][0]["source"]
    assert source["sourceType"] == "pdf_ocr"
    assert source["ocrEngine"] == "unit_pdf_ocr"
    assert source["cell"] == "P1OCRR2C3"
    assert source["pageNumber"] == 1


def test_pdf_ocr_digest_chooses_best_engine_by_table_score(tmp_path, monkeypatch) -> None:
    source_path = tmp_path / "scanned-config.pdf"
    _write_text_pdf(source_path, [])
    page_image = tmp_path / "page-1.png"
    page_image.write_bytes(b"\x89PNG\r\n\x1a\n")
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._render_pdf_pages_for_ocr",
        lambda _path, _output_dir: ([page_image], None),
    )
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._extract_ocr_text_candidates",
        lambda _path, compare_engines=False: (
            [
                ("Seat heating Basic Premium", "legacy_pdf_ocr"),
                (
                    "\n".join(
                        [
                            "Category | Feature | Basic | Premium",
                            "Comfort | Seat heating | - | S",
                            "Safety | 360 camera | - | O",
                        ]
                    ),
                    "paddleocr",
                ),
            ],
            [],
        ),
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "pdf_ocr"
    assert digest["ocrEngine"] == "paddleocr"
    assert digest["status"] == "ready"
    assert digest["ocrEvaluation"]["strategy"] == "highest_config_semantic_score"
    assert digest["ocrEvaluation"]["candidateCount"] == 2
    assert digest["ocrEvaluation"]["selectedEngine"] == "paddleocr"
    assert digest["ocrEvaluation"]["selectedScore"]["columnCount"] == 4
    assert digest["ocrEvaluation"]["selectedReasonDetails"][0] == "paddleocr 识别到可比配置表；legacy_pdf_ocr 未形成可比配置表。"
    assert "paddleocr 选中结果：配置项 2、配置列 2、差异 2" in digest["ocrEvaluation"]["selectedReasonDetails"][1]
    candidates = {candidate["engine"]: candidate for candidate in digest["ocrEngineCandidates"]}
    assert candidates["legacy_pdf_ocr"]["comparableTableDetected"] is False
    assert candidates["legacy_pdf_ocr"]["lineCount"] == 1
    assert candidates["legacy_pdf_ocr"]["textPreview"] == "Seat heating Basic Premium"
    assert candidates["paddleocr"]["selected"] is True
    assert candidates["paddleocr"]["lineCount"] == 3
    assert "Safety | 360 camera | - | O" in candidates["paddleocr"]["textPreview"]
    group = digest["compareGroups"][0]
    assert [trim["trimName"] for trim in group["trims"]] == ["Basic", "Premium"]


def test_pdf_ocr_digest_chooses_higher_scored_table_when_all_engines_detect_tables(tmp_path, monkeypatch) -> None:
    source_path = tmp_path / "scanned-config.pdf"
    _write_text_pdf(source_path, [])
    page_image = tmp_path / "page-1.png"
    page_image.write_bytes(b"\x89PNG\r\n\x1a\n")
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._render_pdf_pages_for_ocr",
        lambda _path, _output_dir: ([page_image], None),
    )
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._extract_ocr_text_candidates",
        lambda _path, compare_engines=False: (
            [
                (
                    "\n".join(
                        [
                            "Feature | Basic | Premium",
                            "Seat heating | - | S",
                            "360 camera | - | O",
                        ]
                    ),
                    "legacy_pdf_ocr",
                ),
                (
                    "\n".join(
                        [
                            "Category | Feature | Basic | Comfort | Premium",
                            "Comfort | Seat heating | S | O | S",
                            "Safety | 360 camera | R | - | O",
                            "Infotainment | Speaker count | 4 | 6 | 8",
                        ]
                    ),
                    "paddleocr",
                ),
            ],
            [],
        ),
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "pdf_ocr"
    assert digest["ocrEngine"] == "paddleocr"
    assert digest["ocrEvaluation"]["selectedEngine"] == "paddleocr"
    assert digest["ocrEvaluation"]["comparableCandidateCount"] == 2
    assert digest["ocrEvaluation"]["selectedScore"]["rowCount"] == 4
    assert digest["ocrEvaluation"]["selectedScore"]["columnCount"] == 5
    assert "paddleocr 配置项 3、配置列 3、差异 3" in digest["ocrEvaluation"]["selectedReasonDetails"][0]
    assert "legacy_pdf_ocr 配置项 2、配置列 2、差异 2" in digest["ocrEvaluation"]["selectedReasonDetails"][0]
    candidates = {candidate["engine"]: candidate for candidate in digest["ocrEngineCandidates"]}
    assert candidates["legacy_pdf_ocr"]["comparableTableDetected"] is True
    assert candidates["legacy_pdf_ocr"]["score"]["columnCount"] == 3
    assert candidates["paddleocr"]["selected"] is True
    group = digest["compareGroups"][0]
    assert [trim["trimName"] for trim in group["trims"]] == ["Comfort", "Premium"]
    assert candidates["legacy_pdf_ocr"]["selected"] is False
    assert any(
        value and value["source"]["ocrEngine"] == "paddleocr"
        for row in group["rows"]
        for value in row["values"]
    )


def test_pdf_ocr_digest_prefers_more_semantic_features_over_wider_noise_table(tmp_path, monkeypatch) -> None:
    source_path = tmp_path / "scanned-config.pdf"
    _write_text_pdf(source_path, [])
    page_image = tmp_path / "page-1.png"
    page_image.write_bytes(b"\x89PNG\r\n\x1a\n")
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._render_pdf_pages_for_ocr",
        lambda _path, _output_dir: ([page_image], None),
    )
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._extract_ocr_text_candidates",
        lambda _path, compare_engines=False: (
            [
                (
                    "\n".join(
                        [
                            "Category | Feature | Basic | Premium",
                            "Comfort | Seat heating | - | S",
                            "Safety | 360 camera | - | O",
                            "Exterior | Roof rack | - | S",
                            "Interior | Wireless charging | - | O",
                        ]
                    ),
                    "legacy_pdf_ocr",
                ),
                (
                    "\n".join(
                        [
                            "Category | Feature | Basic | Comfort | Premium | OCR note",
                            "Comfort | Seat heating | - | S | S | wide",
                            "Safety | 360 camera | - | O | O | wide",
                        ]
                    ),
                    "paddleocr",
                ),
            ],
            [],
        ),
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "pdf_ocr"
    assert digest["ocrEngine"] == "legacy_pdf_ocr"
    assert digest["ocrEvaluation"]["selectedEngine"] == "legacy_pdf_ocr"
    assert digest["ocrEvaluation"]["selectedScore"]["featureCount"] == 4
    candidates = {candidate["engine"]: candidate for candidate in digest["ocrEngineCandidates"]}
    assert candidates["legacy_pdf_ocr"]["score"]["featureCount"] == 4
    assert candidates["paddleocr"]["score"]["featureCount"] == 2
    assert candidates["paddleocr"]["score"]["columnCount"] == 6
    assert candidates["legacy_pdf_ocr"]["selected"] is True
    assert candidates["paddleocr"]["selected"] is False
    assert "legacy_pdf_ocr 配置项 4、配置列 2、差异 4" in digest["ocrEvaluation"]["selectedReasonDetails"][0]
    assert "paddleocr 配置项 2、配置列 4、差异 2" in digest["ocrEvaluation"]["selectedReasonDetails"][0]


def test_image_ocr_digest_detects_simple_config_columns(tmp_path, monkeypatch) -> None:
    source_path = tmp_path / "rival-config.jpg"
    source_path.write_bytes(b"\xff\xd8\xff")
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._extract_ocr_text_candidates",
        lambda _path, compare_engines=False: (
            [
                (
                    "\n".join(
                        [
                            "Category | Feature | Basic | Premium",
                            "Comfort | Seat heating | - | S",
                            "Safety | 360 camera | - | O",
                        ]
                    ),
                    "unit_ocr",
                )
            ],
            [],
        ),
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "image_ocr"
    assert digest["sourceFormat"] == "image_ocr"
    assert digest["ocrEngine"] == "unit_ocr"
    assert digest["ocrEngineCandidates"][0]["engine"] == "unit_ocr"
    assert digest["ocrEngineCandidates"][0]["selected"] is True
    assert digest["status"] == "ready"
    assert digest["summary"]["comparableGroupCount"] == 1
    group = digest["compareGroups"][0]
    assert [trim["trimName"] for trim in group["trims"]] == ["Basic", "Premium"]
    rows_by_feature = {row["featureName"]: row for row in group["rows"]}
    assert rows_by_feature["Seat heating"]["comparisonType"] == "UNIQUE_OR_PARTIAL"
    assert rows_by_feature["360 camera"]["values"][1]["displayValue"] == "选装"
    source = rows_by_feature["Seat heating"]["values"][0]["source"]
    assert source["sourceType"] == "image_ocr"
    assert source["ocrEngine"] == "unit_ocr"
    assert source["cell"] == "OCRR2C3"


def test_image_ocr_headerless_rows_create_temporary_columns() -> None:
    frame = pd.DataFrame(
        [
            ["Drive form", "FWD", "FWD", "FWD"],
            ["Number of", "seats", "5", "5"],
            ["Roof rack", "-", "O", "S"],
            ["Speaker count", "4", "6", "8"],
            ["Seat material", "Fabric", "Leather", "Leather"],
        ]
    )

    digest = build_workbook_digest_from_frames(
        [("OCR Image 1", frame)],
        file_name="headerless-config.png",
        digest_type="image_ocr",
    )

    assert digest["status"] == "ready"
    assert digest["summary"]["comparableGroupCount"] == 1
    assert digest["summary"]["candidateTrimCount"] == 3
    group = digest["compareGroups"][0]
    assert group["sourceKind"] == "ocr_headerless"
    assert group["identityStatus"] == "temporary_ocr_column"
    assert [trim["trimName"] for trim in group["trims"]] == ["OCR Column 1", "OCR Column 2", "OCR Column 3"]
    assert all(trim["identityStatus"] == "temporary_ocr_column" for trim in group["trims"])
    rows_by_feature = {row["featureName"]: row for row in group["rows"]}
    assert rows_by_feature["Drive form"]["comparisonType"] == "COMMON_SAME"
    assert rows_by_feature["Number of"]["reviewFlags"] == ["ocr_possible_feature_text_in_value_cell"]
    assert "特征名换行或单位被切入值列" in rows_by_feature["Number of"]["reviewNotes"][0]
    assert "需核对：OCR 值单元格像配置项文本" in rows_by_feature["Number of"]["businessNote"]
    assert rows_by_feature["Roof rack"]["comparisonType"] == "OPTIONAL_DIFFERENT"
    assert rows_by_feature["Speaker count"]["comparisonType"] == "DIFFERENT_VALUE"
    assert "reviewFlags" not in rows_by_feature["Seat material"]
    assert "OCR 未识别到配置列标题" in rows_by_feature["Roof rack"]["businessNote"]
    assert rows_by_feature["Roof rack"]["values"][1]["source"]["cell"] == "C3"


def test_pdf_ocr_headerless_rows_can_use_category_feature_shape() -> None:
    frame = pd.DataFrame(
        [
            ["Comfort", "Seat heating", "-", "S"],
            ["Safety", "360 camera", "-", "O"],
            ["Infotainment", "Speaker count", "6", "8"],
        ]
    )

    digest = build_workbook_digest_from_frames(
        [("PDF OCR Page 1", frame)],
        file_name="headerless-config.pdf",
        digest_type="pdf_ocr",
    )

    group = digest["compareGroups"][0]
    assert group["sourceKind"] == "ocr_headerless"
    assert [trim["trimName"] for trim in group["trims"]] == ["OCR Column 1", "OCR Column 2"]
    rows_by_feature = {row["featureName"]: row for row in group["rows"]}
    assert rows_by_feature["Seat heating"]["category"] == "Comfort"
    assert rows_by_feature["Seat heating"]["comparisonType"] == "UNIQUE_OR_PARTIAL"
    assert rows_by_feature["360 camera"]["comparisonType"] == "OPTIONAL_DIFFERENT"


def test_headerless_rows_do_not_create_temporary_columns_for_workbook() -> None:
    frame = pd.DataFrame(
        [
            ["Drive form", "FWD", "FWD", "FWD"],
            ["Roof rack", "-", "O", "S"],
            ["Speaker count", "4", "6", "8"],
        ]
    )

    digest = build_workbook_digest_from_frames(
        [("Worksheet", frame)],
        file_name="headerless-config.xlsx",
    )

    assert digest["summary"]["comparableGroupCount"] == 0
    assert digest["compareGroups"] == []


def test_image_ocr_digest_without_compare_group_stays_pending(tmp_path, monkeypatch) -> None:
    source_path = tmp_path / "narrative-screenshot.png"
    source_path.write_bytes(b"\x89PNG\r\n\x1a\n")
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._extract_ocr_text_candidates",
        lambda _path, compare_engines=False: (
            [("Section | Narrative | Not a config matrix", "unit_ocr")],
            [],
        ),
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "image_ocr"
    assert digest["sourceFormat"] == "image_ocr"
    assert digest["status"] == "pending"
    assert digest["compareGroups"] == []
    assert "fewer than 2 comparable features" in digest["message"]
    assert digest["ocrEvaluation"]["selectedEngine"] is None
    assert digest["ocrEngineCandidates"][0]["engine"] == "unit_ocr"
    assert digest["ocrEngineCandidates"][0]["selected"] is False


def test_image_ocr_digest_uses_paddleocr_when_available(tmp_path, monkeypatch) -> None:
    source_path = tmp_path / "rival-config.png"
    source_path.write_bytes(b"\x89PNG\r\n\x1a\n")

    class FakePaddleOCR:
        def __init__(self, **_kwargs) -> None:
            pass

        def ocr(self, _path: str, cls: bool = True) -> list[list[list[object]]]:
            return [
                [
                    [[[0, 0], [1, 0], [1, 1], [0, 1]], ("Category | Feature | Basic | Premium", 0.99)],
                    [[[0, 2], [1, 2], [1, 3], [0, 3]], ("Comfort | Seat heating | - | S", 0.98)],
                    [[[0, 4], [1, 4], [1, 5], [0, 5]], ("Safety | 360 camera | - | O", 0.97)],
                ]
            ]

    fake_module = type("FakePaddleOCRModule", (), {"PaddleOCR": FakePaddleOCR})
    monkeypatch.setitem(sys.modules, "paddleocr", fake_module)
    monkeypatch.delenv("JATO_CONFIG_OCR_COMMAND", raising=False)

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "image_ocr"
    assert digest["ocrEngine"] == "paddleocr"
    assert digest["status"] == "ready"
    group = digest["compareGroups"][0]
    assert [trim["trimName"] for trim in group["trims"]] == ["Basic", "Premium"]
    rows_by_feature = {row["featureName"]: row for row in group["rows"]}
    assert rows_by_feature["Seat heating"]["comparisonType"] == "UNIQUE_OR_PARTIAL"
    assert rows_by_feature["Seat heating"]["values"][0]["source"]["ocrEngine"] == "paddleocr"


def test_image_ocr_digest_chooses_best_engine_by_table_score(tmp_path, monkeypatch) -> None:
    source_path = tmp_path / "rival-config.png"
    source_path.write_bytes(b"\x89PNG\r\n\x1a\n")
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._extract_ocr_text_candidates",
        lambda _path, compare_engines=False: (
            [
                ("Seat heating Basic Premium", "legacy_image_ocr"),
                (
                    "\n".join(
                        [
                            "Category | Feature | Basic | Premium",
                            "Comfort | Seat heating | - | S",
                            "Safety | 360 camera | - | O",
                        ]
                    ),
                    "paddleocr",
                ),
            ],
            [],
        ),
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "image_ocr"
    assert digest["ocrEngine"] == "paddleocr"
    assert digest["ocrEvaluation"]["strategy"] == "highest_config_semantic_score"
    assert digest["ocrEvaluation"]["candidateCount"] == 2
    assert digest["ocrEvaluation"]["selectedEngine"] == "paddleocr"
    assert digest["ocrEvaluation"]["selectedReasonDetails"][0] == "paddleocr 识别到可比配置表；legacy_image_ocr 未形成可比配置表。"
    candidates = {candidate["engine"]: candidate for candidate in digest["ocrEngineCandidates"]}
    assert candidates["legacy_image_ocr"]["comparableTableDetected"] is False
    assert candidates["paddleocr"]["selected"] is True
    assert candidates["paddleocr"]["score"]["columnCount"] == 4


def test_image_ocr_digest_rejects_low_information_table(tmp_path, monkeypatch) -> None:
    source_path = tmp_path / "misread-config.png"
    source_path.write_bytes(b"\x89PNG\r\n\x1a\n")
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._extract_ocr_text_candidates",
        lambda _path, compare_engines=False: (
            [
                (
                    "\n".join(
                        [
                            "Category | Feature | 4447*1824*1588 | 4447*1824*1588",
                            "尺寸 | Wheelbase | 2610 | 2610",
                        ]
                    ),
                    "paddleocr",
                )
            ],
            [],
        ),
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "image_ocr"
    assert digest["status"] == "pending"
    assert digest["compareGroups"] == []
    assert digest["ocrEvaluation"]["reason"] == "no_comparable_table_detected"
    assert digest["ocrEvaluation"]["selectedEngine"] is None
    candidate = digest["ocrEngineCandidates"][0]
    assert candidate["engine"] == "paddleocr"
    assert candidate["selected"] is False
    assert candidate["comparableTableDetected"] is False
    assert candidate["score"]["featureCount"] == 1
    assert candidate["score"]["candidateTrimCount"] == 2
    assert "fewer than 2 comparable features" in candidate["message"]


def test_image_ocr_digest_reports_paddleocr_failure_when_legacy_engine_succeeds(tmp_path, monkeypatch) -> None:
    source_path = tmp_path / "rival-config.png"
    source_path.write_bytes(b"\x89PNG\r\n\x1a\n")
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._extract_ocr_text_candidates",
        lambda _path, compare_engines=False: (
            [
                (
                    "\n".join(
                        [
                            "Category | Feature | Basic | Premium",
                            "Comfort | Seat heating | - | S",
                            "Safety | 360 camera | - | O",
                        ]
                    ),
                    "legacy_image_ocr",
                )
            ],
            ["PaddleOCR is not installed."],
        ),
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "image_ocr"
    assert digest["ocrEngine"] == "legacy_image_ocr"
    assert digest["ocrEvaluation"]["candidateCount"] == 2
    assert digest["ocrEvaluation"]["comparableCandidateCount"] == 1
    assert digest["ocrEvaluation"]["selectedEngine"] == "legacy_image_ocr"
    candidates = {candidate["engine"]: candidate for candidate in digest["ocrEngineCandidates"]}
    assert candidates["legacy_image_ocr"]["selected"] is True
    assert candidates["legacy_image_ocr"]["comparableTableDetected"] is True
    assert candidates["paddleocr"]["selected"] is False
    assert candidates["paddleocr"]["comparableTableDetected"] is False
    assert candidates["paddleocr"]["message"] == "PaddleOCR is not installed."
    assert digest["ocrEvaluation"]["selectedReasonDetails"][0] == "legacy_image_ocr 识别到可比配置表；paddleocr 未形成可比配置表。"


def test_command_ocr_failure_message_includes_engine(monkeypatch) -> None:
    class FailedCommand:
        returncode = 1
        stderr = "bad image"
        stdout = ""

    monkeypatch.setattr(source_digest.subprocess, "run", lambda *_args, **_kwargs: FailedCommand())

    text, message = source_digest._extract_command_ocr_text(["tesseract", "input.png"], "tesseract")

    assert text is None
    assert message == "tesseract OCR execution failed with exit code 1: bad image"


def test_image_ocr_digest_reports_command_failure_when_paddleocr_succeeds(tmp_path, monkeypatch) -> None:
    source_path = tmp_path / "rival-config.png"
    source_path.write_bytes(b"\x89PNG\r\n\x1a\n")
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._extract_ocr_text_candidates",
        lambda _path, compare_engines=False: (
            [
                (
                    "\n".join(
                        [
                            "Category | Feature | Basic | Premium",
                            "Comfort | Seat heating | - | S",
                            "Safety | 360 camera | - | O",
                        ]
                    ),
                    "paddleocr",
                )
            ],
            ["tesseract OCR execution failed with exit code 1: bad image"],
        ),
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "image_ocr"
    assert digest["ocrEngine"] == "paddleocr"
    assert digest["ocrEvaluation"]["candidateCount"] == 2
    assert digest["ocrEvaluation"]["comparableCandidateCount"] == 1
    candidates = {candidate["engine"]: candidate for candidate in digest["ocrEngineCandidates"]}
    assert candidates["paddleocr"]["selected"] is True
    assert candidates["tesseract"]["selected"] is False
    assert candidates["tesseract"]["comparableTableDetected"] is False
    assert candidates["tesseract"]["message"] == "tesseract OCR execution failed with exit code 1: bad image"
    assert digest["ocrEvaluation"]["selectedReasonDetails"][0] == "paddleocr 识别到可比配置表；tesseract 未形成可比配置表。"


def test_image_ocr_digest_without_engine_stays_pending(tmp_path, monkeypatch) -> None:
    source_path = tmp_path / "rival-config.jpg"
    source_path.write_bytes(b"\xff\xd8\xff")
    monkeypatch.setattr(
        "app.services.engineering_config_source_digest._extract_ocr_text_candidates",
        lambda _path, compare_engines=False: ([], ["OCR engine is not configured."]),
    )

    digest = build_source_digest(source_path, source_path.name)

    assert digest is not None
    assert digest["digestType"] == "image_ocr"
    assert digest["sourceFormat"] == "image_ocr"
    assert digest["status"] == "pending"
    assert digest["summary"]["comparableGroupCount"] == 0
    assert digest["compareGroups"] == []
    assert digest["message"] == "OCR engine is not configured."


def test_real_eu_workbook_regression_for_ice_and_bev_if_available() -> None:
    workbook_path = (
        Path(__file__).resolve().parents[4]
        / "02_Config_MetaData"
        / "欧盟在售车型可控资源表20260226.xlsx"
    )
    if not workbook_path.exists():
        pytest.skip("local EU config workbook is not available")

    digest = build_source_digest(workbook_path, workbook_path.name)
    assert digest is not None
    groups = {group["sourceSheet"]: group for group in digest["compareGroups"]}

    ice = groups["T19C MY ICE "]
    assert ice["featureCount"] >= 220
    ice_rows = {row["featureName"]: row for row in ice["rows"]}
    assert "Country / 国家" not in ice_rows
    assert "Configuration version / 配置版型" not in ice_rows
    assert "Material No. / 物料号" not in ice_rows
    assert ice_rows["Number of seats / 座椅"]["comparisonType"] == "COMMON_SAME"
    assert [value["rawValue"] for value in ice_rows["Number of seats / 座椅"]["values"]] == ["5", "5", "5"]
    assert ice_rows["Number of seats / 座椅"]["values"][2]["source"]["mergedRange"] == "D11:F11"
    assert "舒适便利 Comfort&Convenient" in ice["summary"]["categoryCounts"]
    assert "信息娱乐 Information&Entertainment" in ice["summary"]["categoryCounts"]
    assert [trim["profile"]["materialNo"] for trim in ice["trims"]] == [
        "T71607V**MM0001",
        "T71607V**MM0002",
        "T71607V**MM0003",
    ]

    bev = groups["T19C-BEV（2025款）"]
    assert bev["featureCount"] >= 200
    bev_rows = {row["featureName"]: row for row in bev["rows"]}
    assert "Country / 国家" not in bev_rows
    assert "Configuration version / 配置版型" not in bev_rows
    assert "Material No. / 物料号" not in bev_rows
    assert any("热泵空调" in feature_name for feature_name in bev_rows)
    assert any("品牌音响（sony）" in feature_name for feature_name in bev_rows)
    assert bev_rows["Number of seats / 座椅"]["comparisonType"] == "COMMON_SAME"
    assert "舒适便利" in bev["summary"]["categoryCounts"]
    assert "科技配置" in bev["summary"]["categoryCounts"]
