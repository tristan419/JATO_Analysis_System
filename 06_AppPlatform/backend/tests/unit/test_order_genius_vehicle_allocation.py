from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import openpyxl
import pytest
from fastapi import HTTPException

from app.infra import order_genius_repository as order_repo
from app.infra import order_genius_vehicle_repository as vehicle_repo
from app.services import order_genius_vehicle_service as vehicle_service
from app.services.order_genius_vehicle_exporter import generate_vehicle_allocation_excel
from app.services.order_genius_vehicle_import_parser import parse_vehicle_allocation_xlsx
from app.services.order_genius_vehicle_service import (
    _line_items_from_order_quantities,
    _line_payload_from_material,
    build_car_code,
    build_pi_code,
    build_pi_line_code,
    bulk_update_vehicle_units,
    generate_from_order_matrix,
    get_order_matrix_allocation_plan,
    parse_car_code,
    parse_pi_code,
)


def test_vehicle_allocation_code_generation_rules() -> None:
    pi_code = build_pi_code("ro", 2026, 7, 1)
    line_code = build_pi_line_code(pi_code, 2)
    car_code = build_car_code("ro", 2026, 7, 1, 2, 12)

    assert pi_code == "PI-RO-202607-001"
    assert line_code == "PI-RO-202607-001-L02"
    assert car_code == "CAR-RO-2607-001-L02-0012"
    assert parse_pi_code(pi_code) == {
        "countryCode": "RO",
        "scopeCode": "RO",
        "year": 2026,
        "month": 7,
        "sequence": 1,
        "orderMonth": "2026-07",
    }
    assert parse_pi_code(build_pi_code("NORDIC", 2026, 7, 1)) == {
        "countryCode": "NORDIC",
        "scopeCode": "NORDIC",
        "year": 2026,
        "month": 7,
        "sequence": 1,
        "orderMonth": "2026-07",
    }
    assert parse_car_code(car_code) == {
        "countryCode": "RO",
        "yearSuffix": 26,
        "month": 7,
        "piSequence": 1,
        "lineSequence": 2,
        "unitSequence": 12,
    }


def test_vehicle_allocation_export_has_eta_column_after_etd() -> None:
    buffer = generate_vehicle_allocation_excel([
        {
            "piCode": "PI-RO-202607-001",
            "officialPiNo": "OFF-1",
            "carCode": "CAR-RO-2607-001-L02-0012",
            "vin": None,
            "materialCode": "T7151RYKUMH0001",
            "etd": "2026-07-12",
            "eta": "2026-08-02",
            "countryCode": "RO",
        }
    ])
    workbook = openpyxl.load_workbook(buffer, data_only=True)
    try:
        ws = workbook["Vehicle Allocation"]
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        assert headers[headers.index("ETD") + 1] == "ETA"
        assert headers.count("ETD") == 1
        assert ws.cell(row=2, column=headers.index("ETD") + 1).value == "2026-07-12"
        assert ws.cell(row=2, column=headers.index("ETA") + 1).value == "2026-08-02"
    finally:
        workbook.close()


def test_vehicle_allocation_import_parser_maps_eta(tmp_path: Path) -> None:
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.append(["PI Code", "Car Code", "VIN", "ETD", "ETA", "Country"])
    ws.append([
        "PI-RO-202607-001",
        "CAR-RO-2607-001-L02-0012",
        "",
        "2026-07-12",
        "2026-08-02",
        "RO",
    ])
    path = tmp_path / "vehicle-allocation.xlsx"
    workbook.save(path)
    workbook.close()

    rows = parse_vehicle_allocation_xlsx(path)

    assert rows == [
        {
            "sourceRow": 2,
            "pi_code": "PI-RO-202607-001",
            "car_code": "CAR-RO-2607-001-L02-0012",
            "vin": None,
            "etd": "2026-07-12",
            "eta": "2026-08-02",
            "country_code": "RO",
        }
    ]


def test_line_payload_uses_bom_admin_and_country_fob(monkeypatch) -> None:
    sku = SimpleNamespace(
        material_code="T7151RYKUMH0001",
        bom_template="O9-EX-BLACK",
        brand="OMODA",
        model_name="OMODA9",
        version="Exclusive",
        powertrain="PHEV",
        exterior_color_name="Black Warrior",
        exterior_color_code="CL",
        interior_color_name="Black Red",
        interior_colour_code="BR",
    )
    fob = SimpleNamespace(final_fob_eur=21200)
    monkeypatch.setattr(order_repo, "get_sku_by_material_code_any_status", lambda session, material_code: sku)
    monkeypatch.setattr(order_repo, "get_fob_for_country_sku", lambda session, country, material_code: fob)

    payload = _line_payload_from_material(
        session=object(),
        country="RO",
        material_code="T7151RYKUMH0001",
        item={"materialCode": "T7151RYKUMH0001", "quantity": 2},
    )

    assert payload["bom"] == "O9-EX-BLACK"
    assert payload["brand"] == "OMODA"
    assert payload["modelName"] == "OMODA9"
    assert payload["exteriorColorName"] == "Black Warrior"
    assert payload["interiorColorName"] == "Black Red"
    assert payload["fobEur"] == 21200


def test_line_items_can_be_built_from_selection_quantities(monkeypatch) -> None:
    cells = [
        SimpleNamespace(material_code="A", quantity=2, fob_eur=100),
        SimpleNamespace(material_code="B", quantity=1, fob_eur=200),
    ]
    calls = []

    def fake_list_quantities(session, country_code, order_year, order_month, positive_only):
        calls.append((country_code, order_year, order_month, positive_only))
        return cells

    monkeypatch.setattr(order_repo, "list_quantities_for_country_month", fake_list_quantities)

    items = _line_items_from_order_quantities(object(), "RO", 2026, 7)

    assert calls == [("RO", 2026, 7, True)]
    assert items == [
        {"materialCode": "A", "quantity": 2, "fobEur": 100.0},
        {"materialCode": "B", "quantity": 1, "fobEur": 200.0},
    ]


def test_order_matrix_plan_uses_remaining_quantities(monkeypatch) -> None:
    cells = [
        SimpleNamespace(material_code="A", quantity=3, fob_eur=100),
        SimpleNamespace(material_code="B", quantity=1, fob_eur=200),
    ]
    existing_allocation = SimpleNamespace(
        pi_line_allocation_id="alloc-1",
        pi_code="PI-RO-202607-001",
        pi_line_code="PI-RO-202607-001-L01",
        market_country_code="RO",
        order_year=2026,
        order_month=7,
        material_code="A",
        quantity=2,
        fob_eur=100,
    )

    monkeypatch.setattr(
        order_repo,
        "list_quantities_for_country_month",
        lambda session, country_code, order_year, order_month, positive_only: cells,
    )
    monkeypatch.setattr(
        vehicle_repo,
        "list_allocations_for_country_month",
        lambda session, country_code, order_year, order_month: [existing_allocation],
    )
    monkeypatch.setattr(
        vehicle_repo,
        "list_lines_without_allocations_for_country_month",
        lambda session, country_code, order_month: [],
    )
    monkeypatch.setattr(
        vehicle_repo,
        "count_vehicles_for_line_country",
        lambda session, pi_line_code, country_code: 2,
    )
    monkeypatch.setattr(
        vehicle_service,
        "_line_payload_from_material",
        lambda session, country, material_code, item: {
            **item,
            "materialCode": material_code,
            "modelName": f"Model {material_code}",
        },
    )
    monkeypatch.setattr(
        vehicle_service,
        "allocation_to_dict",
        lambda allocation: {
            "materialCode": allocation.material_code,
            "quantity": allocation.quantity,
            "piLineCode": allocation.pi_line_code,
        },
    )

    plan = get_order_matrix_allocation_plan(object(), "ro", 2026, 7)

    assert plan["countryCode"] == "RO"
    assert plan["orderMonth"] == "2026-07"
    assert plan["status"] == "pending"
    assert plan["totals"] == {
        "selectedQuantity": 4,
        "generatedQuantity": 2,
        "generatedVehicleCount": 2,
        "remainingQuantity": 2,
        "overGeneratedQuantity": 0,
    }
    assert {
        item["materialCode"]: item["quantity"]
        for item in plan["remainingLineItems"]
    } == {"A": 1, "B": 1}


def test_explicit_pi_line_items_cannot_exceed_remaining_quantities(monkeypatch) -> None:
    cells = [SimpleNamespace(material_code="A", quantity=3, fob_eur=100)]
    existing_allocation = SimpleNamespace(
        pi_line_allocation_id="alloc-1",
        pi_code="PI-RO-202607-001",
        pi_line_code="PI-RO-202607-001-L01",
        market_country_code="RO",
        order_year=2026,
        order_month=7,
        material_code="A",
        quantity=2,
        fob_eur=100,
    )
    monkeypatch.setattr(
        order_repo,
        "list_quantities_for_country_month",
        lambda session, country_code, order_year, order_month, positive_only: cells,
    )
    monkeypatch.setattr(
        vehicle_repo,
        "list_allocations_for_country_month",
        lambda session, country_code, order_year, order_month: [existing_allocation],
    )
    monkeypatch.setattr(
        vehicle_repo,
        "list_lines_without_allocations_for_country_month",
        lambda session, country_code, order_month: [],
    )
    monkeypatch.setattr(
        vehicle_repo,
        "count_vehicles_for_line_country",
        lambda session, pi_line_code, country_code: 2,
    )
    monkeypatch.setattr(
        vehicle_service,
        "_line_payload_from_material",
        lambda session, country, material_code, item: {**item, "materialCode": material_code},
    )
    monkeypatch.setattr(
        vehicle_service,
        "allocation_to_dict",
        lambda allocation: {"materialCode": allocation.material_code, "quantity": allocation.quantity},
    )

    with pytest.raises(HTTPException) as exc:
        generate_from_order_matrix(
            object(),
            {
                "countryCode": "RO",
                "orderYear": 2026,
                "orderMonth": 7,
                "lineItems": [{"materialCode": "A", "quantity": 2}],
            },
            "tester",
        )

    assert exc.value.status_code == 409
    assert "requested 2, remaining 1" in str(exc.value.detail)


def test_combined_pi_allocations_validate_each_market_country(monkeypatch) -> None:
    cells_by_country = {
        "SE": [SimpleNamespace(material_code="A", quantity=3, fob_eur=100)],
        "FI": [SimpleNamespace(material_code="A", quantity=1, fob_eur=100)],
    }

    monkeypatch.setattr(
        order_repo,
        "list_quantities_for_country_month",
        lambda session, country_code, order_year, order_month, positive_only: cells_by_country[country_code],
    )
    monkeypatch.setattr(
        vehicle_repo,
        "list_allocations_for_country_month",
        lambda session, country_code, order_year, order_month: [],
    )
    monkeypatch.setattr(
        vehicle_repo,
        "list_lines_without_allocations_for_country_month",
        lambda session, country_code, order_month: [],
    )
    monkeypatch.setattr(
        vehicle_service,
        "_line_payload_from_material",
        lambda session, country, material_code, item: {**item, "materialCode": material_code},
    )

    with pytest.raises(HTTPException) as exc:
        generate_from_order_matrix(
            object(),
            {
                "countryCode": "SE",
                "orderYear": 2026,
                "orderMonth": 7,
                "orderingAccountCode": "NORDIC",
                "marketCountryCodes": ["SE", "FI"],
                "lineItems": [
                    {
                        "materialCode": "A",
                        "quantity": 4,
                        "allocations": [
                            {"countryCode": "SE", "quantity": 2},
                            {"countryCode": "FI", "quantity": 2},
                        ],
                    }
                ],
            },
            "tester",
        )

    assert exc.value.status_code == 409
    assert "FI A: requested 2, remaining 1" in str(exc.value.detail)


def test_bulk_vehicle_update_assigns_vins_to_empty_units_in_car_code_order(monkeypatch) -> None:
    vehicles = [
        _fake_vehicle("CAR-RO-2607-001-L01-0001"),
        _fake_vehicle("CAR-RO-2607-001-L01-0002"),
        _fake_vehicle("CAR-RO-2607-001-L01-0003", vin="EXISTINGVIN"),
    ]
    monkeypatch.setattr(vehicle_repo, "get_header_by_code", lambda session, pi_code: SimpleNamespace(pi_code=pi_code))
    monkeypatch.setattr(vehicle_repo, "get_line_by_code", lambda session, pi_line_code: SimpleNamespace(pi_code="PI-RO-202607-001"))
    monkeypatch.setattr(vehicle_repo, "list_vehicles_for_bulk_update", lambda session, pi_code, pi_line_code: vehicles)
    monkeypatch.setattr(vehicle_repo, "get_vehicle_by_vin", lambda session, vin: None)

    session = SimpleNamespace(flush=lambda: None)

    result = bulk_update_vehicle_units(
        session,
        {
            "piCode": "PI-RO-202607-001",
            "piLineCode": "PI-RO-202607-001-L01",
            "vinList": ["vin-a", "vin-b"],
            "fields": {"eta": "2026-08-02", "shipName": "Baltic Star"},
        },
        "tester",
    )

    assert result == {
        "piCode": "PI-RO-202607-001",
        "piLineCode": "PI-RO-202607-001-L01",
        "matchedUnits": 3,
        "updatedUnits": 3,
        "vinAssigned": 2,
        "fieldsUpdated": ["eta", "shipName"],
    }
    assert [vehicle.vin for vehicle in vehicles] == ["VIN-A", "VIN-B", "EXISTINGVIN"]
    assert [vehicle.ship_name for vehicle in vehicles] == ["Baltic Star", "Baltic Star", "Baltic Star"]
    assert [str(vehicle.eta) for vehicle in vehicles] == ["2026-08-02", "2026-08-02", "2026-08-02"]


def test_bulk_vehicle_update_rejects_duplicate_pasted_vins(monkeypatch) -> None:
    monkeypatch.setattr(vehicle_repo, "get_header_by_code", lambda session, pi_code: SimpleNamespace(pi_code=pi_code))
    monkeypatch.setattr(
        vehicle_repo,
        "list_vehicles_for_bulk_update",
        lambda session, pi_code, pi_line_code: [_fake_vehicle("CAR-RO-2607-001-L01-0001")],
    )

    with pytest.raises(HTTPException) as exc:
        bulk_update_vehicle_units(
            SimpleNamespace(flush=lambda: None),
            {
                "piCode": "PI-RO-202607-001",
                "vinList": ["VIN-A", "VIN-A"],
            },
            "tester",
        )

    assert exc.value.status_code == 400
    assert "Duplicate VINs" in str(exc.value.detail)


def _fake_vehicle(car_code: str, vin: str | None = None) -> SimpleNamespace:
    return SimpleNamespace(
        vehicle_unit_id=uuid4(),
        car_code=car_code,
        pi_code="PI-RO-202607-001",
        pi_line_code="PI-RO-202607-001-L01",
        vin=vin,
        dealer_code=None,
        dealer_name=None,
        customer_ref=None,
        ship_name=None,
        remark=None,
        production_date=None,
        etd=None,
        eta=None,
        actual_departure_date=None,
        actual_arrival_date=None,
        ready_for_pickup_date=None,
        allocation_status="unallocated",
        logistics_status="pending",
        row_version=1,
        updated_by=None,
    )
