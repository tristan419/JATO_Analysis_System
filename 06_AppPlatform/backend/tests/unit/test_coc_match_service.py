import zipfile
from pathlib import Path

import pytest
from fastapi import HTTPException
from upload_toolkit.job_engine import load_job_state, persist_job_state, state_path

import app.services.coc_match_service as coc_match_service
from app.services.coc_match_service import (
    CocMatchJobRunner,
    _build_coc_match_failure_result,
    classify_coc_difference,
    find_archive_only_files,
    list_archive_files,
    match_cocs,
    read_excel_rows,
    _list_rar_files_python,
)


def test_match_cocs_reports_missing_excel_rows() -> None:
    rows = [
        {"chassis": "A001", "model": "J7", "country": "CZ"},
        {"chassis": "A002", "model": "J7", "country": "CZ"},
    ]

    matched, missing = match_cocs(rows, {"A001"})

    assert [item["chassis"] for item in matched] == ["A001"]
    assert [item["chassis"] for item in missing] == ["A002"]


def test_find_archive_only_files_reports_extra_archive_files() -> None:
    rows = [
        {"chassis": "A001", "model": "J7", "country": "CZ"},
        {"chassis": "A002", "model": "J7", "country": "CZ"},
    ]

    archive_only = find_archive_only_files(rows, {"A001", "A002", "A003", "A004"})

    assert archive_only == [{"filename": "A003"}, {"filename": "A004"}]


def test_read_excel_rows_finds_vin_header_in_multi_column_sheet(tmp_path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Exported COC registry"])
    sheet.append(["Order", "Model", "VIN", "Country", "Comment"])
    sheet.append([1, "J7", "LVUGTB220TDE99425", "CZ", "ok"])
    sheet.append([2, "J7", "LVUGTB220TDE99426", "CZ", "ok"])
    sheet.append([3, "J7", None, "CZ", "skip"])
    path = tmp_path / "registry.xlsx"
    workbook.save(path)

    assert read_excel_rows(path) == [
        {"chassis": "LVUGTB220TDE99425", "model": "J7", "country": "CZ"},
        {"chassis": "LVUGTB220TDE99426", "model": "J7", "country": "CZ"},
    ]


def test_read_excel_rows_infers_headerless_single_vin_column(tmp_path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append([" LVUGTBHD5TC114629 "])
    sheet.append(["not-a-vin-note"])
    sheet.append(["LNNBDDEH0TG030197"])
    sheet.append([None])
    sheet.append(["LNNBDDEH0TG030457"])
    path = tmp_path / "headerless-vins.xlsx"
    workbook.save(path)

    assert read_excel_rows(path) == [
        {"chassis": "LVUGTBHD5TC114629", "model": "", "country": ""},
        {"chassis": "LNNBDDEH0TG030197", "model": "", "country": ""},
        {"chassis": "LNNBDDEH0TG030457", "model": "", "country": ""},
    ]


def test_classify_coc_difference_detects_one_sided_and_bidirectional() -> None:
    assert classify_coc_difference(0, 0) == "matched"
    assert classify_coc_difference(1, 0) == "missing_archive_files"
    assert classify_coc_difference(0, 1) == "archive_only_files"
    assert classify_coc_difference(1, 1) == "bidirectional_mismatch"


def test_build_coc_match_failure_result_explains_excel_stage() -> None:
    state = {
        "phase": "reading_excel",
        "excelFilename": "vin.xlsx",
        "archiveFilename": "coc.rar",
        "fileExt": ".pdf",
        "country": "SE",
        "month": "2026-07",
    }

    result = _build_coc_match_failure_result(
        state,
        HTTPException(status_code=400, detail="Excel 未找到 VIN / Chassis / 车架号 表头。"),
    )

    assert result["stage"] == "reading_excel"
    assert result["stageLabel"] == "Excel 读取"
    assert result["message"] == "Excel 未找到 VIN / Chassis / 车架号 表头。"
    assert "VIN / Chassis / 车架号" in result["suggestion"]
    assert result["retryable"] is False
    assert result["actionLabel"] == "重新上传"
    assert result["excelFilename"] == "vin.xlsx"
    assert result["archiveFilename"] == "coc.rar"


def test_read_excel_rows_fails_when_vin_column_has_no_valid_rows(tmp_path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["VIN", "Model", "Country"])
    sheet.append([None, "J7", "CZ"])
    sheet.append(["", "J7", "CZ"])
    path = tmp_path / "empty-vin.xlsx"
    workbook.save(path)

    with pytest.raises(HTTPException) as exc:
        read_excel_rows(path)

    assert "没有有效 VIN 数据" in str(exc.value.detail)


def test_coc_match_runner_persists_failure_result_for_excel_error(tmp_path: Path) -> None:
    job_root = tmp_path / "jobs"
    job_id = "coc-match-test"
    job_dir = job_root / job_id
    job_dir.mkdir(parents=True)
    excel_path = job_dir / "excel-bad.xlsx"
    archive_path = job_dir / "archive-coc.zip"
    excel_path.write_text("not an excel workbook", encoding="utf-8")
    with zipfile.ZipFile(archive_path, "w") as zf:
        zf.writestr("A001.pdf", b"pdf")
    persist_job_state(
        state_path(job_root, job_id),
        {
            "jobId": job_id,
            "status": "queued",
            "phase": "pending",
            "country": "SE",
            "month": "2026-07",
            "fileExt": ".pdf",
            "excelFilename": "bad.xlsx",
            "archiveFilename": "coc.zip",
            "failureResult": None,
        },
    )
    runner = CocMatchJobRunner(
        job_id=job_id,
        state_dir=job_root,
        excel_path=excel_path,
        archive_path=archive_path,
        country="SE",
        month="2026-07",
        file_ext=".pdf",
        triggered_by="test",
    )

    with pytest.raises(Exception):
        runner.run()

    state = load_job_state(state_path(job_root, job_id))
    assert state["phase"] == "reading_excel"
    assert state["failureResult"]["stage"] == "reading_excel"
    assert state["failureResult"]["stageLabel"] == "Excel 读取"
    assert state["failureResult"]["retryable"] is False
    assert state["failureResult"]["excelFilename"] == "bad.xlsx"


def test_coc_match_runner_succeeds_with_warning_when_archive_has_no_target_files(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from openpyxl import Workbook

    job_root = tmp_path / "jobs"
    monkeypatch.setattr(coc_match_service, "COC_MATCH_JOB_ROOT", job_root)
    monkeypatch.setattr(coc_match_service, "COC_DB_PATH", job_root / "coc_match_history.db")
    job_id = "coc-match-no-target-files"
    job_dir = job_root / job_id
    job_dir.mkdir(parents=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["VIN", "Model", "Country"])
    sheet.append(["LVUGTB220TDE99425", "J7", "SE"])
    excel_path = job_dir / "excel-vin.xlsx"
    workbook.save(excel_path)

    archive_path = job_dir / "archive-coc.zip"
    with zipfile.ZipFile(archive_path, "w") as zf:
        zf.writestr("notes.txt", b"no pdf here")

    persist_job_state(
        state_path(job_root, job_id),
        {
            "jobId": job_id,
            "status": "queued",
            "phase": "pending",
            "country": "SE",
            "month": "2026-07",
            "fileExt": ".pdf",
            "excelFilename": "vin.xlsx",
            "archiveFilename": "coc.zip",
            "failureResult": None,
        },
    )
    runner = CocMatchJobRunner(
        job_id=job_id,
        state_dir=job_root,
        excel_path=excel_path,
        archive_path=archive_path,
        country="SE",
        month="2026-07",
        file_ext=".pdf",
        triggered_by="test",
    )

    runner.run()

    state = load_job_state(state_path(job_root, job_id))
    assert state["status"] == "success"
    assert state["matchedCount"] == 0
    assert state["missingCount"] == 1
    assert "没有找到 .PDF COC 文件" in state["inputWarning"]


def test_get_coc_match_job_marks_stale_running_job(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    job_id = "coc-match-stale"
    monkeypatch.setattr(coc_match_service, "COC_MATCH_JOB_ROOT", tmp_path)
    monkeypatch.setattr(coc_match_service, "_COC_MATCH_STALE_AFTER_SECONDS", -1)
    persist_job_state(
        state_path(tmp_path, job_id),
        {
            "jobId": job_id,
            "jobType": "match",
            "status": "running",
            "phase": "listing_archive",
            "country": "SE",
            "month": "2026-07",
            "fileExt": ".pdf",
            "excelFilename": "vin.xlsx",
            "archiveFilename": "coc.zip",
            "createdAt": "2000-01-01T00:00:00",
            "startedAt": "2000-01-01T00:00:00",
            "updatedAt": "2000-01-01T00:00:00",
            "failureResult": None,
        },
    )

    payload = coc_match_service.get_coc_match_job(job_id)

    assert payload["status"] == "failed"
    assert payload["phase"] == "failed"
    assert payload["failureResult"]["stage"] == "listing_archive"
    assert payload["failureResult"]["retryable"] is True
    assert payload["failureResult"]["actionLabel"] == "重试"


def test_list_archive_files_reads_zip_without_external_unzip(tmp_path: Path) -> None:
    archive = tmp_path / "coc.zip"
    with zipfile.ZipFile(archive, "w") as zf:
        zf.writestr("nested/A001.pdf", b"pdf")
        zf.writestr("A002.PDF", b"pdf")
        zf.writestr("ignore.txt", b"txt")

    assert list_archive_files(archive, [".pdf"]) == {"A001", "A002"}


def test_rar5_python_fallback_reads_sample_archive_when_available() -> None:
    sample = Path("/Users/litristan/Downloads/COCtrack/CZ/27.rar")
    if not sample.exists():
        return

    names = _list_rar_files_python(sample, [".pdf"])

    assert "LVUGTB220TDE99425" in names
    assert len(names) >= 1
