from openpyxl import load_workbook
from pypdf import PdfReader

from app.services.engineering_config_export_service import (
    generate_engineering_config_compare_pdf,
    generate_engineering_config_compare_xlsx,
)


def test_generate_engineering_config_compare_xlsx_includes_table_and_evidence() -> None:
    payload = {
        "scope": {
            "title": "Basic vs Premium",
            "baseLabel": "Basic",
            "rangeLabel": "全部配置行",
            "targetLabel": "Premium",
            "categoryLabel": "Safety",
            "searchLabel": "blind",
        },
        "summary": {"totalFeatures": 2, "shownFeatures": 2},
        "evidenceSummary": {
            "rowCount": 1,
            "trimCount": 2,
            "valueCount": 2,
            "inferredValueCount": 1,
            "missingValueCount": 0,
            "missingSourceValueCount": 0,
            "sourceIssueRowCount": 0,
            "mergedCellValueCount": 0,
        },
        "trims": [
            {
                "trimId": "basic",
                "fullTrimName": "Basic",
                "dataOrigin": "own_catalog",
                "materialNo": "MAT-BASIC",
                "market": "Germany",
                "modelYear": "2026",
                "sourceFileName": "own-basic.xlsx",
            },
            {
                "trimId": "premium",
                "fullTrimName": "Premium",
                "dataOrigin": "external_or_scraped",
                "salesVersion": "PREM-SV",
                "market": "France",
                "modelYear": "2025",
                "sourceFileName": "rival-premium.pdf",
                "sourceCreatedBy": "alice",
            },
        ],
        "rows": [
            {
                "category": "Safety",
                "featureCode": "blind_spot",
                "featureName": "Blind spot",
                "comparisonType": "UNIQUE_OR_PARTIAL",
                "businessNote": "Premium adds blind spot.",
                "values": [
                    {
                        "rawValue": "",
                        "displayValue": "不配备*",
                        "availability": "NOT_AVAILABLE",
                        "valueState": "blank",
                        "inferred": True,
                        "inferenceReason": "blank_as_not_equipped_by_eu_matrix_policy",
                        "confidence": 0.7,
                        "source": {
                            "sheetName": "T19C MY ICE",
                            "rowNumber": 128,
                            "columnLetter": "D",
                            "cell": "D128",
                            "sourceCell": "D128",
                            "mergedRange": None,
                            "ocrEngine": "paddleocr",
                        },
                    },
                    {
                        "rawValue": "●",
                        "displayValue": "标配",
                        "availability": "STANDARD",
                        "valueState": "marker_value",
                        "source": {
                            "sheetName": "T19C MY ICE",
                            "rowNumber": 128,
                            "columnLetter": "E",
                            "cell": "E128",
                            "sourceCell": "E128",
                            "mergedRange": None,
                        },
                    },
                ],
            }
        ],
    }

    workbook_bytes = generate_engineering_config_compare_xlsx(payload)
    workbook = load_workbook(workbook_bytes)

    assert workbook.sheetnames == ["Config Compare", "Evidence"]
    compare_sheet = workbook["Config Compare"]
    assert compare_sheet["A1"].value == "Basic vs Premium"
    assert "范围 全部配置行" in compare_sheet["A2"].value
    assert "基准 Basic" in compare_sheet["A2"].value
    assert "目标 Premium" in compare_sheet["A2"].value
    assert "大类 Safety" in compare_sheet["A2"].value
    assert "搜索 blind" in compare_sheet["A2"].value
    assert compare_sheet["A3"].value == (
        "证据审计 · 导出行 1 · 配置列 2 · 单元格 2 · 规则推断 1 · 缺值 0 · "
        "缺来源 0 · 来源问题行 0 · 合并格 0"
    )
    assert [compare_sheet.cell(row=4, column=col).value for col in range(1, 7)] == [
        "配置项",
        "大类",
        "差异类型",
        "Basic · 本品 · 物料号 MAT-BASIC · Germany · 2026 · own-basic.xlsx",
        "Premium · 竞品 / 外部 · Sales version PREM-SV · France · 2025 · rival-premium.pdf · 来源人 alice",
        "业务备注",
    ]
    assert compare_sheet["A5"].value == "Blind spot"
    assert compare_sheet["D5"].value == "不配备*"
    assert compare_sheet["E5"].value == "标配"

    evidence_sheet = workbook["Evidence"]
    assert evidence_sheet["A1"].value == "配置项"
    assert evidence_sheet["C2"].value == "Basic · 本品 · 物料号 MAT-BASIC · Germany · 2026 · own-basic.xlsx"
    assert evidence_sheet["C3"].value == "Premium · 竞品 / 外部 · Sales version PREM-SV · France · 2025 · rival-premium.pdf · 来源人 alice"
    assert evidence_sheet["L1"].value == "rowNumber"
    assert evidence_sheet["M1"].value == "columnLetter"
    assert evidence_sheet["Q1"].value == "ocrEngine"
    assert evidence_sheet["D2"].value == "不配备*"
    assert evidence_sheet["H2"].value == "yes"
    assert evidence_sheet["I2"].value == "blank_as_not_equipped_by_eu_matrix_policy"
    assert evidence_sheet["J2"].value == 0.7
    assert evidence_sheet["K2"].value == "T19C MY ICE"
    assert evidence_sheet["L2"].value == 128
    assert evidence_sheet["M2"].value == "D"
    assert evidence_sheet["N2"].value == "D128"
    assert evidence_sheet["Q2"].value == "paddleocr"
    assert evidence_sheet.column_dimensions["Q"].width == 16


def test_generate_engineering_config_compare_xlsx_includes_ai_summary_sheet() -> None:
    payload = {
        "scope": {"title": "Basic vs Premium", "rangeLabel": "全部配置行"},
        "summary": {"totalFeatures": 1, "shownFeatures": 1},
        "trims": [
            {
                "trimId": "basic",
                "fullTrimName": "Basic",
                "dataOrigin": "own_catalog",
                "materialNo": "MAT-BASIC",
                "market": "Germany",
                "modelYear": "2026",
                "sourceFileName": "own-basic.xlsx",
            },
            {
                "trimId": "premium",
                "fullTrimName": "Premium",
                "dataOrigin": "external_or_scraped",
                "salesVersion": "PREM-SV",
                "market": "France",
                "modelYear": "2025",
                "sourceFileName": "rival-premium.pdf",
                "sourceCreatedBy": "alice",
            },
        ],
        "businessSummary": [
            {
                "targetTrimId": "premium",
                "targetLabel": "Premium",
                "headline": "Premium 相比 Basic 的主要升级集中在泊车辅助。",
                "mainUpgrades": ["泊车辅助：倒车影像升级为 360 全景影像"],
                "replacementsOrReductions": ["手动折叠后视镜被电动折叠替代"],
                "evidenceStatus": ["1 项来自规则推断，不是 Excel 原文"],
                "evidenceRefs": [
                    {
                        "section": "mainUpgrades",
                        "itemIndex": 0,
                        "evidenceKey": "premium:ADDED:rear_camera",
                        "featureCode": "rear_camera",
                        "category": "Safety",
                        "reason": "AI 摘要中的泊车辅助升级来自 rear_camera 配置差异。",
                    }
                ],
                "recommendedUse": "引用前核对 evidence。",
            }
        ],
        "businessSummaryUsage": {
            "provider": "deepseek",
            "model": "deepseek-chat",
            "status": "ok",
            "totalTokens": 150,
            "finishReason": "stop",
            "transportFallback": "curl",
        },
        "rows": [
            {
                "category": "Safety",
                "featureCode": "rear_camera",
                "featureName": "Rear camera",
                "comparisonType": "UNIQUE_OR_PARTIAL",
                "values": [None, {"displayValue": "标配", "availability": "STANDARD"}],
            }
        ],
    }

    workbook_bytes = generate_engineering_config_compare_xlsx(payload)
    workbook = load_workbook(workbook_bytes)

    assert workbook.sheetnames == ["Config Compare", "AI Summary", "Evidence"]
    summary_sheet = workbook["AI Summary"]
    assert summary_sheet["A1"].value == "目标配置列"
    assert summary_sheet["A2"].value == "Premium"
    assert summary_sheet["B2"].value == "Premium 相比 Basic 的主要升级集中在泊车辅助。"
    assert "360 全景影像" in summary_sheet["C2"].value
    assert "不是 Excel 原文" in summary_sheet["E2"].value
    assert summary_sheet["F1"].value == "证据引用"
    assert "premium:ADDED:rear_camera" in summary_sheet["F2"].value
    assert "AI 摘要中的泊车辅助升级" in summary_sheet["F2"].value
    assert summary_sheet["H1"].value == "LLM Provider"
    assert summary_sheet["H2"].value == "deepseek"
    assert summary_sheet["I2"].value == "deepseek-chat"
    assert summary_sheet["J2"].value == "ok"
    assert summary_sheet["K2"].value == "150"
    assert summary_sheet["L2"].value == "stop"
    assert summary_sheet["M1"].value == "Transport Fallback"
    assert summary_sheet["M2"].value == "curl"
    assert summary_sheet["N1"].value == "导出口径"
    assert "Basic vs Premium" in summary_sheet["N2"].value
    assert "范围 全部配置行" in summary_sheet["N2"].value


def test_generate_engineering_config_compare_pdf_includes_visible_rows() -> None:
    payload = {
        "scope": {
            "title": "Basic vs Premium",
            "baseLabel": "Basic",
            "rangeLabel": "全部配置行",
            "targetLabel": "Premium",
            "categoryLabel": "Safety",
            "searchLabel": "blind",
        },
        "summary": {"totalFeatures": 2, "shownFeatures": 1},
        "evidenceSummary": {
            "rowCount": 1,
            "trimCount": 2,
            "valueCount": 2,
            "inferredValueCount": 1,
            "missingValueCount": 0,
            "missingSourceValueCount": 0,
            "sourceIssueRowCount": 0,
            "mergedCellValueCount": 0,
        },
        "trims": [
            {
                "trimId": "basic",
                "fullTrimName": "Basic",
                "dataOrigin": "own_catalog",
                "materialNo": "MAT-BASIC",
                "market": "Germany",
                "modelYear": "2026",
                "sourceFileName": "own-basic.xlsx",
            },
            {
                "trimId": "premium",
                "fullTrimName": "Premium",
                "dataOrigin": "external_or_scraped",
                "salesVersion": "PREM-SV",
                "market": "France",
                "modelYear": "2025",
                "sourceFileName": "rival-premium.pdf",
                "sourceCreatedBy": "alice",
            },
        ],
        "rows": [
            {
                "category": "Safety",
                "featureCode": "blind_spot",
                "featureName": "Blind spot",
                "comparisonType": "UNIQUE_OR_PARTIAL",
                "businessNote": "Premium adds blind spot.",
                "values": [
                    {
                        "rawValue": "",
                        "displayValue": "Not available*",
                        "availability": "NOT_AVAILABLE",
                        "valueState": "blank",
                        "inferred": True,
                        "inferenceReason": "blank_as_not_equipped_by_eu_matrix_policy",
                        "confidence": 0.7,
                        "source": {
                            "sheetName": "T19C MY ICE",
                            "rowNumber": 128,
                            "columnLetter": "D",
                            "cell": "D128",
                            "sourceCell": "D128",
                            "ocrEngine": "paddleocr",
                            "mergedRange": None,
                        },
                    },
                    {
                        "rawValue": "standard",
                        "displayValue": "Standard",
                        "availability": "STANDARD",
                        "valueState": "marker_value",
                        "source": {
                            "sheetName": "T19C MY ICE",
                            "cell": "E128",
                            "sourceCell": "E128",
                            "mergedRange": None,
                        },
                    },
                ],
            }
        ],
    }

    pdf_bytes = generate_engineering_config_compare_pdf(payload)
    assert pdf_bytes.getvalue().startswith(b"%PDF-1.4")
    reader = PdfReader(pdf_bytes)
    assert len(reader.pages) >= 1
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Basic vs Premium" in text
    assert "基准 Basic" in text
    assert "目标 Premium" in text
    assert "搜索 blind" in text
    assert "证据审计" in text
    assert "规则推断 1" in text
    assert "来源问题行 0" in text
    assert "Blind spot" in text
    assert "物料号 MAT-BASIC" in text
    assert "own-basic.xlsx" in text
    assert "Sales version PREM-SV" in text
    assert "rival-premium.pdf" in text
    assert "来源人 alice" in text
    assert "Premium adds blind spot." in text
    assert "blank_as_not_equipped_by_eu_matrix_policy" in text
    assert "row=128" in text
    assert "column=D" in text
    assert "ocrEngine=paddleocr" in text


def test_generate_engineering_config_compare_pdf_includes_ai_summary() -> None:
    payload = {
        "scope": {"title": "Basic vs Premium", "rangeLabel": "全部配置行"},
        "summary": {"totalFeatures": 1, "shownFeatures": 1},
        "trims": [
            {"trimId": "basic", "fullTrimName": "Basic"},
            {"trimId": "premium", "fullTrimName": "Premium"},
        ],
        "businessSummary": [
            {
                "targetTrimId": "premium",
                "targetLabel": "Premium",
                "headline": "Premium 相比 Basic 的主要升级集中在泊车辅助。",
                "mainUpgrades": ["泊车辅助：倒车影像升级为 360 全景影像"],
                "replacementsOrReductions": ["手动折叠后视镜被电动折叠替代"],
                "evidenceStatus": ["1 项来自规则推断，不是 Excel 原文"],
                "evidenceRefs": [
                    {
                        "section": "mainUpgrades",
                        "itemIndex": 0,
                        "evidenceKey": "premium:ADDED:rear_camera",
                        "featureCode": "rear_camera",
                        "category": "Safety",
                        "reason": "AI 摘要中的泊车辅助升级来自 rear_camera 配置差异。",
                    }
                ],
                "recommendedUse": "引用前核对 evidence。",
            }
        ],
        "businessSummaryUsage": {
            "provider": "deepseek",
            "model": "deepseek-chat",
            "status": "ok",
            "totalTokens": 150,
            "finishReason": "stop",
            "transportFallback": "curl",
        },
        "rows": [
            {
                "category": "Safety",
                "featureCode": "rear_camera",
                "featureName": "Rear camera",
                "comparisonType": "UNIQUE_OR_PARTIAL",
                "values": [None, {"displayValue": "Standard", "availability": "STANDARD"}],
            }
        ],
    }

    pdf_bytes = generate_engineering_config_compare_pdf(payload)
    reader = PdfReader(pdf_bytes)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert "AI 业务摘要" in text
    assert "Premium 相比 Basic 的主要升级集中在泊车辅助。" in text
    assert "LLM usage: provider=deepseek / model=deepseek-chat / status=ok / tokens=150 / finish=stop" in text
    assert "transportFallback=curl" in text
    assert "premium:ADDED:rear_camera" in text
    assert "AI 摘要中的泊车辅助升级" in text
    assert "360 全景影像" in text
    assert "不是 Excel 原文" in text


def test_generate_engineering_config_compare_pdf_preserves_chinese_text() -> None:
    payload = {
        "scope": {
            "title": "配置对比",
            "rangeLabel": "全部配置行",
            "targetLabel": "尊贵型 Premium",
        },
        "summary": {"totalFeatures": 1, "shownFeatures": 1},
        "trims": [
            {"trimId": "basic", "fullTrimName": "基本型 Basic"},
            {"trimId": "premium", "fullTrimName": "尊贵型 Premium"},
        ],
        "rows": [
            {
                "category": "安全 Safety",
                "featureCode": "rear_camera",
                "featureName": "倒车影像",
                "comparisonType": "UNIQUE_OR_PARTIAL",
                "businessNote": "尊贵型新增倒车影像。",
                "values": [
                    {
                        "rawValue": "",
                        "displayValue": "不配备*",
                        "availability": "NOT_AVAILABLE",
                        "valueState": "blank",
                        "inferred": True,
                        "inferenceReason": "blank_as_not_equipped_by_eu_matrix_policy",
                        "source": {"sheetName": "配置表", "cell": "D10"},
                    },
                    {
                        "rawValue": "●",
                        "displayValue": "标配",
                        "availability": "STANDARD",
                        "valueState": "marker_value",
                        "source": {"sheetName": "配置表", "cell": "E10"},
                    },
                ],
            }
        ],
    }

    pdf_bytes = generate_engineering_config_compare_pdf(payload)
    reader = PdfReader(pdf_bytes)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert "配置对比" in text
    assert "基本型 Basic" in text
    assert "尊贵型 Premium" in text
    assert "倒车影像" in text
    assert "不配备*" in text
    assert "标配" in text
    assert "尊贵型新增倒车影像。" in text
