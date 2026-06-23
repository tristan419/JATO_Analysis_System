from __future__ import annotations

from pathlib import Path

import pytest
from fastapi import HTTPException
from openpyxl import Workbook

from app.services.workbook_table_scanner import extract_material_rows


def _save_workbook(workbook: Workbook, path: Path) -> Path:
    workbook.save(path)
    workbook.close()
    return path


def test_extract_material_rows_scans_all_sheets_and_creates_target_columns(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "随便一个空Sheet"
    sheet1["A1"] = "备注"

    sheet3 = workbook.create_sheet("希腊发运随机Sheet3")
    sheet3["B3"] = "生产日期"
    sheet3["C3"] = "物料号组"
    sheet3["D3"] = "国家"
    sheet3["B4"] = "2026年5月11-5月28"
    sheet3["C4"] = "T7000Z5**MY0013"
    sheet3["D4"] = "GR"

    workbook_path = _save_workbook(workbook, tmp_path / "shipment.xlsx")

    parsed_workbook, targets_by_sheet, rows = extract_material_rows(workbook_path)
    try:
        assert list(targets_by_sheet) == ["希腊发运随机Sheet3"]
        assert len(rows) == 1
        assert rows[0].sheet_name == "希腊发运随机Sheet3"
        assert rows[0].row_number == 4
        assert rows[0].material_group == "T7000Z5**MY0013"
        assert rows[0].country == "GR"
        assert rows[0].wvta_cell == "E4"
        assert rows[0].coc_cell == "F4"

        parsed_sheet = parsed_workbook["希腊发运随机Sheet3"]
        assert parsed_sheet["E3"].value == "WVTA编号"
        assert parsed_sheet["F3"].value == "COC编号"
    finally:
        parsed_workbook.close()


def test_extract_material_rows_infers_material_column_when_header_is_missing(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "没有固定表头"
    sheet["B4"] = "T716015B**MH0001"
    sheet["B5"] = "T716015C**MH0001"

    workbook_path = _save_workbook(workbook, tmp_path / "inferred.xlsx")

    parsed_workbook, targets_by_sheet, rows = extract_material_rows(workbook_path)
    try:
        assert list(targets_by_sheet) == ["没有固定表头"]
        assert [row.material_group for row in rows] == ["T716015B**MH0001", "T716015C**MH0001"]
        assert all(row.header_inferred for row in rows)

        parsed_sheet = parsed_workbook["没有固定表头"]
        assert parsed_sheet["B3"].value == "物料号组"
        assert parsed_sheet["C3"].value == "WVTA编号"
        assert parsed_sheet["D3"].value == "COC编号"
    finally:
        parsed_workbook.close()


def test_extract_material_rows_reports_missing_requested_sheet(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "SheetA"
    workbook_path = _save_workbook(workbook, tmp_path / "missing.xlsx")

    with pytest.raises(HTTPException) as exc_info:
        extract_material_rows(workbook_path, sheet_names=["Sheet3"])

    assert exc_info.value.status_code == 422
    assert "指定 Sheet 不存在" in str(exc_info.value.detail)
