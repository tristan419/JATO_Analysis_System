from pathlib import Path

import openpyxl

from app.services.order_genius_export_service import (
    generate_order_genius_excel,
    generate_order_genius_pi_excel,
)
from app.services.order_quantity_parser import parse_order_quantity_xlsx


def test_order_genius_export_can_be_imported_back_with_current_columns(
    tmp_path: Path,
) -> None:
    months = {str(month): {"quantity": 0} for month in range(1, 13)}
    months["1"]["quantity"] = 4
    months["2"]["quantity"] = 3

    export_buffer = generate_order_genius_excel(
        rows=[
            {
                "brand": "OMODA",
                "modelName": "OMODA 5",
                "version": "Comfort",
                "colour": "White",
                "interiorColorName": "Black",
                "materialCode": "T19C-AB",
                "fobEur": 12345.0,
                "powertrain": "ICE",
                "lifecycleStatus": "active",
                "months": months,
                "ttl": 7,
            }
        ],
        country_code="SE",
        country_name="Sweden",
        year=2026,
    )

    export_path = tmp_path / "order-genius-export.xlsx"
    export_path.write_bytes(export_buffer.getvalue())

    parsed = parse_order_quantity_xlsx(export_path)

    assert parsed.errors == []
    assert parsed.country_code == "SE"
    assert parsed.year == 2026
    assert [row.material_code for row in parsed.rows] == ["T19C-AB"]

    row = parsed.rows[0]
    assert row.model_name == "OMODA 5"
    assert row.version == "Comfort"
    assert row.colour == "White"
    assert [cell.quantity for cell in row.cells[:2]] == [4, 3]
    assert sum(cell.quantity for cell in row.cells) == 7


def test_order_genius_single_month_export_can_be_imported_back(
    tmp_path: Path,
) -> None:
    months = {str(month): {"quantity": 0} for month in range(1, 13)}
    months["1"]["quantity"] = 4
    months["2"]["quantity"] = 3

    export_buffer = generate_order_genius_excel(
        rows=[
            {
                "brand": "OMODA",
                "modelName": "OMODA 5",
                "version": "Comfort",
                "colour": "White",
                "interiorColorName": "Black",
                "materialCode": "T19C-AB",
                "fobEur": 12345.0,
                "powertrain": "ICE",
                "lifecycleStatus": "active",
                "months": months,
                "ttl": 7,
            }
        ],
        country_code="SE",
        country_name="Sweden",
        year=2026,
        selected_months=[2],
    )

    workbook = openpyxl.load_workbook(export_buffer, data_only=True)
    try:
        headers = [
            workbook["ICE"].cell(row=2, column=col).value
            for col in range(1, workbook["ICE"].max_column + 1)
        ]
        assert "Feb" in headers
        assert "Jan" not in headers
        assert "Mar" not in headers
        assert workbook["ICE"].cell(row=3, column=headers.index("TTL") + 1).value == 3
    finally:
        workbook.close()

    export_path = tmp_path / "order-genius-single-month-export.xlsx"
    export_path.write_bytes(export_buffer.getvalue())

    parsed = parse_order_quantity_xlsx(export_path)

    assert parsed.errors == []
    assert parsed.country_code == "SE"
    assert parsed.year == 2026
    assert [row.material_code for row in parsed.rows] == ["T19C-AB"]
    assert [(cell.month, cell.quantity) for cell in parsed.rows[0].cells] == [(2, 3)]


def test_order_genius_pi_export_maps_business_columns() -> None:
    months = {str(month): {"quantity": 0} for month in range(1, 13)}
    months["2"]["quantity"] = 3

    export_buffer = generate_order_genius_pi_excel(
        rows=[
            {
                "brand": "OMODA",
                "modelName": "OMODA9 SHS",
                "version": "Exclusive-AWD",
                "colour": "Silver",
                "colourCode": "KU",
                "colourType": "single",
                "materialCode": "T7151RYKUMH0001",
                "fobEur": 21200.0,
                "powertrain": "PHEV",
                "months": months,
            },
            {
                "brand": "JAECOO",
                "modelName": "JAECOO7 HEV",
                "version": "Exclusive-FWD",
                "colour": "Black & White",
                "colourCode": "BW",
                "colourType": "dual",
                "materialCode": "T71604NBWMH0032",
                "fobEur": 19200.0,
                "powertrain": "HEV",
                "months": months,
            },
        ],
        country_code="SE",
        country_name="Sweden",
        year=2026,
        quantity_month=2,
        nl_fob_by_material_code={
            "T7151RYKUMH0001": 20200.0,
            "T71604NBWMH0032": 17700.0,
        },
    )

    workbook = openpyxl.load_workbook(export_buffer, data_only=True)
    try:
        ws = workbook["PI"]
        headers = [ws.cell(row=3, column=col).value for col in range(1, ws.max_column + 1)]
        values = {
            header: ws.cell(row=4, column=idx + 1).value
            for idx, header in enumerate(headers)
        }
        dual_values = {
            header: ws.cell(row=5, column=idx + 1).value
            for idx, header in enumerate(headers)
        }

        assert values["单双色"] == "单色"
        assert values["产品编号"] == "T7151RYKUMH0001"
        assert values["数量"] == 3
        assert values["单价"] == 21200
        assert values["PIProductcategories"] == "O9 SHS"
        assert values["PIExterior"] == "Silver (KU)"
        assert values["一次内销单价"] == 20200
        assert dual_values["单双色"] == "拼色"
        assert dual_values["PIProductcategories"] == "J7 HEV"
    finally:
        workbook.close()
