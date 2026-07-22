from __future__ import annotations

import gc
import hashlib
import json
import os
import re
import signal
import shlex
import shutil
import subprocess
import sys
import threading
import time
import traceback
from contextlib import ExitStack, contextmanager
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote
from uuid import uuid4

import pandas as pd
from fastapi import HTTPException, UploadFile

from app.core.config import (
    JATO_MONTHLY_UPDATE_JOB_ROOT,
    JATO_MONTHLY_UPDATE_UPLOAD_CHUNK_SIZE_BYTES,
    JATO_MONTHLY_UPDATE_UPLOAD_MAX_BYTES,
    PROJECT_ROOT,
)
from app.services.hermes_pipeline_status_service import write_pipeline_status

MONTHLY_UPDATE_JOB_ROOT = JATO_MONTHLY_UPDATE_JOB_ROOT
RAW_DATA_ROOT = PROJECT_ROOT / "01_RAW_DATA"
BASELINE_ROOT = RAW_DATA_ROOT / "baseline"
PATCHES_ROOT = RAW_DATA_ROOT / "patches"
HISTORY_ARCHIVE_ROOT = RAW_DATA_ROOT / "historyDataArchive"
PREPARE_SCRIPT_PATH = (
    PROJECT_ROOT / "03_Scripts" / "data_pipeline" / "prepare_monthly_raw_update.py"
)
REBUILD_SCRIPT_PATH = (
    PROJECT_ROOT / "03_Scripts" / "data_pipeline" / "rebuild_from_parquet.py"
)
PRECOMPUTE_SUMMARIES_SCRIPT_PATH = (
    PROJECT_ROOT / "03_Scripts" / "data_pipeline" / "precompute_summaries.py"
)
SINGLE_COUNTRY_ETL_SCRIPT_PATH = PROJECT_ROOT / "03_Scripts" / "elt_worker.py"
MONTHLY_WORKER_SCRIPT_PATH = PROJECT_ROOT / "03_Scripts" / "jato_monthly_worker.py"
STATE_FILENAME = "job_state.json"
LOG_FILENAME = "job.log"
UPLOAD_STATE_FILENAME = "upload_state.json"
REVIEW_BUNDLE_FILENAME = "review_bundle.json"
REVIEW_BUNDLE_SCHEMA_VERSION = 3
CANDIDATE_ARTIFACT_STAT_SIGNATURE_VERSION = 2
CANDIDATE_ARTIFACT_FIELDS = (
    ("parquet", "stagingOutputPath"),
    ("manifest", "manifestPath"),
    ("partition", "partitionOutputPath"),
    ("fingerprint", "fingerprintPath"),
    ("refreshReport", "refreshReportPath"),
    ("summaries", "summariesOutputPath"),
)
PARTITION_SCOPED_CANDIDATE_SCOPES = frozenset(
    {"target_country_partition_only", "target_country_partitions_only"}
)
CANCEL_REQUEST_FILENAME = "cancel.request.json"
WORKER_LOCK_FILENAME = "worker.lock"
INGESTION_LOCK_FILENAME = "ingestion.lock"
RECOVERY_LOCK_FILENAME = "recovery.lock"
UPLOAD_INITIATE_LOCK_FILENAME = "upload-initiate.lock"
UPLOAD_STATE_LOCK_FILENAME = "state.lock"
JOB_STATE_LOCK_FILENAME = "state.lock"
ACTIVE_BUNDLE_LOCK_FILENAME = "active-bundle.lock"
ACTIVE_TRANSACTION_FILENAME = "active_transaction.json"
BASELINE_PROMOTION_STATE_FILENAME = "baseline_promotion_state.json"
BASELINE_PROMOTION_LOCK_FILENAME = "baseline-promotion.lock"
BASELINE_INSTALL_JOURNAL_FILENAME = "baseline_install_journal.json"
MAINTENANCE_COORDINATION_LOCK_FILENAME = "maintenance-coordination.lock"
UPLOAD_ASSEMBLY_BUFFER_BYTES = 1024 * 1024
DIGEST_WORKER_STALE_GRACE_SECONDS = 15
DIGEST_WORKER_BASE_TIMEOUT_SECONDS = 10 * 60
DIGEST_WORKER_BASE_SIZE_BYTES = 16 * 1024 * 1024
DIGEST_WORKER_EXTRA_SECONDS_PER_MIB = 4
DIGEST_WORKER_MAX_SECONDS = 45 * 60
DIGEST_EXIT_RECEIPT_GRACE_SECONDS = 5
DIGEST_ATTEMPT_DIRNAME = "digest_attempts"
DIGEST_ATTEMPT_LOG_TAIL_BYTES = 16 * 1024
DIGEST_RETRYABLE_FAILURE_CODES = frozenset(
    {
        "DIGEST_TIMEOUT",
        "DIGEST_WORKER_LOST",
        "DIGEST_WORKER_SIGNALLED",
        "DIGEST_WORKER_EXITED",
        "DIGEST_RESULT_MISSING",
        "DIGEST_WORKER_UNAVAILABLE",
    }
)
UPLOAD_SESSION_STALE_SECONDS = 24 * 60 * 60
BASELINE_XLSX_MAX_ROWS = 1_048_576
BASELINE_EXPORT_BATCH_ROWS = 2_048
SMART_MERGE_SCAN_BATCH_ROWS = 4_096
SMART_MERGE_HASH_BUCKET_COUNT = 128
SMART_MERGE_MAX_BUCKET_ROWS = 16_384
MONTHLY_WORKER_DEFAULT_MEMORY_LIMIT_BYTES = 6 * 1024 * 1024 * 1024
MONTH_PATTERN = re.compile(r"(20\d{2})[-./]?(0?[1-9]|1[0-2])")
CODE_BLOCK_PATTERN = re.compile(r"```bash\s*(.*?)\s*```", re.DOTALL)
ALLOWED_UPLOAD_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
UPLOAD_CHUNK_SIZE_BYTES = JATO_MONTHLY_UPDATE_UPLOAD_CHUNK_SIZE_BYTES
UPLOAD_MAX_BYTES = JATO_MONTHLY_UPDATE_UPLOAD_MAX_BYTES
UPLOAD_INTERNAL_FILENAME_DIGEST_LENGTH = 24
DEFAULT_UPLOAD_SHEET_NAME = "Data Export"
MONTH_COLUMN_PATTERN = re.compile(
    r"^\d{4}\s(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)$",
    re.IGNORECASE,
)
YEAR_COLUMN_PATTERN = re.compile(r"^\d{4}$")
YTD_COLUMN_PATTERN = re.compile(r"^YTD\s+\d{4}\s+\([A-Za-z]{3}\)$", re.IGNORECASE)
BATCH_ID_PATTERN = re.compile(r"^(20\d{2}-\d{2})-r(\d+)$")
RECOVERY_KEY_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._:-]{7,127}$")
REVIEW_REFRESH_REQUEST_ID_PATTERN = re.compile(
    r"^[A-Za-z0-9][A-Za-z0-9._:-]{7,127}$"
)
SMART_MERGE_RESUME_REQUEST_ID_PATTERN = re.compile(
    r"^[A-Za-z0-9][A-Za-z0-9._:-]{7,127}$"
)
RECOVERY_ALLOWED_FAILURE_CATEGORIES = frozenset({"platform", "resource"})
COUNTRY_COLUMN_CANDIDATES = ("国家", "country")
SAFE_CLEANUP_TIER = "safe"
CAUTIOUS_CLEANUP_TIER = "cautious"
PROTECTED_CLEANUP_TIER = "protected"
ALLOWED_CLEANUP_TIERS = {SAFE_CLEANUP_TIER, CAUTIOUS_CLEANUP_TIER}
SALES_DOUBLING_RATIO_MIN = 1.98
SALES_DOUBLING_RATIO_MAX = 2.02
SALES_DOUBLING_MIN_REFERENCE_SALES = 1000.0
SALES_DOUBLING_MIN_ABSOLUTE_DELTA = 1000.0
SALES_DOUBLING_MIN_MONTH_COUNT = 2
SALES_DOUBLING_SAMPLE_LIMIT = 6
DEPRECATED_OPTIONAL_STATIC_COLUMNS = frozenset(
    {
        "Base price",
        "CO2 level - (g/km) combined",
        "MSRP including delivery charge",
        "Maximum power kW",
        "WLTP Emission combined",
        "cargo volume (l)",
        "curb weight (kg)",
    }
)
SC011_LATEST_WASHED_CONFIRMATION_SCOPE = {
    "comparedThrough": "2026 Mar",
    "expectedComparedMonthCount": 39,
    "expectedComparedMonthsSha256": (
        "d9cfcaa2bfd045a14c9ad0663be706bc9234bea64385cf46db11b0fb996e69f7"
    ),
    "sourceUploadSha256": (
        "c12e4e1a58e7d292eb6aef6bdd9c34d0632449536d5351880c35142fa38b0453"
    ),
    "approvedBy": "JATO business owner",
    "approvalReference": "user-confirmed-use-latest-washed-data-2026-07-20",
}
CONFIRMED_SC011_RECLASSIFICATIONS = tuple(
    {
        **SC011_LATEST_WASHED_CONFIRMATION_SCOPE,
        **mapping,
    }
    for mapping in (
        {
            "confirmationId": "CZ-AUDI-Q6-SB-TO-Q6-20260720",
            "country": "捷克",
            "source": {"Make": "AUDI", "Model": "Q6 SPORTBACK E-TRON"},
            "target": {"Make": "AUDI", "Model": "Q6 E-TRON"},
            "expectedMonthlyTransfer": {"2025 Nov": 1},
            "expectedTotal": 1,
        },
        {
            "confirmationId": "CZ-BAIC-X5-TO-X7-20260720",
            "country": "捷克",
            "source": {"Make": "BAIC", "Model": "X5"},
            "target": {"Make": "BAIC", "Model": "X7"},
            "expectedMonthlyTransfer": {"2023 Dec": 1},
            "expectedTotal": 1,
        },
        {
            "confirmationId": "CZ-BYD-DOLPHIN-TO-SURF-20260720",
            "country": "捷克",
            "source": {"Make": "BYD", "Model": "DOLPHIN"},
            "target": {"Make": "BYD", "Model": "DOLPHIN SURF"},
            "expectedMonthlyTransfer": {
                "2025 Nov": 2,
                "2025 Dec": 2,
                "2026 Mar": 2,
            },
            "expectedTotal": 6,
        },
        {
            "confirmationId": "CZ-BYD-SEAL05-TO-SEAL5-20260720",
            "country": "捷克",
            "source": {"Make": "BYD", "Model": "SEAL 05"},
            "target": {"Make": "BYD", "Model": "SEAL 5"},
            "expectedMonthlyTransfer": {
                "2025 Jun": 16,
                "2025 Aug": 1,
                "2025 Sep": 2,
                "2025 Oct": 2,
                "2025 Dec": 3,
                "2026 Jan": 1,
            },
            "expectedTotal": 25,
        },
        {
            "confirmationId": "CZ-FIAT-PANDINA-TO-PANDA-20260720",
            "country": "捷克",
            "source": {"Make": "FIAT", "Model": "PANDINA"},
            "target": {"Make": "FIAT", "Model": "PANDA"},
            "expectedMonthlyTransfer": {"2023 Mar": 1},
            "expectedTotal": 1,
        },
        {
            "confirmationId": "CZ-FORD-CONNECT-TO-CUSTOM-20260720",
            "country": "捷克",
            "source": {"Make": "FORD", "Model": "TOURNEO CONNECT"},
            "target": {"Make": "FORD", "Model": "TOURNEO CUSTOM"},
            "expectedMonthlyTransfer": {"2026 Mar": 1},
            "expectedTotal": 1,
        },
        {
            "confirmationId": "CZ-FORTHING-T5-TO-DFSK-T5-EVO-20260720",
            "country": "捷克",
            "source": {"Make": "FORTHING", "Model": "T5"},
            "target": {"Make": "DFSK", "Model": "T5 EVO"},
            "expectedMonthlyTransfer": {
                "2025 Apr": 3,
                "2025 May": 3,
                "2025 Jun": 8,
                "2025 Jul": 4,
                "2025 Aug": 4,
                "2025 Sep": 6,
                "2025 Nov": 3,
                "2025 Dec": 19,
                "2026 Jan": 1,
            },
            "expectedTotal": 51,
            "approvalReference": "user-confirmed-2026-07-20",
        },
        {
            "confirmationId": "CZ-KGM-TORRES-EVX-TO-TORRES-20260720",
            "country": "捷克",
            "source": {"Make": "KGM", "Model": "TORRES EVX"},
            "target": {"Make": "KGM", "Model": "TORRES"},
            "expectedMonthlyTransfer": {"2025 Dec": 2},
            "expectedTotal": 2,
        },
        {
            "confirmationId": "CZ-MERCEDES-GLC-COUPE-TO-GLC-20260720",
            "country": "捷克",
            "source": {"Make": "MERCEDES", "Model": "GLC COUPE"},
            "target": {"Make": "MERCEDES", "Model": "GLC"},
            "expectedMonthlyTransfer": {"2026 Mar": 38},
            "expectedTotal": 38,
        },
        {
            "confirmationId": "CZ-MINI-MINI-TO-COOPER-20260720",
            "country": "捷克",
            "source": {"Make": "MINI", "Model": "MINI"},
            "target": {"Make": "MINI", "Model": "COOPER"},
            "expectedMonthlyTransfer": {
                "2025 Feb": 3,
                "2025 Apr": 3,
                "2025 Jul": 5,
                "2025 Aug": 4,
                "2025 Sep": 3,
                "2025 Nov": 1,
                "2025 Dec": 1,
            },
            "expectedTotal": 20,
        },
        {
            "confirmationId": "CZ-OMODA-5-TO-5-EV-20260720",
            "country": "捷克",
            "source": {"Make": "OMODA", "Model": "5"},
            "target": {"Make": "OMODA", "Model": "5 EV"},
            "expectedMonthlyTransfer": {"2026 Jan": 1},
            "expectedTotal": 1,
        },
        {
            "confirmationId": "CZ-PEUGEOT-EXPERT-TO-TRAVELLER-20260720",
            "country": "捷克",
            "source": {"Make": "PEUGEOT", "Model": "EXPERT"},
            "target": {"Make": "PEUGEOT", "Model": "TRAVELLER"},
            "expectedMonthlyTransfer": {"2026 Mar": 3},
            "expectedTotal": 3,
        },
        {
            "confirmationId": "CZ-PORSCHE-CAYENNE-COUPE-TO-CAYENNE-20260720",
            "country": "捷克",
            "source": {"Make": "PORSCHE", "Model": "CAYENNE COUPE"},
            "target": {"Make": "PORSCHE", "Model": "CAYENNE"},
            "expectedMonthlyTransfer": {"2026 Mar": 10},
            "expectedTotal": 10,
        },
        {
            "confirmationId": "DK-AUDI-Q6-TO-Q6-SB-20260720",
            "country": "丹麦",
            "source": {"Make": "AUDI", "Model": "Q6 E-TRON"},
            "target": {"Make": "AUDI", "Model": "Q6 SPORTBACK E-TRON"},
            "expectedMonthlyTransfer": {"2026 Mar": 1},
            "expectedTotal": 1,
        },
        {
            "confirmationId": "DK-FORD-TRANSIT-TO-TOURNEO-CUSTOM-20260720",
            "country": "丹麦",
            "source": {"Make": "FORD", "Model": "TRANSIT"},
            "target": {"Make": "FORD", "Model": "TOURNEO CUSTOM"},
            "expectedMonthlyTransfer": {"2026 Jan": 1},
            "expectedTotal": 1,
        },
        {
            "confirmationId": "DK-MERCEDES-GLC-COUPE-TO-GLC-20260720",
            "country": "丹麦",
            "source": {"Make": "MERCEDES", "Model": "GLC COUPE"},
            "target": {"Make": "MERCEDES", "Model": "GLC"},
            "expectedMonthlyTransfer": {
                "2026 Jan": 4,
                "2026 Feb": 5,
                "2026 Mar": 7,
            },
            "expectedTotal": 16,
        },
        {
            "confirmationId": "DK-RENAULT-5-TO-4-20260720",
            "country": "丹麦",
            "source": {"Make": "RENAULT", "Model": "5"},
            "target": {"Make": "RENAULT", "Model": "4"},
            "expectedMonthlyTransfer": {
                "2026 Jan": 87,
                "2026 Mar": 7,
            },
            "expectedTotal": 94,
        },
    )
)
STATIC_CARRY_FORWARD_KEY_CANDIDATES = (
    "国家",
    "Country",
    "Countries",
    "Registration type",
    "Make group",
    "Make",
    "Model group",
    "Model",
    "Version name",
    "Powertrain type",
    "Trim level",
    "Body type",
    "Fuel type",
    "Transmission type",
    "Driven wheels",
    "Battery type",
    "Battery kwh",
    "Useable battery kilowatt hour (kWh)",
    "Battery range",
    "Seating capacity",
    "length (mm)",
    "width (mm)",
    "height (mm)",
    "wheelbase (mm)",
)
HISTORICAL_DIMENSION_ALIASES: tuple[tuple[str, ...], ...] = (
    ("Registration type",),
    ("Segment", "细分市场（按车长）", "细分市场"),
    ("Body type", "Body style"),
    ("Powertrain type", "Powertrain", "动总规整"),
    ("Fuel type", "Fuel"),
    ("Version name", "Trim level"),
)
HISTORICAL_RECLASSIFICATION_DECISIONS = frozenset(
    {"use_latest", "keep_active"}
)
HISTORICAL_RECLASSIFICATION_DECISION_ORDER = (
    "use_latest",
    "keep_active",
)
HISTORICAL_RECLASSIFICATION_VALUE_LIMIT = 8
HISTORICAL_RECLASSIFICATION_EXACT_CHANGE_LIMIT = 20
_WRITE_LOCK = threading.Lock()
_RUNNING_THREADS: dict[str, threading.Thread] = {}
RUNNING_JOB_STATUSES = {"queued", "running"}
PROCESS_TERMINATE_GRACE_SECONDS = 8
RUNNING_LOG_STALE_SECONDS = 15 * 60


class _JobCancelled(RuntimeError):
    pass


class _JobResourceKilled(RuntimeError):
    pass


def _utc_now() -> datetime:
    return datetime.now(UTC)


def _failure_digest_from_exception(
    *,
    phase: str,
    exc: BaseException,
) -> dict[str, Any]:
    message = str(exc)
    signal_match = re.search(r"退出码\s+(-\d+)", message)
    if signal_match:
        return_code = int(signal_match.group(1))
        signal_number = abs(return_code)
        try:
            signal_name = signal.Signals(signal_number).name
        except ValueError:
            signal_name = f"SIGNAL_{signal_number}"
        return {
            "code": "SUBPROCESS_SIGNALED",
            "category": "resource",
            "phase": phase,
            "retryable": False,
            "message": message,
            "sourceFeedback": (
                "这不是已证明的数据错误；子进程被系统信号终止。"
                "请平台管理员先检查 worker cgroup memory.events 与日志，再决定是否重试。"
            ),
            "technicalDetail": {
                "returnCode": return_code,
                "signalNumber": signal_number,
                "signalName": signal_name,
            },
            "nextAction": "inspect_worker_resources",
        }
    memory_markers = (
        "arrowmemoryerror",
        "out of memory",
        "cannot allocate memory",
        "unable to allocate",
        "realloc of size",
        "memory limit",
        "std::bad_alloc",
    )
    memory_detail = f"{type(exc).__name__}: {message}".casefold()
    if isinstance(exc, MemoryError) or any(
        marker in memory_detail for marker in memory_markers
    ):
        smart_merge_failure = (
            phase == "building_review"
            or phase.startswith("smart_merge")
        )
        return {
            "code": "MEMORY_LIMIT_EXCEEDED",
            "category": "resource",
            "phase": phase,
            "retryable": True,
            "message": message,
            "sourceFeedback": (
                "无需重新洗数或重新上传；这不是源数据错误。"
                + (
                    "Candidate 和决策已保留，请仅续跑 Smart Merge。"
                    if smart_merge_failure
                    else "请保留 Candidate 和决策，仅续跑失败阶段。"
                )
            ),
            "technicalDetail": type(exc).__name__,
            "nextAction": (
                "resume_smart_merge"
                if smart_merge_failure
                else "resume_failed_stage"
            ),
        }
    return {
        "code": "JOB_STEP_FAILED",
        "category": "processing",
        "phase": phase,
        "retryable": False,
        "message": message,
        "sourceFeedback": (
            "请根据 Review 的 ruleId、国家、月份和字段反馈修正源文件；"
            "若无 sourceFeedback，再交由平台管理员排查。"
        ),
        "technicalDetail": type(exc).__name__,
        "nextAction": "review_failure_report",
    }


def _relative_to_project(path: Path | None) -> str | None:
    if path is None:
        return None
    try:
        return str(path.resolve().relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path.resolve())


def _project_path(value: str | Path | None) -> Path | None:
    if value is None:
        return None
    normalized = str(value).strip()
    if not normalized:
        return None
    candidate = Path(normalized)
    return candidate if candidate.is_absolute() else PROJECT_ROOT / candidate


def _job_dir(job_id: str) -> Path:
    return MONTHLY_UPDATE_JOB_ROOT / job_id


def _job_state_path(job_id: str) -> Path:
    return _job_dir(job_id) / STATE_FILENAME


def _job_log_path(job_id: str) -> Path:
    return _job_dir(job_id) / LOG_FILENAME


def _job_review_bundle_path(job_id: str) -> Path:
    return _job_dir(job_id) / REVIEW_BUNDLE_FILENAME


def _job_cancel_request_path(job_id: str) -> Path:
    return _job_dir(job_id) / CANCEL_REQUEST_FILENAME


def _job_state_lock_path(job_id: str) -> Path:
    return _job_dir(job_id) / JOB_STATE_LOCK_FILENAME


def _maintenance_dir() -> Path:
    return MONTHLY_UPDATE_JOB_ROOT / "_maintenance"


def _baseline_promotion_state_path() -> Path:
    return _maintenance_dir() / BASELINE_PROMOTION_STATE_FILENAME


def _baseline_promotion_lock_path() -> Path:
    return _maintenance_dir() / BASELINE_PROMOTION_LOCK_FILENAME


def _baseline_install_journal_path() -> Path:
    return _maintenance_dir() / BASELINE_INSTALL_JOURNAL_FILENAME


def _maintenance_coordination_lock_path() -> Path:
    return _maintenance_dir() / MAINTENANCE_COORDINATION_LOCK_FILENAME


def _worker_lock_path() -> Path:
    return MONTHLY_UPDATE_JOB_ROOT / WORKER_LOCK_FILENAME


def _upload_initiate_lock_path() -> Path:
    return _upload_session_root() / UPLOAD_INITIATE_LOCK_FILENAME


def _upload_state_lock_path(upload_id: str) -> Path:
    return _upload_session_dir(upload_id) / UPLOAD_STATE_LOCK_FILENAME


def _active_bundle_lock_path() -> Path:
    return _processed_data_root() / ACTIVE_BUNDLE_LOCK_FILENAME


@contextmanager
def _exclusive_file_lock(path: Path, *, blocking: bool = True) -> Any:
    """Cross-process lock used by multiple Uvicorn and worker processes."""
    try:
        import fcntl
    except ImportError:  # pragma: no cover - production is Linux/POSIX
        yield True
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a+", encoding="utf-8") as handle:
        operation = fcntl.LOCK_EX | (0 if blocking else fcntl.LOCK_NB)
        try:
            fcntl.flock(handle.fileno(), operation)
        except BlockingIOError:
            yield False
            return
        try:
            yield True
        finally:
            fcntl.flock(handle.fileno(), fcntl.LOCK_UN)


@contextmanager
def _exclusive_worker_cycle() -> Any:
    with _exclusive_file_lock(_worker_lock_path(), blocking=False) as acquired:
        yield acquired


def _processed_data_root() -> Path:
    return PROJECT_ROOT / "04_Processed_data"


def _active_data_paths() -> dict[str, Path]:
    processed_root = _processed_data_root()
    return {
        "parquet": processed_root / "jato_full_archive.parquet",
        "manifest": processed_root / "manifest.json",
        "partition": processed_root / "partitioned_dataset_v1",
        "fingerprint": processed_root / "dataset_fingerprint.json",
        "refreshReport": processed_root / "refresh_job_report.json",
        "summaries": processed_root / "summaries",
        "backupRoot": processed_root / ".refresh_backups",
    }


def _active_parquet_schema_reference(
    active_paths: dict[str, Path] | None = None,
) -> Path:
    """Return one immutable active parquet schema source for bounded XLSX ETL."""
    paths = active_paths or _active_data_paths()
    active_parquet = paths["parquet"]
    if active_parquet.is_file():
        return active_parquet
    partition_root = paths["partition"]
    if partition_root.exists():
        for parquet_path in sorted(partition_root.rglob("*.parquet")):
            if parquet_path.is_file():
                return parquet_path
    raise RuntimeError(
        "部分国家刷新缺少 active parquet schema；"
        "拒绝猜测 washed 字段类型。"
    )


def _partial_country_streaming_cli_args(
    *,
    upload_suffix: str,
    active_paths: dict[str, Path],
) -> list[str]:
    if upload_suffix.lower() not in {".xlsx", ".xlsm"}:
        # Legacy .xls remains on the existing pandas path. Its worksheet row
        # limit is 65,536, so it does not carry the 762k-row XLSX allocation
        # risk that this bounded path addresses.
        return []
    active_schema_reference = _active_parquet_schema_reference(active_paths)
    return [
        "--streaming-xlsx",
        "--schema-from-parquet",
        str(active_schema_reference.resolve()),
    ]


ACTIVE_BUNDLE_KEYS = (
    "parquet",
    "manifest",
    "partition",
    "fingerprint",
    "refreshReport",
    "summaries",
)


def _remove_file_or_tree(path: Path) -> None:
    if path.is_dir() and not path.is_symlink():
        shutil.rmtree(path)
    else:
        path.unlink(missing_ok=True)


def _copy_file_or_tree(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if source.is_dir():
        shutil.copytree(source, destination)
    else:
        shutil.copy2(source, destination)


def _rewrite_staged_summaries_manifest_paths(
    *,
    staged_summaries_path: Path,
    active_summaries_path: Path,
) -> None:
    manifest_path = staged_summaries_path / "summaries_manifest.json"
    if not manifest_path.exists():
        return
    manifest = _read_json(manifest_path)
    summaries = manifest.get("summaries")
    if isinstance(summaries, dict):
        for raw_entry in summaries.values():
            if not isinstance(raw_entry, dict):
                continue
            for file_type in ("csv", "parquet"):
                raw_path = str(raw_entry.get(file_type) or "").strip()
                if raw_path:
                    raw_entry[file_type] = str(
                        active_summaries_path / Path(raw_path).name
                    )
    _write_json(manifest_path, manifest)


def _stage_active_bundle_sources(
    *,
    source_paths: dict[str, Path],
    active_paths: dict[str, Path],
    transaction_id: str,
    optional_missing_keys: set[str] | None = None,
) -> tuple[Path, dict[str, Path | None]]:
    """Copy a complete release beside active before changing any live path."""
    optional = optional_missing_keys or set()
    staging_root = _processed_data_root() / (
        f".jato-active-stage-{transaction_id}-{uuid4().hex[:8]}"
    )
    staging_root.mkdir(parents=True, exist_ok=False)
    staged: dict[str, Path | None] = {}
    try:
        for key in ACTIVE_BUNDLE_KEYS:
            source = source_paths[key]
            if not source.exists():
                if key in optional:
                    staged[key] = None
                    continue
                raise RuntimeError(f"active bundle source missing: {key}={source}")
            destination = staging_root / active_paths[key].name
            _copy_file_or_tree(source, destination)
            if key == "summaries":
                _rewrite_staged_summaries_manifest_paths(
                    staged_summaries_path=destination,
                    active_summaries_path=active_paths["summaries"],
                )
            staged[key] = destination
        return staging_root, staged
    except Exception:
        shutil.rmtree(staging_root, ignore_errors=True)
        raise


def _swap_staged_active_bundle(
    *,
    staged_paths: dict[str, Path | None],
    active_paths: dict[str, Path],
    backup_dir: Path,
    transaction_metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Fast, recoverable switch after all heavy copies have completed.

    Existing active paths are renamed into the durable backup. If any later
    rename fails, every already-processed path is restored before returning.
    """
    backup_dir.mkdir(parents=True, exist_ok=False)
    processed_keys: list[str] = []
    installed_keys: set[str] = set()
    backed_up_keys: set[str] = set()
    active_existed = {
        key: active_paths[key].exists()
        for key in ACTIVE_BUNDLE_KEYS
    }
    staging_roots = {
        str(path.parent)
        for path in staged_paths.values()
        if path is not None
    }
    journal_path = backup_dir / ACTIVE_TRANSACTION_FILENAME
    transaction = dict(transaction_metadata or {})
    transaction_id = str(
        transaction.get("transactionId")
        or f"jato-active-{uuid4().hex}"
    )
    transaction["transactionId"] = transaction_id
    journal: dict[str, Any] = {
        "status": "switching",
        "startedAt": _utc_now().isoformat(),
        "transaction": transaction,
        "activeExisted": active_existed,
        "processedKeys": [],
        "installedKeys": [],
        "backedUpKeys": [],
        "stagingRoots": sorted(staging_roots),
    }
    _write_json(journal_path, journal)
    try:
        for key in ACTIVE_BUNDLE_KEYS:
            active_path = active_paths[key]
            backup_path = backup_dir / active_path.name
            staged_path = staged_paths.get(key)
            if active_path.exists():
                os.replace(active_path, backup_path)
                backed_up_keys.add(key)
            if staged_path is not None:
                os.replace(staged_path, active_path)
                installed_keys.add(key)
            processed_keys.append(key)
            journal["processedKeys"] = list(processed_keys)
            journal["installedKeys"] = sorted(installed_keys)
            journal["backedUpKeys"] = sorted(backed_up_keys)
            _write_json(journal_path, journal)
    except Exception:
        for key in reversed(ACTIVE_BUNDLE_KEYS):
            active_path = active_paths[key]
            backup_path = backup_dir / active_path.name
            if key in installed_keys and active_path.exists():
                _remove_file_or_tree(active_path)
            if key in backed_up_keys and backup_path.exists():
                os.replace(backup_path, active_path)
        shutil.rmtree(backup_dir, ignore_errors=True)
        raise
    # Do not mark committed until the caller has durably persisted publication
    # or rollback metadata in the job state. Recovery can then distinguish a
    # completed switch from a worker crash between rename and state write.
    journal["status"] = "switched"
    journal["switchedAt"] = _utc_now().isoformat()
    _write_json(journal_path, journal)
    return {
        "activeTransactionId": transaction_id,
        "processedKeys": processed_keys,
        "installedKeys": sorted(installed_keys),
        "backedUpKeys": sorted(backed_up_keys),
    }


def _active_transaction_has_durable_job_commit(
    journal: dict[str, Any],
) -> bool:
    transaction = journal.get("transaction")
    if not isinstance(transaction, dict):
        return False
    transaction_id = str(transaction.get("transactionId") or "")
    job_id = str(transaction.get("jobId") or "")
    operation_type = str(transaction.get("operationType") or "")
    if not transaction_id or not job_id:
        return False
    try:
        payload = _load_job_state(job_id)
    except HTTPException:
        return False
    publication = payload.get("publication")
    if not isinstance(publication, dict):
        return False
    if operation_type == "publish":
        return bool(
            publication.get("publishedAt")
            and str(publication.get("activeTransactionId") or "")
            == transaction_id
        )
    if operation_type == "rollback":
        return bool(
            publication.get("rolledBackAt")
            and str(publication.get("rollbackActiveTransactionId") or "")
            == transaction_id
        )
    return False


def _commit_active_transaction(
    *,
    journal_path: Path,
    transaction_id: str,
) -> None:
    journal = _read_json(journal_path)
    transaction = journal.get("transaction")
    persisted_transaction_id = (
        str(transaction.get("transactionId") or "")
        if isinstance(transaction, dict)
        else ""
    )
    if persisted_transaction_id != transaction_id:
        raise RuntimeError(
            "active transaction journal identity mismatch: "
            f"expected={transaction_id}, actual={persisted_transaction_id or '-'}"
        )
    if str(journal.get("status") or "") == "committed":
        return
    if str(journal.get("status") or "") != "switched":
        raise RuntimeError(
            "active transaction cannot commit from status "
            f"{journal.get('status') or '-'}"
        )
    journal["status"] = "committed"
    journal["committedAt"] = _utc_now().isoformat()
    _write_json(journal_path, journal)


def _recover_incomplete_active_transactions(
    active_paths: dict[str, Path],
) -> list[str]:
    """Restore the old bundle after a worker died during the rename window."""
    recovered: list[str] = []
    backup_root = active_paths["backupRoot"]
    if not backup_root.exists():
        return recovered
    for journal_path in sorted(
        backup_root.glob(f"*/{ACTIVE_TRANSACTION_FILENAME}")
    ):
        try:
            journal = _read_json(journal_path)
        except Exception:
            continue
        status = str(journal.get("status") or "")
        if status == "switched" and _active_transaction_has_durable_job_commit(
            journal
        ):
            journal["status"] = "committed"
            journal["committedAt"] = _utc_now().isoformat()
            journal["committedByRecovery"] = True
            _write_json(journal_path, journal)
            continue
        if status not in {"switching", "switched"}:
            continue
        backup_dir = journal_path.parent
        active_existed = journal.get("activeExisted")
        existed_map = (
            active_existed if isinstance(active_existed, dict) else {}
        )
        for key in reversed(ACTIVE_BUNDLE_KEYS):
            active_path = active_paths[key]
            backup_path = backup_dir / active_path.name
            if backup_path.exists():
                if active_path.exists():
                    _remove_file_or_tree(active_path)
                os.replace(backup_path, active_path)
            elif not bool(existed_map.get(key)) and active_path.exists():
                _remove_file_or_tree(active_path)
        for raw_root in journal.get("stagingRoots", []):
            if isinstance(raw_root, str) and raw_root:
                shutil.rmtree(Path(raw_root), ignore_errors=True)
        journal["status"] = "recovered"
        journal["recoveredAt"] = _utc_now().isoformat()
        _write_json(journal_path, journal)
        recovered.append(str(backup_dir))
    return recovered


def _recover_incomplete_active_transactions_if_possible() -> list[str]:
    with _exclusive_file_lock(
        _active_bundle_lock_path(),
        blocking=False,
    ) as acquired:
        if not acquired:
            return []
        return _recover_incomplete_active_transactions(_active_data_paths())


def _active_dataset_version() -> str:
    """Fingerprint the active bundle without materializing the archive."""
    paths = _active_data_paths()
    hasher = hashlib.sha256()
    observed = False
    for key in ("manifest", "fingerprint", "refreshReport"):
        path = paths[key]
        if not path.exists():
            continue
        observed = True
        hasher.update(key.encode("utf-8"))
        hasher.update(_sha256_hex_for_path(path).encode("ascii"))
    parquet_path = paths["parquet"]
    if parquet_path.exists():
        observed = True
        stat = parquet_path.stat()
        hasher.update(
            f"parquet:{stat.st_size}:{stat.st_mtime_ns}".encode("ascii")
        )
    partition_root = paths["partition"]
    if partition_root.exists():
        for file_path in sorted(
            path
            for path in partition_root.rglob("*")
            if path.is_file()
        ):
            stat = file_path.stat()
            relative = file_path.relative_to(partition_root)
            hasher.update(
                f"{relative}:{stat.st_size}:{stat.st_mtime_ns}".encode("utf-8")
            )
            observed = True
    summaries_root = paths["summaries"]
    if summaries_root.exists():
        hasher.update(b"summaries")
        hasher.update(_sha256_hex_for_tree(summaries_root).encode("ascii"))
        observed = True
    if not observed:
        raise HTTPException(
            status_code=409,
            detail="active 数据集不存在，无法绑定 Review 基线。",
        )
    return hasher.hexdigest()


def _time_sort_key(label: str) -> tuple[int, int, str]:
    text = str(label).strip()
    if MONTH_COLUMN_PATTERN.match(text):
        parsed = datetime.strptime(text.title(), "%Y %b")
        return (parsed.year, parsed.month, text)
    if YEAR_COLUMN_PATTERN.fullmatch(text):
        return (int(text), 0, text)
    return (9999, 12, text)


def _series_has_data(series: pd.Series) -> bool:
    if series.empty:
        return False
    if (
        pd.api.types.is_string_dtype(series.dtype)
        or pd.api.types.is_object_dtype(series.dtype)
    ):
        normalized = series.astype("string").fillna("").str.strip()
        return bool((normalized != "").any())
    return bool(series.notna().any())


def _series_missing_value_mask(series: pd.Series) -> pd.Series:
    missing = series.isna()
    if (
        pd.api.types.is_string_dtype(series.dtype)
        or pd.api.types.is_object_dtype(series.dtype)
    ):
        missing = missing | series.astype("string").fillna("").str.strip().eq("")
    return missing


def _find_country_column(columns: list[str]) -> str | None:
    lookup = {str(column).strip().lower(): str(column) for column in columns}
    for candidate in COUNTRY_COLUMN_CANDIDATES:
        column = lookup.get(candidate.lower())
        if column:
            return column
    return None


def _detect_month_columns(columns: list[str]) -> list[str]:
    return sorted(
        [
            str(column).strip()
            for column in columns
            if MONTH_COLUMN_PATTERN.match(str(column).strip())
        ],
        key=_time_sort_key,
    )


def _collect_dataset_country_latest_months(path: Path) -> dict[str, str | None]:
    try:
        import pyarrow.parquet as pq

        parquet_file = pq.ParquetFile(path)
        schema_columns = [
            str(column).strip()
            for column in parquet_file.schema_arrow.names
        ]
        country_column = _find_country_column(schema_columns)
        month_columns = _detect_month_columns(schema_columns)
        selected_columns = [
            column
            for column in [country_column, *month_columns]
            if column is not None
        ]
        if country_column is None or not month_columns:
            raise ValueError("country/month columns unavailable")
        display_variants: dict[str, list[str]] = {}
        present_months: dict[str, set[str]] = {}
        for batch in parquet_file.iter_batches(
            batch_size=SMART_MERGE_SCAN_BATCH_ROWS,
            columns=selected_columns,
            use_threads=False,
        ):
            frame = batch.to_pandas(use_threads=False)
            frame.columns = [
                str(column).strip()
                for column in frame.columns
            ]
            normalized = (
                frame[country_column]
                .astype("string")
                .fillna("")
                .str.strip()
            )
            country_keys = normalized.str.casefold()
            for country_key, display in zip(
                country_keys,
                normalized,
                strict=False,
            ):
                key = str(country_key)
                rendered = str(display)
                if not key or not rendered:
                    continue
                variants = display_variants.setdefault(key, [])
                if rendered not in variants:
                    variants.append(rendered)
            for country_key, row_indices in frame.groupby(
                country_keys,
                dropna=False,
                sort=False,
            ).groups.items():
                key = str(country_key)
                if not key:
                    continue
                country_months = present_months.setdefault(key, set())
                country_frame = frame.loc[row_indices]
                for month in month_columns:
                    if (
                        month not in country_months
                        and _series_has_data(country_frame[month])
                    ):
                        country_months.add(month)
            del frame
        ambiguous = {
            key: variants
            for key, variants in display_variants.items()
            if len(variants) > 1
        }
        if ambiguous:
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": "ambiguous_logical_country",
                    "message": (
                        f"{path.name} 中同一逻辑国家存在多个大小写/空格"
                        "展示值，继续会造成国家重复累加；请先统一国家字段。"
                    ),
                    "countries": [
                        {
                            "logicalKey": key,
                            "displayValues": variants,
                        }
                        for key, variants in sorted(ambiguous.items())
                    ],
                },
            )
        return {
            variants[0]: (
                max(
                    present_months.get(country_key, set()),
                    key=_time_sort_key,
                )
                if present_months.get(country_key)
                else None
            )
            for country_key, variants in display_variants.items()
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=409,
            detail=f"无法读取 {path.name} 的国家/月字段，不能执行 publish 校验。",
        ) from exc


def _candidate_fingerprint_id(artifacts: dict[str, Any]) -> str:
    """Bind a human approval to the exact candidate that was reviewed."""
    candidate_scope = str(artifacts.get("candidateScope") or "")
    partial_scope = candidate_scope in PARTITION_SCOPED_CANDIDATE_SCOPES
    required_names = (
        {"parquet", "manifest", "refreshReport"}
        if partial_scope
        else {name for name, _field in CANDIDATE_ARTIFACT_FIELDS}
    )
    hasher = hashlib.sha256()
    for artifact_name, artifact_field in CANDIDATE_ARTIFACT_FIELDS:
        raw_path = str(artifacts.get(artifact_field) or "").strip()
        path = _project_path(raw_path) if raw_path else None
        hasher.update(artifact_name.encode("utf-8"))
        if path is None or not path.exists():
            if artifact_name in required_names:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"缺少 candidate {artifact_name}，"
                        "不能生成完整 Review 指纹。"
                    ),
                )
            hasher.update(b":missing")
            continue
        if path.is_dir():
            hasher.update(_sha256_hex_for_tree(path).encode("ascii"))
        elif path.is_file():
            hasher.update(_sha256_hex_for_path(path).encode("ascii"))
        else:
            raise HTTPException(
                status_code=409,
                detail=f"candidate {artifact_name} 不是普通文件或目录。",
            )
    return hasher.hexdigest()


def _candidate_artifact_stat_signature(artifacts: dict[str, Any]) -> str:
    """Build a deployment-stable metadata drift token for HTTP reads.

    The isolated worker still computes the content SHA used by Review approval,
    and Publish recomputes that content SHA before touching active.  This token
    lets GET/approval detect ordinary staging drift without rereading a large
    parquet and every summary file inside a web worker.  inode and ctime are
    intentionally excluded because immutable-release extraction and chown can
    change them without changing candidate content.  The explicit version keeps
    old signatures fail-closed instead of comparing incompatible formats.
    """
    artifact_entries: list[dict[str, Any]] = []
    for artifact_name, artifact_field in CANDIDATE_ARTIFACT_FIELDS:
        raw_path = str(artifacts.get(artifact_field) or "").strip()
        path = _project_path(raw_path) if raw_path else None
        if path is None:
            artifact_entries.append(
                {
                    "artifact": artifact_name,
                    "path": raw_path or None,
                    "kind": "missing",
                }
            )
            continue
        try:
            if path.is_dir():
                files = sorted(
                    candidate
                    for candidate in path.rglob("*")
                    if candidate.is_file()
                )
                file_entries: list[dict[str, Any]] = []
                for candidate in files:
                    stat = candidate.stat()
                    file_entries.append(
                        {
                            "path": str(candidate.relative_to(path)),
                            "size": stat.st_size,
                            "mtimeNs": stat.st_mtime_ns,
                        }
                    )
                artifact_entries.append(
                    {
                        "artifact": artifact_name,
                        "path": _relative_to_project(path),
                        "kind": "directory",
                        "files": file_entries,
                    }
                )
            elif path.is_file():
                stat = path.stat()
                artifact_entries.append(
                    {
                        "artifact": artifact_name,
                        "path": _relative_to_project(path),
                        "kind": "file",
                        "size": stat.st_size,
                        "mtimeNs": stat.st_mtime_ns,
                    }
                )
            else:
                artifact_entries.append(
                    {
                        "artifact": artifact_name,
                        "path": _relative_to_project(path),
                        "kind": "missing",
                    }
                )
        except OSError as exc:
            raise HTTPException(
                status_code=409,
                detail=f"candidate 产物状态读取失败：{artifact_name} ({exc})",
            ) from exc
    encoded = json.dumps(
        artifact_entries,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return (
        f"v{CANDIDATE_ARTIFACT_STAT_SIGNATURE_VERSION}:"
        f"{hashlib.sha256(encoded).hexdigest()}"
    )


def _candidate_artifact_stat_signature_version(signature: str) -> int | None:
    match = re.fullmatch(r"v([1-9][0-9]*):[0-9a-f]{64}", signature.strip())
    return int(match.group(1)) if match else None


def _review_bundle_contract_error(
    review_bundle: dict[str, Any],
) -> str | None:
    schema_version = review_bundle.get("reviewBundleSchemaVersion")
    if (
        type(schema_version) is not int
        or schema_version != REVIEW_BUNDLE_SCHEMA_VERSION
    ):
        return "legacy_review_bundle_schema"
    signature_metadata_version = review_bundle.get(
        "candidateArtifactStatSignatureVersion"
    )
    if (
        type(signature_metadata_version) is not int
        or signature_metadata_version
        != CANDIDATE_ARTIFACT_STAT_SIGNATURE_VERSION
    ):
        return "legacy_stat_signature_metadata"
    signature = str(
        review_bundle.get("candidateArtifactStatSignature") or ""
    )
    if (
        _candidate_artifact_stat_signature_version(signature)
        != CANDIDATE_ARTIFACT_STAT_SIGNATURE_VERSION
    ):
        return "legacy_stat_signature" if signature else "missing_stat_signature"
    return None


def _validate_candidate_summaries_bundle(
    *,
    summaries_path: Path,
    candidate_manifest_path: Path,
) -> dict[str, Any]:
    """Verify staged dashboard summaries belong to the exact full candidate."""
    summaries_manifest_path = summaries_path / "summaries_manifest.json"
    if not summaries_manifest_path.exists():
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "candidate_bundle_invalid",
                "message": "candidate summaries 缺少 summaries_manifest.json。",
            },
        )
    try:
        summaries_manifest = _read_json(summaries_manifest_path)
        candidate_manifest = _read_json(candidate_manifest_path)
    except Exception as exc:
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "candidate_bundle_invalid",
                "message": f"candidate bundle manifest 无法读取：{exc}",
            },
        ) from exc

    expected_rows = int(candidate_manifest.get("rows", -1) or -1)
    summary_rows = int(summaries_manifest.get("originalRowCount", -1) or -1)
    if expected_rows < 0 or summary_rows != expected_rows:
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "candidate_bundle_invalid",
                "message": (
                    "candidate summaries 与 candidate parquet 行数不一致："
                    f"summaries={summary_rows}, candidate={expected_rows}。"
                ),
            },
        )

    incremental = summaries_manifest.get("incremental")
    incremental_mode = (
        str(incremental.get("mode") or "")
        if isinstance(incremental, dict)
        else ""
    )
    if (
        str(summaries_manifest.get("precomputeMode") or "") != "full"
        or incremental_mode != "full"
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "candidate_bundle_invalid",
                "message": "candidate summaries 不是从完整 candidate 全量生成，拒绝 Publish。",
            },
        )

    summary_entries = summaries_manifest.get("summaries")
    if not isinstance(summary_entries, dict) or not summary_entries:
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "candidate_bundle_invalid",
                "message": "candidate summaries manifest 没有任何汇总文件。",
            },
        )

    root = summaries_path.resolve()
    declared_total_rows = 0
    missing_files: list[str] = []
    invalid_files: list[str] = []
    for summary_name, raw_entry in summary_entries.items():
        if not isinstance(raw_entry, dict):
            invalid_files.append(str(summary_name))
            continue
        declared_rows = int(raw_entry.get("rows", -1) or -1)
        declared_total_rows += max(declared_rows, 0)
        for file_type in ("csv", "parquet"):
            raw_path = str(raw_entry.get(file_type) or "").strip()
            if not raw_path:
                missing_files.append(f"{summary_name}.{file_type}")
                continue
            candidate_path = _project_path(raw_path)
            if candidate_path is None:
                invalid_files.append(f"{summary_name}.{file_type}")
                continue
            resolved = candidate_path.resolve()
            if not resolved.is_relative_to(root):
                invalid_files.append(f"{summary_name}.{file_type}")
                continue
            if not resolved.is_file():
                missing_files.append(f"{summary_name}.{file_type}")
                continue
            if file_type == "parquet":
                try:
                    import pyarrow.parquet as pq

                    actual_rows = int(pq.ParquetFile(resolved).metadata.num_rows)
                except Exception as exc:
                    raise HTTPException(
                        status_code=409,
                        detail={
                            "blockerType": "candidate_bundle_invalid",
                            "message": (
                                f"candidate summary parquet 无法读取："
                                f"{summary_name} ({exc})"
                            ),
                        },
                    ) from exc
                if declared_rows < 0 or actual_rows != declared_rows:
                    invalid_files.append(f"{summary_name}.parquet.rows")

    manifest_total_rows = int(
        summaries_manifest.get("totalSummaryRows", -1) or -1
    )
    if manifest_total_rows != declared_total_rows:
        invalid_files.append("totalSummaryRows")
    if missing_files or invalid_files:
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "candidate_bundle_invalid",
                "message": "candidate summaries 文件不完整或指向 staging 目录之外。",
                "missingFiles": missing_files,
                "invalidFiles": invalid_files,
            },
        )
    return summaries_manifest


def _raise_candidate_bundle_invalid(
    message: str,
    **detail: Any,
) -> None:
    raise HTTPException(
        status_code=409,
        detail={
            "blockerType": "candidate_bundle_invalid",
            "message": message,
            **detail,
        },
    )


def _partition_payload_signature(frame: pd.DataFrame) -> str:
    hash_values = pd.util.hash_pandas_object(
        frame,
        index=False,
        categorize=True,
    )
    return str(int(hash_values.astype("uint64").sum()))


def _stream_parquet_payload_signature(
    path: Path,
    *,
    columns: list[str],
    filter_column: str | None = None,
    filter_values: list[str] | None = None,
) -> tuple[int, str]:
    """Compute the existing pandas signature with bounded Arrow batches."""
    import pyarrow.dataset as ds

    expression = None
    if filter_column is not None:
        expression = ds.field(filter_column).isin(filter_values or [])
    scanner = ds.dataset(path, format="parquet").scanner(
        columns=columns,
        filter=expression,
        batch_size=SMART_MERGE_SCAN_BATCH_ROWS,
        use_threads=False,
        batch_readahead=1,
        fragment_readahead=1,
    )
    checksum = 0
    row_count = 0
    modulus = 1 << 64
    for batch in scanner.to_batches():
        if not batch.num_rows:
            continue
        frame = batch.to_pandas(use_threads=False)
        checksum = (
            checksum + int(_partition_payload_signature(frame))
        ) % modulus
        row_count += int(batch.num_rows)
    return row_count, str(checksum)


def _validate_candidate_full_bundle(
    *,
    parquet_path: Path,
    manifest_path: Path,
    partition_path: Path,
    fingerprint_path: Path,
    refresh_report_path: Path,
    summaries_path: Path,
) -> dict[str, Any]:
    """Fail closed unless all six staged artifacts describe one dataset."""
    try:
        import pyarrow.dataset as ds
        import pyarrow.parquet as pq

        manifest = _read_json(manifest_path)
        partition_manifest = _read_json(partition_path / "manifest.json")
        fingerprint = _read_json(fingerprint_path)
        refresh_report = _read_json(refresh_report_path)
        parquet_file = pq.ParquetFile(parquet_path)
    except Exception as exc:
        _raise_candidate_bundle_invalid(
            f"candidate 六件 bundle 无法读取：{exc}",
        )

    parquet_rows = int(parquet_file.metadata.num_rows)
    parquet_columns = int(parquet_file.metadata.num_columns)
    declared_rows = int(manifest.get("rows", -1) or -1)
    declared_columns = int(manifest.get("columns", -1) or -1)
    if (
        declared_rows != parquet_rows
        or declared_columns != parquet_columns
    ):
        _raise_candidate_bundle_invalid(
            "candidate manifest 与 parquet metadata 不一致。",
            parquetRows=parquet_rows,
            manifestRows=declared_rows,
            parquetColumns=parquet_columns,
            manifestColumns=declared_columns,
        )

    partition_rows = int(partition_manifest.get("rows", -1) or -1)
    partition_columns = int(
        partition_manifest.get("columns", -1) or -1
    )
    partition_keys = partition_manifest.get("partitionColumns")
    if (
        partition_rows != parquet_rows
        or partition_columns != parquet_columns
        or not isinstance(partition_keys, list)
        or len(partition_keys) != 1
    ):
        _raise_candidate_bundle_invalid(
            "candidate partition manifest 与 parquet metadata 不一致。",
            parquetRows=parquet_rows,
            partitionRows=partition_rows,
            parquetColumns=parquet_columns,
            partitionColumns=partition_columns,
            partitionKeys=partition_keys,
        )

    parquet_schema = pq.read_schema(parquet_path)
    country_column = _find_country_column(
        [str(column).strip() for column in parquet_schema.names]
    )
    partition_country_column = str(partition_keys[0]).strip()
    if (
        country_column is None
        or partition_country_column.casefold()
        != country_column.casefold()
    ):
        _raise_candidate_bundle_invalid(
            "candidate partition key 不是 parquet 的国家列。",
            parquetCountryColumn=country_column,
            partitionCountryColumn=partition_country_column,
        )

    parquet_files = sorted(partition_path.rglob("*.parquet"))
    actual_partition_dirs = {
        str(path.parent.relative_to(partition_path))
        for path in parquet_files
    }
    partition_stats = partition_manifest.get("partitionStats")
    if not isinstance(partition_stats, dict) or not partition_stats:
        _raise_candidate_bundle_invalid(
            "candidate partition manifest 缺少 partitionStats。",
        )
    declared_partition_dirs = {str(key) for key in partition_stats}
    declared_file_count = int(
        partition_manifest.get("parquetFileCount", -1) or -1
    )
    declared_dir_count = int(
        partition_manifest.get("partitionDirectoryCount", -1) or -1
    )
    if (
        actual_partition_dirs != declared_partition_dirs
        or declared_file_count != len(parquet_files)
        or declared_dir_count != len(actual_partition_dirs)
    ):
        _raise_candidate_bundle_invalid(
            "candidate partition 文件集合与 partition manifest 不一致。",
            missingPartitionDirs=sorted(
                declared_partition_dirs - actual_partition_dirs
            ),
            unexpectedPartitionDirs=sorted(
                actual_partition_dirs - declared_partition_dirs
            ),
            actualFileCount=len(parquet_files),
            declaredFileCount=declared_file_count,
        )

    country_value_map = _smart_merge_country_value_map(
        parquet_path,
        country_column=country_column,
        path_label="candidate parquet",
    )
    observed_country_rows = sum(
        int(batch.num_rows)
        for batch in parquet_file.iter_batches(
            batch_size=SMART_MERGE_SCAN_BATCH_ROWS,
            columns=[country_column],
            use_threads=False,
        )
    )
    if observed_country_rows != parquet_rows:
        _raise_candidate_bundle_invalid(
            "candidate parquet 国家列行数与 metadata 不一致。",
        )

    expected_partition_dirs = {
        f"{country_column}={quote(str(item['display']), safe='')}"
        for item in country_value_map.values()
    }
    if declared_partition_dirs != expected_partition_dirs:
        _raise_candidate_bundle_invalid(
            "candidate parquet 国家集合与 partition 目录不一致。",
            expectedPartitionDirs=sorted(expected_partition_dirs),
            actualPartitionDirs=sorted(declared_partition_dirs),
        )

    verified_rows = 0
    signature_mismatches: list[dict[str, Any]] = []
    payload_columns = [
        str(column)
        for column in parquet_schema.names
        if str(column) != country_column
    ]
    for partition_dir in sorted(declared_partition_dirs):
        raw_stats = partition_stats.get(partition_dir)
        if not isinstance(raw_stats, dict):
            signature_mismatches.append(
                {"partition": partition_dir, "reason": "missing_stats"}
            )
            continue
        country = unquote(partition_dir.split("=", 1)[-1]).strip()
        country_key = country.casefold()
        country_info = country_value_map.get(country_key)
        raw_values = (
            list(country_info.get("rawValues", []))
            if isinstance(country_info, dict)
            else []
        )
        if not raw_values:
            signature_mismatches.append(
                {"partition": partition_dir, "reason": "country_not_in_parquet"}
            )
            continue
        candidate_rows, candidate_signature = (
            _stream_parquet_payload_signature(
                parquet_path,
                columns=payload_columns,
                filter_column=country_column,
                filter_values=raw_values,
            )
        )
        partition_country_path = partition_path / partition_dir
        try:
            partition_schema = ds.dataset(
                partition_country_path,
                format="parquet",
            ).schema
        except Exception as exc:
            signature_mismatches.append(
                {
                    "partition": partition_dir,
                    "reason": f"partition_schema_unreadable:{exc}",
                }
            )
            continue
        partition_columns_actual = [
            str(column).strip()
            for column in partition_schema.names
        ]
        partition_rows_actual, partition_signature = (
            _stream_parquet_payload_signature(
                partition_country_path,
                columns=partition_columns_actual,
            )
        )
        expected_rows = int(raw_stats.get("rows", -1) or -1)
        expected_signature = str(raw_stats.get("signature") or "")
        if (
            payload_columns != partition_columns_actual
            or candidate_rows != expected_rows
            or partition_rows_actual != expected_rows
            or candidate_signature != expected_signature
            or partition_signature != expected_signature
        ):
            signature_mismatches.append(
                {
                    "partition": partition_dir,
                    "candidateRows": candidate_rows,
                    "partitionRows": partition_rows_actual,
                    "expectedRows": expected_rows,
                    "candidateSignature": candidate_signature,
                    "partitionSignature": partition_signature,
                    "expectedSignature": expected_signature,
                }
            )
        verified_rows += partition_rows_actual

    if signature_mismatches or verified_rows != parquet_rows:
        _raise_candidate_bundle_invalid(
            "candidate partition 内容与 parquet 不一致。",
            verifiedRows=verified_rows,
            parquetRows=parquet_rows,
            signatureMismatches=signature_mismatches[:10],
        )

    parquet_sha256 = _sha256_hex_for_path(parquet_path)
    parquet_size = int(parquet_path.stat().st_size)
    manifest_sha256 = str(manifest.get("sha256") or "").strip().lower()
    manifest_size = int(
        manifest.get(
            "fileSizeBytes",
            manifest.get("outputParquetBytes", -1),
        )
        or -1
    )
    fingerprint_sha256 = str(
        fingerprint.get("sha256")
        or fingerprint.get("datasetSha256")
        or ""
    ).strip().lower()
    fingerprint_rows = int(fingerprint.get("rowCount", -1) or -1)
    fingerprint_columns = int(
        fingerprint.get("columnCount", -1) or -1
    )
    if (
        manifest_sha256 != parquet_sha256
        or fingerprint_sha256 != parquet_sha256
        or manifest_size != parquet_size
        or fingerprint_rows != parquet_rows
        or fingerprint_columns != parquet_columns
    ):
        _raise_candidate_bundle_invalid(
            "candidate manifest/fingerprint 与 parquet 内容指纹不一致。",
            parquetSha256=parquet_sha256,
            manifestSha256=manifest_sha256 or None,
            fingerprintSha256=fingerprint_sha256 or None,
            parquetSizeBytes=parquet_size,
            manifestSizeBytes=manifest_size,
            parquetRows=parquet_rows,
            fingerprintRows=fingerprint_rows,
            parquetColumns=parquet_columns,
            fingerprintColumns=fingerprint_columns,
        )

    full_manifest = refresh_report.get("fullManifest")
    report_rows = (
        int(full_manifest.get("rows", -1) or -1)
        if isinstance(full_manifest, dict)
        else -1
    )
    if (
        str(refresh_report.get("jobStatus") or "") != "success"
        or report_rows != parquet_rows
    ):
        _raise_candidate_bundle_invalid(
            "candidate refresh report 与 parquet 不一致。",
            reportStatus=refresh_report.get("jobStatus"),
            reportRows=report_rows,
            parquetRows=parquet_rows,
        )
    if not fingerprint:
        _raise_candidate_bundle_invalid(
            "candidate fingerprint.json 为空。",
        )

    summaries_manifest = _validate_candidate_summaries_bundle(
        summaries_path=summaries_path,
        candidate_manifest_path=manifest_path,
    )
    return {
        "rows": parquet_rows,
        "columns": parquet_columns,
        "partitionCount": len(actual_partition_dirs),
        "summaryRows": int(
            summaries_manifest.get("totalSummaryRows", 0) or 0
        ),
    }


def _load_parquet_country_subset(path: Path, country: str, *, path_label: str) -> pd.DataFrame:
    """Read exactly one country from a parquet file; never materialize the archive."""
    try:
        import pyarrow.parquet as pq

        schema = pq.read_schema(path)
        country_column = _find_country_column([str(column) for column in schema.names])
        if country_column is None:
            raise HTTPException(status_code=409, detail=f"{path_label} 缺少国家列。")
        country_values = pd.read_parquet(
            path,
            columns=[country_column],
        )[country_column]
        normalized = (
            country_values.astype("string").fillna("").str.strip()
        )
        _logical_country_display_map(
            normalized,
            path_label=path_label,
        )
        requested_key = country.strip().casefold()
        raw_values = list(
            dict.fromkeys(
                str(raw_value)
                for raw_value, logical_key in zip(
                    country_values,
                    normalized.str.casefold(),
                    strict=False,
                )
                if not pd.isna(raw_value)
                and str(logical_key) == requested_key
            )
        )
        if not raw_values:
            return pd.DataFrame(columns=[str(column) for column in schema.names])
        country_filter = (
            (country_column, "==", raw_values[0])
            if len(raw_values) == 1
            else (country_column, "in", raw_values)
        )
        frame = pd.read_parquet(path, filters=[country_filter])
        frame.columns = [str(column).strip() for column in frame.columns]
        return frame
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=409,
            detail=f"读取 {path_label} 的 {country} 分区失败。",
        ) from exc


def _load_active_country_partition_subset(partition_root: Path, country: str) -> pd.DataFrame:
    country_dir = partition_root / f"国家={quote(country, safe='')}"
    if not country_dir.exists():
        raise HTTPException(status_code=409, detail=f"找不到 active 的 {country} 国家分区。")
    try:
        frame = pd.read_parquet(country_dir)
        frame.columns = [str(column).strip() for column in frame.columns]
        if _find_country_column(list(frame.columns)) is None:
            frame["国家"] = country
        return frame
    except Exception as exc:
        raise HTTPException(
            status_code=409,
            detail=f"读取 active 的 {country} 国家分区失败。",
        ) from exc


def _untouched_partition_snapshot(
    *,
    partition_root: Path,
    country: str | None = None,
    countries: list[str] | None = None,
) -> dict[str, Any]:
    """Fingerprint both manifest metadata and physical non-target parquet files."""
    manifest_path = partition_root / "manifest.json"
    manifest = _read_json_if_exists(_relative_to_project(manifest_path)) or {}
    partition_stats = manifest.get("partitionStats")
    partition_columns = manifest.get("partitionColumns")
    if not isinstance(partition_stats, dict):
        return {"status": "unavailable", "reason": "missing_partition_stats"}
    column = (
        str(partition_columns[0])
        if isinstance(partition_columns, list) and partition_columns
        else "国家"
    )
    target_countries = _ordered_distinct_strings(
        [*(countries or []), *([country] if country else [])]
    )
    target_prefixes = [
        f"{column}={quote(target_country, safe='')}"
        for target_country in target_countries
    ]
    untouched = {
        str(key): value
        for key, value in partition_stats.items()
        if not any(str(key).startswith(prefix) for prefix in target_prefixes)
    }
    encoded = json.dumps(untouched, ensure_ascii=False, sort_keys=True).encode("utf-8")
    physical_hasher = hashlib.sha256()
    physical_file_count = 0
    for file_path in sorted(partition_root.rglob("*.parquet")):
        relative = file_path.relative_to(partition_root)
        if any(str(relative).startswith(prefix) for prefix in target_prefixes):
            continue
        physical_hasher.update(str(relative).encode("utf-8"))
        physical_hasher.update(_sha256_hex_for_path(file_path).encode("ascii"))
        physical_file_count += 1
    return {
        "status": "pass",
        "targetPartitionPrefix": target_prefixes[0] if len(target_prefixes) == 1 else None,
        "targetPartitionPrefixes": target_prefixes,
        "untouchedPartitionCount": len(untouched),
        "untouchedPartitionFingerprint": hashlib.sha256(encoded).hexdigest(),
        "untouchedPhysicalFileCount": physical_file_count,
        "untouchedPhysicalFingerprint": physical_hasher.hexdigest(),
    }


def _verify_untouched_partition_stability(*, before: dict[str, Any], after: dict[str, Any]) -> dict[str, Any]:
    if before.get("status") != "pass" or after.get("status") != "pass":
        return {"status": "unavailable", "before": before, "after": after}
    changed = (
        before.get("untouchedPartitionFingerprint")
        != after.get("untouchedPartitionFingerprint")
        or before.get("untouchedPhysicalFingerprint")
        != after.get("untouchedPhysicalFingerprint")
    )
    return {
        "status": "fail" if changed else "pass",
        "untouchedPartitionCount": int(before.get("untouchedPartitionCount", 0) or 0),
        "beforeFingerprint": before.get("untouchedPartitionFingerprint"),
        "afterFingerprint": after.get("untouchedPartitionFingerprint"),
        "untouchedPhysicalFileCount": int(
            before.get("untouchedPhysicalFileCount", 0) or 0
        ),
        "beforePhysicalFingerprint": before.get("untouchedPhysicalFingerprint"),
        "afterPhysicalFingerprint": after.get("untouchedPhysicalFingerprint"),
    }


def _latest_month_from_frame(frame: pd.DataFrame) -> str | None:
    for column in reversed(_detect_month_columns(list(frame.columns))):
        if _series_has_data(frame[column]):
            return column
    return None


def _is_derived_ytd_column(column: str) -> bool:
    return bool(YTD_COLUMN_PATTERN.fullmatch(str(column).strip()))


def _single_country_schema_contract(*, active_frame: pd.DataFrame, candidate_frame: pd.DataFrame) -> dict[str, list[str]]:
    active_months = set(_detect_month_columns(list(active_frame.columns)))
    candidate_months = set(_detect_month_columns(list(candidate_frame.columns)))
    active_static = set(active_frame.columns) - active_months
    candidate_static = set(candidate_frame.columns) - candidate_months
    missing = sorted(active_static - candidate_static)
    null_only: list[str] = []
    derived_ytd: list[str] = []
    deprecated_optional: list[str] = []
    material: list[str] = []
    for column in missing:
        if not _series_has_data(active_frame[column]):
            null_only.append(column)
        elif _is_derived_ytd_column(column):
            derived_ytd.append(column)
        elif column in DEPRECATED_OPTIONAL_STATIC_COLUMNS:
            deprecated_optional.append(column)
        else:
            material.append(column)
    return {
        "missing": missing,
        "missingNullOnly": null_only,
        "missingDerivedYtd": derived_ytd,
        "missingDeprecatedOptional": deprecated_optional,
        "missingMaterial": material,
        "extra": sorted(candidate_static - active_static),
    }


def _single_country_configuration_key_columns(frame: pd.DataFrame) -> list[str]:
    month_columns = set(_detect_month_columns(list(frame.columns)))
    return [
        str(column)
        for column in frame.columns
        if str(column) not in month_columns
        and not YEAR_COLUMN_PATTERN.fullmatch(str(column).strip())
        and not _is_derived_ytd_column(str(column))
    ]


def _single_country_make_model_deltas(
    *,
    active_frame: pd.DataFrame,
    candidate_frame: pd.DataFrame,
    active_sales: pd.DataFrame,
    candidate_sales: pd.DataFrame,
    historical_months: list[str],
) -> tuple[dict[tuple[str, str], pd.Series], list[dict[str, Any]], int]:
    dimensions = ("Make", "Model")
    if any(
        column not in active_frame.columns or column not in candidate_frame.columns
        for column in dimensions
    ):
        return {}, [], 0

    def grouped(frame: pd.DataFrame, sales: pd.DataFrame) -> pd.DataFrame:
        keys = [
            frame[column].astype("string").fillna("").str.strip().rename(column)
            for column in dimensions
        ]
        return sales.groupby(keys, dropna=False).sum()

    active_grouped = grouped(active_frame, active_sales)
    candidate_grouped = grouped(candidate_frame, candidate_sales)
    all_keys = active_grouped.index.union(candidate_grouped.index)
    active_aligned = active_grouped.reindex(all_keys, fill_value=0)
    candidate_aligned = candidate_grouped.reindex(all_keys, fill_value=0)
    delta_frame = candidate_aligned - active_aligned
    delta_vectors: dict[tuple[str, str], pd.Series] = {}
    samples: list[dict[str, Any]] = []

    for raw_key, delta in delta_frame.iterrows():
        if not bool(delta.ne(0).any()):
            continue
        make, model = (str(value).strip() for value in raw_key)
        key = (make, model)
        delta_vectors[key] = delta.astype(float)
        active_row = active_aligned.loc[raw_key]
        candidate_row = candidate_aligned.loc[raw_key]
        for month in historical_months:
            delta_value = float(delta.get(month, 0) or 0)
            if delta_value == 0:
                continue
            samples.append({
                "scope": f"makeModel:{make}/{model}",
                "make": make,
                "model": model,
                "month": month,
                "activeSales": _serialize_numeric_value(active_row.get(month, 0)),
                "candidateSales": _serialize_numeric_value(candidate_row.get(month, 0)),
                "deltaSales": _serialize_numeric_value(delta_value),
            })
    return delta_vectors, samples, len(all_keys)


def _shared_historical_dimension_columns(
    active_frame: pd.DataFrame,
    candidate_frame: pd.DataFrame,
) -> list[tuple[str, str, str]]:
    active_lookup = {
        str(column).strip().casefold(): str(column)
        for column in active_frame.columns
    }
    candidate_lookup = {
        str(column).strip().casefold(): str(column)
        for column in candidate_frame.columns
    }
    shared: list[tuple[str, str, str]] = []
    for aliases in HISTORICAL_DIMENSION_ALIASES:
        active_column = next(
            (
                active_lookup[alias.casefold()]
                for alias in aliases
                if alias.casefold() in active_lookup
            ),
            None,
        )
        candidate_column = next(
            (
                candidate_lookup[alias.casefold()]
                for alias in aliases
                if alias.casefold() in candidate_lookup
            ),
            None,
        )
        if active_column is None or candidate_column is None:
            continue
        shared.append((aliases[0], active_column, candidate_column))
    return shared


def _single_country_analysis_dimension_samples(
    *,
    active_frame: pd.DataFrame,
    candidate_frame: pd.DataFrame,
    active_sales: pd.DataFrame,
    candidate_sales: pd.DataFrame,
    historical_months: list[str],
) -> tuple[list[dict[str, Any]], list[str], int]:
    """Compare published history at stable dashboard-analysis dimensions.

    Make/Model reclassifications are validated separately so a precisely
    confirmed model rename can remain allowed.  This second projection catches
    silent redistribution between powertrain/fuel/segment/body/registration
    slices even when country and Make/Model totals remain unchanged.
    """
    shared = _shared_historical_dimension_columns(
        active_frame,
        candidate_frame,
    )
    if not shared:
        return [], [], 0

    def grouped(
        frame: pd.DataFrame,
        sales: pd.DataFrame,
        *,
        active_side: bool,
    ) -> pd.DataFrame:
        keys = [
            _normalized_static_value_series(
                frame[active_column if active_side else candidate_column]
            ).rename(label)
            for label, active_column, candidate_column in shared
        ]
        return sales.groupby(keys, dropna=False).sum()

    active_grouped = grouped(active_frame, active_sales, active_side=True)
    candidate_grouped = grouped(
        candidate_frame,
        candidate_sales,
        active_side=False,
    )
    all_keys = active_grouped.index.union(candidate_grouped.index)
    active_aligned = active_grouped.reindex(all_keys, fill_value=0)
    candidate_aligned = candidate_grouped.reindex(all_keys, fill_value=0)
    samples: list[dict[str, Any]] = []
    for raw_key, delta in (candidate_aligned - active_aligned).iterrows():
        if not bool(delta.ne(0).any()):
            continue
        key_values = raw_key if isinstance(raw_key, tuple) else (raw_key,)
        dimensions = {
            label: str(value)
            for (label, _active_column, _candidate_column), value in zip(
                shared,
                key_values,
                strict=False,
            )
        }
        active_row = active_aligned.loc[raw_key]
        candidate_row = candidate_aligned.loc[raw_key]
        for month in historical_months:
            delta_value = float(delta.get(month, 0) or 0)
            if delta_value == 0:
                continue
            samples.append(
                {
                    "scope": "analysisDimensions",
                    "dimensions": dimensions,
                    "month": month,
                    "activeSales": _serialize_numeric_value(
                        active_row.get(month, 0)
                    ),
                    "candidateSales": _serialize_numeric_value(
                        candidate_row.get(month, 0)
                    ),
                    "deltaSales": _serialize_numeric_value(delta_value),
                }
            )
    return samples, [label for label, _left, _right in shared], len(all_keys)


def _historical_value_display_map(*series_values: pd.Series) -> dict[str, str]:
    """Keep report labels readable while grouping case/whitespace variants."""
    display: dict[str, str] = {}
    for series in series_values:
        normalized = _normalized_static_value_series(series)
        raw_values = series.astype("string").fillna("").str.strip()
        for normalized_value, raw_value in zip(
            normalized,
            raw_values,
            strict=False,
        ):
            key = str(normalized_value)
            if key in display:
                continue
            rendered = str(raw_value).strip()
            display[key] = rendered if rendered else "?"
    return display


def _build_historical_reclassification_country_report(
    *,
    country: str,
    active_frame: pd.DataFrame,
    candidate_frame: pd.DataFrame,
    active_sales: pd.DataFrame,
    candidate_sales: pd.DataFrame,
    historical_months: list[str],
    country_mismatch_count: int,
    analysis_dimension_samples: list[dict[str, Any]],
    confirmed_make_model_reclassifications: list[dict[str, Any]],
    unconfirmed_make_model_candidates: list[dict[str, Any]],
    unpaired_make_models: list[dict[str, str]],
) -> dict[str, Any] | None:
    """Build a bounded, human-readable projection of historical relabelling.

    The input frames already contain only one country.  Each dimension is
    aggregated and discarded before the next one, so report generation never
    materialises the full archive or a cartesian row-level diff.
    """
    shared = _shared_historical_dimension_columns(
        active_frame,
        candidate_frame,
    )
    dimension_summaries: list[dict[str, Any]] = []
    exact_changes: list[dict[str, Any]] = []
    exact_change_count = 0
    complex_change_count = len(unpaired_make_models)
    payload_truncated = False

    def retain_top_exact_change(change: dict[str, Any]) -> None:
        nonlocal payload_truncated
        exact_changes.append(change)
        exact_changes.sort(
            key=lambda item: (
                -float(item.get("transferredSales") or 0),
                str(item.get("dimension") or ""),
                str(item.get("make") or ""),
                str(item.get("model") or ""),
                str(item.get("oldValue") or ""),
            )
        )
        if (
            len(exact_changes)
            > HISTORICAL_RECLASSIFICATION_EXACT_CHANGE_LIMIT
        ):
            exact_changes.pop()
            payload_truncated = True

    make_available = (
        "Make" in active_frame.columns
        and "Make" in candidate_frame.columns
    )
    model_available = (
        "Model" in active_frame.columns
        and "Model" in candidate_frame.columns
    )
    make_display = (
        _historical_value_display_map(
            active_frame["Make"],
            candidate_frame["Make"],
        )
        if make_available
        else {}
    )
    model_display = (
        _historical_value_display_map(
            active_frame["Model"],
            candidate_frame["Model"],
        )
        if model_available
        else {}
    )

    for label, active_column, candidate_column in shared:
        active_dimension = _normalized_static_value_series(
            active_frame[active_column]
        )
        candidate_dimension = _normalized_static_value_series(
            candidate_frame[candidate_column]
        )
        value_display = _historical_value_display_map(
            active_frame[active_column],
            candidate_frame[candidate_column],
        )
        active_make = (
            _normalized_static_value_series(active_frame["Make"])
            if make_available
            else pd.Series("", index=active_frame.index, dtype="string")
        )
        candidate_make = (
            _normalized_static_value_series(candidate_frame["Make"])
            if make_available
            else pd.Series("", index=candidate_frame.index, dtype="string")
        )
        active_model = (
            _normalized_static_value_series(active_frame["Model"])
            if model_available
            else pd.Series("", index=active_frame.index, dtype="string")
        )
        candidate_model = (
            _normalized_static_value_series(candidate_frame["Model"])
            if model_available
            else pd.Series("", index=candidate_frame.index, dtype="string")
        )

        active_dimension_grouped = active_sales.groupby(
            active_dimension.rename(label),
            dropna=False,
        ).sum()
        candidate_dimension_grouped = candidate_sales.groupby(
            candidate_dimension.rename(label),
            dropna=False,
        ).sum()
        dimension_keys = active_dimension_grouped.index.union(
            candidate_dimension_grouped.index
        )
        dimension_delta = candidate_dimension_grouped.reindex(
            dimension_keys,
            fill_value=0,
        ) - active_dimension_grouped.reindex(
            dimension_keys,
            fill_value=0,
        )
        dimension_changed = dimension_delta.loc[
            dimension_delta.ne(0).any(axis=1)
        ]

        old_aggregates: dict[str, dict[str, Any]] = {}
        new_aggregates: dict[str, dict[str, Any]] = {}
        for raw_value, delta in dimension_changed.iterrows():
            value_key = str(raw_value)
            values = delta.reindex(
                historical_months,
                fill_value=0,
            ).astype(float)
            negative = values.where(values.lt(0), 0)
            positive = values.where(values.gt(0), 0)
            if bool(negative.lt(0).any()):
                old_aggregates[value_key] = {
                    "sales": float(-negative.sum()),
                    "months": {
                        month
                        for month in historical_months
                        if float(negative.get(month, 0) or 0) < 0
                    },
                }
            if bool(positive.gt(0).any()):
                new_aggregates[value_key] = {
                    "sales": float(positive.sum()),
                    "months": {
                        month
                        for month in historical_months
                        if float(positive.get(month, 0) or 0) > 0
                    },
                }

        def top_values(
            values: dict[str, dict[str, Any]],
        ) -> list[dict[str, Any]]:
            nonlocal payload_truncated
            ordered = sorted(
                values.items(),
                key=lambda item: (-float(item[1]["sales"]), item[0]),
            )
            if len(ordered) > HISTORICAL_RECLASSIFICATION_VALUE_LIMIT:
                payload_truncated = True
            return [
                {
                    "value": value_display.get(value, value or "?"),
                    "sales": _serialize_numeric_value(metrics["sales"]),
                    "monthCount": len(metrics["months"]),
                }
                for value, metrics in ordered[
                    :HISTORICAL_RECLASSIFICATION_VALUE_LIMIT
                ]
            ]

        if not dimension_changed.empty:
            negative_delta = dimension_changed.where(
                dimension_changed.lt(0),
                0,
            )
            positive_delta = dimension_changed.where(
                dimension_changed.gt(0),
                0,
            )
            moved_sales = float(positive_delta.sum().sum())
            if moved_sales == 0:
                moved_sales = float(-negative_delta.sum().sum())
            dimension_summaries.append(
                {
                    "dimension": label,
                    "mismatchCellCount": int(
                        dimension_changed.ne(0).sum().sum()
                    ),
                    "movedSales": _serialize_numeric_value(
                        moved_sales
                    ),
                    "oldValues": top_values(old_aggregates),
                    "newValues": top_values(new_aggregates),
                }
            )

        def grouped_for_exact_changes(
            make: pd.Series,
            model: pd.Series,
            dimension: pd.Series,
            sales: pd.DataFrame,
        ) -> pd.DataFrame:
            return sales.groupby(
                [
                    make.rename("Make"),
                    model.rename("Model"),
                    dimension.rename(label),
                ],
                dropna=False,
            ).sum()

        active_grouped = grouped_for_exact_changes(
            active_make,
            active_model,
            active_dimension,
            active_sales,
        )
        candidate_grouped = grouped_for_exact_changes(
            candidate_make,
            candidate_model,
            candidate_dimension,
            candidate_sales,
        )
        all_keys = active_grouped.index.union(candidate_grouped.index)
        active_aligned = active_grouped.reindex(all_keys, fill_value=0)
        candidate_aligned = candidate_grouped.reindex(
            all_keys,
            fill_value=0,
        )
        delta_frame = candidate_aligned - active_aligned
        changed = delta_frame.loc[delta_frame.ne(0).any(axis=1)]

        sources: dict[
            tuple[str, str, tuple[float, ...]],
            list[tuple[str, pd.Series]],
        ] = {}
        targets: dict[
            tuple[str, str, tuple[float, ...]],
            list[tuple[str, pd.Series]],
        ] = {}
        for raw_key, delta in changed.iterrows():
            make_key, model_key, value_key = (
                str(value)
                for value in (
                    raw_key
                    if isinstance(raw_key, tuple)
                    else ("", "", raw_key)
                )
            )
            values = delta.reindex(
                historical_months,
                fill_value=0,
            ).astype(float)
            if bool(values.le(0).all()) and bool(values.lt(0).any()):
                signature = tuple(float(-value) for value in values)
                sources.setdefault(
                    (make_key, model_key, signature),
                    [],
                ).append((value_key, -values))
            elif bool(values.ge(0).all()) and bool(values.gt(0).any()):
                signature = tuple(float(value) for value in values)
                targets.setdefault(
                    (make_key, model_key, signature),
                    [],
                ).append((value_key, values))

        dimension_exact_count = 0
        for signature_key, source_values in sources.items():
            target_values = targets.get(signature_key, [])
            if len(source_values) != 1 or len(target_values) != 1:
                continue
            make_key, model_key, _signature = signature_key
            old_value, transfer = source_values[0]
            new_value, _target_transfer = target_values[0]
            if old_value == new_value:
                continue
            monthly_transfers = [
                {
                    "month": month,
                    "sales": _serialize_numeric_value(
                        transfer.get(month, 0)
                    ),
                }
                for month in historical_months
                if float(transfer.get(month, 0) or 0) != 0
            ]
            exact_change_count += 1
            dimension_exact_count += 1
            retain_top_exact_change(
                {
                    "dimension": label,
                    "make": make_display.get(
                        make_key,
                        make_key or "?",
                    ),
                    "model": model_display.get(
                        model_key,
                        model_key or "?",
                    ),
                    "oldValue": value_display.get(
                        old_value,
                        old_value or "?",
                    ),
                    "newValue": value_display.get(
                        new_value,
                        new_value or "?",
                    ),
                    "transferredSales": _serialize_numeric_value(
                        transfer.sum()
                    ),
                    "affectedMonths": [
                        item["month"]
                        for item in monthly_transfers
                    ],
                    "monthlyTransfers": monthly_transfers,
                    "confidence": "candidate_exact_vector",
                }
            )
        complex_change_count += max(
            0,
            int(len(changed)) - (dimension_exact_count * 2),
        )
        del active_dimension_grouped
        del candidate_dimension_grouped
        del dimension_delta
        del dimension_changed
        del active_grouped
        del candidate_grouped
        del active_aligned
        del candidate_aligned
        del delta_frame
        del changed

    for candidate in unconfirmed_make_model_candidates:
        source = candidate.get("source")
        target = candidate.get("target")
        monthly_transfers = candidate.get("monthlyTransfers")
        if not (
            isinstance(source, dict)
            and isinstance(target, dict)
            and isinstance(monthly_transfers, list)
        ):
            continue
        exact_change_count += 1
        retain_top_exact_change(
            {
                "dimension": "Make/Model",
                "make": str(source.get("Make") or "?"),
                "model": str(source.get("Model") or "?"),
                "oldValue": "/".join(
                    [
                        str(source.get("Make") or "?"),
                        str(source.get("Model") or "?"),
                    ]
                ),
                "newValue": "/".join(
                    [
                        str(target.get("Make") or "?"),
                        str(target.get("Model") or "?"),
                    ]
                ),
                "transferredSales": candidate.get(
                    "transferredSales"
                ),
                "affectedMonths": [
                    str(month)
                    for month in candidate.get(
                        "transferredMonths",
                        [],
                    )
                ],
                "monthlyTransfers": [
                    {
                        "month": str(item.get("month") or ""),
                        "sales": item.get("sales"),
                    }
                    for item in monthly_transfers
                    if isinstance(item, dict)
                ],
                "confidence": "candidate_exact_vector",
            }
        )

    for confirmed_change in confirmed_make_model_reclassifications:
        source = confirmed_change.get("source")
        target = confirmed_change.get("target")
        monthly_transfers = confirmed_change.get("monthlyTransfers")
        if not (
            isinstance(source, dict)
            and isinstance(target, dict)
            and isinstance(monthly_transfers, list)
        ):
            continue
        exact_change_count += 1
        retain_top_exact_change(
            {
                "dimension": "Make/Model",
                "make": str(source.get("Make") or "?"),
                "model": str(source.get("Model") or "?"),
                "oldValue": "/".join(
                    [
                        str(source.get("Make") or "?"),
                        str(source.get("Model") or "?"),
                    ]
                ),
                "newValue": "/".join(
                    [
                        str(target.get("Make") or "?"),
                        str(target.get("Model") or "?"),
                    ]
                ),
                "transferredSales": confirmed_change.get(
                    "transferredSales"
                ),
                "affectedMonths": [
                    str(month)
                    for month in confirmed_change.get(
                        "transferredMonths",
                        [],
                    )
                ],
                "monthlyTransfers": [
                    {
                        "month": str(item.get("month") or ""),
                        "sales": item.get("sales"),
                    }
                    for item in monthly_transfers
                    if isinstance(item, dict)
                ],
                "confidence": "confirmed_upload_bound",
            }
        )

    joint_mismatch_cell_count = len(analysis_dimension_samples)
    joint_moved_sales = sum(
        float(item.get("deltaSales") or 0)
        for item in analysis_dimension_samples
        if float(item.get("deltaSales") or 0) > 0
    )
    monthly_totals_stable = country_mismatch_count == 0
    has_historical_change = bool(
        joint_mismatch_cell_count
        or confirmed_make_model_reclassifications
        or unconfirmed_make_model_candidates
        or unpaired_make_models
        or country_mismatch_count
    )
    if not has_historical_change:
        return None
    decision_required = True
    allowed_decisions = (
        HISTORICAL_RECLASSIFICATION_DECISION_ORDER
        if monthly_totals_stable
        else ("keep_active",)
    )
    return {
        "country": country,
        "comparedThrough": (
            historical_months[-1]
            if historical_months
            else None
        ),
        "historicalMonthCount": len(historical_months),
        "jointMismatchCellCount": joint_mismatch_cell_count,
        "jointMovedSales": _serialize_numeric_value(
            joint_moved_sales
        ),
        "monthlyTotalsStable": monthly_totals_stable,
        "decisionRequired": decision_required,
        "allowedDecisions": list(allowed_decisions),
        "dimensionSummaries": dimension_summaries,
        "exactChanges": exact_changes,
        "exactChangeCount": exact_change_count,
        "complexChangeCount": complex_change_count,
        "truncation": {
            "truncated": payload_truncated,
            "exactChangeLimit": (
                HISTORICAL_RECLASSIFICATION_EXACT_CHANGE_LIMIT
            ),
            "valueLimitPerDirection": (
                HISTORICAL_RECLASSIFICATION_VALUE_LIMIT
            ),
        },
    }


def _sc011_transfer_payload(
    *,
    source: tuple[str, str],
    target: tuple[str, str],
    transfer: pd.Series,
    historical_months: list[str],
) -> dict[str, Any]:
    monthly_transfers = [
        {
            "month": month,
            "sales": _serialize_numeric_value(transfer.get(month, 0)),
        }
        for month in historical_months
        if float(transfer.get(month, 0) or 0) != 0
    ]
    return {
        "source": {"Make": source[0], "Model": source[1]},
        "target": {"Make": target[0], "Model": target[1]},
        "transferredSales": _serialize_numeric_value(transfer.sum()),
        "transferredMonths": [item["month"] for item in monthly_transfers],
        "monthlyTransfers": monthly_transfers,
    }


def _apply_confirmed_sc011_reclassifications(
    *,
    country: str,
    active_latest_month: str,
    historical_months: list[str],
    delta_vectors: dict[tuple[str, str], pd.Series],
    source_upload_sha256: str | None,
) -> tuple[list[dict[str, Any]], dict[tuple[str, str], pd.Series]]:
    residual = {key: value.copy() for key, value in delta_vectors.items()}
    confirmed: list[dict[str, Any]] = []
    normalized_upload_sha256 = str(source_upload_sha256 or "").strip().lower()
    compared_months_sha256 = hashlib.sha256(
        "\n".join(historical_months).encode("utf-8")
    ).hexdigest()
    for rule in CONFIRMED_SC011_RECLASSIFICATIONS:
        if (
            str(rule.get("country") or "") != country
            or str(rule.get("comparedThrough") or "") != active_latest_month
            or int(rule.get("expectedComparedMonthCount") or 0)
            != len(historical_months)
            or str(rule.get("expectedComparedMonthsSha256") or "")
            != compared_months_sha256
            or str(rule.get("sourceUploadSha256") or "").strip().lower()
            != normalized_upload_sha256
        ):
            continue
        source_value = rule.get("source")
        target_value = rule.get("target")
        expected_value = rule.get("expectedMonthlyTransfer")
        if not (
            isinstance(source_value, dict)
            and isinstance(target_value, dict)
            and isinstance(expected_value, dict)
        ):
            continue
        source = (
            str(source_value.get("Make") or "").strip(),
            str(source_value.get("Model") or "").strip(),
        )
        target = (
            str(target_value.get("Make") or "").strip(),
            str(target_value.get("Model") or "").strip(),
        )
        if source == target or source not in residual or target not in residual:
            continue
        if any(month not in historical_months for month in expected_value):
            continue
        expected = pd.Series(0.0, index=historical_months)
        valid_expected = True
        for month, raw_sales in expected_value.items():
            numeric_sales = pd.to_numeric(pd.Series([raw_sales]), errors="coerce").iloc[0]
            if pd.isna(numeric_sales) or float(numeric_sales) < 0:
                valid_expected = False
                break
            expected.loc[str(month)] = float(numeric_sales)
        if not valid_expected:
            continue
        expected_total = float(rule.get("expectedTotal") or 0)
        source_delta = residual[source].reindex(historical_months, fill_value=0)
        target_delta = residual[target].reindex(historical_months, fill_value=0)
        if (
            float(expected.sum()) != expected_total
            or not bool(source_delta.eq(-expected).all())
            or not bool(target_delta.eq(expected).all())
        ):
            continue
        payload = _sc011_transfer_payload(
            source=source,
            target=target,
            transfer=expected,
            historical_months=historical_months,
        )
        payload.update({
            "confirmationId": str(rule.get("confirmationId") or ""),
            "approvedBy": str(rule.get("approvedBy") or ""),
            "approvalReference": str(rule.get("approvalReference") or ""),
            "comparedThrough": active_latest_month,
            "comparedMonthsSha256": compared_months_sha256,
            "sourceUploadSha256": normalized_upload_sha256,
        })
        confirmed.append(payload)
        residual.pop(source)
        residual.pop(target)
    return confirmed, residual


def _unconfirmed_sc011_reclassification_candidates(
    *,
    residual: dict[tuple[str, str], pd.Series],
    historical_months: list[str],
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    sources_by_signature: dict[tuple[str, tuple[float, ...]], list[tuple[str, str]]] = {}
    targets_by_signature: dict[tuple[str, tuple[float, ...]], list[tuple[str, str]]] = {}
    for key, delta in residual.items():
        values = delta.reindex(historical_months, fill_value=0).astype(float)
        if bool(values.le(0).all()) and bool(values.lt(0).any()):
            signature = tuple(float(-value) for value in values)
            sources_by_signature.setdefault((key[0], signature), []).append(key)
        elif bool(values.ge(0).all()) and bool(values.gt(0).any()):
            signature = tuple(float(value) for value in values)
            targets_by_signature.setdefault((key[0], signature), []).append(key)

    candidates: list[dict[str, Any]] = []
    paired_keys: set[tuple[str, str]] = set()
    for signature_key, sources in sources_by_signature.items():
        targets = targets_by_signature.get(signature_key, [])
        if len(sources) != 1 or len(targets) != 1:
            continue
        source = sources[0]
        target = targets[0]
        transfer = -residual[source].reindex(historical_months, fill_value=0)
        candidates.append(
            _sc011_transfer_payload(
                source=source,
                target=target,
                transfer=transfer,
                historical_months=historical_months,
            )
        )
        paired_keys.update({source, target})
    candidates.sort(
        key=lambda item: (
            str(item["source"]["Make"]),
            str(item["source"]["Model"]),
            str(item["target"]["Model"]),
        )
    )
    unpaired = [
        {"Make": make, "Model": model}
        for make, model in sorted(set(residual) - paired_keys)
    ]
    return candidates, unpaired


def _historical_sales_frame(
    frame: pd.DataFrame,
    historical_months: list[str],
) -> pd.DataFrame:
    """Normalize sales columns without materializing a second 2-D block.

    Clean numeric columns may share their buffers with ``frame``. Callers
    must therefore treat the returned frame as read-only.
    """
    numeric_columns: dict[str, pd.Series] = {}
    for month in historical_months:
        numeric = pd.to_numeric(frame[month], errors="coerce")
        if bool(numeric.isna().any()):
            numeric = numeric.fillna(0)
        numeric_columns[month] = numeric
    return pd.DataFrame(
        numeric_columns,
        index=frame.index,
        copy=False,
    )


def _single_country_historical_sales_stability(
    *,
    country: str,
    active_frame: pd.DataFrame,
    candidate_frame: pd.DataFrame,
    active_latest_month: str | None,
    source_upload_sha256: str | None = None,
) -> dict[str, Any]:
    if active_latest_month is None:
        return {"status": "unavailable", "reason": "active_latest_month_missing"}
    active_months = set(_detect_month_columns(list(active_frame.columns)))
    candidate_months = set(_detect_month_columns(list(candidate_frame.columns)))
    historical_months = sorted(
        [month for month in active_months if _time_sort_key(month) <= _time_sort_key(active_latest_month)],
        key=_time_sort_key,
    )
    missing = [month for month in historical_months if month not in candidate_months]
    if missing:
        return {"status": "fail", "reason": "candidate_missing_historical_months", "missingMonths": missing}

    active_sales = _historical_sales_frame(
        active_frame,
        historical_months,
    )
    candidate_sales = _historical_sales_frame(
        candidate_frame,
        historical_months,
    )
    country_samples: list[dict[str, Any]] = []
    make_samples: list[dict[str, Any]] = []

    def compare(
        left: pd.Series,
        right: pd.Series,
        *,
        scope: str,
        destination: list[dict[str, Any]],
    ) -> None:
        for month in historical_months:
            left_value = float(left.get(month, 0) or 0)
            right_value = float(right.get(month, 0) or 0)
            if left_value != right_value:
                destination.append({
                    "scope": scope,
                    "month": month,
                    "activeSales": _serialize_numeric_value(left_value),
                    "candidateSales": _serialize_numeric_value(right_value),
                    "deltaSales": _serialize_numeric_value(right_value - left_value),
                })

    compare(
        active_sales.sum(),
        candidate_sales.sum(),
        scope="country",
        destination=country_samples,
    )
    compared_make_count = 0
    if "Make" in active_frame.columns and "Make" in candidate_frame.columns:
        active_grouped = active_sales.groupby(active_frame["Make"].astype("string").fillna("").str.strip(), dropna=False).sum()
        candidate_grouped = candidate_sales.groupby(candidate_frame["Make"].astype("string").fillna("").str.strip(), dropna=False).sum()
        for make in sorted(set(active_grouped.index) | set(candidate_grouped.index)):
            compared_make_count += 1
            compare(
                active_grouped.loc[make] if make in active_grouped.index else pd.Series(0, index=historical_months),
                candidate_grouped.loc[make] if make in candidate_grouped.index else pd.Series(0, index=historical_months),
                scope=f"make:{make}",
                destination=make_samples,
            )

    make_model_deltas, make_model_samples, compared_make_model_count = (
        _single_country_make_model_deltas(
            active_frame=active_frame,
            candidate_frame=candidate_frame,
            active_sales=active_sales,
            candidate_sales=candidate_sales,
            historical_months=historical_months,
        )
    )
    (
        analysis_dimension_samples,
        compared_analysis_dimensions,
        compared_analysis_group_count,
    ) = _single_country_analysis_dimension_samples(
        active_frame=active_frame,
        candidate_frame=candidate_frame,
        active_sales=active_sales,
        candidate_sales=candidate_sales,
        historical_months=historical_months,
    )
    confirmed: list[dict[str, Any]] = []
    residual = make_model_deltas
    if not country_samples and make_model_deltas:
        confirmed, residual = _apply_confirmed_sc011_reclassifications(
            country=country,
            active_latest_month=active_latest_month,
            historical_months=historical_months,
            delta_vectors=make_model_deltas,
            source_upload_sha256=source_upload_sha256,
        )
    unconfirmed_candidates, unpaired_make_models = (
        _unconfirmed_sc011_reclassification_candidates(
            residual=residual,
            historical_months=historical_months,
        )
    )
    historical_reclassification = (
        _build_historical_reclassification_country_report(
            country=country,
            active_frame=active_frame,
            candidate_frame=candidate_frame,
            active_sales=active_sales,
            candidate_sales=candidate_sales,
            historical_months=historical_months,
            country_mismatch_count=len(country_samples),
            analysis_dimension_samples=analysis_dimension_samples,
            confirmed_make_model_reclassifications=confirmed,
            unconfirmed_make_model_candidates=unconfirmed_candidates,
            unpaired_make_models=unpaired_make_models,
        )
    )

    effective_dimension_samples = (
        make_model_samples if compared_make_model_count else make_samples
    )
    samples = [*country_samples, *effective_dimension_samples]
    if country_samples:
        status = "fail"
        cause = "historical_sales_changed"
    elif residual:
        status = "fail"
        cause = "unconfirmed_make_model_reclassification"
    elif analysis_dimension_samples:
        status = "fail"
        cause = "historical_analysis_dimension_reclassification"
    elif confirmed:
        status = "confirmed"
        cause = "confirmed_make_model_reclassification"
    elif make_samples:
        status = "fail"
        cause = "make_dimension_reclassification"
    else:
        status = "pass"
        cause = None

    residual_mismatch_count = sum(
        int(delta.reindex(historical_months, fill_value=0).ne(0).sum())
        for delta in residual.values()
    )
    impacted_make_models = [
        {"Make": make, "Model": model}
        for make, model in sorted(make_model_deltas)
    ]
    return {
        "status": status,
        "reason": cause,
        "comparedThrough": active_latest_month,
        "comparedMonthCount": len(historical_months),
        "comparedMakeCount": compared_make_count,
        "comparedMakeModelCount": compared_make_model_count,
        "mismatchCount": len(samples),
        "countryMismatchCount": len(country_samples),
        "makeMismatchCount": len(make_samples),
        "makeModelMismatchCount": len(make_model_samples),
        "analysisDimensionMismatchCount": len(
            analysis_dimension_samples
        ),
        "comparedAnalysisDimensions": compared_analysis_dimensions,
        "comparedAnalysisGroupCount": compared_analysis_group_count,
        "unconfirmedMakeModelMismatchCount": residual_mismatch_count,
        "impactedMakes": sorted(
            {
                make
                for make, _model in make_model_deltas
            }
            | {
                str(sample["scope"]).removeprefix("make:")
                for sample in make_samples
            }
        ),
        "impactedMakeModels": impacted_make_models,
        "impactedMonths": sorted(
            {str(sample["month"]) for sample in samples},
            key=_time_sort_key,
        ),
        "confirmedReclassifications": confirmed,
        "unconfirmedReclassificationCandidates": unconfirmed_candidates,
        "unpairedMakeModels": unpaired_make_models,
        "historicalReclassification": historical_reclassification,
        "mismatchSamples": [
            *samples,
            *analysis_dimension_samples,
        ][:20],
    }


def _single_country_source_feedback(*, rule_id: str, country: str, metrics: dict[str, Any]) -> str | None:
    """Translate measured validation failures into copy-ready washer feedback."""
    def columns(name: str) -> str:
        values = metrics.get(name)
        return "、".join(str(value) for value in values) if isinstance(values, list) else ""

    if rule_id == "SC001":
        return f"请重新导出 {country} 的完整月份列。当前文件未识别到有效月份；请保留历史月份和本次最新月份，并确保销量字段为可解析数字。"
    if rule_id == "SC002":
        return f"请检查 {country} 的月份销量值。文件包含月份标题但没有可用数值；请不要把销量列清空、转成文本或用格式符号替代数值。"
    if rule_id == "SC003":
        return f"请提供不早于当前 active {metrics.get('active') or '月份'} 的 {country} 数据。本次最新月份为 {metrics.get('candidate') or '未知'}，发生回退；请确认导出筛选条件包含本次要推进的最新月份。"
    if rule_id == "SC004":
        return f"请去除 {country} 文件中的完全相同配置行（检测到 {metrics.get('duplicateRows', 0)} 行）。只删除所有配置字段均相同的重复记录；价格、Registration type、动力或车身不同的版本不能合并。"
    if rule_id == "SC005":
        return f"请修正 {country} 文件中的负销量（检测到 {metrics.get('negativeSalesCells', 0)} 个单元格）。请按原始 JATO 导出确认更正、冲销或缺失值的处理方式，不能直接把负数绝对值化。"
    if rule_id == "SC006":
        suspicious_months = columns("months")
        months_detail = f"：{suspicious_months}" if suspicious_months else ""
        return f"请检查 {country} 的历史月份是否被重复拼接或重复累计{months_detail}。候选销量接近 active 的两倍；请从原始导出重新生成，不要把历史全量文件再追加到已有历史上。"
    if rule_id == "SC007":
        return f"请仅提供 {country} 的单国数据。系统发现未上传国家的分区也发生变化；请检查洗数流程是否混入其他国家或复写了共享数据。"
    if rule_id == "SC008":
        return "本次无法证明未上传国家未被影响。请保留单国文件的国家范围、源文件哈希和分区清单，以便重新提交时完成隔离校验。"
    if rule_id == "SC009":
        return f"请在 {country} 的洗数后 Data Export 中保留业务字段：{columns('missingMaterialColumns') or '见 Review 明细'}。这些列在当前 active 数据中有实际值，不能用 Retail price 或其他相近字段自动替代；请从原始 JATO 导出补齐后重新输出同一月份文件。"
    if rule_id == "SC010":
        return f"{country} 文件新增字段：{columns('extraColumns') or '见 Review 明细'}。请提供字段定义、单位和是否应保留的确认；新增列不应覆盖或改名现有业务列。"
    if rule_id == "SC011":
        confirmed = metrics.get("confirmedReclassifications")
        unconfirmed = metrics.get("unconfirmedReclassificationCandidates")
        unpaired = metrics.get("unpairedMakeModels")
        confirmed_text = ""
        if isinstance(confirmed, list) and confirmed:
            confirmed_items = []
            for item in confirmed:
                if not isinstance(item, dict):
                    continue
                source = item.get("source") if isinstance(item.get("source"), dict) else {}
                target = item.get("target") if isinstance(item.get("target"), dict) else {}
                confirmed_items.append(
                    f"{source.get('Make')}/{source.get('Model')}→"
                    f"{target.get('Make')}/{target.get('Model')}"
                    f"（{item.get('transferredSales', 0)} 台）"
                )
            if confirmed_items:
                confirmed_text = "已确认：" + "；".join(confirmed_items) + "。"
        if metrics.get("reason") == "confirmed_make_model_reclassification":
            return (
                f"{country} 的国家历史月销量总量与 active 一致，且车型重分类已完成业务确认。"
                f"{confirmed_text}Review 时请核对确认编号、逐月转移量和 candidate 指纹；"
                "不得把该确认扩展到名称相似的其他车型。"
            )
        if metrics.get("reason") == "unconfirmed_make_model_reclassification":
            unconfirmed_items = []
            if isinstance(unconfirmed, list):
                for item in unconfirmed:
                    if not isinstance(item, dict):
                        continue
                    source = item.get("source") if isinstance(item.get("source"), dict) else {}
                    target = item.get("target") if isinstance(item.get("target"), dict) else {}
                    unconfirmed_items.append(
                        f"{source.get('Make')}/{source.get('Model')}→"
                        f"{target.get('Make')}/{target.get('Model')}"
                        f"（{item.get('transferredSales', 0)} 台）"
                    )
            unpaired_items = []
            if isinstance(unpaired, list):
                for item in unpaired:
                    if isinstance(item, dict):
                        unpaired_items.append(
                            f"{item.get('Make')}/{item.get('Model')}"
                        )
            unpaired_text = (
                f"另有未能成对解释的车型：{'、'.join(unpaired_items)}。"
                if unpaired_items
                else ""
            )
            return (
                f"{country} 的国家历史月销量总量与 active 一致，但仍有未确认的 Make/Model 重分类。"
                f"{confirmed_text}"
                f"待确认映射：{'；'.join(unconfirmed_items) or '见 Review 明细'}。"
                f"{unpaired_text}"
                "请逐项确认旧 Make/Model → 新 Make/Model 映射；未确认项目请恢复已发布历史归类。"
            )
        if metrics.get("reason") == "make_dimension_reclassification":
            return f"{country} 的国家历史月销量总量与 active 一致，但 Make 归类发生变化。受影响品牌：{columns('impactedMakes') or '见 Review 明细'}；受影响月份：{columns('impactedMonths') or '见 Review 明细'}。请确认是否调整了品牌/车型映射；若为有意重分类，请提供旧 Make → 新 Make/Model 映射和业务确认，否则恢复已发布历史归类。"
        if (
            metrics.get("reason")
            == "historical_analysis_dimension_reclassification"
        ):
            return (
                f"{country} 的国家与 Make/Model 历史销量总量一致，但"
                "已发布的分析维度切片发生变化。受检维度："
                f"{columns('comparedAnalysisDimensions') or '见 Review 明细'}。"
                "请洗数人员恢复历史月份的 Registration type、Segment、"
                "Body、Powertrain、Fuel、Version/Trim 归类；若确需重分类，"
                "请提供逐月、旧值→新值及销量转移量的业务确认。"
            )
        return f"请恢复 {country} 在 {metrics.get('comparedThrough') or '已有'} 之前的历史销量。系统发现 {metrics.get('mismatchCount', 0)} 处历史销量差异；本次更新只能新增或修正经确认的最新月份，不能重写已发布历史月份。"
    if rule_id == "SC012":
        row_delta = int(metrics.get("rowDelta", 0) or 0)
        stability = metrics.get("historicalSalesStability")
        if isinstance(stability, dict) and stability.get("status") == "confirmed":
            history_note = "历史国家总量守恒，重分类已按确认映射核对"
        elif isinstance(stability, dict) and stability.get("status") == "pass":
            history_note = "历史销量已通过核对"
        else:
            history_note = "历史销量需要一并复核"
        return f"{country} 本次配置行较 active {'减少' if row_delta < 0 else '增加'} {abs(row_delta)} 行，{history_note}。请说明洗数时的去重、零销量配置过滤和车型下架规则，并提供清洗前后配置行数；不要仅为凑行数复制或补造车型。"
    if rule_id == "SC013":
        return f"{country} 缺少旧月份 YTD 派生列：{columns('missingDerivedYtdColumns') or '见 Review 明细'}。这不是发布 blocker；请确认最新月份的 YTD 字段仍由 Jan 至当月销量计算，并在后续导出中保持 YTD 列命名和口径一致。"
    if rule_id == "SC014":
        return f"{country} 缺少已停用的可选静态字段：{columns('missingDeprecatedOptionalColumns') or '见 Review 明细'}。无需为本次月更补齐；系统只会在配置键匹配且 active 旧值一致时沿用旧值，新配置或旧值冲突的配置保持为空，且不会用 Retail price 等相近字段冒充原字段。"
    return None


def _ordered_distinct_strings(values: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for raw_value in values:
        value = str(raw_value).strip()
        if not value or value in seen:
            continue
        seen.add(value)
        ordered.append(value)
    return ordered


def _logical_country_display_map(
    values: pd.Series,
    *,
    path_label: str,
) -> dict[str, str]:
    normalized = values.astype("string").fillna("").str.strip()
    match_keys = normalized.str.casefold()
    display_by_key: dict[str, list[str]] = {}
    for country_key, display in zip(match_keys, normalized, strict=False):
        key = str(country_key)
        rendered = str(display)
        if not key or not rendered:
            continue
        variants = display_by_key.setdefault(key, [])
        if rendered not in variants:
            variants.append(rendered)
    ambiguous = {
        key: variants
        for key, variants in display_by_key.items()
        if len(variants) > 1
    }
    if ambiguous:
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "ambiguous_logical_country",
                "message": (
                    f"{path_label} 中同一逻辑国家存在多个大小写/空格展示值，"
                    "继续会造成国家重复累加；请先统一国家字段。"
                ),
                "sourceFeedback": (
                    "请洗数人员统一同一国家的拼写、大小写和首尾空格；"
                    "一个逻辑国家只能保留一个展示值。"
                ),
                "countries": [
                    {
                        "logicalKey": key,
                        "displayValues": variants,
                    }
                    for key, variants in sorted(ambiguous.items())
                ],
            },
        )
    return {
        key: variants[0]
        for key, variants in display_by_key.items()
    }


def _serialize_numeric_value(value: Any) -> int | float | None:
    if value is None or pd.isna(value):
        return None
    number = float(value)
    if number.is_integer():
        return int(number)
    return number


def _load_monthly_sales_frame(path: Path, *, path_label: str) -> pd.DataFrame:
    if not path.exists():
        raise HTTPException(status_code=409, detail=f"{path_label} 不存在：{path}")
    suffix = path.suffix.lower()
    if suffix == ".parquet":
        try:
            import pyarrow.parquet as pq

            schema_columns = [
                str(column).strip()
                for column in pq.read_schema(path).names
            ]
            country_column = _find_country_column(schema_columns)
            month_columns = _detect_month_columns(schema_columns)
            if country_column is None or not month_columns:
                raise HTTPException(
                    status_code=409,
                    detail=f"{path_label} 缺少国家列或月份列。",
                )
            frame = pd.read_parquet(
                path,
                columns=[country_column, *month_columns],
            )
            frame.columns = [str(column).strip() for column in frame.columns]
            return frame
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(
                status_code=409,
                detail=f"读取 {path_label} 的国家/月字段失败。",
            ) from exc
    if suffix in ALLOWED_UPLOAD_EXTENSIONS:
        frame = _read_excel_with_fallback(path, sheet_name=0)
        frame.columns = [str(column).strip() for column in frame.columns]
        return frame
    raise HTTPException(
        status_code=409,
        detail=f"{path_label} 不是支持的 tabular 数据文件：{path.name}",
    )


def _collect_country_monthly_sales(
    frame: pd.DataFrame,
    *,
    countries: list[str],
    path_label: str,
) -> dict[str, dict[str, int | float]]:
    if frame.empty or not countries:
        return {}
    country_column = _find_country_column(list(frame.columns))
    if country_column is None:
        raise HTTPException(
            status_code=409,
            detail=f"{path_label} 缺少国家列，无法生成逐月销量核对表。",
        )
    month_columns = _detect_month_columns(list(frame.columns))
    if not month_columns:
        raise HTTPException(
            status_code=409,
            detail=f"{path_label} 缺少月份列，无法生成逐月销量核对表。",
        )
    normalized_countries = (
        frame[country_column].astype("string").fillna("").str.strip()
    )
    _logical_country_display_map(
        normalized_countries,
        path_label=path_label,
    )
    country_keys = normalized_countries.str.casefold()
    requested_country_keys = {
        str(country).strip().casefold(): str(country).strip()
        for country in countries
        if str(country).strip()
    }
    working = frame.loc[
        country_keys.isin(requested_country_keys), [country_column, *month_columns]
    ].copy()
    if working.empty:
        return {}
    logical_country_column = "__jato_logical_country"
    working[logical_country_column] = country_keys.loc[working.index]
    for column in month_columns:
        working[column] = pd.to_numeric(working[column], errors="coerce")
    grouped = working.groupby(logical_country_column, dropna=False, sort=False)[
        month_columns
    ].sum(min_count=1)
    result: dict[str, dict[str, int | float]] = {}
    for country in countries:
        country_key = str(country).strip().casefold()
        if country_key not in grouped.index:
            continue
        values = grouped.loc[country_key]
        result[country] = {
            month: serialized
            for month in month_columns
            if (serialized := _serialize_numeric_value(values.get(month))) is not None
        }
    return result


def _raise_for_missing_country_rows(
    *,
    path_label: str,
    row_count: int,
) -> None:
    if row_count <= 0:
        return
    raise HTTPException(
        status_code=409,
        detail={
            "blockerType": "missing_country_rows",
            "message": (
                f"{path_label} 有 {row_count} 行国家字段为空；"
                "为避免月更流程静默丢行，已停止处理。"
            ),
            "sourceFeedback": (
                "请洗数人员补齐每一条数据的 Country/国家字段，"
                "不得留空或仅包含空格。"
            ),
            "rowCount": row_count,
        },
    )


def _collect_country_monthly_sales_from_path(
    path: Path,
    *,
    countries: list[str],
    path_label: str,
) -> dict[str, dict[str, int | float]]:
    """Aggregate requested country/month totals without loading a parquet whole."""
    if path.suffix.lower() != ".parquet":
        return _collect_country_monthly_sales(
            _load_monthly_sales_frame(path, path_label=path_label),
            countries=countries,
            path_label=path_label,
        )
    if not path.exists():
        raise HTTPException(status_code=409, detail=f"{path_label} 不存在：{path}")
    if not countries:
        return {}

    try:
        import pyarrow.parquet as pq

        parquet_file = pq.ParquetFile(path)
        schema_columns = [
            str(column).strip()
            for column in parquet_file.schema_arrow.names
        ]
        country_column = _find_country_column(schema_columns)
        month_columns = _detect_month_columns(schema_columns)
        if country_column is None or not month_columns:
            raise HTTPException(
                status_code=409,
                detail=f"{path_label} 缺少国家列或月份列。",
            )

        requested_country_keys = {
            str(country).strip().casefold(): str(country).strip()
            for country in countries
            if str(country).strip()
        }
        totals: dict[str, dict[str, int | float]] = {}
        display_values: list[str] = []
        seen_display_values: set[str] = set()
        missing_country_rows = 0
        for batch in parquet_file.iter_batches(
            batch_size=SMART_MERGE_SCAN_BATCH_ROWS,
            columns=[country_column, *month_columns],
            use_threads=False,
        ):
            frame = batch.to_pandas(use_threads=False)
            frame.columns = [str(column).strip() for column in frame.columns]
            normalized = (
                frame[country_column]
                .astype("string")
                .fillna("")
                .str.strip()
            )
            missing_country_rows += int(normalized.eq("").sum())
            for display in normalized.unique().tolist():
                value = str(display)
                if value and value not in seen_display_values:
                    seen_display_values.add(value)
                    display_values.append(value)
            country_keys = normalized.str.casefold()
            requested_mask = country_keys.isin(requested_country_keys)
            if not bool(requested_mask.any()):
                continue
            working = frame.loc[requested_mask, month_columns].copy()
            logical_country_column = "__jato_logical_country"
            working[logical_country_column] = country_keys.loc[
                requested_mask
            ].to_numpy()
            for column in month_columns:
                working[column] = pd.to_numeric(
                    working[column], errors="coerce"
                )
            grouped = working.groupby(
                logical_country_column,
                dropna=False,
                sort=False,
            )[month_columns].sum(min_count=1)
            for country_key, values in grouped.iterrows():
                country_totals = totals.setdefault(str(country_key), {})
                for month in month_columns:
                    value = values.get(month)
                    if value is None or pd.isna(value):
                        continue
                    if month in country_totals:
                        country_totals[month] += value
                    else:
                        country_totals[month] = value

        _raise_for_missing_country_rows(
            path_label=path_label,
            row_count=missing_country_rows,
        )
        _logical_country_display_map(
            pd.Series(display_values, dtype="string"),
            path_label=path_label,
        )
        return {
            country: {
                month: serialized
                for month in month_columns
                if (
                    serialized := _serialize_numeric_value(
                        totals.get(str(country).strip().casefold(), {}).get(month)
                    )
                )
                is not None
            }
            for country in countries
            if str(country).strip().casefold() in totals
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=409,
            detail=f"读取 {path_label} 的国家/月字段失败。",
        ) from exc


def _collect_frame_countries(frame: pd.DataFrame, *, path_label: str) -> list[str]:
    country_column = _find_country_column(list(frame.columns))
    if country_column is None:
        raise HTTPException(
            status_code=409,
            detail=f"{path_label} 缺少国家列，无法执行 publish 销量防重校验。",
        )
    values = frame[country_column].astype("string").fillna("").str.strip()
    return list(
        _logical_country_display_map(
            values,
            path_label=path_label,
        ).values()
    )


def _collect_countries_from_path(
    path: Path,
    *,
    path_label: str,
) -> list[str]:
    """Collect country labels without materializing an entire parquet column."""
    if path.suffix.lower() != ".parquet":
        return _collect_frame_countries(
            _load_monthly_sales_frame(path, path_label=path_label),
            path_label=path_label,
        )
    if not path.exists():
        raise HTTPException(status_code=409, detail=f"{path_label} 不存在：{path}")
    try:
        import pyarrow.parquet as pq

        schema_columns = [
            str(column).strip()
            for column in pq.read_schema(path).names
        ]
        country_column = _find_country_column(schema_columns)
        if country_column is None:
            raise HTTPException(
                status_code=409,
                detail=f"{path_label} 缺少国家列，无法执行 publish 销量防重校验。",
            )
        country_value_map = _smart_merge_country_value_map(
            path,
            country_column=country_column,
            path_label=path_label,
        )
        return [
            str(item["display"])
            for item in country_value_map.values()
        ]
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=409,
            detail=f"读取 {path_label} 的国家字段失败。",
        ) from exc


def _is_near_sales_doubling(
    *,
    reference_sales: int | float | None,
    candidate_sales: int | float | None,
) -> tuple[bool, float | None]:
    if reference_sales is None or candidate_sales is None:
        return False, None
    reference = float(reference_sales)
    candidate = float(candidate_sales)
    if reference < SALES_DOUBLING_MIN_REFERENCE_SALES:
        return False, None
    if abs(candidate - reference) < SALES_DOUBLING_MIN_ABSOLUTE_DELTA:
        return False, None
    ratio = candidate / reference if reference else 0.0
    if SALES_DOUBLING_RATIO_MIN <= ratio <= SALES_DOUBLING_RATIO_MAX:
        return True, ratio
    return False, ratio


def _find_publish_sales_doubling_anomalies(
    *,
    active_parquet_path: Path,
    candidate_parquet_path: Path,
) -> list[dict[str, Any]]:
    """Detect likely duplicate country merges before candidate data becomes active."""
    active_countries = _collect_countries_from_path(
        active_parquet_path,
        path_label="当前 active 数据集",
    )
    countries = sorted(active_countries)
    if not countries:
        return []

    active_totals = _collect_country_monthly_sales_from_path(
        active_parquet_path,
        countries=countries,
        path_label="当前 active 数据集",
    )
    candidate_totals = _collect_country_monthly_sales_from_path(
        candidate_parquet_path,
        countries=countries,
        path_label="candidate 数据集",
    )

    anomalies: list[dict[str, Any]] = []
    for country in countries:
        reference_months = active_totals.get(country, {})
        candidate_months = candidate_totals.get(country, {})
        common_months = sorted(
            set(reference_months.keys()) & set(candidate_months.keys()),
            key=_time_sort_key,
        )
        suspicious_months: list[dict[str, Any]] = []
        for month in common_months:
            reference_sales = reference_months.get(month)
            candidate_sales = candidate_months.get(month)
            is_doubled, ratio = _is_near_sales_doubling(
                reference_sales=reference_sales,
                candidate_sales=candidate_sales,
            )
            if not is_doubled:
                continue
            suspicious_months.append(
                {
                    "month": month,
                    "referenceSales": reference_sales,
                    "candidateSales": candidate_sales,
                    "ratio": round(float(ratio or 0), 3),
                }
            )

        if len(suspicious_months) < SALES_DOUBLING_MIN_MONTH_COUNT:
            continue

        recent_months = common_months[-12:]
        reference_rolling = sum(
            float(reference_months.get(month) or 0) for month in recent_months
        )
        candidate_rolling = sum(
            float(candidate_months.get(month) or 0) for month in recent_months
        )
        rolling_ratio = (
            candidate_rolling / reference_rolling if reference_rolling else None
        )
        anomalies.append(
            {
                "country": country,
                "suspiciousMonthCount": len(suspicious_months),
                "sampleMonths": suspicious_months[:SALES_DOUBLING_SAMPLE_LIMIT],
                "referenceRolling12": _serialize_numeric_value(reference_rolling),
                "candidateRolling12": _serialize_numeric_value(candidate_rolling),
                "rolling12Ratio": (
                    round(float(rolling_ratio), 3)
                    if rolling_ratio is not None
                    else None
                ),
            }
        )
    return anomalies


def _find_publish_historical_sales_changes(
    *,
    active_parquet_path: Path,
    candidate_parquet_path: Path,
) -> list[dict[str, Any]]:
    """Block any country/month sales rewrite already present in active.

    This invariant catches a one-unit append just as reliably as a 2x append.
    New months are intentionally outside the comparison window.
    """
    active_countries = _collect_countries_from_path(
        active_parquet_path,
        path_label="当前 active 数据集",
    )
    active_totals = _collect_country_monthly_sales_from_path(
        active_parquet_path,
        countries=active_countries,
        path_label="当前 active 数据集",
    )
    candidate_totals = _collect_country_monthly_sales_from_path(
        candidate_parquet_path,
        countries=active_countries,
        path_label="candidate 数据集",
    )
    changes: list[dict[str, Any]] = []
    for country in active_countries:
        reference_months = active_totals.get(country, {})
        proposed_months = candidate_totals.get(country, {})
        month_changes: list[dict[str, Any]] = []
        for month in sorted(reference_months, key=_time_sort_key):
            reference = reference_months.get(month)
            proposed = proposed_months.get(month)
            if reference is None and proposed is None:
                continue
            if (
                reference is not None
                and proposed is not None
                and abs(float(reference) - float(proposed)) <= 1e-9
            ):
                continue
            month_changes.append(
                {
                    "month": month,
                    "activeSales": reference,
                    "candidateSales": proposed,
                    "deltaSales": (
                        _serialize_numeric_value(float(proposed) - float(reference))
                        if reference is not None and proposed is not None
                        else None
                    ),
                }
            )
        if month_changes:
            changes.append(
                {
                    "country": country,
                    "changedMonthCount": len(month_changes),
                    "sampleMonths": month_changes[:6],
                }
            )
    return changes


def _find_candidate_duplicate_configurations(
    candidate_parquet_path: Path,
    *,
    countries: list[str] | None = None,
) -> list[dict[str, Any]]:
    """Find exact static configuration duplicates one country at a time.

    Monthly/YTD values are deliberately excluded from the key.  Two rows that
    describe the same business configuration can otherwise double only the new
    month and evade both the historical-total guard and the older near-2x
    heuristic.
    """
    candidate_latest = _collect_dataset_country_latest_months(
        candidate_parquet_path
    )
    requested_country_keys = {
        country.strip().casefold()
        for country in (countries or [])
        if country.strip()
    }
    duplicates: list[dict[str, Any]] = []
    for country in candidate_latest:
        if (
            requested_country_keys
            and country.strip().casefold() not in requested_country_keys
        ):
            continue
        frame = _load_parquet_country_subset(
            candidate_parquet_path,
            country,
            path_label=f"candidate 数据集（{country}）",
        )
        key_columns = [
            column
            for column in _single_country_configuration_key_columns(frame)
            if not str(column).startswith("__source_")
        ]
        if not key_columns:
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": "duplicate_configuration_guard_unavailable",
                    "message": (
                        f"{country} candidate 无法识别静态配置字段，"
                        "不能证明新月份没有重复累加。"
                    ),
                    "sourceFeedback": None,
                },
            )
        normalized_keys = _normalized_static_key_frame(
            frame,
            key_columns,
        )
        duplicate_mask = normalized_keys.duplicated(keep=False)
        duplicate_rows = int(duplicate_mask.sum())
        if duplicate_rows <= 0:
            continue
        sample_columns = [
            column
            for column in (
                _find_country_column(list(frame.columns)),
                "Make",
                "Model",
                "Version name",
                "Trim level",
                "Powertrain type",
                "Registration type",
            )
            if column and column in frame.columns
        ]
        sample_frame = (
            frame.loc[duplicate_mask, sample_columns]
            .astype("string")
            .fillna("")
            .drop_duplicates()
            .head(10)
        )
        duplicate_keys = normalized_keys.loc[duplicate_mask]
        duplicate_keys = duplicate_keys.sort_values(
            by=list(duplicate_keys.columns),
            kind="stable",
        ).reset_index(drop=True)
        duplicate_fingerprint = hashlib.sha256()
        duplicate_fingerprint.update(
            json.dumps(
                key_columns,
                ensure_ascii=False,
                separators=(",", ":"),
            ).encode("utf-8")
        )
        duplicate_fingerprint.update(b"\n")
        duplicate_fingerprint.update(
            duplicate_keys.to_csv(
                index=False,
                lineterminator="\n",
            ).encode("utf-8")
        )
        duplicates.append(
            {
                "country": country,
                "duplicateRows": duplicate_rows,
                "duplicateGroupCount": int(
                    duplicate_keys.drop_duplicates().shape[0]
                ),
                "keyColumnCount": len(key_columns),
                "duplicateFingerprint": duplicate_fingerprint.hexdigest(),
                "samples": sample_frame.to_dict(orient="records"),
            }
        )
        del frame
    return duplicates


def _publish_exact_content_value_series(
    series: pd.Series,
) -> pd.Series:
    """Canonicalize storage dtypes while preserving raw string labels."""
    if pd.api.types.is_numeric_dtype(series.dtype):
        numbers = pd.to_numeric(series, errors="coerce")
        return numbers.map(
            lambda value: (
                "null:"
                if pd.isna(value)
                else f"number:{format(float(value), '.17g')}"
            )
        )
    return series.map(
        lambda value: (
            "null:"
            if pd.isna(value)
            else f"string:{value}"
        )
    )


def _canonical_country_content_fingerprint(
    frame: pd.DataFrame,
    columns: list[str],
) -> str:
    """Hash the normalized row multiset without making row order significant."""
    canonical = pd.DataFrame(
        {
            column: _publish_exact_content_value_series(frame[column])
            for column in columns
        }
    )
    row_digests = sorted(
        hashlib.sha256(
            json.dumps(
                row,
                ensure_ascii=False,
                separators=(",", ":"),
            ).encode("utf-8")
        ).digest()
        for row in canonical.itertuples(index=False, name=None)
    )
    fingerprint = hashlib.sha256()
    fingerprint.update(
        json.dumps(
            columns,
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode("utf-8")
    )
    fingerprint.update(b"\n")
    for row_digest in row_digests:
        fingerprint.update(row_digest)
    return fingerprint.hexdigest()


def _logical_country_map(
    countries: list[str],
) -> dict[str, str] | None:
    result: dict[str, str] = {}
    for raw_country in countries:
        country = str(raw_country).strip()
        country_key = country.casefold()
        if not country or country_key in result:
            return None
        result[country_key] = country
    return result


def _publish_duplicate_scope_contract(
    *,
    payload: dict[str, Any],
    active_countries: list[str],
    candidate_countries: list[str],
    refresh_report: dict[str, Any],
) -> dict[str, Any]:
    """Validate the immutable lineage used by the legacy-duplicate exception."""
    artifacts = payload.get("artifacts")
    approval = payload.get("reviewApproval")
    ingest_digest = payload.get("ingestDigest")
    upload_inspection = payload.get("uploadInspection")
    summaries = payload.get("summaries")
    smart_merge = (
        summaries.get("smartMerge")
        if isinstance(summaries, dict)
        and isinstance(summaries.get("smartMerge"), dict)
        else {}
    )
    static_summary = (
        smart_merge.get("deprecatedStaticCarryForward")
        if isinstance(
            smart_merge.get("deprecatedStaticCarryForward"),
            dict,
        )
        else {}
    )
    stored_untouched_checks = (
        static_summary.get("untouchedCountryChecks")
        if isinstance(
            static_summary.get("untouchedCountryChecks"),
            dict,
        )
        else {}
    )
    partition_check = (
        artifacts.get("untouchedPartitionCheck")
        if isinstance(artifacts, dict)
        and isinstance(artifacts.get("untouchedPartitionCheck"), dict)
        else {}
    )
    incremental = (
        refresh_report.get("incremental")
        if isinstance(refresh_report.get("incremental"), dict)
        else {}
    )

    country_scope = (
        [str(country) for country in payload.get("countryScope", [])]
        if isinstance(payload.get("countryScope"), list)
        else []
    )
    ingest_countries = (
        [str(country) for country in ingest_digest.get("countries", [])]
        if isinstance(ingest_digest, dict)
        and isinstance(ingest_digest.get("countries"), list)
        else []
    )
    inspected_countries = (
        [
            str(country)
            for country in upload_inspection.get("countries", [])
        ]
        if isinstance(upload_inspection, dict)
        and isinstance(upload_inspection.get("countries"), list)
        else []
    )
    scope_map = _logical_country_map(country_scope)
    ingest_map = _logical_country_map(ingest_countries)
    inspected_map = _logical_country_map(inspected_countries)
    active_map = _logical_country_map(active_countries)
    candidate_map = _logical_country_map(candidate_countries)
    stored_check_map = _logical_country_map(
        [str(country) for country in stored_untouched_checks]
    )

    job_type = str(payload.get("jobType") or "")
    expected_route = {
        "single_country": "single_country",
        "partial_country": "partial_country",
    }.get(job_type)
    expected_source_scope = {
        "single_country": "target_country_partition_only",
        "partial_country": "target_country_partitions_only",
    }.get(job_type)
    errors: list[str] = []
    if not isinstance(artifacts, dict):
        errors.append("artifacts_missing")
    elif str(artifacts.get("candidateScope") or "") != "full_smart_merge":
        errors.append("candidate_scope_not_full_smart_merge")
    if expected_route is None:
        errors.append("job_type_not_partial")
    if not isinstance(ingest_digest, dict):
        errors.append("ingest_digest_missing")
    elif str(ingest_digest.get("route") or "") != expected_route:
        errors.append("ingest_route_mismatch")
    if not isinstance(upload_inspection, dict):
        errors.append("upload_inspection_missing")
    if (
        scope_map is None
        or ingest_map is None
        or inspected_map is None
        or not scope_map
        or set(scope_map) != set(ingest_map)
        or set(scope_map) != set(inspected_map)
    ):
        errors.append("target_country_lineage_mismatch")
    if active_map is None or candidate_map is None:
        errors.append("dataset_country_identity_ambiguous")
    elif set(active_map) != set(candidate_map):
        errors.append("full_candidate_country_set_changed")
    elif scope_map is not None and not (
        set(scope_map) < set(active_map)
    ):
        errors.append("target_scope_not_strict_active_subset")
    if partition_check.get("status") != "pass":
        errors.append("untouched_partition_check_not_pass")
    if str(incremental.get("scope") or "") != "full_smart_merge":
        errors.append("refresh_scope_not_full_smart_merge")
    if (
        expected_source_scope is None
        or str(incremental.get("sourceCandidateScope") or "")
        != expected_source_scope
    ):
        errors.append("refresh_source_scope_mismatch")
    if isinstance(ingest_digest, dict) and (
        str(ingest_digest.get("candidateScope") or "")
        != expected_source_scope
    ):
        errors.append("digest_candidate_scope_mismatch")

    active_base_fingerprint = str(
        payload.get("activeBaseFingerprint") or ""
    )
    approved_active_fingerprint = (
        str(approval.get("activeBaseFingerprint") or "")
        if isinstance(approval, dict)
        else ""
    )
    digest_active_fingerprint = (
        str(ingest_digest.get("activeDatasetVersion") or "")
        if isinstance(ingest_digest, dict)
        else ""
    )
    if not (
        _valid_sha256(active_base_fingerprint)
        and active_base_fingerprint == approved_active_fingerprint
        and active_base_fingerprint == digest_active_fingerprint
    ):
        errors.append("active_lineage_mismatch")

    expected_untouched_keys = (
        set(active_map) - set(scope_map)
        if active_map is not None and scope_map is not None
        else set()
    )
    if (
        stored_check_map is None
        or set(stored_check_map) != expected_untouched_keys
    ):
        errors.append("smart_merge_untouched_proof_coverage_mismatch")
    return {
        "status": "pass" if not errors else "fail",
        "errors": errors,
        "targetCountryMap": scope_map or {},
        "activeCountryMap": active_map or {},
        "candidateCountryMap": candidate_map or {},
        "untouchedCountryKeys": sorted(expected_untouched_keys),
        "storedUntouchedChecks": stored_untouched_checks,
        "activeBaseFingerprint": active_base_fingerprint or None,
        "approvedActiveFingerprint": approved_active_fingerprint or None,
        "ingestActiveFingerprint": digest_active_fingerprint or None,
        "candidateFingerprint": (
            approval.get("candidateFingerprint")
            if isinstance(approval, dict)
            else None
        ),
    }


def _publish_untouched_country_content_evidence(
    *,
    country: str,
    active_frame: pd.DataFrame,
    candidate_frame: pd.DataFrame,
    stored_check: dict[str, Any] | None,
) -> dict[str, Any]:
    """Re-prove one untouched country against current active at Publish."""
    active_columns = [str(column) for column in active_frame.columns]
    candidate_columns = [str(column) for column in candidate_frame.columns]
    missing_active_columns = [
        column for column in active_columns if column not in candidate_frame
    ]
    candidate_only_columns = [
        column for column in candidate_columns if column not in active_frame
    ]
    non_null_candidate_only_columns = [
        column
        for column in candidate_only_columns
        if bool(candidate_frame[column].notna().any())
    ]
    row_count_equal = len(active_frame) == len(candidate_frame)
    signatures_available = not missing_active_columns
    active_signature: str | None = None
    candidate_signature: str | None = None
    active_fingerprint: str | None = None
    candidate_fingerprint: str | None = None
    if signatures_available:
        active_signature = _canonical_country_content_signature(
            active_frame,
            active_columns,
        )
        candidate_signature = _canonical_country_content_signature(
            candidate_frame,
            active_columns,
        )
        active_fingerprint = _canonical_country_content_fingerprint(
            active_frame,
            active_columns,
        )
        candidate_fingerprint = _canonical_country_content_fingerprint(
            candidate_frame,
            active_columns,
        )

    stored_proof_pass = bool(
        isinstance(stored_check, dict)
        and stored_check.get("status") == "pass"
        and int(stored_check.get("rowCount", -1) or -1)
        == len(active_frame)
        and str(stored_check.get("canonicalSignature") or "")
        == str(active_signature or "")
        and stored_check.get("candidateOnlyColumnsNull") is True
    )
    content_equal = bool(
        row_count_equal
        and signatures_available
        and not non_null_candidate_only_columns
        and active_signature == candidate_signature
        and active_fingerprint == candidate_fingerprint
        and stored_proof_pass
    )
    return {
        "country": country,
        "status": "pass" if content_equal else "fail",
        "activeRows": int(len(active_frame)),
        "candidateRows": int(len(candidate_frame)),
        "rowCountEqual": row_count_equal,
        "missingActiveColumns": missing_active_columns,
        "candidateOnlyColumns": candidate_only_columns,
        "nonNullCandidateOnlyColumns": non_null_candidate_only_columns,
        "activeCanonicalSignature": active_signature,
        "candidateCanonicalSignature": candidate_signature,
        "activeContentFingerprint": active_fingerprint,
        "candidateContentFingerprint": candidate_fingerprint,
        "storedSmartMergeProofPass": stored_proof_pass,
    }


def _publish_duplicate_configuration_assessment(
    *,
    payload: dict[str, Any],
    active_parquet_path: Path,
    candidate_parquet_path: Path,
    refresh_report_path: Path | None = None,
) -> dict[str, Any]:
    """Separate new duplicates from unchanged active legacy duplicates."""
    candidate_duplicates = _find_candidate_duplicate_configurations(
        candidate_parquet_path
    )
    artifacts = payload.get("artifacts")
    ingest_digest = payload.get("ingestDigest")
    scoped_smart_merge_claim = bool(
        str(payload.get("jobType") or "")
        in {"single_country", "partial_country"}
        or (
            isinstance(ingest_digest, dict)
            and str(ingest_digest.get("route") or "")
            in {"single_country", "partial_country"}
        )
    )
    if not scoped_smart_merge_claim:
        return {
            "blocking": [
                {**entry, "duplicateStatus": "candidate_duplicate"}
                for entry in candidate_duplicates
            ],
            "inherited": [],
            "guard": {
                "status": "not_applicable",
                "policy": "full_candidate_zero_duplicates",
            },
        }

    try:
        refresh_report = (
            _read_json(refresh_report_path)
            if refresh_report_path is not None
            else {}
        )
    except Exception:
        refresh_report = {}
    active_latest = _collect_dataset_country_latest_months(
        active_parquet_path
    )
    candidate_latest = _collect_dataset_country_latest_months(
        candidate_parquet_path
    )
    scope_contract = _publish_duplicate_scope_contract(
        payload=payload,
        active_countries=list(active_latest),
        candidate_countries=list(candidate_latest),
        refresh_report=refresh_report,
    )
    if scope_contract["status"] != "pass":
        return {
            "blocking": [
                {
                    "country": None,
                    "duplicateRows": sum(
                        int(entry.get("duplicateRows", 0) or 0)
                        for entry in candidate_duplicates
                    ),
                    "duplicateGroupCount": sum(
                        int(entry.get("duplicateGroupCount", 0) or 0)
                        for entry in candidate_duplicates
                    ),
                    "keyColumnCount": None,
                    "samples": [],
                    "duplicateStatus": "duplicate_guard_scope_invalid",
                    "scopeErrors": scope_contract["errors"],
                },
                *[
                    {**entry, "duplicateStatus": "candidate_duplicate"}
                    for entry in candidate_duplicates
                ],
            ],
            "inherited": [],
            "guard": scope_contract,
        }

    target_country_keys = set(scope_contract["targetCountryMap"])
    target_duplicates: list[dict[str, Any]] = []
    candidate_untouched_duplicates: dict[str, dict[str, Any]] = {}
    for entry in candidate_duplicates:
        country_key = str(entry.get("country") or "").strip().casefold()
        if country_key in target_country_keys:
            target_duplicates.append(
                {**entry, "duplicateStatus": "target_country_duplicate"}
            )
        else:
            candidate_untouched_duplicates[country_key] = entry

    untouched_country_keys = list(
        scope_contract["untouchedCountryKeys"]
    )
    active_duplicates = _find_candidate_duplicate_configurations(
        active_parquet_path,
        countries=[
            scope_contract["activeCountryMap"][country_key]
            for country_key in untouched_country_keys
        ],
    )
    active_by_country = {
        str(entry.get("country") or "").strip().casefold(): entry
        for entry in active_duplicates
    }
    blocking = list(target_duplicates)
    inherited: list[dict[str, Any]] = []
    content_evidence: list[dict[str, Any]] = []
    stored_checks = scope_contract["storedUntouchedChecks"]
    stored_checks_by_key = {
        str(country).strip().casefold(): check
        for country, check in stored_checks.items()
        if isinstance(check, dict)
    }
    for country_key in untouched_country_keys:
        country = scope_contract["activeCountryMap"][country_key]
        active_frame = _load_parquet_country_subset(
            active_parquet_path,
            country,
            path_label=f"Publish active（{country}）",
        )
        candidate_country = scope_contract["candidateCountryMap"][
            country_key
        ]
        candidate_frame = _load_parquet_country_subset(
            candidate_parquet_path,
            candidate_country,
            path_label=f"Publish candidate（{candidate_country}）",
        )
        evidence = _publish_untouched_country_content_evidence(
            country=country,
            active_frame=active_frame,
            candidate_frame=candidate_frame,
            stored_check=stored_checks_by_key.get(country_key),
        )
        content_evidence.append(evidence)
        del active_frame
        del candidate_frame

        candidate_entry = candidate_untouched_duplicates.get(country_key)
        active_entry = active_by_country.get(country_key)
        if evidence["status"] != "pass":
            source_entry = candidate_entry or active_entry or {
                "country": country,
                "duplicateRows": 0,
                "duplicateGroupCount": 0,
                "keyColumnCount": None,
                "samples": [],
            }
            blocking.append(
                {
                    **source_entry,
                    "duplicateStatus": "untouched_country_content_changed",
                    "activeDuplicateRows": (
                        active_entry.get("duplicateRows")
                        if isinstance(active_entry, dict)
                        else 0
                    ),
                    "candidateDuplicateRows": (
                        candidate_entry.get("duplicateRows")
                        if isinstance(candidate_entry, dict)
                        else 0
                    ),
                    "contentEvidence": evidence,
                }
            )
            continue
        if isinstance(candidate_entry, dict) and isinstance(
            active_entry,
            dict,
        ):
            inherited.append(
                {
                    **candidate_entry,
                    "duplicateStatus": "unchanged_active_duplicate",
                    "activeDuplicateFingerprint": active_entry.get(
                        "duplicateFingerprint"
                    ),
                    "contentFingerprint": evidence.get(
                        "candidateContentFingerprint"
                    ),
                }
            )
        elif candidate_entry is not None or active_entry is not None:
            source_entry = candidate_entry or active_entry
            blocking.append(
                {
                    **source_entry,
                    "duplicateStatus": (
                        "untouched_country_duplicate_changed"
                    ),
                    "activeDuplicateRows": (
                        active_entry.get("duplicateRows")
                        if isinstance(active_entry, dict)
                        else 0
                    ),
                    "candidateDuplicateRows": (
                        candidate_entry.get("duplicateRows")
                        if isinstance(candidate_entry, dict)
                        else 0
                    ),
                    "contentEvidence": evidence,
                }
            )
    guard = {
        **scope_contract,
        "status": "pass" if not blocking else "fail",
        "policy": "exact_untouched_country_content_only",
        "targetCountries": list(
            scope_contract["targetCountryMap"].values()
        ),
        "untouchedCountries": [
            scope_contract["activeCountryMap"][country_key]
            for country_key in untouched_country_keys
        ],
        "untouchedContentEvidence": content_evidence,
    }
    guard.pop("storedUntouchedChecks", None)
    return {
        "blocking": blocking,
        "inherited": inherited,
        "guard": guard,
    }


def _load_country_configuration_history_frame(
    path: Path,
    *,
    country: str,
    path_label: str,
) -> pd.DataFrame:
    """Load one country with stable analysis dimensions and monthly sales."""
    try:
        import pyarrow.parquet as pq

        schema_columns = [
            str(column).strip()
            for column in pq.read_schema(path).names
        ]
        country_column = _find_country_column(schema_columns)
        lookup = {
            str(column).strip().casefold(): str(column).strip()
            for column in schema_columns
        }
        make_column = lookup.get("make")
        model_column = lookup.get("model")
        dimension_columns = [
            column
            for aliases in HISTORICAL_DIMENSION_ALIASES
            if (
                column := next(
                    (
                        lookup[alias.casefold()]
                        for alias in aliases
                        if alias.casefold() in lookup
                    ),
                    None,
                )
            )
            is not None
        ]
        month_columns = _detect_month_columns(schema_columns)
        if (
            country_column is None
            or make_column is None
            or model_column is None
            or not month_columns
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": "historical_configuration_guard_unavailable",
                    "message": (
                        f"{path_label} 缺少 Country/Make/Model 或月份列，"
                        "无法证明已发布历史配置未被改写。"
                    ),
                    "sourceFeedback": None,
                },
            )
        requested_key = country.strip().casefold()
        country_value_map = _smart_merge_country_value_map(
            path,
            country_column=country_column,
            path_label=path_label,
        )
        matching_raw_values = list(
            country_value_map.get(requested_key, {}).get("rawValues", [])
        )
        if not matching_raw_values:
            return pd.DataFrame(
                columns=[
                    country_column,
                    make_column,
                    model_column,
                    *dimension_columns,
                    *month_columns,
                ]
            )
        selected_columns = list(
            dict.fromkeys(
                [
                    country_column,
                    make_column,
                    model_column,
                    *dimension_columns,
                    *month_columns,
                ]
            )
        )
        country_filter = (
            (country_column, "==", matching_raw_values[0])
            if len(matching_raw_values) == 1
            else (country_column, "in", matching_raw_values)
        )
        frame = pd.read_parquet(
            path,
            columns=selected_columns,
            filters=[country_filter],
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "historical_configuration_guard_unavailable",
                "message": (
                    f"读取 {path_label} 的单国 Make/Model 历史切片失败，"
                    "为避免扰乱历史已拒绝 Publish。"
                ),
                "sourceFeedback": None,
            },
        ) from exc
    frame.columns = [str(column).strip() for column in frame.columns]
    return frame


def _find_publish_historical_configuration_changes(
    *,
    active_parquet_path: Path,
    candidate_parquet_path: Path,
    source_upload_sha256: str | None = None,
    approved_reclassification_decisions: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Block unapproved Make/Model redistribution inside published months.

    Each country is loaded independently so the isolated worker never needs two
    full wide datasets in memory at once.
    """
    active_latest = _collect_dataset_country_latest_months(active_parquet_path)
    candidate_latest = _collect_dataset_country_latest_months(
        candidate_parquet_path
    )
    candidate_country_by_key = {
        country.casefold(): country
        for country in candidate_latest
    }
    approved_decisions = {
        str(country).strip().casefold(): str(decision).strip().lower()
        for country, decision in (
            approved_reclassification_decisions or {}
        ).items()
    }
    changes: list[dict[str, Any]] = []
    for country, active_latest_month in active_latest.items():
        candidate_country = candidate_country_by_key.get(country.casefold())
        if candidate_country is None:
            continue
        active_frame = _load_country_configuration_history_frame(
            active_parquet_path,
            country=country,
            path_label=f"当前 active 数据集（{country}）",
        )
        candidate_frame = _load_country_configuration_history_frame(
            candidate_parquet_path,
            country=candidate_country,
            path_label=f"candidate 数据集（{country}）",
        )
        stability = _single_country_historical_sales_stability(
            country=country,
            active_frame=active_frame,
            candidate_frame=candidate_frame,
            active_latest_month=active_latest_month,
            source_upload_sha256=source_upload_sha256,
        )
        if stability.get("status") == "pass":
            continue
        if (
            approved_decisions.get(country.casefold()) == "use_latest"
            and int(stability.get("countryMismatchCount") or 0) == 0
            and stability.get("reason")
            in {
                "unconfirmed_make_model_reclassification",
                "historical_analysis_dimension_reclassification",
                "make_dimension_reclassification",
                "confirmed_make_model_reclassification",
            }
        ):
            continue
        changes.append(
            {
                "country": country,
                "reason": stability.get("reason"),
                "comparedThrough": stability.get("comparedThrough"),
                "mismatchCount": int(stability.get("mismatchCount") or 0),
                "makeModelMismatchCount": int(
                    stability.get("makeModelMismatchCount") or 0
                ),
                "analysisDimensionMismatchCount": int(
                    stability.get("analysisDimensionMismatchCount") or 0
                ),
                "comparedAnalysisDimensions": (
                    stability.get("comparedAnalysisDimensions")
                    if isinstance(
                        stability.get("comparedAnalysisDimensions"),
                        list,
                    )
                    else []
                ),
                "impactedMakeModels": (
                    stability.get("impactedMakeModels")
                    if isinstance(
                        stability.get("impactedMakeModels"),
                        list,
                    )
                    else []
                ),
                "mismatchSamples": (
                    stability.get("mismatchSamples")
                    if isinstance(stability.get("mismatchSamples"), list)
                    else []
                ),
                "unconfirmedReclassificationCandidates": (
                    stability.get("unconfirmedReclassificationCandidates")
                    if isinstance(
                        stability.get(
                            "unconfirmedReclassificationCandidates"
                        ),
                        list,
                    )
                    else []
                ),
            }
        )
    return changes


def _render_sales_doubling_anomalies(anomalies: list[dict[str, Any]]) -> str:
    rendered_items: list[str] = []
    for item in anomalies[:5]:
        country = str(item.get("country", ""))
        sample_months = item.get("sampleMonths")
        months = []
        if isinstance(sample_months, list):
            for month_item in sample_months[:3]:
                if isinstance(month_item, dict):
                    months.append(str(month_item.get("month", "")))
        month_text = ",".join(month for month in months if month) or "-"
        ratio = item.get("rolling12Ratio")
        ratio_text = f", rolling12={ratio}x" if ratio is not None else ""
        rendered_items.append(f"{country}({month_text}{ratio_text})")
    extra = " 等" if len(anomalies) > 5 else ""
    return "；".join(rendered_items) + extra


def _resolve_review_reference_dataset(
    artifacts: dict[str, Any]
) -> tuple[Path | None, str]:
    active_parquet_path = _active_data_paths()["parquet"]
    if active_parquet_path.exists():
        return active_parquet_path, "网站当前 active"
    baseline_path = _project_path(str(artifacts.get("baselinePath") or "").strip())
    if baseline_path and baseline_path.exists():
        return baseline_path, "baseline"
    return None, "-"


def _build_country_monthly_sales_summary(
    *,
    countries: list[str],
    candidate_path: Path,
    reference_path: Path | None,
) -> list[dict[str, Any]]:
    candidate_totals = _collect_country_monthly_sales_from_path(
        candidate_path,
        countries=countries,
        path_label="candidate 数据集",
    )
    reference_totals: dict[str, dict[str, int | float]] = {}
    if reference_path is not None:
        reference_totals = _collect_country_monthly_sales_from_path(
            reference_path,
            countries=countries,
            path_label="参考数据集",
        )

    summaries: list[dict[str, Any]] = []
    for country in countries:
        reference_months = reference_totals.get(country, {})
        candidate_months = candidate_totals.get(country, {})
        months = sorted(
            set(reference_months.keys()) | set(candidate_months.keys()),
            key=_time_sort_key,
        )
        if not months:
            continue
        rows: list[dict[str, Any]] = []
        for month in months:
            reference_sales = reference_months.get(month)
            candidate_sales = candidate_months.get(month)
            if reference_sales is not None and candidate_sales is not None:
                delta_sales = _serialize_numeric_value(candidate_sales - reference_sales)
                change_status = "unchanged" if delta_sales == 0 else "changed"
            elif candidate_sales is not None:
                delta_sales = None
                change_status = "added"
            else:
                delta_sales = None
                change_status = "removed"
            rows.append(
                {
                    "month": month,
                    "referenceSales": reference_sales,
                    "candidateSales": candidate_sales,
                    "deltaSales": delta_sales,
                    "changeStatus": change_status,
                }
            )
        summaries.append({"country": country, "rows": rows})
    return summaries


def _build_historical_reclassification_report_from_paths(
    *,
    payload: dict[str, Any],
    countries: list[str],
    active_path: Path | None,
    candidate_path: Path | None,
) -> dict[str, Any]:
    if (
        active_path is None
        or candidate_path is None
        or active_path.suffix.casefold() != ".parquet"
        or candidate_path.suffix.casefold() != ".parquet"
        or not active_path.exists()
        or not candidate_path.exists()
    ):
        return _build_historical_reclassification_report(
            payload=payload,
            current_countries=[],
        )
    active_latest = _collect_dataset_country_latest_months(active_path)
    active_by_key = {
        country.casefold(): (country, latest_month)
        for country, latest_month in active_latest.items()
    }
    candidate_latest = _collect_dataset_country_latest_months(
        candidate_path
    )
    candidate_by_key = {
        country.casefold(): country
        for country in candidate_latest
    }
    upload = (
        payload.get("upload")
        if isinstance(payload.get("upload"), dict)
        else {}
    )
    source_upload_sha256 = str(upload.get("sha256") or "").strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", source_upload_sha256):
        source_upload_sha256 = None
    reports: list[dict[str, Any]] = []
    unavailable_countries: list[dict[str, Any]] = []
    current_stability_by_key: dict[str, dict[str, Any]] = {}
    for requested_country in countries:
        country_key = requested_country.strip().casefold()
        active_entry = active_by_key.get(country_key)
        candidate_country = candidate_by_key.get(country_key)
        if active_entry is None or candidate_country is None:
            current_stability_by_key[country_key] = {
                "status": "unavailable",
                "reason": "country_missing_from_active_or_candidate",
            }
            continue
        active_country, active_latest_month = active_entry
        try:
            active_frame = _load_country_configuration_history_frame(
                active_path,
                country=active_country,
                path_label=f"Review active（{active_country}）",
            )
            candidate_frame = _load_country_configuration_history_frame(
                candidate_path,
                country=candidate_country,
                path_label=f"Review candidate（{candidate_country}）",
            )
        except HTTPException as exc:
            current_stability_by_key[country_key] = {
                "status": "unavailable",
                "reason": "historical_configuration_guard_unavailable",
            }
            unavailable_countries.append(
                {
                    "country": active_country,
                    "detail": exc.detail,
                }
            )
            continue
        stability = _single_country_historical_sales_stability(
            country=active_country,
            active_frame=active_frame,
            candidate_frame=candidate_frame,
            active_latest_month=active_latest_month,
            source_upload_sha256=source_upload_sha256,
        )
        current_stability_by_key[country_key] = {
            "status": str(stability.get("status") or "unavailable"),
            "reason": stability.get("reason"),
        }
        country_report = stability.get("historicalReclassification")
        if isinstance(country_report, dict):
            reports.append(country_report)
        del active_frame
        del candidate_frame
    report = _build_historical_reclassification_report(
        payload=payload,
        current_countries=reports,
    )
    resolution = _historical_reclassification_resolution(payload)
    artifacts = payload.get("artifacts")
    candidate_scope = (
        str(artifacts.get("candidateScope") or "")
        if isinstance(artifacts, dict)
        else ""
    )
    if (
        report.get("status") == "resolved"
        and isinstance(resolution, dict)
        and candidate_scope == "full_smart_merge"
    ):
        decisions = _validated_historical_reclassification_resolution(
            resolution
        )
        validation: list[dict[str, Any]] = []
        raw_resolved_countries = report.get("countries")
        for country_report in (
            raw_resolved_countries
            if isinstance(raw_resolved_countries, list)
            else []
        ):
            if not isinstance(country_report, dict):
                continue
            country = str(country_report.get("country") or "").strip()
            country_key = country.casefold()
            current_stability = current_stability_by_key.get(
                country_key,
                {"status": "unavailable", "reason": "country_not_checked"},
            )
            validation_entry = (
                _historical_keep_active_resolution_validation(
                    country=country,
                    decision=decisions.get(country_key),
                    historical_stability=current_stability,
                )
            )
            if validation_entry is not None:
                validation.append(validation_entry)
        report["resolutionValidation"] = validation
    report["unavailableCountries"] = unavailable_countries[:10]
    if len(unavailable_countries) > 10:
        report["truncation"]["truncated"] = True
    return report


def _find_publish_country_regressions(
    *,
    active_parquet_path: Path,
    candidate_parquet_path: Path,
) -> list[dict[str, str | None]]:
    active_latest = _collect_dataset_country_latest_months(active_parquet_path)
    candidate_latest = _collect_dataset_country_latest_months(candidate_parquet_path)
    candidate_latest_by_key = {
        country.casefold(): month
        for country, month in candidate_latest.items()
    }
    regressions: list[dict[str, str | None]] = []
    for country, active_month in active_latest.items():
        candidate_month = candidate_latest_by_key.get(country.casefold())
        if active_month and not candidate_month:
            regressions.append(
                {
                    "country": country,
                    "activeLatestMonth": active_month,
                    "candidateLatestMonth": None,
                }
            )
            continue
        if (
            active_month
            and candidate_month
            and _time_sort_key(candidate_month) < _time_sort_key(active_month)
        ):
            regressions.append(
                {
                    "country": country,
                    "activeLatestMonth": active_month,
                    "candidateLatestMonth": candidate_month,
                }
            )
    return regressions


def _upload_session_root() -> Path:
    return MONTHLY_UPDATE_JOB_ROOT / "_upload_sessions"


def _upload_session_dir(upload_id: str) -> Path:
    return _upload_session_root() / upload_id


def _upload_session_state_path(upload_id: str) -> Path:
    return _upload_session_dir(upload_id) / UPLOAD_STATE_FILENAME


def _upload_session_chunk_dir(upload_id: str) -> Path:
    return _upload_session_dir(upload_id) / "chunks"


def _upload_digest_attempt_dir(upload_id: str) -> Path:
    return _upload_session_dir(upload_id) / DIGEST_ATTEMPT_DIRNAME


def _new_upload_digest_attempt(
    *,
    upload_id: str,
    attempt_number: int,
) -> dict[str, Any]:
    attempt_id = f"{attempt_number}-{uuid4().hex[:12]}"
    attempt_dir = _upload_digest_attempt_dir(upload_id)
    attempt_dir.mkdir(parents=True, exist_ok=True)
    return {
        "attemptId": attempt_id,
        "attemptNumber": attempt_number,
        "status": "launching",
        "logPath": _relative_to_project(
            attempt_dir / f"{attempt_id}.log"
        ),
        "receiptPath": _relative_to_project(
            attempt_dir / f"{attempt_id}.exit.json"
        ),
        "supervisorPid": None,
        "workerPid": None,
        "launchedAt": _utc_now().isoformat(),
        "supervisorMissingAt": None,
        "exit": None,
    }


def _upload_digest_attempt_artifact_path(
    state: dict[str, Any],
    key: str,
) -> Path | None:
    attempt = state.get("digestAttempt")
    if not isinstance(attempt, dict):
        return None
    candidate = _project_path(attempt.get(key))
    upload_id = str(state.get("uploadId") or "").strip()
    if candidate is None or not upload_id:
        return None
    expected_parent = _upload_digest_attempt_dir(upload_id).resolve()
    resolved = candidate.resolve()
    if resolved.parent != expected_parent:
        return None
    return resolved


def _read_digest_attempt_log_tail(state: dict[str, Any]) -> str | None:
    log_path = _upload_digest_attempt_artifact_path(state, "logPath")
    if log_path is None or not log_path.is_file():
        return None
    try:
        with log_path.open("rb") as handle:
            handle.seek(0, os.SEEK_END)
            size = handle.tell()
            handle.seek(max(size - DIGEST_ATTEMPT_LOG_TAIL_BYTES, 0))
            return handle.read(DIGEST_ATTEMPT_LOG_TAIL_BYTES).decode(
                "utf-8",
                errors="replace",
            )
    except OSError:
        return None


def _digest_attempt_matches(
    state: dict[str, Any],
    attempt_id: str | None,
) -> bool:
    attempt = state.get("digestAttempt")
    if not isinstance(attempt, dict):
        return attempt_id is None
    return bool(
        attempt_id
        and str(attempt.get("attemptId") or "") == attempt_id
    )


def _safe_internal_upload_filename(filename: str) -> str:
    """Return a bounded ASCII storage name while keeping display names in state."""
    normalized = _validate_upload_filename(filename)
    suffix = Path(normalized).suffix.lower()
    filename_digest = hashlib.sha256(normalized.encode("utf-8")).hexdigest()
    return (
        f"upload-{filename_digest[:UPLOAD_INTERNAL_FILENAME_DIGEST_LENGTH]}"
        f"{suffix}"
    )


def _upload_session_assembled_path(upload_id: str, filename: str) -> Path:
    return (
        _upload_session_dir(upload_id)
        / "assembled"
        / _safe_internal_upload_filename(filename)
    )


def _persisted_upload_session_assembled_path(
    state: dict[str, Any],
) -> Path | None:
    """Resolve a persisted path, including legacy display-named assemblies, safely."""
    assembled_value = str(state.get("assembledPath") or "").strip()
    if not assembled_value:
        return None
    upload_id = str(state.get("uploadId") or "").strip()
    candidate = _project_path(assembled_value)
    if not upload_id or candidate is None:
        raise RuntimeError("上传会话的 assembledPath 无效，请重新上传。")
    assembled_root = (_upload_session_dir(upload_id) / "assembled").resolve()
    resolved = candidate.resolve()
    if resolved.parent != assembled_root:
        raise RuntimeError("上传会话的 assembledPath 越界，请重新上传。")
    return resolved


def _job_upload_storage_path(job_id: str, filename: str) -> Path:
    return (
        _job_dir(job_id)
        / "uploads"
        / _safe_internal_upload_filename(filename)
    )


def _iter_upload_session_payloads() -> list[dict[str, Any]]:
    _upload_session_root().mkdir(parents=True, exist_ok=True)
    payloads: list[dict[str, Any]] = []
    for state_path in _upload_session_root().glob(f"*/{UPLOAD_STATE_FILENAME}"):
        try:
            payloads.append(_read_json(state_path))
        except Exception:
            continue
    return payloads


def _reconcile_expired_upload_sessions() -> list[str]:
    """Release abandoned browser uploads without deleting their evidence."""
    expired: list[str] = []
    now = _utc_now()
    for payload in _iter_upload_session_payloads():
        if str(payload.get("status") or "") not in {
            "pending",
            "uploading",
            "ready",
        }:
            continue
        updated_at = str(
            payload.get("updatedAt")
            or payload.get("createdAt")
            or ""
        ).strip()
        try:
            updated = datetime.fromisoformat(updated_at)
            if updated.tzinfo is None:
                updated = updated.replace(tzinfo=UTC)
        except ValueError:
            updated = datetime.fromtimestamp(0, UTC)
        if (now - updated).total_seconds() <= UPLOAD_SESSION_STALE_SECONDS:
            continue
        upload_id = str(payload.get("uploadId") or "").strip()
        if not upload_id:
            continue
        with _exclusive_file_lock(
            _upload_state_lock_path(upload_id),
            blocking=False,
        ) as acquired:
            if not acquired:
                continue
            latest = _load_upload_session(upload_id)
            if str(latest.get("status") or "") not in {
                "pending",
                "uploading",
                "ready",
            }:
                continue
            latest["status"] = "expired"
            latest["completedAt"] = now.isoformat()
            latest["failureDigest"] = {
                "code": "UPLOAD_SESSION_EXPIRED",
                "category": "lifecycle",
                "phase": "upload",
                "retryable": True,
                "message": "上传会话超过 24 小时未继续，已安全释放；active 未修改。",
                "sourceFeedback": None,
                "technicalDetail": {"lastUpdatedAt": updated_at or None},
                "nextAction": "start_new_upload",
            }
            _persist_upload_session(latest)
            expired.append(upload_id)
    return expired


def _active_upload_session_payloads() -> list[dict[str, Any]]:
    _reconcile_expired_upload_sessions()
    active_statuses = {
        "pending",
        "uploading",
        "assembling",
        "digesting",
        "ready",
    }
    return [
        payload
        for payload in _iter_upload_session_payloads()
        if str(payload.get("status") or "") in active_statuses
    ]


def _require_no_active_upload_sessions(
    *,
    action: str,
    excluding_ready_upload_id: str | None = None,
) -> None:
    excluded_upload_id = str(excluding_ready_upload_id or "").strip()
    active_uploads = [
        payload
        for payload in _active_upload_session_payloads()
        if not (
            excluded_upload_id
            and str(payload.get("uploadId") or "") == excluded_upload_id
            and str(payload.get("status") or "") == "ready"
        )
    ]
    if not active_uploads:
        return
    upload_ids = [
        str(payload.get("uploadId") or "")
        for payload in active_uploads[:5]
        if str(payload.get("uploadId") or "")
    ]
    rendered = "、".join(upload_ids) or "未知会话"
    raise HTTPException(
        status_code=409,
        detail=(
            f"存在尚未消费的上传/digest 会话，不能{action}：{rendered}。"
            "请先完成或明确放弃该上传，避免删除分片或与重 worker 抢内存。"
        ),
    )


def _terminal_upload_session_dirs() -> list[Path]:
    terminal_statuses = {"invalid", "consumed", "abandoned", "expired"}
    paths: list[Path] = []
    for payload in _iter_upload_session_payloads():
        if str(payload.get("status") or "") not in terminal_statuses:
            continue
        upload_id = str(payload.get("uploadId") or "").strip()
        if upload_id:
            paths.append(_upload_session_dir(upload_id))
    return paths


def _list_job_state_payloads() -> list[dict[str, Any]]:
    MONTHLY_UPDATE_JOB_ROOT.mkdir(parents=True, exist_ok=True)
    payloads: list[dict[str, Any]] = []
    for state_path in MONTHLY_UPDATE_JOB_ROOT.glob(f"*/{STATE_FILENAME}"):
        try:
            payloads.append(_read_json(state_path))
        except Exception:
            continue
    return payloads


def _load_upload_session(upload_id: str) -> dict[str, Any]:
    path = _upload_session_state_path(upload_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail="上传会话不存在或已失效。")
    return _read_json(path)


def _persist_upload_session(payload: dict[str, Any]) -> None:
    upload_id = str(payload["uploadId"])
    payload["updatedAt"] = _utc_now().isoformat()
    with _WRITE_LOCK:
        _write_json(_upload_session_state_path(upload_id), payload)


def _upload_session_owner(payload: dict[str, Any]) -> str:
    return str(
        payload.get("owner")
        or payload.get("triggeredBy")
        or ""
    ).strip()


def _same_upload_owner(left: str, right: str) -> bool:
    return bool(left.strip()) and left.strip().casefold() == right.strip().casefold()


def _require_upload_session_access(
    payload: dict[str, Any],
    *,
    requested_by: str,
    requested_role: str,
) -> None:
    owner = _upload_session_owner(payload)
    actor = str(requested_by or "").strip()
    role = str(requested_role or "").strip().casefold()
    if role in {"admin", "developer"}:
        return
    if owner and _same_upload_owner(owner, actor):
        return
    raise HTTPException(status_code=403, detail="无权访问该上传会话。")


def _find_upload_session_by_resume_key(
    *,
    resume_key: str,
    filename: str,
    size_bytes: int,
    owner: str,
) -> dict[str, Any] | None:
    if not resume_key:
        return None
    for payload in _iter_upload_session_payloads():
        if str(payload.get("resumeKey", "")) != resume_key:
            continue
        if str(payload.get("filename", "")) != filename:
            continue
        if int(payload.get("sizeBytes", 0) or 0) != size_bytes:
            continue
        if not _same_upload_owner(_upload_session_owner(payload), owner):
            continue
        if str(payload.get("status", "")) not in {
            "pending",
            "uploading",
            "assembling",
            "digesting",
            "ready",
        }:
            continue
        return payload
    return None


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(
        f".{path.name}.{os.getpid()}.{threading.get_ident()}.{uuid4().hex}.tmp"
    )
    try:
        with temp_path.open("w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_path, path)
    finally:
        temp_path.unlink(missing_ok=True)


def _read_json(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise RuntimeError(f"Expected JSON object: {path}")
    return payload


def _invalidate_jato_publish_runtime_caches() -> dict[str, Any]:
    result: dict[str, Any] = {
        "marketScanDeckLocal": {"enabled": False, "clearedCount": 0},
        "marketScanDeckRedis": {"enabled": False, "deletedCount": 0},
        "heroProductDeckLocal": {"enabled": False, "clearedCount": 0},
        "heroProductDeckRedis": {"enabled": False, "deletedCount": 0},
        "datasetToken": {
            "enabled": True,
            "message": "Parquet repository dataset token changes with active data artifacts.",
        },
    }
    try:
        from app.infra.redis_client import get_redis_client
        from app.services.hero_product_analysis_service import (
            clear_hero_product_deck_cache,
            invalidate_hero_product_deck_cache,
        )
        from app.services.market_scan_cache import invalidate_market_scan_deck_cache
        from app.services.market_scan_service import clear_market_scan_local_cache

        redis_client = get_redis_client()
        result["marketScanDeckLocal"] = clear_market_scan_local_cache()
        result["marketScanDeckRedis"] = invalidate_market_scan_deck_cache(redis_client)
        result["heroProductDeckLocal"] = clear_hero_product_deck_cache()
        result["heroProductDeckRedis"] = invalidate_hero_product_deck_cache(redis_client)
    except Exception as exc:
        result["error"] = str(exc)
        result["message"] = (
            "Runtime cache invalidation failed; dataset-token cache keys should still "
            "avoid stale MarketScan/HeroProduct data."
        )
    return result


def _write_jato_publish_cache_invalidation_evidence(
    *,
    job_id: str,
    published_at: datetime,
    triggered_by: str,
    cache_invalidation: dict[str, Any],
    active_paths: dict[str, Path],
) -> None:
    ledger_path = PROJECT_ROOT / "hermes" / "evidence_ledger.jsonl"
    stamp = published_at.strftime("%Y%m%dT%H%M%SZ")
    record = {
        "evidenceId": f"evidence.jato_monthly_update.cache_invalidation.{job_id}.{stamp}",
        "evidenceType": "runtime_cache_invalidation",
        "claim": "JATO monthly publish invalidated MarketScan runtime caches.",
        "sourceRef": f"jato_monthly_update_job::{job_id}",
        "artifactId": "feature.jato_monthly_update",
        "confidence": 1.0,
        "supportCount": 1,
        "contradictionCount": 0,
        "createdAt": published_at.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "details": {
            "publishedBy": triggered_by.strip() or "anonymous",
            "cacheInvalidation": cache_invalidation,
            "activeParquetPath": _relative_to_project(active_paths.get("parquet")),
            "activeManifestPath": _relative_to_project(active_paths.get("manifest")),
            "activePartitionPath": _relative_to_project(active_paths.get("partition")),
        },
    }
    ledger_path.parent.mkdir(parents=True, exist_ok=True)
    with ledger_path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(record, ensure_ascii=False) + "\n")


def _load_job_state(job_id: str) -> dict[str, Any]:
    path = _job_state_path(job_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail="JATO monthly update job not found")
    return _read_json(path)


def _persist_job_state(payload: dict[str, Any]) -> None:
    job_id = str(payload["jobId"])
    cancel_path = _job_cancel_request_path(job_id)
    if cancel_path.exists():
        try:
            cancel_request = _read_json(cancel_path)
        except Exception:
            cancel_request = {}
        cancelled_at = str(
            cancel_request.get("cancelledAt")
            or payload.get("finishedAt")
            or _utc_now().isoformat()
        )
        cancelled_by = str(
            cancel_request.get("cancelledBy") or "anonymous"
        )
        phase_at_cancel = str(
            cancel_request.get("phaseAtCancel")
            or payload.get("phase")
            or "unknown"
        )
        payload["status"] = "cancelled"
        payload["phase"] = "cancelled"
        payload["finishedAt"] = cancelled_at
        payload["error"] = (
            f"Cancelled by {cancelled_by} during {phase_at_cancel}"
        )
        existing_cancellation = payload.get("cancellation")
        payload["cancellation"] = {
            **(
                existing_cancellation
                if isinstance(existing_cancellation, dict)
                else {}
            ),
            **cancel_request,
            "cancelledAt": cancelled_at,
            "cancelledBy": cancelled_by,
            "phaseAtCancel": phase_at_cancel,
        }
    payload["updatedAt"] = _utc_now().isoformat()
    with _WRITE_LOCK:
        _write_json(_job_state_path(job_id), payload)


def _thread_is_alive(job_id: str) -> bool:
    worker = _RUNNING_THREADS.get(job_id)
    return bool(worker and worker.is_alive())


def _current_process_pid(payload: dict[str, Any]) -> int | None:
    current = payload.get("currentProcess")
    if not isinstance(current, dict):
        return None
    try:
        pid = int(current.get("pid") or 0)
    except (TypeError, ValueError):
        return None
    return pid if pid > 0 else None


def _process_exists(pid: int | None) -> bool:
    if pid is None or pid <= 0:
        return False
    proc_stat_path = Path(f"/proc/{pid}/stat")
    if proc_stat_path.exists():
        try:
            raw_stat = proc_stat_path.read_text(encoding="utf-8")
            command_end = raw_stat.rfind(")")
            process_state = (
                raw_stat[command_end + 2 :].split(maxsplit=1)[0]
                if command_end >= 0
                else ""
            )
            if process_state in {"Z", "X"}:
                return False
        except OSError:
            pass
    try:
        os.kill(pid, 0)
        return True
    except ProcessLookupError:
        return False
    except PermissionError:
        return True
    except OSError:
        return False


def _read_process_identity(pid: int) -> dict[str, Any] | None:
    """Read Linux process birth/command identity used to defend PID reuse."""
    stat_path = Path(f"/proc/{pid}/stat")
    cmdline_path = Path(f"/proc/{pid}/cmdline")
    if not stat_path.exists() or not cmdline_path.exists():
        return None
    try:
        raw_stat = stat_path.read_text(encoding="utf-8")
        command_end = raw_stat.rfind(")")
        if command_end < 0:
            return None
        stat_fields = raw_stat[command_end + 2 :].split()
        if len(stat_fields) <= 19:
            return None
        raw_cmdline = cmdline_path.read_bytes()
        arguments = [
            item.decode("utf-8", errors="replace")
            for item in raw_cmdline.split(b"\0")
            if item
        ]
        return {
            "startTimeTicks": str(stat_fields[19]),
            "cmdlineSha256": hashlib.sha256(raw_cmdline).hexdigest(),
            "arguments": arguments,
        }
    except OSError:
        return None


def _process_identity_matches(
    *,
    pid: int,
    expected_identity: dict[str, Any] | None,
    required_command_tokens: tuple[str, ...] = (),
) -> tuple[bool, dict[str, Any] | None]:
    current = _read_process_identity(pid)
    if current is None:
        return False, None
    if not isinstance(expected_identity, dict):
        return False, current
    if (
        str(expected_identity.get("startTimeTicks") or "")
        != str(current.get("startTimeTicks") or "")
        or str(expected_identity.get("cmdlineSha256") or "")
        != str(current.get("cmdlineSha256") or "")
    ):
        return False, current
    arguments = [
        str(argument)
        for argument in current.get("arguments", [])
    ]
    if any(token not in arguments for token in required_command_tokens):
        return False, current
    return True, current


def _process_is_digest_worker_for_upload(pid: int, upload_id: str) -> bool:
    """Avoid terminating an unrelated process if a stale PID was reused."""
    cmdline_path = Path(f"/proc/{pid}/cmdline")
    if not cmdline_path.exists():
        return False
    try:
        arguments = [
            item.decode("utf-8", errors="replace")
            for item in cmdline_path.read_bytes().split(b"\0")
            if item
        ]
    except OSError:
        return False
    return (
        "--digest-upload" in arguments
        and upload_id in arguments
        and any(
            Path(argument).name == MONTHLY_WORKER_SCRIPT_PATH.name
            for argument in arguments
        )
    )


def _set_current_process(
    *,
    job_id: str,
    pid: int,
    label: str,
    command: str,
) -> None:
    payload = _load_job_state(job_id)
    now = _utc_now().isoformat()
    payload["currentProcess"] = {
        "pid": pid,
        "label": label,
        "command": command,
        "startedAt": now,
        "lastHeartbeatAt": now,
        "identity": _read_process_identity(pid),
    }
    _persist_job_state(payload)


def _touch_current_process_heartbeat(
    *,
    job_id: str,
    pid: int,
) -> None:
    payload = _load_job_state(job_id)
    current = payload.get("currentProcess")
    if not isinstance(current, dict) or int(current.get("pid") or 0) != pid:
        return
    current["lastHeartbeatAt"] = _utc_now().isoformat()
    payload["currentProcess"] = current
    _persist_job_state(payload)


def _clear_current_process(
    *,
    job_id: str,
    pid: int | None = None,
) -> None:
    payload = _load_job_state(job_id)
    current = payload.get("currentProcess")
    if pid is not None and isinstance(current, dict) and int(current.get("pid") or 0) != pid:
        return
    payload["currentProcess"] = None
    _persist_job_state(payload)


def _ensure_job_not_cancelled(job_id: str) -> None:
    cancel_path = _job_cancel_request_path(job_id)
    if cancel_path.exists():
        request = _read_json(cancel_path)
        raise _JobCancelled(
            f"Cancelled by {request.get('cancelledBy') or 'anonymous'} "
            f"during {request.get('phaseAtCancel') or 'unknown'}"
        )
    payload = _load_job_state(job_id)
    if str(payload.get("status") or "") == "cancelled":
        raise _JobCancelled(str(payload.get("error") or f"Job {job_id} cancelled."))


def _infer_job_id_from_log_path(log_path: Path) -> str | None:
    job_id = log_path.parent.name
    return job_id if (_job_state_path(job_id)).exists() else None


def _terminate_process_group(
    pid: int,
    *,
    expected_identity: dict[str, Any] | None = None,
    required_command_tokens: tuple[str, ...] = (),
) -> dict[str, Any]:
    result: dict[str, Any] = {
        "pid": pid,
        "sigtermSent": False,
        "sigkillSent": False,
        "processAliveBefore": _process_exists(pid),
        "processAliveAfter": False,
        "identityVerified": False,
    }
    if not result["processAliveBefore"]:
        return result
    identity_verified, observed_identity = _process_identity_matches(
        pid=pid,
        expected_identity=expected_identity,
        required_command_tokens=required_command_tokens,
    )
    result["identityVerified"] = identity_verified
    result["observedIdentity"] = observed_identity
    if not identity_verified:
        result["processAliveAfter"] = _process_exists(pid)
        result["error"] = (
            "process identity could not be verified; refused to signal reused PID"
        )
        return result
    try:
        os.killpg(pid, signal.SIGTERM)
        result["sigtermSent"] = True
    except ProcessLookupError:
        result["processAliveAfter"] = False
        return result
    except OSError as exc:
        result["error"] = str(exc)
        result["processAliveAfter"] = _process_exists(pid)
        return result

    deadline = time.monotonic() + PROCESS_TERMINATE_GRACE_SECONDS
    while time.monotonic() < deadline:
        if not _process_exists(pid):
            result["processAliveAfter"] = False
            return result
        time.sleep(0.2)

    if _process_exists(pid):
        identity_verified, observed_identity = _process_identity_matches(
            pid=pid,
            expected_identity=expected_identity,
            required_command_tokens=required_command_tokens,
        )
        result["identityVerifiedBeforeSigkill"] = identity_verified
        result["observedIdentityBeforeSigkill"] = observed_identity
        if not identity_verified:
            result["processAliveAfter"] = True
            result["error"] = (
                "process identity changed before SIGKILL; refused to signal"
            )
            return result
        try:
            os.killpg(pid, signal.SIGKILL)
            result["sigkillSent"] = True
        except ProcessLookupError:
            pass
        except OSError as exc:
            result["error"] = str(exc)
    result["processAliveAfter"] = _process_exists(pid)
    return result


def _digest_process_termination_confirmed(
    pid: int,
    termination: dict[str, Any] | None,
) -> bool:
    """Only release resource gates after the old digest is conclusively gone."""
    if pid <= 0 or not _process_exists(pid):
        return True
    return bool(
        isinstance(termination, dict)
        and int(termination.get("pid") or 0) == pid
        and termination.get("processAliveAfter") is False
    )


def _terminate_digest_worker_with_evidence(
    *,
    pid: int,
    upload_id: str,
    expected_identity: dict[str, Any] | None,
    attempt_id: str | None = None,
) -> dict[str, Any]:
    required_tokens = (
        ("--supervise-digest-upload", upload_id, attempt_id)
        if attempt_id
        else ("--digest-upload", upload_id)
    )
    try:
        return _terminate_process_group(
            pid,
            expected_identity=expected_identity,
            required_command_tokens=required_tokens,
        )
    except Exception as exc:
        return {
            "pid": pid,
            "processAliveBefore": True,
            "processAliveAfter": _process_exists(pid),
            "identityVerified": False,
            "error": str(exc),
            "technicalDetail": traceback.format_exc(limit=4),
        }


def _job_log_probe(log_path: Path) -> dict[str, Any]:
    if not log_path.exists():
        return {
            "path": _relative_to_project(log_path),
            "exists": False,
            "updatedAt": None,
            "ageSeconds": None,
            "stale": True,
        }
    updated = datetime.fromtimestamp(log_path.stat().st_mtime, UTC)
    age_seconds = max(0, int((_utc_now() - updated).total_seconds()))
    return {
        "path": _relative_to_project(log_path),
        "exists": True,
        "updatedAt": updated.isoformat(),
        "ageSeconds": age_seconds,
        "stale": age_seconds > RUNNING_LOG_STALE_SECONDS,
    }


def _artifact_probe(payload: dict[str, Any]) -> list[dict[str, Any]]:
    artifacts = payload.get("artifacts") if isinstance(payload.get("artifacts"), dict) else {}
    probes: list[dict[str, Any]] = []
    for key in ("planPath", "rawCompareReportPath", "refreshReportPath", "stagingOutputPath"):
        value = artifacts.get(key)
        if not isinstance(value, str) or not value.strip():
            continue
        path = _project_path(value)
        probes.append({
            "key": key,
            "path": value,
            "exists": bool(path and path.exists()),
        })
    return probes


def _normalize_month(value: str) -> str:
    match = MONTH_PATTERN.fullmatch(value.strip())
    if match is None:
        raise HTTPException(status_code=400, detail="月份格式无效，需要 YYYY-MM，例如 2026-02")
    return f"{match.group(1)}-{int(match.group(2)):02d}"


def _parse_batch_id(value: str | None) -> tuple[str, int] | None:
    candidate = str(value or "").strip()
    match = BATCH_ID_PATTERN.fullmatch(candidate)
    if match is None:
        return None
    return match.group(1), int(match.group(2))


def _infer_month_token(value: str) -> str | None:
    dotted_match = re.findall(
        r"(?<!\d)(20\d{2})[._-](0?[1-9]|1[0-2])(?!\d)", value
    )
    if dotted_match:
        year, month = max(dotted_match)
        return f"{year}-{int(month):02d}"
    compact_match = re.findall(r"(?<!\d)(20\d{2})(0[1-9]|1[0-2])(?!\d)", value)
    if compact_match:
        year, month = max(compact_match)
        return f"{year}-{int(month):02d}"
    return None


def _normalize_filename(value: str | None) -> str:
    # Browsers normally submit a basename, but normalize both POSIX and Windows
    # separators so a client-controlled path can never become a storage path.
    raw_value = str(value or "jato-update.xlsx").replace("\\", "/")
    candidate = Path(raw_value).name.strip()
    if not candidate or candidate in {".", ".."}:
        return "jato-update.xlsx"
    return candidate


def _validate_upload(file: UploadFile) -> str:
    return _validate_upload_filename(file.filename or "jato-update.xlsx")


def _validate_upload_filename(filename: str) -> str:
    normalized = _normalize_filename(filename)
    if any(ord(character) < 32 or ord(character) == 127 for character in normalized):
        raise HTTPException(status_code=400, detail="上传文件名包含无效控制字符。")
    try:
        normalized.encode("utf-8")
    except UnicodeEncodeError as exc:
        raise HTTPException(status_code=400, detail="上传文件名不是有效 UTF-8 文本。") from exc
    suffix = Path(normalized).suffix.lower()
    if suffix not in ALLOWED_UPLOAD_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="JATO monthly update 仅支持 Excel 文件（.xlsx/.xlsm/.xls）。",
        )
    return normalized


def _normalize_size_bytes(value: Any) -> int:
    if isinstance(value, bool):
        raise HTTPException(status_code=400, detail="上传文件大小无效。") from None
    rendered = str(value or "").strip()
    if not re.fullmatch(r"[1-9]\d*", rendered):
        raise HTTPException(status_code=400, detail="上传文件大小无效。")
    size_bytes = int(rendered)
    if size_bytes <= 0:
        raise HTTPException(status_code=400, detail="上传文件为空，无法启动月更任务。")
    if size_bytes > UPLOAD_MAX_BYTES:
        raise HTTPException(
            status_code=413,
            detail=(
                f"上传文件超过系统上限 "
                f"{UPLOAD_MAX_BYTES / 1024 / 1024:.0f}MB，请缩小文件或联系管理员。"
            ),
        )
    return size_bytes


def _digest_worker_timeout_seconds(size_bytes: Any) -> int:
    """Scale digest time with compressed upload size, capped at 45 minutes."""
    try:
        normalized_size = max(int(size_bytes or 0), 0)
    except (TypeError, ValueError):
        normalized_size = 0
    extra_bytes = max(normalized_size - DIGEST_WORKER_BASE_SIZE_BYTES, 0)
    extra_mib = (extra_bytes + 1024 * 1024 - 1) // (1024 * 1024)
    return min(
        DIGEST_WORKER_BASE_TIMEOUT_SECONDS
        + extra_mib * DIGEST_WORKER_EXTRA_SECONDS_PER_MIB,
        DIGEST_WORKER_MAX_SECONDS,
    )


def _digest_failure_from_exit_receipt(
    *,
    state: dict[str, Any],
    receipt: dict[str, Any],
    digest_pid: int,
    phase: str,
    digest_process_identity: dict[str, Any] | None,
) -> dict[str, Any]:
    return_code_value = receipt.get("returnCode")
    return_code = (
        int(return_code_value)
        if isinstance(return_code_value, int)
        and not isinstance(return_code_value, bool)
        else None
    )
    termination_reason = str(receipt.get("terminationReason") or "")
    oom_kill_delta = int(receipt.get("oomKillDelta") or 0)
    supervisor_error = str(receipt.get("supervisorError") or "").strip()
    try:
        rss_limit_bytes = int(receipt.get("rssLimitBytes") or 0)
    except (TypeError, ValueError):
        rss_limit_bytes = 0
    rss_limit_gib = rss_limit_bytes / (1024**3)
    if termination_reason == "rss_limit" or return_code == 72:
        code = "DIGEST_MEMORY_LIMIT"
        category = "resource"
        message = (
            "上传文件 digest 的实际 RSS 连续超过"
            f" {rss_limit_gib:.1f} GiB 安全线，"
            "worker 已被隔离 supervisor 停止；active 未修改。"
        )
        next_action = "inspect_worker_memory_or_reduce_source"
    elif oom_kill_delta > 0:
        code = "DIGEST_CGROUP_OOM"
        category = "resource"
        message = (
            "digest 运行期间 cgroup 记录了 OOM kill；"
            "这不是已证明的数据错误，active 未修改。"
        )
        next_action = "inspect_worker_cgroup"
    elif return_code is not None and return_code < 0:
        code = "DIGEST_WORKER_SIGNALLED"
        category = "resource"
        message = (
            "上传文件 digest worker 被系统信号终止"
            f"（{receipt.get('signalName') or return_code}）；active 未修改。"
        )
        next_action = "retry_digest_or_contact_admin"
    elif return_code == 0:
        code = "DIGEST_RESULT_MISSING"
        category = "platform"
        message = (
            "digest worker 正常退出，但没有生成最终 digest 状态；"
            "active 未修改。"
        )
        next_action = "retry_digest_or_contact_admin"
    elif return_code is not None:
        code = "DIGEST_WORKER_EXITED"
        category = "platform"
        message = (
            f"digest worker 异常退出（退出码 {return_code}）；"
            "active 未修改。"
        )
        next_action = "retry_digest_or_contact_admin"
    else:
        code = "DIGEST_WORKER_UNAVAILABLE"
        category = "platform"
        message = (
            "digest supervisor 未能启动或等待 worker；active 未修改。"
        )
        next_action = "contact_admin"
    if supervisor_error and code == "DIGEST_RESULT_MISSING":
        code = "DIGEST_WORKER_UNAVAILABLE"
        message = "digest supervisor 运行失败；active 未修改。"
        next_action = "contact_admin"
    attempt = state.get("digestAttempt")
    attempt_id = (
        str(attempt.get("attemptId") or "")
        if isinstance(attempt, dict)
        else ""
    )
    return {
        "code": code,
        "category": category,
        "phase": phase or "digesting",
        "retryable": code in DIGEST_RETRYABLE_FAILURE_CODES,
        "message": message,
        "sourceFeedback": (
            "这不是已证明的 washed 数据错误；请把退出码、信号、"
            "峰值 RSS 和日志摘要交给平台管理员。"
        ),
        "technicalDetail": {
            "digestPid": digest_pid or None,
            "digestProcessIdentity": digest_process_identity,
            "digestAttemptId": attempt_id or None,
            "fileSizeBytes": int(state.get("sizeBytes") or 0),
            "exitReceipt": receipt,
            "logTail": _read_digest_attempt_log_tail(state),
        },
        "nextAction": next_action,
    }


def _reconcile_digest_attempt_receipt_locked(
    state: dict[str, Any],
) -> dict[str, Any]:
    """Fold one durable supervisor receipt into state under state.lock."""
    attempt = state.get("digestAttempt")
    if not isinstance(attempt, dict):
        return state
    receipt_path = _upload_digest_attempt_artifact_path(
        state,
        "receiptPath",
    )
    if receipt_path is None or not receipt_path.is_file():
        return state
    try:
        receipt = _read_json(receipt_path)
    except Exception:
        return state
    attempt_id = str(attempt.get("attemptId") or "")
    if (
        str(receipt.get("uploadId") or "")
        != str(state.get("uploadId") or "")
        or str(receipt.get("attemptId") or "") != attempt_id
    ):
        return state
    changed = False
    worker_pid = int(receipt.get("workerPid") or 0)
    if worker_pid > 0 and int(attempt.get("workerPid") or 0) != worker_pid:
        attempt["workerPid"] = worker_pid
        state["digestWorkerPid"] = worker_pid
        changed = True
    if str(receipt.get("status") or "") != "finished":
        if str(attempt.get("status") or "") != "running":
            attempt["status"] = "running"
            changed = True
        if changed:
            state["digestAttempt"] = attempt
            _persist_upload_session(state)
        return state

    with _exclusive_file_lock(
        _upload_digest_lock_path(str(state.get("uploadId") or "")),
        blocking=False,
    ) as digest_lock_acquired:
        if not digest_lock_acquired:
            return state

    digest_pid = int(state.get("digestPid") or 0)
    digest_process_identity = (
        state.get("digestProcessIdentity")
        if isinstance(state.get("digestProcessIdentity"), dict)
        else None
    )
    status = str(state.get("status") or "")
    receipt = dict(receipt)
    receipt["logPath"] = attempt.get("logPath")
    attempt["status"] = "finished"
    attempt["exit"] = receipt
    attempt["workerPid"] = worker_pid or None
    attempt["supervisorMissingAt"] = None
    state["digestAttempt"] = attempt
    state["digestPid"] = None
    state["digestWorkerPid"] = None
    state["digestProcessIdentity"] = None
    if status in {"assembling", "digesting"}:
        state["status"] = "invalid"
        state["completedAt"] = _utc_now().isoformat()
        state["failureDigest"] = _digest_failure_from_exit_receipt(
            state=state,
            receipt=receipt,
            digest_pid=digest_pid,
            phase=status,
            digest_process_identity=digest_process_identity,
        )
    _persist_upload_session(state)
    return state


def _normalize_sha256(value: Any, *, detail: str) -> str:
    candidate = str(value or "").strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", candidate):
        raise HTTPException(status_code=400, detail=detail)
    return candidate


def _sha256_hex_for_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha256_hex_for_path(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            hasher.update(chunk)
    return hasher.hexdigest()


def _sha256_hex_for_tree(path: Path) -> str:
    hasher = hashlib.sha256()
    for file_path in sorted(item for item in path.rglob("*") if item.is_file()):
        hasher.update(str(file_path.relative_to(path)).encode("utf-8"))
        hasher.update(_sha256_hex_for_path(file_path).encode("ascii"))
    return hasher.hexdigest()


def _read_excel_with_fallback(
    input_file: Path,
    *,
    sheet_name: str | int,
    nrows: int | None = None,
    usecols: list[str] | None = None,
) -> pd.DataFrame:
    kwargs: dict[str, Any] = {"sheet_name": sheet_name}
    if nrows is not None:
        kwargs["nrows"] = nrows
    if usecols is not None:
        kwargs["usecols"] = usecols
    try:
        return pd.read_excel(input_file, engine="calamine", **kwargs)
    except Exception:
        try:
            return pd.read_excel(input_file, **kwargs)
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail=f"读取上传 Excel 失败：{input_file.name}",
            ) from exc


def _detect_latest_month_from_upload(path: Path) -> str:
    header = _read_excel_with_fallback(
        path,
        sheet_name=DEFAULT_UPLOAD_SHEET_NAME,
        nrows=0,
    )
    header.columns = [str(column).strip() for column in header.columns]
    month_columns = _detect_month_columns(list(header.columns))
    if not month_columns:
        raise HTTPException(
            status_code=400,
            detail="上传文件中未识别到可用月份列，无法自动生成批次。",
        )
    month_frame = _read_excel_with_fallback(
        path,
        sheet_name=DEFAULT_UPLOAD_SHEET_NAME,
        usecols=month_columns,
    )
    month_frame.columns = [str(column).strip() for column in month_frame.columns]
    for column in reversed(month_columns):
        if column in month_frame.columns and _series_has_data(month_frame[column]):
            parsed = datetime.strptime(column.title(), "%Y %b")
            return f"{parsed.year}-{parsed.month:02d}"
    raise HTTPException(
        status_code=400,
        detail="上传文件的月份列均为空，无法自动生成批次。",
    )


def _allocate_batch_id(month: str) -> str:
    normalized_month = _normalize_month(month)
    highest_revision = 0
    for payload in _list_job_state_payloads():
        parsed = _parse_batch_id(payload.get("batchId"))
        if parsed is None:
            continue
        item_month, item_revision = parsed
        if item_month == normalized_month:
            highest_revision = max(highest_revision, item_revision)
    if PATCHES_ROOT.exists():
        for path in PATCHES_ROOT.glob("*"):
            if not path.is_dir():
                continue
            parsed = _parse_batch_id(path.name)
            if parsed is None:
                continue
            item_month, item_revision = parsed
            if item_month == normalized_month:
                highest_revision = max(highest_revision, item_revision)
    return f"{normalized_month}-r{highest_revision + 1}"


def _chunk_file_name(part_number: int) -> str:
    return f"{part_number:05d}.part"


def _expected_chunk_size(*, size_bytes: int, chunk_size: int, total_chunks: int, part_number: int) -> int:
    if part_number < total_chunks:
        return chunk_size
    consumed = chunk_size * (total_chunks - 1)
    return size_bytes - consumed


def get_jato_monthly_update_expected_chunk_size(
    *,
    upload_id: str,
    part_number: int,
    requested_by: str,
    requested_role: str,
) -> int:
    state = _load_upload_session(upload_id)
    _require_upload_session_access(
        state,
        requested_by=requested_by,
        requested_role=requested_role,
    )
    total_chunks = int(state.get("totalChunks", 0))
    if part_number < 1 or part_number > total_chunks:
        raise HTTPException(status_code=400, detail="分片序号超出范围。")
    return _expected_chunk_size(
        size_bytes=int(state.get("sizeBytes", 0)),
        chunk_size=int(state.get("chunkSize", UPLOAD_CHUNK_SIZE_BYTES)),
        total_chunks=total_chunks,
        part_number=part_number,
    )


def _collect_uploaded_chunk_numbers(upload_id: str) -> list[int]:
    chunk_dir = _upload_session_chunk_dir(upload_id)
    if not chunk_dir.exists():
        return []
    chunk_numbers: list[int] = []
    for path in chunk_dir.glob("*.part"):
        try:
            chunk_numbers.append(int(path.stem))
        except ValueError:
            continue
    chunk_numbers.sort()
    return chunk_numbers


def _uploaded_chunk_bytes(upload_id: str) -> int:
    chunk_dir = _upload_session_chunk_dir(upload_id)
    if not chunk_dir.exists():
        return 0
    return sum(
        path.stat().st_size
        for path in chunk_dir.glob("*.part")
        if path.is_file()
    )


def _validate_uploaded_chunks_complete(state: dict[str, Any]) -> tuple[list[int], int]:
    upload_id = str(state["uploadId"])
    total_chunks = int(state.get("totalChunks", 0))
    size_bytes = int(state.get("sizeBytes", 0))
    chunk_size = int(state.get("chunkSize", UPLOAD_CHUNK_SIZE_BYTES))
    if (
        total_chunks <= 0
        or size_bytes <= 0
        or size_bytes > UPLOAD_MAX_BYTES
    ):
        raise HTTPException(status_code=409, detail="上传会话大小元数据无效，请重新上传。")
    received_chunks = _collect_uploaded_chunk_numbers(upload_id)
    expected_numbers = list(range(1, total_chunks + 1))
    if received_chunks != expected_numbers:
        raise HTTPException(status_code=409, detail="上传分片尚未齐全，不能完成组装。")
    uploaded_bytes = 0
    for part_number in expected_numbers:
        chunk_path = (
            _upload_session_chunk_dir(upload_id)
            / _chunk_file_name(part_number)
        )
        expected_size = _expected_chunk_size(
            size_bytes=size_bytes,
            chunk_size=chunk_size,
            total_chunks=total_chunks,
            part_number=part_number,
        )
        actual_size = chunk_path.stat().st_size
        if actual_size != expected_size:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"分片 {part_number} 大小已变化，期望 {expected_size} 字节，"
                    f"实际 {actual_size} 字节；请重新上传。"
                ),
            )
        uploaded_bytes += actual_size
    if uploaded_bytes != size_bytes:
        raise HTTPException(
            status_code=409,
            detail=(
                f"上传总大小不匹配，声明 {size_bytes} 字节，"
                f"实际 {uploaded_bytes} 字节；请重新上传。"
            ),
        )
    return received_chunks, uploaded_bytes


def _serialize_upload_digest_attempt(
    payload: Any,
) -> dict[str, Any] | None:
    """Expose operational progress without leaking host paths or cmdlines."""
    if not isinstance(payload, dict):
        return None
    serialized = {
        key: payload.get(key)
        for key in (
            "attemptId",
            "attemptNumber",
            "status",
            "supervisorPid",
            "workerPid",
            "launchedAt",
            "supervisorMissingAt",
            "workerFinishedAt",
        )
    }
    exit_payload = payload.get("exit")
    serialized["exit"] = (
        {
            key: exit_payload.get(key)
            for key in (
                "status",
                "startedAt",
                "finishedAt",
                "elapsedSeconds",
                "returnCode",
                "signalNumber",
                "signalName",
                "terminationReason",
                "peakRssBytes",
                "rssWarningBytes",
                "rssLimitBytes",
                "rssWarningExceeded",
                "cgroupEventDelta",
                "oomKillDelta",
            )
        }
        if isinstance(exit_payload, dict)
        else None
    )
    return serialized


def _serialize_upload_session(payload: dict[str, Any]) -> dict[str, Any]:
    upload_id = str(payload["uploadId"])
    persisted_chunks = payload.get("receivedChunks")
    if isinstance(persisted_chunks, list) and persisted_chunks:
        received_chunks = [int(item) for item in persisted_chunks]
    else:
        received_chunks = _collect_uploaded_chunk_numbers(upload_id)
    persisted_uploaded_bytes = payload.get("uploadedBytes")
    uploaded_bytes = (
        int(persisted_uploaded_bytes)
        if persisted_uploaded_bytes is not None
        else _uploaded_chunk_bytes(upload_id)
    )
    raw_chunk_digests = payload.get("chunkDigests")
    chunk_digests = (
        {
            str(part_number): str(raw_chunk_digests.get(str(part_number)) or "")
            for part_number in received_chunks
            if re.fullmatch(
                r"[0-9a-f]{64}",
                str(raw_chunk_digests.get(str(part_number)) or ""),
            )
        }
        if isinstance(raw_chunk_digests, dict)
        else {}
    )
    return {
        "uploadId": upload_id,
        "filename": str(payload.get("filename", "")),
        "sizeBytes": int(payload.get("sizeBytes", 0)),
        "chunkSize": int(payload.get("chunkSize", UPLOAD_CHUNK_SIZE_BYTES)),
        "totalChunks": int(payload.get("totalChunks", 0)),
        "receivedChunkCount": len(received_chunks),
        "receivedChunks": received_chunks,
        "chunkDigests": chunk_digests,
        "uploadedBytes": uploaded_bytes,
        "status": str(payload.get("status", "pending")),
        "createdAt": payload.get("createdAt"),
        "updatedAt": payload.get("updatedAt"),
        "completedAt": payload.get("completedAt"),
        "assembledPath": payload.get("assembledPath"),
        "resumeKey": payload.get("resumeKey"),
        "fileSha256": payload.get("fileSha256"),
        "triggeredBy": payload.get("triggeredBy"),
        "ingestDigest": (
            payload.get("ingestDigest")
            if isinstance(payload.get("ingestDigest"), dict)
            else None
        ),
        "failureDigest": (
            payload.get("failureDigest")
            if isinstance(payload.get("failureDigest"), dict)
            else None
        ),
        "consumedJobId": payload.get("consumedJobId"),
        "digestPid": payload.get("digestPid"),
        "digestWorkerPid": payload.get("digestWorkerPid"),
        "digestLaunchedAt": payload.get("digestLaunchedAt"),
        "digestAttempts": int(payload.get("digestAttempts") or 0),
        "digestAttempt": _serialize_upload_digest_attempt(
            payload.get("digestAttempt")
        ),
    }


def _ensure_unique_archive_path(target: Path) -> Path:
    if not target.exists():
        return target
    stem = target.stem
    suffix = target.suffix
    counter = 1
    while True:
        candidate = target.with_name(f"{stem}-{counter}{suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def _move_to_archive(path: Path, archive_dir: Path) -> Path:
    archive_dir.mkdir(parents=True, exist_ok=True)
    target = _ensure_unique_archive_path(archive_dir / path.name)
    shutil.move(str(path), str(target))
    return target


def _latest_baseline_file(paths: list[Path]) -> Path | None:
    if not paths:
        return None
    return max(
        paths,
        key=lambda item: (
            _infer_month_token(item.stem) or "0000-00",
            item.stat().st_mtime,
            item.name,
        ),
    )


def _list_supported_excel_files(root: Path) -> list[Path]:
    if not root.exists():
        return []
    return sorted(
        path
        for path in root.glob("*")
        if path.is_file()
        and path.suffix.lower() in ALLOWED_UPLOAD_EXTENSIONS
        and not path.name.startswith("~$")
    )


def _load_dataset_frame(path: Path) -> pd.DataFrame:
    frame = pd.read_parquet(path)
    frame.columns = [str(column).strip() for column in frame.columns]
    return frame


def _detect_latest_month_from_dataset_frame(
    frame: pd.DataFrame, *, path_label: str
) -> str:
    month_columns = _detect_month_columns(list(frame.columns))
    if not month_columns:
        raise HTTPException(
            status_code=409,
            detail=f"{path_label} 缺少可识别的月份列，不能生成 baseline。",
        )
    for column in reversed(month_columns):
        if column in frame.columns and _series_has_data(frame[column]):
            parsed = datetime.strptime(column.title(), "%Y %b")
            return f"{parsed.year}-{parsed.month:02d}"
    raise HTTPException(
        status_code=409,
        detail=f"{path_label} 的月份列均为空，不能生成 baseline。",
    )


def _collect_dataset_country_count(frame: pd.DataFrame, *, path_label: str) -> int:
    country_column = _find_country_column(list(frame.columns))
    if country_column is None:
        raise HTTPException(
            status_code=409,
            detail=f"{path_label} 缺少国家列，不能生成 baseline。",
        )
    values = frame[country_column].astype("string").fillna("").str.strip()
    count = int(values[values != ""].nunique())
    if count <= 0:
        raise HTTPException(
            status_code=409,
            detail=f"{path_label} 不包含有效国家数据，不能生成 baseline。",
        )
    return count


def _baseline_snapshot_filename(*, latest_month: str, country_count: int) -> str:
    year, month = latest_month.split("-")
    jato_month = f"{year}.{int(month)}"
    country_tag = f"full-{country_count}countries" if country_count > 0 else "full"
    return f"JATO-{jato_month}-{country_tag}-baseline.xlsx"


def _measure_path_usage(path: Path) -> tuple[int, int, int]:
    if not path.exists():
        return (0, 0, 0)
    if path.is_file():
        return (int(path.stat().st_size), 1, 0)
    total_bytes = 0
    file_count = 0
    dir_count = 0
    for root, dirnames, filenames in os.walk(path):
        dir_count += len(dirnames)
        for filename in filenames:
            candidate = Path(root) / filename
            try:
                stat = candidate.stat()
            except FileNotFoundError:
                continue
            total_bytes += int(stat.st_size)
            file_count += 1
    return (total_bytes, file_count, dir_count)


def _measure_combined_usage(paths: list[Path]) -> tuple[int, int, int]:
    total_bytes = 0
    file_count = 0
    dir_count = 0
    for path in paths:
        item_bytes, item_files, item_dirs = _measure_path_usage(path)
        total_bytes += item_bytes
        file_count += item_files
        dir_count += item_dirs
    return (total_bytes, file_count, dir_count)


def _build_storage_metric(*, key: str, label: str, paths: list[Path]) -> dict[str, Any]:
    total_bytes, file_count, dir_count = _measure_combined_usage(paths)
    relative_paths = [
        _relative_to_project(path) or str(path)
        for path in paths
        if path.exists()
    ]
    cleanup_tier = {
        "upload-session-cache": SAFE_CLEANUP_TIER,
        "job-upload-copies": SAFE_CLEANUP_TIER,
        "baseline-archive": CAUTIOUS_CLEANUP_TIER,
        "review-reports": CAUTIOUS_CLEANUP_TIER,
        "staging-outputs": CAUTIOUS_CLEANUP_TIER,
        "refresh-backups": CAUTIOUS_CLEANUP_TIER,
    }.get(key, PROTECTED_CLEANUP_TIER)
    return {
        "key": key,
        "label": label,
        "bytes": total_bytes,
        "fileCount": file_count,
        "dirCount": dir_count,
        "paths": relative_paths,
        "cleanupTier": cleanup_tier,
    }


def _normalize_cleanup_tier(value: Any) -> str:
    tier = str(value or SAFE_CLEANUP_TIER).strip().lower() or SAFE_CLEANUP_TIER
    if tier not in ALLOWED_CLEANUP_TIERS:
        supported = ", ".join(sorted(ALLOWED_CLEANUP_TIERS))
        raise HTTPException(
            status_code=400,
            detail=f"不支持的 cleanupTier: {tier}（支持: {supported}）。",
        )
    return tier


def _remove_cleanup_paths(paths: list[Path]) -> tuple[list[str], int]:
    removed_paths: list[str] = []
    freed_bytes = 0
    for path in paths:
        if not path.exists():
            continue
        path_bytes, _, _ = _measure_path_usage(path)
        freed_bytes += path_bytes
        if path.is_dir():
            shutil.rmtree(path)
        else:
            path.unlink()
        removed_paths.append(_relative_to_project(path) or str(path))
    return removed_paths, freed_bytes


def _child_cleanup_paths(root: Path) -> list[Path]:
    if not root.exists():
        return []
    return sorted(root.glob("*"))


def _job_upload_dirs() -> list[Path]:
    if not MONTHLY_UPDATE_JOB_ROOT.exists():
        return []
    return sorted(
        [
            path
            for path in MONTHLY_UPDATE_JOB_ROOT.glob("*/uploads")
            if path.is_dir()
        ]
    )


def _resolve_latest_baseline() -> tuple[Path | None, str | None]:
    active_baseline = _latest_baseline_file(_list_supported_excel_files(BASELINE_ROOT))
    if active_baseline is not None:
        return active_baseline, "active"

    archived_root = HISTORY_ARCHIVE_ROOT / "baseline"
    archived_baseline = _latest_baseline_file(_list_supported_excel_files(archived_root))
    if archived_baseline is not None:
        return archived_baseline, "archive"

    return None, None


def _require_latest_baseline() -> tuple[Path, str]:
    baseline_path, baseline_source = _resolve_latest_baseline()
    if baseline_path is None or baseline_source is None:
        raise HTTPException(
            status_code=409,
            detail=(
                "未找到可用 baseline。请先在 01_RAW_DATA/baseline/ 放入当前激活 baseline；"
                "如误清理，可从 01_RAW_DATA/historyDataArchive/baseline/ 恢复后重试。"
            ),
        )
    return baseline_path, baseline_source


def _latest_patch_dir(paths: list[Path]) -> Path | None:
    if not paths:
        return None
    return max(
        paths,
        key=lambda item: (
            _infer_month_token(item.name) or "0000-00",
            item.stat().st_mtime,
            item.name,
        ),
    )


def _append_log(log_path: Path, message: str) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as handle:
        handle.write(message)
        if not message.endswith("\n"):
            handle.write("\n")


def _tail_text(path: Path, *, max_lines: int = 160, max_chars: int = 20000) -> str | None:
    if not path.exists():
        return None
    read_bytes = max(max_chars * 4, 64 * 1024)
    with path.open("rb") as handle:
        handle.seek(0, os.SEEK_END)
        file_size = handle.tell()
        handle.seek(max(0, file_size - read_bytes))
        payload = handle.read(read_bytes)
    lines = payload.decode("utf-8", errors="replace").splitlines()
    tail = "\n".join(lines[-max_lines:])
    if len(tail) > max_chars:
        tail = tail[-max_chars:]
    return tail


def _read_text_if_exists(project_relative_path: str | None) -> str | None:
    if not project_relative_path:
        return None
    path = PROJECT_ROOT / project_relative_path
    if not path.exists():
        return None
    return path.read_text(encoding="utf-8", errors="replace")


def _extract_flag_value(command_args: list[str], flag: str) -> str | None:
    try:
        index = command_args.index(flag)
    except ValueError:
        return None
    next_index = index + 1
    if next_index >= len(command_args):
        return None
    return command_args[next_index]


def _parse_plan_markdown(plan_path: Path) -> dict[str, Any]:
    text = plan_path.read_text(encoding="utf-8")
    commands = [
        block.strip()
        for block in CODE_BLOCK_PATTERN.findall(text)
        if block.strip()
    ]
    if len(commands) < 2:
        raise RuntimeError("monthly_update_plan.md 缺少 raw compare / refresh 命令")

    metadata: dict[str, str] = {}
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("- 数据月:"):
            metadata["month"] = stripped.split(":", 1)[1].strip()
        elif stripped.startswith("- 批次:"):
            metadata["batchId"] = stripped.split(":", 1)[1].strip()
        elif stripped.startswith("- 对比:"):
            metadata["compareId"] = stripped.split(":", 1)[1].strip()
        elif stripped.startswith("- baseline:"):
            metadata["baselinePath"] = stripped.split(":", 1)[1].strip()
        elif stripped.startswith("- patch:"):
            metadata["patchPath"] = stripped.split(":", 1)[1].strip()
        elif stripped.startswith("- supplement parquet:"):
            metadata["supplementParquetPath"] = stripped.split(":", 1)[1].strip()

    compare_command = commands[0]
    refresh_command = commands[1]
    compare_args = shlex.split(compare_command)
    refresh_args = shlex.split(refresh_command)

    review_dir = _extract_flag_value(compare_args, "--output-dir")
    refresh_report = _extract_flag_value(refresh_args, "--report")
    staging_output = _extract_flag_value(refresh_args, "--output")
    manifest_path = _extract_flag_value(refresh_args, "--manifest")
    partition_output = _extract_flag_value(refresh_args, "--partition-output")
    fingerprint_path = _extract_flag_value(refresh_args, "--fingerprint")
    summaries_output = _extract_flag_value(refresh_args, "--summaries-output")

    return {
        "month": metadata.get("month"),
        "batchId": metadata.get("batchId"),
        "compareId": metadata.get("compareId", ""),
        "baselinePath": metadata.get("baselinePath"),
        "patchPath": metadata.get("patchPath"),
        "supplementParquetPath": (
            metadata.get("supplementParquetPath")
            or _extract_flag_value(
                refresh_args,
                "--supplement-missing-countries-from-parquet",
            )
        ),
        "compareCommand": compare_command,
        "refreshCommand": refresh_command,
        "reviewDir": review_dir,
        "rawCompareReportPath": (
            f"{review_dir}/raw_compare_report.json" if review_dir else None
        ),
        "stagingOutputPath": staging_output,
        "manifestPath": manifest_path,
        "partitionOutputPath": partition_output,
        "refreshReportPath": refresh_report,
        "fingerprintPath": fingerprint_path,
        "summariesOutputPath": summaries_output,
    }


def _inject_refresh_supplement_arg(
    refresh_command: str,
) -> tuple[str, str | None]:
    refresh_args = shlex.split(refresh_command)
    if (
        "--summaries-output" not in refresh_args
        and "--skip-precompute" not in refresh_args
    ):
        output_value = _extract_flag_value(refresh_args, "--output")
        if output_value:
            summaries_output = str(Path(output_value).parent / "summaries")
            refresh_args.extend(["--summaries-output", summaries_output])
        else:
            refresh_args.append("--skip-precompute")

    active_parquet_path = _active_data_paths()["parquet"]
    if not active_parquet_path.exists():
        return shlex.join(refresh_args), None

    supplement_flag = "--supplement-missing-countries-from-parquet"
    supplement_path = _relative_to_project(active_parquet_path) or str(
        active_parquet_path
    )
    if supplement_flag in refresh_args:
        return (
            shlex.join(refresh_args),
            _extract_flag_value(refresh_args, supplement_flag) or supplement_path,
        )

    refresh_args.extend([supplement_flag, supplement_path])
    return shlex.join(refresh_args), supplement_path


def _command_to_args(command: str) -> list[str]:
    args = shlex.split(command)
    if not args:
        raise RuntimeError("命令为空")
    if args[0] == "python":
        args[0] = sys.executable
    return args


def _read_json_if_exists(project_relative_path: str | None) -> dict[str, Any] | None:
    if not project_relative_path:
        return None
    path = PROJECT_ROOT / project_relative_path
    if not path.exists():
        return None
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload if isinstance(payload, dict) else None


def _summarize_raw_compare_report(report: dict[str, Any]) -> dict[str, Any]:
    findings = report.get("reviewFindings")
    finding_list = findings if isinstance(findings, list) else []
    freshness_entries = report.get("countryFreshnessSummary")
    freshness_list = freshness_entries if isinstance(freshness_entries, list) else []
    scope_summary = report.get("countryScopeSummary")
    scope = scope_summary if isinstance(scope_summary, dict) else {}

    blocker_count = 0
    review_count = 0
    info_count = 0
    for item in finding_list:
        if not isinstance(item, dict):
            continue
        severity = str(item.get("severity", "")).lower()
        if severity == "blocker":
            blocker_count += 1
        elif severity == "review":
            review_count += 1
        elif severity == "info":
            info_count += 1

    advanced_country_count = 0
    regressed_country_count = 0
    new_country_count = 0
    missing_country_count = 0
    for item in freshness_list:
        if not isinstance(item, dict):
            continue
        status = str(item.get("freshnessStatus", ""))
        if status == "advanced":
            advanced_country_count += 1
        elif status == "regressed":
            regressed_country_count += 1
        elif status == "new_country":
            new_country_count += 1
        elif status == "missing_in_candidate":
            missing_country_count += 1

    added_countries = scope.get("addedCountries")
    removed_countries = scope.get("removedCountries")

    return {
        "compareId": str(report.get("compareId", "")),
        "decisionSuggestion": str(report.get("decisionSuggestion", "")),
        "compareKeyMode": str(report.get("compareKeyMode", "")),
        "compareKeyColumns": (
            [str(item) for item in report.get("compareKeyColumns", [])]
            if isinstance(report.get("compareKeyColumns"), list)
            else []
        ),
        "blockerCount": blocker_count,
        "reviewCount": review_count,
        "infoCount": info_count,
        "advancedCountryCount": advanced_country_count,
        "regressedCountryCount": regressed_country_count,
        "newCountryCount": new_country_count,
        "missingCountryCount": missing_country_count,
        "addedCountryCount": len(added_countries) if isinstance(added_countries, list) else 0,
        "removedCountryCount": (
            len(removed_countries) if isinstance(removed_countries, list) else 0
        ),
    }


def _summarize_refresh_report(report: dict[str, Any]) -> dict[str, Any]:
    full_manifest = report.get("fullManifest")
    full = full_manifest if isinstance(full_manifest, dict) else {}
    partition_manifest = report.get("partitionManifest")
    partition = partition_manifest if isinstance(partition_manifest, dict) else {}
    incremental_payload = report.get("incremental")
    incremental = incremental_payload if isinstance(incremental_payload, dict) else {}
    regression_payload = incremental.get("regression")
    regression = regression_payload if isinstance(regression_payload, dict) else {}
    merge_payload = regression.get("mergeKeyRegression")
    merge = merge_payload if isinstance(merge_payload, dict) else {}

    return {
        "jobStatus": str(report.get("jobStatus", "")),
        "jobElapsedSeconds": float(report.get("jobElapsedSeconds", 0) or 0),
        "rowCount": int(full.get("rows", 0) or 0),
        "columnCount": int(full.get("columns", 0) or 0),
        "partitionCount": int(partition.get("parquetFileCount", 0) or 0),
        "changedRows": int(regression.get("changedRows", 0) or 0),
        "changedCountryCount": int(regression.get("changedCountryCount", 0) or 0),
        "fingerprintMatched": bool(incremental.get("fingerprintMatched")),
        "fingerprintUpdated": bool(incremental.get("fingerprintUpdated")),
        "conflictGroupCount": int(merge.get("conflictGroupCount", 0) or 0),
        "conflictRowCount": int(merge.get("conflictRowCount", 0) or 0),
    }


def _parse_status_dt(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            return parsed.replace(tzinfo=UTC)
        return parsed
    except (TypeError, ValueError):
        return None


def _status_duration_seconds(started_at: Any, finished_at: Any) -> int:
    started = _parse_status_dt(started_at)
    finished = _parse_status_dt(finished_at)
    if not started or not finished:
        return 0
    return max(0, int((finished - started).total_seconds()))


def _jato_etl_artifact_refs(state: dict[str, Any]) -> list[str]:
    artifacts = state.get("artifacts") if isinstance(state.get("artifacts"), dict) else {}
    refs: list[str] = []
    for key in (
        "logPath",
        "planPath",
        "rawCompareReportPath",
        "stagingOutputPath",
        "manifestPath",
        "partitionOutputPath",
        "refreshReportPath",
        "fingerprintPath",
    ):
        value = artifacts.get(key)
        if isinstance(value, str) and value.strip():
            refs.append(value.strip())
    return refs


def _write_jato_etl_pipeline_status(state: dict[str, Any]) -> None:
    """Publish JATO monthly update runtime status for Hermes Sentinel."""
    try:
        summaries = state.get("summaries") if isinstance(state.get("summaries"), dict) else {}
        refresh = summaries.get("refresh") if isinstance(summaries.get("refresh"), dict) else {}
        job_status = str(state.get("status") or "unknown").strip().lower()
        status = (
            "success"
            if job_status == "success"
            else "failed"
            if job_status == "failed"
            else "degraded"
            if job_status == "cancelled"
            else "unknown"
        )
        message = str(state.get("error") or "").strip()
        if not message:
            message = (
                f"jobId={state.get('jobId', '')} "
                f"month={state.get('month', '')} phase={state.get('phase', '')}"
            ).strip()

        write_pipeline_status({
            "pipelineId": "jato_etl",
            "status": status,
            "lastRunAt": state.get("finishedAt") or state.get("updatedAt") or state.get("startedAt"),
            "startedAt": state.get("startedAt"),
            "finishedAt": state.get("finishedAt"),
            "exitCode": 0 if status == "success" else 1 if status == "failed" else 130 if status == "degraded" else None,
            "durationSeconds": _status_duration_seconds(state.get("startedAt"), state.get("finishedAt")),
            "recordsProcessed": refresh.get("rowCount", 0),
            "failedCount": 1 if status == "failed" else 0,
            "warningCount": 1 if status == "degraded" else 0,
            "artifactRefs": _jato_etl_artifact_refs(state),
            "source": "app.services.jato_monthly_update_service",
            "message": message,
            "jobId": state.get("jobId"),
            "month": state.get("month"),
            "batchId": state.get("batchId"),
            "phase": state.get("phase"),
            "triggeredBy": state.get("triggeredBy"),
        })
    except Exception:
        return


def _serialize_job_state(
    payload: dict[str, Any],
    *,
    include_log_tail: bool,
) -> dict[str, Any]:
    item = {
        "jobId": str(payload.get("jobId", "")),
        "month": (
            None
            if payload.get("month") in {None, ""}
            else str(payload.get("month"))
        ),
        "batchId": (
            None
            if payload.get("batchId") in {None, ""}
            else str(payload.get("batchId"))
        ),
        "status": str(payload.get("status", "")),
        "phase": str(payload.get("phase", "")),
        "jobType": payload.get("jobType"),
        "country": payload.get("country"),
        "countryScope": (
            [str(country) for country in payload.get("countryScope", [])]
            if isinstance(payload.get("countryScope"), list)
            else []
        ),
        "triggeredBy": str(payload.get("triggeredBy", "")),
        "createdAt": str(payload.get("createdAt", "")),
        "updatedAt": str(payload.get("updatedAt", "")),
        "startedAt": payload.get("startedAt"),
        "finishedAt": payload.get("finishedAt"),
        "error": payload.get("error"),
        "ingestionKey": payload.get("ingestionKey"),
        "ingestDigest": (
            payload.get("ingestDigest")
            if isinstance(payload.get("ingestDigest"), dict)
            else None
        ),
        "failureDigest": (
            payload.get("failureDigest")
            if isinstance(payload.get("failureDigest"), dict)
            else None
        ),
        "duplicateOfJobId": payload.get("duplicateOfJobId"),
        "activeBaseFingerprint": payload.get("activeBaseFingerprint"),
        "upload": payload.get("upload") if isinstance(payload.get("upload"), dict) else None,
        "plan": payload.get("plan") if isinstance(payload.get("plan"), dict) else None,
        "artifacts": (
            payload.get("artifacts")
            if isinstance(payload.get("artifacts"), dict)
            else None
        ),
        "summaries": (
            payload.get("summaries")
            if isinstance(payload.get("summaries"), dict)
            else None
        ),
        "logPath": payload.get("logPath"),
        "publication": (
            payload.get("publication")
            if isinstance(payload.get("publication"), dict)
            else None
        ),
        "currentProcess": (
            payload.get("currentProcess")
            if isinstance(payload.get("currentProcess"), dict)
            else None
        ),
        "runtimeCheck": (
            payload.get("runtimeCheck")
            if isinstance(payload.get("runtimeCheck"), dict)
            else None
        ),
        "cancellation": (
            payload.get("cancellation")
            if isinstance(payload.get("cancellation"), dict)
            else None
        ),
        "reviewApproval": (
            payload.get("reviewApproval")
            if isinstance(payload.get("reviewApproval"), dict)
            else None
        ),
        "pendingOperation": (
            payload.get("pendingOperation")
            if isinstance(payload.get("pendingOperation"), dict)
            else None
        ),
        "smartMergeRecovery": _smart_merge_recovery_view(payload),
    }
    if "recoveryOfJobId" in payload or "recoveryKey" in payload:
        item["recoveryOfJobId"] = payload.get("recoveryOfJobId")
        item["recoveryKey"] = payload.get("recoveryKey")
        item["recoverySource"] = (
            payload.get("recoverySource")
            if isinstance(payload.get("recoverySource"), dict)
            else None
        )
    if include_log_tail:
        log_path = _job_log_path(item["jobId"])
        item["logTail"] = _tail_text(log_path)
    return item


def _sanitize_review_finding(item: Any) -> dict[str, Any] | None:
    if not isinstance(item, dict):
        return None
    metrics = item.get("metrics")
    return {
        "severity": str(item.get("severity", "")),
        "scope": str(item.get("scope", "")),
        "target": str(item.get("target", "")),
        "ruleId": str(item.get("ruleId", "")),
        "message": str(item.get("message", "")),
        "metrics": metrics if isinstance(metrics, dict) else {},
        "suggestedAction": str(item.get("suggestedAction", "")),
        "sourceFeedback": (
            None
            if item.get("sourceFeedback") in {None, ""}
            else str(item.get("sourceFeedback"))
        ),
    }


def _sanitize_conflict_sample(item: Any) -> dict[str, Any] | None:
    if not isinstance(item, dict):
        return None
    business_key = item.get("businessKey")
    changed_fields = item.get("changedFields")
    return {
        "country": str(item.get("country", "")),
        "businessKey": business_key if isinstance(business_key, dict) else {},
        "oldValueDigest": (
            None
            if item.get("oldValueDigest") in {None, ""}
            else str(item.get("oldValueDigest"))
        ),
        "newValueDigest": (
            None
            if item.get("newValueDigest") in {None, ""}
            else str(item.get("newValueDigest"))
        ),
        "changedFields": (
            [str(field) for field in changed_fields]
            if isinstance(changed_fields, list)
            else []
        ),
    }


def _sanitize_overlap_change_summary(item: Any) -> dict[str, Any] | None:
    if not isinstance(item, dict):
        return None
    return {
        "country": str(item.get("country", "")),
        "compareMonths": (
            [str(month) for month in item.get("compareMonths", [])]
            if isinstance(item.get("compareMonths"), list)
            else []
        ),
        "compareKeyColumns": (
            [str(column) for column in item.get("compareKeyColumns", [])]
            if isinstance(item.get("compareKeyColumns"), list)
            else []
        ),
        "addedRecordCount": int(item.get("addedRecordCount", 0) or 0),
        "removedRecordCount": int(item.get("removedRecordCount", 0) or 0),
        "changedRecordCount": int(item.get("changedRecordCount", 0) or 0),
        "unchangedRecordCount": int(item.get("unchangedRecordCount", 0) or 0),
        "changeRate": float(item.get("changeRate", 0) or 0),
        "sampleAddedKeys": (
            [entry for entry in item.get("sampleAddedKeys", []) if isinstance(entry, dict)]
            if isinstance(item.get("sampleAddedKeys"), list)
            else []
        ),
        "sampleRemovedKeys": (
            [entry for entry in item.get("sampleRemovedKeys", []) if isinstance(entry, dict)]
            if isinstance(item.get("sampleRemovedKeys"), list)
            else []
        ),
        "sampleChangedKeys": (
            [entry for entry in item.get("sampleChangedKeys", []) if isinstance(entry, dict)]
            if isinstance(item.get("sampleChangedKeys"), list)
            else []
        ),
    }


def _sanitize_country_freshness_summary(item: Any) -> dict[str, Any] | None:
    if not isinstance(item, dict):
        return None
    return {
        "country": str(item.get("country", "")),
        "oldLatestMonth": (
            None
            if item.get("oldLatestMonth") in {None, ""}
            else str(item.get("oldLatestMonth"))
        ),
        "newLatestMonth": (
            None
            if item.get("newLatestMonth") in {None, ""}
            else str(item.get("newLatestMonth"))
        ),
        "freshnessStatus": str(item.get("freshnessStatus", "")),
        "rowDelta": int(item.get("rowDelta", 0) or 0),
    }


def _sanitize_country_coverage_summary(item: Any) -> dict[str, Any] | None:
    if not isinstance(item, dict):
        return None
    return {
        "country": str(item.get("country", "")),
        "oldMonths": (
            [str(month) for month in item.get("oldMonths", [])]
            if isinstance(item.get("oldMonths"), list)
            else []
        ),
        "newMonths": (
            [str(month) for month in item.get("newMonths", [])]
            if isinstance(item.get("newMonths"), list)
            else []
        ),
        "addedMonths": (
            [str(month) for month in item.get("addedMonths", [])]
            if isinstance(item.get("addedMonths"), list)
            else []
        ),
        "removedMonths": (
            [str(month) for month in item.get("removedMonths", [])]
            if isinstance(item.get("removedMonths"), list)
            else []
        ),
        "overlappingMonths": (
            [str(month) for month in item.get("overlappingMonths", [])]
            if isinstance(item.get("overlappingMonths"), list)
            else []
        ),
        "coverageStatus": str(item.get("coverageStatus", "")),
    }


def _require_no_running_monthly_update_jobs(*, excluding_job_id: str | None = None) -> None:
    running_jobs = [
        str(payload.get("jobId", ""))
        for payload in _list_job_state_payloads()
        if (
            str(payload.get("status", "")) in {"queued", "running"}
            or (
                isinstance(payload.get("pendingOperation"), dict)
                and str(payload["pendingOperation"].get("status") or "")
                in {"queued", "running"}
            )
        )
        and str(payload.get("jobId", "")) != str(excluding_job_id or "")
    ]
    if running_jobs:
        raise HTTPException(
            status_code=409,
            detail="存在运行中的月更任务，请等待完成后再执行 review / publish。",
        )
    baseline_promotion = _load_baseline_promotion_state()
    if (
        isinstance(baseline_promotion, dict)
        and str(baseline_promotion.get("status") or "")
        in {"queued", "running"}
    ):
        raise HTTPException(
            status_code=409,
            detail="正在保存 active baseline，请等待隔离 worker 完成后再执行 publish / rollback。",
        )


@contextmanager
def _monthly_update_resource_start_locks(
    *,
    action: str,
) -> Any:
    """Acquire the shared heavy-resource locks in the canonical order."""
    with _exclusive_file_lock(
        _maintenance_coordination_lock_path(),
        blocking=False,
    ) as coordinated:
        if not coordinated:
            raise HTTPException(
                status_code=409,
                detail=(
                    "JATO 清理或 baseline 操作正在准备中，"
                    f"请稍后再{action}。"
                ),
            )
        with _exclusive_file_lock(
            _upload_initiate_lock_path()
        ) as global_upload:
            if not global_upload:
                raise HTTPException(
                    status_code=503,
                    detail="JATO 全局资源锁暂不可用，请稍后重试。",
                )
            yield


@contextmanager
def _monthly_update_worker_start_window(
    *,
    action: str,
    excluding_ready_upload_id: str | None = None,
    excluding_job_id: str | None = None,
) -> Any:
    """Serialize every heavy-worker check and launch against upload/maintenance."""
    with _monthly_update_resource_start_locks(action=action):
        _require_no_active_upload_sessions(
            action=action,
            excluding_ready_upload_id=excluding_ready_upload_id,
        )
        _require_no_running_monthly_update_jobs(
            excluding_job_id=excluding_job_id,
        )
        yield


def _historical_reclassification_report_fingerprint(
    countries: list[dict[str, Any]],
) -> str:
    return hashlib.sha256(
        json.dumps(
            countries,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()


def _historical_reclassification_allowed_decisions(
    country_report: dict[str, Any],
) -> tuple[str, ...]:
    """Return the only safe choices for one immutable Review report row."""
    monthly_totals_stable = country_report.get("monthlyTotalsStable")
    if monthly_totals_stable is False:
        return ("keep_active",)
    if monthly_totals_stable is True:
        return HISTORICAL_RECLASSIFICATION_DECISION_ORDER
    return ()


def _normalize_historical_reclassification_countries_for_resolution(
    countries: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Upgrade legacy Review rows before binding a decision fingerprint.

    Older cached Reviews marked countries with changed historical country/month
    totals as non-decisionable.  They are safely decisionable only because the
    existing keep_active merge takes all published months from active.  This
    normalization is shared by the endpoint and Smart Merge preflight so the
    exact same country payload is fingerprinted in both places.
    """
    normalized: list[dict[str, Any]] = []
    for raw_country in countries:
        country_report = dict(raw_country)
        if country_report.get("monthlyTotalsStable") is False:
            country_report["decisionRequired"] = True
        if bool(country_report.get("decisionRequired")):
            country_report["allowedDecisions"] = list(
                _historical_reclassification_allowed_decisions(
                    country_report
                )
            )
        normalized.append(country_report)
    return normalized


def _normalized_historical_reclassification_report_for_resolution(
    report: dict[str, Any],
) -> dict[str, Any]:
    raw_countries = report.get("countries")
    if not isinstance(raw_countries, list) or any(
        not isinstance(item, dict) for item in raw_countries
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": (
                    "historical_reclassification_resolution_invalid"
                ),
                "message": "历史重分类报告缺少有效的逐国结构。",
            },
        )
    countries = (
        _normalize_historical_reclassification_countries_for_resolution(
            [dict(item) for item in raw_countries]
        )
    )
    report_fingerprint = (
        _historical_reclassification_report_fingerprint(countries)
    )
    return {
        "status": (
            "decision_required"
            if any(bool(item.get("decisionRequired")) for item in countries)
            else "not_required"
        ),
        "countries": countries,
        "reportFingerprint": report_fingerprint,
        "truncation": (
            dict(report.get("truncation"))
            if isinstance(report.get("truncation"), dict)
            else {}
        ),
    }


def _validated_normalized_historical_reclassification_report(
    report: dict[str, Any],
) -> dict[str, Any]:
    """Validate a cached Review snapshot, then upgrade its decision contract."""
    raw_countries = report.get("countries")
    if not isinstance(raw_countries, list) or any(
        not isinstance(item, dict) for item in raw_countries
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": (
                    "historical_reclassification_resolution_invalid"
                ),
                "message": "Review 历史重分类报告结构无效。",
            },
        )
    countries = [dict(item) for item in raw_countries]
    declared_fingerprint = str(
        report.get("reportFingerprint") or ""
    ).strip()
    computed_fingerprint = (
        _historical_reclassification_report_fingerprint(countries)
    )
    if (
        not re.fullmatch(r"[0-9a-f]{64}", declared_fingerprint)
        or declared_fingerprint != computed_fingerprint
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": (
                    "historical_reclassification_resolution_invalid"
                ),
                "message": (
                    "Review 历史重分类报告与其指纹不一致，"
                    "拒绝绑定决策。"
                ),
            },
        )
    return _normalized_historical_reclassification_report_for_resolution(
        report
    )


def _historical_sales_changed_blocker_country_key(
    finding: dict[str, Any],
) -> str | None:
    """Identify only the legacy SC011 blocker that keep_active can repair."""
    metrics = (
        finding.get("metrics")
        if isinstance(finding.get("metrics"), dict)
        else {}
    )
    blocker_type = str(
        finding.get("blockerType")
        or metrics.get("blockerType")
        or metrics.get("reason")
        or ""
    ).strip()
    mismatch_count = metrics.get("countryMismatchCount")
    target_key = str(finding.get("target") or "").strip().casefold()
    if not (
        finding.get("severity") == "blocker"
        and finding.get("scope") == "country"
        and finding.get("ruleId") == "SC011"
        and blocker_type == "historical_sales_changed"
        and isinstance(mismatch_count, (int, float))
        and not isinstance(mismatch_count, bool)
        and mismatch_count > 0
        and target_key
    ):
        return None
    return target_key


def _historical_reclassification_resolution(
    payload: dict[str, Any],
) -> dict[str, Any] | None:
    value = payload.get("historicalReclassificationResolution")
    return value if isinstance(value, dict) else None


def _historical_reclassification_resolution_fingerprint(
    resolution: dict[str, Any],
) -> str:
    """Seal the exact active/candidate/report/decision recovery contract."""
    raw_decisions = resolution.get("decisions")
    canonical_decisions = sorted(
        (
            {
                "country": str(item.get("country") or "").strip(),
                "decision": str(item.get("decision") or "")
                .strip()
                .lower(),
            }
            for item in (
                raw_decisions if isinstance(raw_decisions, list) else []
            )
            if isinstance(item, dict)
        ),
        key=lambda item: (
            item["country"].casefold(),
            item["country"],
            item["decision"],
        ),
    )
    sealed = {
        "activeBaseFingerprint": str(
            resolution.get("activeBaseFingerprint") or ""
        ).strip().lower(),
        "sourceCandidateFingerprint": str(
            resolution.get("sourceCandidateFingerprint") or ""
        ).strip().lower(),
        "reportFingerprint": str(
            resolution.get("reportFingerprint") or ""
        ).strip().lower(),
        "decisions": canonical_decisions,
    }
    return hashlib.sha256(
        json.dumps(
            sealed,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()


def _smart_merge_recovery_view(
    payload: dict[str, Any],
) -> dict[str, Any] | None:
    """Expose only immutable seals needed by the admin resume control."""
    resolution = _historical_reclassification_resolution(payload)
    if not isinstance(resolution, dict):
        return None
    artifacts = payload.get("artifacts")
    candidate_scope = (
        str(artifacts.get("candidateScope") or "")
        if isinstance(artifacts, dict)
        else ""
    )
    source_candidate_fingerprint = _valid_sha256(
        resolution.get("sourceCandidateFingerprint")
    )
    active_fingerprint = _valid_sha256(
        resolution.get("activeBaseFingerprint")
    )
    report_fingerprint = _valid_sha256(
        resolution.get("reportFingerprint")
    )
    resolution_fingerprint = (
        _historical_reclassification_resolution_fingerprint(resolution)
    )
    pending = _pending_operation(payload)
    pending_active = bool(
        isinstance(pending, dict)
        and str(pending.get("status") or "") in {"queued", "running"}
    )
    publication = payload.get("publication")
    published = bool(
        isinstance(publication, dict)
        and publication.get("publishedAt")
        and not publication.get("rolledBackAt")
    )
    reason: str | None = None
    if pending_active:
        reason = "operation_in_progress"
    elif (
        str(payload.get("status") or "") != "failed"
        or str(payload.get("phase") or "") != "smart_merge_failed"
        or str(payload.get("operation") or "") != "smart_merge"
    ):
        reason = "job_not_smart_merge_failed"
    elif published:
        reason = "candidate_already_published"
    elif isinstance(payload.get("reviewApproval"), dict):
        reason = "review_already_approved"
    elif candidate_scope not in {
        *PARTITION_SCOPED_CANDIDATE_SCOPES,
        "full_smart_merge",
    }:
        reason = "candidate_scope_invalid"
    elif not all(
        (
            source_candidate_fingerprint,
            active_fingerprint,
            report_fingerprint,
        )
    ):
        reason = "recovery_seal_missing"
    elif (
        _valid_sha256(payload.get("activeBaseFingerprint"))
        != active_fingerprint
    ):
        reason = "active_lineage_mismatch"
    elif (
        candidate_scope == "full_smart_merge"
        and str(resolution.get("status") or "") != "resolved"
    ):
        reason = "committed_bundle_not_resolved"
    return {
        "canResume": reason is None,
        "reason": reason,
        "candidateScope": candidate_scope or None,
        "sourceCandidateFingerprint": source_candidate_fingerprint,
        "activeBaseFingerprint": active_fingerprint,
        "reportFingerprint": report_fingerprint,
        "resolutionFingerprint": resolution_fingerprint,
    }


def _historical_reclassification_decision_map(
    resolution: dict[str, Any] | None,
) -> dict[str, str]:
    if not isinstance(resolution, dict):
        return {}
    raw_decisions = resolution.get("decisions")
    if not isinstance(raw_decisions, list):
        return {}
    decisions: dict[str, str] = {}
    for item in raw_decisions:
        if not isinstance(item, dict):
            continue
        country = str(item.get("country") or "").strip()
        decision = str(item.get("decision") or "").strip().lower()
        if (
            country
            and decision in HISTORICAL_RECLASSIFICATION_DECISIONS
        ):
            decisions[country.casefold()] = decision
    return decisions


def _validated_historical_reclassification_resolution(
    resolution: dict[str, Any],
) -> dict[str, str]:
    report = resolution.get("report")
    raw_countries = (
        report.get("countries")
        if isinstance(report, dict)
        else None
    )
    if (
        not isinstance(report, dict)
        or not isinstance(raw_countries, list)
        or any(not isinstance(item, dict) for item in raw_countries)
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": (
                    "historical_reclassification_resolution_invalid"
                ),
                "message": "历史重分类 resolution 缺少原始结构化报告。",
            },
        )
    countries = [dict(item) for item in raw_countries]
    computed_fingerprint = (
        _historical_reclassification_report_fingerprint(countries)
    )
    declared_fingerprint = str(
        resolution.get("reportFingerprint") or ""
    )
    nested_fingerprint = str(report.get("reportFingerprint") or "")
    if (
        not re.fullmatch(r"[0-9a-f]{64}", declared_fingerprint)
        or nested_fingerprint != declared_fingerprint
        or computed_fingerprint != declared_fingerprint
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": (
                    "historical_reclassification_resolution_invalid"
                ),
                "message": (
                    "历史重分类报告内容与声明指纹不一致，"
                    "拒绝应用可能被篡改或过期的决策。"
                ),
                "declaredReportFingerprint": (
                    declared_fingerprint or None
                ),
                "computedReportFingerprint": computed_fingerprint,
            },
        )

    required_by_key: dict[str, str] = {}
    required_reports_by_key: dict[str, dict[str, Any]] = {}
    for item in countries:
        if not bool(item.get("decisionRequired")):
            continue
        country = str(item.get("country") or "").strip()
        country_key = country.casefold()
        if not country or country_key in required_by_key:
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": (
                        "historical_reclassification_resolution_invalid"
                    ),
                    "message": (
                        "原始历史重分类报告包含空国家或重复逻辑国家。"
                    ),
                },
            )
        required_by_key[country_key] = country
        required_reports_by_key[country_key] = item

    raw_decisions = resolution.get("decisions")
    if not isinstance(raw_decisions, list):
        raw_decisions = []
    decisions: dict[str, str] = {}
    for item in raw_decisions:
        if not isinstance(item, dict):
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": (
                        "historical_reclassification_resolution_invalid"
                    ),
                    "message": "历史重分类 decisions 格式无效。",
                },
            )
        country = str(item.get("country") or "").strip()
        country_key = country.casefold()
        decision = str(item.get("decision") or "").strip().lower()
        if (
            not country
            or country_key in decisions
            or decision not in HISTORICAL_RECLASSIFICATION_DECISIONS
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": (
                        "historical_reclassification_resolution_invalid"
                    ),
                    "message": (
                        "历史重分类 decisions 包含空国家、重复国家"
                        "或无效 decision。"
                    ),
                },
            )
        country_report = required_reports_by_key.get(country_key)
        if (
            country_report is not None
            and decision
            not in _historical_reclassification_allowed_decisions(
                country_report
            )
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": (
                        "historical_sales_changed_requires_keep_active"
                    ),
                    "message": (
                        f"{required_by_key[country_key]} 的历史国家/月总量"
                        "发生变化，只能选择 keep_active；"
                        "拒绝 use_latest 改写已发布历史。"
                    ),
                    "country": required_by_key[country_key],
                    "allowedDecisions": ["keep_active"],
                },
            )
        decisions[country_key] = decision
    missing_keys = set(required_by_key) - set(decisions)
    extra_keys = set(decisions) - set(required_by_key)
    if missing_keys or extra_keys:
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": (
                    "historical_reclassification_resolution_invalid"
                ),
                "message": (
                    "历史重分类 decisions 未精确覆盖原报告所有受影响国家。"
                ),
                "missingCountries": [
                    required_by_key[key]
                    for key in sorted(missing_keys)
                ],
                "extraCountries": sorted(extra_keys),
            },
        )
    return decisions


def _build_historical_reclassification_report(
    *,
    payload: dict[str, Any],
    current_countries: list[dict[str, Any]],
) -> dict[str, Any]:
    resolution = _historical_reclassification_resolution(payload)
    resolved_report = (
        resolution.get("report")
        if isinstance(resolution, dict)
        and isinstance(resolution.get("report"), dict)
        else None
    )
    if isinstance(resolved_report, dict):
        decisions = _validated_historical_reclassification_resolution(
            resolution
        )
        stored_countries = resolved_report.get("countries")
        countries = [
            dict(item)
            for item in (
                stored_countries
                if isinstance(stored_countries, list)
                else []
            )
            if isinstance(item, dict)
        ]
    else:
        decisions = {}
        countries = [
            dict(item)
            for item in current_countries
            if isinstance(item, dict)
        ]
    resolution_is_resolved = bool(
        isinstance(resolution, dict)
        and resolution.get("status") == "resolved"
    )
    if resolution_is_resolved and decisions:
        for item in countries:
            country_key = str(item.get("country") or "").strip().casefold()
            if country_key in decisions:
                item["decision"] = decisions[country_key]
    if resolution_is_resolved and countries:
        status = "resolved"
    elif any(bool(item.get("decisionRequired")) for item in countries):
        status = "decision_required"
    else:
        status = "not_required"
    report_fingerprint = (
        str(resolution.get("reportFingerprint") or "")
        if isinstance(resolution, dict)
        else _historical_reclassification_report_fingerprint(countries)
    )
    return {
        "status": status,
        "countries": countries,
        "reportFingerprint": report_fingerprint,
        "truncation": {
            "truncated": any(
                bool(
                    (
                        item.get("truncation")
                        if isinstance(item.get("truncation"), dict)
                        else {}
                    ).get("truncated")
                )
                for item in countries
            ),
            "countryCount": len(countries),
        },
    }


def _resolved_historical_reclassification_decision(
    payload: dict[str, Any],
    country: str,
) -> str | None:
    return _historical_reclassification_decision_map(
        _historical_reclassification_resolution(payload)
    ).get(country.strip().casefold())


def _historical_keep_active_resolution_validation(
    *,
    country: str,
    decision: str | None,
    historical_stability: dict[str, Any],
) -> dict[str, Any] | None:
    """Bind one keep-active decision to this full candidate's history check."""
    if decision != "keep_active":
        return None
    current_status = str(
        historical_stability.get("status") or "unavailable"
    )
    return {
        "country": country,
        "decision": "keep_active",
        "status": "pass" if current_status == "pass" else "fail",
        "currentStabilityStatus": current_status,
        "reason": historical_stability.get("reason"),
    }


def _exact_partial_keep_active_resolution_validation(
    *,
    payload: dict[str, Any],
    raw_validation: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Return one validation per keep-active decision or fail closed."""
    resolution = _historical_reclassification_resolution(payload)
    if not isinstance(resolution, dict):
        return []
    validated_decisions = (
        _validated_historical_reclassification_resolution(resolution)
    )
    raw_decisions = resolution.get("decisions")
    expected: list[tuple[str, str]] = []
    for item in (
        raw_decisions if isinstance(raw_decisions, list) else []
    ):
        if not isinstance(item, dict):
            continue
        country = str(item.get("country") or "").strip()
        country_key = country.casefold()
        if (
            country_key
            and validated_decisions.get(country_key) == "keep_active"
        ):
            expected.append((country_key, country))

    validation_by_key: dict[str, dict[str, Any]] = {}
    duplicate_countries: list[str] = []
    invalid_entries = False
    for item in raw_validation:
        country = str(item.get("country") or "").strip()
        country_key = country.casefold()
        if not country_key or item.get("decision") != "keep_active":
            invalid_entries = True
            continue
        if country_key in validation_by_key:
            duplicate_countries.append(country)
            continue
        validation_by_key[country_key] = item

    expected_by_key = dict(expected)
    missing_keys = set(expected_by_key) - set(validation_by_key)
    extra_keys = set(validation_by_key) - set(expected_by_key)
    if (
        invalid_entries
        or duplicate_countries
        or missing_keys
        or extra_keys
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": (
                    "historical_keep_active_validation_failed"
                ),
                "message": (
                    "完整 Candidate 的 keep_active 逐国复核证明不完整，"
                    "拒绝生成可批准的 Review。"
                ),
                "missingCountries": [
                    expected_by_key[key]
                    for key in sorted(missing_keys)
                ],
                "duplicateCountries": sorted(duplicate_countries),
                "extraCountries": sorted(extra_keys),
            },
        )
    return [validation_by_key[key] for key, _country in expected]


def _build_single_country_review(
    payload: dict[str, Any],
    *,
    candidate_fingerprint: str | None = None,
) -> dict[str, Any]:
    country = str(payload.get("country") or "").strip()
    artifacts = payload.get("artifacts")
    if not country or not isinstance(artifacts, dict):
        raise HTTPException(status_code=409, detail="目标国家任务缺少 Review 所需信息。")
    candidate_path = _project_path(str(artifacts.get("stagingOutputPath") or "").strip())
    active_paths = _active_data_paths()
    if candidate_path is None or not candidate_path.exists():
        raise HTTPException(status_code=409, detail="单国任务尚未生成可 Review 的 candidate。")
    candidate_frame = _load_parquet_country_subset(candidate_path, country, path_label="candidate")
    if candidate_frame.empty:
        raise HTTPException(status_code=409, detail="单国 candidate 不包含目标国家。")
    if active_paths["partition"].exists():
        active_frame = _load_active_country_partition_subset(active_paths["partition"], country)
    elif active_paths["parquet"].exists():
        active_frame = _load_parquet_country_subset(active_paths["parquet"], country, path_label="active")
    else:
        raise HTTPException(status_code=409, detail="缺少 active 数据，不能生成单国 Review。")
    if active_frame.empty:
        raise HTTPException(status_code=409, detail="active 不包含目标国家，不能生成单国 Review。")

    month_columns = _detect_month_columns(list(candidate_frame.columns))
    schema_contract = _single_country_schema_contract(
        active_frame=active_frame,
        candidate_frame=candidate_frame,
    )
    key_columns = _single_country_configuration_key_columns(candidate_frame)
    duplicate_count = int(candidate_frame.duplicated(subset=key_columns, keep=False).sum()) if key_columns else 0
    negative_sales_count = int(sum(
        pd.to_numeric(candidate_frame[column], errors="coerce").lt(0).sum()
        for column in month_columns
    ))
    active_latest = _latest_month_from_frame(active_frame)
    candidate_latest = _latest_month_from_frame(candidate_frame)
    upload_payload = payload.get("upload")
    source_upload_sha256 = None
    if isinstance(upload_payload, dict):
        stored_upload_path = _project_path(
            str(upload_payload.get("storedPath") or "").strip()
        )
        if stored_upload_path is not None and stored_upload_path.is_file():
            source_upload_sha256 = _sha256_hex_for_path(stored_upload_path)
    historical_stability = _single_country_historical_sales_stability(
        country=country,
        active_frame=active_frame,
        candidate_frame=candidate_frame,
        active_latest_month=active_latest,
        source_upload_sha256=source_upload_sha256,
    )
    current_reclassification = historical_stability.get(
        "historicalReclassification"
    )
    historical_reclassification_report = (
        _build_historical_reclassification_report(
            payload=payload,
            current_countries=(
                [current_reclassification]
                if isinstance(current_reclassification, dict)
                else []
            ),
        )
    )
    resolved_reclassification_decision = (
        _resolved_historical_reclassification_decision(
            payload,
            country,
        )
    )
    candidate_scope = str(artifacts.get("candidateScope") or "")
    if (
        historical_reclassification_report.get("status") == "resolved"
        and candidate_scope == "full_smart_merge"
    ):
        resolution_validation = (
            _historical_keep_active_resolution_validation(
                country=country,
                decision=resolved_reclassification_decision,
                historical_stability=historical_stability,
            )
        )
        if resolution_validation is not None:
            historical_reclassification_report[
                "resolutionValidation"
            ] = [resolution_validation]
    reclassification_resolution = (
        _historical_reclassification_resolution(payload)
    )
    active_sales = _collect_country_monthly_sales(active_frame, countries=[country], path_label="active").get(country, {})
    candidate_sales = _collect_country_monthly_sales(candidate_frame, countries=[country], path_label="candidate").get(country, {})
    common_months = sorted(set(active_sales) & set(candidate_sales), key=_time_sort_key)
    doubled_months = [
        month for month in common_months
        if _is_near_sales_doubling(
            reference_sales=active_sales.get(month),
            candidate_sales=candidate_sales.get(month),
        )[0]
    ]
    partition_check = artifacts.get("untouchedPartitionCheck")
    if not isinstance(partition_check, dict):
        partition_check = {"status": "unavailable", "changedPartitions": []}

    findings: list[dict[str, Any]] = []

    def add_finding(severity: str, rule_id: str, message: str, metrics: dict[str, Any]) -> None:
        finding: dict[str, Any] = {
            "severity": severity,
            "scope": "country",
            "target": country,
            "ruleId": rule_id,
            "message": message,
            "metrics": metrics,
            "suggestedAction": "reject_input_batch" if severity == "blocker" else "manual_review_required",
        }
        source_feedback = _single_country_source_feedback(
            rule_id=rule_id,
            country=country,
            metrics=metrics,
        )
        if source_feedback:
            finding["sourceFeedback"] = source_feedback
        findings.append(finding)

    if not month_columns:
        add_finding("blocker", "SC001", "candidate 缺少月份列。", {})
    if schema_contract["missingMaterial"]:
        add_finding("blocker", "SC009", "目标国家 candidate 缺少 active 业务列：" + "、".join(schema_contract["missingMaterial"]) + "。", {
            "missingMaterialColumns": schema_contract["missingMaterial"],
            "missingDerivedYtdColumns": schema_contract["missingDerivedYtd"],
            "missingNullOnlyColumns": schema_contract["missingNullOnly"],
        })
    if schema_contract["missingDeprecatedOptional"]:
        add_finding(
            "review",
            "SC014",
            "目标国家 candidate 缺少已停用的可选静态字段："
            + "、".join(schema_contract["missingDeprecatedOptional"])
            + "。Smart Merge 仅在配置键匹配且 active 旧值一致时沿用旧值。",
            {
                "missingDeprecatedOptionalColumns": schema_contract[
                    "missingDeprecatedOptional"
                ],
                "activePopulatedRowsByColumn": {
                    column: int(
                        (~_series_missing_value_mask(active_frame[column])).sum()
                    )
                    for column in schema_contract["missingDeprecatedOptional"]
                },
                "carryForwardPolicy": "consistent_active_key_values_only",
                "unmatchedPolicy": "leave_null",
            },
        )
    if schema_contract["missingDerivedYtd"]:
        add_finding("review", "SC013", "目标国家 candidate 缺少旧月份 YTD 派生列：" + "、".join(schema_contract["missingDerivedYtd"]) + "。", {
            "missingDerivedYtdColumns": schema_contract["missingDerivedYtd"],
        })
    if schema_contract["extra"]:
        add_finding("review", "SC010", "目标国家 candidate 包含 active 中没有的新业务列：" + "、".join(schema_contract["extra"]) + "。", {
            "extraColumns": schema_contract["extra"],
        })
    if candidate_latest is None:
        add_finding("blocker", "SC002", "candidate 没有有效销量月份。", {})
    if active_latest and candidate_latest and _time_sort_key(candidate_latest) < _time_sort_key(active_latest):
        add_finding("blocker", "SC003", "目标国家最新月份发生回退。", {"active": active_latest, "candidate": candidate_latest})
    if duplicate_count:
        add_finding("blocker", "SC004", "目标国家 candidate 存在完全相同的配置指纹。", {
            "duplicateRows": duplicate_count,
            "keyColumnCount": len(key_columns),
        })
    if negative_sales_count:
        add_finding("blocker", "SC005", "目标国家 candidate 存在负销量。", {"negativeSalesCells": negative_sales_count})
    if len(doubled_months) >= SALES_DOUBLING_MIN_MONTH_COUNT:
        add_finding("blocker", "SC006", "目标国家 candidate 疑似销量翻倍。", {"months": doubled_months})
    keep_active_validation_failed = bool(
        isinstance(reclassification_resolution, dict)
        and reclassification_resolution.get("status") == "resolved"
        and resolved_reclassification_decision == "keep_active"
        and historical_stability.get("status") != "pass"
    )
    if keep_active_validation_failed:
        add_finding(
            "blocker",
            "SC011",
            (
                "keep_active 合并后的 candidate 仍与 active 历史不一致；"
                "拒绝批准 Publish。"
            ),
            {
                **historical_stability,
                "blockerType": (
                    "historical_keep_active_validation_failed"
                ),
                "requiredStatus": "pass",
            },
        )
    elif historical_stability.get("status") == "fail":
        decision_required = bool(
            isinstance(current_reclassification, dict)
            and current_reclassification.get("decisionRequired")
        )
        monthly_totals_stable = bool(
            isinstance(current_reclassification, dict)
            and current_reclassification.get("monthlyTotalsStable")
        )
        if (
            monthly_totals_stable
            and resolved_reclassification_decision == "use_latest"
        ):
            add_finding(
                "review",
                "SC011",
                "已选择以最新 washed 分类替换该国家历史分析维度。",
                historical_stability,
            )
        elif decision_required and not monthly_totals_stable:
            add_finding(
                "review",
                "SC011",
                (
                    "目标国家历史月总量发生变化；"
                    "只能选择 keep_active，以 active 保留全部已发布历史，"
                    "并仅从 candidate 读取之后的新月份。"
                ),
                historical_stability,
            )
        elif decision_required:
            add_finding(
                "review",
                "SC011",
                "目标国家历史销量总量稳定，但分析维度被重新分类；必须先选择采用最新分类或维持 active 历史分类。",
                historical_stability,
            )
        else:
            add_finding(
                "blocker",
                "SC011",
                "目标国家 candidate 改写了 active 已有历史销量。",
                historical_stability,
            )
    elif historical_stability.get("status") == "confirmed":
        add_finding(
            "review",
            "SC011",
            "目标国家 candidate 包含已精确核验并经业务确认的历史车型重分类。",
            historical_stability,
        )
    row_delta = int(len(candidate_frame) - len(active_frame))
    if row_delta:
        add_finding("review", "SC012", "目标国家 candidate 的配置行数与 active 不同。", {
            "activeRows": int(len(active_frame)),
            "candidateRows": int(len(candidate_frame)),
            "rowDelta": row_delta,
            "historicalSalesStability": historical_stability,
        })
    if partition_check.get("status") == "fail":
        add_finding("blocker", "SC007", "未上传国家的分区签名发生变化。", partition_check)
    elif partition_check.get("status") == "unavailable":
        add_finding("review", "SC008", "无法验证未上传国家分区稳定性。", partition_check)
    findings.append({
        "severity": "info",
        "scope": "country",
        "target": country,
        "ruleId": "SC201",
        "message": "candidate 仅读取目标国家分区；未上传国家 active 分区保持只读。",
        "metrics": {"rowCount": int(len(candidate_frame)), "latestMonth": candidate_latest},
        "suggestedAction": "manual_review_required",
    })

    rows = []
    for month in sorted(set(active_sales) | set(candidate_sales), key=_time_sort_key):
        reference_sales = active_sales.get(month)
        proposed_sales = candidate_sales.get(month)
        rows.append({
            "month": month,
            "referenceSales": reference_sales,
            "candidateSales": proposed_sales,
            "deltaSales": _serialize_numeric_value((proposed_sales or 0) - (reference_sales or 0)) if month in active_sales and month in candidate_sales else None,
            "changeStatus": "unchanged" if reference_sales == proposed_sales else "changed",
        })
    return {
        "jobId": str(payload.get("jobId") or ""),
        "reviewDir": None,
        "compareId": f"{payload.get('jobId')}-single-country",
        "decisionSuggestion": "reject_input_batch" if any(item["severity"] == "blocker" for item in findings) else "manual_review_required",
        "compareKeyColumns": key_columns,
        "checklistMarkdown": "\n".join(f"- {item['severity']}: {item['message']}" for item in findings),
        "reviewFindings": findings,
        "sampledCountries": [country],
        "conflictSampleCount": 0,
        "conflictSamples": [],
        "overlapChangeSummary": [],
        "countryFreshnessSummary": [{
            "country": country,
            "oldLatestMonth": active_latest,
            "newLatestMonth": candidate_latest,
            "freshnessStatus": "advanced" if active_latest and candidate_latest and _time_sort_key(candidate_latest) > _time_sort_key(active_latest) else "unchanged_latest",
            "rowDelta": row_delta,
        }],
        "countryCoverageSummary": [{
            "country": country,
            "oldMonths": _detect_month_columns(list(active_frame.columns)),
            "newMonths": month_columns,
            "addedMonths": sorted(set(month_columns) - set(_detect_month_columns(list(active_frame.columns))), key=_time_sort_key),
            "removedMonths": sorted(set(_detect_month_columns(list(active_frame.columns))) - set(month_columns), key=_time_sort_key),
            "overlappingMonths": common_months,
            "coverageStatus": "single_country",
        }],
        "countrySalesReferenceLabel": "网站当前 active",
        "countryMonthlySalesSummary": [{"country": country, "rows": rows}],
        "countryMonthlySalesError": None,
        "timeAxisCheck": {
            "targetCountry": country,
            "activeLatestMonth": active_latest,
            "candidateLatestMonth": candidate_latest,
            "schema": schema_contract,
        },
        "countryScopeSummary": {"targetCountry": country, "untouchedPartitionCheck": partition_check},
        "refreshSummary": _summarize_refresh_report(_read_json_if_exists(str(artifacts.get("refreshReportPath") or "")) or {}),
        "candidateFingerprint": (
            candidate_fingerprint
            if candidate_fingerprint is not None
            else _candidate_fingerprint_id(artifacts)
        ),
        "historicalReclassificationReport": (
            historical_reclassification_report
        ),
        "approval": payload.get("reviewApproval"),
    }


def _build_partial_country_review(
    payload: dict[str, Any],
    *,
    candidate_fingerprint: str | None = None,
) -> dict[str, Any]:
    countries = _ordered_distinct_strings(
        [
            str(country)
            for country in (
                payload.get("countryScope")
                if isinstance(payload.get("countryScope"), list)
                else []
            )
        ]
    )
    if len(countries) < 2:
        raise HTTPException(status_code=409, detail="部分国家任务缺少至少两个目标国家。")

    country_reviews = [
        _build_single_country_review(
            {**payload, "country": country},
            candidate_fingerprint=candidate_fingerprint,
        )
        for country in countries
    ]
    findings = [
        finding
        for review in country_reviews
        for finding in review["reviewFindings"]
    ]
    compare_key_columns = _ordered_distinct_strings(
        [
            column
            for review in country_reviews
            for column in review["compareKeyColumns"]
        ]
    )
    checklist_sections = [
        "\n".join(
            [
                f"## {country}",
                review["checklistMarkdown"],
            ]
        )
        for country, review in zip(countries, country_reviews, strict=True)
    ]
    artifacts = payload.get("artifacts")
    partition_check = (
        artifacts.get("untouchedPartitionCheck")
        if isinstance(artifacts, dict)
        else None
    )
    if not isinstance(partition_check, dict):
        partition_check = {"status": "unavailable"}
    current_reclassification_countries: list[dict[str, Any]] = []
    resolution_validation: list[dict[str, Any]] = []
    seen_reclassification_countries: set[str] = set()
    for review in country_reviews:
        report = review.get("historicalReclassificationReport")
        raw_countries = (
            report.get("countries")
            if isinstance(report, dict)
            and isinstance(report.get("countries"), list)
            else []
        )
        for item in raw_countries:
            if not isinstance(item, dict):
                continue
            country_key = str(item.get("country") or "").strip().casefold()
            if not country_key or country_key in seen_reclassification_countries:
                continue
            seen_reclassification_countries.add(country_key)
            current_reclassification_countries.append(item)
        raw_resolution_validation = (
            report.get("resolutionValidation")
            if isinstance(report, dict)
            and isinstance(report.get("resolutionValidation"), list)
            else []
        )
        resolution_validation.extend(
            dict(item)
            for item in raw_resolution_validation
            if isinstance(item, dict)
        )
    historical_reclassification_report = (
        _build_historical_reclassification_report(
            payload=payload,
            current_countries=current_reclassification_countries,
        )
    )
    if (
        historical_reclassification_report.get("status") == "resolved"
        and isinstance(artifacts, dict)
        and str(artifacts.get("candidateScope") or "")
        == "full_smart_merge"
    ):
        historical_reclassification_report[
            "resolutionValidation"
        ] = _exact_partial_keep_active_resolution_validation(
            payload=payload,
            raw_validation=resolution_validation,
        )
    return {
        "jobId": str(payload.get("jobId") or ""),
        "reviewDir": None,
        "compareId": f"{payload.get('jobId')}-partial-country",
        "decisionSuggestion": (
            "reject_input_batch"
            if any(item["severity"] == "blocker" for item in findings)
            else "manual_review_required"
        ),
        "compareKeyColumns": compare_key_columns,
        "checklistMarkdown": "\n\n".join(checklist_sections),
        "reviewFindings": findings,
        "sampledCountries": countries,
        "conflictSampleCount": 0,
        "conflictSamples": [],
        "overlapChangeSummary": [
            item
            for review in country_reviews
            for item in review["overlapChangeSummary"]
        ],
        "countryFreshnessSummary": [
            item
            for review in country_reviews
            for item in review["countryFreshnessSummary"]
        ],
        "countryCoverageSummary": [
            item
            for review in country_reviews
            for item in review["countryCoverageSummary"]
        ],
        "countrySalesReferenceLabel": "网站当前 active",
        "countryMonthlySalesSummary": [
            item
            for review in country_reviews
            for item in review["countryMonthlySalesSummary"]
        ],
        "countryMonthlySalesError": None,
        "timeAxisCheck": {
            "countries": {
                country: review["timeAxisCheck"]
                for country, review in zip(countries, country_reviews, strict=True)
            },
        },
        "countryScopeSummary": {
            "targetCountries": countries,
            "untouchedPartitionCheck": partition_check,
        },
        "refreshSummary": country_reviews[0]["refreshSummary"],
        "candidateFingerprint": country_reviews[0]["candidateFingerprint"],
        "historicalReclassificationReport": (
            historical_reclassification_report
        ),
        "approval": payload.get("reviewApproval"),
    }


def _review_refresh_operation(
    payload: dict[str, Any],
) -> dict[str, Any] | None:
    value = _pending_operation(payload)
    return (
        value
        if isinstance(value, dict)
        and str(value.get("type") or "") == "review_refresh"
        else None
    )


def _valid_sha256(value: Any) -> str | None:
    normalized = str(value or "").strip().lower()
    return normalized if re.fullmatch(r"[0-9a-f]{64}", normalized) else None


def _review_refresh_eligibility(
    *,
    payload: dict[str, Any],
    cached_candidate_fingerprint: str | None,
    bundle_exists: bool,
) -> tuple[bool, str, str | None]:
    operation = _review_refresh_operation(payload)
    if isinstance(operation, dict) and str(operation.get("status") or "") in {
        "queued",
        "running",
    }:
        return False, "review_refresh_in_progress", None
    if (
        str(payload.get("status") or "") != "success"
        or str(payload.get("phase") or "") != "completed"
    ):
        return False, "job_not_completed", None
    publication = payload.get("publication")
    if isinstance(publication, dict) and publication.get("publishedAt"):
        return False, "candidate_already_published", None
    pending = _pending_operation(payload)
    if isinstance(pending, dict) and str(pending.get("status") or "") in {
        "queued",
        "running",
    }:
        return False, "active_operation_in_progress", None
    active_base_fingerprint = _valid_sha256(
        payload.get("activeBaseFingerprint")
    )
    if active_base_fingerprint is None:
        return False, "active_lineage_missing", None
    try:
        current_active_fingerprint = _active_dataset_version()
    except Exception:
        return False, "active_lineage_unavailable", None
    if current_active_fingerprint != active_base_fingerprint:
        return False, "active_lineage_changed", current_active_fingerprint
    artifacts = payload.get("artifacts")
    if not isinstance(artifacts, dict):
        return False, "candidate_artifacts_missing", current_active_fingerprint
    candidate_scope = str(artifacts.get("candidateScope") or "")
    required_names = (
        {"parquet", "manifest", "refreshReport"}
        if candidate_scope in PARTITION_SCOPED_CANDIDATE_SCOPES
        else {name for name, _field in CANDIDATE_ARTIFACT_FIELDS}
    )
    if any(
        (
            path := _project_path(
                str(artifacts.get(artifact_field) or "").strip()
            )
        )
        is None
        or not path.exists()
        for artifact_name, artifact_field in CANDIDATE_ARTIFACT_FIELDS
        if artifact_name in required_names
    ):
        return False, "candidate_artifacts_missing", current_active_fingerprint
    if bundle_exists and cached_candidate_fingerprint is None:
        return (
            False,
            "candidate_fingerprint_unavailable",
            current_active_fingerprint,
        )
    return True, "review_refresh_available", current_active_fingerprint


def _review_bundle_unavailable_detail(
    *,
    payload: dict[str, Any],
    blocker_type: str,
    reason: str,
    message: str,
    cached_candidate_fingerprint: str | None,
    bundle_exists: bool,
) -> dict[str, Any]:
    can_rebuild, eligibility_reason, current_active_fingerprint = (
        _review_refresh_eligibility(
            payload=payload,
            cached_candidate_fingerprint=cached_candidate_fingerprint,
            bundle_exists=bundle_exists,
        )
    )
    return {
        "blockerType": blocker_type,
        "reason": reason,
        "rebuildBlockerReason": (
            None if can_rebuild else eligibility_reason
        ),
        "message": message,
        "canRebuild": can_rebuild,
        "candidateFingerprint": cached_candidate_fingerprint,
        "activeBaseFingerprint": _valid_sha256(
            payload.get("activeBaseFingerprint")
        ),
        "currentActiveFingerprint": current_active_fingerprint,
        "reviewRefresh": _review_refresh_operation(payload),
    }


def _configured_review_bundle_path(
    *,
    job_id: str,
    artifacts: dict[str, Any],
) -> Path:
    configured = str(artifacts.get("reviewBundlePath") or "").strip()
    return (
        _project_path(configured)
        if configured
        else _job_review_bundle_path(job_id)
    ) or _job_review_bundle_path(job_id)


def _read_cached_review_bundle(
    path: Path,
) -> tuple[dict[str, Any] | None, str | None]:
    if not path.exists():
        return None, None
    try:
        return _read_json(path), None
    except (OSError, UnicodeError, json.JSONDecodeError, RuntimeError):
        return None, "review_bundle_corrupt"


def get_jato_monthly_update_review(
    job_id: str,
    *,
    allow_build: bool = False,
    force_rebuild: bool = False,
    candidate_fingerprint: str | None = None,
) -> dict[str, Any]:
    if force_rebuild and not allow_build:
        raise ValueError("force_rebuild 只允许隔离 worker 构建 Review。")
    payload = _load_job_state(job_id)
    artifacts = payload.get("artifacts")
    if not isinstance(artifacts, dict):
        raise HTTPException(status_code=409, detail="当前任务暂无可 review 的 compare 产物。")
    review_bundle_path = _configured_review_bundle_path(
        job_id=job_id,
        artifacts=artifacts,
    )
    if not force_rebuild and review_bundle_path.exists():
        review_bundle, read_error = _read_cached_review_bundle(
            review_bundle_path
        )
        if read_error or review_bundle is None:
            raise HTTPException(
                status_code=409,
                detail=_review_bundle_unavailable_detail(
                    payload=payload,
                    blocker_type="review_bundle_stale",
                    reason=read_error or "review_bundle_corrupt",
                    message=(
                        "Review bundle 无法解析为可信 JSON object；"
                        "已拒绝在 Web 请求中重建。"
                    ),
                    cached_candidate_fingerprint=None,
                    bundle_exists=True,
                ),
            )
        cached_stat_signature = str(
            review_bundle.get("candidateArtifactStatSignature") or ""
        )
        cached_candidate_fingerprint = _valid_sha256(
            review_bundle.get("candidateFingerprint")
        )
        contract_error = _review_bundle_contract_error(review_bundle)
        if contract_error:
            raise HTTPException(
                status_code=409,
                detail=_review_bundle_unavailable_detail(
                    payload=payload,
                    blocker_type="review_bundle_stale",
                    reason=contract_error,
                    message=(
                        "Review bundle 使用旧版或缺失的 schema / 候选签名；"
                        "请在隔离 worker 中重建 Review。"
                    ),
                    cached_candidate_fingerprint=(
                        cached_candidate_fingerprint
                    ),
                    bundle_exists=True,
                ),
            )
        if cached_candidate_fingerprint is None:
            raise HTTPException(
                status_code=409,
                detail=_review_bundle_unavailable_detail(
                    payload=payload,
                    blocker_type="review_bundle_stale",
                    reason="candidate_fingerprint_unavailable",
                    message=(
                        "Review bundle 缺少可信 candidate 内容指纹；"
                        "拒绝在 Web 请求中接受或重建。"
                    ),
                    cached_candidate_fingerprint=None,
                    bundle_exists=True,
                ),
            )
        current_stat_signature = _candidate_artifact_stat_signature(artifacts)
        if cached_stat_signature != current_stat_signature:
            raise HTTPException(
                status_code=409,
                detail=_review_bundle_unavailable_detail(
                    payload=payload,
                    blocker_type="review_bundle_stale",
                    reason="candidate_metadata_changed",
                    message=(
                        "candidate 的部署稳定元数据与 Review bundle 不一致；"
                        "请由隔离 worker 校验内容指纹并重建 Review。"
                    ),
                    cached_candidate_fingerprint=(
                        cached_candidate_fingerprint
                    ),
                    bundle_exists=True,
                ),
            )
        cached_historical_report = review_bundle.get(
            "historicalReclassificationReport"
        )
        if (
            isinstance(cached_historical_report, dict)
            and cached_historical_report.get("status") != "resolved"
        ):
            review_bundle["historicalReclassificationReport"] = (
                _validated_normalized_historical_reclassification_report(
                    cached_historical_report
                )
            )
        review_bundle["approval"] = payload.get("reviewApproval")
        return review_bundle
    if not allow_build:
        raise HTTPException(
            status_code=409,
            detail=_review_bundle_unavailable_detail(
                payload=payload,
                blocker_type="review_bundle_not_ready",
                reason="review_bundle_missing",
                message=(
                    "Review bundle 尚未由隔离 worker 生成完成；"
                    "Web 请求不会同步读取大型 candidate。"
                ),
                cached_candidate_fingerprint=None,
                bundle_exists=False,
            ),
        )

    raw_compare_report_path = str(artifacts.get("rawCompareReportPath") or "").strip()
    review_dir = str(artifacts.get("reviewDir") or "").strip()
    raw_compare_report = _read_json_if_exists(raw_compare_report_path)
    if raw_compare_report is None:
        job_type = str(payload.get("jobType") or "")
        if job_type == "single_country":
            return _build_single_country_review(
                payload,
                candidate_fingerprint=candidate_fingerprint,
            )
        if job_type == "partial_country":
            return _build_partial_country_review(
                payload,
                candidate_fingerprint=candidate_fingerprint,
            )
        raise HTTPException(status_code=409, detail="当前任务暂无可 review 的 compare 报告。")

    checklist_markdown = _read_text_if_exists(
        f"{review_dir}/review_checklist.md" if review_dir else None
    )
    conflict_payload = _read_json_if_exists(
        f"{review_dir}/conflict_samples.json" if review_dir else None
    )
    refresh_report = _read_json_if_exists(str(artifacts.get("refreshReportPath") or "").strip())

    findings = [
        sanitized
        for sanitized in (
            _sanitize_review_finding(item)
            for item in raw_compare_report.get("reviewFindings", [])
        )
        if sanitized is not None
    ]
    sampled_countries: list[str] = []
    sample_count = 0
    conflict_samples: list[dict[str, Any]] = []
    overlap_change_summary = [
        sanitized
        for sanitized in (
            _sanitize_overlap_change_summary(item)
            for item in raw_compare_report.get("overlapChangeSummary", [])
        )
        if sanitized is not None
    ]
    country_freshness_summary = [
        sanitized
        for sanitized in (
            _sanitize_country_freshness_summary(item)
            for item in raw_compare_report.get("countryFreshnessSummary", [])
        )
        if sanitized is not None
    ]
    country_coverage_summary = [
        sanitized
        for sanitized in (
            _sanitize_country_coverage_summary(item)
            for item in raw_compare_report.get("countryCoverageSummary", [])
        )
        if sanitized is not None
    ]
    if isinstance(conflict_payload, dict):
        sampled = conflict_payload.get("sampledCountries")
        samples = conflict_payload.get("samples")
        if isinstance(sampled, list):
            sampled_countries = [str(item) for item in sampled]
        if isinstance(samples, list):
            sample_count = len(samples)
            conflict_samples = [
                sanitized
                for sanitized in (
                    _sanitize_conflict_sample(item) for item in samples
                )
                if sanitized is not None
            ]

    review_countries = _ordered_distinct_strings(
        [
            *(item.get("country", "") for item in overlap_change_summary),
            *(item.get("country", "") for item in country_freshness_summary),
            *(item.get("country", "") for item in country_coverage_summary),
            *sampled_countries,
        ]
    )
    country_monthly_sales_summary: list[dict[str, Any]] = []
    country_sales_reference_label = "-"
    country_monthly_sales_error: str | None = None
    reference_path: Path | None = None
    candidate_path = _project_path(str(artifacts.get("stagingOutputPath") or "").strip())
    if candidate_path is not None:
        reference_path, country_sales_reference_label = (
            _resolve_review_reference_dataset(artifacts)
        )
    if review_countries:
        if candidate_path is None:
            country_monthly_sales_error = (
                "缺少 candidate parquet 产物，无法生成逐月销量核对表。"
            )
        else:
            try:
                country_monthly_sales_summary = _build_country_monthly_sales_summary(
                    countries=review_countries,
                    candidate_path=candidate_path,
                    reference_path=reference_path,
                )
            except HTTPException as exc:
                country_monthly_sales_error = str(exc.detail)
    historical_report_countries = review_countries
    if (
        candidate_path is not None
        and str(artifacts.get("candidateScope") or "")
        not in {
            "target_country_partition_only",
            "target_country_partitions_only",
        }
    ):
        historical_report_countries = list(
            _collect_dataset_country_latest_months(candidate_path)
        )
    historical_reclassification_report = (
        _build_historical_reclassification_report_from_paths(
            payload=payload,
            countries=historical_report_countries,
            active_path=reference_path,
            candidate_path=candidate_path,
        )
    )
    resolution_validation = (
        historical_reclassification_report.get("resolutionValidation")
        if isinstance(
            historical_reclassification_report.get(
                "resolutionValidation"
            ),
            list,
        )
        else []
    )
    keep_active_validation_by_key = {
        str(item.get("country") or "").strip().casefold(): item
        for item in resolution_validation
        if isinstance(item, dict)
        and item.get("decision") == "keep_active"
        and str(item.get("country") or "").strip()
    }
    if historical_reclassification_report.get("status") == "resolved":
        findings = [
            item
            for item in findings
            if not (
                isinstance(item, dict)
                and (
                    blocker_country_key := (
                        _historical_sales_changed_blocker_country_key(
                            item
                        )
                    )
                )
                and (
                    keep_active_validation_by_key.get(
                        blocker_country_key,
                        {},
                    ).get("status")
                    == "pass"
                )
            )
        ]
        for country_key, validation in (
            keep_active_validation_by_key.items()
        ):
            if validation.get("status") == "pass":
                continue
            findings.append(
                {
                    "severity": "blocker",
                    "scope": "country",
                    "target": validation.get("country"),
                    "ruleId": "SC011",
                    "blockerType": (
                        "historical_keep_active_validation_failed"
                    ),
                    "message": (
                        "keep_active 合并后的 candidate 仍与 active "
                        "历史不一致；拒绝批准 Publish。"
                    ),
                    "metrics": {
                        **validation,
                        "blockerType": (
                            "historical_keep_active_validation_failed"
                        ),
                    },
                    "suggestedAction": "reject_input_batch",
                    "sourceFeedback": (
                        "请保留 active 的全部已发布历史，仅让上传数据"
                        "提供 active 最新月之后的新月份。"
                    ),
                }
            )
    existing_sc011_targets = {
        str(item.get("target") or "").strip().casefold()
        for item in findings
        if isinstance(item, dict)
        and item.get("ruleId") == "SC011"
    }
    for country_report in historical_reclassification_report["countries"]:
        country = str(country_report.get("country") or "").strip()
        if not country or country.casefold() in existing_sc011_targets:
            continue
        monthly_totals_stable = bool(
            country_report.get("monthlyTotalsStable")
        )
        decision_required = bool(
            country_report.get("decisionRequired")
        )
        if decision_required:
            findings.append(
                {
                    "severity": "review",
                    "scope": "country",
                    "target": country,
                    "ruleId": "SC011",
                    "message": (
                        (
                            "历史月总量稳定，但分析维度发生重分类；"
                            "必须逐国选择 use_latest 或 keep_active。"
                        )
                        if monthly_totals_stable
                        else (
                            "历史国家/月总量发生变化；只能选择 "
                            "keep_active，以 active 保留已发布历史。"
                        )
                    ),
                    "metrics": country_report,
                    "suggestedAction": "manual_review_required",
                    "sourceFeedback": (
                        "请核对报告中的旧值→新值、月份和转移销量。"
                    ),
                }
            )
        elif not monthly_totals_stable:
            findings.append(
                {
                    "severity": "blocker",
                    "scope": "country",
                    "target": country,
                    "ruleId": "SC011",
                    "message": "历史国家/月销量总量变化，不能通过分类选择放行。",
                    "metrics": country_report,
                    "suggestedAction": "reject_input_batch",
                    "sourceFeedback": (
                        "请恢复 active 已有月份的国家总销量。"
                    ),
                }
            )
    for unavailable in historical_reclassification_report.get(
        "unavailableCountries",
        [],
    ):
        if not isinstance(unavailable, dict):
            continue
        country = str(unavailable.get("country") or "").strip()
        findings.append(
            {
                "severity": "blocker",
                "scope": "country",
                "target": country,
                "ruleId": "SC011",
                "message": (
                    "缺少 Make/Model、分析维度或月份列，"
                    "无法生成历史重分类报告。"
                ),
                "metrics": {
                    "reason": "historical_configuration_guard_unavailable",
                    "detail": unavailable.get("detail"),
                },
                "suggestedAction": "reject_input_batch",
                "sourceFeedback": (
                    "请保留 Country、Make、Model、分析维度和历史月份列。"
                ),
            }
        )

    return {
        "jobId": job_id,
        "reviewDir": review_dir or None,
        "compareId": str(raw_compare_report.get("compareId", "")),
        "decisionSuggestion": str(raw_compare_report.get("decisionSuggestion", "")),
        "compareKeyColumns": (
            [str(item) for item in raw_compare_report.get("compareKeyColumns", [])]
            if isinstance(raw_compare_report.get("compareKeyColumns"), list)
            else []
        ),
        "checklistMarkdown": checklist_markdown,
        "reviewFindings": findings,
        "sampledCountries": sampled_countries,
        "conflictSampleCount": sample_count,
        "conflictSamples": conflict_samples,
        "overlapChangeSummary": overlap_change_summary,
        "countryFreshnessSummary": country_freshness_summary,
        "countryCoverageSummary": country_coverage_summary,
        "countrySalesReferenceLabel": country_sales_reference_label,
        "countryMonthlySalesSummary": country_monthly_sales_summary,
        "countryMonthlySalesError": country_monthly_sales_error,
        "timeAxisCheck": (
            raw_compare_report.get("timeAxisCheck")
            if isinstance(raw_compare_report.get("timeAxisCheck"), dict)
            else {}
        ),
        "countryScopeSummary": (
            raw_compare_report.get("countryScopeSummary")
            if isinstance(raw_compare_report.get("countryScopeSummary"), dict)
            else {}
        ),
        "refreshSummary": (
            _summarize_refresh_report(refresh_report)
            if isinstance(refresh_report, dict)
            else None
        ),
        "candidateFingerprint": (
            candidate_fingerprint
            if candidate_fingerprint is not None
            else _candidate_fingerprint_id(artifacts)
        ),
        "historicalReclassificationReport": (
            historical_reclassification_report
        ),
        "approval": payload.get("reviewApproval"),
    }


def _cache_jato_monthly_update_review(
    job_id: str,
    *,
    expected_candidate_fingerprint: str | None = None,
    expected_active_fingerprint: str | None = None,
    review_generation_id: str | None = None,
) -> Path:
    """Generate and atomically replace Review in the isolated worker.

    The previous bundle remains in place until the new report has passed the
    candidate-content, metadata-snapshot and active-lineage gates.
    """
    payload = _load_job_state(job_id)
    artifacts = payload.get("artifacts")
    if not isinstance(artifacts, dict):
        raise HTTPException(
            status_code=409,
            detail="当前任务缺少 candidate 产物，不能缓存 Review。",
        )
    bundle_path = _job_review_bundle_path(job_id)
    active_fingerprint = _valid_sha256(payload.get("activeBaseFingerprint"))
    normalized_expected_active = _valid_sha256(expected_active_fingerprint)
    if expected_active_fingerprint is not None and normalized_expected_active is None:
        raise HTTPException(
            status_code=409,
            detail="Review 重建缺少有效 active lineage 指纹。",
        )
    if (
        normalized_expected_active is not None
        and (
            active_fingerprint != normalized_expected_active
            or _active_dataset_version() != normalized_expected_active
        )
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "stale_candidate",
                "message": "Review 重建前 active lineage 已变化。",
            },
        )
    stat_signature_before = _candidate_artifact_stat_signature(artifacts)
    actual_candidate_fingerprint = _candidate_fingerprint_id(artifacts)
    normalized_expected_candidate = _valid_sha256(
        expected_candidate_fingerprint
    )
    if (
        expected_candidate_fingerprint is not None
        and normalized_expected_candidate is None
    ):
        raise HTTPException(
            status_code=409,
            detail="Review 重建缺少有效 candidate 指纹。",
        )
    if (
        normalized_expected_candidate is not None
        and actual_candidate_fingerprint != normalized_expected_candidate
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "candidate_content_drift",
                "message": (
                    "candidate 内容指纹与旧 Review 不一致；"
                    "已保留旧 bundle，未修改 candidate 或 active。"
                ),
                "expectedCandidateFingerprint": (
                    normalized_expected_candidate
                ),
                "actualCandidateFingerprint": (
                    actual_candidate_fingerprint
                ),
            },
        )
    review = get_jato_monthly_update_review(
        job_id,
        allow_build=True,
        force_rebuild=True,
        candidate_fingerprint=actual_candidate_fingerprint,
    )
    latest_payload = _load_job_state(job_id)
    latest_artifacts = latest_payload.get("artifacts")
    if not isinstance(latest_artifacts, dict):
        raise HTTPException(
            status_code=409,
            detail="当前任务缺少 candidate 产物，不能缓存 Review。",
        )
    stat_signature_after = _candidate_artifact_stat_signature(
        latest_artifacts
    )
    if stat_signature_after != stat_signature_before:
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "candidate_content_drift",
                "message": (
                    "candidate 在 Review 重建期间发生变化；"
                    "已保留旧 bundle，未修改 candidate 或 active。"
                ),
            },
        )
    if str(review.get("candidateFingerprint") or "") != (
        actual_candidate_fingerprint
    ):
        raise HTTPException(
            status_code=409,
            detail="Review candidate 指纹与 worker 校验结果不一致。",
        )
    if (
        normalized_expected_active is not None
        and (
            _valid_sha256(latest_payload.get("activeBaseFingerprint"))
            != normalized_expected_active
            or _active_dataset_version() != normalized_expected_active
        )
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "stale_candidate",
                "message": "Review 重建期间 active lineage 已变化。",
            },
        )
    review["reviewBundleSchemaVersion"] = REVIEW_BUNDLE_SCHEMA_VERSION
    review["candidateArtifactStatSignatureVersion"] = (
        CANDIDATE_ARTIFACT_STAT_SIGNATURE_VERSION
    )
    review["candidateArtifactStatSignature"] = stat_signature_after
    review["generatedAt"] = _utc_now().isoformat()
    review["reviewGenerationId"] = review_generation_id
    _write_json(bundle_path, review)
    payload = _load_job_state(job_id)
    artifacts = payload.get("artifacts")
    payload["artifacts"] = artifacts if isinstance(artifacts, dict) else {}
    payload["artifacts"]["reviewBundlePath"] = _relative_to_project(bundle_path)
    _persist_job_state(payload)
    return bundle_path


def _review_refresh_failure_digest(exc: BaseException) -> dict[str, Any]:
    resource_digest = _failure_digest_from_exception(
        phase="review_refresh",
        exc=exc,
    )
    if resource_digest.get("code") == "MEMORY_LIMIT_EXCEEDED":
        return {
            **resource_digest,
            "sourceFeedback": (
                "无需重新洗数或重新上传；Candidate 和 active 均未修改。"
                "请仅重试刷新 Review。"
            ),
            "nextAction": "retry_review_refresh",
        }
    detail: Any = exc.detail if isinstance(exc, HTTPException) else None
    message = (
        str(detail.get("message") or detail)
        if isinstance(detail, dict)
        else str(detail or exc)
    )
    return {
        "code": "REVIEW_REFRESH_BLOCKED",
        "category": "safety_gate" if isinstance(exc, HTTPException) else "processing",
        "phase": "review_refresh",
        "retryable": not isinstance(exc, HTTPException),
        "message": message,
        "sourceFeedback": None,
        "technicalDetail": detail or type(exc).__name__,
        "nextAction": "inspect_review_refresh",
    }


def _review_refresh_bundle_is_durable(
    *,
    payload: dict[str, Any],
    operation: dict[str, Any],
) -> bool:
    operation_id = str(operation.get("operationId") or "")
    expected_candidate_fingerprint = _valid_sha256(
        operation.get("expectedCandidateFingerprint")
    )
    expected_active_fingerprint = _valid_sha256(
        operation.get("expectedActiveFingerprint")
    )
    artifacts = payload.get("artifacts")
    bundle_path = _job_review_bundle_path(str(payload.get("jobId") or ""))
    if (
        not operation_id
        or expected_candidate_fingerprint is None
        or expected_active_fingerprint is None
        or not isinstance(artifacts, dict)
        or not bundle_path.is_file()
    ):
        return False
    try:
        bundle = _read_json(bundle_path)
        return bool(
            str(bundle.get("reviewGenerationId") or "") == operation_id
            and _valid_sha256(bundle.get("candidateFingerprint"))
            == expected_candidate_fingerprint
            and int(bundle.get("reviewBundleSchemaVersion") or 0)
            == REVIEW_BUNDLE_SCHEMA_VERSION
            and int(
                bundle.get("candidateArtifactStatSignatureVersion") or 0
            )
            == CANDIDATE_ARTIFACT_STAT_SIGNATURE_VERSION
            and str(bundle.get("candidateArtifactStatSignature") or "")
            == _candidate_artifact_stat_signature(artifacts)
            and _valid_sha256(payload.get("activeBaseFingerprint"))
            == expected_active_fingerprint
            and _active_dataset_version() == expected_active_fingerprint
        )
    except Exception:
        return False


def _queue_jato_monthly_update_review_refresh_locked(
    *,
    job_id: str,
    triggered_by: str,
    request_id: str,
    expected_candidate_fingerprint: str | None,
) -> dict[str, Any]:
    payload = _load_job_state(job_id)
    existing = _review_refresh_operation(payload)
    normalized_expected = _valid_sha256(expected_candidate_fingerprint)
    if expected_candidate_fingerprint is not None and normalized_expected is None:
        raise HTTPException(
            status_code=400,
            detail="expectedCandidateFingerprint 必须是 64 位 SHA-256。",
        )
    if isinstance(existing, dict):
        existing_status = str(existing.get("status") or "")
        existing_expected = _valid_sha256(
            existing.get("expectedCandidateFingerprint")
        )
        same_candidate = (
            normalized_expected is None
            or existing_expected == normalized_expected
        )
        if existing_status in {"queued", "running"}:
            if not same_candidate:
                raise HTTPException(
                    status_code=409,
                    detail="另一个 candidate 的 Review 重建正在运行。",
                )
            if existing_status == "queued":
                _launch_job_thread(job_id)
            return _serialize_job_state(payload, include_log_tail=False)
        if (
            existing_status == "success"
            and str(existing.get("requestId") or "") == request_id
            and same_candidate
        ):
            return _serialize_job_state(payload, include_log_tail=False)

    if (
        str(payload.get("status") or "") != "success"
        or str(payload.get("phase") or "") != "completed"
    ):
        raise HTTPException(
            status_code=409,
            detail="只有 success/completed 的任务才能重建 Review。",
        )
    publication = payload.get("publication")
    if isinstance(publication, dict) and publication.get("publishedAt"):
        raise HTTPException(
            status_code=409,
            detail="已发布或已回滚的 candidate 不能重建旧 Review。",
        )
    pending = _pending_operation(payload)
    if isinstance(pending, dict) and str(pending.get("status") or "") in {
        "queued",
        "running",
    }:
        raise HTTPException(
            status_code=409,
            detail="Publish/Rollback 正在排队或运行，不能重建 Review。",
        )
    artifacts = payload.get("artifacts")
    if not isinstance(artifacts, dict):
        raise HTTPException(
            status_code=409,
            detail="当前任务缺少 candidate 产物，不能重建 Review。",
        )
    bundle_path = _configured_review_bundle_path(
        job_id=job_id,
        artifacts=artifacts,
    )
    cached_bundle, bundle_read_error = _read_cached_review_bundle(bundle_path)
    if bundle_read_error:
        raise HTTPException(
            status_code=409,
            detail=_review_bundle_unavailable_detail(
                payload=payload,
                blocker_type="review_bundle_stale",
                reason=bundle_read_error,
                message=(
                    "Review bundle 无法解析为可信 JSON object；"
                    "旧 candidate 指纹不可用，拒绝自动重建。"
                ),
                cached_candidate_fingerprint=None,
                bundle_exists=True,
            ),
        )
    cached_candidate_fingerprint = (
        _valid_sha256(cached_bundle.get("candidateFingerprint"))
        if isinstance(cached_bundle, dict)
        else None
    )
    if (
        cached_bundle is not None
        and cached_candidate_fingerprint is None
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "review_bundle_stale",
                "reason": "candidate_fingerprint_unavailable",
                "message": "旧 Review 缺少可信 candidate 指纹，拒绝自动重建。",
                "canRebuild": False,
            },
        )
    if (
        normalized_expected is not None
        and cached_candidate_fingerprint is not None
        and normalized_expected != cached_candidate_fingerprint
    ):
        raise HTTPException(
            status_code=409,
            detail="页面绑定的 candidate 指纹与当前旧 Review 不一致。",
        )
    source_candidate_fingerprint = (
        normalized_expected or cached_candidate_fingerprint
    )
    active_fingerprint = _valid_sha256(payload.get("activeBaseFingerprint"))
    if (
        active_fingerprint is None
        or _active_dataset_version() != active_fingerprint
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "stale_candidate",
                "message": "active lineage 已变化，不能重建旧 candidate 的 Review。",
            },
        )
    current_stat_signature = _candidate_artifact_stat_signature(artifacts)
    if isinstance(cached_bundle, dict):
        cached_signature = str(
            cached_bundle.get("candidateArtifactStatSignature") or ""
        )
        if (
            _review_bundle_contract_error(cached_bundle) is None
            and cached_signature == current_stat_signature
        ):
            raise HTTPException(
                status_code=409,
                detail="当前 Review bundle 已是最新，无需重建。",
            )
    payload["reviewApproval"] = None
    return _queue_active_bundle_operation(
        payload=payload,
        operation_type="review_refresh",
        triggered_by=triggered_by,
        operation_fields={
            "requestId": request_id,
            "phase": "queued",
            "expectedCandidateFingerprint": source_candidate_fingerprint,
            "expectedActiveFingerprint": active_fingerprint,
            "candidateArtifactStatSignature": current_stat_signature,
            "candidateArtifactStatSignatureVersion": (
                CANDIDATE_ARTIFACT_STAT_SIGNATURE_VERSION
            ),
        },
    )


def refresh_jato_monthly_update_review(
    *,
    job_id: str,
    triggered_by: str,
    request_id: str,
    expected_candidate_fingerprint: str | None,
) -> dict[str, Any]:
    normalized_request_id = request_id.strip()
    if not REVIEW_REFRESH_REQUEST_ID_PATTERN.fullmatch(normalized_request_id):
        raise HTTPException(
            status_code=400,
            detail="requestId 格式无效。",
        )
    with _monthly_update_worker_start_window(
        action="重建 Review",
        excluding_job_id=job_id,
    ):
        with _exclusive_file_lock(_job_state_lock_path(job_id)) as acquired:
            if not acquired:
                raise HTTPException(
                    status_code=503,
                    detail="Review 重建状态锁暂不可用，请稍后刷新状态。",
                )
            return _queue_jato_monthly_update_review_refresh_locked(
                job_id=job_id,
                triggered_by=triggered_by,
                request_id=normalized_request_id,
                expected_candidate_fingerprint=(
                    expected_candidate_fingerprint
                ),
            )


def approve_jato_monthly_update_review(
    *,
    job_id: str,
    triggered_by: str,
    decision: str,
    note: str | None = None,
) -> dict[str, Any]:
    normalized_decision = decision.strip().lower()
    if normalized_decision not in {"approve", "reject"}:
        raise HTTPException(status_code=400, detail="review decision 只支持 approve 或 reject。")
    with _exclusive_file_lock(_job_state_lock_path(job_id)) as acquired:
        if not acquired:
            raise HTTPException(
                status_code=503,
                detail="Review 状态锁暂不可用，请稍后重试。",
            )
        payload = _load_job_state(job_id)
        artifacts = payload.get("artifacts")
        if not isinstance(artifacts, dict):
            raise HTTPException(status_code=409, detail="当前任务缺少 candidate，不能确认 Review。")
        pending = _pending_operation(payload)
        if isinstance(pending, dict) and str(
            pending.get("status") or ""
        ) in {"queued", "running"}:
            raise HTTPException(
                status_code=409,
                detail="后台月更操作已排队或运行，不能覆盖 Review 状态。",
            )
        publication = payload.get("publication")
        if isinstance(publication, dict) and publication.get("publishedAt"):
            raise HTTPException(
                status_code=409,
                detail="该 candidate 已进入发布记录，不能重新写入旧 Review 决策。",
            )
        review = get_jato_monthly_update_review(job_id)
        historical_reclassification_report = review.get(
            "historicalReclassificationReport"
        )
        if (
            normalized_decision == "approve"
            and isinstance(historical_reclassification_report, dict)
            and historical_reclassification_report.get("status")
            == "decision_required"
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": (
                        "historical_reclassification_decision_required"
                    ),
                    "message": (
                        "历史分析维度变化尚未逐国选择；请先选择"
                        " use_latest 或 keep_active 并生成新的完整 candidate。"
                    ),
                    "reportFingerprint": (
                        historical_reclassification_report.get(
                            "reportFingerprint"
                        )
                    ),
                },
            )
        findings = review.get("reviewFindings")
        has_blocker = isinstance(findings, list) and any(
            isinstance(item, dict) and item.get("severity") == "blocker"
            for item in findings
        )
        if normalized_decision == "approve" and has_blocker:
            raise HTTPException(status_code=409, detail="Review 存在 blocker，不能批准 Publish。")
        candidate_fingerprint = str(
            review.get("candidateFingerprint") or ""
        ).strip()
        if not re.fullmatch(r"[0-9a-f]{64}", candidate_fingerprint):
            raise HTTPException(
                status_code=409,
                detail="Review bundle 缺少有效 candidate fingerprint，请由 worker 重新生成 Review。",
            )
        candidate_active_fingerprint = str(
            payload.get("activeBaseFingerprint") or ""
        ).strip()
        if not re.fullmatch(
            r"[0-9a-f]{64}",
            candidate_active_fingerprint,
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": "candidate_lineage_missing",
                    "message": (
                        "candidate 未绑定构建时的 active 指纹，不能在审批时"
                        "用当前 active 补写；请重建 candidate。"
                    ),
                },
            )
        current_active_fingerprint = _active_dataset_version()
        if candidate_active_fingerprint != current_active_fingerprint:
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": "stale_candidate",
                    "message": "Review 生成后 active 数据已变化，请重建 candidate 后再审批。",
                    "candidateActiveFingerprint": candidate_active_fingerprint,
                    "currentActiveFingerprint": current_active_fingerprint,
                },
            )
        payload["reviewApproval"] = {
            "decision": "approved" if normalized_decision == "approve" else "rejected",
            "reviewedAt": _utc_now().isoformat(),
            "reviewedBy": triggered_by.strip() or "anonymous",
            "candidateFingerprint": candidate_fingerprint,
            "activeBaseFingerprint": candidate_active_fingerprint,
            "note": str(note or "").strip() or None,
        }
        if (
            normalized_decision == "approve"
            and isinstance(historical_reclassification_report, dict)
            and historical_reclassification_report.get("status")
            == "resolved"
        ):
            resolution = _historical_reclassification_resolution(
                payload
            )
            validated_decisions = (
                _validated_historical_reclassification_resolution(
                    resolution
                )
                if isinstance(resolution, dict)
                else {}
            )
            if not (
                isinstance(resolution, dict)
                and resolution.get("status") == "resolved"
                and bool(validated_decisions)
                and str(
                    resolution.get(
                        "resolvedCandidateFingerprint"
                    )
                    or ""
                )
                == candidate_fingerprint
                and str(resolution.get("reportFingerprint") or "")
                == str(
                    historical_reclassification_report.get(
                        "reportFingerprint"
                    )
                    or ""
                )
            ):
                raise HTTPException(
                    status_code=409,
                    detail={
                        "blockerType": (
                            "historical_reclassification_resolution_stale"
                        ),
                        "message": (
                            "历史分类决策与当前 candidate/report 指纹不一致，"
                            "请重新选择并生成 candidate。"
                        ),
                    },
                )
            expected_keep_active = {
                country_key: country
                for country_key, country in (
                    (
                        str(item.get("country") or "").strip().casefold(),
                        str(item.get("country") or "").strip(),
                    )
                    for item in resolution.get("decisions", [])
                    if isinstance(item, dict)
                    and str(item.get("country") or "").strip()
                )
                if validated_decisions.get(country_key) == "keep_active"
            }
            raw_resolution_validation = (
                historical_reclassification_report.get(
                    "resolutionValidation"
                )
            )
            validated_keep_active: dict[str, dict[str, Any]] = {}
            invalid_keep_active_validation = bool(
                expected_keep_active
                and not isinstance(raw_resolution_validation, list)
            )
            for item in (
                raw_resolution_validation
                if isinstance(raw_resolution_validation, list)
                else []
            ):
                if not isinstance(item, dict):
                    invalid_keep_active_validation = True
                    continue
                country_key = str(
                    item.get("country") or ""
                ).strip().casefold()
                if (
                    not country_key
                    or country_key in validated_keep_active
                    or item.get("decision") != "keep_active"
                ):
                    invalid_keep_active_validation = True
                    continue
                validated_keep_active[country_key] = item
            missing_keep_active = (
                set(expected_keep_active) - set(validated_keep_active)
            )
            extra_keep_active = (
                set(validated_keep_active) - set(expected_keep_active)
            )
            failed_keep_active = [
                expected_keep_active[country_key]
                for country_key, item in validated_keep_active.items()
                if country_key in expected_keep_active
                and item.get("status") != "pass"
            ]
            if (
                invalid_keep_active_validation
                or missing_keep_active
                or extra_keep_active
                or failed_keep_active
            ):
                raise HTTPException(
                    status_code=409,
                    detail={
                        "blockerType": (
                            "historical_keep_active_validation_failed"
                        ),
                        "message": (
                            "keep_active 尚未被当前完整 Candidate 逐国复核通过，"
                            "拒绝批准 Publish。"
                        ),
                        "missingCountries": [
                            expected_keep_active[country_key]
                            for country_key in sorted(
                                missing_keep_active
                            )
                        ],
                        "failedCountries": sorted(
                            failed_keep_active
                        ),
                        "extraCountries": sorted(
                            extra_keep_active
                        ),
                    },
                )
            payload["reviewApproval"][
                "historicalReclassification"
            ] = {
                "reportFingerprint": resolution.get(
                    "reportFingerprint"
                ),
                "resolvedCandidateFingerprint": resolution.get(
                    "resolvedCandidateFingerprint"
                ),
                "decisions": [
                    {
                        "country": item["country"],
                        "decision": validated_decisions[
                            str(item["country"]).strip().casefold()
                        ],
                    }
                    for item in resolution.get("decisions", [])
                    if isinstance(item, dict)
                    and str(item.get("country") or "").strip().casefold()
                    in validated_decisions
                ],
            }
        _persist_job_state(payload)
    _append_log(
        _job_log_path(job_id),
        f"[{_utc_now().isoformat()}] Review {normalized_decision} by {triggered_by.strip() or 'anonymous'}.",
    )
    return _serialize_job_state(payload, include_log_tail=True)


def _resolve_jato_historical_reclassification_with_job_lock(
    *,
    job_id: str,
    triggered_by: str,
    decisions: Any,
) -> dict[str, Any]:
    """Bind per-country choices, then queue the existing isolated rebuild."""
    if not isinstance(decisions, list):
        raise HTTPException(
            status_code=400,
            detail="decisions 必须是逐国决策数组。",
        )
    with _exclusive_file_lock(_job_state_lock_path(job_id)) as acquired:
        if not acquired:
            raise HTTPException(
                status_code=503,
                detail="历史重分类状态锁暂不可用，请稍后重试。",
            )
        payload = _load_job_state(job_id)
        pending = _pending_operation(payload)
        if isinstance(pending, dict) and str(
            pending.get("status") or ""
        ) in {"queued", "running"}:
            raise HTTPException(
                status_code=409,
                detail="后台月更操作正在运行，不能应用历史重分类决策。",
            )
        review = get_jato_monthly_update_review(job_id)
        report = review.get("historicalReclassificationReport")
        if not isinstance(report, dict):
            raise HTTPException(
                status_code=409,
                detail="当前 Review 没有待处理的历史重分类决策。",
            )
        normalized_report = (
            _validated_normalized_historical_reclassification_report(
                report
            )
        )
        if normalized_report.get("status") != "decision_required":
            raise HTTPException(
                status_code=409,
                detail="当前 Review 没有待处理的历史重分类决策。",
            )
        report_countries = normalized_report["countries"]
        required_by_key: dict[str, str] = {}
        required_reports_by_key: dict[str, dict[str, Any]] = {}
        for item in report_countries:
            if not bool(item.get("decisionRequired")):
                continue
            country = str(item.get("country") or "").strip()
            country_key = country.casefold()
            if not country or country_key in required_by_key:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "Review 报告包含重复或空国家，拒绝写入决策。"
                    ),
                )
            required_by_key[country_key] = country
            required_reports_by_key[country_key] = item
        if not required_by_key:
            raise HTTPException(
                status_code=409,
                detail="当前 Review 没有可决策的历史分类变化。",
            )

        findings = review.get("reviewFindings")
        blockers = [
            item
            for item in (
                findings
                if isinstance(findings, list)
                else []
            )
            if isinstance(item, dict)
            and item.get("severity") == "blocker"
        ]
        unresolved_blockers: list[dict[str, Any]] = []
        for blocker in blockers:
            target_key = (
                _historical_sales_changed_blocker_country_key(blocker)
            )
            country_report = required_reports_by_key.get(target_key)
            resolves_with_keep_active = bool(
                target_key
                and country_report is not None
                and country_report.get("monthlyTotalsStable") is False
            )
            if not resolves_with_keep_active:
                unresolved_blockers.append(blocker)
        if unresolved_blockers:
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": "review_blockers_present",
                    "message": (
                        "Review 仍有不可通过历史分类决策解决的 blocker，"
                        "拒绝生成 candidate。"
                    ),
                    "rules": [
                        str(item.get("ruleId") or "")
                        for item in unresolved_blockers
                    ],
                    "blockerTypes": [
                        str(
                            item.get("blockerType")
                            or (
                                item.get("metrics", {}).get(
                                    "blockerType"
                                )
                                if isinstance(item.get("metrics"), dict)
                                else ""
                            )
                            or (
                                item.get("metrics", {}).get("reason")
                                if isinstance(item.get("metrics"), dict)
                                else ""
                            )
                            or ""
                        )
                        for item in unresolved_blockers
                    ],
                },
            )

        submitted: dict[str, str] = {}
        for raw_item in decisions:
            if not isinstance(raw_item, dict):
                raise HTTPException(
                    status_code=400,
                    detail="每项 decision 必须包含 country 和 decision。",
                )
            country = str(raw_item.get("country") or "").strip()
            country_key = country.casefold()
            decision = str(
                raw_item.get("decision") or ""
            ).strip().lower()
            if not country or country_key in submitted:
                raise HTTPException(
                    status_code=400,
                    detail="decisions 国家不能为空且每国只能出现一次。",
                )
            if decision not in HISTORICAL_RECLASSIFICATION_DECISIONS:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"{country} decision 只支持 use_latest 或 "
                        "keep_active。"
                    ),
                )
            country_report = required_reports_by_key.get(country_key)
            if (
                country_report is not None
                and decision
                not in _historical_reclassification_allowed_decisions(
                    country_report
                )
            ):
                raise HTTPException(
                    status_code=409,
                    detail={
                        "blockerType": (
                            "historical_sales_changed_requires_keep_active"
                        ),
                        "message": (
                            f"{required_by_key[country_key]} 的历史国家/月"
                            "总量发生变化，只能选择 keep_active；"
                            "拒绝 use_latest 改写已发布历史。"
                        ),
                        "country": required_by_key[country_key],
                        "allowedDecisions": ["keep_active"],
                    },
                )
            submitted[country_key] = decision
        missing_keys = set(required_by_key) - set(submitted)
        extra_keys = set(submitted) - set(required_by_key)
        if missing_keys or extra_keys:
            raise HTTPException(
                status_code=400,
                detail={
                    "blockerType": (
                        "historical_reclassification_decisions_incomplete"
                    ),
                    "message": "必须严格覆盖所有受影响国家，不能缺少或增加国家。",
                    "missingCountries": [
                        required_by_key[key]
                        for key in sorted(missing_keys)
                    ],
                    "extraCountries": sorted(extra_keys),
                },
            )

        candidate_fingerprint = str(
            review.get("candidateFingerprint") or ""
        ).strip()
        report_fingerprint = str(
            normalized_report.get("reportFingerprint") or ""
        ).strip()
        active_fingerprint = str(
            payload.get("activeBaseFingerprint") or ""
        ).strip()
        if not (
            re.fullmatch(r"[0-9a-f]{64}", candidate_fingerprint)
            and re.fullmatch(r"[0-9a-f]{64}", report_fingerprint)
            and re.fullmatch(r"[0-9a-f]{64}", active_fingerprint)
        ):
            raise HTTPException(
                status_code=409,
                detail="Review 缺少 candidate/active/report 指纹，请重新生成 Review。",
            )
        if _active_dataset_version() != active_fingerprint:
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": "stale_candidate",
                    "message": (
                        "Review 后 active 已变化，不能应用旧的历史分类决策。"
                    ),
                },
            )
        canonical_decisions = [
            {
                "country": required_by_key[key],
                "decision": submitted[key],
            }
            for key in sorted(required_by_key)
        ]
        payload["historicalReclassificationResolution"] = {
            "status": "queued",
            "requestedAt": _utc_now().isoformat(),
            "requestedBy": triggered_by.strip() or "anonymous",
            "activeBaseFingerprint": active_fingerprint,
            "sourceCandidateFingerprint": candidate_fingerprint,
            "reportFingerprint": report_fingerprint,
            "decisions": canonical_decisions,
            "report": normalized_report,
        }
        _persist_job_state(payload)
        try:
            return _create_smart_merge_candidate_locked(
                job_id=job_id,
                triggered_by=triggered_by,
            )
        except Exception as exc:
            failed_payload = _load_job_state(job_id)
            failed_resolution = (
                _historical_reclassification_resolution(failed_payload)
            )
            if isinstance(failed_resolution, dict):
                failed_resolution["status"] = "failed"
                failed_resolution["failedAt"] = _utc_now().isoformat()
                failed_resolution["error"] = str(exc)
                failed_payload[
                    "historicalReclassificationResolution"
                ] = failed_resolution
                _persist_job_state(failed_payload)
            raise


def resolve_jato_historical_reclassification(
    *,
    job_id: str,
    triggered_by: str,
    decisions: Any,
) -> dict[str, Any]:
    if not isinstance(decisions, list):
        raise HTTPException(
            status_code=400,
            detail="decisions 必须是逐国决策数组。",
        )
    with _monthly_update_worker_start_window(
        action="应用历史重分类决策",
        excluding_job_id=job_id,
    ):
        return _resolve_jato_historical_reclassification_with_job_lock(
            job_id=job_id,
            triggered_by=triggered_by,
            decisions=decisions,
        )


def _pending_operation(payload: dict[str, Any]) -> dict[str, Any] | None:
    value = payload.get("pendingOperation")
    return value if isinstance(value, dict) else None


def _queue_active_bundle_operation(
    *,
    payload: dict[str, Any],
    operation_type: str,
    triggered_by: str,
    operation_fields: dict[str, Any] | None = None,
) -> dict[str, Any]:
    existing = _pending_operation(payload)
    if isinstance(existing, dict) and str(existing.get("status") or "") in {
        "queued",
        "running",
    }:
        existing_type = str(existing.get("type") or "")
        if existing_type != operation_type:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"{existing_type or 'active bundle'} 操作仍在"
                    f" {existing.get('status')}，不能同时排队 {operation_type}。"
                ),
            )
        if str(existing.get("status") or "") == "queued":
            # A replay is also a durable wake-up. Multiple drain processes are
            # harmless because the cross-process worker flock admits only one.
            _launch_job_thread(str(payload["jobId"]))
        return _serialize_job_state(payload, include_log_tail=False)

    now = _utc_now().isoformat()
    payload["pendingOperation"] = {
        "operationId": f"jato-{operation_type}-{uuid4().hex[:10]}",
        "type": operation_type,
        "status": "queued",
        "requestedAt": now,
        "requestedBy": triggered_by.strip() or "anonymous",
        "startedAt": None,
        "finishedAt": None,
        "error": None,
        "failureDigest": None,
        **(operation_fields or {}),
    }
    _persist_job_state(payload)
    try:
        _launch_job_thread(str(payload["jobId"]))
    except Exception as exc:
        latest = _load_job_state(str(payload["jobId"]))
        operation = _pending_operation(latest) or {}
        operation["status"] = "failed"
        operation["finishedAt"] = _utc_now().isoformat()
        operation["error"] = str(exc)
        operation["failureDigest"] = (
            _review_refresh_failure_digest(exc)
            if operation_type == "review_refresh"
            else _failure_digest_from_exception(
                phase=f"{operation_type}_queued",
                exc=exc,
            )
        )
        latest["pendingOperation"] = operation
        _persist_job_state(latest)
        raise HTTPException(
            status_code=503,
            detail=f"{operation_type} worker 无法启动：{exc}",
        ) from exc
    return _serialize_job_state(
        _load_job_state(str(payload["jobId"])),
        include_log_tail=False,
    )


def _publish_jato_monthly_update_job_with_job_lock(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    with _exclusive_file_lock(_job_state_lock_path(job_id)) as acquired:
        if not acquired:  # blocking lock always acquires on supported platforms
            raise HTTPException(
                status_code=503,
                detail="Publish 状态锁暂不可用，请稍后重试。",
            )
        return _queue_publish_jato_monthly_update_job_locked(
            job_id=job_id,
            triggered_by=triggered_by,
        )


def publish_jato_monthly_update_job(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    with _monthly_update_worker_start_window(
        action="排队 Publish",
        excluding_job_id=job_id,
    ):
        return _publish_jato_monthly_update_job_with_job_lock(
            job_id=job_id,
            triggered_by=triggered_by,
        )


def _queue_publish_jato_monthly_update_job_locked(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    """Queue heavy Publish work so the HTTP request cannot be the memory owner."""
    _require_no_running_monthly_update_jobs(excluding_job_id=job_id)
    payload = _load_job_state(job_id)
    publication = payload.get("publication")
    if (
        isinstance(publication, dict)
        and publication.get("publishedAt")
        and not publication.get("rolledBackAt")
    ):
        # Safe replay after a lost HTTP response.
        return _serialize_job_state(payload, include_log_tail=False)
    if (
        isinstance(publication, dict)
        and publication.get("publishedAt")
        and publication.get("rolledBackAt")
    ):
        raise HTTPException(
            status_code=409,
            detail="该任务已发布并回滚，不能复用旧 candidate 再发布；请创建新 attempt。",
        )
    if (
        str(payload.get("status") or "") != "success"
        or str(payload.get("phase") or "") != "completed"
    ):
        raise HTTPException(
            status_code=409,
            detail="只有 success/completed 的月更任务才能排队 Publish。",
        )
    artifacts = payload.get("artifacts")
    if not isinstance(artifacts, dict):
        raise HTTPException(status_code=409, detail="当前任务缺少 staging 产物信息。")
    if artifacts.get("candidateScope") in {
        "target_country_partition_only",
        "target_country_partitions_only",
    }:
        raise HTTPException(
            status_code=409,
            detail="部分国家 candidate 只能 Review，不能直接 Publish。",
        )
    approval = payload.get("reviewApproval")
    if not isinstance(approval, dict) or approval.get("decision") != "approved":
        raise HTTPException(status_code=409, detail="必须先批准当前 candidate 的 Review。")
    approved_active_fingerprint = str(
        approval.get("activeBaseFingerprint") or ""
    )
    current_active_fingerprint = _active_dataset_version()
    if (
        not approved_active_fingerprint
        or approved_active_fingerprint != current_active_fingerprint
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "stale_candidate",
                "message": (
                    "审批后 active 数据已变化，旧 candidate 不得覆盖新 active；"
                    "请重新生成 Review。"
                ),
                "approvedActiveFingerprint": approved_active_fingerprint or None,
                "currentActiveFingerprint": current_active_fingerprint,
            },
        )
    return _queue_active_bundle_operation(
        payload=payload,
        operation_type="publish",
        triggered_by=triggered_by,
    )


def _execute_publish_jato_monthly_update_job(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    with _exclusive_file_lock(
        _active_bundle_lock_path(),
        blocking=False,
    ) as acquired:
        if not acquired:
            raise HTTPException(
                status_code=409,
                detail="另一个 Publish/Rollback 正在切换 active bundle，请等待其完成。",
            )
        recovered = _recover_incomplete_active_transactions(
            _active_data_paths()
        )
        if recovered:
            _append_log(
                _job_log_path(job_id),
                (
                    f"[{_utc_now().isoformat()}] recovered incomplete active "
                    f"transactions before publish: {recovered}"
                ),
            )
        return _publish_jato_monthly_update_job_locked(
            job_id=job_id,
            triggered_by=triggered_by,
        )


def _publish_jato_monthly_update_job_locked(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    _require_no_running_monthly_update_jobs(excluding_job_id=job_id)
    payload = _load_job_state(job_id)
    if str(payload.get("status", "")) != "success" or str(payload.get("phase", "")) != "completed":
        raise HTTPException(
            status_code=409,
            detail="只有 success/completed 的月更任务才能执行 publish。",
        )

    publication = payload.get("publication")
    if (
        isinstance(publication, dict)
        and publication.get("publishedAt")
        and not publication.get("rolledBackAt")
    ):
        raise HTTPException(status_code=409, detail="该月更任务已经 publish 过。")
    if (
        isinstance(publication, dict)
        and publication.get("publishedAt")
        and publication.get("rolledBackAt")
    ):
        raise HTTPException(
            status_code=409,
            detail="该月更任务已经回滚，不能复用旧 candidate 再次 publish。",
        )

    summaries = payload.get("summaries")
    refresh_summary = summaries.get("refresh") if isinstance(summaries, dict) else None
    if not isinstance(refresh_summary, dict) or str(refresh_summary.get("jobStatus", "")) != "success":
        raise HTTPException(status_code=409, detail="当前任务 refresh 尚未成功，不能 publish。")

    artifacts = payload.get("artifacts")
    if not isinstance(artifacts, dict):
        raise HTTPException(status_code=409, detail="当前任务缺少 staging 产物信息，不能 publish。")
    if artifacts.get("candidateScope") in {
        "target_country_partition_only",
        "target_country_partitions_only",
    }:
        raise HTTPException(
            status_code=409,
            detail="部分国家 candidate 仅用于 Review，不能直接 Publish。请使用已批准的完整 promotion 流程生成 active 产物。",
        )
    approval = payload.get("reviewApproval")
    if not isinstance(approval, dict) or approval.get("decision") != "approved":
        raise HTTPException(status_code=409, detail="必须先完成并批准当前 candidate 的 Review，才能 publish。")
    if str(approval.get("candidateFingerprint") or "") != _candidate_fingerprint_id(artifacts):
        raise HTTPException(status_code=409, detail="candidate 在 Review 后已变化，请重新 Review 并批准。")
    approved_active_fingerprint = str(
        approval.get("activeBaseFingerprint") or ""
    )
    current_active_fingerprint = _active_dataset_version()
    if (
        not approved_active_fingerprint
        or approved_active_fingerprint != current_active_fingerprint
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "stale_candidate",
                "message": "审批后 active 数据已变化，旧 candidate 不得覆盖新 active；请重新生成 Review。",
                "approvedActiveFingerprint": approved_active_fingerprint or None,
                "currentActiveFingerprint": current_active_fingerprint,
            },
        )

    source_paths = {
        "parquet": _project_path(str(artifacts.get("stagingOutputPath") or "").strip()),
        "manifest": _project_path(str(artifacts.get("manifestPath") or "").strip()),
        "partition": _project_path(str(artifacts.get("partitionOutputPath") or "").strip()),
        "fingerprint": _project_path(str(artifacts.get("fingerprintPath") or "").strip()),
        "refreshReport": _project_path(str(artifacts.get("refreshReportPath") or "").strip()),
        "summaries": _project_path(str(artifacts.get("summariesOutputPath") or "").strip()),
    }
    missing_sources = [
        key
        for key, path in source_paths.items()
        if path is None or not path.exists()
    ]
    if missing_sources:
        raise HTTPException(
            status_code=409,
            detail=f"当前任务缺少 publish 所需产物：{', '.join(missing_sources)}。",
        )
    _validate_candidate_full_bundle(
        parquet_path=source_paths["parquet"],
        manifest_path=source_paths["manifest"],
        partition_path=source_paths["partition"],
        fingerprint_path=source_paths["fingerprint"],
        refresh_report_path=source_paths["refreshReport"],
        summaries_path=source_paths["summaries"],
    )
    active_paths = _active_data_paths()
    duplicate_assessment = _publish_duplicate_configuration_assessment(
        payload=payload,
        active_parquet_path=active_paths["parquet"],
        candidate_parquet_path=source_paths["parquet"],
        refresh_report_path=source_paths["refreshReport"],
    )
    duplicate_configurations = duplicate_assessment["blocking"]
    inherited_duplicate_configurations = duplicate_assessment["inherited"]
    if duplicate_configurations:
        platform_guard_failure = any(
            entry.get("duplicateStatus")
            in {
                "duplicate_guard_scope_invalid",
                "untouched_country_content_changed",
                "untouched_country_duplicate_changed",
            }
            for entry in duplicate_configurations
        )
        rendered = "；".join(
            (
                f"{entry.get('country') or '范围证明'} "
                f"({entry['duplicateRows']} 行/"
                f"{entry['duplicateGroupCount']} 组)"
            )
            for entry in duplicate_configurations[:5]
        )
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "duplicate_configurations",
                "message": (
                    "candidate 存在完全相同的静态配置行，可能只在新月份"
                    f"造成销量重复累加，拒绝 Publish：{rendered}。"
                ),
                "sourceFeedback": (
                    (
                        "未上传国家的完整内容或范围证明与 active 不一致；"
                        "这不是本批源文件"
                        "可修复的范围。请平台管理员检查 Smart Merge 与未触碰"
                        "国家分区，禁止让洗数人员修改未上传国家。"
                    )
                    if platform_guard_failure
                    else (
                        "请洗数人员按全部非月份业务字段去除真正重复的配置行；"
                        "价格、Registration type、动力、车身或版本不同的记录"
                        "不能合并。请同时提供去重前后行数。"
                    )
                ),
                "countries": duplicate_configurations,
                "duplicateConfigurationGuard": (
                    duplicate_assessment.get("guard")
                ),
            },
        )
    if inherited_duplicate_configurations:
        inherited_rendered = "；".join(
            (
                f"{entry['country']} "
                f"({entry['duplicateRows']} 行/"
                f"{entry['duplicateGroupCount']} 组)"
            )
            for entry in inherited_duplicate_configurations
        )
        _append_log(
            _job_log_path(job_id),
            (
                f"[{_utc_now().isoformat()}] Publish: 未上传国家的重复配置"
                "经逐国完整内容指纹证明与 active 完全一致，"
                "按历史遗留只读保留："
                f"{inherited_rendered}。"
            ),
        )

    if active_paths["parquet"].exists():
        historical_sales_changes = _find_publish_historical_sales_changes(
            active_parquet_path=active_paths["parquet"],
            candidate_parquet_path=source_paths["parquet"],
        )
        if historical_sales_changes:
            rendered = "；".join(
                (
                    f"{entry['country']} "
                    f"({entry['changedMonthCount']} 个历史月)"
                )
                for entry in historical_sales_changes[:5]
            )
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": "historical_sales_changed",
                    "message": (
                        "candidate 改写了已经发布的国家/月销量，拒绝 Publish："
                        f"{rendered}。本次月更只能新增最新月份，不能累加或重写历史。"
                    ),
                    "sourceFeedback": (
                        "请洗数人员恢复 active 已有月份的国家总销量；"
                        "本次文件只能推进新月份，不能把历史快照追加、翻倍或重算。"
                    ),
                    "changes": historical_sales_changes,
                },
            )
        upload_payload = (
            payload.get("upload")
            if isinstance(payload.get("upload"), dict)
            else {}
        )
        source_upload_sha256 = str(
            upload_payload.get("sha256") or ""
        ).strip().lower()
        if not re.fullmatch(r"[0-9a-f]{64}", source_upload_sha256):
            stored_upload_path = _project_path(
                str(upload_payload.get("storedPath") or "").strip()
            )
            source_upload_sha256 = (
                _sha256_hex_for_path(stored_upload_path)
                if stored_upload_path is not None
                and stored_upload_path.is_file()
                else None
            )
        approved_reclassification_decisions: dict[str, str] = {}
        approved_reclassification = approval.get(
            "historicalReclassification"
        )
        if (
            isinstance(approved_reclassification, dict)
            and str(
                approved_reclassification.get(
                    "resolvedCandidateFingerprint"
                )
                or ""
            )
            == str(approval.get("candidateFingerprint") or "")
        ):
            approved_reclassification_decisions = (
                _historical_reclassification_decision_map(
                    {
                        "decisions": approved_reclassification.get(
                            "decisions"
                        )
                    }
                )
            )
        historical_configuration_changes = (
            _find_publish_historical_configuration_changes(
                active_parquet_path=active_paths["parquet"],
                candidate_parquet_path=source_paths["parquet"],
                source_upload_sha256=source_upload_sha256,
                approved_reclassification_decisions=(
                    approved_reclassification_decisions
                ),
            )
        )
        if historical_configuration_changes:
            rendered = "；".join(
                (
                    f"{entry['country']} "
                    f"({entry['makeModelMismatchCount']} 个 Make/Model、"
                    f"{entry['analysisDimensionMismatchCount']} 个分析维度历史差异)"
                )
                for entry in historical_configuration_changes[:5]
            )
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": "historical_configuration_changed",
                    "message": (
                        "candidate 在国家总量未变的情况下改写了已发布月份的 "
                        f"Make/Model 或分析维度分布，拒绝 Publish：{rendered}。"
                        "只有与上传文件 SHA 和逐月转移量完全绑定的已确认 "
                        "SC011 重分类才能放行。"
                    ),
                    "sourceFeedback": (
                        "请洗数人员逐项说明旧值→新值映射及每月转移量；"
                        "未确认的 Make/Model、动力、燃料、Segment、Body、"
                        "Registration type 或 Version/Trim 变化请恢复 active 历史归类。"
                    ),
                    "changes": historical_configuration_changes,
                },
            )
        regressions = _find_publish_country_regressions(
            active_parquet_path=active_paths["parquet"],
            candidate_parquet_path=source_paths["parquet"],
        )
        if regressions:
            rendered = ", ".join(
                (
                    f"{entry['country']} "
                    f"({entry['activeLatestMonth']} -> {entry['candidateLatestMonth'] or '-'})"
                )
                for entry in regressions[:5]
            )
            extra = " 等" if len(regressions) > 5 else ""
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": "country_regression",
                    "message": (
                        "publish 会让当前 active 数据回退，请先更换 baseline 或重建 candidate："
                        f"{rendered}{extra}"
                    ),
                    "regressions": [
                        {
                            "country": r["country"],
                            "activeLatestMonth": r["activeLatestMonth"],
                            "candidateLatestMonth": r["candidateLatestMonth"],
                        }
                        for r in regressions
                    ],
                },
            )
        sales_doubling_anomalies = _find_publish_sales_doubling_anomalies(
            active_parquet_path=active_paths["parquet"],
            candidate_parquet_path=source_paths["parquet"],
        )
        if sales_doubling_anomalies:
            rendered = _render_sales_doubling_anomalies(sales_doubling_anomalies)
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": "sales_doubling",
                    "message": (
                        "publish 检测到 candidate 疑似重复合并，多个重叠月份销量约为 "
                        f"当前 active 的 2x：{rendered}。请先重建 candidate 或回滚到正确 active。"
                    ),
                    "anomalies": [
                        {
                            "country": a["country"],
                            "suspiciousMonthCount": a["suspiciousMonthCount"],
                            "sampleMonths": a["sampleMonths"][:3],
                            "rolling12Ratio": a["rolling12Ratio"],
                        }
                        for a in sales_doubling_anomalies
                    ],
                },
            )

    published_at = _utc_now()
    pending_operation = _pending_operation(payload)
    operation_id = (
        str(pending_operation.get("operationId") or "")
        if isinstance(pending_operation, dict)
        else ""
    )
    active_transaction_id = f"publish-{job_id}-{uuid4().hex}"
    backup_dir = active_paths["backupRoot"] / (
        f"manual-promote-{job_id}-{published_at.strftime('%Y%m%d-%H%M%S')}"
    )
    staging_root, staged_paths = _stage_active_bundle_sources(
        source_paths=source_paths,
        active_paths=active_paths,
        transaction_id=f"publish-{job_id}",
    )
    try:
        bundle_switch = _swap_staged_active_bundle(
            staged_paths=staged_paths,
            active_paths=active_paths,
            backup_dir=backup_dir,
            transaction_metadata={
                "transactionId": active_transaction_id,
                "operationType": "publish",
                "operationId": operation_id or None,
                "jobId": job_id,
            },
        )
    finally:
        shutil.rmtree(staging_root, ignore_errors=True)

    try:
        active_fingerprint_after = _active_dataset_version()
        cache_invalidation = _invalidate_jato_publish_runtime_caches()
        payload["publication"] = {
            "publishedAt": published_at.isoformat(),
            "publishedBy": triggered_by.strip() or "anonymous",
            "backupDir": _relative_to_project(backup_dir),
            "activeParquetPath": _relative_to_project(active_paths["parquet"]),
            "activeManifestPath": _relative_to_project(active_paths["manifest"]),
            "activePartitionPath": _relative_to_project(active_paths["partition"]),
            "activeFingerprintPath": _relative_to_project(active_paths["fingerprint"]),
            "activeRefreshReportPath": _relative_to_project(active_paths["refreshReport"]),
            "activeSummariesPath": _relative_to_project(active_paths["summaries"]),
            "summariesState": "candidate_bundle_promoted",
            "publishOperationId": operation_id or None,
            "activeTransactionId": active_transaction_id,
            "activeFingerprintBefore": current_active_fingerprint,
            "activeFingerprintAfter": active_fingerprint_after,
            "bundleSwitch": bundle_switch,
            "cacheInvalidation": cache_invalidation,
            "duplicateConfigurationGuard": duplicate_assessment.get(
                "guard"
            ),
            "inheritedDuplicateConfigurations": [
                {
                    "country": entry.get("country"),
                    "duplicateRows": entry.get("duplicateRows"),
                    "duplicateGroupCount": entry.get(
                        "duplicateGroupCount"
                    ),
                    "keyColumnCount": entry.get("keyColumnCount"),
                    "duplicateFingerprint": entry.get(
                        "duplicateFingerprint"
                    ),
                    "duplicateStatus": entry.get("duplicateStatus"),
                    "activeDuplicateFingerprint": entry.get(
                        "activeDuplicateFingerprint"
                    ),
                    "contentFingerprint": entry.get(
                        "contentFingerprint"
                    ),
                }
                for entry in inherited_duplicate_configurations
            ],
        }
        _persist_job_state(payload)
    except Exception:
        _recover_incomplete_active_transactions(active_paths)
        try:
            _invalidate_jato_publish_runtime_caches()
        except Exception:
            pass
        raise
    try:
        _commit_active_transaction(
            journal_path=backup_dir / ACTIVE_TRANSACTION_FILENAME,
            transaction_id=active_transaction_id,
        )
    except Exception as exc:
        # Job state is already the durable commit record. Recovery will finalize
        # this journal instead of rolling the active bundle back.
        _append_log(
            _job_log_path(job_id),
            (
                f"[{_utc_now().isoformat()}] publish transaction journal "
                f"finalization deferred: {exc}"
            ),
        )
    try:
        _write_jato_publish_cache_invalidation_evidence(
            job_id=job_id,
            published_at=published_at,
            triggered_by=triggered_by,
            cache_invalidation=cache_invalidation,
            active_paths=active_paths,
        )
    except Exception as exc:
        _append_log(
            _job_log_path(job_id),
            f"[{_utc_now().isoformat()}] Hermes evidence write failed after publish: {exc}",
        )
    _append_log(
        _job_log_path(job_id),
        (
            f"[{published_at.isoformat()}] published candidate to active dataset by "
            f"{triggered_by.strip() or 'anonymous'}"
        ),
    )
    return _serialize_job_state(payload, include_log_tail=True)


def _rollback_jato_monthly_update_job_with_job_lock(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    with _exclusive_file_lock(_job_state_lock_path(job_id)) as acquired:
        if not acquired:  # blocking lock always acquires on supported platforms
            raise HTTPException(
                status_code=503,
                detail="Rollback 状态锁暂不可用，请稍后重试。",
            )
        return _queue_rollback_jato_monthly_update_job_locked(
            job_id=job_id,
            triggered_by=triggered_by,
        )


def rollback_jato_monthly_update_job(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    with _monthly_update_worker_start_window(
        action="排队 Rollback",
        excluding_job_id=job_id,
    ):
        return _rollback_jato_monthly_update_job_with_job_lock(
            job_id=job_id,
            triggered_by=triggered_by,
        )


def _queue_rollback_jato_monthly_update_job_locked(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    """Queue rollback copies/switches and make response replay idempotent."""
    _require_no_running_monthly_update_jobs(excluding_job_id=job_id)
    payload = _load_job_state(job_id)
    publication = payload.get("publication")
    if not isinstance(publication, dict) or not publication.get("publishedAt"):
        raise HTTPException(status_code=409, detail="当前任务还没有 publish，不能回滚。")
    if publication.get("rolledBackAt"):
        return _serialize_job_state(payload, include_log_tail=False)

    expected_active_fingerprint = str(
        publication.get("activeFingerprintAfter") or ""
    )
    current_active_fingerprint = _active_dataset_version()
    if (
        not expected_active_fingerprint
        or expected_active_fingerprint != current_active_fingerprint
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "rollback_target_stale",
                "message": (
                    "当前 active 已不是该任务发布的版本，禁止用旧任务覆盖后续发布；"
                    "请先处理最新一次发布。"
                ),
                "expectedActiveFingerprint": (
                    expected_active_fingerprint or None
                ),
                "currentActiveFingerprint": current_active_fingerprint,
            },
        )
    return _queue_active_bundle_operation(
        payload=payload,
        operation_type="rollback",
        triggered_by=triggered_by,
    )


def _execute_rollback_jato_monthly_update_job(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    with _exclusive_file_lock(
        _active_bundle_lock_path(),
        blocking=False,
    ) as acquired:
        if not acquired:
            raise HTTPException(
                status_code=409,
                detail="另一个 Publish/Rollback 正在切换 active bundle，请等待其完成。",
            )
        recovered = _recover_incomplete_active_transactions(
            _active_data_paths()
        )
        if recovered:
            _append_log(
                _job_log_path(job_id),
                (
                    f"[{_utc_now().isoformat()}] recovered incomplete active "
                    f"transactions before rollback: {recovered}"
                ),
            )
        return _rollback_jato_monthly_update_job_locked(
            job_id=job_id,
            triggered_by=triggered_by,
        )


def _rollback_jato_monthly_update_job_locked(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    _require_no_running_monthly_update_jobs(excluding_job_id=job_id)
    payload = _load_job_state(job_id)
    publication = payload.get("publication")
    if not isinstance(publication, dict) or not publication.get("publishedAt"):
        raise HTTPException(status_code=409, detail="当前任务还没有 publish，不能回滚。")
    if publication.get("rolledBackAt"):
        raise HTTPException(status_code=409, detail="当前任务已经执行过回滚。")

    active_paths = _active_data_paths()
    expected_active_fingerprint = str(
        publication.get("activeFingerprintAfter") or ""
    )
    current_active_fingerprint = _active_dataset_version()
    if (
        not expected_active_fingerprint
        or expected_active_fingerprint != current_active_fingerprint
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "rollback_target_stale",
                "message": (
                    "当前 active 已不是该任务发布的版本，禁止乱序回滚覆盖后续发布。"
                ),
                "expectedActiveFingerprint": (
                    expected_active_fingerprint or None
                ),
                "currentActiveFingerprint": current_active_fingerprint,
            },
        )
    backup_dir = _project_path(str(publication.get("backupDir") or "").strip())
    if backup_dir is None or not backup_dir.exists():
        raise HTTPException(status_code=409, detail="找不到 publish 备份目录，不能回滚。")

    restore_sources = {
        "parquet": backup_dir / active_paths["parquet"].name,
        "manifest": backup_dir / active_paths["manifest"].name,
        "partition": backup_dir / active_paths["partition"].name,
        "fingerprint": backup_dir / active_paths["fingerprint"].name,
        "refreshReport": backup_dir / active_paths["refreshReport"].name,
        "summaries": backup_dir / active_paths["summaries"].name,
    }
    missing_sources = [
        key
        for key, path in restore_sources.items()
        if key != "summaries" and not path.exists()
    ]
    if missing_sources:
        raise HTTPException(
            status_code=409,
            detail=f"publish 备份不完整，缺少：{', '.join(missing_sources)}。",
        )

    rolled_back_at = _utc_now()
    pending_operation = _pending_operation(payload)
    operation_id = (
        str(pending_operation.get("operationId") or "")
        if isinstance(pending_operation, dict)
        else ""
    )
    rollback_transaction_id = f"rollback-{job_id}-{uuid4().hex}"
    rollback_backup_dir = active_paths["backupRoot"] / (
        f"restore-pre-{job_id}-{rolled_back_at.strftime('%Y%m%d-%H%M%S')}"
    )
    staging_root, staged_paths = _stage_active_bundle_sources(
        source_paths=restore_sources,
        active_paths=active_paths,
        transaction_id=f"rollback-{job_id}",
        optional_missing_keys={"summaries"},
    )
    try:
        bundle_switch = _swap_staged_active_bundle(
            staged_paths=staged_paths,
            active_paths=active_paths,
            backup_dir=rollback_backup_dir,
            transaction_metadata={
                "transactionId": rollback_transaction_id,
                "operationType": "rollback",
                "operationId": operation_id or None,
                "jobId": job_id,
            },
        )
    finally:
        shutil.rmtree(staging_root, ignore_errors=True)

    try:
        publication["rolledBackAt"] = rolled_back_at.isoformat()
        publication["rolledBackBy"] = triggered_by.strip() or "anonymous"
        publication["rollbackBackupDir"] = _relative_to_project(
            rollback_backup_dir
        )
        publication["rollbackBundleSwitch"] = bundle_switch
        publication["rollbackOperationId"] = operation_id or None
        publication["rollbackActiveTransactionId"] = rollback_transaction_id
        publication["rollbackActiveFingerprintBefore"] = (
            current_active_fingerprint
        )
        publication["rollbackActiveFingerprintAfter"] = (
            _active_dataset_version()
        )
        publication["rollbackCacheInvalidation"] = (
            _invalidate_jato_publish_runtime_caches()
        )
        payload["publication"] = publication
        _persist_job_state(payload)
    except Exception:
        _recover_incomplete_active_transactions(active_paths)
        try:
            _invalidate_jato_publish_runtime_caches()
        except Exception:
            pass
        raise
    try:
        _commit_active_transaction(
            journal_path=rollback_backup_dir / ACTIVE_TRANSACTION_FILENAME,
            transaction_id=rollback_transaction_id,
        )
    except Exception as exc:
        _append_log(
            _job_log_path(job_id),
            (
                f"[{_utc_now().isoformat()}] rollback transaction journal "
                f"finalization deferred: {exc}"
            ),
        )
    _append_log(
        _job_log_path(job_id),
        (
            f"[{rolled_back_at.isoformat()}] restored active dataset from publish backup by "
            f"{triggered_by.strip() or 'anonymous'}"
        ),
    )
    return _serialize_job_state(payload, include_log_tail=True)


def _run_logged_command(
    *,
    label: str,
    args: list[str],
    log_path: Path,
    job_id: str | None = None,
) -> None:
    tracking_job_id = job_id or _infer_job_id_from_log_path(log_path)
    if tracking_job_id:
        _ensure_job_not_cancelled(tracking_job_id)
    rendered_command = " ".join(shlex.quote(arg) for arg in args)
    _append_log(log_path, f"\n=== {label} ===\n$ {rendered_command}")
    env = dict(os.environ)
    env["PYTHONUNBUFFERED"] = "1"
    process = subprocess.Popen(
        args,
        cwd=PROJECT_ROOT,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
        env=env,
        start_new_session=True,
    )
    if tracking_job_id:
        _set_current_process(
            job_id=tracking_job_id,
            pid=process.pid,
            label=label,
            command=rendered_command,
        )
    last_heartbeat = time.monotonic()
    try:
        if process.stdout is not None:
            for line in process.stdout:
                _append_log(log_path, line.rstrip("\n"))
                if tracking_job_id and (time.monotonic() - last_heartbeat) >= 5:
                    _touch_current_process_heartbeat(job_id=tracking_job_id, pid=process.pid)
                    last_heartbeat = time.monotonic()
    finally:
        if process.stdout is not None:
            process.stdout.close()
    return_code = process.wait()
    if tracking_job_id:
        _clear_current_process(job_id=tracking_job_id, pid=process.pid)
    if return_code != 0:
        if tracking_job_id:
            _ensure_job_not_cancelled(tracking_job_id)
        error_message = f"{label} 失败，退出码 {return_code}"
        if return_code < 0:
            raise _JobResourceKilled(error_message)
        raise RuntimeError(error_message)
    if tracking_job_id:
        _ensure_job_not_cancelled(tracking_job_id)


def _prepare_initial_job_state(
    *,
    job_id: str,
    month: str,
    batch_id: str | None = None,
    triggered_by: str,
    upload_filename: str,
    stored_upload_path: Path,
    file_sha256: str | None = None,
    baseline_path: Path | None = None,
    baseline_source: str | None = None,
) -> dict[str, Any]:
    now = _utc_now().isoformat()
    artifacts: dict[str, Any] = {
        "jobDir": _relative_to_project(_job_dir(job_id)),
        "logPath": _relative_to_project(_job_log_path(job_id)),
    }
    if baseline_path is not None:
        artifacts["baselinePath"] = _relative_to_project(baseline_path)
    if baseline_source:
        artifacts["baselineSource"] = baseline_source
    return {
        "jobId": job_id,
        "month": month,
        "batchId": str(batch_id or month),
        "status": "queued",
        "phase": "queued",
        "triggeredBy": triggered_by,
        "createdAt": now,
        "updatedAt": now,
        "startedAt": None,
        "finishedAt": None,
        "error": None,
        "upload": {
            "originalFilename": upload_filename,
            "storedPath": _relative_to_project(stored_upload_path),
            "sizeBytes": stored_upload_path.stat().st_size,
            "sha256": file_sha256,
        },
        "plan": None,
        "artifacts": artifacts,
        "summaries": {},
        "logPath": _relative_to_project(_job_log_path(job_id)),
    }


def _prepare_upload_session_state(
    *,
    upload_id: str,
    filename: str,
    size_bytes: int,
    resume_key: str | None,
    triggered_by: str,
) -> dict[str, Any]:
    now = _utc_now().isoformat()
    owner = triggered_by.strip() or "anonymous"
    total_chunks = max((size_bytes + UPLOAD_CHUNK_SIZE_BYTES - 1) // UPLOAD_CHUNK_SIZE_BYTES, 1)
    return {
        "uploadId": upload_id,
        "filename": filename,
        "sizeBytes": size_bytes,
        "chunkSize": UPLOAD_CHUNK_SIZE_BYTES,
        "totalChunks": total_chunks,
        "receivedChunks": [],
        "chunkDigests": {},
        "uploadedBytes": 0,
        "status": "pending",
        "resumeKey": resume_key,
        "fileSha256": None,
        "owner": owner,
        "triggeredBy": owner,
        "createdAt": now,
        "updatedAt": now,
        "completedAt": None,
        "assembledPath": None,
        "ingestDigest": None,
        "failureDigest": None,
        "consumedJobId": None,
        "digestPid": None,
        "digestWorkerPid": None,
        "digestProcessIdentity": None,
        "digestLaunchedAt": None,
        "digestAttempts": 0,
        "digestAttempt": None,
    }


def _initiate_jato_monthly_update_upload_locked(
    *,
    filename: str,
    size_bytes: Any,
    resume_key: str | None,
    triggered_by: str,
) -> dict[str, Any]:
    normalized_filename = _validate_upload_filename(filename)
    normalized_size = _normalize_size_bytes(size_bytes)
    normalized_resume_key = str(resume_key or "").strip() or None
    owner = triggered_by.strip() or "anonymous"
    _reconcile_expired_upload_sessions()
    existing = _find_upload_session_by_resume_key(
        resume_key=normalized_resume_key or "",
        filename=normalized_filename,
        size_bytes=normalized_size,
        owner=owner,
    )
    if existing is not None:
        return _serialize_upload_session(existing)
    active_uploads = _active_upload_session_payloads()
    if active_uploads:
        active = active_uploads[0]
        if _same_upload_owner(_upload_session_owner(active), owner):
            detail = (
                "你已有一个未消费的上传会话；请继续原上传，"
                "或先明确放弃后再选择新文件。"
            )
        else:
            detail = (
                "已有其他用户的 JATO 上传/digest 正在进行；"
                "为避免并发 digest 击穿内存，请完成后再试。"
            )
        raise HTTPException(status_code=409, detail=detail)
    upload_id = f"jato-upload-{uuid4().hex[:10]}"
    session_dir = _upload_session_dir(upload_id)
    session_dir.mkdir(parents=True, exist_ok=True)
    _upload_session_chunk_dir(upload_id).mkdir(parents=True, exist_ok=True)
    state = _prepare_upload_session_state(
        upload_id=upload_id,
        filename=normalized_filename,
        size_bytes=normalized_size,
        resume_key=normalized_resume_key,
        triggered_by=owner,
    )
    _persist_upload_session(state)
    return _serialize_upload_session(state)


def initiate_jato_monthly_update_upload(
    *,
    filename: str,
    size_bytes: Any,
    resume_key: str | None,
    triggered_by: str,
) -> dict[str, Any]:
    """Create or resume exactly one session across all Uvicorn workers."""
    with _exclusive_file_lock(
        _maintenance_coordination_lock_path(),
        blocking=False,
    ) as coordinated:
        if not coordinated:
            raise HTTPException(
                status_code=409,
                detail="JATO 清理或 baseline 操作正在准备中，请稍后再开始上传。",
            )
        baseline_promotion = _load_baseline_promotion_state()
        if (
            isinstance(baseline_promotion, dict)
            and str(baseline_promotion.get("status") or "")
            in {"queued", "running"}
        ):
            raise HTTPException(
                status_code=409,
                detail="正在保存 active baseline，请完成后再开始新的上传/digest。",
            )
        _require_no_running_monthly_update_jobs()
        with _exclusive_file_lock(_upload_initiate_lock_path()) as acquired:
            if not acquired:  # blocking lock always acquires on supported platforms
                raise HTTPException(
                    status_code=503,
                    detail="上传会话锁暂不可用，请稍后重试。",
                )
            return _initiate_jato_monthly_update_upload_locked(
                filename=filename,
                size_bytes=size_bytes,
                resume_key=resume_key,
                triggered_by=triggered_by,
            )


def get_jato_monthly_update_upload(
    upload_id: str,
    *,
    requested_by: str,
    requested_role: str,
) -> dict[str, Any]:
    state = _load_upload_session(upload_id)
    _require_upload_session_access(
        state,
        requested_by=requested_by,
        requested_role=requested_role,
    )
    if isinstance(state.get("digestAttempt"), dict):
        with _exclusive_file_lock(
            _upload_state_lock_path(upload_id)
        ) as acquired:
            if acquired:
                state = _load_upload_session(upload_id)
                _require_upload_session_access(
                    state,
                    requested_by=requested_by,
                    requested_role=requested_role,
                )
                state = _reconcile_digest_attempt_receipt_locked(state)
    status = str(state.get("status") or "")
    if status in {"assembling", "digesting"}:
        digest_pid = int(state.get("digestPid") or 0)
        existing_failure = state.get("failureDigest")
        if (
            isinstance(existing_failure, dict)
            and str(existing_failure.get("code") or "")
            == "RESOURCE_QUARANTINED"
            and digest_pid > 0
            and _process_exists(digest_pid)
        ):
            return _serialize_upload_session(state)
        launched_at = str(state.get("digestLaunchedAt") or "")
        file_size_bytes = int(state.get("sizeBytes") or 0)
        timeout_seconds = _digest_worker_timeout_seconds(file_size_bytes)
        elapsed_seconds: float | None = None
        worker_missing = digest_pid > 0 and not _process_exists(digest_pid)
        launch_stale = False
        digest_timed_out = False
        if launched_at:
            try:
                launched = datetime.fromisoformat(launched_at)
                elapsed_seconds = (_utc_now() - launched).total_seconds()
                launch_stale = (
                    digest_pid <= 0
                    and elapsed_seconds > DIGEST_WORKER_STALE_GRACE_SECONDS
                )
                digest_timed_out = elapsed_seconds > timeout_seconds
            except ValueError:
                launch_stale = True
        attempt = state.get("digestAttempt")
        if (
            isinstance(attempt, dict)
            and (worker_missing or launch_stale)
            and not digest_timed_out
        ):
            missing_at = str(attempt.get("supervisorMissingAt") or "")
            missing_elapsed: float | None = None
            if missing_at:
                try:
                    missing_elapsed = (
                        _utc_now() - datetime.fromisoformat(missing_at)
                    ).total_seconds()
                except ValueError:
                    missing_elapsed = DIGEST_EXIT_RECEIPT_GRACE_SECONDS + 1
            if missing_elapsed is None:
                with _exclusive_file_lock(
                    _upload_state_lock_path(upload_id)
                ) as acquired:
                    if acquired:
                        state = _load_upload_session(upload_id)
                        state = _reconcile_digest_attempt_receipt_locked(state)
                        if str(state.get("status") or "") not in {
                            "assembling",
                            "digesting",
                        }:
                            return _serialize_upload_session(state)
                        current_attempt = state.get("digestAttempt")
                        if isinstance(current_attempt, dict):
                            current_attempt["supervisorMissingAt"] = (
                                _utc_now().isoformat()
                            )
                            state["digestAttempt"] = current_attempt
                            _persist_upload_session(state)
                return _serialize_upload_session(state)
            if missing_elapsed <= DIGEST_EXIT_RECEIPT_GRACE_SECONDS:
                return _serialize_upload_session(state)
        if worker_missing or launch_stale or digest_timed_out:
            termination: dict[str, Any] | None = None
            if (
                digest_timed_out
                and digest_pid > 0
                and _process_exists(digest_pid)
            ):
                termination = _terminate_digest_worker_with_evidence(
                    pid=digest_pid,
                    upload_id=upload_id,
                    expected_identity=(
                        state.get("digestProcessIdentity")
                        if isinstance(
                            state.get("digestProcessIdentity"),
                            dict,
                        )
                        else None
                    ),
                    attempt_id=(
                        str(state["digestAttempt"].get("attemptId") or "")
                        if isinstance(state.get("digestAttempt"), dict)
                        else None
                    ),
                )
            with ExitStack() as digest_probe_stack:
                digest_lock_held: bool | None = None
                if (
                    (launch_stale and digest_pid <= 0)
                    or (
                        worker_missing
                        and isinstance(state.get("digestAttempt"), dict)
                    )
                ):
                    digest_lock_acquired = digest_probe_stack.enter_context(
                        _exclusive_file_lock(
                            _upload_digest_lock_path(upload_id),
                            blocking=False,
                        )
                    )
                    digest_lock_held = not digest_lock_acquired
                if (
                    digest_lock_held is None
                    and isinstance(state.get("digestAttempt"), dict)
                    and not _process_exists(digest_pid)
                ):
                    digest_lock_acquired = digest_probe_stack.enter_context(
                        _exclusive_file_lock(
                            _upload_digest_lock_path(upload_id),
                            blocking=False,
                        )
                    )
                    digest_lock_held = not digest_lock_acquired
                with _exclusive_file_lock(
                    _upload_state_lock_path(upload_id)
                ) as acquired:
                    if acquired:
                        state = _load_upload_session(upload_id)
                        if str(state.get("status") or "") in {
                            "assembling",
                            "digesting",
                        }:
                            current_digest_pid = int(state.get("digestPid") or 0)
                            failed_process_identity = (
                                state.get("digestProcessIdentity")
                                if isinstance(
                                    state.get("digestProcessIdentity"),
                                    dict,
                                )
                                else None
                            )
                            termination_confirmed = (
                                current_digest_pid == digest_pid
                                and digest_lock_held is not True
                                and _digest_process_termination_confirmed(
                                    current_digest_pid,
                                    termination,
                                )
                            )
                            technical_detail = {
                                "digestPid": current_digest_pid or None,
                                "digestProcessIdentity": failed_process_identity,
                                "digestLaunchedAt": launched_at or None,
                                "digestLockHeld": digest_lock_held,
                                "timeoutSeconds": timeout_seconds,
                                "elapsedSeconds": (
                                    round(elapsed_seconds, 3)
                                    if elapsed_seconds is not None
                                    else None
                                ),
                                "fileSizeBytes": file_size_bytes,
                                "termination": termination,
                                "digestAttempt": state.get("digestAttempt"),
                                "logTail": _read_digest_attempt_log_tail(state),
                            }
                            launch_worker_lost = (
                                launch_stale and digest_pid <= 0
                            )
                            if termination_confirmed:
                                state["status"] = "invalid"
                                state["completedAt"] = _utc_now().isoformat()
                                state["digestPid"] = None
                                state["digestProcessIdentity"] = None
                                state["failureDigest"] = {
                                    "code": (
                                        "DIGEST_WORKER_LOST"
                                        if launch_worker_lost
                                        else (
                                            "DIGEST_TIMEOUT"
                                            if digest_timed_out
                                            else "DIGEST_WORKER_LOST"
                                        )
                                    ),
                                    "category": "resource",
                                    "phase": "digesting",
                                    "retryable": True,
                                    "message": (
                                        (
                                            "上传文件 digest worker 在写回 PID 前退出；"
                                            if launch_worker_lost
                                            else (
                                                "上传文件 digest 超过当前安全时限"
                                                f"（{timeout_seconds / 60:.1f} 分钟），"
                                                "已确认旧 worker 停止；"
                                                if digest_timed_out
                                                else "上传文件 digest worker 在生成报告前退出；"
                                            )
                                        )
                                        + "active 未修改。"
                                    ),
                                    "sourceFeedback": None,
                                    "technicalDetail": technical_detail,
                                    "nextAction": (
                                        "retry_digest_or_contact_admin"
                                    ),
                                }
                            else:
                                state["completedAt"] = None
                                state["failureDigest"] = {
                                    "code": "RESOURCE_QUARANTINED",
                                    "category": "resource",
                                    "phase": "digesting",
                                    "retryable": False,
                                    "message": (
                                        "digest worker 超时或失联，但系统未能确认进程已停止；"
                                        "会话继续占用资源隔离门禁，禁止新上传、任务和清理。"
                                    ),
                                    "sourceFeedback": None,
                                    "technicalDetail": technical_detail,
                                    "nextAction": "contact_admin_verify_digest_process",
                                }
                            _persist_upload_session(state)
    return _serialize_upload_session(state)


def abandon_jato_monthly_update_upload(
    *,
    upload_id: str,
    triggered_by: str,
    triggered_role: str = "editor",
) -> dict[str, Any]:
    """Explicitly end an unconsumed upload so maintenance cannot deadlock."""
    with _exclusive_file_lock(_upload_state_lock_path(upload_id)) as acquired:
        if not acquired:
            raise HTTPException(
                status_code=503,
                detail="上传会话状态锁暂不可用，请稍后重试。",
            )
        state = _load_upload_session(upload_id)
        _require_upload_session_access(
            state,
            requested_by=triggered_by,
            requested_role=triggered_role,
        )
        status = str(state.get("status") or "")
        if status == "consumed":
            raise HTTPException(
                status_code=409,
                detail="该上传已创建月更任务，不能通过放弃上传删除任务审计链。",
            )
        if status in {"abandoned", "expired", "invalid"}:
            return _serialize_upload_session(state)
        with ExitStack() as digest_probe_stack:
            digest_pid = int(state.get("digestPid") or 0)
            digest_lock_held: bool | None = None
            if (
                status in {"assembling", "digesting"}
                and (
                    digest_pid <= 0
                    or (
                        isinstance(state.get("digestAttempt"), dict)
                        and not _process_exists(digest_pid)
                    )
                )
            ):
                digest_lock_acquired = digest_probe_stack.enter_context(
                    _exclusive_file_lock(
                        _upload_digest_lock_path(upload_id),
                        blocking=False,
                    )
                )
                digest_lock_held = not digest_lock_acquired
            digest_identity = (
                state.get("digestProcessIdentity")
                if isinstance(state.get("digestProcessIdentity"), dict)
                else None
            )
            termination: dict[str, Any] | None = None
            if digest_pid > 0 and _process_exists(digest_pid):
                termination = _terminate_digest_worker_with_evidence(
                    pid=digest_pid,
                    upload_id=upload_id,
                    expected_identity=digest_identity,
                    attempt_id=(
                        str(state["digestAttempt"].get("attemptId") or "")
                        if isinstance(state.get("digestAttempt"), dict)
                        else None
                    ),
                )
            if (
                digest_lock_held is None
                and isinstance(state.get("digestAttempt"), dict)
                and not _process_exists(digest_pid)
            ):
                digest_lock_acquired = digest_probe_stack.enter_context(
                    _exclusive_file_lock(
                        _upload_digest_lock_path(upload_id),
                        blocking=False,
                    )
                )
                digest_lock_held = not digest_lock_acquired
            termination_confirmed = (
                digest_lock_held is not True
                and _digest_process_termination_confirmed(
                    digest_pid,
                    termination,
                )
            )
            now = _utc_now().isoformat()
            technical_detail = {
                "digestPid": digest_pid or None,
                "digestProcessIdentity": digest_identity,
                "digestLockHeld": digest_lock_held,
                "termination": termination,
                "abandonRequestedAt": now,
                "abandonRequestedBy": triggered_by.strip() or "anonymous",
            }
            if not termination_confirmed:
                state["completedAt"] = None
                state["failureDigest"] = {
                    "code": "RESOURCE_QUARANTINED",
                    "category": "resource",
                    "phase": status or "digesting",
                    "retryable": False,
                    "message": (
                        "已收到放弃请求，但无法确认 digest worker 已停止；"
                        "会话仍保持 active，禁止释放资源门禁。"
                    ),
                    "sourceFeedback": None,
                    "technicalDetail": technical_detail,
                    "nextAction": "contact_admin_verify_digest_process",
                }
                _persist_upload_session(state)
                raise HTTPException(
                    status_code=409,
                    detail={
                        "code": "RESOURCE_QUARANTINED",
                        "message": state["failureDigest"]["message"],
                        "technicalDetail": technical_detail,
                    },
                )
            state["status"] = "abandoned"
            state["completedAt"] = now
            state["abandonedAt"] = now
            state["abandonedBy"] = triggered_by.strip() or "anonymous"
            state["digestPid"] = None
            state["digestProcessIdentity"] = None
            state["failureDigest"] = {
                "code": "UPLOAD_SESSION_ABANDONED",
                "category": "lifecycle",
                "phase": status or "upload",
                "retryable": True,
                "message": "用户已明确放弃上传；candidate 与 active 均未修改。",
                "sourceFeedback": None,
                "technicalDetail": technical_detail,
                "nextAction": "start_new_upload",
            }
            _persist_upload_session(state)
            return _serialize_upload_session(state)


def _upload_jato_monthly_update_chunk_locked(
    *,
    upload_id: str,
    part_number: int,
    content: bytes,
    chunk_sha256: str,
    requested_by: str,
    requested_role: str,
) -> dict[str, Any]:
    state = _load_upload_session(upload_id)
    _require_upload_session_access(
        state,
        requested_by=requested_by,
        requested_role=requested_role,
    )
    status = str(state.get("status", "pending"))
    if status in {
        "assembling",
        "digesting",
        "ready",
        "invalid",
        "consumed",
        "abandoned",
        "expired",
    }:
        raise HTTPException(status_code=409, detail="上传会话已完成，不能继续写入分片。")

    total_chunks = int(state.get("totalChunks", 0))
    size_bytes = int(state.get("sizeBytes", 0))
    chunk_size = int(state.get("chunkSize", UPLOAD_CHUNK_SIZE_BYTES))
    if part_number < 1 or part_number > total_chunks:
        raise HTTPException(status_code=400, detail="分片序号超出范围。")

    expected_size = _expected_chunk_size(
        size_bytes=size_bytes,
        chunk_size=chunk_size,
        total_chunks=total_chunks,
        part_number=part_number,
    )
    if len(content) != expected_size:
        raise HTTPException(
            status_code=400,
            detail=f"分片大小不匹配，期望 {expected_size} 字节，实际 {len(content)} 字节。",
        )
    normalized_chunk_sha256 = _normalize_sha256(
        chunk_sha256,
        detail="分片 SHA-256 无效。",
    )
    actual_chunk_sha256 = _sha256_hex_for_bytes(content)
    if actual_chunk_sha256 != normalized_chunk_sha256:
        raise HTTPException(status_code=400, detail="分片内容校验失败，请重试。")

    chunk_dir = _upload_session_chunk_dir(upload_id)
    chunk_dir.mkdir(parents=True, exist_ok=True)
    chunk_path = chunk_dir / _chunk_file_name(part_number)
    existing_digests = state.get("chunkDigests")
    digest_map = existing_digests if isinstance(existing_digests, dict) else {}
    if (
        chunk_path.exists()
        and chunk_path.stat().st_size == expected_size
        and str(digest_map.get(str(part_number), "")) == normalized_chunk_sha256
    ):
        state["receivedChunks"] = _collect_uploaded_chunk_numbers(upload_id)
        state["uploadedBytes"] = _uploaded_chunk_bytes(upload_id)
        state["chunkDigests"] = digest_map
        state["status"] = "uploading"
        _persist_upload_session(state)
        return _serialize_upload_session(state)

    if chunk_path.exists():
        raise HTTPException(
            status_code=409,
            detail="该分片已存在但 SHA-256 不同，拒绝覆盖；请新建上传会话。",
        )
    temp_path = chunk_path.with_name(
        f".{chunk_path.name}.{os.getpid()}.{uuid4().hex}.tmp"
    )
    try:
        with temp_path.open("wb") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_path, chunk_path)
    finally:
        temp_path.unlink(missing_ok=True)
    digest_map[str(part_number)] = normalized_chunk_sha256
    state["chunkDigests"] = digest_map
    state["receivedChunks"] = _collect_uploaded_chunk_numbers(upload_id)
    state["uploadedBytes"] = _uploaded_chunk_bytes(upload_id)
    state["status"] = "uploading"
    _persist_upload_session(state)
    return _serialize_upload_session(state)


def upload_jato_monthly_update_chunk(
    *,
    upload_id: str,
    part_number: int,
    content: bytes,
    chunk_sha256: str,
    requested_by: str,
    requested_role: str,
) -> dict[str, Any]:
    """Serialize each session's read/write cycle to prevent lost chunk state."""
    with _exclusive_file_lock(_upload_state_lock_path(upload_id)) as acquired:
        if not acquired:  # blocking lock always acquires on supported platforms
            raise HTTPException(
                status_code=503,
                detail="上传分片状态锁暂不可用，请稍后重试。",
            )
        return _upload_jato_monthly_update_chunk_locked(
            upload_id=upload_id,
            part_number=part_number,
            content=content,
            chunk_sha256=chunk_sha256,
            requested_by=requested_by,
            requested_role=requested_role,
        )


def _upload_digest_lock_path(upload_id: str) -> Path:
    return _upload_session_dir(upload_id) / "digest.lock"


def _assemble_monthly_update_upload(state: dict[str, Any]) -> tuple[Path, str]:
    upload_id = str(state["uploadId"])
    filename = str(state.get("filename", "jato-update.xlsx"))
    assembled_path = _upload_session_assembled_path(upload_id, filename)
    expected_file_size = int(state.get("sizeBytes", 0))
    existing_sha = str(state.get("fileSha256") or "").strip().lower()
    if (
        assembled_path.exists()
        and assembled_path.stat().st_size == expected_file_size
        and re.fullmatch(r"[0-9a-f]{64}", existing_sha)
    ):
        return assembled_path, existing_sha

    total_chunks = int(state.get("totalChunks", 0))
    digest_payload = state.get("chunkDigests")
    digest_map = digest_payload if isinstance(digest_payload, dict) else {}
    assembled_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = assembled_path.with_name(
        f".assemble-{os.getpid()}-{uuid4().hex}.tmp"
    )
    file_hasher = hashlib.sha256()
    total_bytes_written = 0
    try:
        with temp_path.open("wb") as target:
            for part_number in range(1, total_chunks + 1):
                chunk_path = (
                    _upload_session_chunk_dir(upload_id)
                    / _chunk_file_name(part_number)
                )
                expected_digest = str(
                    digest_map.get(str(part_number), "")
                ).strip().lower()
                expected_chunk_size = _expected_chunk_size(
                    size_bytes=expected_file_size,
                    chunk_size=int(
                        state.get("chunkSize", UPLOAD_CHUNK_SIZE_BYTES)
                    ),
                    total_chunks=total_chunks,
                    part_number=part_number,
                )
                if chunk_path.stat().st_size != expected_chunk_size:
                    raise RuntimeError(
                        f"分片 {part_number} 大小已变化，请重新上传。"
                    )
                chunk_hasher = hashlib.sha256()
                chunk_bytes_written = 0
                with chunk_path.open("rb") as source:
                    while True:
                        block = source.read(UPLOAD_ASSEMBLY_BUFFER_BYTES)
                        if not block:
                            break
                        target.write(block)
                        file_hasher.update(block)
                        chunk_hasher.update(block)
                        chunk_bytes_written += len(block)
                        total_bytes_written += len(block)
                        if total_bytes_written > expected_file_size:
                            raise RuntimeError(
                                "组装内容超过声明大小，请重新上传。"
                            )
                if chunk_bytes_written != expected_chunk_size:
                    raise RuntimeError(
                        f"分片 {part_number} 大小已变化，请重新上传。"
                    )
                if expected_digest and chunk_hasher.hexdigest() != expected_digest:
                    raise RuntimeError(
                        f"分片 {part_number} SHA-256 校验失败，请重新上传。"
                    )
            target.flush()
            os.fsync(target.fileno())
        if temp_path.stat().st_size != expected_file_size:
            raise RuntimeError("组装后的文件大小校验失败，请重新上传。")
        os.replace(temp_path, assembled_path)
    finally:
        temp_path.unlink(missing_ok=True)
    return assembled_path, file_hasher.hexdigest()


def _active_country_latest_month(country: str) -> str | None:
    partition_root = _active_data_paths()["partition"]
    country_dir = partition_root / f"国家={quote(country, safe='')}"
    parquet_files = sorted(country_dir.rglob("*.parquet"))
    if not parquet_files:
        return None
    try:
        import pyarrow.parquet as pq

        schema_columns = [
            str(column).strip()
            for column in pq.read_schema(parquet_files[0]).names
        ]
        month_columns = _detect_month_columns(schema_columns)
        if not month_columns:
            return None
        frame = pd.read_parquet(country_dir, columns=month_columns)
        frame.columns = [str(column).strip() for column in frame.columns]
        return _latest_month_from_frame(frame)
    except Exception as exc:
        raise RuntimeError(f"无法读取 active 的 {country} 月份范围：{exc}") from exc


def _capture_active_digest_snapshot(
    countries: list[str],
) -> tuple[list[str], dict[str, str | None], str]:
    """Pin all active reads used by one digest behind the publication lock."""
    with _exclusive_file_lock(_active_bundle_lock_path()) as acquired:
        if not acquired:
            raise RuntimeError("active bundle 锁暂不可用，无法生成一致的 digest 快照。")
        version_before = _active_dataset_version()
        active_countries = _active_partition_country_names()
        active_country_keys = {
            country.casefold()
            for country in active_countries
        }
        latest = {
            country: _active_country_latest_month(country)
            for country in countries
            if country.casefold() in active_country_keys
        }
        version_after = _active_dataset_version()
        if version_before != version_after:
            raise RuntimeError(
                "读取 digest 所需 active 分区时版本发生变化；"
                "系统拒绝把旧月份范围绑定到新 active，请自动重试。"
            )
        return active_countries, latest, version_before


def _build_upload_ingest_digest(
    *,
    path: Path,
    file_sha256: str,
    size_bytes: int,
) -> dict[str, Any]:
    inspection = _inspect_upload_scope(path)
    countries = _ordered_distinct_strings(
        [str(country) for country in inspection.get("countries", [])]
    )
    (
        active_countries,
        active_latest_labels,
        active_dataset_version,
    ) = _capture_active_digest_snapshot(countries)
    blockers: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    def issue(
        code: str,
        message: str,
        *,
        issue_countries: list[str] | None = None,
        source_feedback: str,
    ) -> dict[str, Any]:
        return {
            "code": code,
            "message": message,
            "countries": issue_countries or [],
            "fields": [],
            "sourceFeedback": source_feedback,
        }

    inspection_issues = inspection.get("issues")
    if isinstance(inspection_issues, list):
        for inspection_issue in inspection_issues:
            if not isinstance(inspection_issue, dict):
                continue
            code = str(inspection_issue.get("code") or "").strip()
            message = str(inspection_issue.get("message") or "").strip()
            source_feedback = str(
                inspection_issue.get("sourceFeedback") or ""
            ).strip()
            if not code or not message or not source_feedback:
                continue
            blockers.append(
                {
                    "code": code,
                    "message": message,
                    "countries": _ordered_distinct_strings(
                        [
                            str(country)
                            for country in inspection_issue.get(
                                "countries",
                                [],
                            )
                        ]
                    ),
                    "fields": _ordered_distinct_strings(
                        [
                            str(field)
                            for field in inspection_issue.get("fields", [])
                        ]
                    ),
                    "sourceFeedback": source_feedback,
                }
            )

    route: str | None = None
    if not active_countries:
        blockers.append(
            issue(
                "ACTIVE_PARTITION_UNAVAILABLE",
                "active 国家分区不可用，无法安全判断上传范围。",
                source_feedback="请联系平台管理员恢复 active partition；不要把该文件改成全量后反复重传。",
            )
        )
    else:
        unknown_countries = sorted(set(countries) - set(active_countries))
        if unknown_countries:
            blockers.append(
                issue(
                    "UNKNOWN_COUNTRY",
                    "上传包含 active 中不存在的国家：" + "、".join(unknown_countries),
                    issue_countries=unknown_countries,
                    source_feedback=(
                        "请核对国家名称是否与 JATO active 完全一致；"
                        "若确为新增国家，请走新增国家流程，不能按普通月更处理。"
                    ),
                )
            )
        elif set(countries) == set(active_countries):
            route = "full_batch"
        elif set(countries) < set(active_countries):
            route = "single_country" if len(countries) == 1 else "partial_country"

    active_latest_months: dict[str, str | None] = {}
    advanced_countries: list[str] = []
    unchanged_countries: list[str] = []
    regressed_countries: list[str] = []
    country_latest_payload = inspection.get("countryLatestMonths")
    country_latest_months = (
        {
            str(country): (
                str(month) if month not in {None, ""} else None
            )
            for country, month in country_latest_payload.items()
        }
        if isinstance(country_latest_payload, dict)
        else {}
    )
    if route is not None:
        for country in countries:
            active_label = active_latest_labels.get(country)
            active_month = (
                _normalized_month_from_label(active_label)
                if active_label
                else None
            )
            upload_month = country_latest_months.get(country)
            active_latest_months[country] = active_month
            if active_month is None:
                blockers.append(
                    issue(
                        "ACTIVE_COUNTRY_MONTH_UNAVAILABLE",
                        f"无法识别 {country} 的 active 最新月份。",
                        issue_countries=[country],
                        source_feedback=f"请平台管理员检查 {country} active 分区，不要修改上传数据规避校验。",
                    )
                )
            elif upload_month is None:
                blockers.append(
                    issue(
                        "UPLOAD_COUNTRY_MONTH_EMPTY",
                        f"{country} 的月份列均为空。",
                        issue_countries=[country],
                        source_feedback=f"请洗数人员补齐 {country} 的真实最新月份列和销量，不要填 0 冒充缺失值。",
                    )
                )
            elif upload_month < active_month:
                regressed_countries.append(country)
            elif upload_month == active_month:
                unchanged_countries.append(country)
            else:
                advanced_countries.append(country)

        if regressed_countries:
            blockers.append(
                issue(
                    "COUNTRY_MONTH_REGRESSION",
                    "以下国家最新月份发生回退：" + "、".join(regressed_countries),
                    issue_countries=regressed_countries,
                    source_feedback="请重新导出这些国家截至目标月份的完整 washed 快照；不能用更旧月份覆盖 active。",
                )
            )
        if not advanced_countries and not regressed_countries:
            blockers.append(
                issue(
                    "NO_COUNTRY_ADVANCED",
                    "上传没有推进任何国家的最新月份，属于重复或同月文件。",
                    issue_countries=unchanged_countries,
                    source_feedback="无需重新上传同月文件；请提供至少一个国家晚于 active 的最新月份。",
                )
            )
        elif unchanged_countries:
            warnings.append(
                issue(
                    "MIXED_FRESHNESS",
                    "部分国家保持原月份，其余国家推进：" + "、".join(unchanged_countries),
                    issue_countries=unchanged_countries,
                    source_feedback="这是允许的 mixed freshness；请确认这些国家本月确实没有更新版本。",
                )
            )

    return {
        "schemaVersion": 1,
        "status": "invalid" if blockers else "ready",
        "fileSha256": file_sha256,
        "sizeBytes": size_bytes,
        "sheetName": inspection.get("sheetName"),
        "route": route,
        "candidateScope": {
            "single_country": "target_country_partition_only",
            "partial_country": "target_country_partitions_only",
            "full_batch": "full_candidate",
        }.get(route),
        "countries": countries,
        "countryLatestMonths": country_latest_months,
        "activeLatestMonths": active_latest_months,
        "latestMonth": inspection.get("latestMonth"),
        "dataRowCount": int(inspection.get("dataRowCount", 0) or 0),
        "advancedCountries": advanced_countries,
        "unchangedCountries": unchanged_countries,
        "regressedCountries": regressed_countries,
        "activeDatasetVersion": (
            active_dataset_version
            if active_countries
            else None
        ),
        "blockers": blockers,
        "warnings": warnings,
    }


def run_jato_monthly_update_upload_digest(upload_id: str) -> dict[str, Any]:
    """Assemble and inspect one upload outside the FastAPI request process."""
    attempt_id = str(
        os.getenv("APP_JATO_DIGEST_ATTEMPT_ID", "")
    ).strip() or None
    with _exclusive_file_lock(
        _upload_digest_lock_path(upload_id),
        blocking=False,
    ) as acquired:
        if not acquired:
            return _serialize_upload_session(_load_upload_session(upload_id))
        state = _load_upload_session(upload_id)
        if not _digest_attempt_matches(state, attempt_id):
            return _serialize_upload_session(state)
        if str(state.get("status", "")) in {
            "ready",
            "invalid",
            "consumed",
            "abandoned",
            "expired",
        }:
            return _serialize_upload_session(state)
        try:
            existing_assembled_path = _persisted_upload_session_assembled_path(
                state
            )
            expected_sha256 = str(
                state.get("fileSha256") or ""
            ).strip().lower()
            if (
                existing_assembled_path is not None
                and existing_assembled_path.is_file()
                and re.fullmatch(r"[0-9a-f]{64}", expected_sha256)
            ):
                assembled_path = existing_assembled_path
                file_sha256 = _sha256_hex_for_path(assembled_path)
                if file_sha256 != expected_sha256:
                    raise RuntimeError(
                        "已组装上传文件的 SHA-256 发生变化，请重新上传。"
                    )
            else:
                assembled_path, file_sha256 = (
                    _assemble_monthly_update_upload(state)
                )
            with _exclusive_file_lock(
                _upload_state_lock_path(upload_id)
            ) as state_acquired:
                if not state_acquired:
                    raise RuntimeError("digest 状态锁暂不可用。")
                state = _load_upload_session(upload_id)
                if not _digest_attempt_matches(state, attempt_id):
                    return _serialize_upload_session(state)
                if str(state.get("status") or "") in {
                    "abandoned",
                    "expired",
                }:
                    return _serialize_upload_session(state)
                state["status"] = "digesting"
                state["digestWorkerPid"] = os.getpid()
                if attempt_id:
                    attempt = state.get("digestAttempt")
                    if isinstance(attempt, dict):
                        attempt["status"] = "digesting"
                        attempt["workerPid"] = os.getpid()
                        state["digestAttempt"] = attempt
                else:
                    state["digestPid"] = os.getpid()
                    state["digestProcessIdentity"] = _read_process_identity(
                        os.getpid()
                    )
                state["digestLaunchedAt"] = (
                    state.get("digestLaunchedAt") or _utc_now().isoformat()
                )
                state["assembledPath"] = _relative_to_project(assembled_path)
                state["fileSha256"] = file_sha256
                state["uploadedBytes"] = assembled_path.stat().st_size
                _persist_upload_session(state)
            digest = _build_upload_ingest_digest(
                path=assembled_path,
                file_sha256=file_sha256,
                size_bytes=assembled_path.stat().st_size,
            )
            with _exclusive_file_lock(
                _upload_state_lock_path(upload_id)
            ) as state_acquired:
                if not state_acquired:
                    raise RuntimeError("digest 结果状态锁暂不可用。")
                latest = _load_upload_session(upload_id)
                if not _digest_attempt_matches(latest, attempt_id):
                    return _serialize_upload_session(latest)
                latest_failure = latest.get("failureDigest")
                if str(latest.get("status") or "") in {
                    "abandoned",
                    "expired",
                }:
                    return _serialize_upload_session(latest)
                if (
                    str(latest.get("status") or "") == "invalid"
                    and isinstance(latest_failure, dict)
                    and str(latest_failure.get("code") or "")
                    in DIGEST_RETRYABLE_FAILURE_CODES
                ):
                    return _serialize_upload_session(latest)
                state = latest
                state["ingestDigest"] = digest
                state["failureDigest"] = None
                state["status"] = str(digest["status"])
                state["completedAt"] = _utc_now().isoformat()
                state["digestWorkerPid"] = None
                if attempt_id:
                    attempt = state.get("digestAttempt")
                    if isinstance(attempt, dict):
                        attempt["status"] = "worker_finished"
                        attempt["workerFinishedAt"] = _utc_now().isoformat()
                        state["digestAttempt"] = attempt
                else:
                    state["digestPid"] = None
                    state["digestProcessIdentity"] = None
                _persist_upload_session(state)
            shutil.rmtree(_upload_session_chunk_dir(upload_id), ignore_errors=True)
        except MemoryError:
            raise
        except Exception as exc:
            with _exclusive_file_lock(
                _upload_state_lock_path(upload_id)
            ) as state_acquired:
                if not state_acquired:
                    raise
                latest = _load_upload_session(upload_id)
                if not _digest_attempt_matches(latest, attempt_id):
                    return _serialize_upload_session(latest)
                latest_failure = latest.get("failureDigest")
                if str(latest.get("status") or "") in {
                    "abandoned",
                    "expired",
                }:
                    return _serialize_upload_session(latest)
                if (
                    str(latest.get("status") or "") == "invalid"
                    and isinstance(latest_failure, dict)
                    and str(latest_failure.get("code") or "")
                    in DIGEST_RETRYABLE_FAILURE_CODES
                ):
                    return _serialize_upload_session(latest)
                state = latest
                state["status"] = "invalid"
                state["completedAt"] = _utc_now().isoformat()
                state["digestWorkerPid"] = None
                if attempt_id:
                    attempt = state.get("digestAttempt")
                    if isinstance(attempt, dict):
                        attempt["status"] = "worker_finished"
                        attempt["workerFinishedAt"] = _utc_now().isoformat()
                        state["digestAttempt"] = attempt
                else:
                    state["digestPid"] = None
                    state["digestProcessIdentity"] = None
                state["failureDigest"] = {
                    "code": "UPLOAD_DIGEST_FAILED",
                    "category": "input_validation",
                    "phase": "digesting",
                    "retryable": False,
                    "message": f"上传文件 digest 失败：{exc}",
                    "sourceFeedback": "请确认工作表名为 Data Export、国家列存在、月份列有真实数据；若文件可正常打开仍失败，请把此错误交给平台管理员。",
                    "technicalDetail": traceback.format_exc(limit=8),
                    "nextAction": "fix_source_or_contact_admin",
                }
                _persist_upload_session(state)
        return _serialize_upload_session(state)


def _launch_upload_digest_process(upload_id: str) -> int:
    if not MONTHLY_WORKER_SCRIPT_PATH.exists():
        raise RuntimeError("JATO monthly worker 脚本不存在，无法启动 digest。")
    state = _load_upload_session(upload_id)
    attempt = state.get("digestAttempt")
    if not isinstance(attempt, dict):
        raise RuntimeError("JATO digest attempt 状态缺失，拒绝无审计启动。")
    attempt_id = str(attempt.get("attemptId") or "").strip()
    log_path = _upload_digest_attempt_artifact_path(state, "logPath")
    receipt_path = _upload_digest_attempt_artifact_path(
        state,
        "receiptPath",
    )
    if not attempt_id or log_path is None or receipt_path is None:
        raise RuntimeError("JATO digest attempt 路径无效，拒绝启动。")
    env = dict(os.environ)
    env["APP_JATO_MONTHLY_WORKER_MEMORY_LIMIT_BYTES"] = "0"
    env.setdefault(
        "APP_JATO_DIGEST_RSS_WARNING_BYTES",
        str(1024 * 1024 * 1024),
    )
    env.setdefault(
        "APP_JATO_DIGEST_RSS_LIMIT_BYTES",
        str(1536 * 1024 * 1024),
    )
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("ab", buffering=0) as output:
        process = subprocess.Popen(
            [
                sys.executable,
                str(MONTHLY_WORKER_SCRIPT_PATH),
                "--supervise-digest-upload",
                upload_id,
                "--attempt-id",
                attempt_id,
                "--attempt-log",
                str(log_path),
                "--attempt-receipt",
                str(receipt_path),
            ],
            cwd=PROJECT_ROOT,
            stdin=subprocess.DEVNULL,
            stdout=output,
            stderr=subprocess.STDOUT,
            env=env,
            start_new_session=True,
            close_fds=True,
        )
    return int(process.pid)


def _start_upload_digest_locked(state: dict[str, Any]) -> dict[str, Any]:
    """Persist one digest launch while the caller holds the session state lock."""
    upload_id = str(state["uploadId"])
    state["status"] = "assembling"
    state["ingestDigest"] = None
    state["failureDigest"] = None
    state["completedAt"] = None
    state["digestPid"] = None
    state["digestWorkerPid"] = None
    state["digestProcessIdentity"] = None
    state["digestLaunchedAt"] = _utc_now().isoformat()
    attempt_number = int(state.get("digestAttempts") or 0) + 1
    state["digestAttempts"] = attempt_number
    state["digestAttempt"] = _new_upload_digest_attempt(
        upload_id=upload_id,
        attempt_number=attempt_number,
    )
    _persist_upload_session(state)
    try:
        digest_pid = _launch_upload_digest_process(upload_id)
        latest = _load_upload_session(upload_id)
        if str(latest.get("status") or "") in {"assembling", "digesting"}:
            latest["digestPid"] = digest_pid
            latest["digestProcessIdentity"] = _read_process_identity(
                digest_pid
            )
            attempt = latest.get("digestAttempt")
            if isinstance(attempt, dict):
                attempt["status"] = "running"
                attempt["supervisorPid"] = digest_pid
                attempt["supervisorIdentity"] = latest[
                    "digestProcessIdentity"
                ]
                latest["digestAttempt"] = attempt
            _persist_upload_session(latest)
        return latest
    except Exception as exc:
        state = _load_upload_session(upload_id)
        state["status"] = "invalid"
        state["completedAt"] = _utc_now().isoformat()
        state["failureDigest"] = {
            "code": "DIGEST_WORKER_UNAVAILABLE",
            "category": "platform",
            "phase": "assembling",
            "retryable": True,
            "message": str(exc),
            "sourceFeedback": None,
            "technicalDetail": {
                "traceback": traceback.format_exc(limit=4),
                "digestAttempt": state.get("digestAttempt"),
                "logTail": _read_digest_attempt_log_tail(state),
            },
            "nextAction": "contact_admin",
        }
        state["digestPid"] = None
        state["digestWorkerPid"] = None
        state["digestProcessIdentity"] = None
        _persist_upload_session(state)
        return state


def _complete_jato_monthly_update_upload_locked(
    *,
    upload_id: str,
    requested_by: str,
    requested_role: str,
) -> dict[str, Any]:
    state = _load_upload_session(upload_id)
    _require_upload_session_access(
        state,
        requested_by=requested_by,
        requested_role=requested_role,
    )
    status = str(state.get("status", "pending"))
    if status in {
        "assembling",
        "digesting",
        "ready",
        "invalid",
        "consumed",
        "abandoned",
        "expired",
    }:
        return _serialize_upload_session(state)
    received_chunks, uploaded_bytes = _validate_uploaded_chunks_complete(state)

    state["receivedChunks"] = received_chunks
    state["uploadedBytes"] = uploaded_bytes
    state = _start_upload_digest_locked(state)
    return _serialize_upload_session(state)


def complete_jato_monthly_update_upload(
    *,
    upload_id: str,
    requested_by: str,
    requested_role: str,
) -> dict[str, Any]:
    """Make complete idempotent even when two Web workers receive the replay."""
    with _exclusive_file_lock(_upload_state_lock_path(upload_id)) as acquired:
        if not acquired:  # blocking lock always acquires on supported platforms
            raise HTTPException(
                status_code=503,
                detail="上传完成状态锁暂不可用，请稍后重试。",
            )
        return _complete_jato_monthly_update_upload_locked(
            upload_id=upload_id,
            requested_by=requested_by,
            requested_role=requested_role,
        )


def _retry_jato_monthly_update_upload_digest_locked(
    *,
    upload_id: str,
    requested_by: str,
    requested_role: str,
) -> dict[str, Any]:
    state = _load_upload_session(upload_id)
    _require_upload_session_access(
        state,
        requested_by=requested_by,
        requested_role=requested_role,
    )
    status = str(state.get("status") or "")
    if status in {"assembling", "digesting", "ready", "consumed"}:
        return _serialize_upload_session(state)
    failure = state.get("failureDigest")
    failure_code = (
        str(failure.get("code") or "")
        if isinstance(failure, dict)
        else ""
    )
    retryable_failure = (
        isinstance(failure, dict)
        and failure.get("retryable") is True
        and failure_code in DIGEST_RETRYABLE_FAILURE_CODES
    )
    if status != "invalid" or not retryable_failure:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "DIGEST_RETRY_NOT_ALLOWED",
                "message": (
                    "仅允许恢复已确认停止且可重试的 digest worker；"
                    "输入校验失败必须修正源文件后重新上传。"
                ),
                "failureCode": failure_code or None,
            },
        )

    technical_detail = (
        failure.get("technicalDetail")
        if isinstance(failure, dict)
        else None
    )
    old_digest_pid = 0
    expected_process_identity: dict[str, Any] | None = None
    expected_attempt_id: str | None = None
    if isinstance(technical_detail, dict):
        try:
            old_digest_pid = int(technical_detail.get("digestPid") or 0)
        except (TypeError, ValueError):
            old_digest_pid = 0
        raw_identity = technical_detail.get("digestProcessIdentity")
        expected_process_identity = (
            raw_identity if isinstance(raw_identity, dict) else None
        )
        expected_attempt_id = str(
            technical_detail.get("digestAttemptId") or ""
        ).strip() or None
    if old_digest_pid > 0 and _process_exists(old_digest_pid):
        if expected_process_identity is None:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "DIGEST_RETRY_WORKER_IDENTITY_UNKNOWN",
                    "message": (
                        "原 digest PID 仍存活，但失败记录缺少进程身份；"
                        "无法排除旧 worker 仍在运行，请管理员确认。"
                    ),
                    "digestPid": old_digest_pid,
                },
            )
        identity_matches, current_process_identity = _process_identity_matches(
            pid=old_digest_pid,
            expected_identity=expected_process_identity,
            required_command_tokens=(
                (
                    "--supervise-digest-upload",
                    upload_id,
                    expected_attempt_id,
                )
                if expected_attempt_id
                else ("--digest-upload", upload_id)
            ),
        )
        if identity_matches:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "DIGEST_RETRY_WORKER_STILL_RUNNING",
                    "message": (
                        "原 digest worker 仍在运行，系统拒绝叠加第二个 worker；"
                        "请等待或由管理员先安全终止旧进程。"
                    ),
                    "digestPid": old_digest_pid,
                },
            )
        if current_process_identity is None:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "DIGEST_RETRY_WORKER_IDENTITY_UNKNOWN",
                    "message": (
                        "原 digest PID 仍存活，但当前进程身份不可读；"
                        "无法安全判断 PID 是否复用，请管理员确认。"
                    ),
                    "digestPid": old_digest_pid,
                },
            )

    _require_no_running_monthly_update_jobs()
    _require_no_active_upload_sessions(action="恢复上传 digest")

    expected_size = int(state.get("sizeBytes") or 0)
    assembled_path = _persisted_upload_session_assembled_path(state)
    expected_sha256 = str(state.get("fileSha256") or "").strip().lower()
    if expected_size <= 0:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "DIGEST_RETRY_SOURCE_INVALID",
                "message": "上传文件大小状态无效，请重新上传。",
            },
        )
    if assembled_path is not None and assembled_path.is_file():
        actual_size = assembled_path.stat().st_size
        if (
            actual_size != expected_size
            or not re.fullmatch(r"[0-9a-f]{64}", expected_sha256)
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "DIGEST_RETRY_SOURCE_INVALID",
                    "message": "已组装文件大小或 SHA-256 状态无效，请重新上传。",
                },
            )
        if _sha256_hex_for_path(assembled_path) != expected_sha256:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "DIGEST_RETRY_SOURCE_CHANGED",
                    "message": "已组装文件 SHA-256 已变化，请重新上传。",
                },
            )
        state["uploadedBytes"] = actual_size
    else:
        try:
            received_chunks, uploaded_bytes = (
                _validate_uploaded_chunks_complete(state)
            )
        except HTTPException as exc:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "DIGEST_RETRY_SOURCE_MISSING",
                    "message": (
                        "已组装文件和完整上传分片均不存在，"
                        "无法恢复 digest，请重新上传。"
                    ),
                },
            ) from exc
        state["receivedChunks"] = received_chunks
        state["uploadedBytes"] = uploaded_bytes
        state["assembledPath"] = None
        state["fileSha256"] = None
    state = _start_upload_digest_locked(state)
    return _serialize_upload_session(state)


def retry_jato_monthly_update_upload_digest(
    *,
    upload_id: str,
    requested_by: str,
    requested_role: str,
) -> dict[str, Any]:
    """Restart only a lost/timed-out digest inside the global resource window."""
    with _exclusive_file_lock(
        _maintenance_coordination_lock_path(),
        blocking=False,
    ) as coordinated:
        if not coordinated:
            raise HTTPException(
                status_code=409,
                detail=(
                    "JATO 清理或 baseline 操作正在准备中，"
                    "请稍后再恢复上传 digest。"
                ),
            )
        with _exclusive_file_lock(_upload_initiate_lock_path()) as global_upload:
            if not global_upload:
                raise HTTPException(
                    status_code=503,
                    detail="上传会话全局锁暂不可用，请稍后重试。",
                )
            with _exclusive_file_lock(_upload_state_lock_path(upload_id)) as acquired:
                if not acquired:
                    raise HTTPException(
                        status_code=503,
                        detail="上传 digest 状态锁暂不可用，请稍后重试。",
                    )
                return _retry_jato_monthly_update_upload_digest_locked(
                    upload_id=upload_id,
                    requested_by=requested_by,
                    requested_role=requested_role,
                )


def _parse_month_from_filename(filename: str) -> str | None:
    """Try to extract YYYY-MM from filename. Returns None if not found."""
    import re

    name = Path(filename).stem
    patterns = [
        r"(20\d{2})[._-](\d{1,2})",     # 2026.4, 2026-04, 2026_3
        r"(20\d{2})\s*[年-]\s*(\d{1,2})",  # 2026年4, 2026-3
    ]
    for pat in patterns:
        m = re.search(pat, name)
        if m:
            year = int(m.group(1))
            month_num = int(m.group(2))
            if 1 <= month_num <= 12 and 2020 <= year <= 2030:
                return f"{year}-{month_num:02d}"
    return None


def _build_ingestion_key(ingest_digest: dict[str, Any]) -> str:
    payload = {
        "fileSha256": str(ingest_digest.get("fileSha256") or ""),
        "route": str(ingest_digest.get("route") or ""),
        "activeDatasetVersion": str(
            ingest_digest.get("activeDatasetVersion") or ""
        ),
        "countryLatestMonths": {
            str(country): (
                None if month in {None, ""} else str(month)
            )
            for country, month in sorted(
                (
                    ingest_digest.get("countryLatestMonths")
                    if isinstance(
                        ingest_digest.get("countryLatestMonths"),
                        dict,
                    )
                    else {}
                ).items()
            )
        },
    }
    return hashlib.sha256(
        json.dumps(
            payload,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()


def _normalize_recovery_key(recovery_key: str) -> str:
    normalized = str(recovery_key or "").strip()
    if not RECOVERY_KEY_PATTERN.fullmatch(normalized):
        raise HTTPException(
            status_code=400,
            detail={
                "code": "INVALID_RECOVERY_KEY",
                "message": (
                    "recoveryKey 必须为 8-128 位字母、数字或 ._:- 组成的"
                    "幂等键。"
                ),
            },
        )
    return normalized


def _find_recovery_job_for_source(
    *,
    source_job_id: str,
    recovery_key: str,
) -> dict[str, Any] | None:
    source_matches = [
        payload
        for payload in _list_job_state_payloads()
        if str(payload.get("recoveryOfJobId") or "") == source_job_id
    ]
    matching_key = [
        payload
        for payload in source_matches
        if str(payload.get("recoveryKey") or "") == recovery_key
    ]
    if len(matching_key) == 1:
        return matching_key[0]
    if len(matching_key) > 1:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "RECOVERY_STATE_AMBIGUOUS",
                "message": "同一 recoveryKey 已对应多个恢复任务，拒绝自动选择。",
                "jobIds": [
                    str(payload.get("jobId") or "")
                    for payload in matching_key
                ],
            },
        )
    if source_matches:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "RECOVERY_ALREADY_CREATED",
                "message": "该失败任务已创建过另一个 recovery attempt。",
                "jobIds": [
                    str(payload.get("jobId") or "")
                    for payload in source_matches
                ],
            },
        )
    return None


def _consumed_upload_session_for_job(source_job_id: str) -> dict[str, Any]:
    matches = [
        payload
        for payload in _iter_upload_session_payloads()
        if (
            str(payload.get("status") or "") == "consumed"
            and str(payload.get("consumedJobId") or "") == source_job_id
        )
    ]
    if len(matches) != 1:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "RECOVERY_SOURCE_UPLOAD_NOT_UNIQUE",
                "message": (
                    "恢复要求唯一 consumed upload 绑定原失败任务；"
                    "当前证据缺失或不唯一。"
                ),
                "matchCount": len(matches),
            },
        )
    return matches[0]


def _find_existing_job_for_ingestion_key(
    ingestion_key: str,
) -> dict[str, Any] | None:
    for payload in sorted(
        _list_job_state_payloads(),
        key=lambda item: str(item.get("createdAt") or ""),
        reverse=True,
    ):
        if str(payload.get("ingestionKey") or "") != ingestion_key:
            continue
        if str(payload.get("status") or "") in {
            "queued",
            "running",
            "success",
            "duplicate",
        }:
            return payload
    return None


def _ensure_upload_digest_matches_current_active(
    *,
    upload_id: str,
    state: dict[str, Any],
    state_lock_held: bool = False,
) -> None:
    ingest_digest = state.get("ingestDigest")
    if not isinstance(ingest_digest, dict):
        return
    digest_active_version = str(
        ingest_digest.get("activeDatasetVersion") or ""
    )
    current_active_version = _active_dataset_version()
    if digest_active_version == current_active_version:
        return

    with ExitStack() as state_lock_stack:
        acquired = True
        if not state_lock_held:
            acquired = state_lock_stack.enter_context(
                _exclusive_file_lock(_upload_state_lock_path(upload_id))
            )
        if not acquired:
            raise HTTPException(
                status_code=503,
                detail="上传 digest 状态锁暂不可用，请稍后重试。",
            )
        latest = state if state_lock_held else _load_upload_session(upload_id)
        latest_digest = latest.get("ingestDigest")
        latest_digest_active_version = (
            str(latest_digest.get("activeDatasetVersion") or "")
            if isinstance(latest_digest, dict)
            else ""
        )
        current_active_version = _active_dataset_version()
        if latest_digest_active_version == current_active_version:
            return
        latest_status = str(latest.get("status") or "")
        if latest_status == "consumed":
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "STALE_UPLOAD_DIGEST",
                    "message": (
                        "该上传会话已创建过旧 active 版本的任务；当前 active "
                        "已变化，请重新点击上传同一文件以生成新 digest。"
                    ),
                    "digestActiveDatasetVersion": (
                        latest_digest_active_version or None
                    ),
                    "currentActiveDatasetVersion": (
                        current_active_version or None
                    ),
                    "startNewUploadRequired": True,
                },
            )
        if latest_status in {"assembling", "digesting"}:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "STALE_UPLOAD_DIGEST_REFRESHING",
                    "message": (
                        "active 在上传后发生变化，系统正在用同一文件自动刷新 "
                        "digest；无需重新上传。"
                    ),
                    "startNewUploadRequired": False,
                },
            )
        if latest_status != "ready":
            raise HTTPException(
                status_code=409,
                detail="上传 digest 状态已变化，请刷新后重试。",
            )

        refreshed = _start_upload_digest_locked(latest)
        refreshed_failure = refreshed.get("failureDigest")
        if (
            str(refreshed.get("status") or "") == "invalid"
            and isinstance(refreshed_failure, dict)
            and str(refreshed_failure.get("code") or "")
            == "DIGEST_WORKER_UNAVAILABLE"
        ):
            raise HTTPException(
                status_code=503,
                detail={
                    "code": "DIGEST_WORKER_UNAVAILABLE",
                    "message": str(
                        refreshed_failure.get("message")
                        or "digest worker 无法启动。"
                    ),
                },
            )
    raise HTTPException(
        status_code=409,
        detail={
            "code": "STALE_UPLOAD_DIGEST_REFRESHING",
            "message": (
                "active 在上传后发生变化，系统已使用同一文件自动刷新 digest；"
                "前端将在刷新完成后继续创建任务。"
            ),
            "digestActiveDatasetVersion": (
                digest_active_version or None
            ),
            "currentActiveDatasetVersion": (
                current_active_version or None
            ),
            "startNewUploadRequired": False,
        },
    )


def _queue_monthly_update_job_from_stored_upload(
    *,
    job_id: str,
    triggered_by: str,
    upload_filename: str,
    stored_upload_path: Path,
    month: str,
    file_sha256: str | None = None,
    ingest_digest: dict[str, Any] | None = None,
    ingestion_key: str | None = None,
    recovery_of_job_id: str | None = None,
    recovery_key: str | None = None,
    recovery_source: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if stored_upload_path.stat().st_size <= 0:
        raise HTTPException(status_code=400, detail="上传文件为空，无法启动月更任务。")

    normalized_month = _normalize_month(month)
    batch_id = _allocate_batch_id(normalized_month)

    now = _utc_now().isoformat()
    state: dict[str, Any] = {
        "jobId": job_id,
        "month": normalized_month,
        "batchId": batch_id,
        "status": "queued",
        "phase": "queued",
        "triggeredBy": triggered_by.strip() or "anonymous",
        "createdAt": now,
        "updatedAt": now,
        "startedAt": None,
        "finishedAt": None,
        "error": None,
        "ingestionKey": ingestion_key,
        "ingestDigest": ingest_digest,
        "activeBaseFingerprint": (
            ingest_digest.get("activeDatasetVersion")
            if isinstance(ingest_digest, dict)
            else None
        ),
        "upload": {
            "originalFilename": upload_filename,
            "storedPath": _relative_to_project(stored_upload_path),
            "sizeBytes": stored_upload_path.stat().st_size,
            "sha256": file_sha256,
        },
        "plan": None,
        "artifacts": {
            "jobDir": _relative_to_project(_job_dir(job_id)),
            "logPath": _relative_to_project(_job_log_path(job_id)),
        },
        "summaries": {},
        "logPath": _relative_to_project(_job_log_path(job_id)),
    }
    if recovery_of_job_id is not None or recovery_key is not None:
        state["recoveryOfJobId"] = recovery_of_job_id
        state["recoveryKey"] = recovery_key
        state["recoverySource"] = recovery_source
    if isinstance(ingest_digest, dict):
        countries = _ordered_distinct_strings(
            [str(country) for country in ingest_digest.get("countries", [])]
        )
        route = str(ingest_digest.get("route") or "")
        state["jobType"] = {
            "single_country": "single_country",
            "partial_country": "partial_country",
            "full_batch": "batch",
        }.get(route)
        state["countryScope"] = countries
        state["country"] = countries[0] if len(countries) == 1 else None
        state["uploadInspection"] = {
            "sheetName": ingest_digest.get("sheetName"),
            "countries": countries,
            "countryLatestMonths": ingest_digest.get("countryLatestMonths"),
            "latestMonth": ingest_digest.get("latestMonth"),
            "dataRowCount": ingest_digest.get("dataRowCount"),
        }
    _persist_job_state(state)
    _append_log(
        _job_log_path(job_id),
        f"[{now}] queued monthly update batch {batch_id} for month {normalized_month}",
    )
    _launch_job_thread(job_id)
    return _serialize_job_state(state, include_log_tail=False)

# Large file guard for direct multipart upload (not chunked)
_MAX_DIRECT_UPLOAD_BYTES = 50 * 1024 * 1024  # 50 MB


def _run_job(job_id: str) -> None:
    state = _load_job_state(job_id)
    if str(state.get("status") or "") == "cancelled":
        _RUNNING_THREADS.pop(job_id, None)
        return
    log_path = _job_log_path(job_id)
    upload_payload = state.get("upload")
    if not isinstance(upload_payload, dict):
        raise RuntimeError("任务缺少 upload 信息")
    stored_path_value = upload_payload.get("storedPath")
    if not stored_path_value:
        raise RuntimeError("任务缺少 upload storedPath")
    stored_upload_path = PROJECT_ROOT / str(stored_path_value)

    state["status"] = "running"
    state["startedAt"] = _utc_now().isoformat()

    try:
        ingest_digest = state.get("ingestDigest")
        inspection = state.get("uploadInspection")
        if not isinstance(ingest_digest, dict):
            file_sha256 = str(upload_payload.get("sha256") or "").strip().lower()
            if not re.fullmatch(r"[0-9a-f]{64}", file_sha256):
                file_sha256 = _sha256_hex_for_path(stored_upload_path)
                upload_payload["sha256"] = file_sha256
                state["upload"] = upload_payload
            ingest_digest = _build_upload_ingest_digest(
                path=stored_upload_path,
                file_sha256=file_sha256,
                size_bytes=stored_upload_path.stat().st_size,
            )
            if ingest_digest.get("blockers"):
                raise RuntimeError(
                    "上传 digest 未通过："
                    + "；".join(
                        str(item.get("message") or "")
                        for item in ingest_digest.get("blockers", [])
                        if isinstance(item, dict)
                    )
                )
            state["ingestDigest"] = ingest_digest
            state["activeBaseFingerprint"] = ingest_digest.get(
                "activeDatasetVersion"
            )
            inspection = {
                "sheetName": ingest_digest.get("sheetName"),
                "countries": ingest_digest.get("countries"),
                "countryLatestMonths": ingest_digest.get("countryLatestMonths"),
                "latestMonth": ingest_digest.get("latestMonth"),
                "dataRowCount": ingest_digest.get("dataRowCount"),
            }
            state["uploadInspection"] = inspection
        route = str(ingest_digest.get("route") or "")
        if not isinstance(inspection, dict):
            raise RuntimeError("任务缺少已验证的上传范围 digest。")
    except Exception as exc:
        state["status"] = "failed"
        state["phase"] = "failed"
        state["finishedAt"] = _utc_now().isoformat()
        state["error"] = f"上传范围识别失败：{exc}"
        state["failureDigest"] = _failure_digest_from_exception(
            phase="digesting",
            exc=exc,
        )
        _persist_job_state(state)
        _write_jato_etl_pipeline_status(state)
        _append_log(log_path, f"[{_utc_now().isoformat()}] {state['error']}")
        _RUNNING_THREADS.pop(job_id, None)
        return

    if route in {"single_country", "partial_country"}:
        countries = [
            str(country).strip()
            for country in inspection["countries"]
            if str(country).strip()
        ]
        detected_month = str(inspection["latestMonth"])
        is_single_country = len(countries) == 1
        state["jobType"] = "single_country" if is_single_country else "partial_country"
        state["country"] = countries[0] if is_single_country else None
        state["countryScope"] = countries
        state["month"] = detected_month
        state["batchId"] = (
            f"{detected_month}-{countries[0]}-single"
            if is_single_country
            else f"{detected_month}-partial-{len(countries)}c"
        )
        state["uploadInspection"] = inspection
        _persist_job_state(state)
        _append_log(
            log_path,
            (
                f"[{_utc_now().isoformat()}] 自动识别部分国家上传："
                f"countries={','.join(countries)}, month={detected_month}；"
                "跳过全量 baseline Raw Compare，转入目标分区 Review 路径。"
            ),
        )
        _run_country_partition_job(job_id)
        return
    if route != "full_batch":
        state["status"] = "failed"
        state["phase"] = "failed"
        state["finishedAt"] = _utc_now().isoformat()
        state["error"] = f"不支持或无法确认的上传范围 route={route or '-'}。"
        state["failureDigest"] = _failure_digest_from_exception(
            phase="digesting",
            exc=RuntimeError(state["error"]),
        )
        _persist_job_state(state)
        _write_jato_etl_pipeline_status(state)
        return

    batch_id = str(state.get("batchId") or "")
    if not batch_id:
        raise RuntimeError("任务缺少批次标识")

    try:
        state["phase"] = "preparing"
        _persist_job_state(state)
        artifacts = state.get("artifacts")
        baseline_relative_path = (
            artifacts.get("baselinePath")
            if isinstance(artifacts, dict)
            else None
        )
        baseline_source = (
            str(artifacts.get("baselineSource"))
            if isinstance(artifacts, dict) and artifacts.get("baselineSource")
            else None
        )
        baseline_path = _project_path(str(baseline_relative_path)) if baseline_relative_path else None
        if baseline_path is None:
            baseline_path, baseline_source = _require_latest_baseline()
            state["artifacts"] = artifacts if isinstance(artifacts, dict) else {}
            state["artifacts"]["baselinePath"] = _relative_to_project(baseline_path)
            state["artifacts"]["baselineSource"] = baseline_source
            _persist_job_state(state)
        elif not baseline_path.exists():
            raise RuntimeError(
                "任务绑定的 baseline 不存在："
                f"{_relative_to_project(baseline_path) or str(baseline_path)}"
            )

        if baseline_source == "archive":
            _append_log(
                log_path,
                (
                    f"[{_utc_now().isoformat()}] active baseline 缺失，"
                    "使用 archive baseline: "
                    f"{_relative_to_project(baseline_path) or str(baseline_path)}"
                ),
            )

        if not PREPARE_SCRIPT_PATH.exists():
            raise RuntimeError(f"找不到脚本: {_relative_to_project(PREPARE_SCRIPT_PATH)}")

        _ensure_job_not_cancelled(job_id)
        prepare_args = [
            sys.executable,
            str(PREPARE_SCRIPT_PATH),
            "--month",
            str(state["month"]),
            "--batch-id",
            batch_id,
            "--baseline",
            str(baseline_path),
            "--patch",
            str(stored_upload_path),
        ]
        _run_logged_command(
            label="Prepare monthly update",
            args=prepare_args,
            log_path=log_path,
        )

        plan_path = PROJECT_ROOT / "01_RAW_DATA" / "patches" / batch_id / "monthly_update_plan.md"
        if not plan_path.exists():
            raise RuntimeError("prepare 完成后未生成 monthly_update_plan.md")

        parsed_plan = _parse_plan_markdown(plan_path)
        refresh_command, supplement_parquet_path = _inject_refresh_supplement_arg(
            str(parsed_plan["refreshCommand"])
        )
        parsed_plan["refreshCommand"] = refresh_command
        if supplement_parquet_path:
            parsed_plan["supplementParquetPath"] = supplement_parquet_path
        state["plan"] = {
            "path": _relative_to_project(plan_path),
            "batchId": parsed_plan.get("batchId") or batch_id,
            "compareId": parsed_plan.get("compareId"),
            "compareCommand": parsed_plan.get("compareCommand"),
            "refreshCommand": parsed_plan.get("refreshCommand"),
        }
        artifacts = state.get("artifacts")
        state["artifacts"] = artifacts if isinstance(artifacts, dict) else {}
        state["artifacts"].update(
            {
                "jobDir": _relative_to_project(_job_dir(job_id)),
                "logPath": _relative_to_project(log_path),
                "baselinePath": parsed_plan.get("baselinePath"),
                "stagedPatchPath": parsed_plan.get("patchPath"),
                "supplementParquetPath": parsed_plan.get("supplementParquetPath"),
                "planPath": _relative_to_project(plan_path),
                "reviewDir": parsed_plan.get("reviewDir"),
                "rawCompareReportPath": parsed_plan.get("rawCompareReportPath"),
                "stagingOutputPath": parsed_plan.get("stagingOutputPath"),
                "manifestPath": parsed_plan.get("manifestPath"),
                "partitionOutputPath": parsed_plan.get("partitionOutputPath"),
                "refreshReportPath": parsed_plan.get("refreshReportPath"),
                "fingerprintPath": parsed_plan.get("fingerprintPath"),
                "summariesOutputPath": parsed_plan.get("summariesOutputPath"),
            }
        )
        _persist_job_state(state)

        state["phase"] = "raw_compare"
        _persist_job_state(state)
        _ensure_job_not_cancelled(job_id)
        compare_args = _command_to_args(str(parsed_plan["compareCommand"]))
        _run_logged_command(
            label="Raw compare review",
            args=compare_args,
            log_path=log_path,
        )
        raw_compare_report = _read_json_if_exists(parsed_plan.get("rawCompareReportPath"))
        if raw_compare_report is None:
            raise RuntimeError("raw compare 完成后未找到 raw_compare_report.json")
        summaries = state.get("summaries")
        state["summaries"] = summaries if isinstance(summaries, dict) else {}
        state["summaries"]["rawCompare"] = _summarize_raw_compare_report(raw_compare_report)
        _persist_job_state(state)

        state["phase"] = "refresh"
        _persist_job_state(state)
        _ensure_job_not_cancelled(job_id)
        refresh_args = _command_to_args(str(parsed_plan["refreshCommand"]))
        _run_logged_command(
            label="Candidate refresh",
            args=refresh_args,
            log_path=log_path,
        )
        refresh_report = _read_json_if_exists(parsed_plan.get("refreshReportPath"))
        if refresh_report is None:
            raise RuntimeError("refresh 完成后未找到 refresh_job_report.json")
        state["summaries"]["refresh"] = _summarize_refresh_report(refresh_report)
        state["phase"] = "building_review"
        _persist_job_state(state)
        _cache_jato_monthly_update_review(job_id)
        state = _load_job_state(job_id)
        state["status"] = "success"
        state["phase"] = "completed"
        state["finishedAt"] = _utc_now().isoformat()
        _persist_job_state(state)
        _write_jato_etl_pipeline_status(state)
    except _JobResourceKilled as exc:
        state = _load_job_state(job_id)
        failed_phase = str(state.get("phase") or "unknown")
        state["status"] = "failed"
        state["phase"] = "resource_killed"
        state["finishedAt"] = _utc_now().isoformat()
        state["error"] = str(exc)
        state["failureDigest"] = _failure_digest_from_exception(
            phase=failed_phase,
            exc=exc,
        )
        state["currentProcess"] = None
        _persist_job_state(state)
        _write_jato_etl_pipeline_status(state)
        _append_log(log_path, f"[{_utc_now().isoformat()}] Resource killed: {exc}")
    except _JobCancelled as exc:
        state = _load_job_state(job_id)
        state["status"] = "cancelled"
        state["phase"] = "cancelled"
        state["finishedAt"] = state.get("finishedAt") or _utc_now().isoformat()
        state["error"] = str(exc)
        state["currentProcess"] = None
        _persist_job_state(state)
        _write_jato_etl_pipeline_status(state)
        _append_log(log_path, f"[{_utc_now().isoformat()}] Cancelled: {exc}")
    except Exception as exc:
        failed_phase = str(state.get("phase") or "failed")
        state["status"] = "failed"
        state["phase"] = "failed"
        state["finishedAt"] = _utc_now().isoformat()
        state["error"] = str(exc)
        state["failureDigest"] = _failure_digest_from_exception(
            phase=failed_phase,
            exc=exc,
        )
        _persist_job_state(state)
        _write_jato_etl_pipeline_status(state)
        _append_log(log_path, "\n=== Failed ===")
        _append_log(log_path, str(exc))
        _append_log(log_path, traceback.format_exc())
    finally:
        _RUNNING_THREADS.pop(job_id, None)


def _launch_job_thread(job_id: str) -> None:
    """Wake an isolated worker; never execute Pandas ETL in FastAPI."""
    _ = job_id
    execution_mode = os.getenv(
        "APP_JATO_MONTHLY_EXECUTION_MODE",
        "subprocess",
    ).strip().lower()
    if execution_mode == "external":
        return
    if execution_mode != "subprocess":
        raise RuntimeError(
            "APP_JATO_MONTHLY_EXECUTION_MODE 仅支持 subprocess 或 external。"
        )
    if not MONTHLY_WORKER_SCRIPT_PATH.exists():
        raise RuntimeError("JATO monthly worker 脚本不存在。")
    env = dict(os.environ)
    env.setdefault(
        "APP_JATO_MONTHLY_WORKER_MEMORY_LIMIT_BYTES",
        str(MONTHLY_WORKER_DEFAULT_MEMORY_LIMIT_BYTES),
    )
    env["MALLOC_ARENA_MAX"] = "2"
    env["OMP_NUM_THREADS"] = "1"
    env["OPENBLAS_NUM_THREADS"] = "1"
    env["MKL_NUM_THREADS"] = "1"
    env["NUMEXPR_NUM_THREADS"] = "1"
    subprocess.Popen(
        [sys.executable, str(MONTHLY_WORKER_SCRIPT_PATH), "--drain"],
        cwd=PROJECT_ROOT,
        stdin=subprocess.DEVNULL,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
        start_new_session=True,
        close_fds=True,
    )


def _normalized_smart_merge_resume_request_id(request_id: str) -> str:
    normalized = str(request_id or "").strip()
    if not SMART_MERGE_RESUME_REQUEST_ID_PATTERN.fullmatch(normalized):
        raise HTTPException(
            status_code=400,
            detail={
                "code": "INVALID_SMART_MERGE_RESUME_REQUEST_ID",
                "message": (
                    "requestId 必须为 8-128 位字母、数字或 ._:- "
                    "组成的幂等键。"
                ),
            },
        )
    return normalized


def _normalized_required_resume_sha(value: str, *, field: str) -> str:
    normalized = _valid_sha256(value)
    if normalized is None:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "INVALID_SMART_MERGE_RESUME_SEAL",
                "message": f"{field} 必须是 64 位 SHA-256。",
                "field": field,
            },
        )
    return normalized


def _smart_merge_resume_replay(
    *,
    payload: dict[str, Any],
    request_id: str,
    expected_seals: dict[str, str],
) -> dict[str, Any] | None:
    existing = _pending_operation(payload)
    if not isinstance(existing, dict):
        return None
    existing_status = str(existing.get("status") or "")
    existing_type = str(existing.get("type") or "")
    existing_request_id = str(existing.get("requestId") or "")
    if (
        existing_type == "smart_merge_resume"
        and existing_request_id == request_id
    ):
        if any(
            str(existing.get(field) or "").lower() != value
            for field, value in expected_seals.items()
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "SMART_MERGE_RESUME_REQUEST_CONFLICT",
                    "message": (
                        "相同 requestId 已绑定另一组恢复指纹，"
                        "拒绝重复提交。"
                    ),
                    "jobId": payload.get("jobId"),
                },
            )
        if existing_status == "queued":
            _launch_job_thread(str(payload.get("jobId") or ""))
        return _serialize_job_state(payload, include_log_tail=False)
    if existing_status in {"queued", "running"}:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "SMART_MERGE_RESUME_OPERATION_IN_PROGRESS",
                "message": (
                    f"{existing_type or 'active bundle'} 操作仍在"
                    f" {existing_status}，不能排队新的 Smart Merge 续跑。"
                ),
                "jobId": payload.get("jobId"),
            },
        )
    return None


def _queue_smart_merge_resume_locked(
    *,
    job_id: str,
    triggered_by: str,
    request_id: str,
    expected_source_candidate_fingerprint: str,
    expected_active_fingerprint: str,
    expected_report_fingerprint: str,
    expected_resolution_fingerprint: str,
) -> dict[str, Any]:
    payload = _load_job_state(job_id)
    expected_seals = {
        "expectedSourceCandidateFingerprint": (
            expected_source_candidate_fingerprint
        ),
        "expectedActiveFingerprint": expected_active_fingerprint,
        "expectedReportFingerprint": expected_report_fingerprint,
        "expectedResolutionFingerprint": expected_resolution_fingerprint,
    }
    replay = _smart_merge_resume_replay(
        payload=payload,
        request_id=request_id,
        expected_seals=expected_seals,
    )
    if replay is not None:
        return replay

    recovery = _smart_merge_recovery_view(payload)
    if not isinstance(recovery, dict) or not recovery.get("canResume"):
        raise HTTPException(
            status_code=409,
            detail={
                "code": "SMART_MERGE_RESUME_NOT_ALLOWED",
                "blockerType": "smart_merge_resume_not_allowed",
                "message": "当前任务不满足原地续跑 Smart Merge 的安全条件。",
                "reason": (
                    recovery.get("reason")
                    if isinstance(recovery, dict)
                    else "recovery_contract_missing"
                ),
                "jobId": job_id,
            },
        )
    if any(
        str(recovery.get(field) or "").lower() != value
        for field, value in {
            "sourceCandidateFingerprint": (
                expected_source_candidate_fingerprint
            ),
            "activeBaseFingerprint": expected_active_fingerprint,
            "reportFingerprint": expected_report_fingerprint,
            "resolutionFingerprint": expected_resolution_fingerprint,
        }.items()
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "code": "SMART_MERGE_RESUME_SEAL_CHANGED",
                "blockerType": "smart_merge_resume_seal_changed",
                "message": (
                    "页面读取后 Smart Merge 恢复合同已变化；"
                    "请刷新任务后重新确认。"
                ),
                "jobId": job_id,
            },
        )

    resolution = _historical_reclassification_resolution(payload)
    if not isinstance(resolution, dict):
        raise HTTPException(
            status_code=409,
            detail="任务缺少已保存的历史变化决策，不能续跑。",
        )
    _validated_historical_reclassification_resolution(resolution)
    if _active_dataset_version() != expected_active_fingerprint:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "SMART_MERGE_RESUME_ACTIVE_CHANGED",
                "blockerType": "stale_candidate",
                "message": (
                    "失败后 active lineage 已变化；"
                    "旧 Candidate 与决策不得继续应用。"
                ),
                "jobId": job_id,
            },
        )
    artifacts = payload.get("artifacts")
    if not isinstance(artifacts, dict):
        raise HTTPException(
            status_code=409,
            detail="任务缺少 candidate artifacts，不能续跑。",
        )
    candidate_scope = str(artifacts.get("candidateScope") or "")
    required_names = (
        {"parquet", "manifest", "refreshReport"}
        if candidate_scope in PARTITION_SCOPED_CANDIDATE_SCOPES
        else {name for name, _field in CANDIDATE_ARTIFACT_FIELDS}
    )
    missing = [
        artifact_name
        for artifact_name, artifact_field in CANDIDATE_ARTIFACT_FIELDS
        if artifact_name in required_names
        and (
            (path := _project_path(artifacts.get(artifact_field))) is None
            or not path.exists()
        )
    ]
    if missing:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "SMART_MERGE_RESUME_ARTIFACT_MISSING",
                "message": "Smart Merge 续跑所需 Candidate 产物缺失。",
                "missingArtifacts": missing,
                "jobId": job_id,
            },
        )
    return _queue_active_bundle_operation(
        payload=payload,
        operation_type="smart_merge_resume",
        triggered_by=triggered_by,
        operation_fields={
            "requestId": request_id,
            **expected_seals,
            "candidateArtifactStatSignature": (
                _candidate_artifact_stat_signature(artifacts)
            ),
            "resumeMode": (
                "merge_and_review"
                if candidate_scope in PARTITION_SCOPED_CANDIDATE_SCOPES
                else "review_only"
            ),
        },
    )


def resume_failed_jato_smart_merge(
    *,
    job_id: str,
    triggered_by: str,
    request_id: str,
    expected_source_candidate_fingerprint: str,
    expected_active_fingerprint: str,
    expected_report_fingerprint: str,
    expected_resolution_fingerprint: str,
) -> dict[str, Any]:
    normalized_request_id = _normalized_smart_merge_resume_request_id(
        request_id
    )
    expected_source = _normalized_required_resume_sha(
        expected_source_candidate_fingerprint,
        field="expectedSourceCandidateFingerprint",
    )
    expected_active = _normalized_required_resume_sha(
        expected_active_fingerprint,
        field="expectedActiveFingerprint",
    )
    expected_report = _normalized_required_resume_sha(
        expected_report_fingerprint,
        field="expectedReportFingerprint",
    )
    expected_resolution = _normalized_required_resume_sha(
        expected_resolution_fingerprint,
        field="expectedResolutionFingerprint",
    )
    expected_seals = {
        "expectedSourceCandidateFingerprint": expected_source,
        "expectedActiveFingerprint": expected_active,
        "expectedReportFingerprint": expected_report,
        "expectedResolutionFingerprint": expected_resolution,
    }

    # Fast idempotent replay path: do not make a lost HTTP response wait for
    # unrelated upload/resource gates.
    with _exclusive_file_lock(_job_state_lock_path(job_id)) as acquired:
        if not acquired:
            raise HTTPException(
                status_code=503,
                detail="Smart Merge 续跑状态锁暂不可用。",
            )
        replay = _smart_merge_resume_replay(
            payload=_load_job_state(job_id),
            request_id=normalized_request_id,
            expected_seals=expected_seals,
        )
        if replay is not None:
            return replay

    with _monthly_update_worker_start_window(
        action="仅续跑 Smart Merge",
        excluding_job_id=job_id,
    ):
        with _exclusive_file_lock(_job_state_lock_path(job_id)) as acquired:
            if not acquired:
                raise HTTPException(
                    status_code=503,
                    detail="Smart Merge 续跑状态锁暂不可用。",
                )
            return _queue_smart_merge_resume_locked(
                job_id=job_id,
                triggered_by=triggered_by,
                request_id=normalized_request_id,
                expected_source_candidate_fingerprint=expected_source,
                expected_active_fingerprint=expected_active,
                expected_report_fingerprint=expected_report,
                expected_resolution_fingerprint=expected_resolution,
            )


def _smart_merge_resume_bundle_is_durable(
    *,
    payload: dict[str, Any],
    operation: dict[str, Any],
) -> bool:
    operation_id = str(operation.get("operationId") or "")
    expected_active = _valid_sha256(
        operation.get("expectedActiveFingerprint")
    )
    resolution = _historical_reclassification_resolution(payload)
    artifacts = payload.get("artifacts")
    bundle_path = _job_review_bundle_path(str(payload.get("jobId") or ""))
    if (
        not operation_id
        or expected_active is None
        or str(payload.get("status") or "") != "success"
        or str(payload.get("phase") or "") != "completed"
        or not isinstance(resolution, dict)
        or str(resolution.get("status") or "") != "resolved"
        or not isinstance(artifacts, dict)
        or str(artifacts.get("candidateScope") or "")
        != "full_smart_merge"
        or not bundle_path.is_file()
    ):
        return False
    try:
        bundle = _read_json(bundle_path)
        resolved_candidate = _valid_sha256(
            resolution.get("resolvedCandidateFingerprint")
        )
        return bool(
            resolved_candidate is not None
            and str(bundle.get("reviewGenerationId") or "")
            == operation_id
            and _valid_sha256(bundle.get("candidateFingerprint"))
            == resolved_candidate
            and int(bundle.get("reviewBundleSchemaVersion") or 0)
            == REVIEW_BUNDLE_SCHEMA_VERSION
            and str(bundle.get("candidateArtifactStatSignature") or "")
            == _candidate_artifact_stat_signature(artifacts)
            and _valid_sha256(payload.get("activeBaseFingerprint"))
            == expected_active
            and _active_dataset_version() == expected_active
            and _historical_reclassification_resolution_fingerprint(
                resolution
            )
            == _valid_sha256(
                operation.get("expectedResolutionFingerprint")
            )
        )
    except Exception:
        return False


def _active_operation_failure_digest(
    *,
    operation_type: str,
    exc: BaseException,
) -> dict[str, Any]:
    if operation_type == "review_refresh":
        return _review_refresh_failure_digest(exc)
    if isinstance(exc, HTTPException):
        detail: Any = exc.detail
        message = (
            str(detail.get("message") or detail)
            if isinstance(detail, dict)
            else str(detail)
        )
        return {
            "code": f"{operation_type.upper()}_BLOCKED",
            "category": "safety_gate",
            "phase": f"{operation_type}_validating",
            "retryable": False,
            "message": message,
            "sourceFeedback": (
                str(detail.get("sourceFeedback") or "").strip()
                or None
                if isinstance(detail, dict)
                else None
            ),
            "technicalDetail": detail,
            "nextAction": "review_blocker",
        }
    return _failure_digest_from_exception(
        phase=f"{operation_type}_running",
        exc=exc,
    )


def _run_active_bundle_operation(
    *,
    job_id: str,
    operation_type: str,
) -> None:
    state = _load_job_state(job_id)
    operation = _pending_operation(state)
    if not isinstance(operation, dict):
        return
    if (
        str(operation.get("type") or "") != operation_type
        or str(operation.get("status") or "") != "queued"
    ):
        return
    operation_id = str(operation.get("operationId") or "")
    operation["status"] = "running"
    if operation_type in {"review_refresh", "smart_merge_resume"}:
        operation["phase"] = "verifying_candidate"
    operation["startedAt"] = _utc_now().isoformat()
    operation["error"] = None
    operation["failureDigest"] = None
    state["pendingOperation"] = operation
    _persist_job_state(state)
    _append_log(
        _job_log_path(job_id),
        f"[{_utc_now().isoformat()}] {operation_type} operation started in isolated worker.",
    )
    try:
        actor = str(operation.get("requestedBy") or "anonymous")
        if operation_type == "publish":
            _execute_publish_jato_monthly_update_job(
                job_id=job_id,
                triggered_by=actor,
            )
        elif operation_type == "rollback":
            _execute_rollback_jato_monthly_update_job(
                job_id=job_id,
                triggered_by=actor,
            )
        elif operation_type == "review_refresh":
            artifacts = state.get("artifacts")
            if not isinstance(artifacts, dict):
                raise HTTPException(
                    status_code=409,
                    detail="当前任务缺少 candidate 产物，不能重建 Review。",
                )
            queued_stat_signature = str(
                operation.get("candidateArtifactStatSignature") or ""
            )
            current_stat_signature = _candidate_artifact_stat_signature(
                artifacts
            )
            if (
                _candidate_artifact_stat_signature_version(
                    queued_stat_signature
                )
                != CANDIDATE_ARTIFACT_STAT_SIGNATURE_VERSION
                or current_stat_signature != queued_stat_signature
            ):
                raise HTTPException(
                    status_code=409,
                    detail={
                        "blockerType": "candidate_metadata_changed",
                        "message": (
                            "candidate 与 Review 重建排队时的元数据快照不一致；"
                            "未读取大型内容，也未修改 Candidate 或 active。"
                        ),
                    },
                )
            if _valid_sha256(
                operation.get("expectedCandidateFingerprint")
            ) is None:
                operation["expectedCandidateFingerprint"] = (
                    _candidate_fingerprint_id(artifacts)
                )
            operation["phase"] = "building_review"
            state["pendingOperation"] = operation
            _persist_job_state(state)
            _cache_jato_monthly_update_review(
                job_id,
                expected_candidate_fingerprint=str(
                    operation.get("expectedCandidateFingerprint") or ""
                ),
                expected_active_fingerprint=str(
                    operation.get("expectedActiveFingerprint") or ""
                ),
                review_generation_id=operation_id,
            )
            refreshed_state = _load_job_state(job_id)
            refreshed_operation = _pending_operation(refreshed_state)
            if (
                not isinstance(refreshed_operation, dict)
                or not _review_refresh_bundle_is_durable(
                    payload=refreshed_state,
                    operation=refreshed_operation,
                )
            ):
                raise RuntimeError("Review bundle durable seal 校验失败。")
        elif operation_type == "smart_merge_resume":
            artifacts = state.get("artifacts")
            resolution = _historical_reclassification_resolution(state)
            if not isinstance(artifacts, dict) or not isinstance(
                resolution,
                dict,
            ):
                raise HTTPException(
                    status_code=409,
                    detail="Smart Merge 续跑缺少 Candidate 或决策合同。",
                )
            current_stat_signature = _candidate_artifact_stat_signature(
                artifacts
            )
            if current_stat_signature != str(
                operation.get("candidateArtifactStatSignature") or ""
            ):
                raise HTTPException(
                    status_code=409,
                    detail={
                        "blockerType": "candidate_metadata_changed",
                        "message": (
                            "Candidate 与续跑排队时的元数据快照不一致；"
                            "未执行合并，也未修改 active。"
                        ),
                    },
                )
            expected_active = _valid_sha256(
                operation.get("expectedActiveFingerprint")
            )
            expected_source = _valid_sha256(
                operation.get("expectedSourceCandidateFingerprint")
            )
            expected_report = _valid_sha256(
                operation.get("expectedReportFingerprint")
            )
            expected_resolution = _valid_sha256(
                operation.get("expectedResolutionFingerprint")
            )
            if (
                expected_active is None
                or expected_source is None
                or expected_report is None
                or expected_resolution is None
                or _valid_sha256(state.get("activeBaseFingerprint"))
                != expected_active
                or _valid_sha256(
                    resolution.get("activeBaseFingerprint")
                )
                != expected_active
                or _active_dataset_version() != expected_active
                or _valid_sha256(resolution.get("reportFingerprint"))
                != expected_report
                or _valid_sha256(
                    resolution.get("sourceCandidateFingerprint")
                )
                != expected_source
                or _historical_reclassification_resolution_fingerprint(
                    resolution
                )
                != expected_resolution
            ):
                raise HTTPException(
                    status_code=409,
                    detail={
                        "blockerType": "smart_merge_resume_seal_changed",
                        "message": (
                            "Smart Merge 续跑合同已变化；"
                            "拒绝应用旧 Candidate 或决策。"
                        ),
                    },
                )
            _validated_historical_reclassification_resolution(resolution)
            candidate_scope = str(
                artifacts.get("candidateScope") or ""
            )
            if candidate_scope in PARTITION_SCOPED_CANDIDATE_SCOPES:
                actual_source = _candidate_fingerprint_id(artifacts)
                if actual_source != expected_source:
                    raise HTTPException(
                        status_code=409,
                        detail={
                            "blockerType": "candidate_content_drift",
                            "message": (
                                "原 Candidate 内容指纹已变化；"
                                "未执行 Smart Merge，也未修改 active。"
                            ),
                            "expectedCandidateFingerprint": (
                                expected_source
                            ),
                            "actualCandidateFingerprint": actual_source,
                        },
                    )
                operation["phase"] = "smart_merging"
            elif (
                candidate_scope == "full_smart_merge"
                and str(resolution.get("status") or "") == "resolved"
            ):
                operation["phase"] = "building_review"
            else:
                raise HTTPException(
                    status_code=409,
                    detail="Smart Merge 续跑 Candidate scope 无效。",
                )
            state["pendingOperation"] = operation
            _persist_job_state(state)
            _run_smart_merge(
                job_id,
                review_generation_id=operation_id,
            )
            refreshed_state = _load_job_state(job_id)
            if not (
                str(refreshed_state.get("status") or "") == "success"
                and str(refreshed_state.get("phase") or "")
                == "completed"
                and isinstance(
                    _historical_reclassification_resolution(
                        refreshed_state
                    ),
                    dict,
                )
                and str(
                    (
                        _historical_reclassification_resolution(
                            refreshed_state
                        )
                        or {}
                    ).get("status")
                    or ""
                )
                == "resolved"
            ):
                raise RuntimeError(
                    str(
                        refreshed_state.get("error")
                        or "Smart Merge 续跑未形成 completed Review。"
                    )
                )
            refreshed_operation = _pending_operation(refreshed_state)
            if not isinstance(
                refreshed_operation,
                dict,
            ) or not _smart_merge_resume_bundle_is_durable(
                payload=refreshed_state,
                operation=refreshed_operation,
            ):
                raise RuntimeError("Smart Merge 续跑 durable seal 校验失败。")
        else:
            raise RuntimeError(f"Unsupported active operation: {operation_type}")
        state = _load_job_state(job_id)
        current_operation = _pending_operation(state)
        if (
            not isinstance(current_operation, dict)
            or str(current_operation.get("operationId") or "") != operation_id
            or str(current_operation.get("type") or "") != operation_type
        ):
            _append_log(
                _job_log_path(job_id),
                (
                    f"[{_utc_now().isoformat()}] ignored stale "
                    f"{operation_type} completion for operationId={operation_id or '-'}"
                ),
            )
            return
        operation = current_operation
        operation["status"] = "success"
        if operation_type in {"review_refresh", "smart_merge_resume"}:
            operation["phase"] = "completed"
            if operation_type == "review_refresh":
                operation["resultCandidateFingerprint"] = operation.get(
                    "expectedCandidateFingerprint"
                )
            else:
                resolution = _historical_reclassification_resolution(
                    state
                )
                operation["resultCandidateFingerprint"] = (
                    resolution.get("resolvedCandidateFingerprint")
                    if isinstance(resolution, dict)
                    else None
                )
        operation["finishedAt"] = _utc_now().isoformat()
        operation["error"] = None
        operation["failureDigest"] = None
        state["pendingOperation"] = operation
        _persist_job_state(state)
    except Exception as exc:
        state = _load_job_state(job_id)
        current_operation = _pending_operation(state)
        if (
            not isinstance(current_operation, dict)
            or str(current_operation.get("operationId") or "") != operation_id
            or str(current_operation.get("type") or "") != operation_type
        ):
            _append_log(
                _job_log_path(job_id),
                (
                    f"[{_utc_now().isoformat()}] ignored stale "
                    f"{operation_type} failure for operationId={operation_id or '-'}: "
                    f"{exc}"
                ),
            )
            return
        operation = current_operation
        operation["status"] = "failed"
        operation["finishedAt"] = _utc_now().isoformat()
        operation["error"] = str(exc)
        job_failure_digest = state.get("failureDigest")
        operation["failureDigest"] = (
            job_failure_digest
            if operation_type == "smart_merge_resume"
            and isinstance(job_failure_digest, dict)
            else _active_operation_failure_digest(
                operation_type=operation_type,
                exc=exc,
            )
        )
        state["pendingOperation"] = operation
        _persist_job_state(state)
        _append_log(
            _job_log_path(job_id),
            (
                f"[{_utc_now().isoformat()}] {operation_type} operation failed: "
                f"{exc}"
            ),
        )


def _reconcile_stale_monthly_update_jobs() -> list[str]:
    reconciled: list[str] = []
    for payload in _list_job_state_payloads():
        pending = _pending_operation(payload)
        if (
            isinstance(pending, dict)
            and str(pending.get("status") or "") == "running"
        ):
            worker_pid = int(payload.get("workerPid") or 0)
            child_pid = _current_process_pid(payload)
            if _process_exists(worker_pid):
                continue
            pending_type = str(pending.get("type") or "")
            pending_operation_id = str(
                pending.get("operationId") or ""
            )
            publication = payload.get("publication")
            durable_completion = bool(
                (
                    isinstance(publication, dict)
                    and (
                    (
                        pending_type == "publish"
                        and publication.get("publishedAt")
                        and not publication.get("rolledBackAt")
                        and str(
                            publication.get("publishOperationId") or ""
                        )
                        == pending_operation_id
                    )
                    or (
                        pending_type == "rollback"
                        and publication.get("rolledBackAt")
                        and str(
                            publication.get("rollbackOperationId") or ""
                        )
                        == pending_operation_id
                    )
                )
                )
                or (
                    pending_type == "review_refresh"
                    and _review_refresh_bundle_is_durable(
                        payload=payload,
                        operation=pending,
                    )
                )
                or (
                    pending_type == "smart_merge_resume"
                    and _smart_merge_resume_bundle_is_durable(
                        payload=payload,
                        operation=pending,
                    )
                )
            )
            if durable_completion:
                pending["status"] = "success"
                if pending_type in {
                    "review_refresh",
                    "smart_merge_resume",
                }:
                    pending["phase"] = "completed"
                    if pending_type == "review_refresh":
                        pending["resultCandidateFingerprint"] = (
                            pending.get("expectedCandidateFingerprint")
                        )
                    else:
                        resolution = (
                            _historical_reclassification_resolution(
                                payload
                            )
                        )
                        pending["resultCandidateFingerprint"] = (
                            resolution.get(
                                "resolvedCandidateFingerprint"
                            )
                            if isinstance(resolution, dict)
                            else None
                        )
                pending["finishedAt"] = _utc_now().isoformat()
                pending["error"] = None
                pending["failureDigest"] = None
                pending["recoveredAfterWorkerLoss"] = True
                if pending_type == "review_refresh":
                    artifacts = payload.get("artifacts")
                    if not isinstance(artifacts, dict):
                        artifacts = {}
                    artifacts["reviewBundlePath"] = _relative_to_project(
                        _job_review_bundle_path(
                            str(payload.get("jobId") or "")
                        )
                    )
                    payload["artifacts"] = artifacts
                payload["pendingOperation"] = pending
                payload["workerPid"] = None
                payload["currentProcess"] = None
                _persist_job_state(payload)
                reconciled.append(str(payload.get("jobId") or ""))
                continue
            current_process = payload.get("currentProcess")
            orphan_termination = (
                _terminate_process_group(
                    child_pid,
                    expected_identity=(
                        current_process.get("identity")
                        if isinstance(current_process, dict)
                        and isinstance(
                            current_process.get("identity"),
                            dict,
                        )
                        else None
                    ),
                )
                if _process_exists(child_pid)
                else None
            )
            job_id = str(payload.get("jobId") or "")
            if not job_id:
                continue
            pending["status"] = "failed"
            pending["finishedAt"] = _utc_now().isoformat()
            if pending_type == "review_refresh":
                pending["phase"] = "worker_lost"
                pending["error"] = (
                    "Review 重建 worker 在 durable bundle 提交前退出；"
                    "candidate 与 active 均未修改，可人工重新发起。"
                )
                pending["failureDigest"] = {
                    "code": "REVIEW_REFRESH_WORKER_LOST",
                    "category": "resource",
                    "phase": "review_refresh",
                    "retryable": True,
                    "message": pending["error"],
                    "sourceFeedback": None,
                    "technicalDetail": {"workerPid": worker_pid or None},
                    "nextAction": "retry_review_refresh",
                }
            elif pending_type == "smart_merge_resume":
                pending["phase"] = "worker_lost"
                pending["error"] = (
                    "Smart Merge 续跑 worker 在 durable Review 提交前退出；"
                    "active 未修改，可用新 requestId 再次仅续跑。"
                )
                pending["failureDigest"] = {
                    "code": "SMART_MERGE_RESUME_WORKER_LOST",
                    "category": "resource",
                    "phase": "smart_merge_resume",
                    "retryable": True,
                    "message": pending["error"],
                    "sourceFeedback": None,
                    "technicalDetail": {
                        "workerPid": worker_pid or None,
                        "childPid": child_pid,
                        "orphanTermination": orphan_termination,
                    },
                    "nextAction": "resume_smart_merge",
                }
                payload["status"] = "failed"
                payload["phase"] = "smart_merge_failed"
                payload["error"] = pending["error"]
                payload["failureDigest"] = pending["failureDigest"]
                resolution = _historical_reclassification_resolution(
                    payload
                )
                artifacts = payload.get("artifacts")
                if (
                    isinstance(resolution, dict)
                    and not (
                        isinstance(artifacts, dict)
                        and str(artifacts.get("candidateScope") or "")
                        == "full_smart_merge"
                        and str(resolution.get("status") or "")
                        == "resolved"
                    )
                ):
                    resolution["status"] = "failed"
                    resolution["failedAt"] = _utc_now().isoformat()
                    resolution["error"] = pending["error"]
                    payload[
                        "historicalReclassificationResolution"
                    ] = resolution
            else:
                pending["error"] = (
                    "独立月更 worker 在 active bundle 操作完成前退出；"
                    "下一次操作前将按 transaction 记录检查 active。"
                )
                pending["failureDigest"] = {
                    "code": "WORKER_LOST",
                    "category": "resource",
                    "phase": f"{pending.get('type') or 'active'}_running",
                    "retryable": False,
                    "message": pending["error"],
                    "sourceFeedback": None,
                    "technicalDetail": {
                        "workerPid": worker_pid or None,
                        "childPid": child_pid,
                        "orphanTermination": orphan_termination,
                    },
                    "nextAction": "inspect_worker_and_active_transaction",
                }
            payload["pendingOperation"] = pending
            payload["workerPid"] = None
            payload["currentProcess"] = None
            _persist_job_state(payload)
            _append_log(
                _job_log_path(job_id),
                f"[{_utc_now().isoformat()}] Active operation reconciliation: worker_lost.",
            )
            reconciled.append(job_id)
            continue
        if str(payload.get("status") or "") != "running":
            continue
        worker_pid = int(payload.get("workerPid") or 0)
        child_pid = _current_process_pid(payload)
        if _process_exists(worker_pid):
            continue
        current_process = payload.get("currentProcess")
        orphan_termination = (
            _terminate_process_group(
                child_pid,
                expected_identity=(
                    current_process.get("identity")
                    if isinstance(current_process, dict)
                    and isinstance(
                        current_process.get("identity"),
                        dict,
                    )
                    else None
                ),
            )
            if _process_exists(child_pid)
            else None
        )
        job_id = str(payload.get("jobId") or "")
        if not job_id:
            continue
        smart_merge_worker_lost = (
            str(payload.get("operation") or "") == "smart_merge"
        )
        payload["status"] = "failed"
        payload["phase"] = (
            "smart_merge_failed"
            if smart_merge_worker_lost
            else "worker_lost"
        )
        payload["finishedAt"] = _utc_now().isoformat()
        payload["error"] = (
            "Smart Merge worker 在 durable Review 完成前退出；"
            "Candidate 与决策已保留，未修改 active。"
            if smart_merge_worker_lost
            else (
                "独立月更 worker 在任务完成前退出；上传与 candidate 已保留，"
                "未修改 active。"
            )
        )
        payload["failureDigest"] = {
            "code": (
                "SMART_MERGE_WORKER_LOST"
                if smart_merge_worker_lost
                else "WORKER_LOST"
            ),
            "category": "resource",
            "phase": payload["phase"],
            "retryable": smart_merge_worker_lost,
            "message": payload["error"],
            "sourceFeedback": None,
            "technicalDetail": {
                "workerPid": worker_pid or None,
                "childPid": child_pid,
                "orphanTermination": orphan_termination,
            },
            "nextAction": (
                "resume_smart_merge"
                if smart_merge_worker_lost
                else "inspect_worker_service"
            ),
        }
        if smart_merge_worker_lost:
            resolution = _historical_reclassification_resolution(payload)
            artifacts = payload.get("artifacts")
            if (
                isinstance(resolution, dict)
                and not (
                    isinstance(artifacts, dict)
                    and str(artifacts.get("candidateScope") or "")
                    == "full_smart_merge"
                    and str(resolution.get("status") or "")
                    == "resolved"
                )
            ):
                resolution["status"] = "failed"
                resolution["failedAt"] = _utc_now().isoformat()
                resolution["error"] = payload["error"]
                payload[
                    "historicalReclassificationResolution"
                ] = resolution
        payload["workerPid"] = None
        payload["currentProcess"] = None
        _persist_job_state(payload)
        _write_jato_etl_pipeline_status(payload)
        _append_log(
            _job_log_path(job_id),
            f"[{_utc_now().isoformat()}] Worker reconciliation: worker_lost.",
        )
        reconciled.append(job_id)
    return reconciled


def _reconcile_stale_baseline_promotion() -> bool:
    state = _load_baseline_promotion_state()
    if not isinstance(state, dict) or str(state.get("status") or "") != "running":
        return False
    worker_pid = int(state.get("workerPid") or 0)
    if _process_exists(worker_pid):
        return False
    journal = (
        _read_json(_baseline_install_journal_path())
        if _baseline_install_journal_path().exists()
        else None
    )
    operation_id = str(state.get("operationId") or "")
    journal_matches = (
        isinstance(journal, dict)
        and str(journal.get("operationId") or "") == operation_id
    )
    target_path = (
        Path(str(journal.get("targetPath") or ""))
        if journal_matches
        else None
    )
    target_sha256 = (
        str(journal.get("targetSha256") or "")
        if journal_matches
        else ""
    )
    installed = bool(
        target_path is not None
        and target_path.is_file()
        and re.fullmatch(r"[0-9a-f]{64}", target_sha256)
        and _sha256_hex_for_path(target_path) == target_sha256
    )
    if installed and isinstance(journal, dict):
        for raw_path in journal.get("oldBaselinePaths", []):
            old_path = Path(str(raw_path))
            if old_path != target_path:
                old_path.unlink(missing_ok=True)
        result = journal.get("result")
        if isinstance(result, dict):
            state.update(result)
        state["status"] = "success"
        state["finishedAt"] = _utc_now().isoformat()
        state["error"] = None
        state["failureDigest"] = None
        state["workerPid"] = None
        state["recoveredAfterWorkerLoss"] = True
        journal["status"] = "committed"
        journal["committedAt"] = _utc_now().isoformat()
        journal["updatedAt"] = journal["committedAt"]
        _write_json(_baseline_install_journal_path(), journal)
        _write_json(_baseline_promotion_state_path(), state)
        return True
    state["status"] = "failed"
    state["finishedAt"] = _utc_now().isoformat()
    state["error"] = (
        "baseline 保存 worker 在安装新文件前退出；未发现目标 SHA 已安装，"
        "现有 baseline 保持不变。"
    )
    state["failureDigest"] = {
        "code": "WORKER_LOST",
        "category": "resource",
        "phase": "baseline_promotion_running",
        "retryable": True,
        "message": state["error"],
        "sourceFeedback": None,
        "technicalDetail": {"workerPid": worker_pid or None},
        "nextAction": "retry_baseline_promotion",
    }
    state["workerPid"] = None
    _write_json(_baseline_promotion_state_path(), state)
    return True


def _run_baseline_promotion_operation() -> str | None:
    with _exclusive_file_lock(_baseline_promotion_lock_path()) as acquired:
        if not acquired:
            return None
        state = _load_baseline_promotion_state()
        if (
            not isinstance(state, dict)
            or str(state.get("status") or "") != "queued"
        ):
            return None
        operation_id = str(state.get("operationId") or "")
        state["status"] = "running"
        state["startedAt"] = _utc_now().isoformat()
        state["workerPid"] = os.getpid()
        _write_json(_baseline_promotion_state_path(), state)
    try:
        result = _execute_promote_current_active_to_baseline(
            triggered_by=str(state.get("requestedBy") or "anonymous"),
            expected_active_fingerprint=str(
                state.get("sourceActiveFingerprint") or ""
            ),
            operation_id=operation_id,
        )
        latest = _load_baseline_promotion_state() or {}
        if str(latest.get("operationId") or "") != operation_id:
            return operation_id or None
        latest.update(result)
        latest["status"] = "success"
        latest["finishedAt"] = _utc_now().isoformat()
        latest["error"] = None
        latest["failureDigest"] = None
        latest["workerPid"] = None
        _write_json(_baseline_promotion_state_path(), latest)
    except Exception as exc:
        latest = _load_baseline_promotion_state() or state
        if str(latest.get("operationId") or "") == operation_id:
            latest["status"] = "failed"
            latest["finishedAt"] = _utc_now().isoformat()
            latest["error"] = str(exc)
            latest["failureDigest"] = _failure_digest_from_exception(
                phase="baseline_promotion_running",
                exc=exc,
            )
            latest["workerPid"] = None
            _write_json(_baseline_promotion_state_path(), latest)
    return operation_id or None


def run_jato_monthly_update_worker_once() -> dict[str, Any]:
    """Claim and run at most one job under a cross-process flock."""
    with _exclusive_worker_cycle() as acquired:
        if not acquired:
            return {
                "processedJobId": None,
                "reconciledJobIds": [],
                "skipped": "worker_lock_held",
            }
        recovered_active_transactions = (
            _recover_incomplete_active_transactions_if_possible()
        )
        baseline_reconciled = _reconcile_stale_baseline_promotion()
        reconciled = _reconcile_stale_monthly_update_jobs()
        baseline_promotion = _load_baseline_promotion_state()
        if (
            isinstance(baseline_promotion, dict)
            and str(baseline_promotion.get("status") or "") == "queued"
        ):
            operation_id = _run_baseline_promotion_operation()
            return {
                "processedJobId": operation_id,
                "reconciledJobIds": reconciled,
                "baselinePromotionReconciled": baseline_reconciled,
                "recoveredActiveTransactions": recovered_active_transactions,
            }
        pending_operations = sorted(
            (
                payload
                for payload in _list_job_state_payloads()
                if (
                    isinstance(payload.get("pendingOperation"), dict)
                    and str(payload["pendingOperation"].get("status") or "")
                    == "queued"
                )
            ),
            key=lambda payload: str(
                payload["pendingOperation"].get("requestedAt") or ""
            ),
        )
        queued = sorted(
            (
                payload
                for payload in _list_job_state_payloads()
                if str(payload.get("status") or "") == "queued"
            ),
            key=lambda payload: str(payload.get("createdAt") or ""),
        )
        if not pending_operations and not queued:
            return {
                "processedJobId": None,
                "reconciledJobIds": reconciled,
                "baselinePromotionReconciled": baseline_reconciled,
            }
        state = pending_operations[0] if pending_operations else queued[0]
        job_id = str(state.get("jobId") or "")
        state["workerPid"] = os.getpid()
        state["workerStartedAt"] = _utc_now().isoformat()
        _persist_job_state(state)
        try:
            pending = _pending_operation(state)
            if (
                isinstance(pending, dict)
                and str(pending.get("status") or "") == "queued"
            ):
                _run_active_bundle_operation(
                    job_id=job_id,
                    operation_type=str(pending.get("type") or ""),
                )
            elif str(state.get("operation") or "") == "smart_merge":
                _run_smart_merge(job_id)
            else:
                _run_job(job_id)
        except Exception as exc:
            state = _load_job_state(job_id)
            failed_phase = str(state.get("phase") or "worker")
            state["status"] = "failed"
            state["phase"] = "failed"
            state["finishedAt"] = _utc_now().isoformat()
            state["error"] = str(exc)
            state["failureDigest"] = _failure_digest_from_exception(
                phase=failed_phase,
                exc=exc,
            )
            state["workerPid"] = None
            state["currentProcess"] = None
            _persist_job_state(state)
            _write_jato_etl_pipeline_status(state)
            _append_log(
                _job_log_path(job_id),
                f"[{_utc_now().isoformat()}] Isolated worker failed: {exc}",
            )
        finally:
            state = _load_job_state(job_id)
            state["workerPid"] = None
            _persist_job_state(state)
        return {
            "processedJobId": job_id,
            "reconciledJobIds": reconciled,
            "baselinePromotionReconciled": baseline_reconciled,
            "recoveredActiveTransactions": recovered_active_transactions,
        }


def _active_partition_country_names(partition_root: Path | None = None) -> list[str]:
    root = partition_root or _active_data_paths()["partition"]
    if not root.exists():
        return []
    prefix = "国家="
    return sorted(
        {
            unquote(path.name[len(prefix):])
            for path in root.iterdir()
            if path.is_dir() and path.name.startswith(prefix)
        }
    )


def _normalized_month_from_label(label: str) -> str:
    parsed = datetime.strptime(str(label).strip().title(), "%Y %b")
    return f"{parsed.year}-{parsed.month:02d}"


def _inspect_upload_scope(path: Path) -> dict[str, Any]:
    """Inspect country/month scope without materializing every Excel column."""
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(
                path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
            try:
                if DEFAULT_UPLOAD_SHEET_NAME not in workbook.sheetnames:
                    raise RuntimeError(
                        f"上传 Excel 缺少工作表：{DEFAULT_UPLOAD_SHEET_NAME}。"
                    )
                worksheet = workbook[DEFAULT_UPLOAD_SHEET_NAME]
                rows = worksheet.iter_rows(values_only=True)
                header_row = next(rows, None)
                if header_row is None:
                    raise RuntimeError("上传 Excel 的 Data Export 为空。")
                columns = [str(value).strip() if value is not None else "" for value in header_row]
                duplicate_headers = sorted(
                    {
                        column
                        for column in columns
                        if column and columns.count(column) > 1
                    }
                )
                country_column = _find_country_column(columns)
                if country_column is None:
                    raise RuntimeError("上传 Excel 的 Data Export 缺少国家列。")
                country_index = columns.index(country_column)
                month_columns = _detect_month_columns(columns)
                if not month_columns:
                    raise RuntimeError("上传 Excel 的 Data Export 缺少可识别月份列。")
                month_indexes = [
                    (columns.index(month_column), month_column)
                    for month_column in month_columns
                ]
                countries: list[str] = []
                seen_countries: set[str] = set()
                country_latest_labels: dict[str, str | None] = {}
                data_row_count = 0
                negative_sales_countries: set[str] = set()
                negative_sales_fields: set[str] = set()
                invalid_sales_countries: set[str] = set()
                invalid_sales_fields: set[str] = set()
                orphan_sales_row_count = 0
                for row in rows:
                    raw_country = row[country_index] if country_index < len(row) else None
                    country = str(raw_country).strip() if raw_country is not None else ""
                    row_has_sales = False
                    latest_row_month: str | None = None
                    for month_index, month_label in month_indexes:
                        value = row[month_index] if month_index < len(row) else None
                        if value is None or (
                            isinstance(value, str) and not value.strip()
                        ):
                            continue
                        row_has_sales = True
                        try:
                            numeric_value = (
                                float(value)
                                if not isinstance(value, bool)
                                else float("nan")
                            )
                        except (TypeError, ValueError):
                            numeric_value = float("nan")
                        if pd.isna(numeric_value):
                            if country:
                                invalid_sales_countries.add(country)
                            invalid_sales_fields.add(month_label)
                            continue
                        if numeric_value < 0:
                            if country:
                                negative_sales_countries.add(country)
                            negative_sales_fields.add(month_label)
                        if (
                            latest_row_month is None
                            or _time_sort_key(month_label)
                            > _time_sort_key(latest_row_month)
                        ):
                            latest_row_month = month_label
                    if not country:
                        if row_has_sales:
                            orphan_sales_row_count += 1
                        continue
                    data_row_count += 1
                    if country not in seen_countries:
                        seen_countries.add(country)
                        countries.append(country)
                        country_latest_labels[country] = None
                    if latest_row_month is not None:
                        current = country_latest_labels[country]
                        if (
                            current is None
                            or _time_sort_key(latest_row_month)
                            > _time_sort_key(current)
                        ):
                            country_latest_labels[country] = latest_row_month
                inspection_issues: list[dict[str, Any]] = []
                if duplicate_headers:
                    inspection_issues.append(
                        {
                            "code": "DUPLICATE_COLUMNS",
                            "message": "Data Export 存在重复字段名。",
                            "countries": [],
                            "fields": duplicate_headers,
                            "sourceFeedback": (
                                "请洗数人员删除或重命名重复列；同名月份列不能自动判断。"
                            ),
                        }
                    )
                if invalid_sales_fields:
                    inspection_issues.append(
                        {
                            "code": "NON_NUMERIC_MONTHLY_SALES",
                            "message": "月份销量列包含非数字值。",
                            "countries": sorted(invalid_sales_countries),
                            "fields": sorted(
                                invalid_sales_fields,
                                key=_time_sort_key,
                            ),
                            "sourceFeedback": (
                                "请洗数人员把月份销量清洗为数字或真正空值；"
                                "不要填 N/A、-、文本数字。"
                            ),
                        }
                    )
                if negative_sales_fields:
                    inspection_issues.append(
                        {
                            "code": "NEGATIVE_MONTHLY_SALES",
                            "message": "月份销量列包含负数。",
                            "countries": sorted(negative_sales_countries),
                            "fields": sorted(
                                negative_sales_fields,
                                key=_time_sort_key,
                            ),
                            "sourceFeedback": (
                                "请洗数人员核对负销量；普通月更不接受用负数冲销历史。"
                            ),
                        }
                    )
                if orphan_sales_row_count:
                    inspection_issues.append(
                        {
                            "code": "SALES_WITHOUT_COUNTRY",
                            "message": (
                                f"发现 {orphan_sales_row_count} 行有月份销量但国家为空。"
                            ),
                            "countries": [],
                            "fields": [country_column],
                            "sourceFeedback": (
                                "请洗数人员补齐这些行的国家，或删除不属于数据区的行。"
                            ),
                        }
                    )
            finally:
                workbook.close()
        except RuntimeError:
            raise
        except Exception as exc:
            raise RuntimeError(f"无法轻量读取上传 Excel：{exc}") from exc
    else:
        frame = _read_excel_with_fallback(path, sheet_name=DEFAULT_UPLOAD_SHEET_NAME)
        frame.columns = [str(column).strip() for column in frame.columns]
        country_column = _find_country_column(list(frame.columns))
        if country_column is None:
            raise RuntimeError("上传 Excel 的 Data Export 缺少国家列。")
        month_columns = _detect_month_columns(list(frame.columns))
        if not month_columns:
            raise RuntimeError("上传 Excel 的 Data Export 缺少可识别月份列。")
        frame[country_column] = frame[country_column].astype("string").fillna("").str.strip()
        countries = _ordered_distinct_strings(frame[country_column].tolist())
        country_latest_labels = {}
        for country in countries:
            country_frame = frame.loc[frame[country_column] == country]
            country_latest_labels[country] = _latest_month_from_frame(country_frame)
        data_row_count = int((frame[country_column] != "").sum())
        inspection_issues = []
        month_frame = frame[month_columns]
        numeric_month_frame = month_frame.apply(
            pd.to_numeric,
            errors="coerce",
        )
        present_mask = month_frame.notna() & month_frame.astype("string").ne("")
        invalid_mask = present_mask & numeric_month_frame.isna()
        if bool(invalid_mask.any().any()):
            inspection_issues.append(
                {
                    "code": "NON_NUMERIC_MONTHLY_SALES",
                    "message": "月份销量列包含非数字值。",
                    "countries": [],
                    "fields": [
                        column
                        for column in month_columns
                        if bool(invalid_mask[column].any())
                    ],
                    "sourceFeedback": (
                        "请洗数人员把月份销量清洗为数字或真正空值。"
                    ),
                }
            )
        negative_mask = numeric_month_frame.lt(0)
        if bool(negative_mask.any().any()):
            inspection_issues.append(
                {
                    "code": "NEGATIVE_MONTHLY_SALES",
                    "message": "月份销量列包含负数。",
                    "countries": [],
                    "fields": [
                        column
                        for column in month_columns
                        if bool(negative_mask[column].any())
                    ],
                    "sourceFeedback": (
                        "请洗数人员核对负销量；普通月更不接受用负数冲销历史。"
                    ),
                }
            )

    if not countries:
        raise RuntimeError("上传 Excel 的 Data Export 不包含有效国家。")
    latest_labels = [
        label for label in country_latest_labels.values()
        if label is not None
    ]
    if not latest_labels:
        raise RuntimeError("上传 Excel 的月份列均为空。")
    latest_label = max(latest_labels, key=_time_sort_key)
    return {
        "sheetName": DEFAULT_UPLOAD_SHEET_NAME,
        "countries": countries,
        "countryLatestMonths": {
            country: (
                _normalized_month_from_label(label)
                if label is not None
                else None
            )
            for country, label in country_latest_labels.items()
        },
        "latestMonth": _normalized_month_from_label(latest_label),
        "dataRowCount": data_row_count,
        "issues": inspection_issues,
    }


def _detect_partial_country_upload(path: Path) -> dict[str, Any] | None:
    inspection = _inspect_upload_scope(path)
    active_countries = _active_partition_country_names()
    uploaded_countries = set(inspection["countries"])
    if (
        active_countries
        and uploaded_countries.issubset(set(active_countries))
        and len(uploaded_countries) < len(active_countries)
    ):
        return inspection
    return None


def _detect_single_country_upload(path: Path) -> tuple[str, str] | None:
    """Return (country, month) for a one-country Data Export."""
    try:
        inspection = _inspect_upload_scope(path)
        countries = inspection["countries"]
        if len(countries) != 1:
            return None
        return countries[0], str(inspection["latestMonth"])
    except Exception:
        return None


def _queue_single_country_job(
    *,
    job_id: str,
    country: str,
    month: str,
    triggered_by: str,
    upload_filename: str,
    stored_upload_path: Path,
) -> dict[str, Any]:
    """Create job state and launch single-country background runner."""
    normalized_month = _normalize_month(month)
    normalized_country = country.strip()

    active_paths = _active_data_paths()
    active_base_fingerprint = _active_dataset_version()
    if active_paths["parquet"].exists():
        try:
            active_frame = (
                _load_active_country_partition_subset(active_paths["partition"], normalized_country)
                if active_paths["partition"].exists()
                else _load_parquet_country_subset(
                    active_paths["parquet"], normalized_country, path_label="active"
                )
            )
            active_latest = _latest_month_from_frame(active_frame)
            active_latest_month = (
                _normalized_month_from_label(active_latest)
                if active_latest
                else None
            )
            if active_latest_month and normalized_month <= active_latest_month:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"国家「{normalized_country}」在 active 数据集中已有 {active_latest_month} 月的数据，"
                        f"上传月份 {normalized_month} 不更新。请提供晚于 {active_latest_month} 的月份。"
                    ),
                )
        except HTTPException:
            raise
        except Exception:
            pass
    if _active_dataset_version() != active_base_fingerprint:
        raise HTTPException(
            status_code=409,
            detail=(
                "校验单国月份期间 active 已变化；为避免把旧分区与新版本"
                "绑定，请重新提交该文件。"
            ),
        )

    job_dir = _job_dir(job_id)
    now = _utc_now().isoformat()
    batch_id = f"{normalized_month}-{normalized_country}-single"
    state: dict[str, Any] = {
        "jobId": job_id,
        "month": normalized_month,
        "batchId": batch_id,
        "status": "queued",
        "phase": "queued",
        "jobType": "single_country",
        "expectedCountry": normalized_country,
        "triggeredBy": triggered_by.strip() or "anonymous",
        "country": normalized_country,
        "countryScope": [normalized_country],
        "createdAt": now,
        "updatedAt": now,
        "startedAt": None,
        "finishedAt": None,
        "error": None,
        "upload": {
            "originalFilename": upload_filename,
            "storedPath": _relative_to_project(stored_upload_path),
            "sizeBytes": stored_upload_path.stat().st_size,
            "sha256": _sha256_hex_for_path(stored_upload_path),
        },
        "plan": None,
        "artifacts": {
            "jobDir": _relative_to_project(job_dir),
            "logPath": _relative_to_project(_job_log_path(job_id)),
        },
        "summaries": {},
        "logPath": _relative_to_project(_job_log_path(job_id)),
        "activeBaseFingerprint": active_base_fingerprint,
    }
    _persist_job_state(state)
    _append_log(
        _job_log_path(job_id),
        f"[{now}] 已入队单国家任务：{normalized_country} {normalized_month}",
    )

    _launch_job_thread(job_id)

    return _serialize_job_state(state, include_log_tail=False)


def _create_jato_monthly_update_job_in_start_window(
    *,
    file: UploadFile,
    triggered_by: str,
) -> dict[str, Any]:
    filename = _validate_upload(file)

    file_size = file.size or 0
    if file_size > _MAX_DIRECT_UPLOAD_BYTES:
        raise HTTPException(
            status_code=400,
            detail=(
                f"文件 {file_size / 1024 / 1024:.0f}MB 超过 {_MAX_DIRECT_UPLOAD_BYTES / 1024 / 1024:.0f}MB，"
                "请使用分片上传（chunked upload）。"
            ),
        )

    job_id = f"jato-update-{uuid4().hex[:8]}"
    uploads_dir = _job_dir(job_id) / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    stored_upload_path = _job_upload_storage_path(job_id, filename)

    with stored_upload_path.open("wb") as handle:
        shutil.copyfileobj(file.file, handle)

    month = (
        _parse_month_from_filename(filename)
        or datetime.now().strftime("%Y-%m")
    )
    return _queue_monthly_update_job_from_stored_upload(
        job_id=job_id,
        triggered_by=triggered_by,
        upload_filename=filename,
        stored_upload_path=stored_upload_path,
        month=month,
    )


def create_jato_monthly_update_job(
    *,
    file: UploadFile,
    triggered_by: str,
) -> dict[str, Any]:
    with _monthly_update_worker_start_window(action="启动月更任务"):
        return _create_jato_monthly_update_job_in_start_window(
            file=file,
            triggered_by=triggered_by,
        )


def _create_jato_monthly_update_job_from_upload_in_start_window(
    *,
    upload_id: str,
    triggered_by: str,
    triggered_role: str = "editor",
) -> dict[str, Any]:
    state = _load_upload_session(upload_id)
    _require_upload_session_access(
        state,
        requested_by=triggered_by,
        requested_role=triggered_role,
    )
    consumed_job_id = str(state.get("consumedJobId") or "").strip()
    if consumed_job_id:
        _ensure_upload_digest_matches_current_active(
            upload_id=upload_id,
            state=state,
            state_lock_held=True,
        )
        return get_jato_monthly_update_job(consumed_job_id)
    if str(state.get("status", "")) == "invalid":
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "upload_digest_invalid",
                "message": "上传 digest 未通过，不能创建月更任务。",
                "ingestDigest": state.get("ingestDigest"),
                "failureDigest": state.get("failureDigest"),
            },
        )
    if str(state.get("status", "")) != "ready":
        raise HTTPException(
            status_code=409,
            detail="上传仍在组装或 digest，只有 ready 后才能创建月更任务。",
        )

    filename = _validate_upload_filename(str(state.get("filename", "jato-update.xlsx")))
    assembled_path = (
        _persisted_upload_session_assembled_path(state)
        or _upload_session_assembled_path(upload_id, filename)
    )
    if not assembled_path.exists():
        raise HTTPException(status_code=409, detail="组装后的上传文件不存在，请重新上传。")
    declared_size = int(state.get("sizeBytes", 0) or 0)
    actual_size = assembled_path.stat().st_size
    if (
        declared_size <= 0
        or declared_size > UPLOAD_MAX_BYTES
        or actual_size != declared_size
    ):
        raise HTTPException(
            status_code=409,
            detail=(
                f"组装文件大小与上传声明不一致（声明 {declared_size} 字节，"
                f"实际 {actual_size} 字节），请重新上传。"
            ),
        )
    file_sha256 = _normalize_sha256(
        state.get("fileSha256"),
        detail="上传文件指纹缺失，请重新完成组装。",
    )
    actual_file_sha256 = _sha256_hex_for_path(assembled_path)
    if actual_file_sha256 != file_sha256:
        raise HTTPException(
            status_code=409,
            detail="组装文件 SHA-256 与 digest 不一致，请重新上传。",
        )
    ingest_digest = state.get("ingestDigest")
    if not isinstance(ingest_digest, dict):
        raise HTTPException(status_code=409, detail="上传缺少 ingest digest，不能创建任务。")
    if (
        int(ingest_digest.get("sizeBytes", 0) or 0) != actual_size
        or str(ingest_digest.get("fileSha256") or "").strip().lower()
        != actual_file_sha256
    ):
        raise HTTPException(
            status_code=409,
            detail="上传 digest 的文件大小或 SHA-256 已失配，请重新上传。",
        )
    if ingest_digest.get("blockers"):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "upload_digest_blocked",
                "message": "上传 digest 存在 blocker，不能创建任务。",
                "ingestDigest": ingest_digest,
            },
        )
    _ensure_upload_digest_matches_current_active(
        upload_id=upload_id,
        state=state,
        state_lock_held=True,
    )

    digest_month_raw = str(ingest_digest.get("latestMonth") or "").strip()
    if not digest_month_raw:
        raise HTTPException(
            status_code=400,
            detail="无法从工作簿 digest 识别真实月份。",
        )
    digest_month = _normalize_month(digest_month_raw)

    ingestion_key = _build_ingestion_key(ingest_digest)
    ingestion_lock = MONTHLY_UPDATE_JOB_ROOT / INGESTION_LOCK_FILENAME
    with _exclusive_file_lock(ingestion_lock) as acquired:
        if not acquired:  # blocking lock always acquires; retained for portability
            raise HTTPException(status_code=503, detail="月更摄入锁暂不可用，请稍后重试。")
        state = _load_upload_session(upload_id)
        _require_upload_session_access(
            state,
            requested_by=triggered_by,
            requested_role=triggered_role,
        )
        consumed_job_id = str(state.get("consumedJobId") or "").strip()
        if consumed_job_id:
            _ensure_upload_digest_matches_current_active(
                upload_id=upload_id,
                state=state,
                state_lock_held=True,
            )
            consumed_job = get_jato_monthly_update_job(consumed_job_id)
            if str(consumed_job.get("status") or "") == "queued":
                _launch_job_thread(consumed_job_id)
            return consumed_job
        existing = _find_existing_job_for_ingestion_key(ingestion_key)
        if existing is not None:
            existing_job_id = str(existing.get("jobId") or "")
            if (
                existing_job_id
                and str(existing.get("status") or "") == "queued"
            ):
                _launch_job_thread(existing_job_id)
            state["status"] = "consumed"
            state["consumedJobId"] = existing_job_id
            state["ingestionKey"] = ingestion_key
            _persist_upload_session(state)
            return _serialize_job_state(existing, include_log_tail=False)

        job_id = f"jato-update-{uuid4().hex[:8]}"
        uploads_dir = _job_dir(job_id) / "uploads"
        uploads_dir.mkdir(parents=True, exist_ok=True)
        stored_upload_path = _job_upload_storage_path(job_id, filename)
        try:
            os.link(assembled_path, stored_upload_path)
        except OSError:
            shutil.copy2(assembled_path, stored_upload_path)
        try:
            result = _queue_monthly_update_job_from_stored_upload(
                job_id=job_id,
                triggered_by=triggered_by,
                upload_filename=filename,
                stored_upload_path=stored_upload_path,
                month=digest_month,
                file_sha256=file_sha256,
                ingest_digest=ingest_digest,
                ingestion_key=ingestion_key,
            )
        except Exception:
            shutil.rmtree(_job_dir(job_id), ignore_errors=True)
            raise
        state["status"] = "consumed"
        state["consumedJobId"] = job_id
        state["ingestionKey"] = ingestion_key
        _persist_upload_session(state)
        return result


def create_jato_monthly_update_job_from_upload(
    *,
    upload_id: str,
    triggered_by: str,
    triggered_role: str = "editor",
) -> dict[str, Any]:
    state = _load_upload_session(upload_id)
    _require_upload_session_access(
        state,
        requested_by=triggered_by,
        requested_role=triggered_role,
    )
    action = "从上传创建月更任务"
    with _monthly_update_resource_start_locks(action=action):
        with _exclusive_file_lock(
            _upload_state_lock_path(upload_id)
        ) as state_acquired:
            if not state_acquired:
                raise HTTPException(
                    status_code=503,
                    detail="上传会话状态锁暂不可用，请稍后重试。",
                )
            state = _load_upload_session(upload_id)
            _require_upload_session_access(
                state,
                requested_by=triggered_by,
                requested_role=triggered_role,
            )
            if str(state.get("status") or "") == "ready":
                _require_no_active_upload_sessions(
                    action=action,
                    excluding_ready_upload_id=upload_id,
                )
                _require_no_running_monthly_update_jobs()
            return _create_jato_monthly_update_job_from_upload_in_start_window(
                upload_id=upload_id,
                triggered_by=triggered_by,
                triggered_role=triggered_role,
            )


def _recover_failed_jato_monthly_update_job_in_start_window(
    *,
    source_job_id: str,
    recovery_key: str,
    triggered_by: str,
) -> dict[str, Any]:
    existing_recovery = _find_recovery_job_for_source(
        source_job_id=source_job_id,
        recovery_key=recovery_key,
    )
    if existing_recovery is not None:
        return _serialize_job_state(existing_recovery, include_log_tail=False)

    _require_no_active_upload_sessions(action="创建恢复任务")
    _require_no_running_monthly_update_jobs()

    with _exclusive_file_lock(
        _job_state_lock_path(source_job_id)
    ) as source_acquired:
        if not source_acquired:
            raise HTTPException(
                status_code=503,
                detail="原任务状态锁暂不可用，请稍后重试。",
            )
        source_state = _load_job_state(source_job_id)
        if str(source_state.get("status") or "") != "failed":
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_SOURCE_NOT_FAILED",
                    "message": "只允许为 failed 的月更任务创建恢复 attempt。",
                },
            )
        source_failure = source_state.get("failureDigest")
        source_failure_category = (
            str(source_failure.get("category") or "").strip().lower()
            if isinstance(source_failure, dict)
            else ""
        )
        if source_failure_category not in RECOVERY_ALLOWED_FAILURE_CATEGORIES:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_FAILURE_NOT_ELIGIBLE",
                    "message": (
                        "只允许恢复 resource/platform 类平台失败；"
                        "数据或 processing 失败必须修正源数据。"
                    ),
                    "failureCode": (
                        source_failure.get("code")
                        if isinstance(source_failure, dict)
                        else None
                    ),
                    "failureCategory": source_failure_category or None,
                },
            )
        if isinstance(source_state.get("publication"), dict):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_SOURCE_HAS_PUBLICATION",
                    "message": "原任务已进入过 publish 记录，不能从原文件恢复。",
                },
            )
        if isinstance(source_state.get("reviewApproval"), dict):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_SOURCE_HAS_APPROVAL",
                    "message": "原任务已有 Review 决策，不能作为恢复源。",
                },
            )
        if isinstance(source_state.get("pendingOperation"), dict):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_SOURCE_HAS_PENDING_OPERATION",
                    "message": "原任务已有 Publish/Rollback 操作记录，不能恢复。",
                },
            )

        ingest_digest = source_state.get("ingestDigest")
        if not isinstance(ingest_digest, dict):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_DIGEST_MISSING",
                    "message": "原任务缺少 ingestDigest，不能绕过重新上传校验。",
                },
            )
        if (
            str(ingest_digest.get("status") or "") != "ready"
            or ingest_digest.get("blockers")
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_DIGEST_NOT_READY",
                    "message": "原任务 digest 未通过，不能创建 recovery attempt。",
                },
            )
        if str(ingest_digest.get("route") or "") != "partial_country":
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_ROUTE_NOT_ALLOWED",
                    "message": "受控恢复当前仅支持 partial_country 目标分区路径。",
                },
            )

        source_ingestion_key = str(
            source_state.get("ingestionKey") or ""
        ).strip()
        rebuilt_ingestion_key = _build_ingestion_key(ingest_digest)
        if (
            not re.fullmatch(r"[0-9a-f]{64}", source_ingestion_key)
            or source_ingestion_key != rebuilt_ingestion_key
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_INGESTION_KEY_MISMATCH",
                    "message": "原任务 ingestionKey 与 ingestDigest 不一致，拒绝恢复。",
                },
            )
        duplicate = _find_existing_job_for_ingestion_key(
            source_ingestion_key
        )
        if duplicate is not None:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_INGESTION_ALREADY_ACTIVE",
                    "message": "同一 ingestionKey 已有非 failed 任务，不能再创建恢复任务。",
                    "existingJobId": duplicate.get("jobId"),
                },
            )

        consumed_upload = _consumed_upload_session_for_job(source_job_id)
        upload_digest = consumed_upload.get("ingestDigest")
        if not isinstance(upload_digest, dict) or upload_digest != ingest_digest:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_UPLOAD_DIGEST_MISMATCH",
                    "message": "consumed upload 与原任务的 digest 不一致。",
                },
            )
        if str(consumed_upload.get("ingestionKey") or "") != source_ingestion_key:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_UPLOAD_INGESTION_KEY_MISMATCH",
                    "message": "consumed upload 与原任务的 ingestionKey 不一致。",
                },
            )

        source_upload = source_state.get("upload")
        if not isinstance(source_upload, dict):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_SOURCE_UPLOAD_MISSING",
                    "message": "原任务缺少上传文件记录。",
                },
            )
        stored_path_value = str(source_upload.get("storedPath") or "").strip()
        source_upload_path = _project_path(stored_path_value)
        expected_upload_dir = (_job_dir(source_job_id) / "uploads").resolve()
        if (
            source_upload_path is None
            or not source_upload_path.is_file()
            or source_upload_path.resolve().parent != expected_upload_dir
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_SOURCE_FILE_MISSING",
                    "message": "原任务上传副本不存在或路径越界。",
                },
            )

        try:
            expected_size = int(source_upload.get("sizeBytes") or 0)
            digest_size = int(ingest_digest.get("sizeBytes") or 0)
            session_size = int(consumed_upload.get("sizeBytes") or 0)
        except (TypeError, ValueError):
            expected_size = digest_size = session_size = 0
        expected_sha256 = str(source_upload.get("sha256") or "").lower()
        digest_sha256 = str(ingest_digest.get("fileSha256") or "").lower()
        session_sha256 = str(consumed_upload.get("fileSha256") or "").lower()
        if (
            expected_size <= 0
            or expected_size != digest_size
            or expected_size != session_size
            or not re.fullmatch(r"[0-9a-f]{64}", expected_sha256)
            or expected_sha256 != digest_sha256
            or expected_sha256 != session_sha256
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_SOURCE_METADATA_MISMATCH",
                    "message": "原任务、consumed upload 与 digest 的文件大小/SHA-256 不一致。",
                },
            )

        stat_before = source_upload_path.stat()
        actual_sha256 = _sha256_hex_for_path(source_upload_path)
        stat_after = source_upload_path.stat()
        stable_stat_before = (
            stat_before.st_dev,
            stat_before.st_ino,
            stat_before.st_size,
            stat_before.st_mtime_ns,
        )
        stable_stat_after = (
            stat_after.st_dev,
            stat_after.st_ino,
            stat_after.st_size,
            stat_after.st_mtime_ns,
        )
        if stable_stat_before != stable_stat_after:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_SOURCE_CHANGED_DURING_CHECK",
                    "message": "原任务上传副本在校验期间发生变化。",
                },
            )
        if stat_after.st_size != expected_size or actual_sha256 != expected_sha256:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_SOURCE_FILE_MISMATCH",
                    "message": "原任务上传副本的实际大小/SHA-256 与审计记录不一致。",
                    "actualSizeBytes": stat_after.st_size,
                    "actualSha256": actual_sha256,
                },
            )

        active_base_fingerprint = str(
            source_state.get("activeBaseFingerprint") or ""
        ).strip()
        digest_active_fingerprint = str(
            ingest_digest.get("activeDatasetVersion") or ""
        ).strip()
        current_active_fingerprint = _active_dataset_version()
        if (
            not re.fullmatch(r"[0-9a-f]{64}", active_base_fingerprint)
            or active_base_fingerprint != digest_active_fingerprint
            or active_base_fingerprint != current_active_fingerprint
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_ACTIVE_LINEAGE_CHANGED",
                    "message": "原 digest 绑定的 active lineage 已缺失或变化，拒绝复用。",
                    "sourceActiveBaseFingerprint": (
                        active_base_fingerprint or None
                    ),
                    "digestActiveDatasetVersion": (
                        digest_active_fingerprint or None
                    ),
                    "currentActiveFingerprint": current_active_fingerprint,
                },
            )

        digest_month = _normalize_month(
            str(ingest_digest.get("latestMonth") or "")
        )
        source_month = _normalize_month(str(source_state.get("month") or ""))
        if digest_month != source_month:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RECOVERY_MONTH_MISMATCH",
                    "message": "原任务月份与 digest 真实最新月份不一致。",
                },
            )

        recovery_job_id = f"jato-update-{uuid4().hex[:8]}"
        filename = _validate_upload_filename(
            str(
                source_upload.get("originalFilename")
                or source_upload_path.name
            )
        )
        recovery_upload_path = _job_upload_storage_path(
            recovery_job_id,
            filename,
        )
        recovery_upload_path.parent.mkdir(parents=True, exist_ok=True)
        storage_method = "hardlink"
        try:
            os.link(source_upload_path, recovery_upload_path)
        except OSError:
            storage_method = "copy"
            shutil.copy2(source_upload_path, recovery_upload_path)
        try:
            if recovery_upload_path.stat().st_size != expected_size:
                raise HTTPException(
                    status_code=409,
                    detail={
                        "code": "RECOVERY_COPY_SIZE_MISMATCH",
                        "message": "recovery attempt 文件大小校验失败。",
                    },
                )
            if _sha256_hex_for_path(recovery_upload_path) != expected_sha256:
                raise HTTPException(
                    status_code=409,
                    detail={
                        "code": "RECOVERY_TARGET_SHA_MISMATCH",
                        "message": "recovery attempt 文件 SHA-256 校验失败。",
                    },
                )
            if _active_dataset_version() != active_base_fingerprint:
                raise HTTPException(
                    status_code=409,
                    detail={
                        "code": "RECOVERY_ACTIVE_LINEAGE_CHANGED",
                        "message": "创建 recovery attempt 期间 active lineage 发生变化。",
                    },
                )
            result = _queue_monthly_update_job_from_stored_upload(
                job_id=recovery_job_id,
                triggered_by=triggered_by,
                upload_filename=filename,
                stored_upload_path=recovery_upload_path,
                month=digest_month,
                file_sha256=actual_sha256,
                ingest_digest=ingest_digest,
                ingestion_key=source_ingestion_key,
                recovery_of_job_id=source_job_id,
                recovery_key=recovery_key,
                recovery_source={
                    "uploadId": consumed_upload.get("uploadId"),
                    "validatedAt": _utc_now().isoformat(),
                    "sizeBytes": expected_size,
                    "sha256": actual_sha256,
                    "activeBaseFingerprint": active_base_fingerprint,
                    "storageMethod": storage_method,
                },
            )
        except Exception:
            shutil.rmtree(_job_dir(recovery_job_id), ignore_errors=True)
            raise
        return result


def recover_failed_jato_monthly_update_job(
    *,
    source_job_id: str,
    recovery_key: str,
    triggered_by: str,
) -> dict[str, Any]:
    normalized_recovery_key = _normalize_recovery_key(recovery_key)
    action = "创建月更 recovery attempt"
    recovery_lock = MONTHLY_UPDATE_JOB_ROOT / RECOVERY_LOCK_FILENAME
    with _exclusive_file_lock(recovery_lock) as recovery_acquired:
        if not recovery_acquired:
            raise HTTPException(
                status_code=503,
                detail="月更恢复幂等锁暂不可用，请稍后重试。",
            )
        with _monthly_update_resource_start_locks(action=action):
            ingestion_lock = MONTHLY_UPDATE_JOB_ROOT / INGESTION_LOCK_FILENAME
            with _exclusive_file_lock(ingestion_lock) as ingestion_acquired:
                if not ingestion_acquired:
                    raise HTTPException(
                        status_code=503,
                        detail="月更恢复摄入锁暂不可用，请稍后重试。",
                    )
                return _recover_failed_jato_monthly_update_job_in_start_window(
                    source_job_id=source_job_id,
                    recovery_key=normalized_recovery_key,
                    triggered_by=triggered_by,
                )


def _retry_failed_jato_monthly_update_job_in_start_window(
    *,
    source_job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    source_state = _load_job_state(source_job_id)
    if str(source_state.get("status", "")) != "failed":
        raise HTTPException(
            status_code=409,
            detail="只有 failed 的月更任务才能直接重试。",
        )
    if (
        str(source_state.get("phase") or "") == "smart_merge_failed"
        or str(source_state.get("operation") or "") == "smart_merge"
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "code": "SMART_MERGE_RESUME_REQUIRED",
                "blockerType": "smart_merge_resume_required",
                "message": (
                    "Smart Merge 失败必须在原任务仅续跑；"
                    "禁止复制上传文件或重跑 ETL。"
                ),
                "jobId": source_job_id,
                "nextAction": "resume_smart_merge",
            },
        )

    source_upload = source_state.get("upload")
    if not isinstance(source_upload, dict):
        raise HTTPException(status_code=409, detail="原任务缺少 upload 信息，无法直接重试。")

    stored_path_value = source_upload.get("storedPath")
    if not stored_path_value:
        raise HTTPException(
            status_code=409,
            detail="原任务的上传副本已不存在，请重新上传文件后再试。",
        )

    source_upload_path = _project_path(str(stored_path_value))
    if source_upload_path is None or not source_upload_path.exists():
        raise HTTPException(
            status_code=409,
            detail="原任务的上传副本已不存在，请重新上传文件后再试。",
        )

    filename = _validate_upload_filename(
        str(source_upload.get("originalFilename") or source_upload_path.name)
    )
    job_id = f"jato-update-{uuid4().hex[:8]}"
    uploads_dir = _job_dir(job_id) / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    stored_upload_path = _job_upload_storage_path(job_id, filename)
    shutil.copy2(source_upload_path, stored_upload_path)
    source_month = str(source_state.get("month") or "").strip()
    retry_month = (
        source_month
        or _parse_month_from_filename(filename)
        or datetime.now().strftime("%Y-%m")
    )
    try:
        result = _queue_monthly_update_job_from_stored_upload(
            job_id=job_id,
            triggered_by=triggered_by,
            upload_filename=filename,
            stored_upload_path=stored_upload_path,
            month=retry_month,
            file_sha256=(
                str(source_upload.get("sha256"))
                if source_upload.get("sha256") is not None
                else None
            ),
        )
    except Exception:
        shutil.rmtree(_job_dir(job_id), ignore_errors=True)
        raise

    payload = _load_job_state(job_id)
    artifacts = payload.get("artifacts")
    payload["artifacts"] = artifacts if isinstance(artifacts, dict) else {}
    payload["artifacts"]["retriedFromJobId"] = source_job_id
    _persist_job_state(payload)
    return _serialize_job_state(payload, include_log_tail=True)


def retry_failed_jato_monthly_update_job(
    *,
    source_job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    with _monthly_update_worker_start_window(action="重试失败的月更任务"):
        return _retry_failed_jato_monthly_update_job_in_start_window(
            source_job_id=source_job_id,
            triggered_by=triggered_by,
        )


# ── Country-Partition Upload ───────────────────────────────────────────────────


def _validate_candidate_logical_country_scope(
    *,
    candidate_path: Path,
    expected_countries: list[str],
) -> list[str]:
    expected_by_key: dict[str, str] = {}
    for raw_country in expected_countries:
        country = str(raw_country).strip()
        country_key = country.casefold()
        if not country or country_key in expected_by_key:
            raise RuntimeError(
                "任务绑定 countryScope 包含空国家或重复逻辑国家。"
            )
        expected_by_key[country_key] = country
    actual_countries = list(
        _collect_dataset_country_latest_months(candidate_path)
    )
    actual_by_key = {
        country.strip().casefold(): country.strip()
        for country in actual_countries
        if country.strip()
    }
    missing_keys = set(expected_by_key) - set(actual_by_key)
    extra_keys = set(actual_by_key) - set(expected_by_key)
    if missing_keys or extra_keys:
        raise RuntimeError(
            "部分国家 ETL 输出国家范围与任务绑定范围不一致："
            f"missing={','.join(expected_by_key[key] for key in sorted(missing_keys)) or '-'}，"
            f"extra={','.join(actual_by_key[key] for key in sorted(extra_keys)) or '-'}"
        )
    return [
        actual_by_key[key]
        for key in expected_by_key
    ]


def _run_country_partition_job(job_id: str) -> None:
    """Build a Review-only candidate for a strict subset of active countries."""
    state = _load_job_state(job_id)
    if str(state.get("status") or "") == "cancelled":
        _RUNNING_THREADS.pop(job_id, None)
        return
    log_path = _job_log_path(job_id)

    try:
        state["status"] = "running"
        state["phase"] = "validating"
        _persist_job_state(state)

        upload = state.get("upload") or {}
        stored_path_value = upload.get("storedPath") if isinstance(upload, dict) else None
        stored_upload_path = _project_path(str(stored_path_value or "").strip())
        if stored_upload_path is None or not stored_upload_path.exists():
            raise RuntimeError("上传文件不存在。")
        expected_upload_sha = str(upload.get("sha256") or "").strip().lower()
        if not re.fullmatch(r"[0-9a-f]{64}", expected_upload_sha):
            raise RuntimeError("任务缺少有效上传 SHA-256，拒绝转换未绑定文件。")
        actual_upload_sha = _sha256_hex_for_path(stored_upload_path)
        if actual_upload_sha != expected_upload_sha:
            raise RuntimeError(
                "任务上传副本 SHA-256 已变化，拒绝构建 candidate。"
            )

        countries = _ordered_distinct_strings(
            [
                str(country)
                for country in (
                    state.get("countryScope")
                    if isinstance(state.get("countryScope"), list)
                    else [state.get("country")]
                )
                if country is not None
            ]
        )
        month = str(state.get("month", "")).strip()
        if not countries or not month:
            raise RuntimeError("任务状态缺少 countryScope 或 month。")

        suffix = stored_upload_path.suffix.lower()
        if suffix not in ALLOWED_UPLOAD_EXTENSIONS:
            raise RuntimeError(f"不支持的文件格式：{suffix}，仅支持 Excel。")

        inspection = state.get("uploadInspection")
        if not isinstance(inspection, dict):
            inspection = _inspect_upload_scope(stored_upload_path)
        uploaded_countries = _ordered_distinct_strings(
            [str(country) for country in inspection.get("countries", [])]
        )
        if uploaded_countries != countries:
            raise RuntimeError(
                "上传文件的国家范围与任务绑定范围不一致："
                f"expected={','.join(countries)}，actual={','.join(uploaded_countries) or '-'}"
            )

        _append_log(
            log_path,
            (
                f"[{_utc_now().isoformat()}] 验证通过："
                f"countries={','.join(countries)}, month={month}"
            ),
        )

        # Staging paths under job dir
        job_dir = _job_dir(job_id)
        staging_dir = job_dir / "staging"
        staging_dir.mkdir(parents=True, exist_ok=True)
        staging_output = staging_dir / "country_partition_candidate.parquet"
        manifest_path = staging_dir / "manifest.json"
        report_path = staging_dir / "refresh_job_report.json"
        active_paths = _active_data_paths()
        expected_active_fingerprint = str(
            state.get("activeBaseFingerprint") or ""
        ).strip()
        if (
            not re.fullmatch(r"[0-9a-f]{64}", expected_active_fingerprint)
            or _active_dataset_version() != expected_active_fingerprint
        ):
            raise RuntimeError(
                "目标国家 candidate 开始构建前 active lineage 已缺失或变化；"
                "拒绝混用不同 active 分区，请创建新 attempt。"
            )
        if not active_paths["partition"].exists():
            raise RuntimeError("部分国家刷新需要 active partitioned dataset，当前分区目录不存在。")
        missing_active_countries = [
            country
            for country in countries
            if not (
                active_paths["partition"]
                / f"国家={quote(country, safe='')}"
            ).exists()
        ]
        if missing_active_countries:
            raise RuntimeError(
                "active 分区缺少上传国家："
                + "、".join(missing_active_countries)
            )
        untouched_before = _untouched_partition_snapshot(
            partition_root=active_paths["partition"],
            countries=countries,
        )
        state["phase"] = "refreshing"
        _persist_job_state(state)
        _ensure_job_not_cancelled(job_id)

        refresh_args = [
            sys.executable,
            str(SINGLE_COUNTRY_ETL_SCRIPT_PATH),
            "--input",
            str(stored_upload_path.resolve()),
            "--output",
            str(staging_output.resolve()),
            "--manifest",
            str(manifest_path.resolve()),
            "--job-id",
            job_id,
        ]
        refresh_args.extend(
            _partial_country_streaming_cli_args(
                upload_suffix=suffix,
                active_paths=active_paths,
            )
        )

        _run_logged_command(
            label="部分国家目标分区转换",
            args=refresh_args,
            log_path=log_path,
        )
        _validate_candidate_logical_country_scope(
            candidate_path=staging_output,
            expected_countries=countries,
        )
        country_latest_months = (
            inspection.get("countryLatestMonths")
            if isinstance(inspection.get("countryLatestMonths"), dict)
            else {}
        )
        country_results: list[dict[str, Any]] = []
        total_rows = 0
        candidate_column_count = 0
        for country in countries:
            candidate_frame = _load_parquet_country_subset(
                staging_output,
                country,
                path_label="country-partition candidate",
            )
            if candidate_frame.empty:
                raise RuntimeError(f"candidate 不包含目标国家：{country}")
            candidate_latest = _latest_month_from_frame(candidate_frame)
            expected_month = str(country_latest_months.get(country) or "").strip()
            expected_month_label = (
                datetime.strptime(expected_month, "%Y-%m").strftime("%Y %b")
                if expected_month
                else None
            )
            if candidate_latest != expected_month_label:
                raise RuntimeError(
                    f"{country} candidate 的最新月份与上传识别不一致："
                    f"expected={expected_month_label or '-'}, actual={candidate_latest or '-'}"
                )
            active_frame = _load_active_country_partition_subset(
                active_paths["partition"],
                country,
            )
            active_latest = _latest_month_from_frame(active_frame)
            if (
                active_latest
                and candidate_latest
                and _time_sort_key(candidate_latest) < _time_sort_key(active_latest)
            ):
                raise RuntimeError(
                    f"{country} 最新月份发生回退："
                    f"active={active_latest}, candidate={candidate_latest}"
                )
            total_rows += int(len(candidate_frame))
            candidate_column_count = max(
                candidate_column_count,
                int(len(candidate_frame.columns)),
            )
            country_results.append(
                {
                    "country": country,
                    "rows": int(len(candidate_frame)),
                    "activeLatestMonth": active_latest,
                    "candidateLatestMonth": candidate_latest,
                }
            )
            del active_frame
            del candidate_frame
        expected_row_count = int(inspection.get("dataRowCount", 0) or 0)
        if expected_row_count <= 0 or total_rows != expected_row_count:
            raise RuntimeError(
                "部分国家 ETL 输出行数与 Digest 不一致："
                f"expected={expected_row_count}, actual={total_rows}"
            )
        untouched_after = _untouched_partition_snapshot(
            partition_root=active_paths["partition"],
            countries=countries,
        )
        untouched_partition_check = _verify_untouched_partition_stability(
            before=untouched_before,
            after=untouched_after,
        )
        candidate_scope = (
            "target_country_partition_only"
            if len(countries) == 1
            else "target_country_partitions_only"
        )
        refresh_report = {
            "jobStatus": "success",
            "fullManifest": {
                "rows": total_rows,
                "columns": candidate_column_count,
            },
            "partitionManifest": {"parquetFileCount": len(countries)},
            "incremental": {
                "enabled": True,
                "scope": candidate_scope,
                "targetCountry": countries[0] if len(countries) == 1 else None,
                "targetCountries": countries,
                "countryResults": country_results,
                "untouchedPartitionCheck": untouched_partition_check,
            },
        }
        _write_json(report_path, refresh_report)

        state["artifacts"] = {
            "jobDir": _relative_to_project(job_dir),
            "logPath": _relative_to_project(log_path),
            "baselinePath": None,
            "baselineSource": "active_country_partition",
            "stagedPatchPath": _relative_to_project(stored_upload_path),
            "supplementParquetPath": None,
            "stagingOutputPath": _relative_to_project(staging_output),
            "manifestPath": _relative_to_project(manifest_path),
            "partitionOutputPath": None,
            "refreshReportPath": _relative_to_project(report_path),
            "fingerprintPath": None,
            "reviewDir": None,
            "rawCompareReportPath": None,
            "planPath": None,
            "candidateScope": candidate_scope,
            "untouchedPartitionCheck": untouched_partition_check,
        }
        state["summaries"] = {
            "refresh": _summarize_refresh_report(refresh_report),
            "jobInfo": {
                "country": countries[0] if len(countries) == 1 else None,
                "countries": countries,
                "month": month,
                "type": str(state.get("jobType") or "partial_country"),
            },
        }
        state["plan"] = None
        state["phase"] = "building_review"
        if _active_dataset_version() != expected_active_fingerprint:
            raise RuntimeError(
                "目标国家 candidate 构建期间 active 已变化；"
                "本 attempt 已停止，不能生成可审批 Review。"
            )
        _persist_job_state(state)
        _cache_jato_monthly_update_review(job_id)
        state = _load_job_state(job_id)
        state["status"] = "success"
        state["phase"] = "completed"
        state["finishedAt"] = _utc_now().isoformat()
        _persist_job_state(state)
        _write_jato_etl_pipeline_status(state)
        _append_log(
            log_path,
            (
                f"[{_utc_now().isoformat()}] 部分国家任务完成："
                f"{','.join(countries)} {month}。可前往 Review（不改变 active）。"
            ),
        )

    except _JobResourceKilled as exc:
        state = _load_job_state(job_id)
        failed_phase = str(state.get("phase") or "unknown")
        state["status"] = "failed"
        state["phase"] = "resource_killed"
        state["finishedAt"] = _utc_now().isoformat()
        state["error"] = str(exc)
        state["failureDigest"] = _failure_digest_from_exception(
            phase=failed_phase,
            exc=exc,
        )
        state["currentProcess"] = None
        _persist_job_state(state)
        _write_jato_etl_pipeline_status(state)
        _append_log(log_path, f"[{_utc_now().isoformat()}] Resource killed: {exc}")
    except _JobCancelled as exc:
        state = _load_job_state(job_id)
        state["status"] = "cancelled"
        state["phase"] = "cancelled"
        state["finishedAt"] = state.get("finishedAt") or _utc_now().isoformat()
        state["error"] = str(exc)
        state["currentProcess"] = None
        _persist_job_state(state)
        _write_jato_etl_pipeline_status(state)
        _append_log(log_path, f"[{_utc_now().isoformat()}] Cancelled: {exc}")
    except Exception as exc:
        failed_phase = str(state.get("phase") or "failed")
        state["status"] = "failed"
        state["phase"] = "failed"
        state["finishedAt"] = _utc_now().isoformat()
        state["error"] = str(exc)
        state["failureDigest"] = _failure_digest_from_exception(
            phase=failed_phase,
            exc=exc,
        )
        _persist_job_state(state)
        _write_jato_etl_pipeline_status(state)
        _append_log(log_path, "\n=== Failed ===")
        _append_log(log_path, str(exc))
        _append_log(log_path, traceback.format_exc())
    finally:
        _RUNNING_THREADS.pop(job_id, None)


def _run_single_country_job(job_id: str) -> None:
    """Compatibility wrapper for the explicit single-country API."""
    _run_country_partition_job(job_id)


def _create_single_country_job_in_start_window(
    *,
    country: str,
    month: str,
    file: UploadFile,
    triggered_by: str,
) -> dict[str, Any]:
    """Create a lightweight single-country single-month upload job (no prepare/raw_compare)."""
    filename = _validate_upload(file)
    normalized_month = _normalize_month(month)
    normalized_country = country.strip()
    if not normalized_country:
        raise HTTPException(status_code=400, detail="country 不能为空。")

    job_id = f"jato-sc-{uuid4().hex[:8]}"
    job_dir = _job_dir(job_id)
    uploads_dir = job_dir / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)
    stored_upload_path = _job_upload_storage_path(job_id, filename)

    with stored_upload_path.open("wb") as handle:
        shutil.copyfileobj(file.file, handle)

    return _queue_single_country_job(
        job_id=job_id,
        country=normalized_country,
        month=normalized_month,
        triggered_by=triggered_by,
        upload_filename=filename,
        stored_upload_path=stored_upload_path,
    )


def create_single_country_job(
    *,
    country: str,
    month: str,
    file: UploadFile,
    triggered_by: str,
) -> dict[str, Any]:
    with _monthly_update_worker_start_window(action="启动单国家月更任务"):
        return _create_single_country_job_in_start_window(
            country=country,
            month=month,
            file=file,
            triggered_by=triggered_by,
        )


# ── Smart Merge ──────────────────────────────────────────────────────────────


def _static_carry_forward_key_columns(
    *,
    active_frame: pd.DataFrame,
    candidate_frame: pd.DataFrame,
) -> list[str]:
    shared = [
        column
        for column in STATIC_CARRY_FORWARD_KEY_CANDIDATES
        if column in active_frame.columns and column in candidate_frame.columns
    ]
    if "Make" not in shared or "Model" not in shared:
        return []
    if "Version name" not in shared and "Trim level" not in shared:
        return []
    return shared


def _normalized_static_value_series(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series.dtype):
        numbers = pd.to_numeric(series, errors="coerce")
        return numbers.map(
            lambda value: (
                ""
                if pd.isna(value)
                else format(float(value), ".15g")
            )
        )
    return (
        series.astype("string")
        .fillna("")
        .str.strip()
        .str.casefold()
    )


def _normalized_static_key_frame(
    frame: pd.DataFrame,
    key_columns: list[str],
) -> pd.DataFrame:
    return pd.DataFrame(
        {
            f"__key_{index}": _normalized_static_value_series(frame[column])
            for index, column in enumerate(key_columns)
        },
        index=frame.index,
    )


def _canonical_country_content_signature(
    frame: pd.DataFrame,
    columns: list[str],
) -> str:
    normalized = pd.DataFrame(
        {
            column: _normalized_static_value_series(frame[column])
            for column in columns
        }
    )
    return _partition_payload_signature(normalized)


def _carry_forward_deprecated_static_columns(
    *,
    active_frame: pd.DataFrame,
    candidate_frame: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Retain deprecated values only when an active config key has one value."""
    columns = sorted(
        column
        for column in DEPRECATED_OPTIONAL_STATIC_COLUMNS
        if column in active_frame.columns
        and (
            column not in candidate_frame.columns
            or bool(_series_missing_value_mask(candidate_frame[column]).any())
        )
        and _series_has_data(active_frame[column])
    )
    summary: dict[str, Any] = {
        "enabled": bool(columns),
        "policy": "consistent_active_key_values_only",
        "columns": columns,
        "keyColumns": [],
        "candidateRowCount": int(len(candidate_frame)),
        "matchedConfigurationRowCount": 0,
        "candidateDuplicateKeyRowCount": 0,
        "activeDuplicateKeyRowCount": 0,
        "columnResults": {},
    }
    if not columns:
        return candidate_frame, summary

    key_columns = _static_carry_forward_key_columns(
        active_frame=active_frame,
        candidate_frame=candidate_frame,
    )
    summary["keyColumns"] = key_columns
    if not key_columns:
        summary["enabled"] = False
        summary["reason"] = "stable_configuration_keys_unavailable"
        return candidate_frame, summary

    active_keys = _normalized_static_key_frame(active_frame, key_columns)
    candidate_keys = _normalized_static_key_frame(candidate_frame, key_columns)
    key_names = list(active_keys.columns)
    summary["activeDuplicateKeyRowCount"] = int(
        active_keys.duplicated(keep=False).sum()
    )
    summary["candidateDuplicateKeyRowCount"] = int(
        candidate_keys.duplicated(keep=False).sum()
    )

    candidate_lookup = candidate_keys.copy()
    candidate_lookup["__candidate_index"] = candidate_lookup.index
    result = candidate_frame.copy()
    matched_candidate_indices: set[Any] = set()
    for index, column in enumerate(columns):
        if column not in result.columns:
            result[column] = pd.NA
        carry_column = f"__carry_{index}"
        value_signature_column = f"__carry_signature_{index}"
        active_values = pd.concat(
            [
                active_keys,
                active_frame[column].rename(carry_column),
                _normalized_static_value_series(active_frame[column]).rename(
                    value_signature_column
                ),
            ],
            axis=1,
        )
        active_values = active_values.loc[
            ~_series_missing_value_mask(active_values[carry_column])
        ]
        distinct_values = active_values.drop_duplicates(
            subset=[*key_names, value_signature_column]
        )
        unambiguous_values = distinct_values.loc[
            ~distinct_values.duplicated(subset=key_names, keep=False)
        ]
        matched = candidate_lookup.reset_index(drop=True).merge(
            unambiguous_values[[*key_names, carry_column]],
            how="inner",
            on=key_names,
            validate="many_to_one",
        )
        candidate_indices = matched["__candidate_index"].tolist()
        current_missing = _series_missing_value_mask(
            result.loc[candidate_indices, column]
        )
        rows_to_fill = matched.loc[current_missing.to_numpy()]
        if not rows_to_fill.empty:
            result.loc[
                rows_to_fill["__candidate_index"].tolist(),
                column,
            ] = rows_to_fill[carry_column].tolist()
        matched_candidate_indices.update(rows_to_fill["__candidate_index"].tolist())
        summary["columnResults"][column] = {
            "inheritedRowCount": int(len(rows_to_fill)),
            "ambiguousActiveKeyCount": int(
                distinct_values.loc[
                    distinct_values.duplicated(
                        subset=key_names,
                        keep=False,
                    ),
                    key_names,
                ]
                .drop_duplicates()
                .shape[0]
            ),
            "remainingNullRowCount": int(
                _series_missing_value_mask(result[column]).sum()
            ),
        }
    summary["matchedConfigurationRowCount"] = len(matched_candidate_indices)
    return result, summary


def _smart_merge_dataframes(
    *,
    active_path: Path,
    candidate_path: Path,
    regressed_countries: list[dict[str, str | None]],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Merge active + candidate parquet data for Smart Merge.

    - Regressed countries: use active data (has more recent months)
    - Advanced/unchanged/new countries: use candidate data
    - Countries only in active, missing in candidate: use active data
    """
    active_df = _load_dataset_frame(active_path)
    candidate_df = _load_dataset_frame(candidate_path)
    country_col = _find_country_column(list(active_df.columns))
    if country_col is None:
        raise HTTPException(status_code=409, detail="无法识别 active 数据集的国家列。")

    active_df[country_col] = active_df[country_col].astype("string").fillna("").str.strip()
    candidate_df[country_col] = candidate_df[country_col].astype("string").fillna("").str.strip()
    active_country_display = _logical_country_display_map(
        active_df[country_col],
        path_label="Smart Merge active 数据集",
    )
    candidate_country_display = _logical_country_display_map(
        candidate_df[country_col],
        path_label="Smart Merge candidate 数据集",
    )
    active_country_keys = active_df[country_col].str.casefold()
    candidate_country_keys = candidate_df[country_col].str.casefold()
    candidate_df[country_col] = [
        active_country_display.get(str(country_key), display)
        for country_key, display in zip(
            candidate_country_keys,
            candidate_df[country_col],
            strict=False,
        )
    ]
    candidate_country_keys = candidate_df[country_col].str.casefold()

    regressed_set: set[str] = set()
    for entry in regressed_countries:
        c = str(entry.get("country", "")).strip()
        if c:
            regressed_set.add(c.casefold())

    active_countries = set(active_country_display)
    candidate_countries = set(candidate_country_display)
    missing_from_candidate = active_countries - candidate_countries

    # From candidate: keep rows for non-regressed countries
    candidate_keep = candidate_df[
        ~candidate_country_keys.isin(regressed_set)
    ].copy()
    # From active: keep rows for regressed countries + countries missing in candidate
    active_keep = active_df[
        active_country_keys.isin(regressed_set | missing_from_candidate)
    ].copy()
    candidate_keep_keys = set(
        candidate_keep[country_col].str.casefold().unique()
    )
    candidate_keep, deprecated_static_summary = (
        _carry_forward_deprecated_static_columns(
            active_frame=active_df[
                active_country_keys.isin(candidate_keep_keys)
            ],
            candidate_frame=candidate_keep,
        )
    )

    # Align columns: union of all columns from both dataframes
    all_columns = list(dict.fromkeys(list(active_df.columns) + list(candidate_df.columns)))
    for df in (candidate_keep, active_keep):
        for col in all_columns:
            if col not in df.columns:
                df[col] = None

    merged = pd.concat([candidate_keep, active_keep], ignore_index=True)
    return (
        merged[[col for col in all_columns if col in merged.columns]],
        deprecated_static_summary,
    )


def _keep_active_history_country_frame(
    *,
    active_frame: pd.DataFrame,
    candidate_frame: pd.DataFrame,
    active_latest_month: str | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Use active through its latest month and candidate only afterwards.

    Rows are combined only when every normalized static configuration field is
    exactly equal.  Similar model/version labels are never guessed or merged.
    """
    active_latest = active_latest_month or _latest_month_from_frame(
        active_frame
    )
    if active_latest is None:
        raise HTTPException(
            status_code=409,
            detail="keep_active 无法识别 active 最新月份。",
        )
    all_columns = list(
        dict.fromkeys(
            [
                *[str(column) for column in active_frame.columns],
                *[str(column) for column in candidate_frame.columns],
            ]
        )
    )
    active = active_frame.copy()
    candidate = candidate_frame.copy()
    for frame in (active, candidate):
        for column in all_columns:
            if column not in frame.columns:
                frame[column] = None
    active = active[all_columns]
    candidate = candidate[all_columns]
    key_columns = _single_country_configuration_key_columns(
        pd.DataFrame(columns=all_columns)
    )
    if not key_columns:
        raise HTTPException(
            status_code=409,
            detail="keep_active 无法识别精确静态配置键。",
        )
    month_columns = _detect_month_columns(all_columns)
    historical_months = [
        month
        for month in month_columns
        if _time_sort_key(month) <= _time_sort_key(active_latest)
    ]
    future_months = [
        month
        for month in month_columns
        if _time_sort_key(month) > _time_sort_key(active_latest)
    ]
    if not future_months:
        raise HTTPException(
            status_code=409,
            detail=(
                "keep_active candidate 没有晚于 active 的月份，"
                "不能生成推进后的完整 candidate。"
            ),
        )

    active_expected = (
        active[historical_months]
        .apply(pd.to_numeric, errors="coerce")
        .fillna(0)
        .sum()
    )
    candidate_expected = (
        candidate[future_months]
        .apply(pd.to_numeric, errors="coerce")
        .fillna(0)
        .sum()
    )
    for month in future_months:
        active[month] = 0
    for month in historical_months:
        candidate[month] = 0
    candidate_future_sales = (
        candidate[future_months]
        .apply(pd.to_numeric, errors="coerce")
        .fillna(0)
    )
    candidate = candidate.loc[
        candidate_future_sales.ne(0).any(axis=1)
    ].copy()

    for source_name, source_frame in (
        ("active", active),
        ("candidate_future", candidate),
    ):
        source_keys = _normalized_static_key_frame(
            source_frame,
            key_columns,
        )
        duplicate_mask = source_keys.duplicated(keep=False)
        if bool(duplicate_mask.any()):
            duplicate_group_count = int(
                source_keys.loc[duplicate_mask]
                .drop_duplicates()
                .shape[0]
            )
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": "duplicate_configurations",
                    "message": (
                        f"keep_active 检测到 {source_name} 内部存在"
                        "归一化后完全相同的静态配置；"
                        "拒绝自动相加。"
                    ),
                    "source": source_name,
                    "duplicateRows": int(duplicate_mask.sum()),
                    "duplicateGroupCount": duplicate_group_count,
                    "keyColumnCount": len(key_columns),
                },
            )

    combined = pd.DataFrame(
        {
            column: pd.concat(
                [
                    active[column].reset_index(drop=True),
                    candidate[column].reset_index(drop=True),
                ],
                ignore_index=True,
            )
            for column in all_columns
        }
    )
    normalized_keys = _normalized_static_key_frame(
        combined,
        key_columns,
    )
    group_codes, _unique_keys = pd.factorize(
        pd.MultiIndex.from_frame(normalized_keys),
        sort=False,
    )
    grouped_static = combined.groupby(
        group_codes,
        sort=False,
        dropna=False,
    ).first()
    grouped_sales = (
        combined[month_columns]
        .apply(pd.to_numeric, errors="coerce")
        .fillna(0)
        .groupby(group_codes, sort=False)
        .sum()
    )
    for month in month_columns:
        grouped_static[month] = grouped_sales[month]
    merged = grouped_static[all_columns].reset_index(drop=True)
    merged_keys = _normalized_static_key_frame(merged, key_columns)
    duplicate_rows = int(merged_keys.duplicated(keep=False).sum())
    if duplicate_rows:
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "duplicate_configurations",
                "message": (
                    "keep_active 仅允许完整静态键精确合并；"
                    "结果仍有重复配置，拒绝继续。"
                ),
                "duplicateRows": duplicate_rows,
            },
        )

    merged_numeric = (
        merged[month_columns]
        .apply(pd.to_numeric, errors="coerce")
        .fillna(0)
    )
    mismatches: list[dict[str, Any]] = []
    for month in historical_months:
        expected = float(active_expected.get(month, 0) or 0)
        actual = float(merged_numeric[month].sum())
        if actual != expected:
            mismatches.append(
                {
                    "month": month,
                    "expectedSource": "active",
                    "expectedSales": _serialize_numeric_value(expected),
                    "actualSales": _serialize_numeric_value(actual),
                }
            )
    for month in future_months:
        expected = float(candidate_expected.get(month, 0) or 0)
        actual = float(merged_numeric[month].sum())
        if actual != expected:
            mismatches.append(
                {
                    "month": month,
                    "expectedSource": "candidate",
                    "expectedSales": _serialize_numeric_value(expected),
                    "actualSales": _serialize_numeric_value(actual),
                }
            )
    if mismatches:
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "keep_active_month_boundary_invalid",
                "message": (
                    "keep_active 时间切片校验失败；拒绝生成可能累加或"
                    "改写历史的 candidate。"
                ),
                "mismatches": mismatches[:10],
            },
        )
    return (
        merged,
        {
            "policy": "keep_active",
            "activeLatestMonth": active_latest,
            "historicalMonthsFrom": "active",
            "futureMonthsFrom": "candidate",
            "historicalMonthCount": len(historical_months),
            "futureMonthCount": len(future_months),
            "activeInputRows": int(len(active_frame)),
            "candidateInputRows": int(len(candidate_frame)),
            "outputRows": int(len(merged)),
            "exactStaticKeyColumnCount": len(key_columns),
            "monthBoundaryCheck": "pass",
        },
    )


def _smart_merge_country_value_map(
    path: Path,
    *,
    country_column: str,
    path_label: str,
) -> dict[str, dict[str, Any]]:
    """Collect only distinct country tokens with a bounded Arrow scan."""
    import pyarrow.parquet as pq

    display_variants: dict[str, list[str]] = {}
    raw_values: dict[str, list[str]] = {}
    missing_country_rows = 0
    parquet_file = pq.ParquetFile(path)
    for batch in parquet_file.iter_batches(
        batch_size=SMART_MERGE_SCAN_BATCH_ROWS,
        columns=[country_column],
        use_threads=False,
    ):
        for raw_value in batch.column(0).to_pylist():
            if raw_value is None:
                missing_country_rows += 1
                continue
            raw = str(raw_value)
            display = raw.strip()
            key = display.casefold()
            if not key:
                missing_country_rows += 1
                continue
            variants = display_variants.setdefault(key, [])
            if display not in variants:
                variants.append(display)
            values = raw_values.setdefault(key, [])
            if raw not in values:
                values.append(raw)
    _raise_for_missing_country_rows(
        path_label=path_label,
        row_count=missing_country_rows,
    )
    ambiguous = {
        key: variants
        for key, variants in display_variants.items()
        if len(variants) > 1
    }
    if ambiguous:
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "ambiguous_logical_country",
                "message": (
                    f"{path_label} 中同一逻辑国家存在多个大小写/空格"
                    "展示值，继续会造成国家重复累加；请先统一国家字段。"
                ),
                "countries": [
                    {
                        "logicalKey": key,
                        "displayValues": variants,
                    }
                    for key, variants in sorted(ambiguous.items())
                ],
            },
        )
    return {
        key: {
            "display": variants[0],
            "rawValues": raw_values.get(key, []),
        }
        for key, variants in display_variants.items()
    }


def _iter_smart_merge_country_tables(
    path: Path,
    *,
    country_column: str,
    raw_values: list[str],
) -> Any:
    """Yield one logical country in fixed-size, low-readahead Arrow batches."""
    import pyarrow as pa
    import pyarrow.dataset as ds

    if not raw_values:
        return
    scanner = ds.dataset(path, format="parquet").scanner(
        filter=ds.field(country_column).isin(raw_values),
        batch_size=SMART_MERGE_SCAN_BATCH_ROWS,
        use_threads=False,
        batch_readahead=1,
        fragment_readahead=1,
    )
    for batch in scanner.to_batches():
        if batch.num_rows:
            yield pa.Table.from_batches([batch])


def _align_arrow_table_to_schema(table: Any, schema: Any) -> Any:
    import pyarrow as pa

    arrays: list[Any] = []
    for field in schema:
        if field.name not in table.column_names:
            arrays.append(
                pa.chunked_array(
                    [pa.nulls(table.num_rows, type=field.type)],
                    type=field.type,
                )
            )
            continue
        column = table[field.name]
        if not column.type.equals(field.type):
            column = column.cast(field.type, safe=False)
        arrays.append(column)
    return pa.Table.from_arrays(arrays, schema=schema)


def _empty_smart_merge_static_summary() -> dict[str, Any]:
    return {
        "enabled": False,
        "policy": "consistent_active_key_values_only",
        "columns": [],
        "keyColumns": [],
        "candidateRowCount": 0,
        "matchedConfigurationRowCount": 0,
        "candidateDuplicateKeyRowCount": 0,
        "activeDuplicateKeyRowCount": 0,
        "columnResults": {},
    }


def _merge_smart_merge_static_summary(
    target: dict[str, Any],
    source: dict[str, Any],
) -> None:
    target["enabled"] = bool(
        target.get("enabled") or source.get("enabled")
    )
    for counter_name in (
        "candidateRowCount",
        "matchedConfigurationRowCount",
        "candidateDuplicateKeyRowCount",
        "activeDuplicateKeyRowCount",
    ):
        target[counter_name] = int(target.get(counter_name, 0) or 0) + int(
            source.get(counter_name, 0) or 0
        )
    target["columns"] = sorted(
        {
            *[str(value) for value in target.get("columns", [])],
            *[str(value) for value in source.get("columns", [])],
        }
    )
    if not target.get("keyColumns") and source.get("keyColumns"):
        target["keyColumns"] = [
            str(value) for value in source.get("keyColumns", [])
        ]
    target_results = target.setdefault("columnResults", {})
    for column, raw_result in (
        source.get("columnResults", {})
        if isinstance(source.get("columnResults"), dict)
        else {}
    ).items():
        result = target_results.setdefault(
            str(column),
            {
                "inheritedRowCount": 0,
                "ambiguousActiveKeyCount": 0,
                "remainingNullRowCount": 0,
            },
        )
        if not isinstance(raw_result, dict):
            continue
        for counter_name in result:
            result[counter_name] += int(
                raw_result.get(counter_name, 0) or 0
            )


def _smart_merge_bucket_key_columns(
    *,
    active_schema: Any,
    candidate_schema: Any,
    output_schema: Any,
) -> list[str]:
    active_columns = set(active_schema.names)
    candidate_columns = set(candidate_schema.names)
    shared = [
        column
        for column in STATIC_CARRY_FORWARD_KEY_CANDIDATES
        if column in active_columns and column in candidate_columns
    ]
    if (
        "Make" in shared
        and "Model" in shared
        and ("Version name" in shared or "Trim level" in shared)
    ):
        return shared
    return _single_country_configuration_key_columns(
        pd.DataFrame(columns=output_schema.names)
    )


def _spill_smart_merge_country_to_buckets(
    *,
    source_path: Path,
    country_column: str,
    raw_values: list[str],
    output_country: str | None,
    output_schema: Any,
    bucket_columns: list[str],
    bucket_root: Path,
    source_name: str,
) -> dict[str, Any]:
    import pyarrow as pa
    import pyarrow.parquet as pq

    if not bucket_columns:
        raise HTTPException(
            status_code=409,
            detail="Smart Merge 无法识别稳定的配置分桶键。",
        )
    counts = [0] * SMART_MERGE_HASH_BUCKET_COUNT
    part_numbers = [0] * SMART_MERGE_HASH_BUCKET_COUNT
    max_input_batch_rows = 0
    total_rows = 0
    for raw_table in _iter_smart_merge_country_tables(
        source_path,
        country_column=country_column,
        raw_values=raw_values,
    ):
        max_input_batch_rows = max(
            max_input_batch_rows,
            int(raw_table.num_rows),
        )
        table = _align_arrow_table_to_schema(raw_table, output_schema)
        frame = table.to_pandas(use_threads=False)
        if output_country is not None:
            frame[country_column] = output_country
        normalized_keys = pd.DataFrame(
            {
                f"__bucket_key_{index}": (
                    _normalized_static_value_series(frame[column])
                )
                for index, column in enumerate(bucket_columns)
            }
        )
        bucket_ids = pd.util.hash_pandas_object(
            normalized_keys,
            index=False,
            categorize=True,
        ).map(
            lambda value: int(value) % SMART_MERGE_HASH_BUCKET_COUNT
        )
        normalized_table = pa.Table.from_pandas(
            frame[output_schema.names],
            schema=output_schema,
            preserve_index=False,
            safe=False,
        )
        for bucket_id in sorted(int(value) for value in bucket_ids.unique()):
            positions = [
                int(position)
                for position in bucket_ids.index[bucket_ids == bucket_id]
            ]
            bucket_table = normalized_table.take(
                pa.array(positions, type=pa.int64())
            )
            destination = (
                bucket_root
                / source_name
                / f"bucket-{bucket_id:03d}"
                / f"part-{part_numbers[bucket_id]:06d}.parquet"
            )
            destination.parent.mkdir(parents=True, exist_ok=True)
            pq.write_table(
                bucket_table,
                destination,
                compression="snappy",
                use_dictionary=True,
            )
            part_numbers[bucket_id] += 1
            counts[bucket_id] += int(bucket_table.num_rows)
            total_rows += int(bucket_table.num_rows)
        del normalized_table
        del normalized_keys
        del frame
        del table
    return {
        "counts": counts,
        "rowCount": total_rows,
        "maxInputBatchRows": max_input_batch_rows,
    }


def _read_smart_merge_bucket_frame(
    bucket_path: Path,
    *,
    output_schema: Any,
) -> pd.DataFrame:
    import pyarrow.parquet as pq

    if not bucket_path.exists():
        return output_schema.empty_table().to_pandas()
    table = pq.read_table(bucket_path, use_threads=False)
    table = _align_arrow_table_to_schema(table, output_schema)
    return table.to_pandas(use_threads=False)


def _write_smart_merge_frame(
    *,
    writer: Any,
    frame: pd.DataFrame,
    output_schema: Any,
) -> int:
    import pyarrow as pa

    for field in output_schema:
        if field.name not in frame.columns:
            frame[field.name] = None
    table = pa.Table.from_pandas(
        frame[output_schema.names],
        schema=output_schema,
        preserve_index=False,
        safe=False,
    )
    writer.write_table(table)
    return int(table.num_rows)


def _stream_smart_merge_country_to_writer(
    *,
    writer: Any,
    source_path: Path,
    country_column: str,
    raw_values: list[str],
    output_country: str | None,
    source_schema: Any,
    output_schema: Any,
    verify_untouched: bool,
) -> tuple[int, dict[str, Any] | None, int]:
    import pyarrow as pa

    source_signature = 0
    output_signature = 0
    signature_modulus = 1 << 64
    candidate_only_columns = [
        column
        for column in output_schema.names
        if column not in source_schema.names
    ]
    non_null_candidate_only: set[str] = set()
    row_count = 0
    max_input_batch_rows = 0
    for raw_table in _iter_smart_merge_country_tables(
        source_path,
        country_column=country_column,
        raw_values=raw_values,
    ):
        max_input_batch_rows = max(
            max_input_batch_rows,
            int(raw_table.num_rows),
        )
        if verify_untouched:
            source_frame = raw_table.select(
                source_schema.names
            ).to_pandas(use_threads=False)
            source_signature = (
                source_signature
                + int(
                    _canonical_country_content_signature(
                        source_frame,
                        list(source_schema.names),
                    )
                )
            ) % signature_modulus
        table = _align_arrow_table_to_schema(raw_table, output_schema)
        if output_country is not None:
            frame = table.to_pandas(use_threads=False)
            frame[country_column] = output_country
            table = _align_arrow_table_to_schema(
                pa.Table.from_pandas(
                    frame[output_schema.names],
                    schema=output_schema,
                    preserve_index=False,
                    safe=False,
                ),
                output_schema,
            )
            del frame
        if verify_untouched:
            output_frame = table.select(
                source_schema.names
            ).to_pandas(use_threads=False)
            output_signature = (
                output_signature
                + int(
                    _canonical_country_content_signature(
                        output_frame,
                        list(source_schema.names),
                    )
                )
            ) % signature_modulus
            for column in candidate_only_columns:
                if table[column].null_count != table.num_rows:
                    non_null_candidate_only.add(column)
        writer.write_table(table)
        row_count += int(table.num_rows)
    if not verify_untouched:
        return row_count, None, max_input_batch_rows
    if source_signature != output_signature or non_null_candidate_only:
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "untouched_country_changed",
                "message": "未选择更新的国家在 Smart Merge 输出中发生变化。",
                "nonNullCandidateOnlyColumns": sorted(
                    non_null_candidate_only
                ),
            },
        )
    return (
        row_count,
        {
            "status": "pass",
            "rowCount": row_count,
            "canonicalSignature": str(output_signature),
            "candidateOnlyColumnsNull": True,
        },
        max_input_batch_rows,
    )


def _legacy_smart_merge_parquet_by_country(
    *,
    active_path: Path,
    candidate_path: Path,
    regressed_countries: list[dict[str, str | None]],
    historical_reclassification_decisions: dict[str, str] | None = None,
) -> tuple[int, dict[str, Any]]:
    """Write Smart Merge one logical country at a time.

    This keeps the worker peak near one country partition instead of holding
    active + candidate + two copies + concat for the ~1M × 96 archive.
    """
    import pyarrow as pa
    import pyarrow.parquet as pq

    active_schema = pq.read_schema(active_path)
    candidate_schema = pq.read_schema(candidate_path)
    active_country_column = _find_country_column(
        [str(column).strip() for column in active_schema.names]
    )
    candidate_country_column = _find_country_column(
        [str(column).strip() for column in candidate_schema.names]
    )
    if active_country_column is None or candidate_country_column is None:
        raise HTTPException(
            status_code=409,
            detail="Smart Merge 无法识别 active/candidate 国家列。",
        )
    if active_country_column != candidate_country_column:
        raise HTTPException(
            status_code=409,
            detail=(
                "Smart Merge 的 active/candidate 国家列名不一致；"
                "为避免跨国家错误补值，已拒绝自动兼容。"
            ),
        )
    try:
        output_schema = pa.unify_schemas(
            [active_schema, candidate_schema],
            promote_options="permissive",
        )
    except Exception as exc:
        raise HTTPException(
            status_code=409,
            detail=f"Smart Merge schema 无法安全统一：{exc}",
        ) from exc

    active_latest = _collect_dataset_country_latest_months(active_path)
    candidate_latest = _collect_dataset_country_latest_months(candidate_path)
    active_by_key = {
        country.casefold(): country
        for country in active_latest
    }
    candidate_by_key = {
        country.casefold(): country
        for country in candidate_latest
    }
    regressed_keys = {
        str(entry.get("country") or "").strip().casefold()
        for entry in regressed_countries
        if str(entry.get("country") or "").strip()
    }
    decision_map = {
        str(country).strip().casefold(): str(decision).strip().lower()
        for country, decision in (
            historical_reclassification_decisions or {}
        ).items()
    }
    invalid_decisions = {
        decision
        for decision in decision_map.values()
        if decision not in HISTORICAL_RECLASSIFICATION_DECISIONS
    }
    if invalid_decisions:
        raise HTTPException(
            status_code=409,
            detail="Smart Merge 收到无效历史重分类决策。",
        )
    output_country_keys = [
        *active_by_key,
        *(
            key
            for key in candidate_by_key
            if key not in active_by_key
        ),
    ]
    temp_output = candidate_path.with_name(
        f".{candidate_path.name}.{uuid4().hex}.smart-merge.tmp"
    )
    writer: pq.ParquetWriter | None = None
    row_count = 0
    aggregate_summary: dict[str, Any] = {
        "enabled": False,
        "policy": "consistent_active_key_values_only",
        "streamingByCountry": True,
        "columns": [],
        "keyColumns": [],
        "candidateRowCount": 0,
        "matchedConfigurationRowCount": 0,
        "candidateDuplicateKeyRowCount": 0,
        "activeDuplicateKeyRowCount": 0,
        "columnResults": {},
        "countryResults": {},
        "historicalReclassificationPolicies": {},
        "untouchedCountryChecks": {},
    }
    try:
        writer = pq.ParquetWriter(
            temp_output,
            output_schema,
            compression="snappy",
        )
        for country_key in output_country_keys:
            untouched_check_context: dict[str, Any] | None = None
            use_candidate = (
                country_key in candidate_by_key
                and country_key not in regressed_keys
            )
            if use_candidate:
                candidate_country = candidate_by_key[country_key]
                frame = _load_parquet_country_subset(
                    candidate_path,
                    candidate_country,
                    path_label=f"Smart Merge candidate（{candidate_country}）",
                )
                active_country = active_by_key.get(country_key)
                if active_country is not None:
                    active_frame = _load_parquet_country_subset(
                        active_path,
                        active_country,
                        path_label=f"Smart Merge active（{active_country}）",
                    )
                    if candidate_country_column != active_country_column:
                        frame = frame.rename(
                            columns={
                                candidate_country_column: active_country_column
                            }
                        )
                    frame[active_country_column] = active_country
                    frame, country_summary = (
                        _carry_forward_deprecated_static_columns(
                            active_frame=active_frame,
                            candidate_frame=frame,
                        )
                    )
                    history_policy = decision_map.get(country_key)
                    if history_policy == "keep_active":
                        frame, history_policy_summary = (
                            _keep_active_history_country_frame(
                                active_frame=active_frame,
                                candidate_frame=frame,
                            )
                        )
                    else:
                        history_policy_summary = {
                            "policy": (
                                history_policy or "use_candidate"
                            ),
                            "historicalMonthsFrom": "candidate",
                            "monthBoundaryCheck": (
                                "not_applicable"
                            ),
                        }
                    aggregate_summary[
                        "historicalReclassificationPolicies"
                    ][active_country] = history_policy_summary
                    del active_frame
                else:
                    country_summary = {
                        "enabled": False,
                        "columns": [],
                        "keyColumns": [],
                        "candidateRowCount": int(len(frame)),
                        "matchedConfigurationRowCount": 0,
                        "candidateDuplicateKeyRowCount": 0,
                        "activeDuplicateKeyRowCount": 0,
                        "columnResults": {},
                    }
                output_country = active_country or candidate_country
            else:
                output_country = active_by_key[country_key]
                frame = _load_parquet_country_subset(
                    active_path,
                    output_country,
                    path_label=f"Smart Merge active（{output_country}）",
                )
                country_summary = {
                    "enabled": False,
                    "columns": [],
                    "keyColumns": [],
                    "candidateRowCount": 0,
                    "matchedConfigurationRowCount": 0,
                    "candidateDuplicateKeyRowCount": 0,
                    "activeDuplicateKeyRowCount": 0,
                    "columnResults": {},
                }
                original_columns = [
                    column
                    for column in active_schema.names
                    if column in frame.columns
                ]
                untouched_check_context = {
                    "sourceSignature": (
                        _canonical_country_content_signature(
                            frame,
                            original_columns,
                        )
                    ),
                    "sourceColumns": original_columns,
                    "candidateOnlyColumns": [
                        column
                        for column in candidate_schema.names
                        if column not in active_schema.names
                    ],
                    "rowCount": int(len(frame)),
                }

            aggregate_summary["countryResults"][output_country] = (
                country_summary
            )
            aggregate_summary["enabled"] = bool(
                aggregate_summary["enabled"]
                or country_summary.get("enabled")
            )
            aggregate_summary["candidateRowCount"] += int(
                country_summary.get("candidateRowCount", 0) or 0
            )
            aggregate_summary["matchedConfigurationRowCount"] += int(
                country_summary.get(
                    "matchedConfigurationRowCount",
                    0,
                )
                or 0
            )
            for counter_name in (
                "candidateDuplicateKeyRowCount",
                "activeDuplicateKeyRowCount",
            ):
                aggregate_summary[counter_name] += int(
                    country_summary.get(counter_name, 0) or 0
                )
            aggregate_summary["columns"] = sorted(
                set(aggregate_summary["columns"])
                | {
                    str(column)
                    for column in country_summary.get("columns", [])
                }
            )
            if not aggregate_summary["keyColumns"]:
                aggregate_summary["keyColumns"] = [
                    str(column)
                    for column in country_summary.get("keyColumns", [])
                ]
            for column, raw_result in (
                country_summary.get("columnResults", {})
                if isinstance(
                    country_summary.get("columnResults"),
                    dict,
                )
                else {}
            ).items():
                result = aggregate_summary["columnResults"].setdefault(
                    str(column),
                    {
                        "inheritedRowCount": 0,
                        "ambiguousActiveKeyCount": 0,
                        "remainingNullRowCount": 0,
                    },
                )
                if isinstance(raw_result, dict):
                    for counter_name in result:
                        result[counter_name] += int(
                            raw_result.get(counter_name, 0) or 0
                        )

            for field in output_schema:
                if field.name not in frame.columns:
                    frame[field.name] = None
            frame = frame[output_schema.names]
            table = pa.Table.from_pandas(
                frame,
                schema=output_schema,
                preserve_index=False,
                safe=False,
            )
            if untouched_check_context is not None:
                source_columns = untouched_check_context[
                    "sourceColumns"
                ]
                output_original_frame = table.select(
                    source_columns
                ).to_pandas()
                output_signature = (
                    _canonical_country_content_signature(
                        output_original_frame,
                        source_columns,
                    )
                )
                non_null_candidate_only_columns = [
                    column
                    for column in untouched_check_context[
                        "candidateOnlyColumns"
                    ]
                    if table[column].null_count != table.num_rows
                ]
                if (
                    output_signature
                    != untouched_check_context["sourceSignature"]
                    or non_null_candidate_only_columns
                ):
                    raise HTTPException(
                        status_code=409,
                        detail={
                            "blockerType": (
                                "untouched_country_changed"
                            ),
                            "message": (
                                f"{output_country} 未选择更新但输出内容变化。"
                            ),
                            "nonNullCandidateOnlyColumns": (
                                non_null_candidate_only_columns
                            ),
                        },
                    )
                aggregate_summary["untouchedCountryChecks"][
                    output_country
                ] = {
                    "status": "pass",
                    "rowCount": untouched_check_context[
                        "rowCount"
                    ],
                    "canonicalSignature": output_signature,
                    "candidateOnlyColumnsNull": True,
                }
            writer.write_table(table)
            row_count += int(table.num_rows)
            del table
            del frame
        writer.close()
        writer = None
        if row_count <= 0:
            raise HTTPException(
                status_code=409,
                detail="Smart Merge 结果为空，拒绝覆盖 candidate。",
            )
        os.replace(temp_output, candidate_path)
    finally:
        if writer is not None:
            writer.close()
        temp_output.unlink(missing_ok=True)
    return row_count, aggregate_summary


def _smart_merge_parquet_streaming(
    *,
    active_path: Path,
    candidate_path: Path,
    regressed_countries: list[dict[str, str | None]],
    historical_reclassification_decisions: dict[str, str] | None = None,
    output_path: Path | None = None,
) -> tuple[int, dict[str, Any]]:
    """Build a complete candidate with bounded Arrow scans and disk spill."""
    import pyarrow as pa
    import pyarrow.parquet as pq

    active_schema = pq.read_schema(active_path)
    candidate_schema = pq.read_schema(candidate_path)
    active_country_column = _find_country_column(
        [str(column).strip() for column in active_schema.names]
    )
    candidate_country_column = _find_country_column(
        [str(column).strip() for column in candidate_schema.names]
    )
    if active_country_column is None or candidate_country_column is None:
        raise HTTPException(
            status_code=409,
            detail="Smart Merge 无法识别 active/candidate 国家列。",
        )
    if active_country_column != candidate_country_column:
        raise HTTPException(
            status_code=409,
            detail=(
                "Smart Merge 的 active/candidate 国家列名不一致；"
                "为避免跨国家错误补值，已拒绝自动兼容。"
            ),
        )
    try:
        output_schema = pa.unify_schemas(
            [active_schema, candidate_schema],
            promote_options="permissive",
        )
    except Exception as exc:
        raise HTTPException(
            status_code=409,
            detail=f"Smart Merge schema 无法安全统一：{exc}",
        ) from exc

    active_latest = _collect_dataset_country_latest_months(active_path)
    candidate_latest = _collect_dataset_country_latest_months(candidate_path)
    active_country_values = _smart_merge_country_value_map(
        active_path,
        country_column=active_country_column,
        path_label="Smart Merge active 数据集",
    )
    candidate_country_values = _smart_merge_country_value_map(
        candidate_path,
        country_column=candidate_country_column,
        path_label="Smart Merge candidate 数据集",
    )
    active_by_key = {
        country.casefold(): str(
            active_country_values[country.casefold()]["display"]
        )
        for country in active_latest
    }
    candidate_by_key = {
        country.casefold(): str(
            candidate_country_values[country.casefold()]["display"]
        )
        for country in candidate_latest
    }
    regressed_keys = {
        str(entry.get("country") or "").strip().casefold()
        for entry in regressed_countries
        if str(entry.get("country") or "").strip()
    }
    decision_map = {
        str(country).strip().casefold(): str(decision).strip().lower()
        for country, decision in (
            historical_reclassification_decisions or {}
        ).items()
    }
    if any(
        decision not in HISTORICAL_RECLASSIFICATION_DECISIONS
        for decision in decision_map.values()
    ):
        raise HTTPException(
            status_code=409,
            detail="Smart Merge 收到无效历史重分类决策。",
        )

    output_country_keys = [
        *active_by_key,
        *(key for key in candidate_by_key if key not in active_by_key),
    ]
    destination_path = output_path or candidate_path
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    temp_output = destination_path.with_name(
        f".{destination_path.name}.{uuid4().hex}.smart-merge.tmp"
    )
    spill_root = destination_path.with_name(
        f".{destination_path.name}.{uuid4().hex}.smart-merge-spill"
    )
    bucket_columns = _smart_merge_bucket_key_columns(
        active_schema=active_schema,
        candidate_schema=candidate_schema,
        output_schema=output_schema,
    )
    writer: pq.ParquetWriter | None = None
    row_count = 0
    aggregate_summary: dict[str, Any] = {
        **_empty_smart_merge_static_summary(),
        "streamingByCountry": True,
        "countryResults": {},
        "historicalReclassificationPolicies": {},
        "untouchedCountryChecks": {},
        "resourceProfile": {
            "scanBatchRows": SMART_MERGE_SCAN_BATCH_ROWS,
            "hashBucketCount": SMART_MERGE_HASH_BUCKET_COUNT,
            "maxBucketRows": 0,
            "maxInputBatchRows": 0,
            "spilledRows": 0,
        },
    }

    try:
        writer = pq.ParquetWriter(
            temp_output,
            output_schema,
            compression="snappy",
        )
        for country_key in output_country_keys:
            use_candidate = (
                country_key in candidate_by_key
                and country_key not in regressed_keys
            )
            country_summary = _empty_smart_merge_static_summary()
            if not use_candidate:
                output_country = active_by_key[country_key]
                direct_rows, untouched_check, max_batch = (
                    _stream_smart_merge_country_to_writer(
                        writer=writer,
                        source_path=active_path,
                        country_column=active_country_column,
                        raw_values=list(
                            active_country_values[country_key]["rawValues"]
                        ),
                        output_country=None,
                        source_schema=active_schema,
                        output_schema=output_schema,
                        verify_untouched=True,
                    )
                )
                row_count += direct_rows
                if untouched_check is not None:
                    aggregate_summary["untouchedCountryChecks"][
                        output_country
                    ] = untouched_check
                aggregate_summary["resourceProfile"][
                    "maxInputBatchRows"
                ] = max(
                    int(
                        aggregate_summary["resourceProfile"][
                            "maxInputBatchRows"
                        ]
                    ),
                    max_batch,
                )
            else:
                candidate_country = candidate_by_key[country_key]
                active_country = active_by_key.get(country_key)
                if active_country is None:
                    output_country = candidate_country
                    direct_rows, _untouched, max_batch = (
                        _stream_smart_merge_country_to_writer(
                            writer=writer,
                            source_path=candidate_path,
                            country_column=candidate_country_column,
                            raw_values=list(
                                candidate_country_values[country_key][
                                    "rawValues"
                                ]
                            ),
                            output_country=None,
                            source_schema=candidate_schema,
                            output_schema=output_schema,
                            verify_untouched=False,
                        )
                    )
                    row_count += direct_rows
                    country_summary["candidateRowCount"] = direct_rows
                    aggregate_summary["resourceProfile"][
                        "maxInputBatchRows"
                    ] = max(
                        int(
                            aggregate_summary["resourceProfile"][
                                "maxInputBatchRows"
                            ]
                        ),
                        max_batch,
                    )
                else:
                    output_country = active_country
                    country_spill_root = spill_root / hashlib.sha256(
                        country_key.encode("utf-8")
                    ).hexdigest()[:16]
                    active_spill = _spill_smart_merge_country_to_buckets(
                        source_path=active_path,
                        country_column=active_country_column,
                        raw_values=list(
                            active_country_values[country_key]["rawValues"]
                        ),
                        output_country=None,
                        output_schema=output_schema,
                        bucket_columns=bucket_columns,
                        bucket_root=country_spill_root,
                        source_name="active",
                    )
                    candidate_spill = _spill_smart_merge_country_to_buckets(
                        source_path=candidate_path,
                        country_column=candidate_country_column,
                        raw_values=list(
                            candidate_country_values[country_key]["rawValues"]
                        ),
                        output_country=active_country,
                        output_schema=output_schema,
                        bucket_columns=bucket_columns,
                        bucket_root=country_spill_root,
                        source_name="candidate",
                    )
                    resource_profile = aggregate_summary["resourceProfile"]
                    resource_profile["spilledRows"] += int(
                        active_spill["rowCount"]
                    ) + int(candidate_spill["rowCount"])
                    resource_profile["maxInputBatchRows"] = max(
                        int(resource_profile["maxInputBatchRows"]),
                        int(active_spill["maxInputBatchRows"]),
                        int(candidate_spill["maxInputBatchRows"]),
                    )
                    history_policy = decision_map.get(country_key)
                    history_summary: dict[str, Any] = {
                        "policy": history_policy or "use_candidate",
                        "historicalMonthsFrom": "candidate",
                        "monthBoundaryCheck": "not_applicable",
                    }
                    for bucket_id in range(
                        SMART_MERGE_HASH_BUCKET_COUNT
                    ):
                        bucket_rows = int(
                            active_spill["counts"][bucket_id]
                        ) + int(candidate_spill["counts"][bucket_id])
                        if bucket_rows <= 0:
                            continue
                        resource_profile["maxBucketRows"] = max(
                            int(resource_profile["maxBucketRows"]),
                            bucket_rows,
                        )
                        if bucket_rows > SMART_MERGE_MAX_BUCKET_ROWS:
                            raise HTTPException(
                                status_code=503,
                                detail={
                                    "blockerType": (
                                        "smart_merge_bucket_resource_guard"
                                    ),
                                    "message": (
                                        f"{active_country} 的单个配置桶有"
                                        f" {bucket_rows} 行，已在进入 pandas"
                                        " 前停止；Candidate 与 active 均未修改。"
                                    ),
                                    "country": active_country,
                                    "bucketRows": bucket_rows,
                                    "maxBucketRows": (
                                        SMART_MERGE_MAX_BUCKET_ROWS
                                    ),
                                    "sourceFeedback": None,
                                    "nextAction": "split_smart_merge_bucket",
                                },
                            )
                        active_frame = _read_smart_merge_bucket_frame(
                            country_spill_root
                            / "active"
                            / f"bucket-{bucket_id:03d}",
                            output_schema=output_schema,
                        )
                        candidate_frame = _read_smart_merge_bucket_frame(
                            country_spill_root
                            / "candidate"
                            / f"bucket-{bucket_id:03d}",
                            output_schema=output_schema,
                        )
                        candidate_frame, bucket_summary = (
                            _carry_forward_deprecated_static_columns(
                                active_frame=active_frame,
                                candidate_frame=candidate_frame,
                            )
                        )
                        _merge_smart_merge_static_summary(
                            country_summary,
                            bucket_summary,
                        )
                        if history_policy == "keep_active":
                            merged_frame, bucket_history = (
                                _keep_active_history_country_frame(
                                    active_frame=active_frame,
                                    candidate_frame=candidate_frame,
                                    active_latest_month=active_latest.get(
                                        active_country
                                    ),
                                )
                            )
                            if "activeInputRows" not in history_summary:
                                history_summary = {
                                    **bucket_history,
                                    "activeInputRows": 0,
                                    "candidateInputRows": 0,
                                    "outputRows": 0,
                                }
                            for counter_name in (
                                "activeInputRows",
                                "candidateInputRows",
                                "outputRows",
                            ):
                                history_summary[counter_name] += int(
                                    bucket_history.get(counter_name, 0) or 0
                                )
                        else:
                            merged_frame = candidate_frame
                        row_count += _write_smart_merge_frame(
                            writer=writer,
                            frame=merged_frame,
                            output_schema=output_schema,
                        )
                        del active_frame
                        del candidate_frame
                        del merged_frame
                    aggregate_summary[
                        "historicalReclassificationPolicies"
                    ][active_country] = history_summary
                    _remove_file_or_tree(country_spill_root)

            aggregate_summary["countryResults"][output_country] = (
                country_summary
            )
            _merge_smart_merge_static_summary(
                aggregate_summary,
                country_summary,
            )
        writer.close()
        writer = None
        if row_count <= 0:
            raise HTTPException(
                status_code=409,
                detail="Smart Merge 结果为空，拒绝覆盖 candidate。",
            )
        os.replace(temp_output, destination_path)
    finally:
        if writer is not None:
            writer.close()
        temp_output.unlink(missing_ok=True)
        _remove_file_or_tree(spill_root)
    return row_count, aggregate_summary


def _cache_smart_merge_review(
    job_id: str,
    *,
    review_generation_id: str | None,
    expected_active_fingerprint: str,
) -> Path:
    review_kwargs: dict[str, Any] = {
        "expected_active_fingerprint": expected_active_fingerprint,
    }
    if review_generation_id is not None:
        review_kwargs["review_generation_id"] = review_generation_id
    return _cache_jato_monthly_update_review(job_id, **review_kwargs)


def _release_worker_memory_before_review(*, log_path: Path) -> None:
    """Best-effort release of merge allocations before Pandas Review work."""
    notes: list[str] = []
    gc.collect()
    try:
        import pyarrow as pa

        pa.default_memory_pool().release_unused()
        notes.append("arrow_pool=released")
    except Exception as exc:  # pragma: no cover - platform-dependent fallback
        notes.append(f"arrow_pool={type(exc).__name__}")
    if sys.platform.startswith("linux"):
        try:
            import ctypes

            malloc_trim = getattr(ctypes.CDLL(None), "malloc_trim", None)
            if malloc_trim is not None:
                malloc_trim.argtypes = [ctypes.c_size_t]
                malloc_trim.restype = ctypes.c_int
                notes.append(f"malloc_trim={int(malloc_trim(0))}")
            else:  # pragma: no cover - non-glibc Linux
                notes.append("malloc_trim=unavailable")
        except Exception as exc:  # pragma: no cover - platform-dependent
            notes.append(f"malloc_trim={type(exc).__name__}")
    gc.collect()
    _append_log(
        log_path,
        (
            f"[{_utc_now().isoformat()}] Smart Merge: Review 前释放临时内存 "
            + ", ".join(notes)
            + "。"
        ),
    )


def _run_smart_merge(
    job_id: str,
    *,
    review_generation_id: str | None = None,
) -> None:
    """Background runner: smart merge → rebuild partitions/manifest/fingerprint."""
    state = _load_job_state(job_id)
    if str(state.get("status") or "") == "cancelled":
        _RUNNING_THREADS.pop(job_id, None)
        return
    log_path = _job_log_path(job_id)
    working_bundle_dir: Path | None = None
    bundle_committed = False

    try:
        state["status"] = "running"
        state["phase"] = "smart_merging"
        resolution = _historical_reclassification_resolution(state)
        initial_artifacts = state.get("artifacts")
        committed_bundle_resume = bool(
            isinstance(initial_artifacts, dict)
            and str(initial_artifacts.get("candidateScope") or "")
            == "full_smart_merge"
            and isinstance(resolution, dict)
            and str(resolution.get("status") or "") == "resolved"
        )
        if isinstance(resolution, dict) and not committed_bundle_resume:
            resolution["status"] = "running"
            resolution["startedAt"] = _utc_now().isoformat()
            state["historicalReclassificationResolution"] = resolution
        _persist_job_state(state)
        _append_log(log_path, f"[{_utc_now().isoformat()}] Smart Merge: 开始合并数据...")

        active_paths = _active_data_paths()
        artifacts = state.get("artifacts", {}) or {}
        candidate_path = _project_path(str(artifacts.get("stagingOutputPath") or "").strip())

        if candidate_path is None or not candidate_path.exists():
            raise RuntimeError("找不到 candidate staging parquet。")
        if not active_paths["parquet"].exists():
            raise RuntimeError("找不到 active 数据集，无法执行 Smart Merge。")
        resolution = _historical_reclassification_resolution(state)
        candidate_scope = str(artifacts.get("candidateScope") or "")
        if (
            candidate_scope == "full_smart_merge"
            and isinstance(resolution, dict)
            and resolution.get("status") == "resolved"
        ):
            durable_paths = {
                "parquet": candidate_path,
                "manifest": _project_path(artifacts.get("manifestPath")),
                "partition": _project_path(
                    artifacts.get("partitionOutputPath")
                ),
                "fingerprint": _project_path(
                    artifacts.get("fingerprintPath")
                ),
                "refreshReport": _project_path(
                    artifacts.get("refreshReportPath")
                ),
                "summaries": _project_path(
                    artifacts.get("summariesOutputPath")
                ),
            }
            if any(path is None for path in durable_paths.values()):
                raise RuntimeError(
                    "已提交 Smart Merge bundle 缺少路径，拒绝重复合并。"
                )
            _validate_candidate_full_bundle(
                parquet_path=durable_paths["parquet"],
                manifest_path=durable_paths["manifest"],
                partition_path=durable_paths["partition"],
                fingerprint_path=durable_paths["fingerprint"],
                refresh_report_path=durable_paths["refreshReport"],
                summaries_path=durable_paths["summaries"],
            )
            bundle_committed = True
            if _active_dataset_version() != str(
                state.get("activeBaseFingerprint") or ""
            ):
                raise RuntimeError(
                    "Smart Merge bundle 完成后 active lineage 已变化；"
                    "不能补建旧 Review。"
                )
            state["phase"] = "building_review"
            _persist_job_state(state)
            _release_worker_memory_before_review(log_path=log_path)
            _cache_smart_merge_review(
                job_id,
                review_generation_id=review_generation_id,
                expected_active_fingerprint=str(
                    state.get("activeBaseFingerprint") or ""
                ),
            )
            state = _load_job_state(job_id)
            resolution = _historical_reclassification_resolution(state)
            if isinstance(resolution, dict):
                resolution.pop("reviewBuildError", None)
                resolution.pop("reviewBuildFailedAt", None)
                state["historicalReclassificationResolution"] = resolution
            state["status"] = "success"
            state["phase"] = "completed"
            state["finishedAt"] = _utc_now().isoformat()
            state["error"] = None
            state["failureDigest"] = None
            _persist_job_state(state)
            return
        validated_reclassification_decisions: dict[str, str] = {}
        if isinstance(resolution, dict):
            validated_reclassification_decisions = (
                _validated_historical_reclassification_resolution(
                    resolution
                )
            )
            source_candidate_fingerprint = str(
                resolution.get("sourceCandidateFingerprint") or ""
            )
            current_candidate_fingerprint = (
                _candidate_fingerprint_id(artifacts)
            )
            if (
                not source_candidate_fingerprint
                or source_candidate_fingerprint
                != current_candidate_fingerprint
            ):
                raise RuntimeError(
                    "历史分类决策后 candidate 内容已变化；"
                    "旧决策不得应用，请重新生成 Review。"
                )

        regressions = _find_publish_country_regressions(
            active_parquet_path=active_paths["parquet"],
            candidate_parquet_path=candidate_path,
        )
        requires_full_rebuild = candidate_scope in {
            "target_country_partition_only",
            "target_country_partitions_only",
        }
        requires_decision_rebuild = bool(
            validated_reclassification_decisions
        )
        if (
            not regressions
            and not requires_full_rebuild
            and not requires_decision_rebuild
        ):
            _append_log(log_path, f"[{_utc_now().isoformat()}] Smart Merge: 无回归国家，不需要合并。")
            state["phase"] = "building_review"
            _persist_job_state(state)
            _cache_smart_merge_review(
                job_id,
                review_generation_id=review_generation_id,
                expected_active_fingerprint=str(
                    state.get("activeBaseFingerprint") or ""
                ),
            )
            state = _load_job_state(job_id)
            state["status"] = "success"
            state["phase"] = "completed"
            state["finishedAt"] = _utc_now().isoformat()
            state["error"] = None
            state["failureDigest"] = None
            state["summaries"] = {
                **(state.get("summaries") or {}),
                "smartMerge": {
                    "mergedAt": _utc_now().isoformat(),
                    "regressedCountryCount": 0,
                    "regressedCountries": [],
                    "totalRowCount": 0,
                },
            }
            _persist_job_state(state)
            return

        expected_active_fingerprint = str(
            state.get("activeBaseFingerprint") or ""
        ).strip()
        job_dir = _job_dir(job_id)
        working_bundle_dir = job_dir / "smart_merge_candidate_bundle"
        _remove_file_or_tree(working_bundle_dir)
        working_bundle_dir.mkdir(parents=True, exist_ok=False)
        merged_candidate_path = (
            working_bundle_dir / "jato_full_archive.parquet"
        )
        with _exclusive_file_lock(_active_bundle_lock_path()) as acquired:
            if not acquired:
                raise RuntimeError(
                    "active bundle 正在切换，Smart Merge 无法固定同一快照。"
                )
            if (
                not re.fullmatch(
                    r"[0-9a-f]{64}",
                    expected_active_fingerprint,
                )
                or _active_dataset_version()
                != expected_active_fingerprint
            ):
                raise RuntimeError(
                    "Smart Merge 开始前 active lineage 已变化；"
                    "请创建新 attempt。"
                )
            row_count, deprecated_static_summary = (
                _smart_merge_parquet_streaming(
                    active_path=active_paths["parquet"],
                    candidate_path=candidate_path,
                    regressed_countries=regressions,
                    historical_reclassification_decisions=(
                        validated_reclassification_decisions
                    ),
                    output_path=merged_candidate_path,
                )
            )
        regressed_names = sorted(r["country"] for r in regressions if r.get("country"))
        _append_log(
            log_path,
            f"[{_utc_now().isoformat()}] Smart Merge: 合并完成，共 {row_count} 行。"
            f" 回归国家({len(regressed_names)}): {', '.join(regressed_names)}",
        )
        if deprecated_static_summary.get("enabled"):
            inherited_rows = sum(
                int(item.get("inheritedRowCount", 0) or 0)
                for item in deprecated_static_summary.get(
                    "columnResults",
                    {},
                ).values()
                if isinstance(item, dict)
            )
            _append_log(
                log_path,
                f"[{_utc_now().isoformat()}] Smart Merge: 已停用静态字段仅在配置键匹配且 active 旧值一致时沿用，"
                f"columns={len(deprecated_static_summary.get('columns', []))}, "
                f"inheritedCells={inherited_rows}。",
            )

        # Build every derived artifact in an unreferenced directory. The job
        # state remains bound to the original candidate until all six files
        # validate, so a later subprocess failure cannot leave a half-bundle.
        partition_output = working_bundle_dir / "partitioned_dataset_v1"
        manifest_path = working_bundle_dir / "manifest.json"
        fingerprint_path = working_bundle_dir / "dataset_fingerprint.json"
        refresh_report_path = working_bundle_dir / "refresh_job_report.json"
        summaries_output = working_bundle_dir / "summaries"

        if not REBUILD_SCRIPT_PATH.exists():
            raise RuntimeError(f"找不到重建脚本: {REBUILD_SCRIPT_PATH}")
        if not PRECOMPUTE_SUMMARIES_SCRIPT_PATH.exists():
            raise RuntimeError(
                f"找不到 summaries 脚本: {PRECOMPUTE_SUMMARIES_SCRIPT_PATH}"
            )

        rebuild_args = [
            sys.executable,
            str(REBUILD_SCRIPT_PATH),
            "--input-parquet",
            str(merged_candidate_path),
            "--output-dir",
            str(working_bundle_dir / "rebuild"),
            "--partition-output",
            str(partition_output),
            "--manifest",
            str(manifest_path),
            "--fingerprint",
            str(fingerprint_path),
        ]

        state["phase"] = "smart_merge_rebuild"
        _persist_job_state(state)
        _ensure_job_not_cancelled(job_id)
        _run_logged_command(
            label="Smart Merge rebuild",
            args=rebuild_args,
            log_path=log_path,
        )
        import pyarrow.parquet as pq

        rebuilt_parquet = pq.ParquetFile(merged_candidate_path)
        rebuilt_partition_manifest = _read_json(
            partition_output / "manifest.json"
        )
        smart_merge_refresh_report = {
            "jobStatus": "success",
            "jobElapsedSeconds": 0,
            "fullManifest": {
                "rows": int(rebuilt_parquet.metadata.num_rows),
                "columns": int(rebuilt_parquet.metadata.num_columns),
            },
            "partitionManifest": {
                "parquetFileCount": int(
                    rebuilt_partition_manifest.get(
                        "parquetFileCount",
                        0,
                    )
                    or 0
                ),
                "partitionDirectoryCount": int(
                    rebuilt_partition_manifest.get(
                        "partitionDirectoryCount",
                        0,
                    )
                    or 0
                ),
            },
            "incremental": {
                "enabled": False,
                "scope": "full_smart_merge",
                "sourceCandidateScope": candidate_scope,
                "regressedCountries": regressed_names,
            },
        }
        _write_json(refresh_report_path, smart_merge_refresh_report)
        state["phase"] = "smart_merge_summaries"
        _persist_job_state(state)
        temp_summaries_output = summaries_output.with_name(
            f".{summaries_output.name}.smart-merge-{job_id}.tmp"
        )
        _remove_file_or_tree(temp_summaries_output)
        try:
            _run_logged_command(
                label="Smart Merge summaries",
                args=[
                    sys.executable,
                    str(PRECOMPUTE_SUMMARIES_SCRIPT_PATH),
                    "--parquet",
                    str(merged_candidate_path),
                    "--output-dir",
                    str(temp_summaries_output),
                ],
                log_path=log_path,
            )
            _rewrite_staged_summaries_manifest_paths(
                staged_summaries_path=temp_summaries_output,
                active_summaries_path=summaries_output,
            )
            _remove_file_or_tree(summaries_output)
            os.replace(temp_summaries_output, summaries_output)
        finally:
            _remove_file_or_tree(temp_summaries_output)
        _validate_candidate_full_bundle(
            parquet_path=merged_candidate_path,
            manifest_path=manifest_path,
            partition_path=partition_output,
            fingerprint_path=fingerprint_path,
            refresh_report_path=refresh_report_path,
            summaries_path=summaries_output,
        )
        if _active_dataset_version() != expected_active_fingerprint:
            raise RuntimeError(
                "Smart Merge 构建 Review 期间 active 已变化；"
                "candidate 已保留但不能审批，请创建新 attempt。"
            )

        committed_artifacts = {
            **artifacts,
            "stagingOutputPath": _relative_to_project(
                merged_candidate_path
            ),
            "partitionOutputPath": _relative_to_project(
                partition_output
            ),
            "manifestPath": _relative_to_project(manifest_path),
            "fingerprintPath": _relative_to_project(fingerprint_path),
            "refreshReportPath": _relative_to_project(
                refresh_report_path
            ),
            "summariesOutputPath": _relative_to_project(
                summaries_output
            ),
            "candidateScope": "full_smart_merge",
        }
        committed_artifacts.pop("reviewBundlePath", None)
        state["summaries"] = {
            **(state.get("summaries") or {}),
            "refresh": _summarize_refresh_report(
                smart_merge_refresh_report
            ),
            "smartMerge": {
                "mergedAt": _utc_now().isoformat(),
                "regressedCountryCount": len(regressions),
                "regressedCountries": regressed_names,
                "totalRowCount": row_count,
                "deprecatedStaticCarryForward": deprecated_static_summary,
            },
        }
        resolution = _historical_reclassification_resolution(state)
        if isinstance(resolution, dict):
            resolution["status"] = "resolved"
            resolution["resolvedAt"] = _utc_now().isoformat()
            resolution["resolvedCandidateFingerprint"] = (
                _candidate_fingerprint_id(committed_artifacts)
            )
            state["historicalReclassificationResolution"] = resolution
        state["artifacts"] = committed_artifacts
        state["reviewApproval"] = None
        state["phase"] = "building_review"
        _job_review_bundle_path(job_id).unlink(missing_ok=True)
        _persist_job_state(state)
        bundle_committed = True
        del rebuilt_parquet
        _release_worker_memory_before_review(log_path=log_path)
        _cache_smart_merge_review(
            job_id,
            review_generation_id=review_generation_id,
            expected_active_fingerprint=expected_active_fingerprint,
        )
        state = _load_job_state(job_id)
        resolution = _historical_reclassification_resolution(state)
        if isinstance(resolution, dict):
            resolution.pop("reviewBuildError", None)
            resolution.pop("reviewBuildFailedAt", None)
            state["historicalReclassificationResolution"] = resolution
        state["status"] = "success"
        state["phase"] = "completed"
        state["finishedAt"] = _utc_now().isoformat()
        state["error"] = None
        state["failureDigest"] = None
        _persist_job_state(state)
        _append_log(
            log_path,
            f"[{_utc_now().isoformat()}] Smart Merge: 全部完成。请进入 Review → Publish。",
        )

    except _JobCancelled as exc:
        state = _load_job_state(job_id)
        state["status"] = "cancelled"
        state["phase"] = "cancelled"
        state["finishedAt"] = state.get("finishedAt") or _utc_now().isoformat()
        state["error"] = str(exc)
        state["currentProcess"] = None
        _persist_job_state(state)
        _write_jato_etl_pipeline_status(state)
        _append_log(log_path, f"[{_utc_now().isoformat()}] Cancelled: {exc}")
    except Exception as exc:
        state = _load_job_state(job_id)
        failed_phase = str(state.get("phase") or "smart_merge")
        state["status"] = "failed"
        state["phase"] = "smart_merge_failed"
        state["finishedAt"] = _utc_now().isoformat()
        state["error"] = str(exc)
        state["failureDigest"] = _failure_digest_from_exception(
            phase=failed_phase,
            exc=exc,
        )
        resolution = _historical_reclassification_resolution(state)
        if isinstance(resolution, dict):
            if bundle_committed:
                # The complete candidate bundle is durable.  Preserve the
                # resolved decisions so a safe resume only rebuilds Review.
                resolution["reviewBuildError"] = str(exc)
                resolution["reviewBuildFailedAt"] = (
                    _utc_now().isoformat()
                )
            else:
                resolution["status"] = "failed"
                resolution["failedAt"] = _utc_now().isoformat()
                resolution["error"] = str(exc)
            state["historicalReclassificationResolution"] = resolution
        _persist_job_state(state)
        _append_log(log_path, "\n=== Smart Merge Failed ===")
        _append_log(log_path, str(exc))
        _append_log(log_path, traceback.format_exc())
    finally:
        if working_bundle_dir is not None and not bundle_committed:
            _remove_file_or_tree(working_bundle_dir)
        _RUNNING_THREADS.pop(job_id, None)


def _launch_smart_merge_thread(job_id: str) -> None:
    _launch_job_thread(job_id)


def _create_smart_merge_candidate_with_job_lock(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    with _exclusive_file_lock(_job_state_lock_path(job_id)) as acquired:
        if not acquired:
            raise HTTPException(
                status_code=503,
                detail="Smart Merge 状态锁暂不可用，请稍后重试。",
            )
        return _create_smart_merge_candidate_locked(
            job_id=job_id,
            triggered_by=triggered_by,
        )


def create_smart_merge_candidate(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    with _monthly_update_worker_start_window(
        action="排队 Smart Merge",
        excluding_job_id=job_id,
    ):
        return _create_smart_merge_candidate_with_job_lock(
            job_id=job_id,
            triggered_by=triggered_by,
        )


def _create_smart_merge_candidate_locked(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    """Smart Merge: merge regressed countries from active into candidate, then rebuild artifacts."""
    _require_no_running_monthly_update_jobs(excluding_job_id=job_id)
    payload = _load_job_state(job_id)
    pending = _pending_operation(payload)
    if (
        isinstance(pending, dict)
        and str(pending.get("status") or "") in {"queued", "running"}
    ):
        raise HTTPException(
            status_code=409,
            detail="当前任务已有 Publish/Rollback 操作在队列中，不能同时 Smart Merge。",
        )

    if (
        str(payload.get("status", "")) != "success"
        or str(payload.get("phase", "")) != "completed"
    ):
        raise HTTPException(
            status_code=409,
            detail="只有 success/completed 的月更任务才能执行 Smart Merge。",
        )

    publication = payload.get("publication")
    if isinstance(publication, dict) and publication.get("publishedAt") and not publication.get("rolledBackAt"):
        raise HTTPException(status_code=409, detail="该月更任务已经 publish 过，不能执行 Smart Merge。")

    summaries = payload.get("summaries")
    if isinstance(summaries, dict) and summaries.get("smartMerge") is not None:
        raise HTTPException(status_code=409, detail="该任务已执行过 Smart Merge。")

    active_paths = _active_data_paths()
    artifacts = payload.get("artifacts")
    if not isinstance(artifacts, dict):
        raise HTTPException(status_code=409, detail="当前任务缺少 staging 产物信息。")

    candidate_path = _project_path(str(artifacts.get("stagingOutputPath") or "").strip())
    if candidate_path is None or not candidate_path.exists():
        raise HTTPException(status_code=409, detail="找不到 candidate staging parquet，不能执行 Smart Merge。")
    if not active_paths["parquet"].exists():
        raise HTTPException(status_code=409, detail="找不到 active 数据集，不能执行 Smart Merge。")
    review = get_jato_monthly_update_review(job_id)
    historical_report = review.get("historicalReclassificationReport")
    resolution = _historical_reclassification_resolution(payload)
    if (
        isinstance(historical_report, dict)
        and historical_report.get("status") == "decision_required"
    ):
        raw_historical_countries = historical_report.get("countries")
        normalized_current_report = (
            _normalized_historical_reclassification_report_for_resolution(
                historical_report
            )
        )
        current_raw_fingerprint = (
            _historical_reclassification_report_fingerprint(
                [
                    dict(item)
                    for item in raw_historical_countries
                    if isinstance(item, dict)
                ]
            )
            if isinstance(raw_historical_countries, list)
            else ""
        )
        if (
            not isinstance(resolution, dict)
            or resolution.get("status") != "queued"
            or str(
                resolution.get("sourceCandidateFingerprint") or ""
            )
            != str(review.get("candidateFingerprint") or "")
            or str(resolution.get("reportFingerprint") or "")
            != str(
                normalized_current_report.get("reportFingerprint") or ""
            )
            or str(historical_report.get("reportFingerprint") or "")
            != current_raw_fingerprint
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "blockerType": (
                        "historical_reclassification_decision_required"
                    ),
                    "message": (
                        "请先通过 historical-reclassification-resolution"
                        " 提交所有受影响国家的决策。"
                    ),
                },
            )
        _validated_historical_reclassification_resolution(resolution)
    candidate_active_fingerprint = str(
        payload.get("activeBaseFingerprint") or ""
    ).strip()
    current_active_fingerprint = _active_dataset_version()
    if (
        not re.fullmatch(r"[0-9a-f]{64}", candidate_active_fingerprint)
        or candidate_active_fingerprint != current_active_fingerprint
    ):
        raise HTTPException(
            status_code=409,
            detail={
                "blockerType": "stale_candidate",
                "message": (
                    "部分国家 candidate 的 active lineage 已缺失或变化；"
                    "不能用当前 active 补写旧 Review，请创建新 attempt。"
                ),
                "candidateActiveFingerprint": (
                    candidate_active_fingerprint or None
                ),
                "currentActiveFingerprint": current_active_fingerprint,
            },
        )

    payload["triggeredBy"] = triggered_by.strip() or "anonymous"
    payload["status"] = "queued"
    payload["phase"] = "queued"
    payload["operation"] = "smart_merge"
    payload["startedAt"] = None
    payload["finishedAt"] = None
    payload["error"] = None
    payload["reviewApproval"] = None
    artifacts.pop("reviewBundlePath", None)
    payload["artifacts"] = artifacts
    _job_review_bundle_path(job_id).unlink(missing_ok=True)
    _persist_job_state(payload)

    _launch_smart_merge_thread(job_id)
    return _serialize_job_state(payload, include_log_tail=False)


def _build_runtime_check(payload: dict[str, Any]) -> dict[str, Any]:
    job_id = str(payload.get("jobId") or "")
    pid = _current_process_pid(payload)
    worker_pid = int(payload.get("workerPid") or 0)
    process_alive = _process_exists(pid)
    worker_alive = _process_exists(worker_pid)
    thread_alive = _thread_is_alive(job_id)
    checked_at = _utc_now().isoformat()
    return {
        "checkedAt": checked_at,
        "statusAtCheck": str(payload.get("status") or ""),
        "phaseAtCheck": str(payload.get("phase") or ""),
        "threadAlive": thread_alive,
        "workerPid": worker_pid or None,
        "workerAlive": worker_alive,
        "processPid": pid,
        "processAlive": process_alive,
        "log": _job_log_probe(_job_log_path(job_id)),
        "artifacts": _artifact_probe(payload),
    }


def recheck_jato_monthly_update_job(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    _recover_incomplete_active_transactions_if_possible()
    reconciled = _reconcile_stale_monthly_update_jobs()
    with _exclusive_file_lock(_job_state_lock_path(job_id)) as acquired:
        if not acquired:
            raise HTTPException(
                status_code=503,
                detail="任务复核状态锁暂不可用，请稍后重试。",
            )
        payload = _load_job_state(job_id)
        runtime_check = _build_runtime_check(payload)
        pending = _pending_operation(payload)
        should_wake = (
            str(payload.get("status") or "") == "queued"
            or (
                isinstance(pending, dict)
                and str(pending.get("status") or "") == "queued"
            )
        )
    wake_error = None
    if should_wake:
        with _monthly_update_worker_start_window(
            action="唤醒排队中的月更任务",
            excluding_job_id=job_id,
        ):
            with _exclusive_file_lock(
                _job_state_lock_path(job_id)
            ) as acquired:
                if not acquired:
                    raise HTTPException(
                        status_code=503,
                        detail="任务复核状态锁暂不可用，请稍后重试。",
                    )
                payload = _load_job_state(job_id)
                runtime_check = _build_runtime_check(payload)
                pending = _pending_operation(payload)
                should_wake = (
                    str(payload.get("status") or "") == "queued"
                    or (
                        isinstance(pending, dict)
                        and str(pending.get("status") or "") == "queued"
                    )
                )
                if should_wake:
                    try:
                        _launch_job_thread(job_id)
                    except Exception as exc:
                        wake_error = str(exc)
    # Runtime diagnostics are response-only. Persisting a stale whole payload
    # here used to erase publication/rollback fields written concurrently.
    payload["runtimeCheck"] = {
        **runtime_check,
        "resolvedAs": (
            "worker_rewoken"
            if should_wake and wake_error is None
            else "worker_wake_failed"
            if should_wake
            else "reconciled"
            if job_id in reconciled
            else "observed"
        ),
        "wakeError": wake_error,
        "checkedBy": triggered_by.strip() or "anonymous",
    }
    return _serialize_job_state(payload, include_log_tail=True)


def cancel_jato_monthly_update_job(
    *,
    job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    payload = _load_job_state(job_id)
    status = str(payload.get("status") or "")
    if status not in RUNNING_JOB_STATUSES:
        raise HTTPException(
            status_code=409,
            detail="只有 queued/running 的月更任务可以终止。",
        )

    now = _utc_now().isoformat()
    actor = triggered_by.strip() or "anonymous"
    phase = str(payload.get("phase") or "unknown")
    _write_json(
        _job_cancel_request_path(job_id),
        {
            "cancelledAt": now,
            "cancelledBy": actor,
            "phaseAtCancel": phase,
        },
    )
    pid = _current_process_pid(payload)
    current_process = payload.get("currentProcess")
    termination = _terminate_process_group(
        pid,
        expected_identity=(
            current_process.get("identity")
            if isinstance(current_process, dict)
            and isinstance(current_process.get("identity"), dict)
            else None
        ),
    ) if pid is not None else {
        "pid": None,
        "sigtermSent": False,
        "sigkillSent": False,
        "processAliveBefore": False,
        "processAliveAfter": False,
        "message": "No current subprocess recorded.",
    }
    payload["status"] = "cancelled"
    payload["phase"] = "cancelled"
    payload["finishedAt"] = now
    payload["error"] = f"Cancelled by {actor} during {phase}"
    payload["currentProcess"] = None
    payload["cancellation"] = {
        "cancelledAt": now,
        "cancelledBy": actor,
        "phaseAtCancel": phase,
        "termination": termination,
    }
    payload["runtimeCheck"] = _build_runtime_check(payload)
    _persist_job_state(payload)
    _write_jato_etl_pipeline_status(payload)
    _append_log(
        _job_log_path(job_id),
        (
            f"[{now}] Cancelled by {actor} during {phase}. "
            f"termination={json.dumps(termination, ensure_ascii=False)}"
        ),
    )
    return _serialize_job_state(payload, include_log_tail=True)


def list_jato_monthly_update_jobs(*, limit: int = 20) -> dict[str, Any]:
    _recover_incomplete_active_transactions_if_possible()
    items = [
        _serialize_job_state(payload, include_log_tail=False)
        for payload in _list_job_state_payloads()
    ]

    items.sort(key=lambda item: str(item.get("createdAt", "")), reverse=True)
    sliced = items[:limit]
    return {"rows": len(sliced), "items": sliced}


def get_jato_monthly_update_job(job_id: str) -> dict[str, Any]:
    payload = _load_job_state(job_id)
    return _serialize_job_state(payload, include_log_tail=True)


def get_jato_monthly_update_maintenance_status() -> dict[str, Any]:
    _recover_incomplete_active_transactions_if_possible()
    _reconcile_stale_baseline_promotion()
    baseline_promotion = _load_baseline_promotion_state()
    if (
        isinstance(baseline_promotion, dict)
        and str(baseline_promotion.get("status") or "") == "queued"
    ):
        try:
            _launch_job_thread("_maintenance-baseline-promotion")
        except Exception:
            pass
        baseline_promotion = _load_baseline_promotion_state()
    baseline_path, baseline_source = _resolve_latest_baseline()
    patch_dirs = sorted([path for path in PATCHES_ROOT.glob("*") if path.is_dir()])
    latest_patch_dir = _latest_patch_dir(patch_dirs)
    processed_root = _processed_data_root()
    active_paths = _active_data_paths()
    metrics = [
        _build_storage_metric(
            key="active-baseline",
            label="Active baseline",
            paths=[BASELINE_ROOT],
        ),
        _build_storage_metric(
            key="baseline-archive",
            label="Archived baselines",
            paths=[HISTORY_ARCHIVE_ROOT / "baseline"],
        ),
        _build_storage_metric(
            key="patch-batches",
            label="Patch batches",
            paths=[PATCHES_ROOT],
        ),
        _build_storage_metric(
            key="upload-session-cache",
            label="Upload session cache",
            paths=[_upload_session_root()],
        ),
        _build_storage_metric(
            key="job-upload-copies",
            label="Job upload copies",
            paths=_job_upload_dirs(),
        ),
        _build_storage_metric(
            key="review-reports",
            label="Raw compare reviews",
            paths=[processed_root / "reviews" / "raw_compare"],
        ),
        _build_storage_metric(
            key="staging-outputs",
            label="Staging outputs",
            paths=[processed_root / "staging"],
        ),
        _build_storage_metric(
            key="active-dataset",
            label="Active dataset",
            paths=[
                active_paths["parquet"],
                active_paths["manifest"],
                active_paths["partition"],
                active_paths["fingerprint"],
                active_paths["refreshReport"],
            ],
        ),
        _build_storage_metric(
            key="refresh-backups",
            label="Refresh backups",
            paths=[active_paths["backupRoot"]],
        ),
    ]
    return {
        "checkedAt": _utc_now().isoformat(),
        "activeBaselinePath": _relative_to_project(baseline_path),
        "activeBaselineSource": baseline_source,
        "latestPatchBatch": latest_patch_dir.name if latest_patch_dir is not None else None,
        "jobCount": len(_list_job_state_payloads()),
        "uploadSessionCount": len(_iter_upload_session_payloads()),
        "baselinePromotion": _serialize_baseline_promotion_state(
            baseline_promotion
        ),
        "trackedStorageBytes": sum(int(item["bytes"]) for item in metrics),
        "storageMetrics": metrics,
    }


def _load_baseline_promotion_state() -> dict[str, Any] | None:
    path = _baseline_promotion_state_path()
    if not path.exists():
        return None
    try:
        payload = _read_json(path)
    except Exception:
        return None
    return payload


def _serialize_baseline_promotion_state(
    payload: dict[str, Any] | None,
) -> dict[str, Any] | None:
    if not isinstance(payload, dict):
        return None
    return {
        "operationId": str(payload.get("operationId") or ""),
        "status": str(payload.get("status") or ""),
        "requestedAt": payload.get("requestedAt"),
        "requestedBy": payload.get("requestedBy"),
        "startedAt": payload.get("startedAt"),
        "finishedAt": payload.get("finishedAt"),
        "error": payload.get("error"),
        "failureDigest": (
            payload.get("failureDigest")
            if isinstance(payload.get("failureDigest"), dict)
            else None
        ),
        "sourceActiveFingerprint": payload.get("sourceActiveFingerprint"),
        "promotedAt": payload.get("promotedAt"),
        "triggeredBy": payload.get("triggeredBy"),
        "sourceParquetPath": payload.get("sourceParquetPath"),
        "baselinePath": payload.get("baselinePath"),
        "detectedLatestMonth": payload.get("detectedLatestMonth"),
        "countryCount": int(payload.get("countryCount") or 0),
        "rowCount": int(payload.get("rowCount") or 0),
        "archivedBaselineCount": int(
            payload.get("archivedBaselineCount") or 0
        ),
        "archivedBaselines": (
            [str(value) for value in payload.get("archivedBaselines", [])]
            if isinstance(payload.get("archivedBaselines"), list)
            else []
        ),
    }


def promote_current_active_to_baseline(*, triggered_by: str) -> dict[str, Any]:
    with _exclusive_file_lock(
        _maintenance_coordination_lock_path(),
        blocking=False,
    ) as coordinated:
        if not coordinated:
            raise HTTPException(
                status_code=409,
                detail="JATO 上传或清理操作正在准备中，请稍后再保存 baseline。",
            )
        _require_no_active_upload_sessions(action="保存 active baseline")
        return _promote_current_active_to_baseline_coordinated(
            triggered_by=triggered_by,
        )


def _promote_current_active_to_baseline_coordinated(
    *,
    triggered_by: str,
) -> dict[str, Any]:
    """Queue a heavy parquet→xlsx baseline export in the isolated worker."""
    running_jobs = [
        str(payload.get("jobId", ""))
        for payload in _list_job_state_payloads()
        if (
            str(payload.get("status", "")) in {"queued", "running"}
            or (
                isinstance(payload.get("pendingOperation"), dict)
                and str(payload["pendingOperation"].get("status") or "")
                in {"queued", "running"}
            )
        )
    ]
    if running_jobs:
        raise HTTPException(
            status_code=409,
            detail="存在运行中的月更任务，请等待完成后再保存新的 baseline。",
        )

    active_parquet_path = _active_data_paths()["parquet"]
    if not active_parquet_path.exists():
        raise HTTPException(
            status_code=409,
            detail="当前 active parquet 不存在，不能生成 baseline。",
        )
    with _exclusive_file_lock(
        _active_bundle_lock_path(),
        blocking=False,
    ) as active_acquired:
        if not active_acquired:
            raise HTTPException(
                status_code=409,
                detail="active bundle 正在切换，请完成后再保存 baseline。",
            )
        _recover_incomplete_active_transactions(_active_data_paths())
        source_active_fingerprint = _active_dataset_version()
    with _exclusive_file_lock(_baseline_promotion_lock_path()) as acquired:
        if not acquired:
            raise HTTPException(
                status_code=503,
                detail="baseline 保存状态锁暂不可用，请稍后重试。",
            )
        existing = _load_baseline_promotion_state()
        if isinstance(existing, dict) and str(
            existing.get("status") or ""
        ) in {"queued", "running"}:
            if str(existing.get("status") or "") == "queued":
                _launch_job_thread("_maintenance-baseline-promotion")
            return _serialize_baseline_promotion_state(existing) or {}

        now = _utc_now().isoformat()
        state = {
            "operationId": f"jato-baseline-{uuid4().hex[:10]}",
            "status": "queued",
            "requestedAt": now,
            "requestedBy": triggered_by.strip() or "anonymous",
            "startedAt": None,
            "finishedAt": None,
            "error": None,
            "failureDigest": None,
            "sourceActiveFingerprint": source_active_fingerprint,
        }
        _write_json(_baseline_promotion_state_path(), state)
        try:
            _launch_job_thread("_maintenance-baseline-promotion")
        except Exception as exc:
            state["status"] = "failed"
            state["finishedAt"] = _utc_now().isoformat()
            state["error"] = str(exc)
            state["failureDigest"] = _failure_digest_from_exception(
                phase="baseline_promotion_queued",
                exc=exc,
            )
            _write_json(_baseline_promotion_state_path(), state)
            raise HTTPException(
                status_code=503,
                detail=f"baseline 保存 worker 无法启动：{exc}",
            ) from exc
        return _serialize_baseline_promotion_state(state) or {}


def _execute_promote_current_active_to_baseline(
    *,
    triggered_by: str,
    expected_active_fingerprint: str,
    operation_id: str | None = None,
) -> dict[str, Any]:
    maintenance_dir = _maintenance_dir()
    maintenance_dir.mkdir(parents=True, exist_ok=True)
    snapshot_path = (
        maintenance_dir / f"baseline-source-{uuid4().hex}.parquet"
    )
    active_parquet_path = _active_data_paths()["parquet"]
    with _exclusive_file_lock(_active_bundle_lock_path()) as acquired:
        if not acquired:
            raise HTTPException(
                status_code=409,
                detail="active bundle 正在切换，请稍后重新保存 baseline。",
            )
        _recover_incomplete_active_transactions(_active_data_paths())
        current_active_fingerprint = _active_dataset_version()
        if current_active_fingerprint != expected_active_fingerprint:
            raise HTTPException(
                status_code=409,
                detail=(
                    "baseline 保存排队后 active 已变化；"
                    "为避免导出混合快照，本次操作已停止。"
                ),
            )
        if not active_parquet_path.exists():
            raise HTTPException(
                status_code=409,
                detail="当前 active parquet 不存在，不能生成 baseline。",
            )
        try:
            os.link(active_parquet_path, snapshot_path)
        except OSError:
            shutil.copy2(active_parquet_path, snapshot_path)

    temp_export_path = maintenance_dir / f"baseline-export-{uuid4().hex}.xlsx"
    try:
        import pyarrow.parquet as pq
        from openpyxl import Workbook

        parquet_file = pq.ParquetFile(snapshot_path)
        row_count = int(parquet_file.metadata.num_rows)
        if row_count + 1 > BASELINE_XLSX_MAX_ROWS:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"active 有 {row_count:,} 行，超过 Excel 单工作表"
                    f" {BASELINE_XLSX_MAX_ROWS - 1:,} 条数据上限；"
                    "不能生成会截断历史的 baseline。"
                ),
            )
        columns = [str(column).strip() for column in pq.read_schema(snapshot_path).names]
        country_column = _find_country_column(columns)
        month_columns = _detect_month_columns(columns)
        if country_column is None or not month_columns:
            raise HTTPException(
                status_code=409,
                detail="active 缺少国家列或月份列，不能生成 baseline。",
            )
        country_index = columns.index(country_column)
        month_indices = {
            column: columns.index(column)
            for column in month_columns
        }
        country_values: set[str] = set()
        month_has_data = {column: False for column in month_columns}

        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet(DEFAULT_UPLOAD_SHEET_NAME)
        worksheet.append(columns)
        for batch in parquet_file.iter_batches(
            batch_size=BASELINE_EXPORT_BATCH_ROWS
        ):
            column_values = [
                batch.column(index).to_pylist()
                for index in range(batch.num_columns)
            ]
            for row in zip(*column_values, strict=False):
                country = str(row[country_index] or "").strip()
                if country:
                    country_values.add(country)
                for month, index in month_indices.items():
                    value = row[index]
                    if value is not None and (
                        not isinstance(value, str)
                        or bool(value.strip())
                    ):
                        month_has_data[month] = True
                worksheet.append(row)
        workbook.save(temp_export_path)
        workbook.close()
        if not country_values:
            raise HTTPException(
                status_code=409,
                detail="active 不包含有效国家数据，不能生成 baseline。",
            )
        _logical_country_display_map(
            pd.Series(sorted(country_values), dtype="string"),
            path_label="active baseline snapshot",
        )
        latest_label = next(
            (
                month
                for month in reversed(month_columns)
                if month_has_data[month]
            ),
            None,
        )
        if latest_label is None:
            raise HTTPException(
                status_code=409,
                detail="active 的月份列均为空，不能生成 baseline。",
            )
        parsed_latest = datetime.strptime(latest_label.title(), "%Y %b")
        latest_month = f"{parsed_latest.year}-{parsed_latest.month:02d}"
        country_count = len(
            {country.casefold() for country in country_values}
        )
    except Exception as exc:
        temp_export_path.unlink(missing_ok=True)
        snapshot_path.unlink(missing_ok=True)
        if isinstance(exc, HTTPException):
            raise
        raise HTTPException(
            status_code=500,
            detail="导出 baseline xlsx 失败，请确认后端具备 openpyxl 能力。",
        ) from exc
    finally:
        snapshot_path.unlink(missing_ok=True)

    BASELINE_ROOT.mkdir(parents=True, exist_ok=True)
    existing_baselines = _list_supported_excel_files(BASELINE_ROOT)
    archived_baselines: list[str] = []
    archived_baseline_mappings: list[dict[str, str]] = []
    for path in existing_baselines:
        archive_root = HISTORY_ARCHIVE_ROOT / "baseline"
        archive_root.mkdir(parents=True, exist_ok=True)
        target = _ensure_unique_archive_path(archive_root / path.name)
        shutil.copy2(path, target)
        archived_baselines.append(_relative_to_project(target) or str(target))
        archived_baseline_mappings.append(
            {
                "source": str(path),
                "archive": str(target),
            }
        )

    baseline_path = BASELINE_ROOT / _baseline_snapshot_filename(
        latest_month=latest_month,
        country_count=country_count,
    )
    baseline_path.parent.mkdir(parents=True, exist_ok=True)
    baseline_sha256 = _sha256_hex_for_path(temp_export_path)
    promoted_at = _utc_now().isoformat()
    result = {
        "promotedAt": promoted_at,
        "triggeredBy": triggered_by.strip() or "anonymous",
        "sourceParquetPath": _relative_to_project(active_parquet_path),
        "baselinePath": _relative_to_project(baseline_path),
        "detectedLatestMonth": latest_month,
        "countryCount": country_count,
        "rowCount": row_count,
        "archivedBaselineCount": len(archived_baselines),
        "archivedBaselines": archived_baselines,
    }
    journal = {
        "operationId": operation_id,
        "status": "installing",
        "sourceActiveFingerprint": expected_active_fingerprint,
        "targetPath": str(baseline_path),
        "targetSha256": baseline_sha256,
        "tempExportPath": str(temp_export_path),
        "oldBaselinePaths": [str(path) for path in existing_baselines],
        "archivedBaselineMappings": archived_baseline_mappings,
        "result": result,
        "updatedAt": _utc_now().isoformat(),
    }
    _write_json(_baseline_install_journal_path(), journal)
    os.replace(temp_export_path, baseline_path)
    journal["status"] = "installed"
    journal["updatedAt"] = _utc_now().isoformat()
    _write_json(_baseline_install_journal_path(), journal)
    for old_path in existing_baselines:
        if old_path != baseline_path:
            old_path.unlink(missing_ok=True)
    journal["status"] = "committed"
    journal["committedAt"] = _utc_now().isoformat()
    journal["updatedAt"] = journal["committedAt"]
    _write_json(_baseline_install_journal_path(), journal)

    return result


def run_jato_monthly_update_cleanup(*, triggered_by: str, cleanup_tier: str = SAFE_CLEANUP_TIER) -> dict[str, Any]:
    with _exclusive_file_lock(
        _maintenance_coordination_lock_path(),
        blocking=False,
    ) as coordinated:
        if not coordinated:
            raise HTTPException(
                status_code=409,
                detail="JATO 上传或 baseline 操作正在准备中，请稍后再清理。",
            )
        _require_no_active_upload_sessions(action="执行一键清理")
        return _run_jato_monthly_update_cleanup_coordinated(
            triggered_by=triggered_by,
            cleanup_tier=cleanup_tier,
        )


def _run_jato_monthly_update_cleanup_coordinated(
    *,
    triggered_by: str,
    cleanup_tier: str = SAFE_CLEANUP_TIER,
) -> dict[str, Any]:
    normalized_cleanup_tier = _normalize_cleanup_tier(cleanup_tier)
    job_payloads = _list_job_state_payloads()
    running_jobs = [
        str(payload.get("jobId", ""))
        for payload in job_payloads
        if (
            str(payload.get("status", "")) in {"queued", "running"}
            or (
                isinstance(payload.get("pendingOperation"), dict)
                and str(payload["pendingOperation"].get("status") or "")
                in {"queued", "running"}
            )
        )
    ]
    if running_jobs:
        raise HTTPException(
            status_code=409,
            detail="存在运行中的月更任务，暂时不能执行一键清理。",
        )
    baseline_promotion = _load_baseline_promotion_state()
    if (
        isinstance(baseline_promotion, dict)
        and str(baseline_promotion.get("status") or "")
        in {"queued", "running"}
    ):
        raise HTTPException(
            status_code=409,
            detail="正在保存 active baseline，请等待完成后再执行一键清理。",
        )

    baseline_files = _list_supported_excel_files(BASELINE_ROOT)
    latest_baseline = _latest_baseline_file(baseline_files)
    archived_baselines: list[str] = []
    if latest_baseline is not None:
        for path in baseline_files:
            if path == latest_baseline:
                continue
            target = _move_to_archive(path, HISTORY_ARCHIVE_ROOT / "baseline")
            archived_baselines.append(_relative_to_project(target) or str(target))

    patch_dirs = sorted([path for path in PATCHES_ROOT.glob("*") if path.is_dir()])
    latest_patch_dir = _latest_patch_dir(patch_dirs)
    archived_patch_dirs: list[str] = []
    if latest_patch_dir is not None:
        for path in patch_dirs:
            if path == latest_patch_dir:
                continue
            target = _move_to_archive(path, HISTORY_ARCHIVE_ROOT / "patches")
            archived_patch_dirs.append(_relative_to_project(target) or str(target))

    removed_job_upload_dirs: list[str] = []
    removed_job_upload_bytes = 0
    for payload in job_payloads:
        if str(payload.get("status", "")) not in {"success", "failed"}:
            continue
        job_id = str(payload.get("jobId", "")).strip()
        if not job_id:
            continue
        upload_dir = _job_dir(job_id) / "uploads"
        if not upload_dir.exists():
            continue
        path_bytes, _, _ = _measure_path_usage(upload_dir)
        removed_job_upload_bytes += path_bytes
        shutil.rmtree(upload_dir)
        removed_job_upload_dirs.append(_relative_to_project(upload_dir) or str(upload_dir))
        upload_payload = payload.get("upload")
        if isinstance(upload_payload, dict):
            upload_payload["storedPath"] = None
        _persist_job_state(payload)

    removed_upload_session_dirs, removed_upload_session_bytes = _remove_cleanup_paths(
        _terminal_upload_session_dirs()
    )
    deleted_review_dirs: list[str] = []
    deleted_review_bytes = 0
    deleted_staging_dirs: list[str] = []
    deleted_staging_bytes = 0
    deleted_refresh_backup_dirs: list[str] = []
    deleted_refresh_backup_bytes = 0
    deleted_archived_baselines: list[str] = []
    deleted_archived_baseline_bytes = 0
    deleted_archived_patch_dirs: list[str] = []
    deleted_archived_patch_bytes = 0
    if normalized_cleanup_tier == CAUTIOUS_CLEANUP_TIER:
        processed_root = _processed_data_root()
        active_paths = _active_data_paths()
        deleted_review_dirs, deleted_review_bytes = _remove_cleanup_paths(
            _child_cleanup_paths(processed_root / "reviews" / "raw_compare")
        )
        deleted_staging_dirs, deleted_staging_bytes = _remove_cleanup_paths(
            _child_cleanup_paths(processed_root / "staging")
        )
        deleted_refresh_backup_dirs, deleted_refresh_backup_bytes = _remove_cleanup_paths(
            _child_cleanup_paths(active_paths["backupRoot"])
        )
        deleted_archived_baselines, deleted_archived_baseline_bytes = _remove_cleanup_paths(
            _list_supported_excel_files(HISTORY_ARCHIVE_ROOT / "baseline")
        )
        deleted_archived_patch_dirs, deleted_archived_patch_bytes = _remove_cleanup_paths(
            [path for path in _child_cleanup_paths(HISTORY_ARCHIVE_ROOT / "patches") if path.is_dir()]
        )

    cleaned_at = _utc_now().isoformat()
    freed_bytes = (
        removed_job_upload_bytes
        + removed_upload_session_bytes
        + deleted_review_bytes
        + deleted_staging_bytes
        + deleted_refresh_backup_bytes
        + deleted_archived_baseline_bytes
        + deleted_archived_patch_bytes
    )
    return {
        "cleanedAt": cleaned_at,
        "triggeredBy": triggered_by.strip() or "anonymous",
        "cleanupTier": normalized_cleanup_tier,
        "activeBaselinePath": _relative_to_project(latest_baseline),
        "activePatchMonth": latest_patch_dir.name if latest_patch_dir is not None else None,
        "freedBytes": int(freed_bytes),
        "archivedBaselineCount": len(archived_baselines),
        "archivedBaselines": archived_baselines,
        "archivedPatchDirCount": len(archived_patch_dirs),
        "archivedPatchDirs": archived_patch_dirs,
        "removedUploadSessionDirCount": len(removed_upload_session_dirs),
        "removedUploadSessionDirs": removed_upload_session_dirs,
        "removedJobUploadDirCount": len(removed_job_upload_dirs),
        "removedJobUploadDirs": removed_job_upload_dirs,
        "deletedReviewDirCount": len(deleted_review_dirs),
        "deletedReviewDirs": deleted_review_dirs,
        "deletedStagingDirCount": len(deleted_staging_dirs),
        "deletedStagingDirs": deleted_staging_dirs,
        "deletedRefreshBackupDirCount": len(deleted_refresh_backup_dirs),
        "deletedRefreshBackupDirs": deleted_refresh_backup_dirs,
        "deletedArchivedBaselineCount": len(deleted_archived_baselines),
        "deletedArchivedBaselines": deleted_archived_baselines,
        "deletedArchivedPatchDirCount": len(deleted_archived_patch_dirs),
        "deletedArchivedPatchDirs": deleted_archived_patch_dirs,
    }
