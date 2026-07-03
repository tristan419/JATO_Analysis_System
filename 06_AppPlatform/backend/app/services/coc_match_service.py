"""COC match service — Excel registry vs ZIP/RAR file comparison.

Adapted from COCtrack (match_coc.py). Uses upload_toolkit for background job
processing and file utilities.
"""

import json
import re
import sqlite3
import shutil
import subprocess
import threading
import uuid
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile
from upload_toolkit.file_utils import allowed_extension
from upload_toolkit.job_engine import BaseJobRunner, append_log, list_job_payloads, load_job_state, persist_job_state, state_path, log_path

from app.core.config import COC_MATCH_JOB_ROOT

# ── Constants ───────────────────────────────────────────────────────

ALLOWED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
ALLOWED_ARCHIVE_EXTENSIONS = {".zip", ".rar"}
_RUNNING_THREADS: dict[str, "CocMatchJobRunner"] = {}
_COC_UPLOAD_CHUNK_SIZE = 8 * 1024 * 1024  # 8 MB
_COC_UPLOAD_THRESHOLD_BYTES = 50 * 1024 * 1024  # 50 MB

# ── Excel → rows ───────────────────────────────────────────────────

_VIN_HEADER_CANDIDATES = {
    "vin",
    "vinno",
    "vinnumber",
    "vin码",
    "vin号",
    "vehicleidentificationnumber",
    "chassis",
    "chassisno",
    "chassisnumber",
    "车架号",
    "底盘号",
}
_MODEL_HEADER_CANDIDATES = {"model", "车型", "车型名", "modelname", "vehiclemodel"}
_COUNTRY_HEADER_CANDIDATES = {"country", "国家", "market", "市场"}
_HEADER_SCAN_ROWS = 20
_HEADERLESS_VIN_SCAN_ROWS = 200
_VIN_VALUE_RE = re.compile(r"^[A-HJ-NPR-Z0-9]{17}$", re.IGNORECASE)


def _normalize_excel_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", text)


def _find_header_index(row: tuple[object, ...], candidates: set[str]) -> int | None:
    for index, value in enumerate(row):
        normalized = _normalize_excel_header(value)
        if not normalized:
            continue
        if normalized in candidates:
            return index
        if candidates is _VIN_HEADER_CANDIDATES:
            if normalized.startswith("vin") or normalized.endswith("vin") or "车架" in normalized or "底盘" in normalized:
                return index
    return None


def _normalize_vin_value(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def _looks_like_vin(value: object) -> bool:
    return bool(_VIN_VALUE_RE.fullmatch(_normalize_vin_value(value)))


def _infer_headerless_vin_column(ws: Any) -> tuple[int, int] | None:
    """Infer a VIN column from headerless sheets containing raw VIN rows."""
    stats: dict[int, dict[str, int | None]] = {}
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=_HEADERLESS_VIN_SCAN_ROWS, values_only=True),
        start=1,
    ):
        for column_index, value in enumerate(row):
            text = str(value or "").strip()
            if not text:
                continue
            column_stats = stats.setdefault(
                column_index,
                {"vin_count": 0, "non_empty_count": 0, "first_vin_row": None},
            )
            column_stats["non_empty_count"] = int(column_stats["non_empty_count"] or 0) + 1
            if _looks_like_vin(value):
                column_stats["vin_count"] = int(column_stats["vin_count"] or 0) + 1
                if column_stats["first_vin_row"] is None:
                    column_stats["first_vin_row"] = row_number

    candidates: list[tuple[int, int, int]] = []
    for column_index, column_stats in stats.items():
        vin_count = int(column_stats["vin_count"] or 0)
        non_empty_count = int(column_stats["non_empty_count"] or 0)
        first_vin_row = column_stats["first_vin_row"]
        if not isinstance(first_vin_row, int) or vin_count <= 0:
            continue
        if vin_count >= 2 or vin_count == non_empty_count:
            candidates.append((vin_count, column_index, first_vin_row))

    if not candidates:
        return None
    _vin_count, column_index, first_vin_row = max(
        candidates,
        key=lambda item: (item[0], -item[1], -item[2]),
    )
    return first_vin_row, column_index


def read_excel_rows(excel_path: Path) -> list[dict[str, str]]:
    """Read COC registry Excel. Returns list of {chassis, model, country}."""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    header_row_number: int | None = None
    vin_index: int | None = None
    model_index: int | None = None
    country_index: int | None = None

    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=_HEADER_SCAN_ROWS, values_only=True), start=1):
        hit = _find_header_index(row, _VIN_HEADER_CANDIDATES)
        if hit is not None:
            header_row_number = idx
            vin_index = hit
            model_index = _find_header_index(row, _MODEL_HEADER_CANDIDATES)
            country_index = _find_header_index(row, _COUNTRY_HEADER_CANDIDATES)
            break

    inferred_headerless_vin = False
    if header_row_number is None or vin_index is None:
        inferred = _infer_headerless_vin_column(ws)
        if inferred is None:
            raise HTTPException(
                status_code=400,
                detail="Excel 未找到 VIN / Chassis / 车架号 表头，也未识别到无表头 VIN 列。",
            )
        header_row_number = inferred[0] - 1
        vin_index = inferred[1]
        inferred_headerless_vin = True

    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=header_row_number + 1, values_only=True):
        chassis = row[vin_index] if vin_index < len(row) else None
        chassis_text = _normalize_vin_value(chassis) if inferred_headerless_vin else str(chassis or "").strip()
        if not chassis_text:
            continue
        if inferred_headerless_vin and not _looks_like_vin(chassis_text):
            continue
        model = (
            str(row[model_index]).strip()
            if model_index is not None and model_index < len(row) and row[model_index]
            else ""
        )
        country = (
            str(row[country_index]).strip()
            if country_index is not None and country_index < len(row) and row[country_index]
            else ""
        )
        rows.append({
            "chassis": chassis_text,
            "model": model,
            "country": country,
        })
    return rows


# ── Archive → file set ─────────────────────────────────────────────

def list_archive_files(
    archive_path: Path,
    extensions: list[str] | None = None,
) -> set[str]:
    """List files in RAR or ZIP archive. Returns set of basenames without extension.

    Auto-detects archive type by file extension.
    """
    if extensions is None:
        extensions = [".pdf"]
    ext_lower = archive_path.suffix.lower()
    if ext_lower == ".rar":
        return _list_rar_files(archive_path, extensions)
    elif ext_lower == ".zip":
        return _list_zip_files(archive_path, extensions)
    else:
        raise ValueError(f"不支持的压缩包格式: {ext_lower}")


def _list_rar_files(rar_path: Path, extensions: list[str]) -> set[str]:
    for command in (
        ["lsar", str(rar_path)],
        ["bsdtar", "-tf", str(rar_path)],
        ["unar", "-l", str(rar_path)],
        ["unrar", "lb", str(rar_path)],
        ["7z", "l", "-ba", str(rar_path)],
        ["7zz", "l", "-ba", str(rar_path)],
    ):
        listed = _list_archive_files_with_command(command, extensions)
        if listed:
            return listed

    listed = _list_rar_files_python(rar_path, extensions)
    if listed:
        return listed

    raise ValueError(
        "无法读取 RAR 文件列表：服务器缺少可用 RAR 工具，且内置解析未能识别文件名。"
    )


def _list_archive_files_with_command(
    command: list[str],
    extensions: list[str],
) -> set[str]:
    if shutil.which(command[0]) is None:
        return set()

    try:
        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            check=False,
        )
    except OSError:
        return set()
    if result.returncode != 0:
        return set()

    names: set[str] = set()
    for line in result.stdout.splitlines():
        stem = _archive_member_stem(line.strip(), extensions)
        if stem:
            names.add(stem)
    return names


def _list_zip_files(zip_path: Path, extensions: list[str]) -> set[str]:
    names: set[str] = set()
    with zipfile.ZipFile(zip_path) as zf:
        for member in zf.namelist():
            stem = _archive_member_stem(member, extensions)
            if stem:
                names.add(stem)
    return names


def _archive_member_stem(member_name: str, extensions: list[str]) -> str | None:
    """Return archive member basename without suffix when it matches extensions."""
    normalized = member_name.strip().replace("\\", "/")
    if not normalized or normalized.endswith("/"):
        return None
    lower = normalized.lower()
    matched_ext = next((ext for ext in extensions if lower.endswith(ext.lower())), None)
    if matched_ext is None:
        return None
    basename = normalized.rsplit("/", 1)[-1]
    return basename[: -len(matched_ext)]


def _read_rar_vint(data: bytes, offset: int) -> tuple[int, int]:
    value = 0
    shift = 0
    start = offset
    while offset < len(data):
        byte = data[offset]
        offset += 1
        value |= (byte & 0x7F) << shift
        if not byte & 0x80:
            return value, offset
        shift += 7
        if offset - start > 10:
            break
    raise ValueError("Invalid RAR variable integer")


def _decode_rar_name(raw: bytes) -> str:
    for encoding in ("utf-8", "cp437", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _list_rar_files_python(rar_path: Path, extensions: list[str]) -> set[str]:
    data = rar_path.read_bytes()
    if data.startswith(b"Rar!\x1a\x07\x01\x00"):
        return _list_rar5_files_python(data, extensions)
    if data.startswith(b"Rar!\x1a\x07\x00"):
        return _list_rar4_files_python(data, extensions)
    return set()


def _list_rar5_files_python(data: bytes, extensions: list[str]) -> set[str]:
    names: set[str] = set()
    offset = 8
    while offset + 5 < len(data):
        header_start = offset
        offset += 4  # CRC32
        try:
            header_size, offset = _read_rar_vint(data, offset)
            size_field_end = offset
            header_end = header_start + 4 + (size_field_end - header_start - 4) + header_size
            header_type, offset = _read_rar_vint(data, offset)
            header_flags, offset = _read_rar_vint(data, offset)
            if header_flags & 0x0001:
                _extra_size, offset = _read_rar_vint(data, offset)
            data_size = 0
            if header_flags & 0x0002:
                data_size, offset = _read_rar_vint(data, offset)
        except (IndexError, ValueError):
            break

        if header_type == 2:
            try:
                file_flags, cursor = _read_rar_vint(data, offset)
                _unpacked_size, cursor = _read_rar_vint(data, cursor)
                _attributes, cursor = _read_rar_vint(data, cursor)
                if file_flags & 0x0002:
                    cursor += 4
                if file_flags & 0x0004:
                    cursor += 4
                _compression_info, cursor = _read_rar_vint(data, cursor)
                _host_os, cursor = _read_rar_vint(data, cursor)
                name_size, cursor = _read_rar_vint(data, cursor)
                raw_name = data[cursor : cursor + name_size]
            except (IndexError, ValueError):
                raw_name = b""
            stem = _archive_member_stem(_decode_rar_name(raw_name), extensions)
            if stem:
                names.add(stem)

        next_offset = header_end + data_size
        if next_offset <= header_start:
            break
        offset = next_offset
    return names


def _list_rar4_files_python(data: bytes, extensions: list[str]) -> set[str]:
    names: set[str] = set()
    offset = 7
    while offset + 7 <= len(data):
        header_start = offset
        try:
            header_type = data[offset + 2]
            header_flags = int.from_bytes(data[offset + 3 : offset + 5], "little")
            header_size = int.from_bytes(data[offset + 5 : offset + 7], "little")
        except IndexError:
            break

        cursor = offset + 7
        add_size = 0
        if header_flags & 0x8000 and cursor + 4 <= len(data):
            add_size = int.from_bytes(data[cursor : cursor + 4], "little")

        if header_type == 0x74 and cursor + 25 <= len(data):
            pack_size = int.from_bytes(data[cursor : cursor + 4], "little")
            name_size = int.from_bytes(data[cursor + 19 : cursor + 21], "little")
            name_cursor = cursor + 25
            if header_flags & 0x0100:
                name_cursor += 8
            raw_name = data[name_cursor : name_cursor + name_size]
            stem = _archive_member_stem(_decode_rar_name(raw_name), extensions)
            if stem:
                names.add(stem)
            add_size = pack_size

        if header_type == 0x7B:
            break
        next_offset = header_start + header_size + add_size
        if next_offset <= header_start:
            break
        offset = next_offset
    return names


# ── Matching ───────────────────────────────────────────────────────

def match_cocs(
    rows: list[dict[str, str]],
    file_set: set[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    matched: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []
    for r in rows:
        r["has_pdf"] = r["chassis"] in file_set
        if r["has_pdf"]:
            matched.append(r)
        else:
            missing.append(r)
    return matched, missing


def find_archive_only_files(
    rows: list[dict[str, str]],
    file_set: set[str],
) -> list[dict[str, str]]:
    """Return archive filenames that do not have a row in the Excel registry."""
    excel_chassis = {row["chassis"] for row in rows}
    return [{"filename": filename} for filename in sorted(file_set - excel_chassis)]


def classify_coc_difference(
    missing_count: int,
    archive_only_count: int,
) -> str:
    if missing_count > 0 and archive_only_count > 0:
        return "bidirectional_mismatch"
    if missing_count > 0:
        return "missing_archive_files"
    if archive_only_count > 0:
        return "archive_only_files"
    return "matched"


# ── SQLite DB ──────────────────────────────────────────────────────

COC_DB_PATH = COC_MATCH_JOB_ROOT / "coc_match_history.db"
_COC_DB_LOCK = threading.RLock()
_COC_DB_BUSY_TIMEOUT_MS = 30_000


def _init_coc_db() -> sqlite3.Connection:
    COC_MATCH_JOB_ROOT.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(COC_DB_PATH), timeout=_COC_DB_BUSY_TIMEOUT_MS / 1000)
    conn.execute(f"PRAGMA busy_timeout={_COC_DB_BUSY_TIMEOUT_MS}")
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            country TEXT NOT NULL,
            month TEXT NOT NULL,
            total INTEGER,
            matched INTEGER,
            missing INTEGER,
            created_at TEXT,
            UNIQUE(country, month)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS coc_status (
            run_id INTEGER REFERENCES runs(id),
            chassis TEXT NOT NULL,
            model TEXT,
            has_pdf INTEGER,
            UNIQUE(run_id, chassis)
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_coc_chassis ON coc_status(chassis)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_coc_run ON coc_status(run_id)")
    conn.commit()
    return conn


def _save_run(conn: sqlite3.Connection, country: str, month: str, rows: list[dict[str, Any]]) -> int:
    total = len(rows)
    matched = sum(1 for r in rows if r["has_pdf"])
    missing = total - matched

    try:
        conn.execute("BEGIN IMMEDIATE")
        old = conn.execute(
            "SELECT id FROM runs WHERE country=? AND month=?", (country, month)
        ).fetchone()
        if old:
            old_id = old[0]
            conn.execute("DELETE FROM coc_status WHERE run_id=?", (old_id,))
            conn.execute("DELETE FROM runs WHERE id=?", (old_id,))

        conn.execute(
            "INSERT INTO runs (country, month, total, matched, missing, created_at) VALUES (?,?,?,?,?,?)",
            (country, month, total, matched, missing, datetime.now().isoformat()),
        )
        run_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        conn.executemany(
            "INSERT INTO coc_status (run_id, chassis, model, has_pdf) VALUES (?,?,?,?)",
            [(run_id, r["chassis"], r["model"], 1 if r["has_pdf"] else 0) for r in rows],
        )
        conn.commit()
        return run_id
    except Exception:
        conn.rollback()
        raise


def _get_previous_run(conn: sqlite3.Connection, country: str, current_month: str):
    return conn.execute(
        "SELECT id, month, total, matched, missing FROM runs "
        "WHERE country=? AND month < ? ORDER BY month DESC LIMIT 1",
        (country, current_month),
    ).fetchone()


def _get_diff(
    conn: sqlite3.Connection,
    prev_run_id: int | None,
    current_rows: list[dict[str, Any]],
) -> dict[str, str]:
    if not prev_run_id:
        return {}

    prev_chassis = {
        row[0]: bool(row[1])
        for row in conn.execute(
            "SELECT chassis, has_pdf FROM coc_status WHERE run_id=?", (prev_run_id,)
        )
    }

    diff: dict[str, str] = {}
    for r in current_rows:
        c = r["chassis"]
        cur = r["has_pdf"]
        prev = prev_chassis.get(c)
        if prev is None:
            diff[c] = "new_entry"
        elif prev and not cur:
            diff[c] = "lost_pdf"
        elif not prev and cur:
            diff[c] = "gained_pdf"
        elif prev and cur:
            diff[c] = "kept_pdf"
        else:
            diff[c] = "still_missing"
    return diff


# ── HTML Report ────────────────────────────────────────────────────

def build_html_report(
    rows: list[dict[str, Any]],
    matched: list[dict[str, Any]],
    missing: list[dict[str, Any]],
    archive_only_files: list[dict[str, str]],
    country: str,
    month: str,
    prev_run: Any,
    diff: dict[str, str],
) -> str:
    total = len(rows)
    has = len(matched)
    no = len(missing)
    archive_only_count = len(archive_only_files)
    difference_type = classify_coc_difference(no, archive_only_count)
    rate = has / total * 100 if total else 0
    difference_label = {
        "matched": "完全一致",
        "missing_archive_files": "Excel 有、压缩包缺失",
        "archive_only_files": "压缩包有、Excel 缺码",
        "bidirectional_mismatch": "双向不一致",
    }[difference_type]

    # ── Missing by Model ──
    missing_by_model: defaultdict[str, list] = defaultdict(list)
    for r in missing:
        missing_by_model[r["model"]].append(r)

    missing_rows_html = ""
    for model in sorted(missing_by_model):
        items = missing_by_model[model]
        missing_rows_html += f"""
        <tr class="model-header">
            <td colspan="4"><strong>{model}</strong> — {len(items)} missing</td>
        </tr>"""
        for r in sorted(items, key=lambda x: x["chassis"]):
            d = diff.get(r["chassis"], "")
            tag = ""
            css = "missing"
            if d == "lost_pdf":
                tag = ' <span class="diff-tag lost">上期有</span>'
                css = "missing lost"
            elif d == "new_entry":
                tag = ' <span class="diff-tag new">新注册</span>'
                css = "missing new-entry"
            missing_rows_html += f"""
        <tr class="{css}">
            <td>{r['chassis']}{tag}</td>
            <td>{r['model']}</td>
            <td>{r['country']}</td>
            <td>NO</td>
        </tr>"""

    # ── Matched (top 50) ──
    matched_rows_html = ""
    for r in sorted(matched, key=lambda x: x["chassis"])[:50]:
        matched_rows_html += f"""
        <tr class="matched">
            <td>{r['chassis']}</td>
            <td>{r['model']}</td>
            <td>{r['country']}</td>
        </tr>"""

    # ── Lost chassis ──
    lost_chassis = [c for c, v in diff.items() if v == "lost_pdf"]
    lost_rows_html = ""
    if lost_chassis:
        lost_info = {r["chassis"]: r for r in rows if r["chassis"] in set(lost_chassis)}
        for c in sorted(lost_chassis):
            info = lost_info.get(c, {"model": "?", "country": country})
            lost_rows_html += f"""
        <tr class="missing lost">
            <td>{c}</td>
            <td>{info['model']}</td>
            <td>{info['country']}</td>
        </tr>"""

    archive_only_rows_html = ""
    for item in archive_only_files:
        archive_only_rows_html += f"""
        <tr class="archive-only">
            <td>{item['filename']}</td>
            <td>压缩包内存在，但 Excel 注册表无对应底盘号</td>
        </tr>"""

    # ── All data JSON (for JS) ──
    all_data_json = json.dumps(rows, ensure_ascii=False)
    archive_only_json = json.dumps(archive_only_files, ensure_ascii=False)

    # ── Diff section ──
    diff_section = ""
    if prev_run:
        gained = sum(1 for v in diff.values() if v == "gained_pdf")
        lost = sum(1 for v in diff.values() if v == "lost_pdf")
        new_entries = sum(1 for v in diff.values() if v == "new_entry")
        prev_month = prev_run[1]
        prev_total = prev_run[2]
        prev_matched = prev_run[3]
        prev_rate = prev_matched / prev_total * 100 if prev_total else 0
        diff_section = f"""
        <div class="diff-section">
            <h2>环比变化（vs {prev_month}）</h2>
            <div class="cards" style="grid-template-columns: repeat(5, 1fr);">
                <div class="card">
                    <div class="number" style="color:#16a34a">+{gained}</div>
                    <div class="label">新补上 PDF</div>
                </div>
                <div class="card">
                    <div class="number" style="color:#dc2626">-{lost}</div>
                    <div class="label">新缺失 PDF</div>
                </div>
                <div class="card">
                    <div class="number" style="color:#6b7280">+{new_entries}</div>
                    <div class="label">新注册 COC</div>
                </div>
                <div class="card">
                    <div class="number" style="color:#2563eb">{prev_rate:.1f}% → {rate:.1f}%</div>
                    <div class="label">覆盖率变化</div>
                </div>
                <div class="card">
                    <div class="number">{prev_total} → {total}</div>
                    <div class="label">注册量变化</div>
                </div>
            </div>
        </div>"""

        if lost_chassis:
            diff_section += f"""
        <div class="section" id="section-lost">
            <div class="section-header">本期新缺失 PDF（上月有，本月无）—— {len(lost_chassis)} 个</div>
            <table>
                <thead><tr><th>Chassis Number</th><th>Model</th><th>Country</th></tr></thead>
                <tbody>{lost_rows_html}</tbody>
            </table>
        </div>"""

    # ── JS code (built separately to avoid f-string escape hell) ──
    js_code = """<script>
const allData = """ + all_data_json + """;
const archiveOnlyData = """ + archive_only_json + """;

function renderAll(data) {
    const tbody = document.getElementById("all-tbody");
    tbody.innerHTML = data.map(r => `
        <tr class="${r.has_pdf ? 'matched' : 'missing'}">
            <td>${r.chassis}</td>
            <td>${r.model}</td>
            <td>${r.country}</td>
            <td>${r.has_pdf ? 'YES' : 'NO'}</td>
        </tr>`).join("");
}
renderAll(allData);

document.getElementById("search").addEventListener("input", function() {
    const q = this.value.toLowerCase();
    document.querySelectorAll("table tbody tr").forEach(tr => {
        const text = tr.textContent.toLowerCase();
        tr.classList.toggle("hidden", !text.includes(q));
    });
    document.querySelectorAll(".model-header").forEach(h => {
        let next = h.nextElementSibling;
        let hasVisible = false;
        while (next && !next.classList.contains("model-header")) {
            if (!next.classList.contains("hidden")) { hasVisible = true; break; }
            next = next.nextElementSibling;
        }
        h.classList.toggle("hidden", !hasVisible);
    });
});

function downloadCSV(filename, rows) {
    const headers = ["Chassis number", "Model", "Country"];
    const lines = [headers.join(",")];
    rows.forEach(r => {
        lines.push([r.chassis, r.model, r.country].map(v => '"' + v.replace(/"/g, '""') + '"').join(","));
    });
    const BOM = "\\uFEFF";
    const NL = "\\n";
    const blob = new Blob([BOM + lines.join(NL)], { type: "text/csv;charset=utf-8" });
    const url = URL.createObjectURL(blob);
    const a = document.createElement("a");
    a.href = url;
    a.download = filename;
    a.click();
    URL.revokeObjectURL(url);
}

function exportMissing() {
    const missing = allData.filter(r => !r.has_pdf);
    downloadCSV("missing_coc.csv", missing);
}

function exportArchiveOnly() {
    const rows = archiveOnlyData.map(r => ({
        chassis: r.filename,
        model: "archive-only",
        country: ""
    }));
    downloadCSV("archive_only_coc.csv", rows);
}

function exportAll() {
    downloadCSV("all_coc.csv", allData);
}

function showTab(tab) {
    ["missing","matched","all","lost","extra"].forEach(t => {
        const s = document.getElementById("section-"+t);
        const b = document.getElementById("tab-"+t);
        if (s) s.classList.add("hidden");
        if (b) b.classList.remove("active");
    });
    const s = document.getElementById("section-"+tab);
    const b = document.getElementById("tab-"+tab);
    if (s) s.classList.remove("hidden");
    if (b) b.classList.add("active");
    document.getElementById("search").value = "";
    document.querySelectorAll("table tbody tr").forEach(tr => tr.classList.remove("hidden"));
    document.querySelectorAll(".model-header").forEach(h => h.classList.remove("hidden"));
}
</script>"""

    title = f"COC 匹配报告 — {country} {month}"

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; background: #f5f5f5; color: #333; }}
.container {{ max-width: 1200px; margin: 0 auto; padding: 24px; }}
h1 {{ font-size: 24px; margin-bottom: 8px; color: #1a1a1a; }}
h2 {{ font-size: 18px; margin-bottom: 16px; color: #1a1a1a; }}
.subtitle {{ color: #666; font-size: 14px; margin-bottom: 24px; }}
.cards {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 16px; margin-bottom: 32px; }}
.card {{ background: white; border-radius: 12px; padding: 24px; text-align: center; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
.card .number {{ font-size: 36px; font-weight: 700; }}
.card .label {{ font-size: 14px; color: #666; margin-top: 4px; }}
.card.total .number {{ color: #1a1a1a; }}
.card.has .number {{ color: #16a34a; }}
.card.missing .number {{ color: #dc2626; }}
.card.extra .number {{ color: #c2410c; }}
.card.rate .number {{ color: #2563eb; }}
.progress-bar {{ background: white; border-radius: 12px; padding: 20px 24px; margin-bottom: 24px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
.progress-label {{ display: flex; justify-content: space-between; margin-bottom: 8px; font-size: 14px; }}
.progress-track {{ height: 12px; background: #e5e5e5; border-radius: 6px; overflow: hidden; }}
.progress-fill {{ height: 100%; background: #16a34a; border-radius: 6px; transition: width 0.5s; }}
.progress-fill.warning {{ background: #f59e0b; }}
.diff-section {{ margin-bottom: 16px; }}
.search-bar {{ margin-bottom: 24px; }}
.search-bar input {{ width: 100%; padding: 12px 16px; border: 1px solid #ddd; border-radius: 8px; font-size: 15px; }}
.search-bar input:focus {{ outline: none; border-color: #2563eb; box-shadow: 0 0 0 3px rgba(37,99,235,0.1); }}
.toolbar {{ display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 12px; margin-bottom: 16px; }}
.tabs {{ display: flex; gap: 8px; }}
.tab {{ padding: 8px 20px; border-radius: 8px; border: 1px solid #ddd; background: white; cursor: pointer; font-size: 14px; }}
.tab.active {{ background: #2563eb; color: white; border-color: #2563eb; }}
.btn-export {{ padding: 10px 24px; border: none; border-radius: 8px; font-size: 14px; font-weight: 600; cursor: pointer; transition: all 0.15s; }}
.btn-export.danger {{ background: #dc2626; color: white; }}
.btn-export.danger:hover {{ background: #b91c1c; }}
.btn-export.default {{ background: #2563eb; color: white; }}
.btn-export.default:hover {{ background: #1d4ed8; }}
.section {{ background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.08); margin-bottom: 32px; }}
.section-header {{ padding: 16px 24px; border-bottom: 1px solid #eee; font-weight: 600; }}
table {{ width: 100%; border-collapse: collapse; }}
th, td {{ text-align: left; padding: 10px 24px; font-size: 14px; border-bottom: 1px solid #f0f0f0; }}
th {{ background: #fafafa; font-weight: 600; color: #666; font-size: 13px; text-transform: uppercase; }}
tr.missing {{ background: #fef2f2; }}
tr.missing td:first-child {{ color: #dc2626; font-weight: 500; }}
tr.missing.lost {{ background: #fde8e8; }}
tr.missing.new-entry {{ background: #fff7ed; }}
tr.missing.new-entry td:first-child {{ color: #c2410c; }}
tr.archive-only {{ background: #fff7ed; }}
tr.archive-only td:first-child {{ color: #c2410c; font-weight: 500; }}
tr.model-header td {{ background: #fafafa; padding: 8px 24px; font-size: 13px; color: #666; }}
.hidden {{ display: none; }}
.diff-tag {{ font-size: 11px; padding: 1px 6px; border-radius: 4px; margin-left: 6px; font-weight: 500; }}
.diff-tag.lost {{ background: #fee2e2; color: #dc2626; }}
.diff-tag.new {{ background: #ffedd5; color: #c2410c; }}
.footer {{ text-align: center; color: #999; font-size: 13px; margin-top: 32px; }}
</style>
</head>
<body>
<div class="container">
<h1>{title}</h1>
<p class="subtitle">注册表 {total} 条 · 已有 PDF {has} 条 · Excel 缺码文件 {archive_only_count} 条 · 差异类型 {difference_label} · 覆盖率 {rate:.1f}%</p>

<div class="cards">
    <div class="card total"><div class="number">{total}</div><div class="label">注册表总数</div></div>
    <div class="card has"><div class="number">{has}</div><div class="label">已有 PDF</div></div>
    <div class="card missing"><div class="number">{no}</div><div class="label">缺失 PDF</div></div>
    <div class="card extra"><div class="number">{archive_only_count}</div><div class="label">Excel 缺码</div></div>
    <div class="card rate"><div class="number">{rate:.1f}%</div><div class="label">覆盖率</div></div>
</div>

<div class="progress-bar">
    <div class="progress-label">
        <span>差异类型</span><span>{difference_label}</span>
    </div>
    <div style="font-size:13px;color:#666;line-height:1.5;">
        单边缺失或双向不一致都会继续生成完整报告；请分别查看“缺失”和“Excel 缺码”页签。
    </div>
</div>

<div class="progress-bar">
    <div class="progress-label">
        <span>PDF 覆盖率</span><span>{has} / {total} ({rate:.1f}%)</span>
    </div>
    <div class="progress-track">
        <div class="progress-fill{' warning' if rate < 90 else ''}" style="width:{rate}%"></div>
    </div>
</div>

{diff_section}

<div class="toolbar">
    <div class="tabs">
        <button class="tab active" id="tab-missing" onclick="showTab('missing')">缺失 <strong>({no})</strong></button>
        <button class="tab" id="tab-extra" onclick="showTab('extra')">Excel 缺码 <strong>({archive_only_count})</strong></button>
        <button class="tab" id="tab-matched" onclick="showTab('matched')">已有 <strong>({has})</strong></button>
        <button class="tab" id="tab-all" onclick="showTab('all')">全部 <strong>({total})</strong></button>"""

    if lost_chassis:
        html += f"""
        <button class="tab" id="tab-lost" onclick="showTab('lost')">本期缺失 <strong>({len(lost_chassis)})</strong></button>"""

    html += f"""
    </div>
    <button class="btn-export danger" onclick="exportMissing()">导出缺失 COC</button>
    <button class="btn-export default" onclick="exportArchiveOnly()">导出 Excel 缺码</button>
    <button class="btn-export default" onclick="exportAll()">导出全部</button>
</div>

<div class="search-bar">
    <input type="text" id="search" placeholder="搜索底盘号 / Model ...">
</div>

<div class="section" id="section-missing">
    <div class="section-header">缺失 PDF 的 COC — 按 Model 分组</div>
    <table>
        <thead><tr><th>Chassis Number</th><th>Model</th><th>Country</th><th>Has PDF</th></tr></thead>
        <tbody>{missing_rows_html}</tbody>
    </table>
</div>

<div class="section hidden" id="section-extra">
    <div class="section-header">压缩包中存在、但 Excel 注册表中没有的文件</div>
    <table>
        <thead><tr><th>Archive Filename</th><th>Reason</th></tr></thead>
        <tbody>{archive_only_rows_html}</tbody>
    </table>
</div>

<div class="section hidden" id="section-matched">
    <div class="section-header">已有 PDF 的 COC（显示前 50 条，搜索可过滤全部）</div>
    <table>
        <thead><tr><th>Chassis Number</th><th>Model</th><th>Country</th></tr></thead>
        <tbody id="matched-tbody">{matched_rows_html}</tbody>
    </table>
</div>

<div class="section hidden" id="section-all">
    <div class="section-header">全部 COC</div>
    <table>
        <thead><tr><th>Chassis Number</th><th>Model</th><th>Country</th><th>Has PDF</th></tr></thead>
        <tbody id="all-tbody"></tbody>
    </table>
</div>

<div class="footer">Country: {country} · Month: {month} · Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}</div>
</div>

""" + js_code + """
</body>
</html>"""

    return html


# ── Background Job Runner ──────────────────────────────────────────

class CocMatchJobRunner(BaseJobRunner):
    """Background job runner for COC matching."""

    def __init__(
        self,
        job_id: str,
        state_dir: Path,
        excel_path: Path,
        archive_path: Path,
        country: str,
        month: str,
        file_ext: str,
        triggered_by: str,
    ) -> None:
        super().__init__(job_id, state_dir)
        self.excel_path = excel_path
        self.archive_path = archive_path
        self.country = country
        self.month = month
        self.file_ext = file_ext
        self.triggered_by = triggered_by

    def run(self) -> None:
        state = self.load_state()
        state["status"] = "running"
        state["startedAt"] = datetime.now().isoformat()
        self.persist_state(state)

        self.log(f"Country: {self.country}, Month: {self.month}, Ext: {self.file_ext}")
        self.log(f"Excel: {self.excel_path}")
        self.log(f"Archive: {self.archive_path}")

        # Step 1: Read Excel
        self.log("Reading Excel...")
        extensions = [self.file_ext]
        rows = read_excel_rows(self.excel_path)
        self.log(f"  {len(rows)} rows loaded")

        # Step 2: List archive
        self.log(f"Listing archive (ext={self.file_ext})...")
        file_set = list_archive_files(self.archive_path, extensions)
        self.log(f"  {len(file_set)} unique filenames")

        # Step 3: Match
        matched, missing = match_cocs(rows, file_set)
        archive_only_files = find_archive_only_files(rows, file_set)
        difference_type = classify_coc_difference(len(missing), len(archive_only_files))
        self.log(
            f"  Matched: {len(matched)}, Missing: {len(missing)}, "
            f"Archive-only: {len(archive_only_files)}, Difference: {difference_type}"
        )

        # Step 4/5: Save run history and compare with previous month.
        # SQLite has a single writer; serialize in-process writes and let
        # busy_timeout handle another backend process holding the file lock.
        with _COC_DB_LOCK:
            conn = _init_coc_db()
            try:
                run_id = _save_run(conn, self.country, self.month, rows)
                prev = _get_previous_run(conn, self.country, self.month)
                diff = _get_diff(conn, prev[0] if prev else None, rows) if prev else {}
            finally:
                conn.close()
        if prev:
            gained = sum(1 for v in diff.values() if v == "gained_pdf")
            g_lost = sum(1 for v in diff.values() if v == "lost_pdf")
            self.log(f"  vs {prev[1]}: +{gained} gained, -{g_lost} lost")

        # Step 6: Build HTML report
        html = build_html_report(
            rows,
            matched,
            missing,
            archive_only_files,
            self.country,
            self.month,
            prev,
            diff,
        )
        report_path = self.state_dir / self.job_id / "report.html"
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(html, encoding="utf-8")
        self.log(f"Report: {report_path}")

        # Step 7: Update state with results
        state = self.load_state()
        state["status"] = "success"
        state["phase"] = "completed"
        state["finishedAt"] = datetime.now().isoformat()
        state["totalRows"] = len(rows)
        state["matchedCount"] = len(matched)
        state["missingCount"] = len(missing)
        state["extraFileCount"] = len(archive_only_files)
        state["differenceType"] = difference_type
        state["hasBidirectionalMismatch"] = difference_type == "bidirectional_mismatch"
        state["coverageRate"] = round(len(matched) / len(rows) * 100, 1) if rows else 0
        if prev:
            state["previousRun"] = {"month": prev[1], "matched": prev[3], "total": prev[2]}
            state["diffSummary"] = {
                "gained": sum(1 for v in diff.values() if v == "gained_pdf"),
                "lost": sum(1 for v in diff.values() if v == "lost_pdf"),
                "newEntries": sum(1 for v in diff.values() if v == "new_entry"),
            }
        self.persist_state(state)
        self.log("Done!")


# ── API-facing Functions ───────────────────────────────────────────

def _normalize_filename(filename: str) -> str:
    return filename.strip().replace("\\", "/").split("/")[-1]


def create_coc_match_job(
    *,
    excel_file: UploadFile,
    archive_file: UploadFile,
    country: str,
    month: str | None,
    file_ext: str,
    triggered_by: str,
) -> dict[str, Any]:
    """Create a COC match job. Validates files, saves them, queues background job."""
    country = country.strip().upper()
    if not country:
        raise HTTPException(status_code=400, detail="国家代码不能为空。")
    if len(country) > 10:
        raise HTTPException(status_code=400, detail="国家代码过长。")

    month = month or datetime.now().strftime("%Y-%m")

    if not file_ext.startswith("."):
        file_ext = f".{file_ext}"
    if file_ext not in {".pdf", ".xml"}:
        raise HTTPException(status_code=400, detail="文件类型仅支持 .pdf 或 .xml。")

    excel_name = _normalize_filename(excel_file.filename or "upload.xlsx")
    archive_name = _normalize_filename(archive_file.filename or "archive.zip")

    if not allowed_extension(excel_name, allowed=ALLOWED_EXCEL_EXTENSIONS):
        raise HTTPException(
            status_code=400,
            detail=f"Excel 文件格式不支持: {excel_name}（仅支持 .xlsx/.xlsm/.xls）",
        )

    archive_suffix = Path(archive_name).suffix.lower()
    if archive_suffix not in ALLOWED_ARCHIVE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"压缩包格式不支持: {archive_name}（仅支持 .zip/.rar）",
        )

    job_id = f"coc-match-{uuid.uuid4().hex[:8]}"
    state_dir = COC_MATCH_JOB_ROOT / job_id
    state_dir.mkdir(parents=True, exist_ok=True)

    excel_dst = state_dir / f"excel-{excel_name}"
    archive_dst = state_dir / f"archive-{archive_name}"

    # Save uploaded files
    excel_bytes = excel_file.file.read()
    archive_bytes = archive_file.file.read()
    excel_dst.write_bytes(excel_bytes)
    archive_dst.write_bytes(archive_bytes)

    # Initialize job state
    initial_state: dict[str, Any] = {
        "jobId": job_id,
        "jobType": "match",
        "status": "queued",
        "phase": "pending",
        "country": country,
        "month": month,
        "fileExt": file_ext,
        "excelFilename": excel_name,
        "archiveFilename": archive_name,
        "triggeredBy": triggered_by,
        "totalRows": None,
        "matchedCount": None,
        "missingCount": None,
        "extraFileCount": None,
        "differenceType": None,
        "hasBidirectionalMismatch": False,
        "coverageRate": None,
        "previousRun": None,
        "diffSummary": None,
        "error": None,
        "createdAt": datetime.now().isoformat(),
        "startedAt": None,
        "finishedAt": None,
    }
    persist_job_state(state_path(COC_MATCH_JOB_ROOT, job_id), initial_state)

    # Launch background job
    runner = CocMatchJobRunner(
        job_id=job_id,
        state_dir=COC_MATCH_JOB_ROOT,
        excel_path=excel_dst,
        archive_path=archive_dst,
        country=country,
        month=month,
        file_ext=file_ext,
        triggered_by=triggered_by,
    )
    _RUNNING_THREADS[job_id] = runner
    runner.start()
    _trim_job_history()

    return load_job_state(state_path(COC_MATCH_JOB_ROOT, job_id))


_MAX_HISTORY = 100


def _trim_job_history() -> None:
    """Keep at most _MAX_HISTORY jobs. Delete oldest by createdAt."""
    payloads = list_job_payloads(COC_MATCH_JOB_ROOT)
    if len(payloads) <= _MAX_HISTORY:
        return
    payloads.sort(key=lambda p: str(p.get("createdAt", "")))
    for old in payloads[:-_MAX_HISTORY]:
        jid = str(old.get("jobId", ""))
        if jid:
            path = COC_MATCH_JOB_ROOT / jid
            if path.exists():
                import shutil
                shutil.rmtree(path, ignore_errors=True)


# ── Chunked Upload Session Management ──────────────────────────────

_COC_UPLOAD_SESSION_ROOT = COC_MATCH_JOB_ROOT / "_upload_sessions"


def initiate_coc_match_upload(
    *,
    filename: str,
    size_bytes: int,
    resume_key: str | None = None,
    triggered_by: str = "anonymous",
) -> dict[str, Any]:
    """Initiate a chunked upload session for a COC match file.

    For small files (under 50 MB) the caller should use the simple POST
    endpoint instead; this is only for large archives.
    """
    suffix = Path(filename).suffix.lower()
    if suffix in ALLOWED_EXCEL_EXTENSIONS:
        pass  # valid
    elif suffix in ALLOWED_ARCHIVE_EXTENSIONS:
        pass  # valid
    else:
        raise HTTPException(status_code=400, detail=f"文件格式不支持: {filename}")

    from upload_toolkit.upload_engine import (
        create_upload_session as _toolkit_create,
        find_resumable_session,
    )

    # Resume check
    if resume_key:
        existing = find_resumable_session(
            _COC_UPLOAD_SESSION_ROOT,
            resume_key=resume_key,
            filename=filename,
            size_bytes=size_bytes,
        )
        if existing:
            return existing

    return _toolkit_create(
        _COC_UPLOAD_SESSION_ROOT,
        filename=filename,
        size_bytes=size_bytes,
        chunk_size=_COC_UPLOAD_CHUNK_SIZE,
        resume_key=resume_key or None,
        triggered_by=triggered_by,
    )


def upload_coc_match_chunk(
    upload_id: str,
    part_number: int,
    content: bytes,
    chunk_sha256: str | None = None,
) -> dict[str, Any]:
    from upload_toolkit.upload_engine import receive_chunk as _toolkit_chunk

    return _toolkit_chunk(
        _COC_UPLOAD_SESSION_ROOT,
        upload_id,
        part_number,
        content,
        chunk_sha256=chunk_sha256,
        expected_chunk_size=_COC_UPLOAD_CHUNK_SIZE,
    )


def complete_coc_match_upload(upload_id: str) -> dict[str, Any]:
    from upload_toolkit.upload_engine import complete_upload_session as _toolkit_complete

    return _toolkit_complete(_COC_UPLOAD_SESSION_ROOT, upload_id)


def _get_assembled_path(upload_id: str) -> Path:
    """Get the assembled file path for a completed upload session."""
    from upload_toolkit.upload_engine import get_upload_session as _toolkit_get

    state = _toolkit_get(_COC_UPLOAD_SESSION_ROOT, upload_id)
    if state.get("status") != "completed":
        raise HTTPException(status_code=400, detail="上传未完成。")
    return _COC_UPLOAD_SESSION_ROOT / upload_id / "assembled" / str(state.get("filename", "upload.bin"))


_MEGA_CHUNK_SIZE = 50 * 1024 * 1024  # 50 MB threshold for auto-chunking


def create_coc_match_job_from_upload(
    *,
    excel_upload_id: str,
    archive_upload_id: str,
    excel_filename: str,
    archive_filename: str,
    country: str,
    month: str | None,
    file_ext: str,
    triggered_by: str,
) -> dict[str, Any]:
    """Create a COC match job from two completed chunked uploads."""
    country = country.strip().upper()
    if not country:
        raise HTTPException(status_code=400, detail="国家代码不能为空。")
    if len(country) > 10:
        raise HTTPException(status_code=400, detail="国家代码过长。")

    month = month or datetime.now().strftime("%Y-%m")

    if not file_ext.startswith("."):
        file_ext = f".{file_ext}"
    if file_ext not in {".pdf", ".xml"}:
        raise HTTPException(status_code=400, detail="文件类型仅支持 .pdf 或 .xml。")

    excel_path = _get_assembled_path(excel_upload_id)
    archive_path = _get_assembled_path(archive_upload_id)

    job_id = f"coc-match-{uuid.uuid4().hex[:8]}"
    state_dir = COC_MATCH_JOB_ROOT / job_id
    state_dir.mkdir(parents=True, exist_ok=True)

    # Copy assembled files into job directory
    excel_name = _normalize_filename(excel_filename)
    archive_name = _normalize_filename(archive_filename)
    excel_dst = state_dir / f"excel-{excel_name}"
    archive_dst = state_dir / f"archive-{archive_name}"
    excel_dst.write_bytes(excel_path.read_bytes())
    archive_dst.write_bytes(archive_path.read_bytes())

    # Initialize job state
    initial_state: dict[str, Any] = {
        "jobId": job_id,
        "jobType": "match",
        "status": "queued",
        "phase": "pending",
        "country": country,
        "month": month,
        "fileExt": file_ext,
        "excelFilename": excel_name,
        "archiveFilename": archive_name,
        "triggeredBy": triggered_by,
        "totalRows": None,
        "matchedCount": None,
        "missingCount": None,
        "extraFileCount": None,
        "differenceType": None,
        "hasBidirectionalMismatch": False,
        "coverageRate": None,
        "previousRun": None,
        "diffSummary": None,
        "error": None,
        "createdAt": datetime.now().isoformat(),
        "startedAt": None,
        "finishedAt": None,
    }
    persist_job_state(state_path(COC_MATCH_JOB_ROOT, job_id), initial_state)

    runner = CocMatchJobRunner(
        job_id=job_id,
        state_dir=COC_MATCH_JOB_ROOT,
        excel_path=excel_dst,
        archive_path=archive_dst,
        country=country,
        month=month,
        file_ext=file_ext,
        triggered_by=triggered_by,
    )
    _RUNNING_THREADS[job_id] = runner
    runner.start()
    _trim_job_history()

    return load_job_state(state_path(COC_MATCH_JOB_ROOT, job_id))


def list_coc_match_jobs(limit: int = 20, country: str | None = None) -> dict[str, Any]:
    payloads = [
        payload
        for payload in list_job_payloads(COC_MATCH_JOB_ROOT)
        if str(payload.get("jobType", "match")) == "match"
    ]
    if country:
        country_filter = country.strip().upper()
        payloads = [
            payload
            for payload in payloads
            if str(payload.get("country", "")).strip().upper() == country_filter
        ]
    payloads.sort(key=lambda p: str(p.get("createdAt", "")), reverse=True)
    return {"items": payloads[:limit]}


def get_coc_match_job(job_id: str) -> dict[str, Any]:
    sp = state_path(COC_MATCH_JOB_ROOT, job_id)
    if not sp.exists():
        raise HTTPException(status_code=404, detail=f"COC 匹配任务不存在: {job_id}")
    payload = load_job_state(sp)
    if str(payload.get("jobType", "match")) != "match":
        raise HTTPException(status_code=404, detail=f"COC 匹配任务不存在: {job_id}")
    return payload


def get_coc_match_report_path(job_id: str) -> Path:
    report = COC_MATCH_JOB_ROOT / job_id / "report.html"
    if not report.exists():
        raise HTTPException(status_code=404, detail="报告尚未生成。")
    return report


def retry_failed_coc_match_job(
    *,
    source_job_id: str,
    triggered_by: str,
) -> dict[str, Any]:
    source_state = get_coc_match_job(source_job_id)
    if source_state.get("status") not in ("failed",):
        raise HTTPException(status_code=400, detail="仅支持重试失败的任务。")

    excel_dst = COC_MATCH_JOB_ROOT / source_job_id / f"excel-{source_state.get('excelFilename', 'upload.xlsx')}"
    archive_dst = COC_MATCH_JOB_ROOT / source_job_id / f"archive-{source_state.get('archiveFilename', 'archive.zip')}"

    if not excel_dst.exists() or not archive_dst.exists():
        raise HTTPException(status_code=400, detail="原始上传文件已不存在，无法重试。")

    # Re-upload using stored files
    from fastapi import UploadFile

    class _ReplayFile:
        def __init__(self, path: Path, name: str):
            self._path = path
            self.filename = name

        def read(self):
            return self._path.read_bytes()

    excel_replay = UploadFile(
        filename=source_state.get("excelFilename", "upload.xlsx"),
        file=None,
    )
    archive_replay = UploadFile(
        filename=source_state.get("archiveFilename", "archive.zip"),
        file=None,
    )
    # Manually set the file attribute since UploadFile expects a file-like
    excel_replay.file = _ReplayFile(excel_dst, source_state.get("excelFilename", "upload.xlsx"))
    archive_replay.file = _ReplayFile(archive_dst, source_state.get("archiveFilename", "archive.zip"))

    return create_coc_match_job(
        excel_file=excel_replay,
        archive_file=archive_replay,
        country=str(source_state.get("country", "")),
        month=str(source_state.get("month", "")),
        file_ext=str(source_state.get("fileExt", ".pdf")),
        triggered_by=triggered_by,
    )
