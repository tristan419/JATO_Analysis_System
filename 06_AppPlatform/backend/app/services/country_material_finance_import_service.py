"""Parse country material finance / CBU imports into BOM-template updates."""

from __future__ import annotations

import io
import base64
import json
import re
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import openpyxl

from app.services import news_digest_service
from app.services.ordering_normalization import clean_text


FinanceImportField = str

DEFAULT_FIELDS: list[FinanceImportField] = [
    "materialCode",
    "fobEur",
    "retailPriceEur",
    "wholesalePriceEur",
    "dealerPriceEur",
    "costEur",
    "memo",
]

ASCII_HEADER_ALIASES: dict[str, FinanceImportField] = {
    "material": "materialCode",
    "materialcode": "materialCode",
    "bom": "materialCode",
    "bomcode": "materialCode",
    "bomtemplate": "materialCode",
    "template": "materialCode",
    "fob": "fobEur",
    "fobeur": "fobEur",
    "retail": "retailPriceEur",
    "retailprice": "retailPriceEur",
    "retailpriceeur": "retailPriceEur",
    "wholesale": "wholesalePriceEur",
    "wholesaleprice": "wholesalePriceEur",
    "wholesalepriceeur": "wholesalePriceEur",
    "dealer": "dealerPriceEur",
    "dealerprice": "dealerPriceEur",
    "dealerpriceeur": "dealerPriceEur",
    "cost": "costEur",
    "costeur": "costEur",
    "cbu": "costEur",
    "unitmargin": "vehicleMarginEur",
    "vehiclemargin": "vehicleMarginEur",
    "vehiclemargineur": "vehicleMarginEur",
    "margin": "vehicleMarginEur",
    "marginrate": "vehicleMarginRate",
    "marginpercent": "vehicleMarginRate",
    "marginpct": "vehicleMarginRate",
    "unitprofit": "vehicleProfitEur",
    "vehicleprofit": "vehicleProfitEur",
    "vehicleprofiteur": "vehicleProfitEur",
    "profit": "vehicleProfitEur",
    "profitrate": "vehicleProfitRate",
    "profitpercent": "vehicleProfitRate",
    "profitpct": "vehicleProfitRate",
    "fobdelta": "fobDeltaEur",
    "fobadjust": "fobDeltaEur",
    "fobadjustment": "fobDeltaEur",
    "margindelta": "marginDeltaEur",
    "marginadjust": "marginDeltaEur",
    "marginadjustment": "marginDeltaEur",
    "note": "memo",
    "memo": "memo",
    "remark": "memo",
}

RATE_FIELDS = {"vehicleMarginRate", "vehicleProfitRate", "marginRate"}
TEXT_FIELDS = {"memo", "materialCode"}


def _normalize_header(value: str) -> str:
    return value.strip().lower().replace("%", "percent").replace("Δ", "delta").replace("△", "delta")


def _compact_ascii(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def _field_from_header(value: str) -> FinanceImportField | None:
    header = _normalize_header(value)
    if not header:
        return None
    if "物料" in header or "料号" in header:
        return "materialCode"
    if "边际" in header and ("增" in header or "调" in header or "差" in header):
        return "marginDeltaEur"
    if "fob" in header and ("增" in header or "调" in header or "差" in header or "delta" in header):
        return "fobDeltaEur"
    if "单车边际" in header or "车辆边际" in header:
        return "vehicleMarginEur"
    if "边际率" in header:
        return "vehicleMarginRate"
    if "单车利润" in header or "车辆利润" in header:
        return "vehicleProfitEur"
    if "利润率" in header:
        return "vehicleProfitRate"
    if "成本" in header:
        return "costEur"
    if "零售" in header or "建议售价" in header:
        return "retailPriceEur"
    if "批发" in header:
        return "wholesalePriceEur"
    if "经销" in header:
        return "dealerPriceEur"
    if "备注" in header or "说明" in header:
        return "memo"
    compact = _compact_ascii(header)
    return ASCII_HEADER_ALIASES.get(compact)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _split_line(line: str) -> list[str]:
    if "\t" in line:
        return line.split("\t")
    if "," in line:
        return line.split(",")
    return re.split(r"\s{2,}", line.strip())


def _parse_number(value: str, *, rate: bool) -> tuple[float | None, str]:
    text = value.strip()
    if not text:
        return None, ""
    negative = text.startswith("(") and text.endswith(")")
    normalized = (
        text.replace("€", "")
        .replace("EUR", "")
        .replace("eur", "")
        .replace("%", "")
        .replace(",", "")
        .replace(" ", "")
        .replace("(", "")
        .replace(")", "")
    )
    try:
        parsed = float(normalized)
    except ValueError:
        return None, f"{value} is not numeric"
    if negative:
        parsed = -parsed
    if rate and abs(parsed) > 1:
        parsed = round(parsed / 100, 6)
    return parsed, ""


def _build_update(
    country_code: str,
    cells: list[str],
    fields: list[FinanceImportField | None],
    *,
    source_mode: str,
    source_payload: dict[str, Any],
) -> tuple[dict[str, Any] | None, str]:
    update: dict[str, Any] = {
        "countryCode": country_code,
        "sourceMode": source_mode,
        "sourcePayload": source_payload,
    }
    errors: list[str] = []
    changed = False
    for index, field_name in enumerate(fields):
        if not field_name or field_name == "materialCode":
            continue
        raw_value = cells[index] if index < len(cells) else ""
        if field_name in TEXT_FIELDS:
            text = clean_text(raw_value)
            if text:
                update[field_name] = text
                changed = True
            continue
        parsed, error = _parse_number(raw_value, rate=field_name in RATE_FIELDS)
        if error:
            errors.append(f"{field_name}: {error}")
        elif parsed is not None:
            update[field_name] = parsed
            changed = True
    if errors:
        return None, "; ".join(errors)
    return (update if changed else None), ""


def parse_country_material_finance_cells(
    rows: list[list[str]],
    country_code: str,
    *,
    source_mode: str,
    source_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    country = clean_text(country_code).upper()
    non_empty_rows = [
        (index + 1, [cell.strip() for cell in row])
        for index, row in enumerate(rows)
        if any(cell.strip() for cell in row)
    ]
    if not non_empty_rows:
        return {"rows": [], "warnings": ["No rows found in import source."]}

    first_line_number, first_cells = non_empty_rows[0]
    mapped_headers = [_field_from_header(cell) for cell in first_cells]
    has_header = "materialCode" in mapped_headers
    fields = mapped_headers if has_header else DEFAULT_FIELDS
    data_rows = non_empty_rows[1:] if has_header else non_empty_rows
    first_payload = source_payload or {}
    parsed_rows: list[dict[str, Any]] = []

    for line_number, cells in data_rows:
        material_index = fields.index("materialCode") if "materialCode" in fields else 0
        material_code = clean_text(cells[material_index] if material_index < len(cells) else "").upper()
        payload = {
            **first_payload,
            "hasHeader": has_header,
            "sourceLineNumber": line_number,
        }
        update, parse_error = _build_update(
            country,
            cells,
            fields,
            source_mode=source_mode,
            source_payload=payload,
        )
        error = parse_error
        if not material_code:
            error = "Missing material code"
        elif "**" not in material_code:
            error = f"{material_code} is not a BOM template"
        elif update is None and not error:
            error = f"{material_code} has no finance values"
        parsed_rows.append(
            {
                "lineNumber": line_number,
                "materialCode": material_code,
                "update": update,
                "error": error,
            }
        )

    warnings: list[str] = []
    if not has_header:
        warnings.append(
            f"No header detected on line {first_line_number}; using Material Code, FOB, Retail, Wholesale, Dealer, Cost, Note."
        )
    return {"rows": parsed_rows, "warnings": warnings}


def parse_country_material_finance_text(
    text: str,
    country_code: str,
    *,
    source_mode: str = "uploaded",
    source_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    rows = [
        [_cell_text(cell) for cell in _split_line(line)]
        for line in text.splitlines()
        if line.strip()
    ]
    return parse_country_material_finance_cells(
        rows,
        country_code,
        source_mode=source_mode,
        source_payload=source_payload,
    )


def parse_country_material_finance_xlsx(
    file_bytes: bytes,
    country_code: str,
    *,
    file_name: str,
) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = [
        [_cell_text(cell) for cell in row]
        for row in sheet.iter_rows(values_only=True)
    ]
    result = parse_country_material_finance_cells(
        rows,
        country_code,
        source_mode="uploaded",
        source_payload={"entryMode": "xlsx_upload", "fileName": file_name, "sheetName": sheet.title},
    )
    result["warnings"] = [
        f"Parsed sheet: {sheet.title}",
        *result.get("warnings", []),
    ]
    return result


def _extract_gemini_text(payload: dict[str, Any]) -> str:
    parts: list[str] = []
    for candidate in payload.get("candidates", []) or []:
        if not isinstance(candidate, dict):
            continue
        content = candidate.get("content")
        if not isinstance(content, dict):
            continue
        for part in content.get("parts", []) or []:
            if isinstance(part, dict) and isinstance(part.get("text"), str):
                parts.append(part["text"])
    return "\n".join(parts).strip()


def _parse_country_material_finance_image_with_gemini(
    file_bytes: bytes,
    country_code: str,
    *,
    file_name: str,
    mime_type: str,
) -> dict[str, Any] | None:
    api_key = news_digest_service._gemini_api_key()  # noqa: SLF001
    if not api_key:
        return None

    prompt = (
        "Read this CBU/margin table image and return only a tab-separated table. "
        "Use this exact header: Material Code\tFOB\tUnit Margin\tMargin %\t"
        "Unit Profit\tProfit %\tFOB Delta\tMargin Delta\tNote. "
        "Use BOM-template material codes containing ** when the image shows colour-specific material codes; "
        "preserve missing values as empty cells."
    )
    request_body = {
        "contents": [
            {
                "role": "user",
                "parts": [
                    {"text": prompt},
                    {
                        "inline_data": {
                            "mime_type": mime_type or "image/png",
                            "data": base64.b64encode(file_bytes).decode("ascii"),
                        }
                    },
                ],
            }
        ],
        "generationConfig": {
            "temperature": 0.0,
        },
    }
    request = Request(
        (
            "https://generativelanguage.googleapis.com/v1beta/"
            f"models/{news_digest_service.DEFAULT_GEMINI_MODEL}:generateContent"
            f"?key={api_key}"
        ),
        data=json.dumps(request_body).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urlopen(request, timeout=45) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except (HTTPError, URLError, TimeoutError, ValueError) as exc:
        return {
            "rows": [],
            "warnings": [f"Gemini image parsing failed: {exc}"],
        }
    text = _extract_gemini_text(payload)
    if not text:
        return {"rows": [], "warnings": ["Gemini image parsing returned no table text."]}
    result = parse_country_material_finance_text(
        text,
        country_code,
        source_mode="uploaded",
        source_payload={"entryMode": "image_gemini", "fileName": file_name},
    )
    result["warnings"] = ["Gemini image table parsed; verify values before applying.", *result.get("warnings", [])]
    return result


def parse_country_material_finance_image(
    file_bytes: bytes,
    country_code: str,
    *,
    file_name: str,
    mime_type: str = "image/png",
) -> dict[str, Any]:
    gemini_result = _parse_country_material_finance_image_with_gemini(
        file_bytes,
        country_code,
        file_name=file_name,
        mime_type=mime_type,
    )
    if gemini_result is not None:
        return gemini_result

    try:
        from PIL import Image
        import pytesseract
    except Exception:
        return {
            "rows": [],
            "warnings": [
                "Image parsing is wired to the same digest flow, but neither Gemini vision nor local OCR is configured on this server.",
                "Configure GEMINI_API_KEY/GOOGLE_API_KEY or install OCR dependencies to parse image tables.",
            ],
        }

    try:
        image = Image.open(io.BytesIO(file_bytes))
        text = pytesseract.image_to_string(image)
    except Exception as exc:
        return {
            "rows": [],
            "warnings": [f"Image OCR failed: {exc}"],
        }
    if not text.strip():
        return {"rows": [], "warnings": ["Image OCR returned no text."]}
    result = parse_country_material_finance_text(
        text,
        country_code,
        source_mode="uploaded",
        source_payload={"entryMode": "image_ocr", "fileName": file_name},
    )
    result["warnings"] = ["Image OCR text parsed; verify values before applying.", *result.get("warnings", [])]
    return result
