"""Exports for engineering config compare tables."""

from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="172033", end_color="172033", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="172033")
NORMAL_FONT = Font(name="Calibri", size=11)
MUTED_FONT = Font(name="Calibri", size=10, color="64748B")
DIFF_FILL = PatternFill(start_color="FFF7E6", end_color="FFF7E6", fill_type="solid")
STRIPED_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)
LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


COMPARE_HEADERS = ["配置项", "大类", "差异类型"]
PDF_PAGE_WIDTH = 842
PDF_PAGE_HEIGHT = 595
PDF_MARGIN_X = 36
PDF_MARGIN_TOP = 34
PDF_LINE_HEIGHT = 13
PDF_BOTTOM_Y = 34
EVIDENCE_HEADERS = [
    "配置项",
    "大类",
    "配置列",
    "显示值",
    "原始值",
    "valueState",
    "availability",
    "inferred",
    "inferenceReason",
    "confidence",
    "sheetName",
    "rowNumber",
    "columnLetter",
    "cell",
    "sourceCell",
    "mergedRange",
    "ocrEngine",
]


def _scope_context_parts(scope: dict[str, Any]) -> list[str]:
    parts = [
        f"范围 {scope.get('rangeLabel')}" if scope.get("rangeLabel") else None,
        f"基准 {scope.get('baseLabel')}" if scope.get("baseLabel") else None,
        f"目标 {scope.get('targetLabel')}" if scope.get("targetLabel") else None,
        f"大类 {scope.get('categoryLabel')}" if scope.get("categoryLabel") else None,
        f"搜索 {scope.get('searchLabel')}" if scope.get("searchLabel") else None,
    ]
    return [str(part) for part in parts if part]


def generate_engineering_config_compare_xlsx(payload: dict[str, Any]) -> io.BytesIO:
    trims = _list_of_dicts(payload.get("trims"))
    rows = _list_of_dicts(payload.get("rows"))
    scope = _dict(payload.get("scope"))
    summary = _dict(payload.get("summary"))
    evidence_summary = _evidence_summary(payload)
    business_summary = _business_summary_items(payload)
    business_summary_usage = _business_summary_usage(payload)

    wb = openpyxl.Workbook()
    compare_sheet = wb.active
    compare_sheet.title = "Config Compare"
    if business_summary:
        summary_sheet = wb.create_sheet("AI Summary")
    evidence_sheet = wb.create_sheet("Evidence")

    _write_compare_sheet(compare_sheet, trims, rows, scope, summary, evidence_summary)
    if business_summary:
        _write_business_summary_sheet(summary_sheet, business_summary, business_summary_usage, scope)
    _write_evidence_sheet(evidence_sheet, trims, rows)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_engineering_config_compare_pdf(payload: dict[str, Any]) -> io.BytesIO:
    trims = _list_of_dicts(payload.get("trims"))
    rows = _list_of_dicts(payload.get("rows"))
    scope = _dict(payload.get("scope"))
    summary = _dict(payload.get("summary"))
    evidence_summary = _evidence_summary(payload)
    business_summary = _business_summary_items(payload)
    business_summary_usage = _business_summary_usage(payload)

    lines = _build_pdf_lines(trims, rows, scope, summary, evidence_summary, business_summary, business_summary_usage)
    pages = _paginate_pdf_lines(lines)
    return _write_simple_pdf(pages)


def compare_export_filename(payload: dict[str, Any]) -> str:
    explicit = str(payload.get("fileName") or "").strip()
    if explicit:
        return _safe_filename(explicit, fallback="config-compare.xlsx")
    trims = _list_of_dicts(payload.get("trims"))
    trim_label = "-vs-".join(_safe_token(_trim_label(trim)) for trim in trims[:4])
    if not trim_label:
        trim_label = "config-compare"
    return f"{trim_label[:90]}-{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"


def compare_pdf_export_filename(payload: dict[str, Any]) -> str:
    explicit = str(payload.get("fileName") or "").strip()
    if explicit:
        base = explicit.replace("\\", "/").rsplit("/", 1)[-1].strip()
        base = re.sub(r"\.(xlsx|pdf)$", "", base, flags=re.IGNORECASE)
        return _safe_filename(base, fallback="config-compare.pdf").replace(".xlsx", ".pdf")
    xlsx_name = compare_export_filename(payload)
    return re.sub(r"\.xlsx$", ".pdf", xlsx_name, flags=re.IGNORECASE)


def _write_compare_sheet(
    ws,
    trims: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    scope: dict[str, Any],
    summary: dict[str, Any],
    evidence_summary: dict[str, Any],
) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(5, 4 + len(trims)))
    title = ws.cell(row=1, column=1, value=str(payload_title(scope) or "Engineering Config Compare"))
    title.font = TITLE_FONT
    title.alignment = LEFT_ALIGN

    meta_parts = [
        f"导出时间 {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        f"当前展示 {len(rows)}",
        f"总配置项 {summary.get('totalFeatures')}" if summary.get("totalFeatures") is not None else None,
        *_scope_context_parts(scope),
    ]
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(5, 4 + len(trims)))
    meta = ws.cell(row=2, column=1, value=" · ".join(part for part in meta_parts if part))
    meta.font = MUTED_FONT
    meta.alignment = LEFT_ALIGN

    evidence_summary_text = _evidence_summary_line(evidence_summary)
    if evidence_summary_text:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(5, 4 + len(trims)))
        evidence_meta = ws.cell(row=3, column=1, value=evidence_summary_text)
        evidence_meta.font = MUTED_FONT
        evidence_meta.alignment = LEFT_ALIGN

    headers = [*COMPARE_HEADERS, *[_trim_export_label(trim) for trim in trims], "业务备注"]
    _write_header(ws, 4, headers)

    for row_index, row in enumerate(rows, start=5):
        values = _list(row.get("values"))
        comparison_type = str(row.get("comparisonType") or "")
        row_fill = DIFF_FILL if comparison_type != "COMMON_SAME" else (STRIPED_FILL if row_index % 2 == 0 else None)
        row_values = [
            str(row.get("featureName") or ""),
            str(row.get("category") or ""),
            _comparison_label(comparison_type),
            *[_cell_display(_dict_or_none(values[index])) for index in range(len(trims))],
            str(row.get("businessNote") or ""),
        ]
        for column_index, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, len(rows) + 4)}"
    _set_widths(ws, [34, 18, 16, *([24] * len(trims)), 42])


def _write_evidence_sheet(ws, trims: list[dict[str, Any]], rows: list[dict[str, Any]]) -> None:
    _write_header(ws, 1, EVIDENCE_HEADERS)
    excel_row = 2
    for row in rows:
        values = _list(row.get("values"))
        for trim_index, trim in enumerate(trims):
            cell_value = _dict_or_none(values[trim_index]) if trim_index < len(values) else None
            source = _dict(cell_value.get("source")) if cell_value else {}
            fields = [
                str(row.get("featureName") or ""),
                str(row.get("category") or ""),
                _trim_export_label(trim),
                _cell_display(cell_value),
                str(cell_value.get("rawValue") or "") if cell_value else "",
                str(cell_value.get("valueState") or "") if cell_value else "",
                str(cell_value.get("availability") or "") if cell_value else "",
                "yes" if cell_value and cell_value.get("inferred") else "",
                str(cell_value.get("inferenceReason") or source.get("inferenceReason") or "") if cell_value else "",
                cell_value.get("confidence") if cell_value else "",
                str(source.get("sheetName") or ""),
                source.get("rowNumber") if source.get("rowNumber") is not None else "",
                str(source.get("columnLetter") or ""),
                str(source.get("cell") or ""),
                str(source.get("sourceCell") or ""),
                str(source.get("mergedRange") or ""),
                str(source.get("ocrEngine") or ""),
            ]
            for column_index, value in enumerate(fields, start=1):
                cell = ws.cell(row=excel_row, column=column_index, value=value)
                cell.font = NORMAL_FONT
                cell.alignment = LEFT_ALIGN
                cell.border = THIN_BORDER
            excel_row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EVIDENCE_HEADERS))}{max(1, excel_row - 1)}"
    _set_widths(ws, [34, 18, 28, 18, 18, 16, 18, 10, 34, 12, 24, 12, 14, 12, 12, 18, 16])


def _write_business_summary_sheet(
    ws,
    summaries: list[dict[str, Any]],
    usage: dict[str, Any] | None = None,
    scope: dict[str, Any] | None = None,
) -> None:
    _write_header(ws, 1, [
        "目标配置列",
        "AI 结论",
        "主要升级",
        "减少或替换",
        "证据状态",
        "证据引用",
        "建议用途",
        "LLM Provider",
        "LLM Model",
        "LLM Status",
        "LLM Tokens",
        "Finish Reason",
        "Transport Fallback",
        "导出口径",
    ])
    usage_values = _business_summary_usage_values(usage)
    scope_line = _business_summary_scope_line(scope or {})
    for row_index, item in enumerate(summaries, start=2):
        values = [
            str(item.get("targetLabel") or item.get("targetTrimId") or ""),
            str(item.get("headline") or ""),
            "\n".join(_string_list(item.get("mainUpgrades"))),
            "\n".join(_string_list(item.get("replacementsOrReductions"))),
            "\n".join(_string_list(item.get("evidenceStatus"))),
            "\n".join(_business_summary_evidence_ref_lines(item)),
            str(item.get("recommendedUse") or ""),
            *usage_values,
            scope_line,
        ]
        for column_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{max(1, len(summaries) + 1)}"
    _set_widths(ws, [28, 48, 46, 42, 42, 48, 42, 18, 22, 16, 14, 18, 20, 58])


def _business_summary_scope_line(scope: dict[str, Any]) -> str:
    parts = [
        payload_title(scope),
        *_scope_context_parts(scope),
    ]
    return " · ".join(part for part in parts if part)


def payload_title(scope: dict[str, Any]) -> str:
    title = str(scope.get("title") or "").strip()
    return title


def _build_pdf_lines(
    trims: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    scope: dict[str, Any],
    summary: dict[str, Any],
    evidence_summary: dict[str, Any],
    business_summary: list[dict[str, Any]] | None = None,
    business_summary_usage: dict[str, Any] | None = None,
) -> list[str]:
    title = payload_title(scope) or "Engineering Config Compare"
    meta_parts = [
        f"Exported {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        f"Rows {len(rows)}",
        f"Total features {summary.get('totalFeatures')}" if summary.get("totalFeatures") is not None else None,
        *[_ascii_safe(part) for part in _scope_context_parts(scope)],
    ]
    evidence_summary_text = _evidence_summary_line(evidence_summary)
    lines = [
        _ascii_safe(title),
        " | ".join(part for part in meta_parts if part),
        _ascii_safe(evidence_summary_text) if evidence_summary_text else None,
        "",
        "Columns: " + " | ".join(_ascii_safe(_trim_export_label(trim)) for trim in trims),
        "",
    ]
    lines = [line for line in lines if line is not None]
    if business_summary:
        lines.extend(_business_summary_pdf_lines(business_summary, business_summary_usage))
        lines.append("")
    for row_number, row in enumerate(rows, start=1):
        values = _list(row.get("values"))
        feature_name = _ascii_safe(str(row.get("featureName") or ""))
        category = _ascii_safe(str(row.get("category") or ""))
        comparison_type = _comparison_label(str(row.get("comparisonType") or ""))
        lines.extend(_wrap_pdf_text(f"{row_number}. {feature_name}", 108))
        lines.append(_ascii_safe(f"Category: {category} | Type: {comparison_type}"))
        for index, trim in enumerate(trims):
            cell = _dict_or_none(values[index]) if index < len(values) else None
            display = _ascii_safe(_cell_display(cell))
            prefix = _ascii_safe(_trim_export_label(trim))
            lines.extend(_wrap_pdf_text(f"- {prefix}: {display}", 112))
            source = _dict(cell.get("source")) if cell else {}
            if cell and (cell.get("inferred") or source.get("mergedRange") or source.get("cell")):
                evidence_parts = [
                    "inferred" if cell.get("inferred") else None,
                    f"reason={_ascii_safe(str(cell.get('inferenceReason') or source.get('inferenceReason') or ''))}" if cell.get("inferred") else None,
                    f"row={source.get('rowNumber')}" if source.get("rowNumber") is not None else None,
                    f"column={source.get('columnLetter')}" if source.get("columnLetter") else None,
                    f"cell={source.get('cell')}" if source.get("cell") else None,
                    f"sourceCell={source.get('sourceCell')}" if source.get("sourceCell") else None,
                    f"merged={source.get('mergedRange')}" if source.get("mergedRange") else None,
                    f"ocrEngine={source.get('ocrEngine')}" if source.get("ocrEngine") else None,
                ]
                evidence = " ; ".join(part for part in evidence_parts if part)
                if evidence:
                    lines.extend(_wrap_pdf_text(f"  evidence: {evidence}", 112))
        business_note = _ascii_safe(str(row.get("businessNote") or ""))
        if business_note:
            lines.extend(_wrap_pdf_text(f"Note: {business_note}", 112))
        lines.append("-" * 96)
    return lines


def _business_summary_pdf_lines(
    summaries: list[dict[str, Any]],
    usage: dict[str, Any] | None = None,
) -> list[str]:
    lines = ["AI 业务摘要"]
    usage_line = _business_summary_usage_line(usage)
    if usage_line:
        lines.extend(_wrap_pdf_text(usage_line, 112))
    for index, item in enumerate(summaries, start=1):
        target = str(item.get("targetLabel") or item.get("targetTrimId") or f"Target {index}")
        headline = str(item.get("headline") or "")
        lines.extend(_wrap_pdf_text(f"{index}. {target}: {headline}", 112))
        upgrades = _string_list(item.get("mainUpgrades"))
        if upgrades:
            lines.append("  主要升级:")
            for upgrade in upgrades[:5]:
                lines.extend(_wrap_pdf_text(f"  - {upgrade}", 112))
        reductions = _string_list(item.get("replacementsOrReductions"))
        if reductions:
            lines.append("  减少或替换:")
            for reduction in reductions[:4]:
                lines.extend(_wrap_pdf_text(f"  - {reduction}", 112))
        evidence = _string_list(item.get("evidenceStatus"))
        if evidence:
            lines.append("  证据状态:")
            for warning in evidence[:3]:
                lines.extend(_wrap_pdf_text(f"  - {warning}", 112))
        evidence_refs = _business_summary_evidence_ref_lines(item)
        if evidence_refs:
            lines.append("  证据引用:")
            for evidence_ref in evidence_refs[:8]:
                lines.extend(_wrap_pdf_text(f"  - {evidence_ref}", 112))
        recommended = str(item.get("recommendedUse") or "").strip()
        if recommended:
            lines.extend(_wrap_pdf_text(f"  建议用途: {recommended}", 112))
    lines.append("-" * 96)
    return lines


def _paginate_pdf_lines(lines: list[str]) -> list[list[str]]:
    max_lines = max(1, int((PDF_PAGE_HEIGHT - PDF_MARGIN_TOP - PDF_BOTTOM_Y) / PDF_LINE_HEIGHT))
    return [lines[index:index + max_lines] for index in range(0, len(lines), max_lines)] or [[]]


def _write_simple_pdf(pages: list[list[str]]) -> io.BytesIO:
    objects: list[bytes] = []
    page_object_ids: list[int] = []
    to_unicode = _pdf_to_unicode_cmap(_pdf_document_chars(pages))

    def add_object(body: bytes) -> int:
        objects.append(body)
        return len(objects)

    add_object(b"<< /Type /Catalog /Pages 2 0 R >>")
    add_object(b"<< /Type /Pages /Kids [] /Count 0 >>")
    font_object_id = add_object(
        b"<< /Type /Font /Subtype /Type0 /BaseFont /STSong-Light /Encoding /UniGB-UCS2-H "
        b"/DescendantFonts [4 0 R] /ToUnicode 5 0 R >>"
    )
    add_object(
        b"<< /Type /Font /Subtype /CIDFontType0 /BaseFont /STSong-Light "
        b"/CIDSystemInfo << /Registry (Adobe) /Ordering (GB1) /Supplement 2 >> /DW 1000 >>"
    )
    add_object(
        b"<< /Length " + str(len(to_unicode)).encode("ascii") + b" >>\nstream\n" + to_unicode + b"\nendstream"
    )

    for page_index, page_lines in enumerate(pages, start=1):
        content = _pdf_page_content(page_lines, page_index, len(pages))
        content_object_id = add_object(
            b"<< /Length " + str(len(content)).encode("ascii") + b" >>\nstream\n" + content + b"\nendstream"
        )
        page_object_id = add_object(
            (
                f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {PDF_PAGE_WIDTH} {PDF_PAGE_HEIGHT}] "
                f"/Resources << /Font << /F1 {font_object_id} 0 R >> >> /Contents {content_object_id} 0 R >>"
            ).encode("ascii")
        )
        page_object_ids.append(page_object_id)

    kids = " ".join(f"{page_id} 0 R" for page_id in page_object_ids)
    objects[1] = f"<< /Type /Pages /Kids [{kids}] /Count {len(page_object_ids)} >>".encode("ascii")

    output = io.BytesIO()
    output.write(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for object_id, body in enumerate(objects, start=1):
        offsets.append(output.tell())
        output.write(f"{object_id} 0 obj\n".encode("ascii"))
        output.write(body)
        output.write(b"\nendobj\n")
    xref_offset = output.tell()
    output.write(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.write(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.write(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.write(
        (
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref_offset}\n%%EOF\n"
        ).encode("ascii")
    )
    output.seek(0)
    return output


def _pdf_page_content(lines: list[str], page_index: int, page_count: int) -> bytes:
    commands = [
        "0.96 0.97 0.99 rg 0 0 842 595 re f",
        "0.09 0.13 0.20 rg 0 562 842 33 re f",
        f"1 1 1 rg BT /F1 12 Tf 36 574 Td {_pdf_text('Engineering Config Compare')} Tj ET",
    ]
    y = PDF_PAGE_HEIGHT - PDF_MARGIN_TOP - 20
    for index, line in enumerate(lines):
        font_size = 11 if index == 0 and page_index == 1 else 8
        commands.append(f"0.10 0.14 0.20 rg BT /F1 {font_size} Tf {PDF_MARGIN_X} {y} Td {_pdf_text(line)} Tj ET")
        y -= PDF_LINE_HEIGHT
    footer = f"Page {page_index} / {page_count}"
    commands.append(f"0.40 0.45 0.52 rg BT /F1 8 Tf 744 18 Td {_pdf_text(footer)} Tj ET")
    return "\n".join(commands).encode("ascii")


def _pdf_text(value: str) -> str:
    return "<" + _pdf_bmp_text(value).encode("utf-16-be", errors="replace").hex().upper() + ">"


def _pdf_bmp_text(value: str) -> str:
    return "".join(char if ord(char) <= 0xFFFF else "□" for char in value)


def _pdf_document_chars(pages: list[list[str]]) -> set[str]:
    chars: set[str] = set("Engineering Config Compare")
    for page_index, page_lines in enumerate(pages, start=1):
        chars.update(_pdf_bmp_text(f"Page {page_index} / {len(pages)}"))
        for line in page_lines:
            chars.update(_pdf_bmp_text(line))
    return chars


def _pdf_to_unicode_cmap(chars: set[str]) -> bytes:
    entries = []
    for char in sorted(chars, key=ord):
        codepoint = ord(char)
        if codepoint > 0xFFFF:
            continue
        hex_code = f"{codepoint:04X}"
        entries.append(f"<{hex_code}> <{hex_code}>")

    lines = [
        "/CIDInit /ProcSet findresource begin",
        "12 dict begin",
        "begincmap",
        "/CIDSystemInfo << /Registry (Adobe) /Ordering (UCS) /Supplement 0 >> def",
        "/CMapName /Adobe-Identity-UCS def",
        "/CMapType 2 def",
        "1 begincodespacerange",
        "<0000> <FFFF>",
        "endcodespacerange",
    ]
    for index in range(0, len(entries), 100):
        chunk = entries[index:index + 100]
        lines.append(f"{len(chunk)} beginbfchar")
        lines.extend(chunk)
        lines.append("endbfchar")
    lines.extend(
        [
            "endcmap",
            "CMapName currentdict /CMap defineresource pop",
            "end",
            "end",
        ]
    )
    return "\n".join(lines).encode("ascii")


def _wrap_pdf_text(value: str, max_chars: int) -> list[str]:
    text = " ".join(str(value).split())
    if not text:
        return [""]
    lines: list[str] = []
    current = ""
    for part in text.split(" "):
        if not current:
            current = part
        elif len(current) + 1 + len(part) <= max_chars:
            current = f"{current} {part}"
        else:
            lines.append(current)
            current = part
    if current:
        lines.append(current)
    return lines


def _ascii_safe(value: str) -> str:
    return value.replace("\n", " ").replace("\r", " ").strip()


def _write_header(ws, row_number: int, headers: list[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_number, column=column_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _set_widths(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _trim_label(trim: dict[str, Any]) -> str:
    return str(trim.get("fullTrimName") or trim.get("trimName") or trim.get("trimId") or "配置列")


def _trim_origin_label(trim: dict[str, Any]) -> str:
    data_origin = str(trim.get("dataOrigin") or "")
    if data_origin == "own_catalog" or trim.get("materialNo") or trim.get("hasMaterialNo"):
        return "本品"
    if data_origin == "external_or_scraped":
        return "竞品 / 外部"
    return "身份待确认"


def _trim_identity_label(trim: dict[str, Any]) -> str:
    material_no = _clean_trim_context_value(trim.get("materialNo") or trim.get("vehicleCode"))
    if material_no:
        return f"物料号 {material_no}"
    sales_version = _clean_trim_context_value(trim.get("salesVersion"))
    if sales_version:
        return f"Sales version {sales_version}"
    identity_key = _clean_trim_context_value(trim.get("identityKey"))
    if identity_key:
        return f"Identity {identity_key}"
    source = _clean_trim_context_value(trim.get("sourceFileName") or trim.get("sourceUploadId"))
    if source:
        return f"来源 {source}"
    return "车型 / 市场锚点"


def _trim_context_values(trim: dict[str, Any]) -> list[str]:
    values = [
        trim.get("market") or trim.get("country"),
        trim.get("modelYear"),
        trim.get("sourceFileName") or trim.get("sourceUploadId"),
        f"来源人 {trim.get('sourceCreatedBy')}" if trim.get("sourceCreatedBy") else None,
    ]
    return [cleaned for value in values if (cleaned := _clean_trim_context_value(value))]


def _trim_export_label(trim: dict[str, Any]) -> str:
    origin = _trim_origin_label(trim)
    identity = _trim_identity_label(trim)
    context = " · ".join(_trim_context_values(trim))
    parts = [
        _trim_label(trim),
        origin if origin != "身份待确认" else "",
        identity if identity != "车型 / 市场锚点" else "",
        context,
    ]
    unique_parts: list[str] = []
    seen: set[str] = set()
    for part in parts:
        cleaned = _clean_trim_context_value(part)
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        unique_parts.append(cleaned)
    return " · ".join(unique_parts)


def _clean_trim_context_value(value: Any) -> str:
    return " ".join(str(value or "").split())


def _cell_display(cell: dict[str, Any] | None) -> str:
    if not cell:
        return "待确认"
    display = str(cell.get("displayValue") or "").strip()
    if display:
        return display
    availability = str(cell.get("availability") or "").strip()
    if availability == "STANDARD":
        return "标配"
    if availability == "OPTIONAL":
        return "选装"
    if availability == "NOT_AVAILABLE":
        return "不配备*" if cell.get("inferred") else "不配备"
    if availability == "NOT_APPLICABLE":
        return "不适用"
    if availability == "UNKNOWN":
        return str(cell.get("rawValue") or "待确认")
    return str(cell.get("rawValue") or availability or "")


def _comparison_label(value: str) -> str:
    return {
        "COMMON_SAME": "共同配置",
        "DIFFERENT_VALUE": "值不同",
        "UNIQUE_TO_TRIM": "独有配置",
        "PARTIAL_AVAILABLE": "部分具备",
        "MISSING_OR_UNKNOWN": "缺失 / 未知",
        "MISSING_UNKNOWN": "待确认",
        "AVAILABILITY_DIFFERENT": "可用性差异",
        "OPTIONAL_DIFFERENT": "选装差异",
        "UNIQUE_OR_PARTIAL": "部分具备",
    }.get(value, value)


def _safe_filename(value: str, fallback: str) -> str:
    name = value.replace("\\", "/").rsplit("/", 1)[-1].strip()
    if not name:
        return fallback
    if not name.lower().endswith(".xlsx"):
        name = f"{name}.xlsx"
    return name


def _safe_token(value: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "-", value).strip("-")
    return cleaned or "config"


def _business_summary_items(payload: dict[str, Any]) -> list[dict[str, Any]]:
    return _list_of_dicts(payload.get("businessSummary"))[:8]


def _business_summary_usage(payload: dict[str, Any]) -> dict[str, Any]:
    return _dict(payload.get("businessSummaryUsage"))


def _evidence_summary(payload: dict[str, Any]) -> dict[str, Any]:
    return _dict(payload.get("evidenceSummary"))


def _evidence_summary_line(summary: dict[str, Any]) -> str:
    if not summary:
        return ""
    fields = [
        ("rowCount", "导出行"),
        ("trimCount", "配置列"),
        ("valueCount", "单元格"),
        ("inferredValueCount", "规则推断"),
        ("missingValueCount", "缺值"),
        ("missingSourceValueCount", "缺来源"),
        ("sourceIssueRowCount", "来源问题行"),
        ("mergedCellValueCount", "合并格"),
    ]
    parts = [
        f"{label} {summary.get(key)}"
        for key, label in fields
        if summary.get(key) not in (None, "")
    ]
    return "证据审计 · " + " · ".join(parts) if parts else ""


def _business_summary_usage_values(usage: dict[str, Any] | None) -> list[str]:
    usage = usage or {}
    return [
        str(usage.get("provider") or ""),
        str(usage.get("model") or ""),
        str(usage.get("status") or ""),
        str(usage.get("totalTokens") or ""),
        str(usage.get("finishReason") or ""),
        str(usage.get("transportFallback") or ""),
    ]


def _business_summary_usage_line(usage: dict[str, Any] | None) -> str:
    values = _business_summary_usage_values(usage)
    provider, model, status, total_tokens, finish_reason, transport_fallback = values
    if not any(values):
        return ""
    parts = [
        f"provider={provider}" if provider else None,
        f"model={model}" if model else None,
        f"status={status}" if status else None,
        f"tokens={total_tokens}" if total_tokens else None,
        f"finish={finish_reason}" if finish_reason else None,
        f"transportFallback={transport_fallback}" if transport_fallback else None,
    ]
    return "LLM usage: " + " / ".join(part for part in parts if part)


def _business_summary_evidence_ref_lines(item: dict[str, Any]) -> list[str]:
    lines: list[str] = []
    for ref in _list_of_dicts(item.get("evidenceRefs"))[:16]:
        section = str(ref.get("section") or "").strip()
        item_index = ref.get("itemIndex")
        evidence_key = str(ref.get("evidenceKey") or "").strip()
        feature_code = str(ref.get("featureCode") or "").strip()
        category = str(ref.get("category") or "").strip()
        reason = str(ref.get("reason") or "").strip()
        if not evidence_key:
            continue
        location = f"{section}[{item_index}]" if section and item_index is not None else section or "summary"
        feature_text = f" · {feature_code}" if feature_code else ""
        category_text = f" · {category}" if category else ""
        reason_text = f" · {reason}" if reason else ""
        lines.append(f"{location}: {evidence_key}{feature_text}{category_text}{reason_text}")
    return lines


def _string_list(value: Any) -> list[str]:
    return [str(item or "").strip() for item in _list(value) if str(item or "").strip()]


def _dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _dict_or_none(value: Any) -> dict[str, Any] | None:
    return value if isinstance(value, dict) else None


def _list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def _list_of_dicts(value: Any) -> list[dict[str, Any]]:
    return [item for item in _list(value) if isinstance(item, dict)]
