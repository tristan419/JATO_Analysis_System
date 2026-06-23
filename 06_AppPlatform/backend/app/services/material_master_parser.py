"""Parse Material Master XLSX files into structured SKU rows.

Expected format: multi-sheet workbook where each sheet is one model/powertrain.
Header row has: No., Code Name, Full Name, Configuration, BOM,
Exterior Color, Interior Color, then country FOB columns (e.g. "Croatia FOB").

Data uses carry-forward: model info appears on the first row of a group,
subsequent rows only fill colour + FOB values.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from app.services.ordering_normalization import (
    infer_colour_tier,
    normalize_brand,
    normalize_brand_text,
)


DUAL_COLOUR_PATTERNS = [
    r"/",
    r"&",
    r"dual",
    r"two.tone",
    r"two tone",
    r"black roof",
    r"bi.colou?r",
    r"双色",  # 双色
]


def _detect_colour_type(colour_name: str) -> str:
    lower = colour_name.lower().strip()
    for pattern in DUAL_COLOUR_PATTERNS:
        if re.search(pattern, lower):
            return "dual"
    return "single"


def _detect_edition_tag(raw_colour_name: str) -> str | None:
    m = re.search(r"[（(](black\s*edition)[)）]", raw_colour_name, re.IGNORECASE)
    return m.group(1).strip().title() if m else None


def _detect_colour_tier(
    colour_name: str, colour_type: str, edition_tag: str | None = None
) -> str:
    return infer_colour_tier(colour_name, colour_type, edition_tag)


# ── BOM-pattern → interior inference rules ───────────────────
# Key: regex matching the FULL BOM template (with ** as literal chars).
# The pattern includes the model-specific prefix before '**' so the same
# tail code can resolve to different interiors on different models.
# Value: (interior_name, interior_code)
_BOM_INTERIOR_RULES: list[tuple[str, str, str]] = [
    # JAECOO8 SHS — Luxury-AWD (5座) & Premium-AWD (7座)
    (r"T6481..\*\*LX0002$", "Black-Black", "BB"),
    (r"T6481..\*\*LX0008$", "Black-Black", "BB"),
    (r"T6481..\*\*LX0004$", "Black-Black", "BB"),
    (r"T6481..\*\*LX0009$", "Black-Black", "BB"),
    (r"T6481..\*\*LX0003$", "Black-Brown", "BR"),
    (r"T6481..\*\*LX0010$", "Black-Brown", "BR"),
    # JAECOO8 SHS different prefix (F3 instead of QN)
    (r"T6481F3\*\*LX0008$", "Black-Black", "BB"),
    # OMODA7 ICE — Luxury-AWD & Premium-AWD
    (r"T7161..\*\*MQ00001$", "Black-Black", "BB"),
    (r"T7161..\*\*MQ0001$", "Black-Black", "BB"),
    (r"T7161..\*\*MQ0002$", "Black-Black", "BB"),
    # OMODA7 SHS (PHEV) — Comfort/Luxury/Premium FWD
    (r"T7151..\*\*MH0001$", "Black", "BK"),
    (r"T7151..\*\*MH0002$", "Black", "BK"),
    (r"T7151..\*\*MH0003$", "Black", "BK"),
    (r"T7151N.\*\*MH0001$", "Black", "BK"),
    # JAECOO7 SHS — Premium-FWD
    (r"T71604.\*\*MH0031$", "Black-Black", "BB"),
    (r"T71604.\*\*MH0032$", "Black-Black", "BB"),
    (r"T71604.\*\*MH0035$", "Black-Brown", "BR"),
    # JAECOO7 SHS — Luxury-FWD
    (r"T7160RG\*\*MH0001$", "Black-Black", "BB"),
    # JAECOO7 HEV — Exclusive-FWD
    (r"T71607.\*\*MM0002$", "Black-Black", "BB"),
    (r"T71607.\*\*MM0007$", "Black-Black", "BB"),
    (r"T71607.\*\*MM0008$", "Black-Black", "BB"),
    # JAECOO5 BEV — Select-FWD
    (r"T7000Z5\*\*MY0013$", "Black-Black", "BB"),
    (r"T7000Z5\*\*MY0021$", "Black-Black", "BB"),
    (r"T7000Z5\*\*MY0022$", "Black-Black", "BB"),
    # JAECOO5 ICE — Exclusive-FWD & Premium-FWD
    (r"T71611C\*\*MM0013$", "Black-Black", "BB"),
    (r"T71611C\*\*MM0014$", "Black-Gray", "GY"),
    (r"T71611C\*\*MM0015$", "Black-Black", "BB"),
    # OMODA9 SHS
    (r"T6480J1\*\*LX0014$", "Black-Black", "BB"),
    (r"T6480J1\*\*LX0017$", "Black-Red", "BR"),
    (r"T6480J1\*\*LX0018$", "Black-Black", "BB"),
    # OMODA5 ICE
    (r"T5260CX\*\*MQ0002$", "Black-Black", "BB"),
    # JAECOO5 ICE (T516 variants)
    (r"T516385\*\*MM0004$", "Black-Black", "BB"),
    (r"T516385\*\*MM0005$", "Black-Black", "BB"),
    # OMODA7 ICE (T71506 variants)
    (r"T71506.\*\*MH0008$", "Black-Black", "BB"),
    (r"T71506.\*\*MH0009$", "Black-Black", "BB"),
    (r"T71506.\*\*MH0011$", "Black-Black", "BB"),
    # JAECOO7 HEV (T716015 variants)
    (r"T716015\*\*MH0009$", "Black-Black", "BB"),
    (r"T716015\*\*MH0010$", "Black-Black", "BB"),
    # OMODA7 ICE (T7000SW variants)
    (r"T7000SW\*\*MY0001$", "Black-Black", "BB"),
    (r"T7000SW\*\*MY0002$", "Black-Black", "BB"),
    # JAECOO8 SHS (T7160RG variants, different prefix from JAECOO7)
    (r"T7160RG\*\*MH0001$", "Black-Black", "BB"),
    # JAECOO5 ICE (T71611C additional)
    (r"T71611C\*\*MM0015$", "Black-Black", "BB"),
]


def _infer_interior_from_tail_code(bom_template: str | None) -> tuple[str, str] | None:
    """Try to infer interior (name, code) by matching full BOM pattern against known rules."""
    if not bom_template or "**" not in bom_template:
        return None
    for pattern, name, code in _BOM_INTERIOR_RULES:
        if re.search(pattern, bom_template):
            return name, code
    return None


def _extract_interior_colour_code(interior_name: str | None) -> str | None:
    if not interior_name or not interior_name.strip():
        return None
    words = re.split(r"[\s\-_/&]+", interior_name.strip())
    if len(words) >= 2:
        return (words[0][0] + words[-1][0]).upper()
    return words[0][:2].upper() if words else None


def _extract_colour_code(colour_name: str) -> str:
    """Extract colour code from a name like 'Phantom gray（GV）' or 'White (BW)'.

    Supports both Chinese （） and ASCII () brackets.
    For descriptions like 'Matte black（Black edition）', falls back to
    generating a code from the colour name initials.
    """
    # Chinese brackets
    m = re.search(r"（([^）]+)）", colour_name)
    if m:
        code = m.group(1).strip().upper()
        # If it looks like a real code (short, no spaces), use it
        if len(code) <= 4 and " " not in code:
            return code
        # Otherwise try ASCII brackets
    m = re.search(r"\(([^)]+)\)", colour_name)
    if m:
        code = m.group(1).strip().upper()
        if len(code) <= 4 and " " not in code:
            return code
    # Fallback: generate code from colour name initials
    # e.g. "Phantom gray" → "PG", "Carbon crystal black" → "CB"
    words = re.sub(r"[（(][^)）]+[)）]", "", colour_name).strip().split()
    if len(words) >= 2:
        return (words[0][0] + words[-1][0]).upper()
    elif words:
        return words[0][:2].upper()
    return ""


def _clean_colour_name(colour_name: str) -> str:
    """Remove colour code suffix from colour name."""
    name = re.sub(r"（[A-Za-z0-9]+）", "", colour_name)
    name = re.sub(r"\([A-Za-z0-9]+\)", "", name)
    return name.strip()


def _extract_bom_template(bom_raw: str) -> str | None:
    """Extract the first BOM line containing '**' from potentially multi-line text.
    Strips trailing Chinese/comments. Returns just the BOM pattern.
    """
    if not bom_raw:
        return None
    lines = bom_raw.strip().split("\n")
    for line in lines:
        line = line.strip()
        if "**" in line:
            m = re.match(r"([A-Za-z0-9]+\*\*[A-Za-z0-9]+)", line)
            if m:
                return m.group(1)
            return line
    for line in lines:
        line = line.strip()
        if line:
            return line
    return None


def _extract_bom_remark(bom_raw: str) -> str | None:
    """Extract remark from BOM cell parentheses. e.g. 'T7000SW**MY0001（法规升级）' → '法规升级'"""
    if not bom_raw:
        return None
    lines = bom_raw.strip().split("\n")
    for line in lines:
        if "**" in line:
            m = re.search(r"[（(]([^)）]+)[)）]", line)
            if m:
                return m.group(1).strip()
    return None


# Country alias mapping for normalisation
COUNTRY_ALIAS_MAP: dict[str, str] = {
    # Typo/alias → correct name
    "crotia": "Croatia",
    "hu": "Hungary",
    "波的": "Poland",
    "latvia": "Latvia",
    "austria": "Austria",
    "sweden": "Sweden",
    "greece": "Greece",
    "hungary": "Hungary",
    "croatia": "Croatia",
    # ISO codes → country names (for fallback mapping)
    "hr": "Croatia",
    "cz": "Czech Republic",
    "se": "Sweden",
    "ro": "Romania",
    "fi": "Finland",
    "gr": "Greece",
    "at": "Austria",
    "hu_iso": "Hungary",
    "lv": "Latvia",
    "pl": "Poland",
}


def _normalise_country_name(name: str) -> str:
    """Map country column names to standard names."""
    return COUNTRY_ALIAS_MAP.get(name.lower().strip(), name.strip())


def parse_material_master_xlsx(file_path: Path) -> dict:
    """Parse an OMODA & JAECOO Material Master XLSX file.

    Returns::
        {
            "rows": [ {...}, ... ],
            "warnings": [...],
            "sheet_names": [...],
        }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_rows: list[dict] = []
    warnings: list[str] = []
    interior_by_bom_template: dict[str, str] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 3:
            warnings.append(f"Sheet '{sheet_name}' has too few rows — skipping")
            continue

        # Read all rows
        raw_rows: list[list] = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            raw_rows.append(list(row))

        # Find header row — it has "No." or "BOM" or "Code Name"
        header_idx = -1
        for i, row in enumerate(raw_rows):
            first_cell = str(row[0]).strip() if row[0] is not None else ""
            rest = [str(c).strip().lower() if c is not None else "" for c in row[1:]]
            rest_text = " ".join(rest)
            if first_cell == "No." or "code name" in rest_text or "bom" in rest_text:
                header_idx = i
                break
        if header_idx < 0:
            warnings.append(f"Sheet '{sheet_name}' — could not find header row")
            continue

        headers: list[str] = [str(c).strip() if c else "" for c in raw_rows[header_idx]]

        # Identify column indices
        def col_idx(*names: str) -> int:
            for name in names:
                name_lower = name.lower()
                for j, h in enumerate(headers):
                    if h.lower() == name_lower or name_lower in h.lower():
                        return j
            return -1

        idx_no = col_idx("No.")
        idx_code_name = col_idx("Code Name")
        idx_full_name = col_idx("Full Name")
        idx_config = col_idx("Configuration")
        idx_bom = col_idx("BOM")
        idx_ext_colour = col_idx("Exterior Color")
        idx_int_colour = col_idx("Interior Color")

        # Country FOB columns: any header ending with "FOB" or containing " FOB"
        country_fob_cols: list[tuple[str, int]] = []
        for j, h in enumerate(headers):
            hl = h.lower().strip()
            if hl.endswith("fob") or " fob" in hl:
                country_name = re.sub(r"\s*fob\s*$", "", h, flags=re.IGNORECASE).strip()
                country_fob_cols.append((country_name, j))

        if idx_ext_colour < 0:
            warnings.append(
                f"Sheet '{sheet_name}' — no Exterior Color column found"
            )
            continue

        # ── Pass 1: collect interior → BOM-template mappings ──────
        # Must run BEFORE the main pass because interior rows may appear
        # after colour rows in the Excel — a forward-reference problem.
        _pass1_model: dict[str, str] = {"code_name": "", "full_name": "", "configuration": "", "bom": ""}
        for i in range(header_idx + 1, len(raw_rows)):
            row = raw_rows[i]
            if idx_int_colour < 0 or len(row) <= max(idx_int_colour, idx_bom if idx_bom >= 0 else 0):
                continue
            int_val = row[idx_int_colour]
            if int_val is None or not str(int_val).strip():
                continue
            # Carry-forward model fields
            if idx_code_name >= 0 and idx_code_name < len(row) and row[idx_code_name]:
                _pass1_model["code_name"] = str(row[idx_code_name]).strip()
            if idx_full_name >= 0 and idx_full_name < len(row) and row[idx_full_name]:
                _pass1_model["full_name"] = str(row[idx_full_name]).strip()
            if idx_config >= 0 and idx_config < len(row) and row[idx_config]:
                _pass1_model["configuration"] = str(row[idx_config]).strip()
            if idx_bom >= 0 and idx_bom < len(row) and row[idx_bom]:
                _pass1_model["bom"] = str(row[idx_bom]).strip()
            bom_raw = _pass1_model.get("bom", "")
            bom_tmpl = _extract_bom_template(bom_raw) if bom_raw else None
            raw_int = str(int_val).strip()
            if bom_tmpl and "**" in bom_tmpl and raw_int:
                existing = interior_by_bom_template.get(bom_tmpl)
                if existing and existing != raw_int:
                    warnings.append(
                        f"Sheet '{sheet_name}' row {i + 1}: BOM template {bom_tmpl} "
                        f"has conflicting interiors '{existing}' and '{raw_int}'"
                    )
                interior_by_bom_template[bom_tmpl] = raw_int

        # ── Pass 2: parse data rows with carry-forward for model fields
        current_model: dict[str, str] = {
            "code_name": "",
            "full_name": "",
            "configuration": "",
            "bom": "",
            "interior_color": "",
            "interior_bom_template": "",
        }

        for i in range(header_idx + 1, len(raw_rows)):
            row = raw_rows[i]
            if len(row) <= idx_ext_colour:
                continue

            # Skip empty/separator rows
            non_empty = sum(
                1 for c in row if c is not None and str(c).strip()
            )
            if non_empty == 0:
                continue

            no_val = row[idx_no] if idx_no >= 0 and idx_no < len(row) else None
            code_val = (
                row[idx_code_name]
                if idx_code_name >= 0 and idx_code_name < len(row)
                else None
            )
            full_val = (
                row[idx_full_name]
                if idx_full_name >= 0 and idx_full_name < len(row)
                else None
            )
            config_val = (
                row[idx_config]
                if idx_config >= 0 and idx_config < len(row)
                else None
            )
            bom_val = (
                row[idx_bom]
                if idx_bom >= 0 and idx_bom < len(row)
                else None
            )

            no_str = str(no_val).strip() if no_val is not None else ""
            code_str = str(code_val).strip() if code_val is not None else ""
            full_str = str(full_val).strip() if full_val is not None else ""
            config_str = str(config_val).strip() if config_val is not None else ""
            bom_str = str(bom_val).strip() if bom_val is not None else ""

            # Update current model — only overwrite fields that have values
            if code_str:
                current_model["code_name"] = code_str
            if full_str:
                current_model["full_name"] = full_str
            if config_str:
                current_model["configuration"] = config_str
            if bom_str:
                current_model["bom"] = bom_str

            # BOM template is the stable binding key for colour variants.
            bom_raw = current_model.get("bom", "")
            bom_template = _extract_bom_template(bom_raw) if bom_raw else None
            bom_remark = _extract_bom_remark(bom_raw) if bom_raw else None

            raw_interior = (
                str(row[idx_int_colour]).strip()
                if idx_int_colour >= 0 and row[idx_int_colour] is not None
                else ""
            )
            if raw_interior:
                current_model["interior_color"] = raw_interior
                if bom_template and "**" in bom_template:
                    existing = interior_by_bom_template.get(bom_template)
                    if existing and existing != raw_interior:
                        warnings.append(
                            f"Sheet '{sheet_name}' row {i + 1}: BOM template {bom_template} "
                            f"has conflicting interiors '{existing}' and '{raw_interior}'"
                        )
                    interior_by_bom_template[bom_template] = raw_interior

            # Extract colour
            raw_colour = (
                str(row[idx_ext_colour]).strip()
                if row[idx_ext_colour] is not None
                else ""
            )
            if not raw_colour:
                continue  # rows without colour are skipped

            colour_code = _extract_colour_code(raw_colour)
            # Auto-generated codes (no brackets in original Excel) need user confirmation
            colour_code_confirmed = bool(re.search(r"[（(][A-Za-z0-9]{1,4}[)）]", raw_colour))
            colour_name = _clean_colour_name(raw_colour)
            colour_type = _detect_colour_type(raw_colour)
            edition_tag = _detect_edition_tag(raw_colour)
            colour_tier = _detect_colour_tier(colour_name, colour_type, edition_tag)

            # Interior colour resolution (multi-layer fallback):
            # 1. Direct value from Excel row (already read above)
            # 2. Pass-1 map: BOM template → interior (most reliable for merged cells)
            # 3. Carry-forward: last seen interior in current_model
            # 4. Tail-code inference from known suffix patterns
            if not raw_interior:
                if bom_template and "**" in bom_template:
                    raw_interior = interior_by_bom_template.get(bom_template, "")
                if not raw_interior:
                    raw_interior = current_model.get("interior_color", "")
                if not raw_interior and bom_template:
                    inferred = _infer_interior_from_tail_code(bom_template)
                    if inferred:
                        raw_interior, _inferred_code = inferred
            interior_colour_code = _extract_interior_colour_code(raw_interior) if raw_interior else None
            interior_package = raw_interior or None

            # Generate material code
            material_code = ""
            row_warnings: list[str] = []
            if bom_template and colour_code and "**" in bom_template:
                material_code = bom_template.replace("**", colour_code)
            elif bom_template and "**" in bom_template:
                row_warnings.append(f"No colour code for '{raw_colour}', BOM ** not replaced")
                material_code = bom_template
            elif bom_template:
                material_code = bom_template
                if "**" not in bom_template:
                    row_warnings.append("BOM has no '**' placeholder")

            if not material_code:
                row_warnings.append("Empty material_code")

            # Full name gives model/version/powertrain
            full_name = current_model.get("full_name", "")
            config = current_model.get("configuration", "")

            # Extract base FOB values per country
            base_fob_eur: float | None = None
            country_fobs: dict[str, float] = {}
            for country_name, cidx in country_fob_cols:
                val = row[cidx] if cidx < len(row) else None
                if val is not None:
                    try:
                        country_fobs[country_name] = float(
                            str(val).replace(",", "").replace("€", "").strip()
                        )
                    except (ValueError, TypeError):
                        pass

            # Use first available country FOB as base
            if country_fobs:
                base_fob_eur = list(country_fobs.values())[0]

            all_rows.append({
                "row_index": i + 1,
                "sheet_name": sheet_name,
                "brand": "",  # detected from model name
                "model_name": full_name or current_model.get("code_name", ""),
                "version": config or "",
                "exterior_color_name": colour_name,
                "exterior_color_code": colour_code,
                "exterior_color_type": colour_type,
                "colour_code_confirmed": colour_code_confirmed,
                "colour_tier": colour_tier,
                "edition_tag": edition_tag,
                "interior_color_name": raw_interior or None,
                "interior_colour_code": interior_colour_code,
                "interior_package": interior_package,
                "bom_template": bom_template,
                "remark": bom_remark or "",
                "material_code": material_code,
                "base_fob_eur": base_fob_eur,
                "powertrain": None,  # from sheet name or code_name
                "country_fobs": country_fobs,
                "warnings": row_warnings,
            })

    wb.close()

    # Post-processing: detect brand from model name (with typo tolerance)
    for row in all_rows:
        row["model_name"] = normalize_brand_text(row.get("model_name", ""))
        brand = normalize_brand(
            f"{row.get('model_name', '')} {row.get('sheet_name', '')}"
        )
        if brand in {"JAECOO", "OMODA"}:
            row["brand"] = brand
        else:
            row["brand"] = ""

        # Extract powertrain from model name or sheet
        from app.services.powertrain_normalizer import normalize_powertrain

        sn = row["sheet_name"].upper()
        mn = row.get("model_name", "").upper()
        raw_pt = None
        # Longest-first: PHEV/SHS/MHEV before HEV/EV to avoid substring collision
        for pt in ("SHS", "PHEV", "MHEV", "HEV", "BEV", "EV", "ICE"):
            if pt in sn or pt in mn:
                raw_pt = pt
                break
        if raw_pt:
            row["powertrain"] = normalize_powertrain(raw_pt)

    total_warnings = warnings + [
        f"Row {r['row_index']} ({r['sheet_name']}): {w}"
        for r in all_rows
        for w in r["warnings"]
    ]

    return {
        "rows": all_rows,
        "warnings": total_warnings,
        "sheet_names": wb.sheetnames,
    }
