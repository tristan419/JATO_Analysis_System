"""COC fill service - Excel material groups vs WVTA/COC PDF table."""

from __future__ import annotations

import re
import shutil
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from upload_toolkit.file_utils import allowed_extension
from upload_toolkit.job_engine import (
    BaseJobRunner,
    list_job_payloads,
    load_job_state,
    persist_job_state,
    state_path,
)

from app.core.config import COC_MATCH_JOB_ROOT
from app.services.coc_match_service import _normalize_filename
from app.services.workbook_table_scanner import (
    MaterialGroupRow,
    cell_text,
    display_names,
    extract_material_rows,
    normalize_material_group,
    parse_date,
    target_columns_for_sheet,
)


ALLOWED_FILL_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
ALLOWED_FILL_PDF_EXTENSIONS = {".pdf"}
_COC_FILL_UPLOAD_CHUNK_SIZE = 8 * 1024 * 1024
_FILL_UPLOAD_SESSION_ROOT = COC_MATCH_JOB_ROOT / "_fill_upload_sessions"
_RUNNING_FILL_THREADS: dict[str, "CocFillJobRunner"] = {}
_MAX_FILL_HISTORY = 50

WVTA_RE = re.compile(r"e\d+\*2018/858\*[^\s]+", re.IGNORECASE)
COC_RE = re.compile(r"\d{5}-\d{2}&[^\s]*C[O0]C[^\s]*", re.IGNORECASE)


@dataclass(frozen=True)
class WvtaCocRecord:
    material_group: str
    wvta_no: str
    coc_no: str
    brand: str | None
    model: str | None
    powertrain: str | None
    version: str | None
    sales_name: str | None
    valid_from: date | None
    valid_to: date | None
    comments: str | None
    page_number: int
    table_row_number: int


@dataclass(frozen=True)
class FillDecision:
    material_group: str
    sheet_name: str
    row_number: int
    status: str
    candidate_count: int
    reason: str
    confidence: float
    selected_record: WvtaCocRecord | None
    candidate_records: tuple[WvtaCocRecord, ...]
    written_wvta: str | None
    written_coc: str | None


def _candidate_text(cells: list[str], start: int, width: int = 5) -> str:
    return " ".join(cell for cell in cells[start : start + width] if cell)


def _record_from_pdf_row(
    cells: list[str],
    *,
    material_index: int,
    material_group: str,
    page_number: int,
    row_number: int,
) -> WvtaCocRecord | None:
    tail = _candidate_text(cells, material_index + 1)
    wvta_source = cells[material_index + 1] if material_index + 1 < len(cells) else ""
    coc_source = cells[material_index + 2] if material_index + 2 < len(cells) else ""
    wvta_match = WVTA_RE.search(wvta_source) or WVTA_RE.search(tail)
    coc_match = COC_RE.search(coc_source) or COC_RE.search(tail)
    if wvta_match is None or coc_match is None:
        return None

    def at(offset: int) -> str | None:
        index = material_index + offset
        if index >= len(cells):
            return None
        return cells[index] or None

    return WvtaCocRecord(
        material_group=material_group,
        wvta_no=wvta_match.group(0),
        coc_no=coc_match.group(0),
        brand=at(-4),
        model=at(-3),
        powertrain=at(-2),
        version=at(-1),
        sales_name=at(7),
        valid_from=parse_date(at(8), day_first=True),
        valid_to=parse_date(at(9), day_first=True),
        comments=at(10),
        page_number=page_number,
        table_row_number=row_number,
    )


def extract_wvta_coc_records(pdf_path: Path) -> list[WvtaCocRecord]:
    try:
        import pdfplumber
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="服务器缺少 pdfplumber，无法解析 WVTA PDF。",
        ) from exc

    records: list[WvtaCocRecord] = []
    try:
        with pdfplumber.open(pdf_path) as document:
            for page_index, page in enumerate(document.pages, start=1):
                tables = page.extract_tables()
                for table in tables:
                    for row_index, row in enumerate(table, start=1):
                        cells = [str(cell or "").replace("\n", " ").strip() for cell in row]
                        for cell_index, cell in enumerate(cells):
                            material = normalize_material_group(cell)
                            if not material:
                                continue
                            record = _record_from_pdf_row(
                                cells,
                                material_index=cell_index,
                                material_group=material,
                                page_number=page_index,
                                row_number=row_index,
                            )
                            if record is not None:
                                records.append(record)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"WVTA PDF 表格解析失败: {exc}") from exc
    if not records:
        raise HTTPException(status_code=422, detail="WVTA PDF 未识别到可抽取的物料号/WVTA/COC 表格。")
    return records


_COUNTRY_ALIASES = {
    "PL": {"PL", "POLAND", "波兰"},
    "IT": {"IT", "ITALY", "意大利"},
    "FR": {"FR", "FRANCE", "法国"},
    "DE": {"DE", "GERMANY", "德国"},
    "GR": {"GR", "GREECE", "希腊"},
    "GQ": {"GR", "GREECE", "希腊"},
}


def _country_tokens(country: str | None) -> set[str]:
    if not country:
        return set()
    cleaned = country.strip().upper()
    result = {cleaned}
    result.update(_COUNTRY_ALIASES.get(cleaned, set()))
    return result


def _comment_special_country(comments: str | None) -> str | None:
    text = comments or ""
    special_markers = {
        "PL": ("波兰专用", "波兰专属", "POLAND"),
        "IT": ("意大利专用", "意大利专属", "ITALY"),
        "FR": ("法国专用", "法国专属", "FRANCE"),
        "DE": ("德国专用", "德国专属", "GERMANY"),
    }
    upper_text = text.upper()
    for country_code, markers in special_markers.items():
        if any(marker.upper() in upper_text for marker in markers):
            return country_code
    return None


def _record_matches_country(row: MaterialGroupRow, record: WvtaCocRecord) -> bool:
    special_country = _comment_special_country(record.comments)
    if special_country is None:
        return True
    tokens = _country_tokens(row.country)
    if not tokens:
        return False
    return bool(tokens & _COUNTRY_ALIASES.get(special_country, {special_country}))


def _date_ranges_overlap(
    row_start: date | None,
    row_end: date | None,
    record_start: date | None,
    record_end: date | None,
) -> bool:
    if row_start is None and row_end is None:
        return True
    effective_row_start = row_start or row_end
    effective_row_end = row_end or row_start
    if effective_row_start is None or effective_row_end is None:
        return True
    if record_start is not None and effective_row_end < record_start:
        return False
    if record_end is not None and effective_row_start > record_end:
        return False
    return True


def _unique_records(records: list[WvtaCocRecord]) -> list[WvtaCocRecord]:
    unique: list[WvtaCocRecord] = []
    seen: set[tuple[str, str, str | None]] = set()
    for record in records:
        key = (record.wvta_no, record.coc_no, record.comments)
        if key in seen:
            continue
        seen.add(key)
        unique.append(record)
    return unique


def resolve_fill_decision(
    row: MaterialGroupRow,
    records_by_material: dict[str, list[WvtaCocRecord]],
    *,
    overwrite_existing: bool,
    conflict_strategy: str = "strict",
) -> FillDecision:
    candidates = records_by_material.get(row.material_group, [])
    if not candidates:
        return FillDecision(
            material_group=row.material_group,
            sheet_name=row.sheet_name,
            row_number=row.row_number,
            status="not_found",
            candidate_count=0,
            reason="PDF 已解析，但未找到精确物料号组；可能 PDF 中不存在该物料号，或物料号写法不一致。",
            confidence=0,
            selected_record=None,
            candidate_records=(),
            written_wvta=None,
            written_coc=None,
        )
    unique_candidates = tuple(_unique_records(candidates))
    if (row.existing_wvta or row.existing_coc) and not overwrite_existing:
        return FillDecision(
            material_group=row.material_group,
            sheet_name=row.sheet_name,
            row_number=row.row_number,
            status="skipped_existing",
            candidate_count=len(candidates),
            reason="目标单元格已有值，当前策略为只填空白。",
            confidence=1,
            selected_record=None,
            candidate_records=unique_candidates,
            written_wvta=None,
            written_coc=None,
        )

    selected_candidates = list(candidates)
    if conflict_strategy == "date_country":
        country_filtered = [record for record in selected_candidates if _record_matches_country(row, record)]
        if country_filtered:
            selected_candidates = country_filtered
        date_filtered = [
            record
            for record in selected_candidates
            if _date_ranges_overlap(
                row.production_date_start,
                row.production_date_end,
                record.valid_from,
                record.valid_to,
            )
        ]
        if date_filtered:
            selected_candidates = date_filtered

    unique_pairs = {(record.wvta_no, record.coc_no) for record in selected_candidates}
    if len(unique_pairs) == 1:
        record = selected_candidates[0]
        confidence = 1.0 if len(candidates) == 1 else 0.86
        return FillDecision(
            material_group=row.material_group,
            sheet_name=row.sheet_name,
            row_number=row.row_number,
            status="filled",
            candidate_count=len(candidates),
            reason="精确物料号匹配" if len(candidates) == 1 else "多候选经日期/国家规则收敛。",
            confidence=confidence,
            selected_record=record,
            candidate_records=tuple(_unique_records(selected_candidates)),
            written_wvta=record.wvta_no,
            written_coc=record.coc_no,
        )

    return FillDecision(
        material_group=row.material_group,
        sheet_name=row.sheet_name,
        row_number=row.row_number,
        status="ambiguous",
        candidate_count=len(candidates),
        reason="PDF 存在多个 WVTA/COC 候选，无法唯一确定。",
        confidence=0.3,
        selected_record=None,
        candidate_records=tuple(_unique_records(selected_candidates)),
        written_wvta=None,
        written_coc=None,
    )


def _record_payload(record: WvtaCocRecord | None) -> dict[str, Any] | None:
    if record is None:
        return None
    return {
        "materialGroup": record.material_group,
        "wvtaNo": record.wvta_no,
        "cocNo": record.coc_no,
        "brand": record.brand,
        "model": record.model,
        "powertrain": record.powertrain,
        "version": record.version,
        "salesName": record.sales_name,
        "validFrom": record.valid_from.isoformat() if record.valid_from else None,
        "validTo": record.valid_to.isoformat() if record.valid_to else None,
        "comments": record.comments,
        "pageNumber": record.page_number,
        "tableRowNumber": record.table_row_number,
    }


def _decision_payload(decision: FillDecision) -> dict[str, Any]:
    return {
        "materialGroup": decision.material_group,
        "sheetName": decision.sheet_name,
        "rowNumber": decision.row_number,
        "status": decision.status,
        "candidateCount": decision.candidate_count,
        "reason": decision.reason,
        "confidence": decision.confidence,
        "selectedRecord": _record_payload(decision.selected_record),
        "candidateRecords": [
            record_payload
            for record_payload in (_record_payload(record) for record in decision.candidate_records)
            if record_payload is not None
        ],
        "writtenWvta": decision.written_wvta,
        "writtenCoc": decision.written_coc,
    }


def _preview_groups_payload(
    decisions: list[FillDecision],
    *,
    per_sheet_limit: int = 1000,
) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    for decision in decisions:
        group = groups.setdefault(
            decision.sheet_name,
            {
                "sheetName": decision.sheet_name,
                "totalRows": 0,
                "filledCount": 0,
                "notFoundCount": 0,
                "ambiguousCount": 0,
                "skippedExistingCount": 0,
                "invalidSourceCount": 0,
                "statusCounts": {},
                "decisions": [],
                "previewLimit": per_sheet_limit,
                "truncated": False,
            },
        )
        group["totalRows"] += 1
        status_counts = group["statusCounts"]
        status_counts[decision.status] = status_counts.get(decision.status, 0) + 1
        if decision.status == "filled":
            group["filledCount"] += 1
        elif decision.status == "not_found":
            group["notFoundCount"] += 1
        elif decision.status == "ambiguous":
            group["ambiguousCount"] += 1
        elif decision.status == "skipped_existing":
            group["skippedExistingCount"] += 1
        elif decision.status == "invalid_source":
            group["invalidSourceCount"] += 1

        preview_decisions = group["decisions"]
        if len(preview_decisions) < per_sheet_limit:
            preview_decisions.append(_decision_payload(decision))
        else:
            group["truncated"] = True
    return list(groups.values())


def _write_result_sheet(workbook: Any, decisions: list[FillDecision]) -> None:
    sheet_name = "COC填充结果"
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    sheet = workbook.create_sheet(sheet_name)
    headers = [
        "状态",
        "Sheet",
        "行号",
        "物料号组",
        "WVTA编号",
        "COC编号",
        "候选数",
        "原因",
        "PDF页码",
        "PDF表格行",
        "备注",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for decision in decisions:
        record = decision.selected_record
        sheet.append(
            [
                decision.status,
                decision.sheet_name,
                decision.row_number,
                decision.material_group,
                decision.written_wvta or "",
                decision.written_coc or "",
                decision.candidate_count,
                decision.reason,
                record.page_number if record else "",
                record.table_row_number if record else "",
                record.comments if record else "",
            ]
        )
    for column_index, width in enumerate([16, 18, 10, 22, 28, 32, 10, 42, 10, 12, 48], start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def fill_coc_workbook(
    *,
    workbook_path: Path,
    pdf_path: Path,
    output_path: Path,
    overwrite_existing: bool = False,
    conflict_strategy: str = "strict",
    include_result_sheet: bool = False,
    sheet_names: list[str] | None = None,
) -> dict[str, Any]:
    workbook, targets_by_sheet, material_rows = extract_material_rows(
        workbook_path,
        sheet_names=sheet_names,
    )
    if not material_rows:
        if sheet_names:
            requested = [name.strip() for name in sheet_names if name.strip()]
            raise HTTPException(
                status_code=422,
                detail=(
                    f"指定 Sheet 未识别到物料号组列或包含 ** 的物料号组：{display_names(requested)}。"
                    "请留空处理全部 Sheet，或改填包含物料号组数据的 Sheet。"
                ),
            )
        raise HTTPException(status_code=422, detail="Excel 未识别到物料号组列或包含 ** 的物料号组。")

    pdf_records = extract_wvta_coc_records(pdf_path)
    records_by_material: dict[str, list[WvtaCocRecord]] = {}
    for record in pdf_records:
        records_by_material.setdefault(record.material_group, []).append(record)

    decisions: list[FillDecision] = []
    for row in material_rows:
        decision = resolve_fill_decision(
            row,
            records_by_material,
            overwrite_existing=overwrite_existing,
            conflict_strategy=conflict_strategy,
        )
        decisions.append(decision)
        if decision.status != "filled" or decision.selected_record is None:
            continue
        sheet = workbook[row.sheet_name]
        targets = targets_by_sheet[row.sheet_name]
        sheet.cell(row.row_number, targets.wvta_col).value = decision.selected_record.wvta_no
        sheet.cell(row.row_number, targets.coc_col).value = decision.selected_record.coc_no

    if include_result_sheet:
        _write_result_sheet(workbook, decisions)
    elif "COC填充结果" in workbook.sheetnames:
        workbook.remove(workbook["COC填充结果"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()

    status_counts: dict[str, int] = {}
    for decision in decisions:
        status_counts[decision.status] = status_counts.get(decision.status, 0) + 1

    unique_materials = {row.material_group for row in material_rows}
    return {
        "totalRows": len(material_rows),
        "uniqueMaterialCount": len(unique_materials),
        "pdfRecordCount": len(pdf_records),
        "filledCount": status_counts.get("filled", 0),
        "notFoundCount": status_counts.get("not_found", 0),
        "ambiguousCount": status_counts.get("ambiguous", 0),
        "skippedExistingCount": status_counts.get("skipped_existing", 0),
        "invalidSourceCount": status_counts.get("invalid_source", 0),
        "sheetCount": len(targets_by_sheet),
        "sheetNames": list(targets_by_sheet.keys()),
        "statusCounts": status_counts,
        "previewGroups": _preview_groups_payload(decisions),
        "decisions": [_decision_payload(decision) for decision in decisions[:300]],
    }


class CocFillJobRunner(BaseJobRunner):
    """Background job runner for filling WVTA/COC values into workbooks."""

    def __init__(
        self,
        job_id: str,
        state_dir: Path,
        excel_path: Path,
        pdf_path: Path,
        overwrite_existing: bool,
        conflict_strategy: str,
        include_result_sheet: bool,
        sheet_names: list[str] | None,
        triggered_by: str,
    ) -> None:
        super().__init__(job_id, state_dir)
        self.excel_path = excel_path
        self.pdf_path = pdf_path
        self.overwrite_existing = overwrite_existing
        self.conflict_strategy = conflict_strategy
        self.include_result_sheet = include_result_sheet
        self.sheet_names = sheet_names
        self.triggered_by = triggered_by

    def run(self) -> None:
        state = self.load_state()
        state["status"] = "running"
        state["phase"] = "extracting"
        state["startedAt"] = datetime.now().isoformat()
        self.persist_state(state)

        self.log(f"Excel: {self.excel_path}")
        self.log(f"PDF: {self.pdf_path}")
        self.log(f"Overwrite: {self.overwrite_existing}, Strategy: {self.conflict_strategy}")
        self.log(f"Include result sheet: {self.include_result_sheet}")
        if self.sheet_names:
            self.log(f"Sheet scope: {', '.join(self.sheet_names)}")

        output_path = self.state_dir / self.job_id / "filled.xlsx"
        summary = fill_coc_workbook(
            workbook_path=self.excel_path,
            pdf_path=self.pdf_path,
            output_path=output_path,
            overwrite_existing=self.overwrite_existing,
            conflict_strategy=self.conflict_strategy,
            include_result_sheet=self.include_result_sheet,
            sheet_names=self.sheet_names,
        )
        self.log(
            "Filled rows: "
            f"{summary['filledCount']}, missing: {summary['notFoundCount']}, "
            f"ambiguous: {summary['ambiguousCount']}, skipped: {summary['skippedExistingCount']}"
        )

        state = self.load_state()
        state.update(summary)
        state["status"] = "success"
        state["phase"] = "completed"
        state["finishedAt"] = datetime.now().isoformat()
        state["outputFilename"] = "filled.xlsx"
        self.persist_state(state)


def _validate_excel_filename(filename: str) -> None:
    if not allowed_extension(filename, allowed=ALLOWED_FILL_EXCEL_EXTENSIONS):
        raise HTTPException(status_code=400, detail="发运清单仅支持 .xlsx/.xlsm。")


def _validate_pdf_filename(filename: str) -> None:
    if not allowed_extension(filename, allowed=ALLOWED_FILL_PDF_EXTENSIONS):
        raise HTTPException(status_code=400, detail="WVTA 关联文件仅支持 .pdf。")


def _base_fill_state(
    *,
    job_id: str,
    excel_name: str,
    pdf_name: str,
    overwrite_existing: bool,
    conflict_strategy: str,
    include_result_sheet: bool,
    sheet_names: list[str] | None,
    triggered_by: str,
) -> dict[str, Any]:
    return {
        "jobId": job_id,
        "jobType": "fill",
        "status": "queued",
        "phase": "pending",
        "excelFilename": excel_name,
        "pdfFilename": pdf_name,
        "overwriteExisting": overwrite_existing,
        "conflictStrategy": conflict_strategy,
        "includeResultSheet": include_result_sheet,
        "sheetNames": sheet_names or [],
        "triggeredBy": triggered_by,
        "totalRows": None,
        "uniqueMaterialCount": None,
        "pdfRecordCount": None,
        "filledCount": None,
        "notFoundCount": None,
        "ambiguousCount": None,
        "skippedExistingCount": None,
        "invalidSourceCount": None,
        "sheetCount": None,
        "statusCounts": {},
        "decisions": [],
        "outputFilename": None,
        "error": None,
        "createdAt": datetime.now().isoformat(),
        "startedAt": None,
        "finishedAt": None,
    }


def _start_fill_runner(
    *,
    job_id: str,
    excel_path: Path,
    pdf_path: Path,
    overwrite_existing: bool,
    conflict_strategy: str,
    include_result_sheet: bool,
    sheet_names: list[str] | None,
    triggered_by: str,
) -> dict[str, Any]:
    runner = CocFillJobRunner(
        job_id=job_id,
        state_dir=COC_MATCH_JOB_ROOT,
        excel_path=excel_path,
        pdf_path=pdf_path,
        overwrite_existing=overwrite_existing,
        conflict_strategy=conflict_strategy,
        include_result_sheet=include_result_sheet,
        sheet_names=sheet_names,
        triggered_by=triggered_by,
    )
    _RUNNING_FILL_THREADS[job_id] = runner
    runner.start()
    return load_job_state(state_path(COC_MATCH_JOB_ROOT, job_id))


def create_coc_fill_job(
    *,
    excel_file: UploadFile,
    pdf_file: UploadFile,
    overwrite_existing: bool,
    conflict_strategy: str,
    triggered_by: str,
    include_result_sheet: bool = False,
    sheet_names: list[str] | None = None,
) -> dict[str, Any]:
    excel_name = _normalize_filename(excel_file.filename or "shipment.xlsx")
    pdf_name = _normalize_filename(pdf_file.filename or "wvta.pdf")
    _validate_excel_filename(excel_name)
    _validate_pdf_filename(pdf_name)

    job_id = f"coc-fill-{uuid.uuid4().hex[:8]}"
    job_dir = COC_MATCH_JOB_ROOT / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    excel_dst = job_dir / f"excel-{excel_name}"
    pdf_dst = job_dir / f"wvta-{pdf_name}"
    excel_dst.write_bytes(excel_file.file.read())
    pdf_dst.write_bytes(pdf_file.file.read())

    initial_state = _base_fill_state(
        job_id=job_id,
        excel_name=excel_name,
        pdf_name=pdf_name,
        overwrite_existing=overwrite_existing,
        conflict_strategy=conflict_strategy,
        include_result_sheet=include_result_sheet,
        sheet_names=sheet_names,
        triggered_by=triggered_by,
    )
    persist_job_state(state_path(COC_MATCH_JOB_ROOT, job_id), initial_state)
    state = _start_fill_runner(
        job_id=job_id,
        excel_path=excel_dst,
        pdf_path=pdf_dst,
        overwrite_existing=overwrite_existing,
        conflict_strategy=conflict_strategy,
        include_result_sheet=include_result_sheet,
        sheet_names=sheet_names,
        triggered_by=triggered_by,
    )
    _trim_coc_fill_history()
    return state


def initiate_coc_fill_upload(
    *,
    filename: str,
    size_bytes: int,
    resume_key: str | None = None,
    triggered_by: str = "anonymous",
) -> dict[str, Any]:
    suffix = Path(filename).suffix.lower()
    if suffix not in (ALLOWED_FILL_EXCEL_EXTENSIONS | ALLOWED_FILL_PDF_EXTENSIONS):
        raise HTTPException(status_code=400, detail="COC 填充仅支持 .xlsx/.xlsm/.pdf。")

    from upload_toolkit.upload_engine import (
        create_upload_session as _toolkit_create,
        find_resumable_session,
    )

    if resume_key:
        existing = find_resumable_session(
            _FILL_UPLOAD_SESSION_ROOT,
            resume_key=resume_key,
            filename=filename,
            size_bytes=size_bytes,
        )
        if existing:
            return existing

    return _toolkit_create(
        _FILL_UPLOAD_SESSION_ROOT,
        filename=filename,
        size_bytes=size_bytes,
        chunk_size=_COC_FILL_UPLOAD_CHUNK_SIZE,
        resume_key=resume_key or None,
        triggered_by=triggered_by,
    )


def upload_coc_fill_chunk(
    upload_id: str,
    part_number: int,
    content: bytes,
    chunk_sha256: str | None = None,
) -> dict[str, Any]:
    from upload_toolkit.upload_engine import receive_chunk as _toolkit_chunk

    return _toolkit_chunk(
        _FILL_UPLOAD_SESSION_ROOT,
        upload_id,
        part_number,
        content,
        chunk_sha256=chunk_sha256,
        expected_chunk_size=_COC_FILL_UPLOAD_CHUNK_SIZE,
    )


def complete_coc_fill_upload(upload_id: str) -> dict[str, Any]:
    from upload_toolkit.upload_engine import complete_upload_session as _toolkit_complete

    return _toolkit_complete(_FILL_UPLOAD_SESSION_ROOT, upload_id)


def _get_fill_assembled_path(upload_id: str) -> Path:
    from upload_toolkit.upload_engine import get_upload_session as _toolkit_get

    state = _toolkit_get(_FILL_UPLOAD_SESSION_ROOT, upload_id)
    if state.get("status") != "completed":
        raise HTTPException(status_code=400, detail="上传未完成。")
    return _FILL_UPLOAD_SESSION_ROOT / upload_id / "assembled" / str(state.get("filename", "upload.bin"))


def create_coc_fill_job_from_upload(
    *,
    excel_upload_id: str,
    pdf_upload_id: str,
    excel_filename: str,
    pdf_filename: str,
    overwrite_existing: bool,
    conflict_strategy: str,
    triggered_by: str,
    include_result_sheet: bool = False,
    sheet_names: list[str] | None = None,
) -> dict[str, Any]:
    excel_name = _normalize_filename(excel_filename)
    pdf_name = _normalize_filename(pdf_filename)
    _validate_excel_filename(excel_name)
    _validate_pdf_filename(pdf_name)

    excel_path = _get_fill_assembled_path(excel_upload_id)
    pdf_path = _get_fill_assembled_path(pdf_upload_id)

    job_id = f"coc-fill-{uuid.uuid4().hex[:8]}"
    job_dir = COC_MATCH_JOB_ROOT / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    excel_dst = job_dir / f"excel-{excel_name}"
    pdf_dst = job_dir / f"wvta-{pdf_name}"
    shutil.copyfile(excel_path, excel_dst)
    shutil.copyfile(pdf_path, pdf_dst)

    initial_state = _base_fill_state(
        job_id=job_id,
        excel_name=excel_name,
        pdf_name=pdf_name,
        overwrite_existing=overwrite_existing,
        conflict_strategy=conflict_strategy,
        include_result_sheet=include_result_sheet,
        sheet_names=sheet_names,
        triggered_by=triggered_by,
    )
    persist_job_state(state_path(COC_MATCH_JOB_ROOT, job_id), initial_state)
    state = _start_fill_runner(
        job_id=job_id,
        excel_path=excel_dst,
        pdf_path=pdf_dst,
        overwrite_existing=overwrite_existing,
        conflict_strategy=conflict_strategy,
        include_result_sheet=include_result_sheet,
        sheet_names=sheet_names,
        triggered_by=triggered_by,
    )
    _trim_coc_fill_history()
    return state


def _trim_coc_fill_history() -> None:
    payloads = [
        payload
        for payload in list_job_payloads(COC_MATCH_JOB_ROOT)
        if str(payload.get("jobType", "")) == "fill"
    ]
    if len(payloads) <= _MAX_FILL_HISTORY:
        return
    payloads.sort(key=lambda payload: str(payload.get("createdAt", "")))
    for old in payloads[:-_MAX_FILL_HISTORY]:
        job_id = str(old.get("jobId", ""))
        if not job_id:
            continue
        job_path = COC_MATCH_JOB_ROOT / job_id
        if job_path.exists():
            shutil.rmtree(job_path, ignore_errors=True)


def list_coc_fill_jobs(limit: int = 50) -> dict[str, Any]:
    _trim_coc_fill_history()
    payloads = [
        payload
        for payload in list_job_payloads(COC_MATCH_JOB_ROOT)
        if str(payload.get("jobType", "")) == "fill"
    ]
    payloads.sort(key=lambda payload: str(payload.get("createdAt", "")), reverse=True)
    return {"items": payloads[:limit]}


def get_coc_fill_job(job_id: str) -> dict[str, Any]:
    sp = state_path(COC_MATCH_JOB_ROOT, job_id)
    if not sp.exists():
        raise HTTPException(status_code=404, detail=f"COC 填充任务不存在: {job_id}")
    payload = load_job_state(sp)
    if payload.get("jobType") != "fill":
        raise HTTPException(status_code=404, detail=f"COC 填充任务不存在: {job_id}")
    return payload


def get_coc_fill_workbook_path(job_id: str) -> Path:
    state = get_coc_fill_job(job_id)
    if state.get("status") != "success":
        raise HTTPException(status_code=404, detail="填充结果尚未生成。")
    output = COC_MATCH_JOB_ROOT / job_id / str(state.get("outputFilename") or "filled.xlsx")
    if not output.exists():
        raise HTTPException(status_code=404, detail="填充结果文件不存在。")
    return output


_STATUS_COUNT_FIELDS = {
    "filled": "filledCount",
    "not_found": "notFoundCount",
    "ambiguous": "ambiguousCount",
    "skipped_existing": "skippedExistingCount",
    "invalid_source": "invalidSourceCount",
}

_AMBIGUOUS_REASON = "PDF 存在多个 WVTA/COC 候选，无法唯一确定。"
_MANUAL_CANDIDATE_REASON = "人工选择 PDF 候选。"
_MANUAL_TEXT_REASON = "人工填写 WVTA/COC。"


def _override_key(payload: dict[str, Any]) -> tuple[str, int, str]:
    return (
        str(payload.get("sheetName") or payload.get("sheet_name") or ""),
        int(payload.get("rowNumber") or payload.get("row_number") or 0),
        normalize_material_group(payload.get("materialGroup") or payload.get("material_group")),
    )


def _preview_decision_for_override(
    state: dict[str, Any],
    override: dict[str, Any],
) -> dict[str, Any] | None:
    target = _override_key(override)
    for group in state.get("previewGroups") or []:
        if not isinstance(group, dict):
            continue
        for decision in group.get("decisions") or []:
            if isinstance(decision, dict) and _override_key(decision) == target:
                return decision
    return None


def _candidate_for_override(
    decision: dict[str, Any],
    override: dict[str, Any],
) -> dict[str, Any] | None:
    wvta_no = str(override.get("wvtaNo") or override.get("wvta_no") or "").strip()
    coc_no = str(override.get("cocNo") or override.get("coc_no") or "").strip()
    page_number = override.get("pageNumber") or override.get("page_number")
    table_row_number = override.get("tableRowNumber") or override.get("table_row_number")
    for candidate in decision.get("candidateRecords") or []:
        if not isinstance(candidate, dict):
            continue
        if str(candidate.get("wvtaNo") or "").strip() != wvta_no:
            continue
        if str(candidate.get("cocNo") or "").strip() != coc_no:
            continue
        if page_number is not None and int(candidate.get("pageNumber") or 0) != int(page_number):
            continue
        if table_row_number is not None and int(candidate.get("tableRowNumber") or 0) != int(table_row_number):
            continue
        return candidate
    return None


def _manual_values_for_override(override: dict[str, Any]) -> tuple[str, str]:
    wvta_no = str(override.get("wvtaNo") or override.get("wvta_no") or "").strip()
    coc_no = str(override.get("cocNo") or override.get("coc_no") or "").strip()
    if not wvta_no or not coc_no:
        raise HTTPException(status_code=422, detail="手工填写需要同时提供 WVTA 和 COC。")
    return wvta_no, coc_no


def _adjust_status_counts(container: dict[str, Any], old_status: str, new_status: str) -> None:
    if old_status == new_status:
        return
    status_counts = container.setdefault("statusCounts", {})
    if isinstance(status_counts, dict):
        status_counts[old_status] = max(0, int(status_counts.get(old_status, 0)) - 1)
        status_counts[new_status] = int(status_counts.get(new_status, 0)) + 1
    old_field = _STATUS_COUNT_FIELDS.get(old_status)
    if old_field:
        container[old_field] = max(0, int(container.get(old_field) or 0) - 1)
    new_field = _STATUS_COUNT_FIELDS.get(new_status)
    if new_field:
        container[new_field] = int(container.get(new_field) or 0) + 1


def _apply_candidate_to_decision_payload(
    decision: dict[str, Any],
    candidate: dict[str, Any],
    *,
    previous_wvta: str | None = None,
    previous_coc: str | None = None,
) -> str:
    old_status = str(decision.get("status") or "")
    old_reason = str(decision.get("reason") or "")
    decision["status"] = "filled"
    decision["reason"] = _MANUAL_CANDIDATE_REASON
    decision["confidence"] = 1
    decision["selectedRecord"] = candidate
    decision["writtenWvta"] = str(candidate.get("wvtaNo") or "")
    decision["writtenCoc"] = str(candidate.get("cocNo") or "")
    decision["manualPreviousWvta"] = previous_wvta or ""
    decision["manualPreviousCoc"] = previous_coc or ""
    decision["manualPreviousStatus"] = old_status
    decision["manualPreviousReason"] = old_reason
    return old_status


def _apply_manual_text_to_decision_payload(
    decision: dict[str, Any],
    *,
    wvta_no: str,
    coc_no: str,
    previous_wvta: str | None = None,
    previous_coc: str | None = None,
) -> str:
    old_status = str(decision.get("status") or "")
    old_reason = str(decision.get("reason") or "")
    decision["status"] = "filled"
    decision["reason"] = _MANUAL_TEXT_REASON
    decision["confidence"] = 1
    decision["selectedRecord"] = None
    decision["writtenWvta"] = wvta_no
    decision["writtenCoc"] = coc_no
    decision["manualPreviousWvta"] = previous_wvta or ""
    decision["manualPreviousCoc"] = previous_coc or ""
    decision["manualPreviousStatus"] = old_status
    decision["manualPreviousReason"] = old_reason
    return old_status


def _revert_manual_decision_payload(decision: dict[str, Any]) -> str:
    old_status = str(decision.get("status") or "")
    previous_status = str(decision.get("manualPreviousStatus") or "ambiguous")
    previous_reason = str(decision.get("manualPreviousReason") or _AMBIGUOUS_REASON)
    decision["status"] = previous_status
    decision["reason"] = previous_reason
    decision["confidence"] = 0 if previous_status == "not_found" else 0.3
    decision["selectedRecord"] = None
    decision["writtenWvta"] = None
    decision["writtenCoc"] = None
    decision.pop("manualPreviousWvta", None)
    decision.pop("manualPreviousCoc", None)
    decision.pop("manualPreviousStatus", None)
    decision.pop("manualPreviousReason", None)
    return old_status


def _update_result_sheet_override(
    workbook: Any,
    *,
    sheet_name: str,
    row_number: int,
    material_group: str,
    candidate: dict[str, Any],
    reason: str,
) -> None:
    if "COC填充结果" not in workbook.sheetnames:
        return
    result_sheet = workbook["COC填充结果"]
    for result_row in range(2, (result_sheet.max_row or 0) + 1):
        source_sheet = str(result_sheet.cell(result_row, 2).value or "")
        try:
            source_row = int(result_sheet.cell(result_row, 3).value or 0)
        except (TypeError, ValueError):
            source_row = 0
        source_material = normalize_material_group(result_sheet.cell(result_row, 4).value)
        if source_sheet != sheet_name or source_row != row_number or source_material != material_group:
            continue
        result_sheet.cell(result_row, 1).value = "filled"
        result_sheet.cell(result_row, 5).value = str(candidate.get("wvtaNo") or "")
        result_sheet.cell(result_row, 6).value = str(candidate.get("cocNo") or "")
        result_sheet.cell(result_row, 8).value = reason
        result_sheet.cell(result_row, 9).value = candidate.get("pageNumber") or ""
        result_sheet.cell(result_row, 10).value = candidate.get("tableRowNumber") or ""
        result_sheet.cell(result_row, 11).value = candidate.get("comments") or ""
        return


def _revert_result_sheet_override(
    workbook: Any,
    *,
    sheet_name: str,
    row_number: int,
    material_group: str,
    status: str,
    reason: str,
) -> None:
    if "COC填充结果" not in workbook.sheetnames:
        return
    result_sheet = workbook["COC填充结果"]
    for result_row in range(2, (result_sheet.max_row or 0) + 1):
        source_sheet = str(result_sheet.cell(result_row, 2).value or "")
        try:
            source_row = int(result_sheet.cell(result_row, 3).value or 0)
        except (TypeError, ValueError):
            source_row = 0
        source_material = normalize_material_group(result_sheet.cell(result_row, 4).value)
        if source_sheet != sheet_name or source_row != row_number or source_material != material_group:
            continue
        result_sheet.cell(result_row, 1).value = status
        result_sheet.cell(result_row, 5).value = ""
        result_sheet.cell(result_row, 6).value = ""
        result_sheet.cell(result_row, 8).value = reason
        result_sheet.cell(result_row, 9).value = ""
        result_sheet.cell(result_row, 10).value = ""
        result_sheet.cell(result_row, 11).value = ""
        return


def _sync_flat_decision_payload(
    state: dict[str, Any],
    override: dict[str, Any],
    candidate: dict[str, Any],
    *,
    previous_wvta: str | None = None,
    previous_coc: str | None = None,
) -> None:
    target = _override_key(override)
    for decision in state.get("decisions") or []:
        if isinstance(decision, dict) and _override_key(decision) == target:
            _apply_candidate_to_decision_payload(
                decision,
                candidate,
                previous_wvta=previous_wvta,
                previous_coc=previous_coc,
            )
            return


def _sync_flat_manual_text_payload(
    state: dict[str, Any],
    override: dict[str, Any],
    *,
    wvta_no: str,
    coc_no: str,
    previous_wvta: str | None = None,
    previous_coc: str | None = None,
) -> None:
    target = _override_key(override)
    for decision in state.get("decisions") or []:
        if isinstance(decision, dict) and _override_key(decision) == target:
            _apply_manual_text_to_decision_payload(
                decision,
                wvta_no=wvta_no,
                coc_no=coc_no,
                previous_wvta=previous_wvta,
                previous_coc=previous_coc,
            )
            return


def _sync_flat_decision_revert(state: dict[str, Any], override: dict[str, Any]) -> None:
    target = _override_key(override)
    for decision in state.get("decisions") or []:
        if isinstance(decision, dict) and _override_key(decision) == target:
            _revert_manual_decision_payload(decision)
            return


def apply_coc_fill_overrides(
    job_id: str,
    overrides: list[dict[str, Any]],
) -> dict[str, Any]:
    state = get_coc_fill_job(job_id)
    if state.get("status") != "success":
        raise HTTPException(status_code=400, detail="只能在填充任务完成后选择候选。")
    if not overrides:
        return state

    seen_targets: set[tuple[str, int, str]] = set()
    prepared: list[tuple[dict[str, Any], dict[str, Any], dict[str, Any], dict[str, Any], str]] = []
    for override in overrides:
        if not isinstance(override, dict):
            raise HTTPException(status_code=422, detail="候选选择参数格式错误。")
        target = _override_key(override)
        if target in seen_targets:
            raise HTTPException(status_code=422, detail="同一行不能重复提交候选选择。")
        seen_targets.add(target)
        decision = _preview_decision_for_override(state, override)
        if decision is None:
            raise HTTPException(status_code=404, detail="当前预览中未找到该行，请重新运行填充任务。")
        status = str(decision.get("status") or "")
        if status == "ambiguous":
            candidate = _candidate_for_override(decision, override)
            if candidate is None:
                raise HTTPException(status_code=422, detail="所选 WVTA/COC 不属于该行的 PDF 候选。")
            mode = "candidate"
        elif status == "not_found":
            wvta_no, coc_no = _manual_values_for_override(override)
            candidate = {
                "wvtaNo": wvta_no,
                "cocNo": coc_no,
                "pageNumber": "",
                "tableRowNumber": "",
                "comments": "用户手工填写",
            }
            mode = "manual_text"
        else:
            raise HTTPException(status_code=409, detail="只有冲突行或未命中行支持人工填写。")
        group = next(
            (
                item
                for item in state.get("previewGroups") or []
                if isinstance(item, dict) and str(item.get("sheetName") or "") == _override_key(override)[0]
            ),
            None,
        )
        if not isinstance(group, dict):
            raise HTTPException(status_code=404, detail="当前预览中未找到该 Sheet。")
        prepared.append((override, decision, candidate, group, mode))

    output_path = get_coc_fill_workbook_path(job_id)
    workbook = load_workbook(output_path)
    previous_values: dict[tuple[str, int, str], tuple[str | None, str | None]] = {}
    try:
        for override, _decision, candidate, _group, mode in prepared:
            sheet_name, row_number, material_group = _override_key(override)
            if sheet_name not in workbook.sheetnames:
                raise HTTPException(status_code=404, detail=f"输出 Excel 中未找到 Sheet：{sheet_name}")
            sheet = workbook[sheet_name]
            if row_number < 1 or row_number > (sheet.max_row or 0):
                raise HTTPException(status_code=422, detail=f"{sheet_name} 第 {row_number} 行不存在。")
            targets = target_columns_for_sheet(sheet)
            if targets is None:
                raise HTTPException(status_code=422, detail=f"{sheet_name} 未识别到物料号组列。")
            previous_values[(sheet_name, row_number, material_group)] = (
                cell_text(sheet.cell(row_number, targets.wvta_col).value),
                cell_text(sheet.cell(row_number, targets.coc_col).value),
            )
            sheet.cell(row_number, targets.wvta_col).value = str(candidate.get("wvtaNo") or "")
            sheet.cell(row_number, targets.coc_col).value = str(candidate.get("cocNo") or "")
            _update_result_sheet_override(
                workbook,
                sheet_name=sheet_name,
                row_number=row_number,
                material_group=material_group,
                candidate=candidate,
                reason=_MANUAL_CANDIDATE_REASON if mode == "candidate" else _MANUAL_TEXT_REASON,
            )
        workbook.save(output_path)
    finally:
        workbook.close()

    for override, decision, candidate, group, mode in prepared:
        previous_wvta, previous_coc = previous_values.get(_override_key(override), (None, None))
        if mode == "candidate":
            old_status = _apply_candidate_to_decision_payload(
                decision,
                candidate,
                previous_wvta=previous_wvta,
                previous_coc=previous_coc,
            )
        else:
            old_status = _apply_manual_text_to_decision_payload(
                decision,
                wvta_no=str(candidate.get("wvtaNo") or ""),
                coc_no=str(candidate.get("cocNo") or ""),
                previous_wvta=previous_wvta,
                previous_coc=previous_coc,
            )
        _adjust_status_counts(group, old_status, "filled")
        _adjust_status_counts(state, old_status, "filled")
        if mode == "candidate":
            _sync_flat_decision_payload(
                state,
                override,
                candidate,
                previous_wvta=previous_wvta,
                previous_coc=previous_coc,
            )
        else:
            _sync_flat_manual_text_payload(
                state,
                override,
                wvta_no=str(candidate.get("wvtaNo") or ""),
                coc_no=str(candidate.get("cocNo") or ""),
                previous_wvta=previous_wvta,
                previous_coc=previous_coc,
            )

    state["manualOverrideCount"] = int(state.get("manualOverrideCount") or 0) + len(prepared)
    state["updatedAt"] = datetime.now().isoformat()
    persist_job_state(state_path(COC_MATCH_JOB_ROOT, job_id), state)
    return state


def revert_coc_fill_overrides(
    job_id: str,
    overrides: list[dict[str, Any]],
) -> dict[str, Any]:
    state = get_coc_fill_job(job_id)
    if state.get("status") != "success":
        raise HTTPException(status_code=400, detail="只能在填充任务完成后撤回人工选择。")
    if not overrides:
        return state

    seen_targets: set[tuple[str, int, str]] = set()
    prepared: list[tuple[dict[str, Any], dict[str, Any], dict[str, Any]]] = []
    for override in overrides:
        if not isinstance(override, dict):
            raise HTTPException(status_code=422, detail="撤回参数格式错误。")
        target = _override_key(override)
        if target in seen_targets:
            raise HTTPException(status_code=422, detail="同一行不能重复提交撤回。")
        seen_targets.add(target)
        decision = _preview_decision_for_override(state, override)
        if decision is None:
            raise HTTPException(status_code=404, detail="当前预览中未找到该行。")
        if str(decision.get("status") or "") != "filled" or str(decision.get("reason") or "") not in {
            _MANUAL_CANDIDATE_REASON,
            _MANUAL_TEXT_REASON,
        }:
            raise HTTPException(status_code=409, detail="只有人工填写或人工选择的候选支持撤回。")
        group = next(
            (
                item
                for item in state.get("previewGroups") or []
                if isinstance(item, dict) and str(item.get("sheetName") or "") == target[0]
            ),
            None,
        )
        if not isinstance(group, dict):
            raise HTTPException(status_code=404, detail="当前预览中未找到该 Sheet。")
        prepared.append((override, decision, group))

    output_path = get_coc_fill_workbook_path(job_id)
    workbook = load_workbook(output_path)
    try:
        for override, decision, _group in prepared:
            sheet_name, row_number, material_group = _override_key(override)
            if sheet_name not in workbook.sheetnames:
                raise HTTPException(status_code=404, detail=f"输出 Excel 中未找到 Sheet：{sheet_name}")
            sheet = workbook[sheet_name]
            if row_number < 1 or row_number > (sheet.max_row or 0):
                raise HTTPException(status_code=422, detail=f"{sheet_name} 第 {row_number} 行不存在。")
            targets = target_columns_for_sheet(sheet)
            if targets is None:
                raise HTTPException(status_code=422, detail=f"{sheet_name} 未识别到物料号组列。")
            previous_wvta = str(decision.get("manualPreviousWvta") or "")
            previous_coc = str(decision.get("manualPreviousCoc") or "")
            sheet.cell(row_number, targets.wvta_col).value = previous_wvta or None
            sheet.cell(row_number, targets.coc_col).value = previous_coc or None
            previous_status = str(decision.get("manualPreviousStatus") or "ambiguous")
            previous_reason = str(decision.get("manualPreviousReason") or _AMBIGUOUS_REASON)
            _revert_result_sheet_override(
                workbook,
                sheet_name=sheet_name,
                row_number=row_number,
                material_group=material_group,
                status=previous_status,
                reason=previous_reason,
            )
        workbook.save(output_path)
    finally:
        workbook.close()

    for override, decision, group in prepared:
        previous_status = str(decision.get("manualPreviousStatus") or "ambiguous")
        old_status = _revert_manual_decision_payload(decision)
        _adjust_status_counts(group, old_status, previous_status)
        _adjust_status_counts(state, old_status, previous_status)
        _sync_flat_decision_revert(state, override)

    state["manualOverrideCount"] = max(0, int(state.get("manualOverrideCount") or 0) - len(prepared))
    state["updatedAt"] = datetime.now().isoformat()
    persist_job_state(state_path(COC_MATCH_JOB_ROOT, job_id), state)
    return state
