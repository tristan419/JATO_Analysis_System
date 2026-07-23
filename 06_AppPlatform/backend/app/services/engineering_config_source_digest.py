"""Lightweight digest for engineering config source snapshots.

This module intentionally does not write extracted values into the canonical
engineering config tables. It only produces a reviewable preview from a source
file so the upload panel can show what the file appears to contain.
"""

from __future__ import annotations

import hashlib
import importlib
import os
import re
import shlex
import shutil
import subprocess
import tempfile
import warnings
from pathlib import Path
from typing import Any

import pandas as pd

from app.services.config_availability import classify_availability
from app.services.config_text_normalization import (
    clean_config_cell,
    normalise_config_space,
    normalize_config_feature_key,
    normalize_config_value_for_compare,
    normalize_header_phrase,
)

MAX_SHEETS = 24
SHEET_PREVIEW_ROWS = 160
SHEET_PREVIEW_COLUMNS = 48
TABULAR_PREVIEW_ROWS = 160
MAX_COMPARE_TRIMS = 8
OCR_HEADERLESS_DIGEST_TYPES = {"image_ocr", "pdf_ocr"}
OCR_HEADERLESS_MIN_ROWS = 3
OCR_MIN_COMPARABLE_FEATURES = 2
WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
OPENPYXL_WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm"}
TABULAR_EXTENSIONS = {".csv", ".tsv", ".html", ".htm"}
PDF_EXTENSIONS = {".pdf"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}
OcrCandidateScore = tuple[int, int, int, int, int, int, int, int, int, int, int, int]
OCR_CONFIG_SEMANTIC_STRATEGY = "highest_config_semantic_score"
ROW_VARIANT_HEADER_TOKENS = {"type", "trim", "version", "车型", "配置", "版本"}
SIMPLE_CATEGORY_HEADER_TOKENS = {"category", "类别", "大类", "配置大类"}
SIMPLE_FEATURE_HEADER_TOKENS = {"feature", "feature name", "配置项", "配置名称", "参数", "项目", "item", "name"}
PRICE_LIST_IDENTITY_HEADER_TOKENS = {
    "brand",
    "make",
    "manufacturer",
    "model",
    "model name",
    "modelname",
    "trim",
    "grade",
    "version",
    "variant",
    "configuration",
    "sales version",
    "country",
    "market",
    "model year",
    "modelyear",
    "year",
    "powertrain",
    "drivetrain",
    "engine",
    "fuel",
    "energy",
    "energy type",
    "material no",
    "material no.",
    "material number",
    "sku",
    "code",
    "品牌",
    "厂牌",
    "车型",
    "车系",
    "型号",
    "配置",
    "版型",
    "版本",
    "国家",
    "市场",
    "年款",
    "年份",
    "动力",
    "驱动",
    "燃料",
    "能源",
    "物料号",
}
PRICE_LIST_VALUE_HEADER_TOKENS = {
    "price",
    "msrp",
    "rrp",
    "retail",
    "list price",
    "target price",
    "transaction price",
    "offer price",
    "otr",
    "on the road",
    "currency",
    "售价",
    "价格",
    "指导价",
    "建议零售价",
    "零售价",
    "目标价",
    "成交价",
    "报价",
    "币种",
    "货币",
}
SOURCE_MISSING_TOKENS = {""}
SOURCE_EQUIPPED_TOKENS = {"●", "有", "yes", "y", "true", "标配", "标准", "standard", "s", "✓", "✔"}
SOURCE_OPTIONAL_TOKENS = {"选装", "可选", "optional", "o", "○", "◯"}
SOURCE_NOT_EQUIPPED_TOKENS = {"-", "--", "---", "—", "/", "无", "no", "false", "不配备", "×", "✘", "✕"}
SOURCE_NOT_APPLICABLE_TOKENS = {"n/a", "na", "not applicable", "不适用"}
SOURCE_CANCELLED_TOKENS = {"取消", "停用", "删除", "减少", "cancelled", "canceled", "removed", "delete"}
EU_PROFILE_FIELD_SPECS = (
    ("country", ("国家", "country")),
    ("configurationVersion", ("配置版型", "configuration version", "configuration model", "configure the pattern")),
    ("materialNo", ("物料号", "material no", "material no.", "item number")),
    ("familyIdentifier", ("家族识别码", "family identifier", "interpolation family identifier", "family identification code")),
    ("variantVersion", ("版本号", "variant/version", "variant version")),
)
EU_CONFIG_INDEX_SHEETS = {"总表", "配置修订记录", "停用物料号", "欧盟车型数据图"}
EU_CONFIG_NOTE_COLUMN_TOKENS = {
    "删除",
    "减少",
    "备注",
    "增加或更改",
    "来源",
    "说明",
}
EU_CONFIG_CANCEL_TOKENS = {"取消", "停用"}
EU_BLANK_INFERENCE_BLOCKED_CATEGORY_TOKENS = {
    "基本参数",
    "basic parameter",
    "尺寸",
    "重量",
    "dimension",
    "weight",
    "性能",
    "performance",
    "动力",
    "powertrain",
    "电机",
    "电池",
    "battery",
}
MODEL_STEM_STOP_WORDS = {
    "basic",
    "parameters",
    "parameter",
    "基础参数",
    "配置",
    "config",
    "configuration",
}

PRICE_LIST_IDENTITY_FIELD_ALIASES = {
    "brand": ("brand", "make", "manufacturer", "品牌", "厂牌"),
    "model": ("model", "model name", "modelname", "车型", "车系", "型号"),
    "trim": ("trim", "grade", "version", "variant", "configuration", "配置", "版型", "版本"),
    "market": ("market", "country", "市场", "国家"),
    "modelYear": ("model year", "modelyear", "year", "年款", "年份"),
    "powertrain": ("powertrain", "drivetrain", "engine", "fuel", "energy", "energy type", "动力", "驱动", "燃料", "能源"),
    "materialNo": ("material no", "material no.", "material number", "物料号"),
    "salesVersion": ("sales version", "sku", "code", "销售版本", "销售代码"),
}


def _clean_cell(value: object) -> str:
    return clean_config_cell(value)


def _clean_row(row: list[object]) -> list[str]:
    return [_clean_cell(value) for value in row]


def _non_empty_count(row: list[str]) -> int:
    return sum(1 for value in row if value)


def _looks_numeric(value: str) -> bool:
    return bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?", value.replace(",", "")))


def _looks_numeric_value(value: str) -> bool:
    return bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?", value.replace(",", "")))


def _excel_column_letter(column_number: int) -> str:
    result = ""
    current = column_number
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result or "A"


def _slug(value: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "-", value).strip("-")
    return cleaned[:48] or "field"


def _compact_name(parts: list[str]) -> str:
    result: list[str] = []
    for part in parts:
        cleaned = part.strip()
        if cleaned and cleaned not in result:
            result.append(cleaned)
    return " / ".join(result)


def _infer_model_name(file_name: str, sheets: list[tuple[str, pd.DataFrame]]) -> str:
    stem_tokens = re.split(r"[\s_-]+", re.sub(r"\([^)]*\)", "", Path(file_name).stem))
    model_tokens = [
        token
        for token in stem_tokens
        if token and token.lower() not in MODEL_STEM_STOP_WORDS
    ]
    if model_tokens:
        return " ".join(model_tokens[:4]).strip()

    for _sheet_name, frame in sheets:
        for raw_row in frame.head(4).fillna("").values.tolist():
            for value in _clean_row(raw_row):
                lower = value.lower()
                if (
                    value
                    and len(value) <= 80
                    and "category" not in lower
                    and "parameter" not in lower
                    and "类别" not in value
                    and "参数" not in value
                    and value.upper() not in {"BEV", "PHEV", "HEV", "ICE"}
                ):
                    return value
    stem = Path(file_name).stem
    stem = re.sub(r"\([^)]*\)", "", stem)
    return re.sub(r"[_-]+", " ", stem).strip() or stem


def _preview_frame(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.iloc[:SHEET_PREVIEW_ROWS, :SHEET_PREVIEW_COLUMNS].copy()


def _clean_frame_rows(frame: pd.DataFrame, *, preview_only: bool = False) -> list[list[str]]:
    source = _preview_frame(frame) if preview_only else frame
    return [_clean_row(row) for row in source.fillna("").values.tolist()]


def _normalise_space(value: str) -> str:
    return normalise_config_space(value)


def _feature_code(feature_key: str) -> str:
    normalized_key = normalize_config_feature_key(feature_key) or feature_key
    digest = hashlib.sha1(normalized_key.encode("utf-8")).hexdigest()[:16]
    return f"digest_{digest}"


def _cell(rows: list[list[str]], row_index: int, col_index: int) -> str:
    if row_index < 0 or row_index >= len(rows):
        return ""
    row = rows[row_index]
    if col_index < 0 or col_index >= len(row):
        return ""
    return row[col_index]


def _eu_header_indices(rows: list[list[str]]) -> tuple[int, int, int] | None:
    for index, row in enumerate(rows[:12]):
        zh_label = row[1] if len(row) > 1 else ""
        en_label = (row[2] if len(row) > 2 else "").lower()
        if "配置版型" in zh_label or "configuration version" in en_label:
            return max(0, index - 1), index, min(index + 1, len(rows) - 1)
        if "国家" in zh_label and "country" in en_label and index + 2 < len(rows):
            return index, index + 1, index + 2
    return None


def _is_eu_config_workbook(sheets: list[tuple[str, pd.DataFrame]]) -> bool:
    sheet_names = {sheet_name.strip() for sheet_name, _frame in sheets}
    if "总表" not in sheet_names:
        return False
    for sheet_name, frame in sheets:
        if sheet_name.strip() in EU_CONFIG_INDEX_SHEETS:
            continue
        rows = _clean_frame_rows(frame, preview_only=True)
        if _eu_header_indices(rows) is not None:
            return True
    return False


def _eu_sheet_title(sheet_name: str, rows: list[list[str]]) -> str:
    for row in rows[:2]:
        for value in row:
            cleaned = _normalise_space(value)
            if cleaned and len(cleaned) <= 140:
                return cleaned
    return sheet_name.strip()


def _eu_column_has_note_header(*values: str) -> bool:
    joined = " ".join(_normalise_space(value) for value in values if value)
    return any(token in joined for token in EU_CONFIG_NOTE_COLUMN_TOKENS)


def _eu_column_is_cancelled(*values: str) -> bool:
    joined = " ".join(_normalise_space(value) for value in values if value)
    return any(token in joined for token in EU_CONFIG_CANCEL_TOKENS)


def _eu_profile_field(row: list[str]) -> str | None:
    zh_label = _normalise_space(row[1]) if len(row) > 1 else ""
    en_label = _normalise_space(row[2]).lower() if len(row) > 2 else ""
    label = f"{zh_label} {en_label}".strip()
    if not label:
        return None
    for field, tokens in EU_PROFILE_FIELD_SPECS:
        if any(token.lower() in label.lower() for token in tokens):
            return field
    return None


def _raw_value_is_unknown(value: str) -> bool:
    return _source_value_semantics(value)["availability"] == "UNKNOWN"


def _source_value_semantics(raw_value: str) -> dict[str, Any]:
    cleaned = _normalise_space(raw_value)
    lowered = cleaned.lower()
    if lowered in SOURCE_MISSING_TOKENS:
        return {
            "availability": "UNKNOWN",
            "normalizedValue": None,
            "unit": None,
            "valueState": "blank",
            "displayValue": "待确认",
        }
    if lowered in SOURCE_CANCELLED_TOKENS or any(token in lowered for token in SOURCE_CANCELLED_TOKENS if len(token) > 2):
        return {
            "availability": "CANCELLED_OR_REMOVED",
            "normalizedValue": cleaned or None,
            "unit": None,
            "valueState": "cancelled_or_removed",
            "displayValue": cleaned or "取消 / 删除",
        }
    if lowered in SOURCE_NOT_APPLICABLE_TOKENS:
        return {
            "availability": "NOT_APPLICABLE",
            "normalizedValue": None,
            "unit": None,
            "valueState": "not_applicable",
            "displayValue": "不适用",
        }
    if lowered in SOURCE_NOT_EQUIPPED_TOKENS:
        return {
            "availability": "NOT_AVAILABLE",
            "normalizedValue": None,
            "unit": None,
            "valueState": "marker_value",
            "displayValue": "不配备",
        }
    if lowered in SOURCE_OPTIONAL_TOKENS:
        return {
            "availability": "OPTIONAL",
            "normalizedValue": "选装",
            "unit": None,
            "valueState": "marker_value",
            "displayValue": "选装",
        }
    if lowered in SOURCE_EQUIPPED_TOKENS:
        return {
            "availability": "STANDARD",
            "normalizedValue": "标配",
            "unit": None,
            "valueState": "marker_value",
            "displayValue": "标配",
        }

    availability, normalized_value, unit = classify_availability(raw_value)
    if availability == "STANDARD":
        value_state = "marker_value"
        display_value = "标配"
    elif availability == "OPTIONAL":
        value_state = "marker_value"
        display_value = "选装"
    elif availability == "NOT_AVAILABLE":
        value_state = "marker_value"
        display_value = "不配备"
    elif availability == "NOT_APPLICABLE":
        value_state = "not_applicable"
        display_value = "不适用"
    elif availability == "UNKNOWN":
        value_state = "blank"
        display_value = "待确认"
    else:
        value_state = "numeric_value" if normalized_value and _looks_numeric_value(normalized_value) else "text_value"
        display_value = normalized_value or cleaned
    return {
        "availability": availability,
        "normalizedValue": normalized_value,
        "unit": unit,
        "valueState": value_state,
        "displayValue": display_value,
    }


def _source_value_signature(value: str | dict[str, Any]) -> tuple[str, str, str]:
    semantics = _source_value_semantics(value) if isinstance(value, str) else value
    compare_value = normalize_config_value_for_compare(
        str(semantics.get("displayValue") or semantics.get("normalizedValue") or "")
    )
    return (
        str(semantics["availability"]),
        compare_value or "",
        "",
    )


def _source_value_is_present(value: str) -> bool:
    return _source_value_semantics(value)["availability"] in {"STANDARD", "OPTIONAL", "VALUE"}


def _source_cell_payload(
    frame: pd.DataFrame,
    sheet_name: str,
    row_index: int,
    col_index: int,
) -> dict[str, Any]:
    source_cells = frame.attrs.get("source_cells", {})
    source = source_cells.get((row_index, col_index))
    if source:
        return dict(source)
    column_number = col_index + 1
    row_number = row_index + 1
    column_letter = _excel_column_letter(column_number)
    return {
        "sheetName": sheet_name,
        "rowNumber": row_number,
        "columnNumber": column_number,
        "columnLetter": column_letter,
        "cell": f"{column_letter}{row_number}",
        "sourceCell": f"{column_letter}{row_number}",
        "mergedRange": None,
    }


def _eu_category_allows_blank_inference(category: str) -> bool:
    normalized = _normalise_space(category).lower()
    if not normalized:
        return False
    return not any(token in normalized for token in EU_BLANK_INFERENCE_BLOCKED_CATEGORY_TOKENS)


def _source_value_has_explicit_config(semantics: dict[str, Any]) -> bool:
    return str(semantics.get("availability") or "") in {"STANDARD", "OPTIONAL", "VALUE"}


def _eu_blank_inference_for_values(
    raw_values: list[str],
    sources: list[dict[str, Any]],
    *,
    category: str,
) -> list[dict[str, Any] | None]:
    if not _eu_category_allows_blank_inference(category):
        return [None for _value in raw_values]

    semantics = [_source_value_semantics(value) for value in raw_values]
    has_explicit_config = any(_source_value_has_explicit_config(item) for item in semantics)
    if not has_explicit_config:
        return [None for _value in raw_values]

    inferences: list[dict[str, Any] | None] = []
    for value, source in zip(raw_values, sources, strict=False):
        item = _source_value_semantics(value)
        if item["availability"] != "UNKNOWN":
            inferences.append(None)
            continue
        if source.get("mergedRange"):
            inferences.append(None)
            continue
        inferences.append(
            {
                "availability": "NOT_AVAILABLE",
                "normalizedValue": None,
                "unit": None,
                "valueState": "blank",
                "displayValue": "不配备*",
                "inferred": True,
                "inferenceReason": "blank_as_not_equipped_by_eu_matrix_policy",
                "confidence": 0.7,
            }
        )
    return inferences


def _eu_feature_name(row: list[str]) -> str:
    return _compact_name([
        _normalise_space(row[2]) if len(row) > 2 else "",
        _normalise_space(row[1]) if len(row) > 1 else "",
    ])


def _eu_trim_columns(
    rows: list[list[str]],
    *,
    market_row_index: int,
    config_row_index: int,
    material_row_index: int,
) -> list[tuple[int, dict[str, Any]]]:
    max_columns = max((len(row) for row in rows), default=0)
    trims: list[tuple[int, dict[str, Any]]] = []
    active_market = ""
    for col_idx in range(3, max_columns):
        raw_market = _cell(rows, market_row_index, col_idx)
        config_version = _cell(rows, config_row_index, col_idx)
        material_no = _cell(rows, material_row_index, col_idx)
        if _eu_column_has_note_header(raw_market, config_version, material_no):
            continue
        if raw_market:
            active_market = raw_market
        market = raw_market or active_market

        value_count = 0
        for row_index in range(material_row_index + 1, len(rows)):
            row = rows[row_index]
            if _eu_feature_name(row) and _cell(rows, row_index, col_idx):
                value_count += 1
        is_cancelled = _eu_column_is_cancelled(config_version, material_no)
        if not any((market, config_version, material_no)):
            continue
        if not (config_version or material_no) and value_count < 2 and not is_cancelled:
            continue
        if value_count == 0 and not is_cancelled:
            continue

        trim_name = _normalise_space(config_version or market or material_no or f"Trim {col_idx - 2}")
        clean_material_no = "" if is_cancelled and material_no == "取消" else _normalise_space(material_no)
        trim_id = f"eu-config-c{col_idx}"
        trims.append(
            (
                col_idx,
                {
                    "trimId": trim_id,
                    "trimName": trim_name,
                    "fullTrimName": _compact_name([trim_name, clean_material_no]),
                    "modelName": "",
                    "market": _normalise_space(market),
                    "country": _normalise_space(market),
                    "materialNo": clean_material_no or None,
                    "salesVersion": trim_name,
                    "hasMaterialNo": bool(clean_material_no),
                    "dataOrigin": "own_catalog" if clean_material_no else "external_or_scraped",
                    "sourceColumn": col_idx + 1,
                    "sourceStatus": "cancelled" if is_cancelled else "active",
                },
            )
        )
        if len(trims) >= MAX_COMPARE_TRIMS:
            break
    return trims


def _eu_trim_profiles(
    rows: list[list[str]],
    trim_columns: list[tuple[int, dict[str, Any]]],
    *,
    market_row_index: int,
    material_row_index: int,
) -> tuple[dict[int, dict[str, str]], set[int]]:
    profiles: dict[int, dict[str, str]] = {col_idx: {} for col_idx, _trim in trim_columns}
    profile_rows: set[int] = set()
    scan_end = min(len(rows), material_row_index + 5)
    for row_index in range(market_row_index, scan_end):
        row = rows[row_index]
        field = _eu_profile_field(row)
        if field is None:
            continue
        profile_rows.add(row_index)
        active_value = ""
        for col_idx, _trim in trim_columns:
            raw_value = _cell(rows, row_index, col_idx)
            if raw_value:
                active_value = raw_value
            value = raw_value or active_value
            cleaned = _normalise_space(value)
            if cleaned:
                profiles[col_idx][field] = cleaned
    return profiles, profile_rows


def _eu_model_sheet_group(sheet_name: str, frame: pd.DataFrame) -> dict[str, Any] | None:
    rows = _clean_frame_rows(frame)
    header_indices = _eu_header_indices(rows)
    if header_indices is None:
        return None
    market_row_index, config_row_index, material_row_index = header_indices
    trim_columns = _eu_trim_columns(
        rows,
        market_row_index=market_row_index,
        config_row_index=config_row_index,
        material_row_index=material_row_index,
    )
    if not trim_columns:
        return None

    model_name = sheet_name.strip()
    title = _eu_sheet_title(sheet_name, rows)
    group_id = f"eu-config-{_slug(sheet_name)}"
    trims: list[dict[str, Any]] = []
    profiles, profile_row_indices = _eu_trim_profiles(
        rows,
        trim_columns,
        market_row_index=market_row_index,
        material_row_index=material_row_index,
    )
    for col_idx, payload in trim_columns:
        profile = profiles.get(col_idx, {})
        material_no = profile.get("materialNo") or payload.get("materialNo")
        sales_version = profile.get("configurationVersion") or payload.get("salesVersion")
        trim = {
            **payload,
            "trimId": f"{group_id}-c{col_idx}",
            "trimName": sales_version or payload["trimName"],
            "fullTrimName": _compact_name([sales_version or payload["trimName"], material_no or ""]),
            "modelName": model_name,
            "market": profile.get("country") or payload.get("market"),
            "country": profile.get("country") or payload.get("country"),
            "materialNo": material_no or None,
            "salesVersion": sales_version or payload.get("salesVersion"),
            "hasMaterialNo": bool(material_no),
            "dataOrigin": "own_catalog" if material_no else payload.get("dataOrigin"),
            "profile": profile,
            "sourceSheet": sheet_name,
        }
        trims.append(trim)

    compare_rows: list[dict[str, Any]] = []
    feature_start_index = (max(profile_row_indices) + 1) if profile_row_indices else (material_row_index + 1)
    active_category = ""
    for row_index in range(0, feature_start_index):
        if _cell(rows, row_index, 0):
            active_category = _cell(rows, row_index, 0)
    trim_ids = [str(trim["trimId"]) for trim in trims]
    for row_index in range(feature_start_index, len(rows)):
        row = rows[row_index]
        if len(row) > 0 and row[0]:
            active_category = row[0]
        feature_name = _eu_feature_name(row)
        if not feature_name:
            continue
        raw_values = [_cell(rows, row_index, col_idx) for col_idx, _trim in trim_columns]
        sources = [
            _source_cell_payload(frame, sheet_name, row_index, col_idx)
            for col_idx, _trim in trim_columns
        ]
        inference_semantics = _eu_blank_inference_for_values(
            raw_values,
            sources,
            category=active_category or sheet_name,
        )
        values = [
            _value_payload(
                group_id,
                row_index + 1,
                trim_ids[index],
                value,
                source=sources[index],
                semantics_override=inference_semantics[index],
            )
            for index, value in enumerate(raw_values)
        ]
        comparison_type = _comparison_type(values)
        feature_name_zh = _normalise_space(row[1]) if len(row) > 1 else ""
        feature_name_en = _normalise_space(row[2]) if len(row) > 2 else ""
        feature_key = "|".join([sheet_name.strip(), active_category or sheet_name, feature_name_zh, feature_name_en])
        unique_trim_ids = [
            trim_ids[index]
            for index, value in enumerate(values)
            if value["availability"] in {"STANDARD", "OPTIONAL", "VALUE"}
        ] if comparison_type in {"UNIQUE_TO_TRIM", "PARTIAL_AVAILABLE", "UNIQUE_OR_PARTIAL"} else []
        compare_rows.append(
            {
                "category": active_category or sheet_name,
                "featureKey": feature_key,
                "featureCode": _feature_code(feature_key),
                "featureName": feature_name,
                "comparisonType": comparison_type,
                "uniqueTrimIds": unique_trim_ids,
                "businessNote": _business_note(comparison_type),
                "values": values,
            }
        )

    return _finalize_compare_group(
        group_id=group_id,
        title=f"{model_name} · {title}" if title != model_name else model_name,
        source_sheet=sheet_name,
        model_name=model_name,
        trims=trims,
        rows=compare_rows,
    )


def _eu_config_workbook_digest(
    sheets: list[tuple[str, pd.DataFrame]],
    *,
    file_name: str,
) -> dict[str, Any]:
    sheet_payloads = [_sheet_digest(sheet_name, frame) for sheet_name, frame in sheets[:MAX_SHEETS]]
    compare_groups: list[dict[str, Any]] = []
    for sheet_name, frame in sheets[:MAX_SHEETS]:
        if sheet_name.strip() in EU_CONFIG_INDEX_SHEETS:
            continue
        group = _eu_model_sheet_group(sheet_name, frame)
        if group is not None:
            compare_groups.append(group)

    return {
        "digestType": "workbook",
        "workbookFormat": "eu_config_resource_table",
        "status": "ready",
        "fileName": file_name,
        "modelName": Path(file_name).stem,
        "summary": {
            "sheetCount": len(sheets),
            "tableCount": len(sheet_payloads),
            "candidateTrimCount": sum(group["trimCount"] for group in compare_groups),
            "comparableGroupCount": len([group for group in compare_groups if group["trimCount"] >= 2]),
            "featureCount": sum(group["featureCount"] for group in compare_groups),
            "differenceCount": sum(group["differenceCount"] for group in compare_groups),
        },
        "sheets": sheet_payloads,
        "compareGroups": compare_groups,
    }


def _sheet_digest(sheet_name: str, frame: pd.DataFrame) -> dict[str, Any]:
    rows = _clean_frame_rows(frame, preview_only=True)
    sample_rows = [row for row in rows if _non_empty_count(row) > 0][:6]
    return {
        "name": sheet_name,
        "rowCount": int(frame.shape[0]),
        "columnCount": int(frame.shape[1]),
        "nonEmptyCellCount": int(frame.fillna("").astype(str).ne("").sum().sum()),
        "sampleRows": sample_rows,
    }


def _comparison_type(values: list[str | dict[str, Any]]) -> str:
    semantics = [
        _source_value_semantics(value) if isinstance(value, str) else value
        for value in values
    ]
    availabilities = [str(value.get("availability") or "") for value in semantics]
    signatures = {_source_value_signature(value) for value in semantics}

    if len(signatures) == 1:
        return "COMMON_SAME"
    if any(availability == "CANCELLED_OR_REMOVED" for availability in availabilities):
        return "CANCELLED_OR_REMOVED"
    if any(availability == "UNKNOWN" for availability in availabilities):
        return "MISSING_UNKNOWN"
    if any(availability == "NOT_APPLICABLE" for availability in availabilities):
        return "NOT_APPLICABLE"
    if any(availability == "OPTIONAL" for availability in availabilities) and len(set(availabilities)) > 1:
        return "OPTIONAL_DIFFERENT"

    present_count = sum(1 for availability in availabilities if availability in {"STANDARD", "OPTIONAL", "VALUE"})
    not_equipped_count = sum(1 for availability in availabilities if availability == "NOT_AVAILABLE")
    if present_count > 0 and not_equipped_count > 0:
        return "UNIQUE_OR_PARTIAL"

    if len(set(availabilities)) > 1:
        return "AVAILABILITY_DIFFERENT"
    return "DIFFERENT_VALUE"


def _business_note(comparison_type: str) -> str:
    if comparison_type == "COMMON_SAME":
        return "样例中这些变体参数一致。"
    if comparison_type == "DIFFERENT_VALUE":
        return "同 model 不同 trim / option 之间参数不同，可用于配置差异预览。"
    if comparison_type in {"UNIQUE_TO_TRIM", "PARTIAL_AVAILABLE", "UNIQUE_OR_PARTIAL"}:
        return "部分版本具备，部分版本不配备，可作为确认配置差异。"
    if comparison_type == "OPTIONAL_DIFFERENT":
        return "标配、选装或不配备状态不同，可用于解释版本差异。"
    if comparison_type == "AVAILABILITY_DIFFERENT":
        return "配置可用性不同，可作为确认配置差异。"
    if comparison_type in {"MISSING_UNKNOWN", "MISSING_OR_UNKNOWN"}:
        return "存在空白或未知值，需要确认来源；不能直接解释成不配备。"
    if comparison_type == "NOT_APPLICABLE":
        return "样例中包含不适用配置，需按车型/法规语境解释。"
    if comparison_type == "CANCELLED_OR_REMOVED":
        return "样例中包含取消、删除或减少信息，不应直接解释为常规配置。"
    return "存在待确认配置语义。"


def _business_note_with_review_notes(comparison_type: str, review_notes: list[str] | None = None) -> str:
    note = _business_note(comparison_type)
    first_review_note = next((item.strip() for item in review_notes or [] if item.strip()), "")
    if not first_review_note:
        return note
    return f"{note} 需核对：{first_review_note}"


def _compare_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    common_same = sum(1 for row in rows if row["comparisonType"] == "COMMON_SAME")
    different = sum(1 for row in rows if row["comparisonType"] == "DIFFERENT_VALUE")
    unique = sum(1 for row in rows if row["comparisonType"] == "UNIQUE_TO_TRIM")
    partial = sum(1 for row in rows if row["comparisonType"] in {"PARTIAL_AVAILABLE", "UNIQUE_OR_PARTIAL"})
    availability_different = sum(
        1
        for row in rows
        if row["comparisonType"] in {"AVAILABILITY_DIFFERENT", "OPTIONAL_DIFFERENT", "UNIQUE_OR_PARTIAL"}
    )
    optional_different = sum(1 for row in rows if row["comparisonType"] == "OPTIONAL_DIFFERENT")
    missing = sum(1 for row in rows if row["comparisonType"] in {"MISSING_UNKNOWN", "MISSING_OR_UNKNOWN"})
    not_applicable = sum(1 for row in rows if row["comparisonType"] == "NOT_APPLICABLE")
    cancelled_or_removed = sum(1 for row in rows if row["comparisonType"] == "CANCELLED_OR_REMOVED")
    confirmed_difference_count = sum(
        1
        for row in rows
        if row["comparisonType"] not in {"COMMON_SAME", "MISSING_UNKNOWN", "MISSING_OR_UNKNOWN"}
    )
    inferred_difference_count = sum(
        1
        for row in rows
        if row["comparisonType"] not in {"COMMON_SAME", "MISSING_UNKNOWN", "MISSING_OR_UNKNOWN"}
        and any(value and value.get("inferred") for value in row.get("values", []))
    )
    raw_confirmed_difference_count = confirmed_difference_count - inferred_difference_count
    category_counts: dict[str, int] = {}
    for row in rows:
        category = str(row.get("category") or "")
        if category:
            category_counts[category] = category_counts.get(category, 0) + 1
    categories = sorted(
        {
            str(row.get("category") or "")
            for row in rows
            if row["comparisonType"] not in {"COMMON_SAME", "MISSING_UNKNOWN", "MISSING_OR_UNKNOWN"} and row.get("category")
        }
    )
    return {
        "totalFeatures": len(rows),
        "shownFeatures": len(rows),
        "commonSameCount": common_same,
        "differentValueCount": different,
        "valueDifferentCount": different,
        "availabilityDifferentCount": availability_different,
        "optionalDifferentCount": optional_different,
        "confirmedDifferenceCount": confirmed_difference_count,
        "rawConfirmedDifferenceCount": raw_confirmed_difference_count,
        "inferredDifferenceCount": inferred_difference_count,
        "uniqueFeatureCount": unique,
        "partialAvailableCount": partial,
        "uniqueOrPartialCount": unique + partial,
        "missingOrUnknownCount": missing,
        "notApplicableCount": not_applicable,
        "cancelledOrRemovedCount": cancelled_or_removed,
        "differenceCount": confirmed_difference_count,
        "categoryCounts": category_counts,
        "differenceCategories": categories,
    }


def _value_payload(
    group_id: str,
    row_index: int,
    trim_id: str,
    raw_value: str,
    *,
    source: dict[str, Any] | None = None,
    semantics_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    semantics = semantics_override or _source_value_semantics(raw_value)
    source_payload = dict(source) if source is not None else None
    if source_payload is not None and semantics.get("inferenceReason"):
        source_payload["inferenceReason"] = semantics["inferenceReason"]
        source_payload["confidence"] = semantics.get("confidence")
    return {
        "valueId": f"{group_id}:{row_index}:{trim_id}",
        "rawValue": raw_value,
        "normalizedValue": semantics["normalizedValue"],
        "availability": semantics["availability"],
        "unit": semantics["unit"],
        "valueState": semantics["valueState"],
        "displayValue": semantics.get("displayValue"),
        "inferred": bool(semantics.get("inferred")),
        "inferenceReason": semantics.get("inferenceReason"),
        "confidence": semantics.get("confidence"),
        "source": source_payload,
    }


def _finalize_compare_group(
    *,
    group_id: str,
    title: str,
    source_sheet: str,
    model_name: str,
    trims: list[dict[str, Any]],
    rows: list[dict[str, Any]],
) -> dict[str, Any] | None:
    if not trims or not rows:
        return None
    summary = _compare_summary(rows)
    return {
        "groupId": group_id,
        "title": title,
        "sourceSheet": source_sheet,
        "modelName": model_name,
        "trimCount": len(trims),
        "featureCount": len(rows),
        "differenceCount": summary["differenceCount"],
        "trims": trims,
        "rows": rows,
        "summary": summary,
    }


def _row_variant_group(
    sheet_name: str,
    frame: pd.DataFrame,
    model_name: str,
) -> dict[str, Any] | None:
    rows = _clean_frame_rows(frame)
    header_index: int | None = None
    for index, row in enumerate(rows[:30]):
        first = row[0].lower() if row else ""
        if first in ROW_VARIANT_HEADER_TOKENS and _non_empty_count(row) >= 3:
            header_index = index
            break
    if header_index is None:
        return None

    header = rows[header_index]
    parent = rows[header_index - 1] if header_index > 0 else ["" for _ in header]
    feature_names: list[tuple[int, str]] = []
    active_parent = ""
    for col_idx in range(1, len(header)):
        if col_idx < len(parent) and parent[col_idx]:
            active_parent = parent[col_idx]
        label = header[col_idx] if col_idx < len(header) else ""
        if not label:
            continue
        feature_name = _compact_name([active_parent, label])
        feature_names.append((col_idx, feature_name))
    if len(feature_names) < 2:
        return None

    trim_source_rows: list[tuple[int, str, list[str]]] = []
    for row_index in range(header_index + 1, len(rows)):
        row = rows[row_index]
        if not row or not row[0]:
            continue
        values = [row[col_idx] if col_idx < len(row) else "" for col_idx, _name in feature_names]
        if _non_empty_count(values) == 0:
            continue
        trim_source_rows.append((row_index + 1, row[0], values))
        if len(trim_source_rows) >= MAX_COMPARE_TRIMS:
            break
    if len(trim_source_rows) < 2:
        return None

    group_id = f"workbook-row-{_slug(sheet_name)}"
    trims = [
        {
            "trimId": f"{group_id}-r{excel_row}",
            "trimName": trim_name,
            "fullTrimName": trim_name,
            "modelName": model_name,
            "sourceSheet": sheet_name,
            "sourceRow": excel_row,
        }
        for excel_row, trim_name, _values in trim_source_rows
    ]

    compare_rows: list[dict[str, Any]] = []
    for feature_index, (_col_idx, feature_name) in enumerate(feature_names):
        raw_values = [values[feature_index] if feature_index < len(values) else "" for _row, _trim, values in trim_source_rows]
        comparison_type = _comparison_type(raw_values)
        trim_ids = [str(trim["trimId"]) for trim in trims]
        unique_trim_ids = [
            trim_ids[index]
            for index, value in enumerate(raw_values)
            if _source_value_is_present(value)
        ] if comparison_type in {"UNIQUE_TO_TRIM", "PARTIAL_AVAILABLE", "UNIQUE_OR_PARTIAL"} else []
        compare_rows.append(
            {
                "category": sheet_name,
                "featureCode": f"digest_{_slug(feature_name)}",
                "featureName": feature_name,
                "comparisonType": comparison_type,
                "uniqueTrimIds": unique_trim_ids,
                "businessNote": _business_note(comparison_type),
                "values": [
                    _value_payload(
                        group_id,
                        trim_source_rows[index][0],
                        trim_ids[index],
                        value,
                        source=_source_cell_payload(frame, sheet_name, trim_source_rows[index][0] - 1, feature_names[feature_index][0]),
                    )
                    for index, value in enumerate(raw_values)
                ],
            }
        )

    return _finalize_compare_group(
        group_id=group_id,
        title=f"{model_name} · {sheet_name}",
        source_sheet=sheet_name,
        model_name=model_name,
        trims=trims,
        rows=compare_rows,
    )


def _column_variant_group(
    sheet_name: str,
    frame: pd.DataFrame,
    model_name: str,
) -> dict[str, Any] | None:
    rows = _clean_frame_rows(frame)
    if frame.attrs.get("prefer_simple_columns") and rows:
        first = _normalise_space(rows[0][0]).lower() if rows[0] else ""
        second = _normalise_space(rows[0][1]).lower() if len(rows[0]) > 1 else ""
        if (
            first in SIMPLE_CATEGORY_HEADER_TOKENS
            and second in SIMPLE_FEATURE_HEADER_TOKENS
        ):
            return None
    header_index: int | None = None
    for index, row in enumerate(rows[:20]):
        first = row[0].lower() if row else ""
        if first in ROW_VARIANT_HEADER_TOKENS:
            continue
        if len(row) >= 4 and row[0] and row[1] and _non_empty_count(row[3:]) >= 1:
            header_index = index
            break
    if header_index is None:
        return None

    header = rows[header_index]
    variant_columns = [
        (col_idx, header[col_idx])
        for col_idx in range(3, len(header))
        if header[col_idx]
    ]
    if len(variant_columns) < 2:
        return None
    if all(_looks_numeric(trim_name) for _col_idx, trim_name in variant_columns):
        return None

    group_id = f"workbook-column-{_slug(sheet_name)}"
    trims = [
        {
            "trimId": f"{group_id}-c{col_idx}",
            "trimName": trim_name,
            "fullTrimName": trim_name,
            "modelName": model_name,
            "sourceSheet": sheet_name,
            "sourceColumn": col_idx + 1,
        }
        for col_idx, trim_name in variant_columns[:MAX_COMPARE_TRIMS]
    ]

    compare_rows: list[dict[str, Any]] = []
    active_category = ""
    for source_row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if len(row) > 0 and row[0]:
            active_category = row[0]
        feature_name = _compact_name([
            row[2] if len(row) > 2 else "",
            row[1] if len(row) > 1 else "",
        ])
        if not feature_name:
            continue
        raw_values = [row[col_idx] if col_idx < len(row) else "" for col_idx, _name in variant_columns[:MAX_COMPARE_TRIMS]]
        if _non_empty_count(raw_values) == 0:
            continue
        comparison_type = _comparison_type(raw_values)
        trim_ids = [str(trim["trimId"]) for trim in trims]
        compare_rows.append(
            {
                "category": active_category or sheet_name,
                "featureCode": f"digest_{_slug(feature_name)}",
                "featureName": feature_name,
                "comparisonType": comparison_type,
                "uniqueTrimIds": [
                    trim_ids[index]
                    for index, value in enumerate(raw_values)
                    if _source_value_is_present(value)
                ] if comparison_type in {"UNIQUE_TO_TRIM", "PARTIAL_AVAILABLE", "UNIQUE_OR_PARTIAL"} else [],
                "businessNote": _business_note(comparison_type),
                "values": [
                    _value_payload(
                        group_id,
                        source_row_index,
                        trim_ids[index],
                        value,
                        source=_source_cell_payload(frame, sheet_name, source_row_index - 1, variant_columns[index][0]),
                    )
                    for index, value in enumerate(raw_values)
                ],
            }
        )
    return _finalize_compare_group(
        group_id=group_id,
        title=f"{model_name} · {sheet_name}",
        source_sheet=sheet_name,
        model_name=model_name,
        trims=trims,
        rows=compare_rows,
    )


def _simple_column_variant_group(
    sheet_name: str,
    frame: pd.DataFrame,
    model_name: str,
) -> dict[str, Any] | None:
    rows = _clean_frame_rows(frame)
    header_index: int | None = None
    category_col: int | None = None
    feature_col = 0
    trim_start_col = 1
    for index, row in enumerate(rows[:20]):
        if _non_empty_count(row) < 3:
            continue
        first = _normalise_space(row[0]).lower() if row else ""
        second = _normalise_space(row[1]).lower() if len(row) > 1 else ""
        if first in ROW_VARIANT_HEADER_TOKENS:
            continue
        if first in SIMPLE_CATEGORY_HEADER_TOKENS and second in SIMPLE_FEATURE_HEADER_TOKENS and _non_empty_count(row[2:]) >= 2:
            header_index = index
            category_col = 0
            feature_col = 1
            trim_start_col = 2
            break
        if first in SIMPLE_FEATURE_HEADER_TOKENS and _non_empty_count(row[1:]) >= 2:
            header_index = index
            category_col = None
            feature_col = 0
            trim_start_col = 1
            break
    if header_index is None:
        return None

    header = rows[header_index]
    variant_columns = [
        (col_idx, header[col_idx])
        for col_idx in range(trim_start_col, len(header))
        if header[col_idx]
    ][:MAX_COMPARE_TRIMS]
    if len(variant_columns) < 2:
        return None

    group_id = f"tabular-simple-{_slug(sheet_name)}"
    trims = [
        {
            "trimId": f"{group_id}-c{col_idx}",
            "trimName": trim_name,
            "fullTrimName": trim_name,
            "modelName": model_name,
            "sourceSheet": sheet_name,
            "sourceColumn": col_idx + 1,
        }
        for col_idx, trim_name in variant_columns
    ]

    compare_rows: list[dict[str, Any]] = []
    active_category = sheet_name
    for source_row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if category_col is not None and category_col < len(row) and row[category_col]:
            active_category = row[category_col]
        feature_name = row[feature_col] if feature_col < len(row) else ""
        feature_name = _normalise_space(feature_name)
        if not feature_name:
            continue
        raw_values = [row[col_idx] if col_idx < len(row) else "" for col_idx, _name in variant_columns]
        if _non_empty_count(raw_values) == 0:
            continue
        comparison_type = _comparison_type(raw_values)
        trim_ids = [str(trim["trimId"]) for trim in trims]
        feature_key = "|".join([sheet_name, active_category or sheet_name, feature_name])
        compare_rows.append(
            {
                "category": active_category or sheet_name,
                "featureKey": feature_key,
                "featureCode": _feature_code(feature_key),
                "featureName": feature_name,
                "comparisonType": comparison_type,
                "uniqueTrimIds": [
                    trim_ids[index]
                    for index, value in enumerate(raw_values)
                    if _source_value_is_present(value)
                ] if comparison_type in {"UNIQUE_TO_TRIM", "PARTIAL_AVAILABLE", "UNIQUE_OR_PARTIAL"} else [],
                "businessNote": _business_note(comparison_type),
                "values": [
                    _value_payload(
                        group_id,
                        source_row_index,
                        trim_ids[index],
                        value,
                        source=_source_cell_payload(frame, sheet_name, source_row_index - 1, variant_columns[index][0]),
                    )
                    for index, value in enumerate(raw_values)
                ],
            }
        )

    return _finalize_compare_group(
        group_id=group_id,
        title=f"{model_name} · {sheet_name}",
        source_sheet=sheet_name,
        model_name=model_name,
        trims=trims,
        rows=compare_rows,
    )


def _ocr_headerless_value_quality(value: str) -> int:
    cleaned = _normalise_space(value)
    if not cleaned:
        return 1
    semantics = _source_value_semantics(cleaned)
    if semantics["availability"] in {"STANDARD", "OPTIONAL", "NOT_AVAILABLE", "NOT_APPLICABLE"}:
        return 6
    if semantics["availability"] == "VALUE" and semantics["valueState"] == "numeric_value":
        return 5
    if _looks_numeric_value(cleaned):
        return 5
    if re.fullmatch(r"[A-Z0-9+./-]{2,16}", cleaned):
        return 4
    if len(cleaned) <= 48:
        return 3
    if len(cleaned) <= 96:
        return 2
    return 0


def _trim_trailing_empty_cells(row: list[str]) -> list[str]:
    trimmed = list(row)
    while trimmed and not trimmed[-1]:
        trimmed.pop()
    return trimmed


def _ocr_headerless_feature_quality(value: str) -> int:
    cleaned = _normalise_space(value)
    if not cleaned:
        return 0
    lowered = cleaned.lower()
    if lowered in SIMPLE_CATEGORY_HEADER_TOKENS or lowered in SIMPLE_FEATURE_HEADER_TOKENS:
        return 0
    if lowered in ROW_VARIANT_HEADER_TOKENS:
        return 0
    value_quality = _ocr_headerless_value_quality(cleaned)
    if value_quality >= 5:
        return 0
    if re.fullmatch(r"[A-Z0-9+./-]{2,12}", cleaned):
        return 0
    if len(cleaned) < 3:
        return 0
    if len(cleaned) > 140:
        return 1
    return 4 if len(cleaned) <= 80 else 2


def _ocr_headerless_value_count(rows: list[list[str]], value_start_col: int) -> int | None:
    counts: dict[int, int] = {}
    for row in rows:
        if len(row) <= value_start_col:
            continue
        count = min(len(row) - value_start_col, MAX_COMPARE_TRIMS)
        if count < 2:
            continue
        counts[count] = counts.get(count, 0) + 1
    if not counts:
        return None
    return max(counts, key=lambda count: (counts[count], count))


def _ocr_headerless_row_review(feature_name: str, raw_values: list[str]) -> tuple[list[str], list[str]]:
    flags: list[str] = []
    notes: list[str] = []
    normalized_values = [_normalise_space(value) for value in raw_values]
    value_qualities = [_ocr_headerless_value_quality(value) for value in normalized_values]
    structured_value_count = sum(1 for quality in value_qualities if quality >= 4)
    feature_like_values = [
        value
        for value in normalized_values
        if value and _ocr_headerless_feature_quality(value) >= 4
    ]
    if feature_like_values and structured_value_count > 0:
        flags.append("ocr_possible_feature_text_in_value_cell")
        sample = ", ".join(feature_like_values[:2])
        notes.append(f"OCR 值单元格像配置项文本（{sample}），可能是特征名换行或单位被切入值列。")
    if _ocr_headerless_feature_quality(feature_name) <= 2:
        flags.append("ocr_low_feature_label_confidence")
        notes.append("OCR 配置项名称置信度偏低，建议在在线表格中核对原图。")
    if any(len(value) > 96 for value in normalized_values):
        flags.append("ocr_long_value_cell")
        notes.append("OCR 值单元格过长，可能混入备注或相邻列文本。")
    return flags, notes


def _ocr_headerless_attempt(
    *,
    sheet_name: str,
    rows: list[list[str]],
    category_col: int | None,
    feature_col: int,
    value_start_col: int,
) -> dict[str, Any] | None:
    value_count = _ocr_headerless_value_count(rows, value_start_col)
    if value_count is None:
        return None

    kept_rows: list[dict[str, Any]] = []
    structured_value_count = 0
    total_value_count = 0
    value_quality_total = 0
    feature_quality_total = 0
    feature_value_like_penalty = 0
    long_value_penalty = 0

    for source_row_index, row in enumerate(rows, start=1):
        if len(row) <= feature_col:
            continue
        feature_name = _normalise_space(row[feature_col])
        feature_quality = _ocr_headerless_feature_quality(feature_name)
        if feature_quality <= 0:
            continue
        raw_values = [
            row[col_idx] if col_idx < len(row) else ""
            for col_idx in range(value_start_col, value_start_col + value_count)
        ]
        if _non_empty_count(raw_values) == 0:
            continue
        value_qualities = [_ocr_headerless_value_quality(value) for value in raw_values]
        structured_value_count += sum(1 for quality in value_qualities if quality >= 4)
        total_value_count += len(value_qualities)
        value_quality_total += sum(value_qualities)
        feature_quality_total += feature_quality
        if _ocr_headerless_value_quality(feature_name) >= 4:
            feature_value_like_penalty += 1
        long_value_penalty += sum(1 for value in raw_values if len(_normalise_space(value)) > 120)
        review_flags, review_notes = _ocr_headerless_row_review(feature_name, raw_values)
        category = sheet_name
        if category_col is not None and category_col < len(row) and row[category_col]:
            category = _normalise_space(row[category_col])
        kept_row = {
            "sourceRowIndex": source_row_index,
            "category": category or sheet_name,
            "featureName": feature_name,
            "rawValues": raw_values,
        }
        if review_flags:
            kept_row["reviewFlags"] = review_flags
            kept_row["reviewNotes"] = review_notes
        kept_rows.append(kept_row)

    if len(kept_rows) < OCR_HEADERLESS_MIN_ROWS or total_value_count == 0:
        return None

    structured_ratio = round(structured_value_count * 1000 / total_value_count)
    if structured_value_count < 2 and structured_ratio < 350:
        return None
    if long_value_penalty > len(kept_rows):
        return None

    unique_features = len({str(row["featureName"]) for row in kept_rows})
    score = (
        len(kept_rows),
        structured_ratio,
        structured_value_count,
        value_quality_total,
        feature_quality_total,
        unique_features,
        -feature_value_like_penalty,
        -long_value_penalty,
        value_count,
    )
    return {
        "score": score,
        "valueCount": value_count,
        "valueStartCol": value_start_col,
        "rows": kept_rows,
    }


def _ocr_headerless_column_group(
    sheet_name: str,
    frame: pd.DataFrame,
    model_name: str,
) -> dict[str, Any] | None:
    rows = [
        trimmed
        for row in _clean_frame_rows(frame)
        for trimmed in [_trim_trailing_empty_cells(row)]
        if _non_empty_count(trimmed) >= 3
    ]
    if len(rows) < OCR_HEADERLESS_MIN_ROWS:
        return None

    attempts = [
        _ocr_headerless_attempt(
            sheet_name=sheet_name,
            rows=rows,
            category_col=None,
            feature_col=0,
            value_start_col=1,
        ),
        _ocr_headerless_attempt(
            sheet_name=sheet_name,
            rows=rows,
            category_col=0,
            feature_col=1,
            value_start_col=2,
        ),
    ]
    candidates = [attempt for attempt in attempts if attempt is not None]
    if not candidates:
        return None

    candidate = max(candidates, key=lambda attempt: attempt["score"])
    value_count = int(candidate["valueCount"])
    value_start_col = int(candidate["valueStartCol"])
    group_id = f"ocr-headerless-{_slug(sheet_name)}"
    trims = [
        {
            "trimId": f"{group_id}-c{index + 1}",
            "trimName": f"OCR Column {index + 1}",
            "fullTrimName": f"OCR Column {index + 1} · 待补配置列身份",
            "modelName": model_name,
            "sourceSheet": sheet_name,
            "sourceColumn": value_start_col + index + 1,
            "identityStatus": "temporary_ocr_column",
            "identityNote": "OCR 未识别到配置列标题，当前列名为临时身份。",
        }
        for index in range(value_count)
    ]
    trim_ids = [str(trim["trimId"]) for trim in trims]

    compare_rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(candidate["rows"], start=1):
        raw_values = [str(value) for value in row["rawValues"]]
        comparison_type = _comparison_type(raw_values)
        feature_name = str(row["featureName"])
        category = str(row["category"] or sheet_name)
        feature_key = "|".join([sheet_name, category, feature_name])
        compare_rows.append(
            {
                "category": category,
                "featureKey": feature_key,
                "featureCode": _feature_code(feature_key),
                "featureName": feature_name,
                "comparisonType": comparison_type,
                "uniqueTrimIds": [
                    trim_ids[index]
                    for index, value in enumerate(raw_values)
                    if _source_value_is_present(value)
                ] if comparison_type in {"UNIQUE_TO_TRIM", "PARTIAL_AVAILABLE", "UNIQUE_OR_PARTIAL"} else [],
                "businessNote": (
                    f"{_business_note_with_review_notes(comparison_type, row.get('reviewNotes'))} "
                    "OCR 未识别到配置列标题，已按列位置生成临时配置列；请在 FloatingDeck 补齐配置列身份后再引用。"
                ),
                "values": [
                    _value_payload(
                        group_id,
                        row_index,
                        trim_ids[index],
                        value,
                        source=_source_cell_payload(
                            frame,
                            sheet_name,
                            int(row["sourceRowIndex"]) - 1,
                            value_start_col + index,
                        ),
                    )
                    for index, value in enumerate(raw_values)
                ],
                **(
                    {
                        "reviewFlags": row["reviewFlags"],
                        "reviewNotes": row["reviewNotes"],
                    }
                    if row.get("reviewFlags")
                    else {}
                ),
            }
        )

    group = _finalize_compare_group(
        group_id=group_id,
        title=f"{model_name} · {sheet_name} · OCR 临时列",
        source_sheet=sheet_name,
        model_name=model_name,
        trims=trims,
        rows=compare_rows,
    )
    if group is not None:
        group["sourceKind"] = "ocr_headerless"
        group["identityStatus"] = "temporary_ocr_column"
        group["identityNote"] = "OCR 未识别到配置列标题，已按列位置生成临时配置列。"
    return group


def _normalized_header_token(value: str) -> str:
    return normalize_header_phrase(value)


def _header_contains_any(value: str, tokens: set[str] | tuple[str, ...]) -> bool:
    normalized = _normalized_header_token(value)
    if not normalized:
        return False
    return any(token.lower() in normalized for token in tokens)


def _price_list_header_kind(value: str) -> str | None:
    if _header_contains_any(value, PRICE_LIST_VALUE_HEADER_TOKENS):
        return "value"
    if _header_contains_any(value, PRICE_LIST_IDENTITY_HEADER_TOKENS):
        return "identity"
    return None


def _price_list_field_for_header(value: str) -> str | None:
    normalized = _normalized_header_token(value)
    for field, aliases in PRICE_LIST_IDENTITY_FIELD_ALIASES.items():
        if any(normalized == alias.lower() for alias in aliases):
            return field
    candidates: list[tuple[str, str]] = []
    for field, aliases in PRICE_LIST_IDENTITY_FIELD_ALIASES.items():
        candidates.extend((field, alias.lower()) for alias in aliases)
    for field, alias in sorted(candidates, key=lambda item: len(item[1]), reverse=True):
        if alias in normalized:
            return field
    return None


def _price_list_header_index(rows: list[list[str]]) -> int | None:
    for row_index, row in enumerate(rows[:20]):
        if _non_empty_count(row) < 4:
            continue
        kinds = [_price_list_header_kind(value) for value in row]
        identity_count = sum(1 for kind in kinds if kind == "identity")
        value_count = sum(1 for kind in kinds if kind == "value")
        has_model = any(_price_list_field_for_header(value) == "model" for value in row)
        has_trim = any(_price_list_field_for_header(value) == "trim" for value in row)
        if identity_count >= 2 and value_count >= 1 and (has_model or has_trim):
            return row_index
    return None


def _price_list_identity_columns(header: list[str]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for column_index, label in enumerate(header):
        field = _price_list_field_for_header(label)
        if field and field not in columns:
            columns[field] = column_index
    return columns


def _price_list_value_columns(header: list[str], identity_columns: dict[str, int]) -> list[tuple[int, str]]:
    identity_column_indexes = set(identity_columns.values())
    value_columns: list[tuple[int, str]] = []
    for column_index, label in enumerate(header):
        if column_index in identity_column_indexes:
            continue
        cleaned = _normalise_space(label)
        if not cleaned:
            continue
        if _header_contains_any(cleaned, PRICE_LIST_VALUE_HEADER_TOKENS):
            value_columns.append((column_index, cleaned))
    return value_columns


def _price_list_row_value(row: list[str], column_index: int | None) -> str:
    if column_index is None or column_index >= len(row):
        return ""
    return _normalise_space(row[column_index])


PriceListGroupKey = tuple[str, str, str, str, str]


def _price_list_group_key(
    row: list[str],
    identity_columns: dict[str, int],
    *,
    fallback_model_name: str,
) -> PriceListGroupKey:
    brand = _price_list_row_value(row, identity_columns.get("brand"))
    model = _price_list_row_value(row, identity_columns.get("model")) or fallback_model_name
    market = _price_list_row_value(row, identity_columns.get("market"))
    model_year = _price_list_row_value(row, identity_columns.get("modelYear"))
    powertrain = _price_list_row_value(row, identity_columns.get("powertrain"))
    return brand, model, market, model_year, powertrain


def _price_list_trim_payload(
    *,
    group_id: str,
    row: list[str],
    source_row_number: int,
    identity_columns: dict[str, int],
    fallback_model_name: str,
) -> dict[str, Any]:
    brand = _price_list_row_value(row, identity_columns.get("brand"))
    model = _price_list_row_value(row, identity_columns.get("model")) or fallback_model_name
    trim_name = (
        _price_list_row_value(row, identity_columns.get("trim"))
        or _price_list_row_value(row, identity_columns.get("salesVersion"))
        or f"Row {source_row_number}"
    )
    market = _price_list_row_value(row, identity_columns.get("market"))
    model_year = _price_list_row_value(row, identity_columns.get("modelYear"))
    powertrain = _price_list_row_value(row, identity_columns.get("powertrain"))
    material_no = _price_list_row_value(row, identity_columns.get("materialNo"))
    sales_version = _price_list_row_value(row, identity_columns.get("salesVersion"))
    full_trim_name = _compact_name([brand, model, trim_name])
    profile = {
        "brand": brand,
        "modelYear": model_year,
        "country": market,
        "powertrain": powertrain,
        "materialNo": material_no,
        "configurationVersion": sales_version or trim_name,
    }
    return {
        "trimId": f"{group_id}-r{source_row_number}",
        "trimName": trim_name,
        "fullTrimName": full_trim_name or trim_name,
        "modelName": model,
        "market": market,
        "country": market,
        "modelYear": model_year or None,
        "energyType": powertrain or None,
        "materialNo": material_no or None,
        "salesVersion": sales_version or trim_name,
        "hasMaterialNo": bool(material_no),
        "dataOrigin": "own_catalog" if material_no else "external_or_scraped",
        "sourceRow": source_row_number,
        "sourceStatus": "active",
        "profile": {key: value for key, value in profile.items() if value},
    }


def _price_list_groups(
    sheet_name: str,
    frame: pd.DataFrame,
    model_name: str,
) -> list[dict[str, Any]]:
    rows = _clean_frame_rows(frame)
    header_index = _price_list_header_index(rows)
    if header_index is None:
        return []
    header = rows[header_index]
    identity_columns = _price_list_identity_columns(header)
    value_columns = _price_list_value_columns(header, identity_columns)
    if not value_columns:
        return []

    row_groups: dict[PriceListGroupKey, list[tuple[int, list[str]]]] = {}
    for row_index in range(header_index + 1, len(rows)):
        row = rows[row_index]
        if _non_empty_count(row) < 2:
            continue
        if not any(_price_list_row_value(row, column_index) for column_index, _label in value_columns):
            continue
        trim_name = _price_list_row_value(row, identity_columns.get("trim"))
        model_value = _price_list_row_value(row, identity_columns.get("model")) or model_name
        if not (trim_name or model_value):
            continue
        key = _price_list_group_key(row, identity_columns, fallback_model_name=model_name)
        row_groups.setdefault(key, []).append((row_index, row))

    groups: list[dict[str, Any]] = []
    for group_number, (key, grouped_rows) in enumerate(row_groups.items(), start=1):
        if len(grouped_rows) < 2:
            continue
        brand, model, market, model_year, powertrain = key
        group_id = f"price-list-{_slug(sheet_name)}-{group_number}"
        limited_rows = grouped_rows[:MAX_COMPARE_TRIMS]
        trims = [
            _price_list_trim_payload(
                group_id=group_id,
                row=row,
                source_row_number=row_index + 1,
                identity_columns=identity_columns,
                fallback_model_name=model_name,
            )
            for row_index, row in limited_rows
        ]
        trim_ids = [str(trim["trimId"]) for trim in trims]
        compare_rows: list[dict[str, Any]] = []
        for column_index, feature_name in value_columns:
            raw_values = [
                row[column_index] if column_index < len(row) else ""
                for _row_index, row in limited_rows
            ]
            if _non_empty_count(raw_values) == 0:
                continue
            comparison_type = _comparison_type(raw_values)
            feature_key = "|".join([sheet_name, "Pricing", model, market, model_year, powertrain, feature_name])
            compare_rows.append(
                {
                    "category": "价格 Pricing",
                    "featureKey": feature_key,
                    "featureCode": _feature_code(feature_key),
                    "featureName": feature_name,
                    "comparisonType": comparison_type,
                    "uniqueTrimIds": [
                        trim_ids[index]
                        for index, value in enumerate(raw_values)
                        if _source_value_is_present(value)
                    ] if comparison_type in {"UNIQUE_TO_TRIM", "PARTIAL_AVAILABLE", "UNIQUE_OR_PARTIAL"} else [],
                    "businessNote": "价格单字段，可与配置项一起进入在线编辑、导出和来源证据核对。",
                    "values": [
                        _value_payload(
                            group_id,
                            limited_rows[index][0] + 1,
                            trim_ids[index],
                            value,
                            source=_source_cell_payload(frame, sheet_name, limited_rows[index][0], column_index),
                        )
                        for index, value in enumerate(raw_values)
                    ],
                }
            )
        title = _compact_name([brand, model, market, model_year, powertrain, "价格单"]) or f"{model_name} · 价格单"
        group = _finalize_compare_group(
            group_id=group_id,
            title=title,
            source_sheet=sheet_name,
            model_name=model or model_name,
            trims=trims,
            rows=compare_rows,
        )
        if group is not None:
            group["sourceKind"] = "price_list"
            groups.append(group)
    return groups


def build_workbook_digest_from_frames(
    sheets: list[tuple[str, pd.DataFrame]],
    *,
    file_name: str,
    digest_type: str = "workbook",
) -> dict[str, Any]:
    if _is_eu_config_workbook(sheets):
        return _eu_config_workbook_digest(sheets, file_name=file_name)

    model_name = _infer_model_name(file_name, sheets)
    sheet_payloads = [_sheet_digest(sheet_name, frame) for sheet_name, frame in sheets[:MAX_SHEETS]]
    compare_groups: list[dict[str, Any]] = []
    for sheet_name, frame in sheets[:MAX_SHEETS]:
        price_groups = _price_list_groups(sheet_name, frame, model_name)
        if price_groups:
            compare_groups.extend(price_groups)
            continue
        sheet_group_count = len(compare_groups)
        for group in (
            _row_variant_group(sheet_name, frame, model_name),
            _column_variant_group(sheet_name, frame, model_name),
            _simple_column_variant_group(sheet_name, frame, model_name),
        ):
            if group is not None:
                compare_groups.append(group)
        if digest_type in OCR_HEADERLESS_DIGEST_TYPES and len(compare_groups) == sheet_group_count:
            group = _ocr_headerless_column_group(sheet_name, frame, model_name)
            if group is not None:
                compare_groups.append(group)

    return {
        "digestType": digest_type,
        "status": "ready",
        "fileName": file_name,
        "modelName": model_name,
        "summary": {
            "sheetCount": len(sheets),
            "tableCount": len(sheet_payloads),
            "candidateTrimCount": sum(group["trimCount"] for group in compare_groups),
            "comparableGroupCount": len([group for group in compare_groups if group["trimCount"] >= 2]),
            "featureCount": sum(group["featureCount"] for group in compare_groups),
            "differenceCount": sum(group["differenceCount"] for group in compare_groups),
        },
        "sheets": sheet_payloads,
        "compareGroups": compare_groups,
    }


def _read_xlsx_workbook(file_path: Path) -> list[tuple[str, pd.DataFrame]]:
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(file_path, data_only=True, read_only=False)
    try:
        sheets: list[tuple[str, pd.DataFrame]] = []
        for worksheet in workbook.worksheets[:MAX_SHEETS]:
            max_row = worksheet.max_row or 0
            max_column = worksheet.max_column or 0
            values = [
                [
                    worksheet.cell(row=row_index, column=column_index).value
                    for column_index in range(1, max_column + 1)
                ]
                for row_index in range(1, max_row + 1)
            ]
            source_cells: dict[tuple[int, int], dict[str, Any]] = {}
            for row_index in range(1, max_row + 1):
                for column_index in range(1, max_column + 1):
                    column_letter = _excel_column_letter(column_index)
                    cell_ref = f"{column_letter}{row_index}"
                    source_cells[(row_index - 1, column_index - 1)] = {
                        "sheetName": worksheet.title,
                        "rowNumber": row_index,
                        "columnNumber": column_index,
                        "columnLetter": column_letter,
                        "cell": cell_ref,
                        "sourceCell": cell_ref,
                        "mergedRange": None,
                    }
            for merged_range in worksheet.merged_cells.ranges:
                merged_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
                if merged_value is None:
                    continue
                source_column_letter = _excel_column_letter(merged_range.min_col)
                source_cell = f"{source_column_letter}{merged_range.min_row}"
                merged_range_label = str(merged_range)
                for row_index in range(merged_range.min_row, merged_range.max_row + 1):
                    while len(values) < row_index:
                        values.append([None] * max_column)
                    row_values = values[row_index - 1]
                    if len(row_values) < merged_range.max_col:
                        row_values.extend([None] * (merged_range.max_col - len(row_values)))
                    for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                        cell_value = row_values[column_index - 1]
                        if cell_value is None or cell_value == "":
                            row_values[column_index - 1] = merged_value
                        column_letter = _excel_column_letter(column_index)
                        cell_ref = f"{column_letter}{row_index}"
                        source_cells[(row_index - 1, column_index - 1)] = {
                            "sheetName": worksheet.title,
                            "rowNumber": row_index,
                            "columnNumber": column_index,
                            "columnLetter": column_letter,
                            "cell": cell_ref,
                            "sourceCell": source_cell,
                            "mergedRange": merged_range_label,
                        }
            frame = pd.DataFrame(values)
            frame.attrs["source_cells"] = source_cells
            sheets.append((worksheet.title, frame))
        return sheets
    finally:
        workbook.close()


def _read_workbook(file_path: Path) -> list[tuple[str, pd.DataFrame]]:
    if file_path.suffix.lower() in OPENPYXL_WORKBOOK_EXTENSIONS:
        return _read_xlsx_workbook(file_path)

    try:
        workbook = pd.ExcelFile(file_path, engine="calamine")
        engine: str | None = "calamine"
    except Exception:
        workbook = pd.ExcelFile(file_path)
        engine = None
    result: list[tuple[str, pd.DataFrame]] = []
    for sheet_name in workbook.sheet_names[:MAX_SHEETS]:
        if engine:
            frame = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=None,
                engine=engine,
            )
        else:
            frame = workbook.parse(sheet_name=sheet_name, header=None)
        result.append((sheet_name, frame))
    return result


def _read_tabular(file_path: Path, file_name: str) -> dict[str, Any]:
    extension = Path(file_name).suffix.lower()
    if extension in {".html", ".htm"}:
        html_tables = pd.read_html(file_path, header=None)
        sheets = [
            (f"HTML Table {index + 1}", _normalise_tabular_frame(frame))
            for index, frame in enumerate(html_tables[:MAX_SHEETS])
        ]
    else:
        separator = "\t" if extension == ".tsv" else ","
        frame = pd.read_csv(
            file_path,
            sep=separator,
            header=None,
            dtype=str,
            keep_default_na=False,
            nrows=TABULAR_PREVIEW_ROWS,
        )
        sheets = [(Path(file_name).stem or "CSV", frame)]
    return build_workbook_digest_from_frames(sheets, file_name=file_name, digest_type="tabular")


def _split_pdf_table_line(line: str) -> list[str] | None:
    for delimiter in ("\t", "|", ",", ";"):
        if delimiter not in line:
            continue
        parts = [_normalise_space(part) for part in line.split(delimiter)]
        while parts and not parts[0]:
            parts.pop(0)
        while parts and not parts[-1]:
            parts.pop()
        if len([part for part in parts if part]) >= 3:
            return parts

    parts = [part.strip() for part in re.split(r"\s{2,}", line.strip()) if part.strip()]
    if len(parts) >= 3:
        return [_normalise_space(part) for part in parts]
    return None


def _pdf_table_source_payload(
    *,
    page_number: int,
    table_number: int,
    row_number: int,
    column_number: int,
) -> dict[str, Any]:
    source_cell = (
        f"P{page_number}T{table_number}R{row_number}C{column_number}"
    )
    return {
        "sheetName": f"PDF Page {page_number}",
        "pageNumber": page_number,
        "tableNumber": table_number,
        "rowNumber": row_number,
        "columnNumber": column_number,
        "columnLetter": _excel_column_letter(column_number),
        "cell": source_cell,
        "sourceCell": source_cell,
        "mergedRange": None,
        "sourceType": "pdf_text",
    }


def _pdf_table_price_row(
    row: list[str],
    *,
    carried_motor: str,
) -> tuple[str, str, list[str]] | None:
    if len(row) < 7:
        return None
    motor = row[0] or carried_motor
    identity = [motor, row[1], row[2], row[3]]
    values = row[4:]
    if not motor or _non_empty_count(values) == 0:
        return None
    feature_name = _compact_name(
        ["MSRP", *identity]
    )
    return "Prices", feature_name, values


_PDF_PRICE_VALUE_PATTERN = re.compile(
    r"^(?:[–—-]|\d{1,3}(?:[’'.,]\d{3})+(?:\.-)?)$"
)


def _pdf_text_price_standard_rows(
    page_text: str,
    *,
    variants: tuple[str, ...],
    page_number: int,
    table_number: int,
) -> list[tuple[list[str], list[dict[str, Any]]]]:
    rows: list[tuple[list[str], list[dict[str, Any]]]] = []
    for line_number, raw_line in enumerate(page_text.splitlines(), start=1):
        tokens = raw_line.split()
        value_indexes = [
            index
            for index, token in enumerate(tokens)
            if _PDF_PRICE_VALUE_PATTERN.fullmatch(token.strip())
        ]
        if len(value_indexes) < len(variants):
            continue
        selected_indexes = value_indexes[-len(variants) :]
        if selected_indexes != list(
            range(selected_indexes[0], selected_indexes[0] + len(variants))
        ):
            continue
        identity_tokens = tokens[: selected_indexes[0]]
        identity_text = _normalise_space(" ".join(identity_tokens))
        normalized_identity = identity_text.casefold()
        if not re.search(r"\b\d{2,3}/\d{2,3}\b", identity_text):
            continue
        if not any(
            token in normalized_identity
            for token in ("2wd", "4×4", "4x4", "awd", "fwd")
        ):
            continue
        if not any(
            token in normalized_identity
            for token in ("automat", "automatic", "manuell", "manual")
        ):
            continue
        values = [tokens[index] for index in selected_indexes]
        sources = [
            _pdf_table_source_payload(
                page_number=page_number,
                table_number=table_number,
                row_number=line_number,
                column_number=5 + index,
            )
            for index in range(len(variants))
        ]
        rows.append((["Prices", f"MSRP / {identity_text}", *values], sources))
    return rows


def _pdf_standard_rows_value_count(
    rows: list[tuple[list[str], list[dict[str, Any]]]],
) -> int:
    return sum(_non_empty_count(row[2:]) for row, _sources in rows)


def _pdf_table_standard_rows(
    table: list[list[object] | None],
    *,
    page_number: int,
    table_number: int,
    page_text: str = "",
) -> tuple[tuple[str, ...], list[tuple[list[str], list[dict[str, Any]]]]] | None:
    cleaned_rows = [
        [_clean_cell(value) for value in (row or [])]
        for row in table
        if row
    ]
    if not cleaned_rows:
        return None
    header = cleaned_rows[0]
    if len(header) < 3:
        return None

    price_layout = (
        len(header) >= 7
        and _normalise_space(header[0]).lower() in {"motor", "engine"}
        and _normalise_space(header[2]).lower() in {"antrieb", "drivetrain", "drive"}
        and _normalise_space(header[3]).lower() in {"getriebe", "transmission"}
    )
    trim_start = 4 if price_layout else 1
    variants = tuple(
        _normalise_space(value)
        for value in header[trim_start:]
        if _normalise_space(value)
    )
    if len(variants) < 2:
        return None
    if not any(any(character.isalpha() for character in variant) for variant in variants):
        return None

    standard_rows: list[tuple[list[str], list[dict[str, Any]]]] = []
    active_category = _normalise_space(header[0]) or f"PDF Page {page_number}"
    carried_motor = ""
    for source_row_number, row in enumerate(cleaned_rows[1:], start=2):
        if price_layout:
            if row and row[0]:
                carried_motor = row[0]
            parsed = _pdf_table_price_row(row, carried_motor=carried_motor)
            if parsed is None:
                continue
            category, feature_name, values = parsed
            source_value_start = 5
        else:
            padded = [*row, *([""] * max(0, len(header) - len(row)))]
            if tuple(_normalise_space(value) for value in padded[1 : 1 + len(variants)]) == variants:
                active_category = _normalise_space(padded[0]) or active_category
                continue
            feature_name = _normalise_space(padded[0])
            values = padded[1 : 1 + len(variants)]
            category = active_category
            source_value_start = 2
            if not feature_name or _non_empty_count(values) == 0:
                continue

        normalized_values = [
            _normalise_space(value)
            for value in values[: len(variants)]
        ]
        if len(normalized_values) < len(variants):
            normalized_values.extend([""] * (len(variants) - len(normalized_values)))
        source_payloads = [
            _pdf_table_source_payload(
                page_number=page_number,
                table_number=table_number,
                row_number=source_row_number,
                column_number=source_value_start + index,
            )
            for index in range(len(variants))
        ]
        standard_rows.append(
            (
                [category, feature_name, *normalized_values],
                source_payloads,
            )
        )
    if price_layout and page_text.strip():
        text_rows = _pdf_text_price_standard_rows(
            page_text,
            variants=variants,
            page_number=page_number,
            table_number=table_number,
        )
        if _pdf_standard_rows_value_count(text_rows) > _pdf_standard_rows_value_count(
            standard_rows
        ):
            standard_rows = text_rows
    if not standard_rows:
        return None
    return variants, standard_rows


def _pdfplumber_table_frames(file_path: Path) -> list[tuple[str, pd.DataFrame]]:
    import pdfplumber

    grouped_rows: dict[
        tuple[str, ...],
        list[tuple[list[str], list[dict[str, Any]]]],
    ] = {}
    grouped_pages: dict[tuple[str, ...], set[int]] = {}
    with pdfplumber.open(file_path) as document:
        for page_number, page in enumerate(document.pages[:MAX_SHEETS], start=1):
            page_text = page.extract_text() or ""
            for table_number, table in enumerate(page.extract_tables(), start=1):
                parsed = _pdf_table_standard_rows(
                    table,
                    page_number=page_number,
                    table_number=table_number,
                    page_text=page_text,
                )
                if parsed is None:
                    continue
                variants, rows = parsed
                grouped_rows.setdefault(variants, []).extend(rows)
                grouped_pages.setdefault(variants, set()).add(page_number)

    sheets: list[tuple[str, pd.DataFrame]] = []
    for group_index, (variants, rows) in enumerate(grouped_rows.items(), start=1):
        pages = sorted(grouped_pages[variants])
        page_label = (
            f"{pages[0]}"
            if len(pages) == 1
            else f"{pages[0]}-{pages[-1]}"
        )
        sheet_name = f"PDF Pages {page_label} Table Group {group_index}"
        frame_rows = [["category", "feature", *variants]]
        source_cells: dict[tuple[int, int], dict[str, Any]] = {}
        for row_index, (row, sources) in enumerate(rows, start=1):
            frame_rows.append(row)
            for source_index, source in enumerate(sources, start=2):
                source_cells[(row_index, source_index)] = source
        frame = pd.DataFrame(frame_rows)
        frame.attrs["source_cells"] = source_cells
        frame.attrs["prefer_simple_columns"] = True
        sheets.append((sheet_name, frame))
    return sheets


def _pending_text_digest(file_name: str, digest_type: str, message: str) -> dict[str, Any]:
    digest = {
        "digestType": digest_type,
        "status": "pending",
        "fileName": file_name,
        "modelName": Path(file_name).stem,
        "summary": {
            "sheetCount": 0,
            "tableCount": 0,
            "candidateTrimCount": 0,
            "comparableGroupCount": 0,
            "featureCount": 0,
            "differenceCount": 0,
        },
        "sheets": [],
        "compareGroups": [],
        "message": message,
    }
    if digest_type in {"pdf_text", "pdf_ocr", "image_ocr", "tabular"}:
        digest["sourceFormat"] = digest_type
    return digest


def _mark_pending_without_compare_groups(digest: dict[str, Any], message: str) -> None:
    if digest.get("compareGroups"):
        return
    digest["status"] = "pending"
    digest["message"] = message


def _text_table_frame_from_lines(
    lines: list[str],
    *,
    sheet_name: str,
    source_type: str,
    page_number: int | None = None,
    ocr_engine: str | None = None,
) -> pd.DataFrame | None:
    rows: list[list[str]] = []
    source_cells: dict[tuple[int, int], dict[str, Any]] = {}
    for text_line_index, line in enumerate(lines[:TABULAR_PREVIEW_ROWS]):
        columns = _split_pdf_table_line(line)
        if columns is None:
            continue
        row_index = len(rows)
        rows.append(columns)
        source_line_number = text_line_index + 1
        for column_index, _value in enumerate(columns, start=1):
            column_letter = _excel_column_letter(column_index)
            if source_type == "pdf_text" and page_number is not None:
                source_cell = f"P{page_number}R{source_line_number}C{column_index}"
            elif source_type == "pdf_ocr" and page_number is not None:
                source_cell = f"P{page_number}OCRR{source_line_number}C{column_index}"
            else:
                source_cell = f"OCRR{source_line_number}C{column_index}"
            source_payload = {
                "sheetName": sheet_name,
                "rowNumber": source_line_number,
                "columnNumber": column_index,
                "columnLetter": column_letter,
                "cell": source_cell,
                "sourceCell": source_cell,
                "mergedRange": None,
                "sourceType": source_type,
            }
            if page_number is not None:
                source_payload["pageNumber"] = page_number
            if ocr_engine:
                source_payload["ocrEngine"] = ocr_engine
            source_cells[(row_index, column_index - 1)] = source_payload
    if not rows:
        return None
    frame = pd.DataFrame(rows)
    frame.attrs["source_cells"] = source_cells
    return frame


def _read_pdf_text(file_path: Path, file_name: str) -> dict[str, Any]:
    from pypdf import PdfReader

    table_sheets = _pdfplumber_table_frames(file_path)
    if table_sheets:
        digest = build_workbook_digest_from_frames(
            table_sheets,
            file_name=file_name,
            digest_type="pdf_text",
        )
        digest["sourceFormat"] = "pdf_text"
        _mark_pending_without_compare_groups(
            digest,
            "PDF tables were extracted, but no comparable trim table was detected.",
        )
        if digest.get("compareGroups"):
            return digest

    reader = PdfReader(str(file_path))
    sheets: list[tuple[str, pd.DataFrame]] = []
    has_extractable_text = False
    for page_index, page in enumerate(reader.pages[:MAX_SHEETS]):
        page_number = page_index + 1
        sheet_name = f"PDF Page {page_number}"
        text = page.extract_text() or ""
        if text.strip():
            has_extractable_text = True
        frame = _text_table_frame_from_lines(
            text.splitlines(),
            sheet_name=sheet_name,
            source_type="pdf_text",
            page_number=page_number,
        )
        if frame is not None:
            sheets.append((sheet_name, frame))
    if not sheets:
        if has_extractable_text:
            return _pending_text_digest(
                file_name,
                "pdf_text",
                "PDF text was extracted, but no comparable trim table was detected.",
            )
        return _read_pdf_ocr(file_path, file_name)

    digest = build_workbook_digest_from_frames(sheets, file_name=file_name, digest_type="pdf_text")
    digest["sourceFormat"] = "pdf_text"
    _mark_pending_without_compare_groups(
        digest,
        "PDF text was extracted, but no comparable trim table was detected.",
    )
    return digest


def _render_pdf_pages_for_ocr(file_path: Path, output_dir: Path) -> tuple[list[Path], str | None]:
    try:
        pdfium = importlib.import_module("pypdfium2")
    except Exception:
        return [], "No extractable table text was detected in this PDF. Install pypdfium2 plus PaddleOCR/tesseract for scanned PDF OCR."

    try:
        document = pdfium.PdfDocument(str(file_path))
    except Exception as exc:
        return [], f"PDF page rendering failed: {exc}"

    image_paths: list[Path] = []
    try:
        page_count = min(len(document), MAX_SHEETS)
        for page_index in range(page_count):
            page = document[page_index]
            bitmap = page.render(scale=2)
            image = bitmap.to_pil()
            image_path = output_dir / f"page-{page_index + 1}.png"
            image.save(image_path)
            image_paths.append(image_path)
            close_page = getattr(page, "close", None)
            if callable(close_page):
                close_page()
    except Exception as exc:
        return [], f"PDF page rendering failed: {exc}"
    finally:
        close_document = getattr(document, "close", None)
        if callable(close_document):
            close_document()
    if not image_paths:
        return [], "PDF page rendering produced no images."
    return image_paths, None


def _read_pdf_ocr(file_path: Path, file_name: str) -> dict[str, Any]:
    sheets: list[tuple[str, pd.DataFrame]] = []
    ocr_engines: list[str] = []
    ocr_candidate_reports: list[dict[str, Any]] = []
    last_message: str | None = None
    with tempfile.TemporaryDirectory(prefix="jato_pdf_ocr_") as temp_dir:
        image_paths, render_message = _render_pdf_pages_for_ocr(file_path, Path(temp_dir))
        if not image_paths:
            return _pending_text_digest(
                file_name,
                "pdf_text",
                render_message or "No extractable table text was detected in this PDF.",
            )
        for page_index, image_path in enumerate(image_paths, start=1):
            sheet_name = f"PDF OCR Page {page_index}"
            frame, engine_or_message, message, candidate_reports = _select_best_pdf_ocr_frame(
                image_path,
                sheet_name=sheet_name,
                page_number=page_index,
            )
            ocr_candidate_reports.extend(candidate_reports)
            if frame is None:
                last_message = message or engine_or_message
                continue
            if engine_or_message:
                ocr_engines.append(engine_or_message)
            sheets.append((sheet_name, frame))
    if not sheets:
        digest = _pending_text_digest(
            file_name,
            "pdf_ocr",
            last_message or "PDF pages were rendered for OCR, but no comparable table rows were detected.",
        )
        digest["sourceFormat"] = "pdf_ocr"
        if ocr_engines:
            digest["ocrEngine"] = ocr_engines[0]
        if ocr_candidate_reports:
            digest["ocrEngineCandidates"] = ocr_candidate_reports
            _attach_ocr_evaluation(digest, ocr_candidate_reports)
        return digest

    digest = build_workbook_digest_from_frames(sheets, file_name=file_name, digest_type="pdf_ocr")
    digest["sourceFormat"] = "pdf_ocr"
    if ocr_engines:
        digest["ocrEngine"] = ocr_engines[0]
    if ocr_candidate_reports:
        digest["ocrEngineCandidates"] = ocr_candidate_reports
        _attach_ocr_evaluation(digest, ocr_candidate_reports)
    _mark_pending_without_compare_groups(
        digest,
        "PDF OCR text was extracted, but no comparable trim table was detected.",
    )
    return digest


def _configured_custom_ocr_command(file_path: Path) -> tuple[list[str], str] | None:
    raw_command = os.environ.get("JATO_CONFIG_OCR_COMMAND", "").strip()
    if raw_command:
        args = [part.replace("{input}", str(file_path)) for part in shlex.split(raw_command)]
        if not any(str(file_path) in part for part in args):
            args.append(str(file_path))
        engine = Path(args[0]).name if args else "custom_ocr"
        return args, engine
    return None


def _configured_tesseract_command(file_path: Path) -> tuple[list[str], str] | None:
    executable = shutil.which(os.environ.get("JATO_TESSERACT_COMMAND", "tesseract"))
    if executable is None:
        return None
    return [executable, str(file_path), "stdout", "--psm", "6"], "tesseract"


def _extract_command_ocr_text(args: list[str], engine: str) -> tuple[str | None, str | None]:
    try:
        completed = subprocess.run(
            args,
            check=False,
            capture_output=True,
            text=True,
            timeout=45,
        )
    except Exception as exc:
        return None, f"{engine} OCR execution failed: {exc}"
    if completed.returncode != 0:
        error = (completed.stderr or "").strip()
        return None, f"{engine} OCR execution failed with exit code {completed.returncode}: {error}"
    text = (completed.stdout or "").strip()
    if not text:
        return None, f"{engine} OCR produced no text."
    return text, engine


def _paddleocr_line_text(line: Any) -> str | None:
    if isinstance(line, str):
        cleaned = line.strip()
        return cleaned or None
    if isinstance(line, dict):
        if isinstance(line.get("rec_text"), str):
            return line["rec_text"].strip() or None
        if isinstance(line.get("text"), str):
            return line["text"].strip() or None
        rec_texts = line.get("rec_texts")
        if isinstance(rec_texts, list):
            joined = " ".join(str(item).strip() for item in rec_texts if str(item).strip())
            return joined or None
        return None
    if isinstance(line, tuple) and line and isinstance(line[0], str):
        return line[0].strip() or None
    if isinstance(line, list):
        if len(line) >= 2 and isinstance(line[1], tuple) and line[1] and isinstance(line[1][0], str):
            return line[1][0].strip() or None
        if len(line) >= 2 and isinstance(line[1], list) and line[1] and isinstance(line[1][0], str):
            return line[1][0].strip() or None
    return None


def _numeric_sequence(value: Any) -> list[float]:
    if hasattr(value, "tolist"):
        value = value.tolist()
    if isinstance(value, int | float):
        return [float(value)]
    if not isinstance(value, (list, tuple)):
        return []
    numbers: list[float] = []
    for item in value:
        numbers.extend(_numeric_sequence(item))
    return numbers


def _paddleocr_box_bounds(value: Any) -> tuple[float, float, float, float] | None:
    numbers = _numeric_sequence(value)
    if len(numbers) < 4:
        return None
    if len(numbers) == 4:
        left, top, right, bottom = numbers
        return min(left, right), min(top, bottom), max(left, right), max(top, bottom)
    xs = numbers[0::2]
    ys = numbers[1::2]
    if not xs or not ys:
        return None
    return min(xs), min(ys), max(xs), max(ys)


def _median_number(values: list[float], fallback: float) -> float:
    if not values:
        return fallback
    sorted_values = sorted(values)
    midpoint = len(sorted_values) // 2
    if len(sorted_values) % 2 == 1:
        return sorted_values[midpoint]
    return (sorted_values[midpoint - 1] + sorted_values[midpoint]) / 2


def _paddleocr_geometry_items(result: Any) -> list[tuple[str, float, float, float, float]]:
    items: list[tuple[str, float, float, float, float]] = []
    if isinstance(result, dict) or callable(getattr(result, "get", None)):
        rec_texts = result.get("rec_texts")
        boxes = result.get("rec_boxes")
        if boxes is None:
            boxes = result.get("rec_polys")
        if boxes is None:
            boxes = result.get("dt_polys")
        if hasattr(boxes, "tolist"):
            boxes = boxes.tolist()
        if isinstance(rec_texts, list) and isinstance(boxes, list):
            for text_value, box_value in zip(rec_texts, boxes, strict=False):
                text = str(text_value).strip()
                bounds = _paddleocr_box_bounds(box_value)
                if text and bounds is not None:
                    items.append((text, *bounds))
        if items:
            return items
    if isinstance(result, (list, tuple)):
        for item in result:
            items.extend(_paddleocr_geometry_items(item))
    return items


def _collect_paddleocr_table_lines_from_geometry(result: Any) -> list[str]:
    items = _paddleocr_geometry_items(result)
    if len(items) < 3:
        return []
    heights = [max(1.0, bottom - top) for _text, _left, top, _right, bottom in items]
    row_threshold = max(8.0, _median_number(heights, 12.0) * 0.75)
    rows: list[list[tuple[str, float, float, float, float]]] = []
    for item in sorted(items, key=lambda entry: ((entry[2] + entry[4]) / 2, entry[1])):
        center_y = (item[2] + item[4]) / 2
        for row in rows:
            row_center = sum((entry[2] + entry[4]) / 2 for entry in row) / len(row)
            if abs(center_y - row_center) <= row_threshold:
                row.append(item)
                break
        else:
            rows.append([item])
    lines: list[str] = []
    for row in rows:
        sorted_row = sorted(row, key=lambda entry: entry[1])
        cells = [text for text, *_bounds in sorted_row if text]
        if len(cells) >= 3:
            lines.append(" | ".join(cells))
    return lines


def _collect_paddleocr_text_lines(result: Any) -> list[str]:
    table_lines = _collect_paddleocr_table_lines_from_geometry(result)
    if table_lines:
        return table_lines
    lines: list[str] = []
    direct_line = _paddleocr_line_text(result)
    if direct_line:
        return [direct_line]
    if isinstance(result, dict):
        for key in ("rec_texts", "texts", "ocr_texts"):
            values = result.get(key)
            if isinstance(values, list):
                lines.extend(str(value).strip() for value in values if str(value).strip())
        if lines:
            return lines
    if isinstance(result, (list, tuple)):
        for item in result:
            lines.extend(_collect_paddleocr_text_lines(item))
    return lines


def _extract_paddleocr_text(file_path: Path) -> tuple[str | None, str | None]:
    try:
        paddleocr_module = importlib.import_module("paddleocr")
    except Exception:
        return None, "PaddleOCR is not installed."
    paddle_ocr_class = getattr(paddleocr_module, "PaddleOCR", None)
    if paddle_ocr_class is None:
        return None, "PaddleOCR package does not expose PaddleOCR."

    lang = os.environ.get("JATO_PADDLEOCR_LANG", "ch")
    init_attempts = [
        {"lang": lang, "use_angle_cls": True, "show_log": False},
        {"lang": lang},
        {},
    ]
    last_error: Exception | None = None
    for kwargs in init_attempts:
        try:
            ocr = paddle_ocr_class(**kwargs)
            break
        except TypeError as exc:
            last_error = exc
        except Exception as exc:
            last_error = exc
    else:
        return None, f"PaddleOCR initialization failed: {last_error}"

    try:
        if hasattr(ocr, "ocr"):
            try:
                result = ocr.ocr(str(file_path), cls=True)
            except TypeError:
                result = ocr.ocr(str(file_path))
        elif hasattr(ocr, "predict"):
            result = ocr.predict(str(file_path))
        else:
            return None, "PaddleOCR object has no ocr or predict method."
    except Exception as exc:
        return None, f"PaddleOCR execution failed: {exc}"

    lines = _collect_paddleocr_text_lines(result)
    text = "\n".join(line for line in lines if line).strip()
    if not text:
        return None, "PaddleOCR produced no text."
    return text, "paddleocr"


def _extract_ocr_text_candidates(file_path: Path, *, compare_engines: bool = False) -> tuple[list[tuple[str, str]], list[str]]:
    candidates: list[tuple[str, str]] = []
    messages: list[str] = []

    custom_command = _configured_custom_ocr_command(file_path)
    if custom_command is not None:
        text, engine_or_message = _extract_command_ocr_text(*custom_command)
        if text and engine_or_message:
            candidates.append((text, engine_or_message))
            if not compare_engines:
                return candidates, messages
        elif engine_or_message:
            messages.append(engine_or_message)

    paddle_text, paddle_engine_or_message = _extract_paddleocr_text(file_path)
    if paddle_text and paddle_engine_or_message:
        candidates.append((paddle_text, paddle_engine_or_message))
        if not compare_engines:
            return candidates, messages
    elif paddle_engine_or_message:
        messages.append(paddle_engine_or_message)

    tesseract_command = _configured_tesseract_command(file_path)
    if tesseract_command is not None:
        text, engine_or_message = _extract_command_ocr_text(*tesseract_command)
        if text and engine_or_message:
            candidates.append((text, engine_or_message))
            if not compare_engines:
                return candidates, messages
        elif engine_or_message:
            messages.append(engine_or_message)

    if not candidates and (not messages or messages == ["PaddleOCR is not installed."]):
        messages.append("OCR engine is not configured. Install PaddleOCR or tesseract, or set JATO_CONFIG_OCR_COMMAND.")
    return candidates, messages


def _ocr_int_field(payload: dict[str, Any], key: str) -> int:
    value = payload.get(key)
    return int(value) if isinstance(value, int | float) else 0


def _ocr_table_frame_score(
    frame: pd.DataFrame,
    *,
    sheet_name: str,
    source_type: str,
) -> OcrCandidateScore:
    row_count, column_count = frame.shape
    non_empty_count = int(frame.astype(str).map(lambda value: bool(value.strip())).sum().sum())
    table_shape_score = 1 if row_count >= 2 and column_count >= 3 else 0
    digest = build_workbook_digest_from_frames(
        [(sheet_name, frame)],
        file_name=f"{sheet_name}.ocr",
        digest_type=source_type,
    )
    summary = digest.get("summary", {})
    comparable_group_count = _ocr_int_field(summary, "comparableGroupCount")
    total_feature_count = _ocr_int_field(summary, "featureCount")
    total_difference_count = _ocr_int_field(summary, "differenceCount")
    total_candidate_trim_count = _ocr_int_field(summary, "candidateTrimCount")
    compare_groups = [group for group in digest.get("compareGroups", []) if isinstance(group, dict)]
    best_group = max(
        compare_groups,
        key=lambda group: (
            _ocr_int_field(group, "featureCount"),
            _ocr_int_field(group, "differenceCount"),
            _ocr_int_field(group, "trimCount"),
        ),
        default=None,
    )
    feature_count = _ocr_int_field(best_group, "featureCount") if best_group is not None else 0
    difference_count = _ocr_int_field(best_group, "differenceCount") if best_group is not None else 0
    candidate_trim_count = _ocr_int_field(best_group, "trimCount") if best_group is not None else 0
    semantic_score = 1 if feature_count >= OCR_MIN_COMPARABLE_FEATURES and candidate_trim_count >= 2 else 0
    return (
        semantic_score,
        feature_count,
        difference_count,
        candidate_trim_count,
        comparable_group_count,
        total_feature_count,
        total_difference_count,
        total_candidate_trim_count,
        table_shape_score,
        row_count,
        column_count,
        non_empty_count,
    )


def _ocr_score_payload(score: OcrCandidateScore | None) -> dict[str, int]:
    (
        semantic_score,
        feature_count,
        difference_count,
        candidate_trim_count,
        comparable_group_count,
        total_feature_count,
        total_difference_count,
        total_candidate_trim_count,
        table_shape_score,
        row_count,
        column_count,
        non_empty_count,
    ) = score or (0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
    return {
        "semanticScore": semantic_score,
        "comparableGroupCount": comparable_group_count,
        "featureCount": feature_count,
        "differenceCount": difference_count,
        "candidateTrimCount": candidate_trim_count,
        "totalFeatureCount": total_feature_count,
        "totalDifferenceCount": total_difference_count,
        "totalCandidateTrimCount": total_candidate_trim_count,
        "tableShapeScore": table_shape_score,
        "rowCount": row_count,
        "columnCount": column_count,
        "nonEmptyCount": non_empty_count,
    }


def _ocr_score_is_comparable(score: OcrCandidateScore) -> bool:
    return (
        score[0] > 0
        and score[1] >= OCR_MIN_COMPARABLE_FEATURES
        and score[3] >= 2
    )


def _ocr_text_preview(text: str | None, *, max_length: int = 240) -> str | None:
    if not text:
        return None
    preview = " ".join(line.strip() for line in text.splitlines() if line.strip())
    if not preview:
        return None
    return preview if len(preview) <= max_length else f"{preview[:max_length].rstrip()}..."


def _ocr_text_line_count(text: str | None) -> int:
    if not text:
        return 0
    return sum(1 for line in text.splitlines() if line.strip())


def _ocr_candidate_report(
    *,
    engine: str,
    source_type: str,
    sheet_name: str,
    page_number: int | None,
    score: OcrCandidateScore | None,
    text: str | None = None,
    selected: bool = False,
    message: str | None = None,
    comparable_table_detected: bool | None = None,
) -> dict[str, Any]:
    report: dict[str, Any] = {
        "engine": engine,
        "sourceType": source_type,
        "sheetName": sheet_name,
        "selected": selected,
        "comparableTableDetected": (
            score is not None if comparable_table_detected is None else comparable_table_detected
        ),
        "score": _ocr_score_payload(score),
    }
    text_preview = _ocr_text_preview(text)
    if text_preview:
        report["textPreview"] = text_preview
        report["lineCount"] = _ocr_text_line_count(text)
    if page_number is not None:
        report["pageNumber"] = page_number
    if message:
        report["message"] = message
    return report


def _ocr_failure_engine_for_message(message: str) -> str | None:
    normalized = message.strip().lower()
    if normalized.startswith("paddleocr"):
        return "paddleocr"
    if " ocr execution failed" in normalized or " ocr produced no text" in normalized:
        engine = message.strip().split(" ", 1)[0].strip()
        return engine or None
    return None


def _append_ocr_extraction_failure_reports(
    reports: list[dict[str, Any]],
    messages: list[str],
    *,
    source_type: str,
    sheet_name: str,
    page_number: int | None,
) -> None:
    existing = {
        (str(report.get("engine")), str(report.get("message", "")))
        for report in reports
    }
    for message in messages:
        engine = _ocr_failure_engine_for_message(message)
        if engine is None:
            continue
        key = (engine, message)
        if key in existing:
            continue
        reports.append(_ocr_candidate_report(
            engine=engine,
            source_type=source_type,
            sheet_name=sheet_name,
            page_number=page_number,
            score=None,
            message=message,
        ))
        existing.add(key)


def _ocr_evaluation_payload(candidate_reports: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not candidate_reports:
        return None

    selected_candidates = [candidate for candidate in candidate_reports if candidate.get("selected")]
    selected_candidate = selected_candidates[0] if selected_candidates else None
    selected_engines = sorted(
        {str(candidate.get("engine")) for candidate in selected_candidates if candidate.get("engine")}
    )
    payload: dict[str, Any] = {
        "strategy": OCR_CONFIG_SEMANTIC_STRATEGY,
        "reason": OCR_CONFIG_SEMANTIC_STRATEGY if selected_candidate else "no_comparable_table_detected",
        "candidateCount": len(candidate_reports),
        "comparableCandidateCount": sum(
            1 for candidate in candidate_reports if candidate.get("comparableTableDetected")
        ),
        "selectedCandidateCount": len(selected_candidates),
        "selectedEngine": selected_candidate.get("engine") if selected_candidate else None,
        "selectedEngines": selected_engines,
        "selectedScore": selected_candidate.get("score") if selected_candidate else None,
        "selectedSheetName": selected_candidate.get("sheetName") if selected_candidate else None,
    }
    if selected_candidate and selected_candidate.get("pageNumber") is not None:
        payload["selectedPageNumber"] = selected_candidate["pageNumber"]
    reason_details = _ocr_selection_reason_details(candidate_reports, selected_candidate)
    if reason_details:
        payload["selectedReasonDetails"] = reason_details
    return payload


def _attach_ocr_evaluation(digest: dict[str, Any], candidate_reports: list[dict[str, Any]]) -> None:
    evaluation = _ocr_evaluation_payload(candidate_reports)
    if evaluation is not None:
        digest["ocrEvaluation"] = evaluation


def _ocr_score_detail(score: dict[str, Any]) -> str:
    feature_count = _ocr_int_field(score, "featureCount")
    trim_count = _ocr_int_field(score, "candidateTrimCount")
    difference_count = _ocr_int_field(score, "differenceCount")
    row_count = _ocr_int_field(score, "rowCount")
    column_count = _ocr_int_field(score, "columnCount")
    non_empty_count = _ocr_int_field(score, "nonEmptyCount")
    parts = [
        f"配置项 {feature_count}",
        f"配置列 {trim_count}",
        f"差异 {difference_count}",
    ]
    if row_count > 0 or column_count > 0:
        parts.append(f"表格 {row_count} x {column_count}")
    if non_empty_count > 0:
        parts.append(f"非空 {non_empty_count}")
    return "、".join(parts)


def _ocr_selection_reason_details(
    candidate_reports: list[dict[str, Any]],
    selected_candidate: dict[str, Any] | None,
) -> list[str]:
    if not selected_candidate:
        return []
    selected_engine = str(selected_candidate.get("engine") or "selected OCR")
    selected_score = selected_candidate.get("score") if isinstance(selected_candidate.get("score"), dict) else {}
    details: list[str] = []
    failed_candidates = [
        candidate
        for candidate in candidate_reports
        if candidate is not selected_candidate and not candidate.get("comparableTableDetected")
    ]
    if failed_candidates:
        failed_engines = ", ".join(str(candidate.get("engine") or "unknown") for candidate in failed_candidates[:3])
        details.append(f"{selected_engine} 识别到可比配置表；{failed_engines} 未形成可比配置表。")

    comparable_others = [
        candidate
        for candidate in candidate_reports
        if candidate is not selected_candidate and candidate.get("comparableTableDetected")
    ]
    if comparable_others:
        runner_up = max(
            comparable_others,
            key=lambda candidate: (
                _ocr_int_field(candidate.get("score") if isinstance(candidate.get("score"), dict) else {}, "semanticScore"),
                _ocr_int_field(candidate.get("score") if isinstance(candidate.get("score"), dict) else {}, "featureCount"),
                _ocr_int_field(candidate.get("score") if isinstance(candidate.get("score"), dict) else {}, "differenceCount"),
                _ocr_int_field(candidate.get("score") if isinstance(candidate.get("score"), dict) else {}, "candidateTrimCount"),
            ),
        )
        runner_score = runner_up.get("score") if isinstance(runner_up.get("score"), dict) else {}
        details.append(
            f"{selected_engine} {_ocr_score_detail(selected_score)}；"
            f"{runner_up.get('engine') or 'runner-up'} {_ocr_score_detail(runner_score)}。"
        )
    elif selected_score:
        details.append(f"{selected_engine} 选中结果：{_ocr_score_detail(selected_score)}。")

    return details[:3]


def _select_best_ocr_frame(
    image_path: Path,
    *,
    sheet_name: str,
    source_type: str,
    page_number: int | None = None,
) -> tuple[pd.DataFrame | None, str | None, str | None, list[dict[str, Any]]]:
    candidates, messages = _extract_ocr_text_candidates(image_path, compare_engines=True)
    extraction_messages = list(messages)
    reports: list[dict[str, Any]] = []
    best_frame: pd.DataFrame | None = None
    best_engine: str | None = None
    best_score: OcrCandidateScore | None = None
    best_report_index: int | None = None
    for text, engine in candidates:
        frame = _text_table_frame_from_lines(
            text.splitlines(),
            sheet_name=sheet_name,
            source_type=source_type,
            page_number=page_number,
            ocr_engine=engine,
        )
        if frame is None:
            message = f"{engine} OCR text did not contain comparable table rows."
            messages.append(message)
            reports.append(_ocr_candidate_report(
                engine=engine,
                source_type=source_type,
                sheet_name=sheet_name,
                page_number=page_number,
                score=None,
                text=text,
                message=message,
            ))
            continue
        score = _ocr_table_frame_score(frame, sheet_name=sheet_name, source_type=source_type)
        if not _ocr_score_is_comparable(score):
            message = (
                f"{engine} OCR text produced a low-confidence config table "
                f"with fewer than {OCR_MIN_COMPARABLE_FEATURES} comparable features."
            )
            messages.append(message)
            reports.append(_ocr_candidate_report(
                engine=engine,
                source_type=source_type,
                sheet_name=sheet_name,
                page_number=page_number,
                score=score,
                text=text,
                message=message,
                comparable_table_detected=False,
            ))
            continue
        reports.append(_ocr_candidate_report(
            engine=engine,
            source_type=source_type,
            sheet_name=sheet_name,
            page_number=page_number,
            score=score,
            text=text,
        ))
        if best_score is None or score > best_score:
            best_frame = frame
            best_engine = engine
            best_score = score
            best_report_index = len(reports) - 1
    _append_ocr_extraction_failure_reports(
        reports,
        extraction_messages,
        source_type=source_type,
        sheet_name=sheet_name,
        page_number=page_number,
    )
    if best_report_index is not None:
        reports[best_report_index]["selected"] = True
    if best_frame is None:
        return None, None, "; ".join(message for message in messages if message), reports
    return best_frame, best_engine, None, reports


def _select_best_pdf_ocr_frame(
    image_path: Path,
    *,
    sheet_name: str,
    page_number: int,
) -> tuple[pd.DataFrame | None, str | None, str | None, list[dict[str, Any]]]:
    return _select_best_ocr_frame(
        image_path,
        sheet_name=sheet_name,
        source_type="pdf_ocr",
        page_number=page_number,
    )


def _select_best_image_ocr_frame(
    image_path: Path,
    *,
    sheet_name: str,
) -> tuple[pd.DataFrame | None, str | None, str | None, list[dict[str, Any]]]:
    return _select_best_ocr_frame(
        image_path,
        sheet_name=sheet_name,
        source_type="image_ocr",
    )


def _extract_ocr_text(file_path: Path) -> tuple[str | None, str | None]:
    candidates, messages = _extract_ocr_text_candidates(file_path)
    if candidates:
        return candidates[0]
    return None, "; ".join(message for message in messages if message) or "OCR did not return text."


def _read_image_ocr(file_path: Path, file_name: str) -> dict[str, Any]:
    sheet_name = "OCR Image 1"
    frame, engine_or_message, message, candidate_reports = _select_best_image_ocr_frame(file_path, sheet_name=sheet_name)
    if frame is None:
        digest = _pending_text_digest(file_name, "image_ocr", message or engine_or_message or "OCR did not return text.")
        if candidate_reports:
            digest["ocrEngineCandidates"] = candidate_reports
            _attach_ocr_evaluation(digest, candidate_reports)
        return digest
    digest = build_workbook_digest_from_frames([(sheet_name, frame)], file_name=file_name, digest_type="image_ocr")
    digest["sourceFormat"] = "image_ocr"
    digest["ocrEngine"] = engine_or_message
    if candidate_reports:
        digest["ocrEngineCandidates"] = candidate_reports
        _attach_ocr_evaluation(digest, candidate_reports)
    _mark_pending_without_compare_groups(
        digest,
        "OCR text was extracted, but no comparable trim table was detected.",
    )
    return digest


def _normalise_tabular_frame(frame: pd.DataFrame) -> pd.DataFrame:
    if isinstance(frame.columns, pd.MultiIndex):
        column_values = [" / ".join(_clean_cell(part) for part in column if _clean_cell(part)) for column in frame.columns]
    else:
        column_values = [_clean_cell(column) for column in frame.columns]
    if any(column_values) and not all(value.isdigit() for value in column_values):
        header_frame = pd.DataFrame([column_values])
        return pd.concat([header_frame, frame.astype(str)], ignore_index=True)
    return frame


def build_source_digest(file_path: str | Path, file_name: str) -> dict[str, Any] | None:
    source_path = Path(file_path)
    extension = Path(file_name).suffix.lower()
    if extension in WORKBOOK_EXTENSIONS:
        sheets = _read_workbook(source_path)
        return build_workbook_digest_from_frames(sheets, file_name=file_name)
    if extension in TABULAR_EXTENSIONS:
        return _read_tabular(source_path, file_name)
    if extension in PDF_EXTENSIONS:
        return _read_pdf_text(source_path, file_name)
    if extension in IMAGE_EXTENSIONS:
        return _read_image_ocr(source_path, file_name)
    return None
