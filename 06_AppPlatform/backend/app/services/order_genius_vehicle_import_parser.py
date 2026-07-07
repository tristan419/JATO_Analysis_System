"""Excel parser for Order Genius PI vehicle allocation imports."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


HEADER_MAP = {
    "pi code": "pi_code",
    "official pi no": "official_pi_no",
    "car code": "car_code",
    "vin": "vin",
    "bom": "bom",
    "material code": "material_code",
    "brand": "brand",
    "model": "model_name",
    "model name": "model_name",
    "version": "version",
    "powertrain": "powertrain",
    "exterior colour": "exterior_color_name",
    "exterior color": "exterior_color_name",
    "interior colour": "interior_color_name",
    "interior color": "interior_color_name",
    "order date": "order_date",
    "production date": "production_date",
    "etd": "etd",
    "eta": "eta",
    "ship name": "ship_name",
    "country": "country_code",
    "country code": "country_code",
    "dealer code": "dealer_code",
    "dealer name": "dealer_name",
    "customer ref": "customer_ref",
    "allocation status": "allocation_status",
    "logistics status": "logistics_status",
    "ready for pickup date": "ready_for_pickup_date",
    "shipping schedule url": "shipping_schedule_url",
    "feishu tracking url": "feishu_tracking_url",
    "remark": "remark",
}

VIN_HEADER_LABELS = {"vin", "vin code", "vin no", "vin number", "车架号", "车辆识别码"}


def parse_vehicle_allocation_xlsx(file_path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_row = _find_header_row(ws)
        if header_row is None:
            raise ValueError("Missing PI vehicle allocation headers")
        header_map = _header_map(ws, header_row)
        rows: list[dict[str, Any]] = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            item: dict[str, Any] = {"sourceRow": row_idx}
            has_value = False
            for key, col_idx in header_map.items():
                raw = ws.cell(row=row_idx, column=col_idx).value
                value = _normalise_cell(raw)
                if value not in (None, ""):
                    has_value = True
                item[key] = value
            if has_value:
                rows.append(item)
        return rows
    finally:
        wb.close()


def parse_vehicle_vin_list_xlsx(file_path: Path) -> list[str]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        vins: list[str] = []
        for ws in wb.worksheets:
            vin_column = _find_vin_column(ws)
            if vin_column:
                for row_idx in range(vin_column["header_row"] + 1, ws.max_row + 1):
                    vins.extend(_cell_tokens(ws.cell(row=row_idx, column=vin_column["column"]).value))
                continue

            for row_idx in range(1, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    vins.extend(_cell_tokens(ws.cell(row=row_idx, column=col_idx).value))
        return vins
    finally:
        wb.close()


def _find_header_row(ws) -> int | None:
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        values = {
            str(ws.cell(row=row_idx, column=col_idx).value or "").strip().lower()
            for col_idx in range(1, ws.max_column + 1)
        }
        if "pi code" in values and ("car code" in values or "vin" in values):
            return row_idx
    return None


def _find_vin_column(ws) -> dict[str, int] | None:
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        for col_idx in range(1, ws.max_column + 1):
            raw = ws.cell(row=row_idx, column=col_idx).value
            label = str(raw or "").strip().lower()
            if label in VIN_HEADER_LABELS:
                return {"header_row": row_idx, "column": col_idx}
    return None


def _header_map(ws, row_idx: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        raw = ws.cell(row=row_idx, column=col_idx).value
        if raw is None:
            continue
        key = HEADER_MAP.get(str(raw).strip().lower())
        if key:
            result[key] = col_idx
    return result


def _cell_tokens(value: object) -> list[str]:
    normalised = _normalise_cell(value)
    if normalised in (None, ""):
        return []
    if not isinstance(normalised, str):
        normalised = str(normalised)
    tokens = [
        token.strip().upper()
        for token in normalised.replace(",", " ").replace(";", " ").split()
        if token.strip()
    ]
    return [
        token
        for token in tokens
        if token.lower() not in VIN_HEADER_LABELS
    ]


def _normalise_cell(value: object) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return value
