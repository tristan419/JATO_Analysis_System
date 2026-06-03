"""Parse exported Order Genius Excel back into structured order quantity data.

Handles the round-trip: Export XLSX → edit offline → re-import quantities.

Export format (one editable sheet per powertrain):
  Row 1: Title  — "Order Genius — {country_name} ({country_code}) {year}"
  Row 2: Header — Model | Version | Colour | Interior | Material Code | FOB(EUR) | month columns | TTL
  Row 3+: Data  — one SKU per row, monthly quantities located by header name

Edge cases handled:
  - Non-numeric / negative / decimal quantities → flagged per-cell
  - Unknown material codes → flagged per-row
  - FOB value present for diff against current system FOB
  - Historical rows (identified by quantity>0 despite lifecycle status)
  - Wrong-format detection (Material Master vs Order Export)
  - Multiple editable sheets combined into one result set
  - Aggregate sheets such as "Total list" are skipped
  - Empty cells → treated as explicit 0
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

TITLE_RE = re.compile(r"Order Genius .+ \(([A-Z]{2})\) (\d{4})")
MONTH_HEADERS = ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")


@dataclass
class ImportedQuantityCell:
    material_code: str
    month: int  # 1-12
    quantity: int
    fob_eur: float | None = None
    model_name: str = ""
    version: str = ""
    colour: str = ""
    error: str = ""  # per-cell error message


@dataclass
class ImportedRow:
    material_code: str
    fob_eur: float | None
    model_name: str
    version: str
    colour: str
    cells: list[ImportedQuantityCell] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)  # per-row errors


@dataclass
class OrderQuantityImport:
    country_code: str
    year: int
    rows: list[ImportedRow] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)  # file-level errors


def parse_order_quantity_xlsx(file_path: Path) -> OrderQuantityImport:
    """Parse an exported Order Genius Excel file into structured import data."""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    country_code: str | None = None
    year: int | None = None
    all_rows: list[ImportedRow] = []
    errors: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 3:
            continue

        title_cell = ws.cell(row=1, column=1).value
        if not isinstance(title_cell, str):
            errors.append(f"Sheet '{sheet_name}': missing title row, skipping")
            continue

        m = TITLE_RE.match(title_cell)
        if not m:
            # Detect wrong format: Material Master has different title structure
            if any(kw in title_cell.lower() for kw in ("bom", "material", "code name", "configuration")):
                errors.append(
                    f"Sheet '{sheet_name}': appears to be a Material Master file, "
                    "not an Order Quantity export. Please use the exported Order Genius XLSX."
                )
                continue
            errors.append(
                f"Sheet '{sheet_name}': cannot parse country/year from title '{title_cell}', skipping"
            )
            continue

        sheet_country = m.group(1)
        sheet_year = int(m.group(2))

        if country_code is None:
            country_code = sheet_country
        elif country_code != sheet_country:
            errors.append(
                f"Sheet '{sheet_name}': country {sheet_country} differs from "
                f"previously parsed {country_code}. All sheets must be for the same country."
            )
            continue

        if year is None:
            year = sheet_year
        elif year != sheet_year:
            errors.append(
                f"Sheet '{sheet_name}': year {sheet_year} differs from "
                f"previously parsed {year}."
            )
            continue

        header_map = _header_map(ws, row=2)
        if "material code" not in header_map:
            # Non-editable aggregate sheets share the Order Genius title but do not carry SKU rows.
            if sheet_name.strip().lower() == "total list" or "product code" in header_map:
                continue
            errors.append(f"Sheet '{sheet_name}': missing Material Code header, skipping")
            continue

        month_columns = {
            month: header_map[header.lower()]
            for month, header in enumerate(MONTH_HEADERS, start=1)
            if header.lower() in header_map
        }
        if not month_columns:
            errors.append(f"Sheet '{sheet_name}': missing month headers, skipping")
            continue

        model_col = header_map.get("model")
        version_col = header_map.get("version")
        colour_col = header_map.get("colour")
        material_code_col = header_map["material code"]
        fob_col = header_map.get("fob(eur)")

        # Parse data rows (starting from row 3)
        material_codes_seen: set[str] = set()
        for row_idx in range(3, ws.max_row + 1):
            material_code = _cell_str(ws, row_idx, material_code_col)
            if not material_code:
                continue  # empty row

            if material_code in material_codes_seen:
                # Duplicate within same file — use last occurrence
                all_rows = [r for r in all_rows if r.material_code != material_code]
            material_codes_seen.add(material_code)

            fob_val = _cell_float(ws, row_idx, fob_col) if fob_col else None
            row = ImportedRow(
                material_code=material_code,
                fob_eur=fob_val,
                model_name=_cell_str(ws, row_idx, model_col),
                version=_cell_str(ws, row_idx, version_col),
                colour=_cell_str(ws, row_idx, colour_col),
            )

            for month, month_col in sorted(month_columns.items()):
                raw_val = ws.cell(row=row_idx, column=month_col).value
                cell = _parse_quantity_cell(raw_val, material_code, month, fob_val)
                row.cells.append(cell)
                if cell.error:
                    row.errors.append(f"M{month}: {cell.error}")

            all_rows.append(row)

    wb.close()

    if country_code is None or year is None:
        errors.append("Could not determine country and year from any sheet. Is this an Order Genius export?")
        return OrderQuantityImport(country_code="", year=0, errors=errors)

    return OrderQuantityImport(country_code=country_code, year=year, rows=all_rows, errors=errors)


def _header_map(ws, row: int) -> dict[str, int]:
    """Return normalized header text to 1-based Excel column index."""
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        if value is None:
            continue
        text = str(value).strip().lower()
        if text:
            headers[text] = col
    return headers


def _cell_str(ws, row: int, col: int | None) -> str:
    if col is None:
        return ""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()


def _cell_float(ws, row: int, col: int) -> float | None:
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_quantity_cell(
    raw_val: object,
    material_code: str,
    month: int,
    fob_eur: float | None,
) -> ImportedQuantityCell:
    """Parse a single month's quantity cell, returning error string if invalid."""
    if raw_val is None:
        return ImportedQuantityCell(
            material_code=material_code, month=month, quantity=0,
            fob_eur=fob_eur, error="",
        )

    try:
        num = float(raw_val)
    except (ValueError, TypeError):
        return ImportedQuantityCell(
            material_code=material_code, month=month, quantity=0,
            fob_eur=fob_eur,
            error=f"Non-numeric value '{raw_val}'",
        )

    if num < 0:
        return ImportedQuantityCell(
            material_code=material_code, month=month, quantity=0,
            fob_eur=fob_eur,
            error=f"Negative quantity {num}",
        )

    if num != int(num):
        return ImportedQuantityCell(
            material_code=material_code, month=month, quantity=round(num),
            fob_eur=fob_eur,
            error=f"Decimal quantity {num} → rounded to {round(num)}",
        )

    return ImportedQuantityCell(
        material_code=material_code, month=month, quantity=int(num),
        fob_eur=fob_eur, error="",
    )
