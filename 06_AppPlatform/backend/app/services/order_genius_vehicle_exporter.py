"""Excel export for Order Genius PI vehicle allocation."""

from __future__ import annotations

import io
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "PI Code",
    "Official PI No",
    "Car Code",
    "VIN",
    "BOM",
    "Material Code",
    "Brand",
    "Model",
    "Version",
    "Powertrain",
    "Exterior Colour",
    "Interior Colour",
    "Order Date",
    "Production Date",
    "ETD",
    "ETA",
    "Ship Name",
    "Country",
    "Dealer Code",
    "Dealer Name",
    "Customer Ref",
    "Allocation Status",
    "Logistics Status",
    "Ready for Pickup Date",
    "Shipping Schedule URL",
    "Feishu Tracking URL",
    "Remark",
]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Calibri", size=11)
STRIPED_FILL = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def generate_vehicle_allocation_excel(vehicles: list[dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicle Allocation"

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, vehicle in enumerate(vehicles, start=2):
        row_values = [
            vehicle.get("piCode"),
            vehicle.get("officialPiNo"),
            vehicle.get("carCode"),
            vehicle.get("vin"),
            vehicle.get("bom"),
            vehicle.get("materialCode"),
            vehicle.get("brand"),
            vehicle.get("modelName"),
            vehicle.get("version"),
            vehicle.get("powertrain"),
            vehicle.get("exteriorColorName"),
            vehicle.get("interiorColorName"),
            vehicle.get("orderDate"),
            vehicle.get("productionDate"),
            vehicle.get("etd"),
            vehicle.get("eta"),
            vehicle.get("shipName"),
            vehicle.get("countryCode"),
            vehicle.get("dealerCode"),
            vehicle.get("dealerName"),
            vehicle.get("customerRef"),
            vehicle.get("allocationStatus"),
            vehicle.get("logisticsStatus"),
            vehicle.get("readyForPickupDate"),
            vehicle.get("shippingScheduleUrl"),
            vehicle.get("feishuTrackingUrl"),
            vehicle.get("remark"),
        ]
        row_fill = STRIPED_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_excel_value(value))
            cell.font = NORMAL_FONT
            if row_fill:
                cell.fill = row_fill
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(1, len(vehicles) + 1)}"
    for col_idx, header in enumerate(HEADERS, 1):
        width = max(12, min(28, len(header) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _excel_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value
