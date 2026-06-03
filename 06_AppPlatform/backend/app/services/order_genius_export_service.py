"""Excel export for Order Genius — generates RO-June style workbooks.

One sheet per powertrain, columns: Model | Version | Colour |
Material Code | FOB(EUR) | Jan | Feb | ... | Dec | TTL
"""

from __future__ import annotations

import io
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
STRIPED_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
HISTORICAL_FONT = Font(name="Calibri", size=11, strikethrough=True, color="808080")
NORMAL_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

HEADERS = [
    "Model", "Version", "Colour", "Interior", "Material Code", "FOB(EUR)",
    *MONTH_NAMES, "TTL",
]

PI_HEADERS = [
    "单双色",
    "质保政策",
    "产品编号",
    "数量",
    "单价",
    "单车运费",
    "单车保险",
    "PIProductcategories",
    "PIPower",
    "PIVersion",
    "PIInterior",
    "PIExterior",
    "单车支持",
    "一次内销单价",
    "一次内销单车运费",
    "一次内销单车保险",
]


def _apply_cell_style(cell, font=None, fill=None, alignment=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def generate_order_genius_excel(
    rows: list[dict],
    country_code: str,
    country_name: str,
    year: int,
    include_historical: bool = True,
    include_row_numbers: bool = False,
    column_labels: dict[str, str] | None = None,
    selected_months: list[int] | None = None,
) -> io.BytesIO:
    """Generate an Order Genius Excel workbook.

    Args:
        rows: Matrix rows from ``build_matrix()``.
        country_code: e.g. "RO"
        country_name: Display name for the country.
        year: Order year.
        include_historical: If True, include historical rows with quantity.
        include_row_numbers: If True, prepend a "No." column.
        column_labels: Optional override for header labels (e.g. {"FOB(EUR)": "Price(Eur)"}).
        selected_months: Optional subset of month numbers to export.

    Returns:
        A BytesIO buffer containing the .xlsx file.
    """
    labels = dict(column_labels) if column_labels else {}

    def _header(name: str) -> str:
        return labels.get(name, name)

    export_months = list(dict.fromkeys(selected_months or list(range(1, 13))))
    if not export_months or any(month < 1 or month > 12 for month in export_months):
        raise ValueError("selected_months must contain month numbers from 1 to 12")
    month_names = [MONTH_NAMES[month - 1] for month in export_months]

    # Build effective header list
    effective_headers: list[str] = []
    if include_row_numbers:
        effective_headers.append(_header("No."))
    effective_headers.extend([
        _header("Model"),
        _header("Version"),
        _header("Colour"),
        _header("Interior"),
        _header("Material Code"),
        _header("FOB(EUR)"),
    ])
    for m in month_names:
        effective_headers.append(_header(m))
    effective_headers.append(_header("TTL"))
    n_cols = len(effective_headers)

    wb = openpyxl.Workbook()

    # Group rows by powertrain
    groups: dict[str, list[dict]] = {}
    for row in rows:
        pt = row.get("powertrain") or "Other"
        groups.setdefault(pt, []).append(row)

    sorted_keys = sorted(
        [k for k in groups if k != "Other"]
    ) + (["Other"] if "Other" in groups else [])

    if not sorted_keys:
        ws = wb.active
        ws.title = "No Data"
        ws.cell(row=1, column=1, value=f"No order data for {country_name} ({country_code}) {year}")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    wb.remove(wb.active)

    # ── Total list aggregate sheet (first sheet) ──────────────────
    _build_total_list_sheet(wb, rows, country_code, country_name, year,
                            effective_headers, n_cols, month_names,
                            include_row_numbers, export_months)

    # ── Per-powertrain sheets ─────────────────────────────────────
    for pt_key in sorted_keys:
        pt_rows = groups[pt_key]
        ws = wb.create_sheet(title=pt_key[:31])

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = ws.cell(row=1, column=1,
                             value=f"Order Genius — {country_name} ({country_code}) {year}")
        title_cell.font = Font(name="Calibri", size=14, bold=True)

        # Header row
        for col_idx, header in enumerate(effective_headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            _apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL,
                              alignment=CENTER_ALIGN, border=THIN_BORDER)

        # Data rows
        for row_idx, row in enumerate(pt_rows):
            excel_row = row_idx + 3
            is_historical = row.get("lifecycleStatus") == "historical"
            row_font = HISTORICAL_FONT if is_historical else NORMAL_FONT
            row_fill = STRIPED_FILL if row_idx % 2 == 1 else None

            col = 1
            if include_row_numbers:
                cell = ws.cell(row=excel_row, column=col, value=row_idx + 1)
                _apply_cell_style(cell, font=row_font, fill=row_fill, border=THIN_BORDER)
                col += 1

            values = [
                row.get("modelName", ""),
                row.get("version", ""),
                row.get("colour", ""),
                row.get("interiorColorName", ""),
                row.get("materialCode", ""),
                row.get("fobEur"),
            ]
            for val in values:
                cell = ws.cell(row=excel_row, column=col, value=val)
                _apply_cell_style(cell, font=row_font, fill=row_fill, border=THIN_BORDER)
                col += 1

            # Month columns
            months = row.get("months", {})
            for month_num in export_months:
                month_data = months.get(str(month_num), {})
                qty = month_data.get("quantity", 0) if month_data else 0
                cell = ws.cell(row=excel_row, column=col, value=qty)
                _apply_cell_style(cell, font=row_font, fill=row_fill,
                                  alignment=CENTER_ALIGN, border=THIN_BORDER)
                col += 1

            # TTL column
            ttl_cell = ws.cell(row=excel_row, column=col, value=_selected_month_ttl(row, export_months))
            _apply_cell_style(ttl_cell, font=row_font, fill=row_fill,
                              alignment=CENTER_ALIGN, border=THIN_BORDER)

        # Freeze header + left columns (adjust for optional No. column)
        freeze_col = get_column_letter(7 if not include_row_numbers else 8)
        ws.freeze_panes = f"{freeze_col}3"

        # Auto-fit column widths
        for col_idx in range(1, n_cols + 1):
            max_width = len(str(effective_headers[col_idx - 1])) + 2
            for r_idx in range(3, len(pt_rows) + 3):
                val = ws.cell(row=r_idx, column=col_idx).value
                if val is not None:
                    max_width = max(max_width, len(str(val)) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_order_genius_pi_excel(
    rows: list[dict],
    country_code: str,
    country_name: str,
    year: int,
    quantity_month: int | None,
    nl_fob_by_material_code: dict[str, float | None],
) -> io.BytesIO:
    """Generate a PI-ready workbook from Order Genius matrix rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PI"

    title_month = f"M{quantity_month:02d}" if quantity_month else "TTL"
    title = f"PI Export — {country_name} ({country_code}) {year} — Quantity: {title_month}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PI_HEADERS))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", size=14, bold=True)
    title_cell.alignment = LEFT_ALIGN

    manual_note = "预留列可手动填写；Excel 中可直接拖动复制。"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(PI_HEADERS))
    note_cell = ws.cell(row=2, column=1, value=manual_note)
    note_cell.font = Font(name="Calibri", size=10, color="64748B")
    note_cell.alignment = LEFT_ALIGN

    for col_idx, header in enumerate(PI_HEADERS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        _apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL,
                          alignment=CENTER_ALIGN, border=THIN_BORDER)

    manual_cols = {2, 6, 7, 9, 10, 11, 13, 15, 16}
    manual_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for row_idx, row in enumerate(rows, start=4):
        material_code = row.get("materialCode", "")
        values = [
            _pi_colour_kind(row),
            "",
            material_code,
            _pi_quantity(row, quantity_month),
            row.get("fobEur"),
            "",
            "",
            _pi_product_category(row),
            "/",
            "/",
            "/",
            _pi_exterior(row),
            "",
            nl_fob_by_material_code.get(material_code),
            "",
            "",
        ]
        row_fill = STRIPED_FILL if (row_idx - 4) % 2 == 1 else None
        for col_idx, value in enumerate(values, 1):
            fill = manual_fill if col_idx in manual_cols else row_fill
            alignment = RIGHT_ALIGN if col_idx in {4, 5, 6, 7, 13, 14, 15, 16} else LEFT_ALIGN
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply_cell_style(cell, font=NORMAL_FONT, fill=fill,
                              alignment=alignment, border=THIN_BORDER)
            if col_idx in {4, 5, 6, 7, 13, 14, 15, 16}:
                cell.number_format = "#,##0"

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(PI_HEADERS))}{max(3, len(rows) + 3)}"

    widths = {
        1: 10, 2: 22, 3: 20, 4: 10, 5: 12, 6: 12, 7: 12, 8: 18,
        9: 10, 10: 12, 11: 12, 12: 18, 13: 12, 14: 16, 15: 18, 16: 18,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _pi_quantity(row: dict, quantity_month: int | None) -> int:
    if quantity_month:
        month_data = row.get("months", {}).get(str(quantity_month), {})
        return month_data.get("quantity", 0) if month_data else 0
    return _selected_month_ttl(row, list(range(1, 13)))


def _pi_colour_kind(row: dict) -> str:
    colour = row.get("colour", "") or ""
    colour_type = (row.get("colourType") or "").lower()
    colour_tier = (row.get("colourTier") or "").lower()
    is_dual = colour_type == "dual" or colour_tier == "dual" or bool(re.search(r"[&／]", colour))
    return "拼色" if is_dual else "单色"


def _pi_powertrain_label(row: dict) -> str:
    text = f"{row.get('modelName', '')} {row.get('powertrain', '')}".upper()
    if "SHS" in text:
        return "SHS"
    if "PHEV" in text or "PLUG" in text:
        return "PHEV"
    if "HEV" in text:
        return "HEV"
    if "BEV" in text or re.search(r"\bEV\b", text):
        return "EV"
    if "ICE" in text:
        return "ICE"
    return (row.get("powertrain") or "").upper()


def _pi_product_category(row: dict) -> str:
    brand = (row.get("brand") or "").upper()
    model_name = row.get("modelName") or ""
    if "OMODA" in brand or "OMODA" in model_name.upper():
        prefix = "O"
    elif "JAECOO" in brand or "JAECOO" in model_name.upper():
        prefix = "J"
    else:
        prefix = (brand[:1] or model_name[:1]).upper()
    model_number = re.search(r"\d+", model_name)
    model_part = f"{prefix}{model_number.group(0)}" if model_number else model_name.strip()
    powertrain = _pi_powertrain_label(row)
    return f"{model_part} {powertrain}".strip()


def _pi_exterior(row: dict) -> str:
    colour = (row.get("colour") or "").strip()
    code = (row.get("colourCode") or "").strip()
    if colour and code and f"({code})" not in colour:
        return f"{colour} ({code})"
    return colour or code


def _selected_month_ttl(row: dict, export_months: list[int]) -> int:
    months = row.get("months", {})
    return sum(
        (months.get(str(month), {}) or {}).get("quantity", 0) or 0
        for month in export_months
    )


def _build_total_list_sheet(
    wb,
    rows: list[dict],
    country_code: str,
    country_name: str,
    year: int,
    effective_headers: list[str],
    n_cols: int,
    month_names: list[str],
    include_row_numbers: bool,
    export_months: list[int],
) -> None:
    """Create a 'Total list' aggregate sheet grouped by product identity + powertrain."""
    ws = wb.create_sheet(title="Total list")

    # Group by product key: brand|modelName|version|powertrain
    agg: dict[str, dict] = {}
    for row in rows:
        key = f"{row.get('brand', '')}|{row.get('modelName', '')}|{row.get('version', '')}|{row.get('powertrain', '')}"
        if key not in agg:
            agg[key] = {
                "brand": row.get("brand", ""),
                "modelName": row.get("modelName", ""),
                "version": row.get("version", ""),
                "powertrain": row.get("powertrain", ""),
                "months": {str(m): 0 for m in export_months},
                "ttl": 0,
            }
        months = row.get("months", {})
        for m in export_months:
            md = months.get(str(m), {})
            qty = md.get("quantity", 0) if md else 0
            agg[key]["months"][str(m)] += qty
            agg[key]["ttl"] += qty

    agg_rows = list(agg.values())
    # Sort by brand, model, version
    agg_rows.sort(key=lambda r: (r["brand"], r["modelName"], r["version"]))

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1,
                         value=f"Order Genius — {country_name} ({country_code}) {year} — Total List")
    title_cell.font = Font(name="Calibri", size=14, bold=True)

    # Header row (skip Model/Version/Colour/Material Code/FOB, add Product Code + Powertrain)
    total_headers = ["No.", "Product Code", "Powertrain"] if include_row_numbers else ["Product Code", "Powertrain"]
    total_headers += month_names + ["TTL"]

    for col_idx, header in enumerate(total_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        _apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL,
                          alignment=CENTER_ALIGN, border=THIN_BORDER)

    for row_idx, ar in enumerate(agg_rows):
        excel_row = row_idx + 3
        row_fill = STRIPED_FILL if row_idx % 2 == 1 else None
        col = 1

        if include_row_numbers:
            c = ws.cell(row=excel_row, column=col, value=row_idx + 1)
            _apply_cell_style(c, font=NORMAL_FONT, fill=row_fill, border=THIN_BORDER)
            col += 1

        # Product Code = brand + modelName (e.g. "OMODA OMODA5")
        product_code = f"{ar['brand']} {ar['modelName']}".strip()
        c = ws.cell(row=excel_row, column=col, value=product_code)
        _apply_cell_style(c, font=NORMAL_FONT, fill=row_fill, border=THIN_BORDER)
        col += 1

        c = ws.cell(row=excel_row, column=col, value=ar["powertrain"])
        _apply_cell_style(c, font=NORMAL_FONT, fill=row_fill, border=THIN_BORDER)
        col += 1

        for m in export_months:
            qty = ar["months"][str(m)]
            c = ws.cell(row=excel_row, column=col, value=qty if qty else 0)
            _apply_cell_style(c, font=NORMAL_FONT, fill=row_fill,
                              alignment=CENTER_ALIGN, border=THIN_BORDER)
            col += 1

        c = ws.cell(row=excel_row, column=col, value=ar["ttl"])
        _apply_cell_style(c, font=NORMAL_FONT, fill=row_fill,
                          alignment=CENTER_ALIGN, border=THIN_BORDER)

    ws.freeze_panes = "C3"

    # Auto-fit
    for col_idx in range(1, len(total_headers) + 1):
        max_width = len(str(total_headers[col_idx - 1])) + 2
        for r_idx in range(3, len(agg_rows) + 3):
            val = ws.cell(row=r_idx, column=col_idx).value
            if val is not None:
                max_width = max(max_width, len(str(val)) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width, 30)
