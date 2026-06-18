"""Reusable workbook table scanning helpers.

The scanner keeps Excel-specific mechanics out of business services: multi-sheet
selection, header inference, target column creation, source cell references, and
lightweight date parsing. Domain services still own matching and persistence.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


HEADER_SCAN_ROWS = 30
MATERIAL_GROUP_RE = re.compile(r"\b[A-Z0-9]{5,}\*\*[A-Z0-9]{4,}\b", re.IGNORECASE)
WVTA_RE = re.compile(r"e\d+\*2018/858\*[^\s]+", re.IGNORECASE)
COC_RE = re.compile(r"\d{5}-\d{2}&[^\s]*C[O0]C[^\s]*", re.IGNORECASE)


@dataclass(frozen=True)
class MaterialGroupRow:
    sheet_name: str
    row_number: int
    material_group: str
    material_no: str | None
    model: str | None
    country: str | None
    production_date_raw: str | None
    production_date_start: date | None
    production_date_end: date | None
    wvta_cell: str
    coc_cell: str
    existing_wvta: str | None
    existing_coc: str | None
    header_inferred: bool


@dataclass(frozen=True)
class TargetColumns:
    header_row: int
    material_col: int
    wvta_col: int
    coc_col: int
    header_inferred: bool
    metadata_cols: dict[str, int]


def _normalize_material_group(value: object) -> str:
    text = str(value or "").strip().upper()
    match = MATERIAL_GROUP_RE.search(text)
    return match.group(0).upper() if match else ""


def _normalize_header(value: object) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", str(value or "").strip().lower())


def _display_names(names: list[str] | set[str]) -> str:
    return "、".join(str(name) for name in names) if names else "-"


def _cell_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _parse_date(value: object, *, day_first: bool = False) -> date | None:
    if value in (None, "", "/"):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text or text == "/":
        return None

    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:10], pattern).date()
        except ValueError:
            pass

    match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", text)
    if match:
        first, second, year = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
        day, month = (first, second) if day_first else (second, first)
        try:
            return date(year, month, day)
        except ValueError:
            return None

    match = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})", text)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    return None


def _parse_excel_date_range(value: object) -> tuple[date | None, date | None]:
    if value in (None, ""):
        return None, None
    if isinstance(value, (date, datetime)):
        parsed = _parse_date(value)
        return parsed, parsed
    text = str(value).strip()
    if not text:
        return None, None

    match = re.search(
        r"(\d{4})年(\d{1,2})月(\d{1,2})日?\s*[-~至到]\s*(?:(\d{1,2})月)?(\d{1,2})日?",
        text,
    )
    if match:
        year = int(match.group(1))
        start_month = int(match.group(2))
        start_day = int(match.group(3))
        end_month = int(match.group(4) or start_month)
        end_day = int(match.group(5))
        try:
            return date(year, start_month, start_day), date(year, end_month, end_day)
        except ValueError:
            return None, None

    parsed = _parse_date(text)
    return parsed, parsed


def _header_metadata_columns(sheet: Worksheet, header_row: int) -> dict[str, int]:
    result: dict[str, int] = {}
    aliases = {
        "material_no": {"物料号", "materialcode", "materialno"},
        "model": {"车型", "车型名", "model", "modelname"},
        "country": {"国家", "country", "market", "市场"},
        "production_date": {"生产日期", "时间", "日期", "productiondate", "date"},
    }
    for col in range(1, (sheet.max_column or 0) + 1):
        normalized = _normalize_header(sheet.cell(header_row, col).value)
        if not normalized:
            continue
        for key, values in aliases.items():
            if normalized in values and key not in result:
                result[key] = col
    return result


def _find_target_from_headers(sheet: Worksheet, header_row: int) -> tuple[int | None, int | None]:
    wvta_col: int | None = None
    coc_col: int | None = None
    for col in range(1, (sheet.max_column or 0) + 1):
        normalized = _normalize_header(sheet.cell(header_row, col).value)
        if not normalized:
            continue
        if "wvta" in normalized and wvta_col is None:
            wvta_col = col
        if "coc" in normalized and coc_col is None:
            coc_col = col
    return wvta_col, coc_col


def _find_target_from_values(
    sheet: Worksheet,
    header_row: int,
    material_col: int,
) -> tuple[int | None, int | None]:
    wvta_col: int | None = None
    coc_col: int | None = None
    for row in range(header_row + 1, min(sheet.max_row or 0, header_row + 25) + 1):
        for col in range(material_col + 1, (sheet.max_column or 0) + 1):
            text = str(sheet.cell(row, col).value or "").strip()
            if not text:
                continue
            if wvta_col is None and WVTA_RE.search(text):
                wvta_col = col
            if coc_col is None and COC_RE.search(text):
                coc_col = col
            if wvta_col is not None and coc_col is not None:
                return wvta_col, coc_col
    return wvta_col, coc_col


def _ensure_target_columns(sheet: Worksheet, header_row: int, material_col: int) -> tuple[int, int]:
    start_col = max(sheet.max_column or material_col, material_col) + 1
    sheet.cell(header_row, start_col).value = "WVTA编号"
    sheet.cell(header_row, start_col + 1).value = "COC编号"
    sheet.cell(header_row, start_col).font = Font(bold=True)
    sheet.cell(header_row, start_col + 1).font = Font(bold=True)
    return start_col, start_col + 1


def _find_material_header(sheet: Worksheet) -> tuple[int, int] | None:
    for row in range(1, min(sheet.max_row or 0, HEADER_SCAN_ROWS) + 1):
        for col in range(1, (sheet.max_column or 0) + 1):
            normalized = _normalize_header(sheet.cell(row, col).value)
            if normalized == "物料号组" or normalized == "materialgroup":
                return row, col
    return None


def _infer_material_header(sheet: Worksheet) -> tuple[int, int] | None:
    hit_counts: dict[int, tuple[int, int]] = {}
    for row in range(1, (sheet.max_row or 0) + 1):
        for col in range(1, (sheet.max_column or 0) + 1):
            material = _normalize_material_group(sheet.cell(row, col).value)
            if not material:
                continue
            count, first_row = hit_counts.get(col, (0, row))
            hit_counts[col] = (count + 1, min(first_row, row))
    if not hit_counts:
        return None
    material_col, (_count, first_row) = max(hit_counts.items(), key=lambda item: item[1][0])
    header_row = max(1, first_row - 1)
    if not sheet.cell(header_row, material_col).value:
        sheet.cell(header_row, material_col).value = "物料号组"
    return header_row, material_col


def _target_columns_for_sheet(sheet: Worksheet) -> TargetColumns | None:
    header_inferred = False
    header = _find_material_header(sheet)
    if header is None:
        header = _infer_material_header(sheet)
        header_inferred = header is not None
    if header is None:
        return None

    header_row, material_col = header
    wvta_col, coc_col = _find_target_from_headers(sheet, header_row)
    if wvta_col is None or coc_col is None:
        value_wvta_col, value_coc_col = _find_target_from_values(sheet, header_row, material_col)
        wvta_col = wvta_col or value_wvta_col
        coc_col = coc_col or value_coc_col
    if wvta_col is None or coc_col is None:
        wvta_col, coc_col = _ensure_target_columns(sheet, header_row, material_col)
    if not sheet.cell(header_row, wvta_col).value:
        sheet.cell(header_row, wvta_col).value = "WVTA编号"
    if not sheet.cell(header_row, coc_col).value:
        sheet.cell(header_row, coc_col).value = "COC编号"

    return TargetColumns(
        header_row=header_row,
        material_col=material_col,
        wvta_col=wvta_col,
        coc_col=coc_col,
        header_inferred=header_inferred,
        metadata_cols=_header_metadata_columns(sheet, header_row),
    )


def _material_rows_for_sheet(sheet: Worksheet, targets: TargetColumns) -> list[MaterialGroupRow]:
    rows: list[MaterialGroupRow] = []
    metadata = targets.metadata_cols
    for row in range(targets.header_row + 1, (sheet.max_row or 0) + 1):
        material = _normalize_material_group(sheet.cell(row, targets.material_col).value)
        if not material:
            continue
        production_raw = _cell_text(
            sheet.cell(row, metadata["production_date"]).value
            if "production_date" in metadata
            else None
        )
        production_start, production_end = _parse_excel_date_range(production_raw)
        wvta_cell = f"{get_column_letter(targets.wvta_col)}{row}"
        coc_cell = f"{get_column_letter(targets.coc_col)}{row}"
        rows.append(
            MaterialGroupRow(
                sheet_name=sheet.title,
                row_number=row,
                material_group=material,
                material_no=_cell_text(sheet.cell(row, metadata["material_no"]).value)
                if "material_no" in metadata
                else None,
                model=_cell_text(sheet.cell(row, metadata["model"]).value)
                if "model" in metadata
                else None,
                country=_cell_text(sheet.cell(row, metadata["country"]).value)
                if "country" in metadata
                else None,
                production_date_raw=production_raw,
                production_date_start=production_start,
                production_date_end=production_end,
                wvta_cell=wvta_cell,
                coc_cell=coc_cell,
                existing_wvta=_cell_text(sheet.cell(row, targets.wvta_col).value),
                existing_coc=_cell_text(sheet.cell(row, targets.coc_col).value),
                header_inferred=targets.header_inferred,
            )
        )
    return rows


def _extract_material_rows(
    workbook_path: Path,
    sheet_names: list[str] | None = None,
) -> tuple[Any, dict[str, TargetColumns], list[MaterialGroupRow]]:
    workbook = load_workbook(workbook_path)
    requested_names = [name.strip() for name in sheet_names or [] if name.strip()]
    available_sheets = [sheet.title for sheet in workbook.worksheets if sheet.title != "COC填充结果"]
    available_by_key = {name.casefold(): name for name in available_sheets}
    missing_names = [name for name in requested_names if name.casefold() not in available_by_key]
    if missing_names:
        workbook.close()
        raise HTTPException(
            status_code=422,
            detail=(
                f"指定 Sheet 不存在：{_display_names(missing_names)}。"
                f"当前 Excel 包含：{_display_names(available_sheets)}。"
            ),
        )
    requested_sheets = {available_by_key[name.casefold()] for name in requested_names}
    targets_by_sheet: dict[str, TargetColumns] = {}
    rows: list[MaterialGroupRow] = []
    for sheet in workbook.worksheets:
        if sheet.title == "COC填充结果":
            continue
        if requested_sheets and sheet.title not in requested_sheets:
            continue
        targets = _target_columns_for_sheet(sheet)
        if targets is None:
            continue
        sheet_rows = _material_rows_for_sheet(sheet, targets)
        if not sheet_rows:
            continue
        targets_by_sheet[sheet.title] = targets
        rows.extend(sheet_rows)
    return workbook, targets_by_sheet, rows


normalize_material_group = _normalize_material_group
display_names = _display_names
cell_text = _cell_text
parse_date = _parse_date
target_columns_for_sheet = _target_columns_for_sheet
extract_material_rows = _extract_material_rows

__all__ = [
    "MaterialGroupRow",
    "TargetColumns",
    "cell_text",
    "display_names",
    "extract_material_rows",
    "normalize_material_group",
    "parse_date",
    "target_columns_for_sheet",
]
